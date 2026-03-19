import warnings
warnings.filterwarnings("ignore", category=UserWarning, module="openpyxl")
from flask import Flask, render_template, request, redirect, url_for, flash, jsonify, send_file, send_from_directory, Response
from flask_sqlalchemy import SQLAlchemy
from flask_login import LoginManager, UserMixin, login_user, login_required, logout_user, current_user
from werkzeug.security import generate_password_hash, check_password_hash
from werkzeug.utils import secure_filename
from openpyxl import load_workbook, Workbook
import pandas as pd
import os
import sys
import threading
import subprocess
import uuid
import shutil
from pathlib import Path
from datetime import datetime, timedelta
from types import SimpleNamespace
import json
from collections import defaultdict
import csv
from io import BytesIO
import re
import time
import math

PROJECT_ROOT = Path(__file__).resolve().parents[1]
if str(PROJECT_ROOT) not in sys.path:
    sys.path.insert(0, str(PROJECT_ROOT))

from config import (
    FRONTEND_DB_PATH,
    FRONTEND_INSTANCE_DIR,
    FRONTEND_UPLOAD_DIR,
    PIPELINE_RUNS_DIR,
    PIPELINE_TEMPLATE_DIR,
    PROJECT_ROOT,
    SECRET_KEY,
    STAGE1_INPUT_BACKUP_DIR,
    STAGE1_INPUT_DIR,
    STAGE2_EF_XLSX,
    STAGE2_MAPPING_DIR,
    STAGE2_OUTPUT_DIR,
)

APP_DIR = Path(__file__).resolve().parent
INSTANCE_DIR = FRONTEND_INSTANCE_DIR

# Use centralized configuration so file-system paths do not change between
# local development and server deployments.
app = Flask(__name__, instance_path=str(INSTANCE_DIR), instance_relative_config=True)
app.config['SECRET_KEY'] = SECRET_KEY
INSTANCE_DIR.mkdir(parents=True, exist_ok=True)
app.config['SQLALCHEMY_DATABASE_URI'] = 'sqlite:///' + str(FRONTEND_DB_PATH)
app.config['SQLALCHEMY_TRACK_MODIFICATIONS'] = False
app.config['UPLOAD_FOLDER'] = str(FRONTEND_UPLOAD_DIR)

# ---- Stage2 mapping (web single-company runner) ----
STAGE2_MAPPING_OUTPUT_DIR = STAGE2_OUTPUT_DIR
_STAGE2_MAP_LOCK = threading.Lock()
_MAPPING_RUNS: dict[str, dict[str, object]] = {}  # legacy in-memory (kept for backward compatibility)

# ---- Company canonical names + countries (web mapping) ----
# User-provided source-of-truth list (deduplicated).
_COMPANY_COUNTRY_CANONICAL: dict[str, str] = {
    "BIMMS": "Portugal",
    "CTS Finland": "Finland",
    "CTS-VDC Services": "Ireland",
    "DC Piping": "Portugal",
    "GT Nordics": "Norway",
    "Mecwide Nordics": "Norway",
    "Porvelox": "Portugal",
    "Caerus Nordics": "Norway",
    "CTS Sweden": "Sweden",
    "Navitas Portugal": "Portugal",
    "NEP Switchboards": "Norway",
    "CTS Denmark": "Denmark",
    "CTS Group": "Switzerland",
    "CTS Nordics": "Norway",
    "Fortica": "Norway",
    "MC Prefab": "Sweden",
    "Navitas Norway": "Norway",
    "QEC": "Norway",
    "SD Nordics": "Norway",
    "CTS Security Solutions": "Sweden",
    "Velox": "Norway",
    "CTS EU": "Portugal",
    "Gapit": "Norway",
}


def _canonical_company_name_and_country(name: str) -> tuple[str, str | None]:
    raw = (name or "").strip()
    if not raw:
        return "", None

    def norm(s: str) -> str:
        return "".join(ch.lower() for ch in (s or "") if ch.isalnum())

    n = norm(raw)
    # Build synonyms (file stems vs display names)
    synonyms: dict[str, str] = {
        norm("CTS-VDC"): "CTS-VDC Services",
        norm("CTS VDC"): "CTS-VDC Services",
        norm("CTS-VDC Services"): "CTS-VDC Services",
        norm("CTS Group HQ"): "CTS Group",
        norm("CTS Group"): "CTS Group",
        norm("NordicEPOD"): "Nordic EPOD",
        norm("Nordic EPOD"): "Nordic EPOD",
        norm("Caerus Nordics"): "Caerus Nordics",
        norm("CTS EU"): "CTS EU",
        norm("GT Nordics"): "GT Nordics",
        norm("Mecwide Nordics"): "Mecwide Nordics",
        norm("Navitas Norway"): "Navitas Norway",
        norm("Navitas Portugal"): "Navitas Portugal",
        norm("CTS Denmark"): "CTS Denmark",
        norm("CTS Finland"): "CTS Finland",
        norm("CTS Sweden"): "CTS Sweden",
        norm("CTS Nordics"): "CTS Nordics",
        norm("CTS Security Solutions"): "CTS Security Solutions",
        norm("DC Piping"): "DC Piping",
        norm("MC Prefab"): "MC Prefab",
        norm("Porvelox"): "Porvelox",
        norm("Fortica"): "Fortica",
        norm("Gapit"): "Gapit",
        norm("QEC"): "QEC",
        norm("SD Nordics"): "SD Nordics",
        norm("Velox"): "Velox",
        norm("BIMMS"): "BIMMS",
        norm("NEP Switchboards"): "NEP Switchboards",
        norm("NEP Switchboards AS"): "NEP Switchboards",
    }

    canonical = synonyms.get(n, raw)
    # Country lookup uses canonical keys; if missing, best-effort match by normalized canonical names
    country = _COMPANY_COUNTRY_CANONICAL.get(canonical)
    if country is None:
        want = norm(canonical)
        for k, v in _COMPANY_COUNTRY_CANONICAL.items():
            if norm(k) == want:
                country = v
                canonical = k
                break

    return canonical, country

db = SQLAlchemy(app)
login_manager = LoginManager()
login_manager.init_app(app)
login_manager.login_view = 'login'

# Database models
class User(UserMixin, db.Model):
    id = db.Column(db.Integer, primary_key=True)
    email = db.Column(db.String(120), unique=True, nullable=False)
    password_hash = db.Column(db.String(200), nullable=False)
    company_name = db.Column(db.String(200), nullable=False)
    is_admin = db.Column(db.Boolean, default=False)
    created_at = db.Column(db.DateTime, default=datetime.utcnow)

class EmissionFactor(db.Model):
    id = db.Column(db.Integer, primary_key=True)
    category = db.Column(db.String(200), nullable=False)
    subcategory = db.Column(db.String(200), nullable=False)
    factor = db.Column(db.Float, nullable=False)
    unit = db.Column(db.String(100))
    year = db.Column(db.Integer)
    description = db.Column(db.Text)
    extra_data = db.Column(db.Text)  # Tüm satır verisi JSON olarak


class PipelineRun(db.Model):
    id = db.Column(db.Integer, primary_key=True)
    user_id = db.Column(db.Integer, db.ForeignKey('user.id'), nullable=False)
    created_at = db.Column(db.DateTime, default=datetime.utcnow)
    status = db.Column(db.String(30), default="queued")  # queued, running, succeeded, failed
    input_dir = db.Column(db.String(800), nullable=False)
    run_dir = db.Column(db.String(800), nullable=False)
    stage1_output = db.Column(db.String(800))
    stage2_output_dir = db.Column(db.String(800))
    log_path = db.Column(db.String(800))
    exit_code = db.Column(db.Integer)
    error_message = db.Column(db.Text)


class MappingRun(db.Model):
    """
    Web-triggered single-company Stage2 mapping runs (persisted).
    """
    id = db.Column(db.String(32), primary_key=True)  # run_id (hex)
    user_id = db.Column(db.Integer, db.ForeignKey('user.id'), nullable=False)
    company_name = db.Column(db.String(200), nullable=False)
    sheet_name = db.Column(db.String(200), nullable=False)
    created_at = db.Column(db.DateTime, default=datetime.utcnow)
    status = db.Column(db.String(30), default="queued")  # queued, running, succeeded, failed
    input_path = db.Column(db.String(900))
    output_path = db.Column(db.String(900))
    error_message = db.Column(db.Text)


class MappingRunSummary(db.Model):
    """
    Lightweight persisted totals derived from a MappingRun output workbook.
    Used to power Home / Carbon Accounting dashboards.
    """
    id = db.Column(db.Integer, primary_key=True)
    run_id = db.Column(db.String(32), unique=True, nullable=False)
    company_name = db.Column(db.String(200), nullable=False)
    sheet_name = db.Column(db.String(200), nullable=False)
    scope = db.Column(db.Integer)  # 1/2/3 or NULL
    tco2e_total = db.Column(db.Float, default=0.0)
    rows_count = db.Column(db.Integer, default=0)
    created_at = db.Column(db.DateTime, default=datetime.utcnow)


def _ensure_db_tables() -> None:
    """
    Best-effort table creation for environments that don't run app.py directly.
    Safe to call within request context.
    """
    try:
        db.create_all()
    except Exception:
        pass


def _infer_scope_from_sheet(sheet_name: str) -> int | None:
    s = (sheet_name or "").strip().lower()
    if "scope 1" in s or s.startswith("scope1"):
        return 1
    if "scope 2" in s or s.startswith("scope2"):
        return 2
    if "scope 3" in s or s.startswith("scope3"):
        return 3
    return None


def _find_tco2e_column(df: "pd.DataFrame") -> str | None:
    if df is None or getattr(df, "columns", None) is None:
        return None

    def norm(x: str) -> str:
        return "".join(ch.lower() for ch in (x or "") if ch.isalnum())

    candidates: list[str] = []
    for c in list(df.columns):
        n = norm(str(c))
        if n in {"emissionstco2e", "emissionstco2etonnes", "tco2e", "totalemissionstco2e"}:
            candidates.append(str(c))
            continue
        if n in ("co2e", "co2et", "co2etonnes", "co2eton", "co2etonne"):
            candidates.append(str(c))
            continue
        if n.startswith("emissionstco2e") or n.startswith("co2et") or n.startswith("co2e"):
            candidates.append(str(c))
            continue
        if ("co2e" in n) and ("calculated" in n):
            candidates.append(str(c))

    if not candidates:
        return None

    best_col = None
    best_score = (-1, -1.0)
    for c in candidates:
        try:
            ser = df[c]
        except Exception:
            continue

        numeric_count = 0
        total_abs = 0.0
        try:
            for v in ser.tolist():
                f = _parse_float_loose(v)
                if f is None:
                    continue
                numeric_count += 1
                total_abs += abs(float(f))
        except Exception:
            continue

        if (numeric_count, total_abs) > best_score:
            best_col = str(c)
            best_score = (numeric_count, total_abs)

    return best_col


def _sum_tco2e(df: "pd.DataFrame") -> tuple[float, int, str | None]:
    """
    Returns (total_tco2e, rows_count, used_column).
    """
    col = _find_tco2e_column(df)
    if not col:
        return 0.0, int(len(df.index)) if df is not None else 0, None
    try:
        ser_raw = df[col]
        # Robust numeric parsing for Excel exports (comma decimals, NBSP, etc.)
        try:
            if getattr(ser_raw, "dtype", None) == "object":
                ser_raw = (
                    ser_raw.astype(str)
                    .str.replace("\u00a0", "", regex=False)
                    .str.replace(" ", "", regex=False)
                    .str.replace(",", ".", regex=False)
                )
        except Exception:
            pass
        ser = pd.to_numeric(ser_raw, errors="coerce")
        total = float(ser.fillna(0).sum())
    except Exception:
        total = 0.0
    return total, int(len(df.index)) if df is not None else 0, col


def _parse_float_loose(v) -> float | None:
    if v is None:
        return None
    try:
        if pd.isna(v):
            return None
    except Exception:
        pass
    if isinstance(v, (int, float)):
        try:
            out = float(v)
            if math.isnan(out) or math.isinf(out):
                return None
            return out
        except Exception:
            return None
    try:
        s = str(v).strip()
    except Exception:
        return None
    if not s:
        return None
    s = s.replace("\u00a0", "").replace(" ", "")
    # Extract first numeric token (handles "12.3t", "1,23", "≈ 5.1", etc.)
    import re as _re
    m = _re.search(r"[+-]?\d+(?:[.,]\d+)?(?:e[+-]?\d+)?", s, flags=_re.IGNORECASE)
    if not m:
        return None
    token = m.group(0).replace(",", ".")
    try:
        out = float(token)
        if math.isnan(out) or math.isinf(out):
            return None
        return out
    except Exception:
        return None


def _sum_tco2e_from_xlsx(xlsx_path: str | Path, sheet_name: str | None) -> tuple[float, int, str | None]:
    """
    Sum tCO2e values from an Excel workbook using openpyxl with data_only=True.
    This is more reliable than pandas when cells contain formulas (Excel-visible numbers).
    Returns (total_tco2e, rows_count, used_header_name).
    """
    try:
        p = Path(xlsx_path)
    except Exception:
        p = None
    if not p or not str(p) or not os.path.exists(str(p)):
        return 0.0, 0, None

    try:
        wb = load_workbook(str(p), read_only=True, data_only=True, keep_links=False)
    except Exception:
        return 0.0, 0, None

    target_ws = None
    want = (sheet_name or "").strip().lower()
    if want:
        for n in wb.sheetnames:
            if str(n).strip().lower() == want:
                target_ws = wb[n]
                break
    if target_ws is None and wb.sheetnames:
        target_ws = wb[wb.sheetnames[0]]
    if target_ws is None:
        return 0.0, 0, None

    def norm(x: str) -> str:
        return "".join(ch.lower() for ch in (x or "") if ch.isalnum())

    header_row, headers = _detect_header_row_and_headers(target_ws, max_scan_rows=50)
    candidates: list[tuple[int, str]] = []
    for i, h in enumerate(headers, start=1):
        nh = norm(str(h))
        if nh in (
            "emissionstco2e",
            "emissionstco2etonnes",
            "tco2e",
            "totalemissionstco2e",
            "co2e",
            "co2et",
            "co2etonnes",
        ):
            candidates.append((i, str(h)))
            continue
        if nh.startswith("emissionstco2e") or nh.startswith("co2et") or nh.startswith("co2e"):
            candidates.append((i, str(h)))
            continue
        # Some providers include long headers containing "Calculated CO2e"
        if ("co2e" in nh) and ("calculated" in nh):
            candidates.append((i, str(h)))

    if not candidates:
        return 0.0, 0, None

    rows_count = 0
    max_r = int(getattr(target_ws, "max_row", 1) or 1)
    # First pass: count non-empty rows (for display) + compute per-candidate sums.
    best = {"used": None, "total": 0.0, "numeric_count": 0}
    for r in range(header_row + 1, max_r + 1):
        # Skip fully empty rows to keep counts consistent
        row_has_any = False
        try:
            for c in range(1, min(int(getattr(target_ws, "max_column", 1) or 1), 40) + 1):
                v2 = target_ws.cell(row=r, column=c).value
                if v2 is not None and str(v2).strip() != "":
                    row_has_any = True
                    break
        except Exception:
            row_has_any = True
        if not row_has_any:
            continue

        rows_count += 1
        for col_idx, used in candidates:
            val = target_ws.cell(row=r, column=col_idx).value
            f = _parse_float_loose(val)
            if f is None:
                continue
            # Accumulate into a temporary bucket per candidate by encoding in dict key
            k = f"{col_idx}:{used}"
            # stash totals in best dict dynamically
            # (avoid allocating large per-column arrays)
            tot_key = f"tot::{k}"
            cnt_key = f"cnt::{k}"
            best[tot_key] = float(best.get(tot_key, 0.0)) + float(f)
            best[cnt_key] = int(best.get(cnt_key, 0)) + 1

    # Choose the candidate column with the highest numeric_count; ties broken by total magnitude.
    chosen = None
    chosen_total = 0.0
    chosen_cnt = 0
    for col_idx, used in candidates:
        k = f"{col_idx}:{used}"
        cnt = int(best.get(f"cnt::{k}", 0))
        tot = float(best.get(f"tot::{k}", 0.0))
        if cnt > chosen_cnt or (cnt == chosen_cnt and abs(tot) > abs(chosen_total)):
            chosen = used
            chosen_total = tot
            chosen_cnt = cnt

    if chosen is None:
        return 0.0, int(rows_count), None

    return float(chosen_total), int(rows_count), str(chosen)


def _detect_data_year_from_xlsx(xlsx_path: str | Path, sheet_name: str | None, scan_rows: int = 500) -> int | None:
    """
    Best-effort: infer reporting year from a period column like
    'Reporting period (month, year)' with values like \"Jan'-2025\".
    Returns the most common year found (or max if tie), else None.
    """
    try:
        p = Path(xlsx_path)
    except Exception:
        return None
    if not p or not os.path.exists(str(p)):
        return None
    try:
        wb = load_workbook(str(p), read_only=True, data_only=True, keep_links=False)
    except Exception:
        return None

    want = (sheet_name or "").strip().lower()
    ws = None
    if want:
        for n in wb.sheetnames:
            if str(n).strip().lower() == want:
                ws = wb[n]
                break
    if ws is None and wb.sheetnames:
        ws = wb[wb.sheetnames[0]]
    if ws is None:
        return None

    def norm(x: str) -> str:
        return "".join(ch.lower() for ch in (x or "") if ch.isalnum())

    header_row, headers = _detect_header_row_and_headers(ws, max_scan_rows=50)
    period_col = None
    for i, h in enumerate(headers, start=1):
        nh = norm(str(h))
        if nh in ("reportingperiodmonthyear", "reportingperiod", "reportingperiodmonth", "monthyear", "periodmonthyear"):
            period_col = i
            break
        if ("period" in nh and "year" in nh) or nh.startswith("reportingperiod"):
            period_col = i
            break
    if period_col is None:
        return None

    import re as _re
    counts: dict[int, int] = {}
    max_r = int(getattr(ws, "max_row", 1) or 1)
    end = min(max_r, header_row + int(scan_rows))
    for r in range(header_row + 1, end + 1):
        v = ws.cell(row=r, column=period_col).value
        if v is None:
            continue
        s = str(v)
        m = _re.search(r"(19|20)\d{2}", s)
        if not m:
            continue
        y = int(m.group(0))
        counts[y] = counts.get(y, 0) + 1
    if not counts:
        return None
    best = sorted(counts.items(), key=lambda kv: (kv[1], kv[0]), reverse=True)[0][0]
    return int(best)


def _parse_period_value(value) -> datetime | None:
    if value is None:
        return None
    try:
        if isinstance(value, pd.Timestamp):
            if pd.isna(value):
                return None
            return datetime(int(value.year), int(value.month), 1)
    except Exception:
        pass
    if isinstance(value, datetime):
        return datetime(value.year, value.month, 1)

    s = str(value).strip()
    if not s:
        return None

    import re as _re
    month_map = {
        "jan": 1, "feb": 2, "mar": 3, "apr": 4, "may": 5, "jun": 6,
        "jul": 7, "aug": 8, "sep": 9, "oct": 10, "nov": 11, "dec": 12,
    }

    # Many uploaded workbooks store "reporting period (month, year)"
    # as the first day of each month, e.g. 01/03/2025 for Mar 2025.
    m = _re.match(r"^\s*(\d{1,2})[/-](\d{1,2})[/-]((?:19|20)\d{2})\s*$", s)
    if m:
        first = int(m.group(1))
        second = int(m.group(2))
        year = int(m.group(3))
        if first == 1 and 1 <= second <= 12:
            return datetime(year, second, 1)
        if second == 1 and 1 <= first <= 12:
            return datetime(year, first, 1)

    m = _re.search(r"(?i)\b(jan|feb|mar|apr|may|jun|jul|aug|sep|oct|nov|dec)[a-z]*[\'\s\-_/,]*(19|20)\d{2}\b", s)
    if m:
        mon = month_map[m.group(1).lower()[:3]]
        year_match = _re.search(r"(19|20)\d{2}", s)
        if year_match:
            return datetime(int(year_match.group(0)), mon, 1)

    m = _re.search(r"\b((?:19|20)\d{2})[-/](0?[1-9]|1[0-2])\b", s)
    if m:
        return datetime(int(m.group(1)), int(m.group(2)), 1)

    try:
        dt = pd.to_datetime([s], errors="coerce", dayfirst=False)[0]
        if pd.isna(dt):
            return None
        return datetime(int(dt.year), int(dt.month), 1)
    except Exception:
        return None


def _find_period_column(df: "pd.DataFrame", sample_rows: int = 120) -> str | None:
    if df is None or getattr(df, "columns", None) is None or len(df.columns) == 0:
        return None

    def norm(x: str) -> str:
        return "".join(ch.lower() for ch in (x or "") if ch.isalnum())

    preferred_exact = {
        "date",
        "reportingperiodmonthyear",
        "reportingperiod",
        "reportingperiodmonth",
        "monthyear",
        "periodmonthyear",
        "reportingmonth",
    }

    for c in list(df.columns):
        n = norm(str(c))
        if n in preferred_exact:
            return str(c)
        if ("date" in n) or ("period" in n and ("month" in n or "year" in n)):
            return str(c)

    first_col = str(df.columns[0])
    try:
        sample = [v for v in df[first_col].tolist()[:sample_rows] if str(v).strip() != ""]
        if sample:
            parsed = sum(1 for v in sample if _parse_period_value(v) is not None)
            if parsed >= max(2, int(len(sample) * 0.5)):
                return first_col
    except Exception:
        pass

    best_col = None
    best_score = (-1, -1.0)
    for c in list(df.columns):
        try:
            sample = [v for v in df[str(c)].tolist()[:sample_rows] if str(v).strip() != ""]
        except Exception:
            continue
        if not sample:
            continue
        parsed = sum(1 for v in sample if _parse_period_value(v) is not None)
        ratio = parsed / max(len(sample), 1)
        if parsed >= 3 and ratio >= 0.55 and (parsed, ratio) > best_score:
            best_col = str(c)
            best_score = (parsed, ratio)
    return best_col


def _read_sheet_df_from_workbook(xlsx_path: str | Path, sheet_name: str | None) -> "pd.DataFrame | None":
    try:
        sheets = pd.read_excel(xlsx_path, sheet_name=None, engine="openpyxl")
    except Exception:
        return None

    want = (sheet_name or "").strip().lower()
    if want:
        for k, v in sheets.items():
            if str(k).strip().lower() == want:
                return v

    if len(sheets) == 1:
        return next(iter(sheets.values()))

    if want:
        want_prefix = want[:20]
        for k, v in sheets.items():
            if str(k).strip().lower().startswith(want_prefix):
                return v

    return next(iter(sheets.values())) if sheets else None


def _build_period_profile_from_df(df: "pd.DataFrame") -> dict[str, object]:
    total_tco2e, rows_count, used_col = _sum_tco2e(df)
    if not used_col:
        return {
            "points": [],
            "total": 0.0,
            "rows_count": int(rows_count or 0),
            "min_date": None,
            "max_date": None,
        }

    period_col = _find_period_column(df)
    if not period_col:
        return {
            "points": [],
            "total": float(total_tco2e or 0.0),
            "rows_count": int(rows_count or 0),
            "min_date": None,
            "max_date": None,
        }

    buckets: dict[str, float] = defaultdict(float)
    for _, row in df.iterrows():
        try:
            raw_date = row.get(period_col)
            raw_value = row.get(used_col)
        except Exception:
            continue
        dt = _parse_period_value(raw_date)
        val = _parse_float_loose(raw_value)
        if dt is None or val is None:
            continue
        key = dt.strftime("%Y-%m")
        buckets[key] += float(val)

    if not buckets:
        return {
            "points": [],
            "total": float(total_tco2e or 0.0),
            "rows_count": int(rows_count or 0),
            "min_date": None,
            "max_date": None,
        }

    points = []
    for key in sorted(buckets.keys()):
        year, month = key.split("-")
        dt = datetime(int(year), int(month), 1)
        points.append(
            {
                "key": key,
                "label": dt.strftime("%b %Y"),
                "date": dt,
                "value": round(float(buckets[key]), 6),
            }
        )

    return {
        "points": points,
        "total": round(sum(float(p["value"]) for p in points), 6),
        "rows_count": int(rows_count or 0),
        "min_date": points[0]["date"] if points else None,
        "max_date": points[-1]["date"] if points else None,
    }


def _format_period_label(points: list[dict[str, object]], fallback_dt: datetime | None) -> str:
    if points:
        start = points[0].get("date")
        end = points[-1].get("date")
        if isinstance(start, datetime) and isinstance(end, datetime):
            if start.year == end.year and start.month == end.month:
                return start.strftime("%Y-%m")
            return f"{start.strftime('%Y-%m')} to {end.strftime('%Y-%m')}"
    if isinstance(fallback_dt, datetime):
        return fallback_dt.strftime("%Y-%m-%d")
    return ""


def _period_profile_for_summary(sub: "MappingRunSummary", run_cache: dict[str, "MappingRun | None"] | None = None) -> dict[str, object]:
    run_cache = run_cache if run_cache is not None else {}
    rid = str(getattr(sub, "run_id", "") or "")
    mr = run_cache.get(rid)
    if rid and rid not in run_cache:
        mr = MappingRun.query.get(rid)
        run_cache[rid] = mr

    if mr and getattr(mr, "output_path", None) and os.path.exists(str(mr.output_path)):
        df = _read_sheet_df_from_workbook(mr.output_path, getattr(sub, "sheet_name", None))
        if df is not None:
            profile = _build_period_profile_from_df(df)
            if profile.get("points"):
                return profile

    fallback_dt = getattr(sub, "created_at", None)
    fallback_point = None
    if isinstance(fallback_dt, datetime):
        fallback_point = {
            "key": fallback_dt.strftime("%Y-%m"),
            "label": fallback_dt.strftime("%b %Y"),
            "date": datetime(fallback_dt.year, fallback_dt.month, 1),
            "value": round(float(getattr(sub, "tco2e_total", 0.0) or 0.0), 6),
        }

    return {
        "points": [fallback_point] if fallback_point else [],
        "total": round(float(getattr(sub, "tco2e_total", 0.0) or 0.0), 6),
        "rows_count": int(getattr(sub, "rows_count", 0) or 0),
        "min_date": fallback_point["date"] if fallback_point else None,
        "max_date": fallback_point["date"] if fallback_point else None,
    }


def _company_candidate_keys(raw_company_name: str) -> list[str]:
    keys: list[str] = []
    raw = (raw_company_name or "").strip()
    if raw:
        keys.append(raw)
    canon, _country = _canonical_company_name_and_country(raw)
    if canon and canon not in keys:
        keys.append(canon)
    try:
        p = _resolve_company_file(canon or raw)
        if p and p.stem and p.stem not in keys:
            keys.append(p.stem)
    except Exception:
        pass
    return keys


def _count_company_schema_sheets(company_name: str) -> int:
    company_file = _resolve_company_file(company_name)
    if not company_file or not company_file.exists():
        return 0
    try:
        wb = load_workbook(company_file, read_only=True, data_only=True, keep_links=False)
        return int(len([s for s in (wb.sheetnames or []) if not _is_hidden_schema_sheet(s)]))
    except Exception:
        return 0


HIDDEN_SCHEMA_SHEETS = {"readme", "company information"}


def _is_hidden_schema_sheet(sheet_name: str) -> bool:
    return (sheet_name or "").strip().lower() in HIDDEN_SCHEMA_SHEETS


def _infer_column_rule(header: str) -> dict[str, str]:
    raw = (header or "").strip()
    low = raw.lower()
    compact = "".join(ch for ch in low if ch.isalnum())

    date_markers = (
        "reportingperiod",
        "release date",
        "attachment date",
        "start date",
        "end date",
        "travel date",
        "date duration",
    )
    number_keywords = (
        "spend",
        "amount",
        "consumption",
        "activity volume",
        "activity amount",
        "fuel consumption",
        "distance",
        "km",
        "kwh",
        "mwh",
        "litre",
        "liter",
        "volume",
        "quantity",
        "qty",
        "weight",
        "ton",
        "tonne",
        "hours",
        "days",
        "months",
        "employee count",
        "headcount",
        "ef_value",
    )
    text_exclusions = ("id", "unit", "currency", "supplier", "source", "description", "name", "country")

    if any(marker in low for marker in date_markers) or compact == "date":
        return {"type": "date", "format": "YYYY-MM-DD", "placeholder": "YYYY-MM-DD"}

    if any(ex in low for ex in text_exclusions):
        return {"type": "text"}

    if any(key in low for key in number_keywords):
        return {"type": "number", "format": "decimal", "placeholder": "0"}

    return {"type": "text"}


def _normalize_date_like(value: str) -> str | None:
    s = (value or "").strip()
    if not s:
        return ""
    try:
        dt = pd.to_datetime([s], errors="coerce", dayfirst=False)[0]
        if pd.isna(dt):
            return None
        return dt.strftime("%Y-%m-%d")
    except Exception:
        return None


def _normalize_number_like(value: str) -> str | None:
    s = (value or "").strip()
    if not s:
        return ""
    f = _parse_float_loose(s)
    if f is None:
        return None
    return str(f)


def _validate_and_normalize_rows(headers: list[str], rows: list[list[str]]) -> tuple[list[list[str]], list[str]]:
    rules = [_infer_column_rule(h) for h in headers]
    out: list[list[str]] = []
    errors: list[str] = []
    for ridx, row in enumerate(rows, start=1):
        normalized: list[str] = []
        for cidx, value in enumerate(row):
            rule = rules[cidx] if cidx < len(rules) else {"type": "text"}
            text = "" if value is None else str(value).strip()
            if rule.get("type") == "date":
                norm = _normalize_date_like(text)
                if norm is None:
                    errors.append(f"Row {ridx}, column '{headers[cidx]}': expected date format YYYY-MM-DD")
                    normalized.append(text)
                else:
                    normalized.append(norm)
            elif rule.get("type") == "number":
                norm = _normalize_number_like(text)
                if norm is None:
                    errors.append(f"Row {ridx}, column '{headers[cidx]}': expected a numeric value")
                    normalized.append(text)
                else:
                    normalized.append(norm)
            else:
                normalized.append(text)
        out.append(normalized)
    return out, errors


def _latest_sheet_totals_for_company(company_keys: list[str]) -> list[MappingRunSummary]:
    """
    For a company (possibly multiple aliases), return the latest summary per sheet.
    """
    if not company_keys:
        return []
    q = (
        MappingRunSummary.query.filter(MappingRunSummary.company_name.in_(company_keys))
        .order_by(MappingRunSummary.created_at.desc())
        .all()
    )
    seen: set[str] = set()
    out: list[MappingRunSummary] = []
    for r in q:
        k = (r.sheet_name or "").strip().lower()
        if not k or k in seen:
            continue
        seen.add(k)
        out.append(r)
    return out


def _backfill_mapping_summaries(max_runs: int = 200) -> None:
    return    """
    Best-effort backfill for existing MappingRun rows created before summaries existed.
    Safe to call frequently; only fills missing summaries.
    """
    try:
        _ensure_db_tables()
        existing = {r.run_id for r in MappingRunSummary.query.with_entities(MappingRunSummary.run_id).all()}
        runs = (
            MappingRun.query.filter_by(status="succeeded")
            .order_by(MappingRun.created_at.desc())
            .limit(int(max_runs))
            .all()
        )
        changed = False
        for mr in runs:
            rid = getattr(mr, "id", None)
            if not rid:
                continue
            summ_existing = MappingRunSummary.query.filter_by(run_id=rid).first()
            # If a summary exists but is zero, re-evaluate once (common when column name differed like 'co2e')
            if summ_existing and float(getattr(summ_existing, "tco2e_total", 0.0) or 0.0) != 0.0:
                continue
            p = getattr(mr, "output_path", None)
            if not p or not os.path.exists(str(p)):
                continue
            total_tco2e, rows_count, _used_col = _sum_tco2e_from_xlsx(p, getattr(mr, "sheet_name", None))
            if _used_col is None:
                # Fallback: try pandas if workbook read fails
                try:
                    sheets = pd.read_excel(p, sheet_name=None, engine="openpyxl")
                except Exception:
                    continue
                mapped_df = None
                want = (getattr(mr, "sheet_name", "") or "").strip().lower()
                if want:
                    for k, v in sheets.items():
                        if str(k).strip().lower() == want:
                            mapped_df = v
                            break
                if mapped_df is None:
                    mapped_df = next(iter(sheets.values())) if sheets else None
                if mapped_df is None:
                    continue
                total_tco2e, rows_count, _used_col = _sum_tco2e(mapped_df)
            scope = _infer_scope_from_sheet(getattr(mr, "sheet_name", "") or "")
            if summ_existing:
                summ_existing.company_name = getattr(mr, "company_name", "") or summ_existing.company_name
                summ_existing.sheet_name = getattr(mr, "sheet_name", "") or summ_existing.sheet_name
                summ_existing.scope = scope
                summ_existing.tco2e_total = float(total_tco2e or 0.0)
                summ_existing.rows_count = int(rows_count or 0)
                if not summ_existing.created_at:
                    summ_existing.created_at = getattr(mr, "created_at", None) or datetime.utcnow()
                changed = True
            else:
                summ = MappingRunSummary(
                    run_id=rid,
                    company_name=getattr(mr, "company_name", "") or "",
                    sheet_name=getattr(mr, "sheet_name", "") or "",
                    scope=scope,
                    tco2e_total=float(total_tco2e or 0.0),
                    rows_count=int(rows_count or 0),
                    created_at=getattr(mr, "created_at", None) or datetime.utcnow(),
                )
                db.session.add(summ)
                existing.add(rid)
                changed = True
        if changed:
            db.session.commit()
    except Exception:
        try:
            db.session.rollback()
        except Exception:
            pass

@login_manager.user_loader
def load_user(user_id):
    return db.session.get(User, int(user_id))


def _norm_name(s: str) -> str:
    return "".join(ch.lower() for ch in (s or "") if ch.isalnum())


def _list_stage1_input_files() -> list[Path]:
    out: list[Path] = []
    try:
        for p in sorted(STAGE1_INPUT_DIR.glob("*.xls*"), key=lambda x: x.name.lower()):
            if p.name.startswith("~$"):
                continue
            if p.suffix.lower() not in (".xlsx", ".xlsm"):
                continue
            out.append(p)
    except Exception:
        out = []
    return out


def _company_file_map() -> dict[str, Path]:
    """
    Map a stable company key -> input Excel path.
    Key is the filename stem (e.g. "BIMMS" from "BIMMS.xlsx").
    """
    m: dict[str, Path] = {}
    for p in _list_stage1_input_files():
        m[p.stem] = p
    return m


def _company_canonical_file_map() -> dict[str, Path]:
    """
    Map canonical company name -> input Excel path.
    If multiple files map to the same canonical name, keep the first.
    """
    out: dict[str, Path] = {}
    for stem, p in _company_file_map().items():
        canon, _country = _canonical_company_name_and_country(stem)
        key = canon or stem
        if key not in out:
            out[key] = p
    return out


def _resolve_company_file(company_key: str) -> Path | None:
    company_key = (company_key or "").strip()
    if not company_key:
        return None

    # First: canonical map (supports web display names)
    canon, _country = _canonical_company_name_and_country(company_key)
    cm = _company_canonical_file_map()
    if canon and canon in cm:
        return cm[canon]

    m = _company_file_map()
    if company_key in m:
        return m[company_key]

    want = _norm_name(company_key)
    if not want:
        return None

    # Best-effort normalized match (handles spaces/dashes differences)
    for k, p in m.items():
        if _norm_name(k) == want:
            return p

    return None


def _detect_header_row_and_headers(ws, max_scan_rows: int = 20) -> tuple[int, list[str]]:
    """
    Return (header_row_index_1based, headers[]).
    Scans first N rows and picks the earliest row with the highest number of non-empty cells (>=2).
    This helps when sheets have a few instruction rows before the actual header row.
    """
    best_row_idx = 1
    best_values = None
    best_score = -1

    max_r = min(int(getattr(ws, "max_row", 1) or 1), max_scan_rows)
    max_c = int(getattr(ws, "max_column", 1) or 1)
    max_c = min(max_c, 200)  # guard

    for r in range(1, max_r + 1):
        values = []
        score = 0
        last_non_empty = 0
        for c in range(1, max_c + 1):
            v = ws.cell(row=r, column=c).value
            s = "" if v is None else str(v).strip()
            values.append(s)
            if s:
                score += 1
                last_non_empty = c
        if score >= 2 and score > best_score:
            best_score = score
            best_row_idx = r
            best_values = values[:last_non_empty] if last_non_empty else values

    if best_values is None:
        # Fallback: use first row up to last non-empty cell
        r = 1
        last_non_empty = 0
        values = []
        for c in range(1, max_c + 1):
            v = ws.cell(row=r, column=c).value
            s = "" if v is None else str(v).strip()
            values.append(s)
            if s:
                last_non_empty = c
        best_values = values[:last_non_empty] if last_non_empty else values

    headers: list[str] = []
    for i, h in enumerate(best_values, start=1):
        hh = (h or "").strip()
        headers.append(hh if hh else f"Column {i}")

    return best_row_idx, headers


def _read_header_row_raw(ws, header_row: int, max_columns: int = 200) -> list[str]:
    """
    Read header row values as strings, up to the last non-empty cell.
    Does not invent placeholder names.
    """
    max_c = int(getattr(ws, "max_column", 1) or 1)
    max_c = min(max_c, max_columns)
    values: list[str] = []
    last_non_empty = 0
    for c in range(1, max_c + 1):
        v = ws.cell(row=header_row, column=c).value
        s = "" if v is None else str(v).strip()
        values.append(s)
        if s:
            last_non_empty = c
    return values[:last_non_empty] if last_non_empty else values


def _write_schema_only_workbook(source_file: Path, dest_file: Path) -> None:
    """
    Create a schema-only workbook:
    - same sheet names
    - blank sheet contents except the detected header row (written at the same row index)
    This avoids slow row-deletion on very large historical files.
    """
    src_wb = load_workbook(source_file, read_only=True, data_only=True, keep_links=False)
    out_wb = Workbook()
    try:
        out_wb.remove(out_wb.active)
    except Exception:
        pass

    for name in src_wb.sheetnames:
        src_ws = src_wb[name]
        header_row, _headers = _detect_header_row_and_headers(src_ws)
        raw_headers = _read_header_row_raw(src_ws, header_row)
        ws = out_wb.create_sheet(title=name)
        for j, v in enumerate(raw_headers, start=1):
            vv = (v or "").strip()
            ws.cell(row=header_row, column=j).value = vv if vv else None

    out_wb.save(dest_file)


def _ensure_company_backup_and_initialize(company_file: Path) -> None:
    """
    One-time operation per company file:
    - Create a backup of the original (historical 2025 filled) workbook
    - Clear all data rows below detected headers (keeps instruction rows + header row)
    This ensures Stage1 pipeline reads only newly submitted web rows.
    """
    backup_path = STAGE1_INPUT_BACKUP_DIR / company_file.name
    if not backup_path.exists():
        try:
            shutil.copy2(company_file, backup_path)
        except Exception:
            # If backup fails, do not proceed with destructive changes.
            raise

        tmp_path = company_file.with_suffix(".webtmp.xlsx")
        try:
            if tmp_path.exists():
                tmp_path.unlink()
        except Exception:
            pass

        _write_schema_only_workbook(backup_path, tmp_path)
        os.replace(str(tmp_path), str(company_file))


_EF_CACHE: dict[str, object] = {"mtime_ns": None, "rows": None, "sheets": None, "scopes": None, "sources": None}
_SCHEMA_CACHE_LOCK = threading.Lock()
_SCHEMA_CACHE: dict[tuple[str, int | None], dict[str, object]] = {}


def _schema_cache_key(company_file: Path) -> tuple[str, int | None]:
    try:
        st = company_file.stat()
        mtime_ns = getattr(st, "st_mtime_ns", None) or int(st.st_mtime * 1_000_000_000)
    except Exception:
        mtime_ns = None
    return (str(company_file.resolve()), mtime_ns)


def _invalidate_schema_cache(company_file: Path | None = None) -> None:
    global _SCHEMA_CACHE
    with _SCHEMA_CACHE_LOCK:
        if company_file is None:
            _SCHEMA_CACHE = {}
            return
        target = str(company_file.resolve())
        _SCHEMA_CACHE = {k: v for k, v in _SCHEMA_CACHE.items() if k[0] != target}


def _get_schema_cache_entry(company_file: Path) -> dict[str, object]:
    key = _schema_cache_key(company_file)
    with _SCHEMA_CACHE_LOCK:
        cached = _SCHEMA_CACHE.get(key)
        if cached is None:
            cached = {"sheets": None, "headers": {}}
            _SCHEMA_CACHE[key] = cached
        return cached


def _get_visible_sheet_names(company_file: Path) -> list[str]:
    cache_entry = _get_schema_cache_entry(company_file)
    cached_sheets = cache_entry.get("sheets")
    if isinstance(cached_sheets, list):
        return cached_sheets

    wb = load_workbook(company_file, read_only=True, data_only=True, keep_links=False)
    sheets = [s for s in list(wb.sheetnames) if not _is_hidden_schema_sheet(s)]
    cache_entry["sheets"] = sheets
    return sheets


def _get_sheet_headers_and_rules(company_file: Path, sheet: str) -> tuple[int, list[str], dict[str, str]]:
    cache_entry = _get_schema_cache_entry(company_file)
    headers_cache = cache_entry.setdefault("headers", {})
    if isinstance(headers_cache, dict) and sheet in headers_cache:
        cached = headers_cache[sheet]
        if isinstance(cached, tuple) and len(cached) == 3:
            return cached  # type: ignore[return-value]

    wb = load_workbook(company_file, read_only=True, data_only=True, keep_links=False)
    if sheet not in wb.sheetnames:
        raise KeyError(sheet)
    ws = wb[sheet]
    header_row, headers = _detect_header_row_and_headers(ws)
    rules = {h: _infer_column_rule(h) for h in headers}
    result = (header_row, headers, rules)
    if isinstance(headers_cache, dict):
        headers_cache[sheet] = result
    return result


def _load_stage2_emission_factors() -> dict[str, object]:
    """
    Load mapping emission factors from the Stage2 Excel file (each sheet = category).
    Returns a dict with:
      - rows: list[dict] where keys match the mapping headers
      - sheets: list[str]
      - scopes: list[str]
      - sources: list[str]
    """
    global _EF_CACHE
    try:
        st = STAGE2_EF_XLSX.stat()
        mtime_ns = getattr(st, "st_mtime_ns", None) or int(st.st_mtime * 1_000_000_000)
    except Exception:
        mtime_ns = None

    if _EF_CACHE.get("rows") is not None and _EF_CACHE.get("mtime_ns") == mtime_ns:
        return _EF_CACHE  # type: ignore[return-value]

    rows: list[dict[str, object]] = []
    sheets: list[str] = []
    scopes_set: set[str] = set()
    sources_set: set[str] = set()

    if not STAGE2_EF_XLSX.exists():
        _EF_CACHE = {"mtime_ns": mtime_ns, "rows": [], "sheets": [], "scopes": [], "sources": []}
        return _EF_CACHE  # type: ignore[return-value]

    wb = load_workbook(STAGE2_EF_XLSX, read_only=True, data_only=True, keep_links=False)
    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        sheets.append(sheet_name)

        # Read headers from first row (fast in read_only mode)
        first = next(ws.iter_rows(min_row=1, max_row=1, values_only=True), None)
        headers = [("" if v is None else str(v).strip()) for v in (first or [])]
        while headers and headers[-1] == "":
            headers.pop()

        if not headers:
            continue

        # Expect the mapping headers; tolerate extra columns
        wanted = {
            "ef_name",
            "ef_description",
            "scope",
            "ef_category",
            "ef_id",
            "ef_value",
            "ef_unit",
            "ef_source",
            "Emission Factor Category",
        }

        col_idx = {h: i for i, h in enumerate(headers) if h}

        # Stream rows efficiently
        for row_idx, row_values in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
            values = list(row_values[: len(headers)]) if row_values else []
            if not values:
                continue

            any_val = False
            for v in values:
                if v is None:
                    continue
                if str(v).strip() != "":
                    any_val = True
                    break
            if not any_val:
                continue

            row: dict[str, object] = {"sheet": sheet_name, "_row": row_idx}
            for h, idx in col_idx.items():
                if h in wanted:
                    row[h] = values[idx] if idx < len(values) else None

            # Normalize some fields to strings for filtering/search
            for k in ("ef_name", "ef_description", "scope", "ef_category", "ef_id", "ef_unit", "ef_source", "Emission Factor Category"):
                if k in row and row[k] is not None:
                    row[k] = str(row[k]).strip()

            if row.get("scope"):
                scopes_set.add(str(row["scope"]))
            if row.get("ef_source"):
                sources_set.add(str(row["ef_source"]))

            rows.append(row)

    scopes = sorted([s for s in scopes_set if s], key=lambda x: x.lower())
    sources = sorted([s for s in sources_set if s], key=lambda x: x.lower())
    sheets_sorted = sorted([s for s in sheets if s], key=lambda x: x.lower())

    _EF_CACHE = {"mtime_ns": mtime_ns, "rows": rows, "sheets": sheets_sorted, "scopes": scopes, "sources": sources}
    return _EF_CACHE  # type: ignore[return-value]


def list_pipeline_templates_for_user(user_company: str, is_admin: bool):
    files = []
    try:
        for p in sorted(PIPELINE_TEMPLATE_DIR.glob("*.xls*"), key=lambda x: x.name.lower()):
            if p.name.startswith("~$"):
                continue
            files.append(p)
    except Exception:
        files = []

    if is_admin:
        return files

    want = _norm_name(user_company)
    if not want:
        return []
    # Match by filename containing normalized company name (best-effort)
    out = []
    for p in files:
        if want in _norm_name(p.stem):
            out.append(p)
    return out


def _safe_user_can_access_run(run: "PipelineRun") -> bool:
    return bool(current_user.is_admin) or (run.user_id == current_user.id)


def _run_pipeline_background(run_id: int) -> None:
    with app.app_context():
        run = PipelineRun.query.get(run_id)
        if not run:
            return

        run.status = "running"
        db.session.commit()

        run_dir = Path(run.run_dir)
        input_dir = Path(run.input_dir)
        log_path = run_dir / "pipeline.log"
        run.log_path = str(log_path)
        db.session.commit()

        stage2_out_dir = PROJECT_ROOT / "engine" / "stage2_mapping" / "output"
        before = set()
        try:
            if stage2_out_dir.exists():
                before = {p.resolve() for p in stage2_out_dir.glob("*.xlsx")}
        except Exception:
            before = set()

        cmd = [
            sys.executable,
            str(PROJECT_ROOT / "run_pipeline.py"),
            "all",
            "--stage1-input-folder",
            str(input_dir),
            "--stage1-work-dir",
            str(run_dir / "stage1"),
        ]

        rc = 1
        try:
            run_dir.mkdir(parents=True, exist_ok=True)
            with open(log_path, "w", encoding="utf-8") as f:
                f.write("CMD: " + " ".join(cmd) + "\n")
                f.write("START: " + datetime.now().isoformat() + "\n\n")
                f.flush()
                proc = subprocess.run(
                    cmd,
                    cwd=str(PROJECT_ROOT),
                    stdout=f,
                    stderr=subprocess.STDOUT,
                    text=True,
                    env={
                        **os.environ,
                        "PYTHONUTF8": os.environ.get("PYTHONUTF8", "1"),
                        "PYTHONIOENCODING": os.environ.get("PYTHONIOENCODING", "utf-8"),
                        "PYTHONUNBUFFERED": os.environ.get("PYTHONUNBUFFERED", "1"),
                    },
                )
                rc = int(proc.returncode)
        except Exception as exc:
            run.status = "failed"
            run.exit_code = -1
            run.error_message = str(exc)
            db.session.commit()
            return

        run.exit_code = rc

        # Stage1 output: last chained file (if present)
        try:
            stage1_final = run_dir / "stage1" / "stage1_05_translated.xlsx"
            if stage1_final.exists():
                run.stage1_output = str(stage1_final)
        except Exception:
            pass

        # Collect new Stage2 outputs and copy into run dir
        try:
            after = set()
            if stage2_out_dir.exists():
                after = {p.resolve() for p in stage2_out_dir.glob("*.xlsx")}
            new_files = sorted([p for p in after if p not in before], key=lambda p: p.stat().st_mtime)
            if new_files:
                dst = run_dir / "stage2_output"
                dst.mkdir(parents=True, exist_ok=True)
                for p in new_files:
                    shutil.copy2(p, dst / p.name)
                run.stage2_output_dir = str(dst)
        except Exception:
            pass

        run.status = "succeeded" if rc == 0 else "failed"
        db.session.commit()

# Company list
COMPANIES = [
    "Navitas (Portugal, Nordics) -Portugal",
    "Navitas (Portugal, Nordics) - Norway and Finland",
    "Caerus Architects",
    "SD Nordics",
    "CTS-VDC",
    "CTS Nordics AS Norway",
    "CTS Sweden",
    "CTS Denmark",
    "QEC",
    "CTS Finland OY",
    "Velox",
    "Mecwide",
    "Porvelox",
    "DC Piping",
    "MC Prefab",
    "Gapit Nordics",
    "Nordic EPOD",
    "CTS EU Portugal (CTS Nordics Eng)",
    "BIMMS",
    "Commissioning Services",
    "NEP Switchboards AS Norway"
]

# Utility: Calculate emissions from excel data (veritabanından faktör çeker)
def calculate_emissions_from_excel(file_path, template_name):
    from sqlalchemy import and_
    total_emission = 0
    by_category = {}
    try:
        xls = pd.ExcelFile(file_path)
        # Önce 'Activity Based' sheet var mı kontrol et
        if 'Activity Based' in xls.sheet_names:
            sheet = 'Activity Based'
        else:
            sheet = 0  # ilk sheet
        header_row = 0
        if template_name == 'Scope 1 Fuel Usage (Mobile Combustion)':
            header_row = 4
        elif template_name == 'Scope 2 Electricity':
            header_row = 0
        elif template_name == 'Scope 2 District Heating':
            header_row = 0
        elif template_name == 'Scope 3 Category 7 Employee Commuting':
            header_row = 0
        elif template_name.startswith('Scope 3'):
            header_row = 0
        elif template_name == 'Water Tracker':
            header_row = 0
        else:
            header_row = 0
        df = xls.parse(sheet, header=header_row)
        if not isinstance(df, pd.DataFrame):
            return {
                'total_emission': 0,
                'by_category': {}
            }
        df.columns = [str(c).strip().lower() for c in df.columns]
        # Scope 1 Fuel Usage
        if template_name == 'Scope 1 Fuel Usage (Mobile Combustion)':
            fuel_col = next((c for c in df.columns if 'fuel' in c), None)
            litre_col = next((c for c in df.columns if 'litre' in c), None)
            if fuel_col and litre_col:
                for _, row in df.iterrows():
                    fuel = str(row[fuel_col]).strip()
                    try:
                        litres = float(row[litre_col])
                    except (ValueError, TypeError):
                        litres = 0
                    # Veritabanından faktör çek
                    factor_obj = EmissionFactor.query.filter_by(category=template_name, subcategory=fuel).first()
                    factor = factor_obj.factor if factor_obj else 0
                    emission = litres * factor
                    if fuel not in by_category:
                        by_category[fuel] = 0
                    by_category[fuel] += emission
                    total_emission += emission
        # Scope 2 Electricity
        elif template_name == 'Scope 2 Electricity':
            cons_col = next((c for c in df.columns if 'consumption' in c or 'kwh' in c), None)
            factor_col = next((c for c in df.columns if 'emission factor' in c), None)
            if cons_col:
                for _, row in df.iterrows():
                    try:
                        cons = float(row[cons_col])
                    except (ValueError, TypeError):
                        cons = 0
                    if factor_col and row.get(factor_col):
                        factor = float(row[factor_col])
                    else:
                        factor_obj = EmissionFactor.query.filter_by(category=template_name, subcategory='Electricity').first()
                        factor = factor_obj.factor if factor_obj else 0
                    emission = cons * factor
                    by_category['Electricity'] = by_category.get('Electricity', 0) + emission
                    total_emission += emission
        # Scope 2 District Heating
        elif template_name == 'Scope 2 District Heating':
            cons_col = next((c for c in df.columns if 'consumption' in c), None)
            factor_col = next((c for c in df.columns if 'emission factor' in c), None)
            if cons_col:
                for _, row in df.iterrows():
                    try:
                        cons = float(row[cons_col])
                    except (ValueError, TypeError):
                        cons = 0
                    if factor_col and row.get(factor_col):
                        factor = float(row[factor_col])
                    else:
                        factor_obj = EmissionFactor.query.filter_by(category=template_name, subcategory='District Heating').first()
                        factor = factor_obj.factor if factor_obj else 0
                    emission = cons * factor
                    by_category['District Heating'] = by_category.get('District Heating', 0) + emission
                    total_emission += emission
        # Scope 3 Category 7 Employee Commuting
        elif template_name == 'Scope 3 Category 7 Employee Commuting':
            mode_col = next((c for c in df.columns if 'mode of transport' in c), None)
            km_col = next((c for c in df.columns if 'km travelled per month' in c), None)
            if mode_col and km_col:
                for _, row in df.iterrows():
                    mode = str(row[mode_col]).strip().lower()
                    try:
                        km = float(row[km_col])
                    except (ValueError, TypeError):
                        km = 0
                    factor_obj = EmissionFactor.query.filter_by(category=template_name, subcategory=mode).first()
                    factor = factor_obj.factor if factor_obj else 0
                    emission = km * factor
                    if mode not in by_category:
                        by_category[mode] = 0
                    by_category[mode] += emission
                    total_emission += emission
        # Scope 3 (Spend-based veya Activity-based)
        elif template_name.startswith('Scope 3'):
            spend_col = next((c for c in df.columns if 'spend' in c), None)
            factor_col = next((c for c in df.columns if 'emission factor' in c), None)
            cons_col = next((c for c in df.columns if 'consumption' in c or 'amount' in c), None)
            if spend_col and factor_col:
                for _, row in df.iterrows():
                    try:
                        spend = float(row[spend_col])
                        factor = float(row[factor_col])
                        emission = spend * factor
                        by_category['Spend-based'] = by_category.get('Spend-based', 0) + emission
                        total_emission += emission
                    except (ValueError, TypeError):
                        continue
            elif cons_col and factor_col:
                for _, row in df.iterrows():
                    try:
                        cons = float(row[cons_col])
                        factor = float(row[factor_col])
                        emission = cons * factor
                        by_category['Activity-based'] = by_category.get('Activity-based', 0) + emission
                        total_emission += emission
                    except (ValueError, TypeError):
                        continue
        # Water Tracker
        elif template_name == 'Water Tracker':
            cons_col = next((c for c in df.columns if 'consumption' in c or 'amount' in c), None)
            factor_col = next((c for c in df.columns if 'emission factor' in c), None)
            if cons_col:
                for _, row in df.iterrows():
                    try:
                        cons = float(row[cons_col])
                    except (ValueError, TypeError):
                        cons = 0
                    if factor_col and row.get(factor_col):
                        factor = float(row[factor_col])
                    else:
                        factor_obj = EmissionFactor.query.filter_by(category=template_name, subcategory='Water').first()
                        factor = factor_obj.factor if factor_obj else 0
                    emission = cons * factor
                    by_category['Water'] = by_category.get('Water', 0) + emission
                    total_emission += emission
        else:
            # Diğer template'ler için dummy değer
            return {
                'total_emission': 1234.56,
                'by_category': {
                    'Dummy': 1234.56
                }
            }
    except Exception as e:
        print(f'Excel emission calculation error: {e}')
        return {'total_emission': 0, 'by_category': {}}
    return {'total_emission': round(total_emission, 2), 'by_category': {k: round(v, 2) for k, v in by_category.items()}}

def _safe_float(v) -> float:
    try:
        out = float(v or 0)
        if math.isnan(out) or math.isinf(out):
            return 0.0
        return out
    except Exception:
        return 0.0

def _infer_scopes(template_name: str, total_emission: float) -> tuple[float, float, float]:
    """Infer scope1/2/3 totals for a single submission based on template name."""
    name = (template_name or "").lower()
    s1 = s2 = s3 = 0.0
    if "scope 1" in name:
        s1 = total_emission
    elif "scope 2" in name:
        s2 = total_emission
    elif "scope 3" in name:
        s3 = total_emission
    return s1, s2, s3

def _parse_date_yyyy_mm_dd(s: str | None) -> datetime | None:
    s = (s or "").strip()
    if not s:
        return None
    try:
        return datetime.strptime(s, "%Y-%m-%d")
    except Exception:
        return None

def get_emission_factor_name(extra_data_json):
    if not extra_data_json or extra_data_json == 'null':
        return ''
    try:
        data = json.loads(extra_data_json) if isinstance(extra_data_json, str) else extra_data_json
        for key in data:
            if key.strip().lower() == 'item prettyid':
                return data[key]
        return ''
    except Exception:
        return ''

app.jinja_env.filters['get_emission_factor_name'] = get_emission_factor_name

@app.route('/')
def index():
    return render_template('index.html')

@app.route('/assets/mouse-symbol')
def mouse_symbol_image():
    symbol_path = APP_DIR / "images" / "Symbol for mouse.png"
    if not symbol_path.exists():
        return ("Mouse symbol image not found.", 404)
    return send_file(str(symbol_path), mimetype="image/png")

@app.route('/login', methods=['GET', 'POST'])
def login():
    if request.method == 'POST':
        email = request.form['email']
        password = request.form['password']
        
        user = User.query.filter_by(email=email).first()
        if user and check_password_hash(user.password_hash, password):
            login_user(user)
            return redirect(url_for('dashboard'))
        else:
            flash('Invalid email or password')
    
    return render_template('login.html')

@app.route('/register', methods=['GET', 'POST'])
def register():
    if request.method == 'POST':
        email = request.form['email']
        password = request.form['password']
        company_name = request.form['company_name']
        
        if User.query.filter_by(email=email).first():
            flash('Email already registered')
            return render_template('register.html', companies=COMPANIES)
        
        user = User(
            email=email,
            password_hash=generate_password_hash(password),
            company_name=company_name
        )
        db.session.add(user)
        db.session.commit()
        
        flash('Registration successful! Please login.')
        return redirect(url_for('login'))  
    return render_template('register.html', companies=COMPANIES)




@app.route('/dashboard')
@login_required
def dashboard():
    _ensure_db_tables()
    if current_user.is_admin:
        #_backfill_mapping_summaries()
        # Admin analytics dashboard (mapping-summary based)
        company_filter = request.args.get('company', '').strip().lower()
        template_filter = request.args.get('template', '').strip().lower()
        include_categories = (request.args.get('include_categories', '1').strip() != '0')
        all_time = (request.args.get('all_time', '').strip() in ('1', 'true', 'on', 'yes'))
        date_from = _parse_date_yyyy_mm_dd(request.args.get('from'))
        date_to = _parse_date_yyyy_mm_dd(request.args.get('to'))

        # Default window: last 365 days (unless all_time is explicitly requested)
        if not all_time and not date_from and not date_to:
            date_to = datetime.now()
            date_from = date_to - timedelta(days=365)

        q = MappingRunSummary.query.order_by(MappingRunSummary.created_at.desc())
        if company_filter:
            q = q.filter(MappingRunSummary.company_name.ilike(f"%{company_filter}%"))
        if template_filter:
            q = q.filter(MappingRunSummary.sheet_name.ilike(f"%{template_filter}%"))

        summaries = q.all()
        # Only count the latest run per company+sheet
        latest_summaries: list[MappingRunSummary] = []
        seen_latest: set[tuple[str, str]] = set()
        for s in summaries:
            key = ((s.company_name or "").strip().lower(), (s.sheet_name or "").strip().lower())
            if not key[0] or not key[1] or key in seen_latest:
                continue
            seen_latest.add(key)
            latest_summaries.append(s)

        company_totals = defaultdict(lambda: {"total": 0.0, "scope1": 0.0, "scope2": 0.0, "scope3": 0.0, "count": 0})
        month_totals = defaultdict(float)  # YYYY-MM -> total
        year_totals = defaultdict(float)   # YYYY -> total
        category_totals = defaultdict(float)
        run_cache: dict[str, MappingRun | None] = {}

        grand = {"total": 0.0, "scope1": 0.0, "scope2": 0.0, "scope3": 0.0, "count": 0}
        for sub in latest_summaries:
            profile = _period_profile_for_summary(sub, run_cache=run_cache)
            points = list(profile.get("points") or [])
            if date_from:
                points = [p for p in points if isinstance(p.get("date"), datetime) and p["date"] >= date_from]
            if date_to:
                points = [p for p in points if isinstance(p.get("date"), datetime) and p["date"] < (date_to + timedelta(days=1))]
            if (date_from or date_to) and not points:
                continue

            total = _safe_float(sum(float(p.get("value") or 0.0) for p in points) if points else profile.get("total", 0.0))
            scope_num = int(getattr(sub, "scope", 0) or 0)
            s1 = total if scope_num == 1 else 0.0
            s2 = total if scope_num == 2 else 0.0
            s3 = total if scope_num == 3 else 0.0

            company = (sub.company_name or "").strip() or "(unknown)"
            company_totals[company]["total"] += total
            company_totals[company]["scope1"] += s1
            company_totals[company]["scope2"] += s2
            company_totals[company]["scope3"] += s3
            company_totals[company]["count"] += 1

            grand["total"] += total
            grand["scope1"] += s1
            grand["scope2"] += s2
            grand["scope3"] += s3
            grand["count"] += 1

            for point in points:
                dt = point.get("date")
                if not isinstance(dt, datetime):
                    continue
                month_key = dt.strftime("%Y-%m")
                year_key = dt.strftime("%Y")
                month_totals[month_key] += float(point.get("value") or 0.0)
                year_totals[year_key] += float(point.get("value") or 0.0)

            if include_categories:
                category_totals[str(sub.sheet_name or "(unknown)")] += total

        company_rows = [
            {
                "company": k,
                "total": round(v["total"], 2),
                "scope1": round(v["scope1"], 2),
                "scope2": round(v["scope2"], 2),
                "scope3": round(v["scope3"], 2),
                "count": v["count"],
            }
            for k, v in company_totals.items()
        ]
        company_rows.sort(key=lambda r: r["total"], reverse=True)

        month_rows = [{"month": k, "total": round(v, 2)} for k, v in month_totals.items()]
        month_rows.sort(key=lambda r: r["month"])

        year_rows = [{"year": k, "total": round(v, 2)} for k, v in year_totals.items()]
        year_rows.sort(key=lambda r: r["year"])

        category_rows = [{"category": k, "total": round(v, 2)} for k, v in category_totals.items()]
        category_rows.sort(key=lambda r: r["total"], reverse=True)

        # Chart payloads (keep top N for readability)
        top_companies = company_rows[:12]
        top_categories = category_rows[:12]
        chart_payload = {
            "companies": {
                "labels": [r["company"] for r in top_companies],
                "values": [r["total"] for r in top_companies],
            },
            "months": {
                "labels": [r["month"] for r in month_rows][-24:],
                "values": [r["total"] for r in month_rows][-24:],
            },
            "years": {
                "labels": [r["year"] for r in year_rows],
                "values": [r["total"] for r in year_rows],
            },
            "categories": {
                "labels": [r["category"] for r in top_categories],
                "values": [r["total"] for r in top_categories],
            },
        }

        return render_template(
            "dashboard_admin.html",
            user=current_user,
            filters={
                "company": request.args.get("company", ""),
                "template": request.args.get("template", ""),
                "from": request.args.get("from", ""),
                "to": request.args.get("to", ""),
                "include_categories": "1" if include_categories else "0",
                "all_time": "1" if all_time else "0",
            },
            grand={
                "total": round(grand["total"], 2),
                "scope1": round(grand["scope1"], 2),
                "scope2": round(grand["scope2"], 2),
                "scope3": round(grand["scope3"], 2),
                "count": grand["count"],
                "companies": len(company_totals),
            },
            company_rows=company_rows,
            month_rows=month_rows,
            year_rows=year_rows,
            category_rows=category_rows,
            chart_payload=chart_payload,
        )

    mapping_runs = (
        MappingRun.query.filter_by(user_id=current_user.id)
        .order_by(MappingRun.created_at.desc())
        .limit(10)
        .all()
    )
    total_mapping_runs = MappingRun.query.filter_by(user_id=current_user.id).count()
    thirty_days_ago = datetime.now() - timedelta(days=30)
    recent_mapping_runs_count = MappingRun.query.filter(
        MappingRun.user_id == current_user.id,
        MappingRun.created_at >= thirty_days_ago
    ).count()

    # Excel schema-driven templates (Stage1 input folder) with canonical display names
    canon_map = _company_canonical_file_map()
    all_companies = [{"key": k, "label": k} for k in sorted(canon_map.keys(), key=lambda x: x.lower())]
    default_company = None

    if current_user.is_admin:
        default_company = (all_companies[0]["key"] if all_companies else None)
        companies = all_companies
    else:
        canon, _country = _canonical_company_name_and_country(current_user.company_name)
        resolved = _resolve_company_file(canon or current_user.company_name)
        if resolved:
            default_company = canon or resolved.stem
            companies = [{"key": default_company, "label": default_company}]
        else:
            companies = []

    return render_template(
        "dashboard.html",
        user=current_user,
        companies=companies,
        default_company=default_company,
        mapping_runs=mapping_runs,
        total_mapping_runs=total_mapping_runs,
        recent_mapping_runs_count=recent_mapping_runs_count,
    )


def _user_can_access_company_file(company_file: Path) -> bool:
    if not company_file:
        return False
    if bool(getattr(current_user, "is_admin", False)):
        return True
    user_file = _resolve_company_file(getattr(current_user, "company_name", "") or "")
    if not user_file:
        return False
    try:
        return user_file.resolve() == company_file.resolve()
    except Exception:
        return str(user_file) == str(company_file)


@app.route("/api/excel_schema/companies", methods=["GET"])
@login_required
def api_excel_schema_companies():
    if current_user.is_admin:
        companies = [{"key": k, "label": k} for k in sorted(_company_canonical_file_map().keys(), key=lambda x: x.lower())]
        return jsonify({"companies": companies})

    canon, _country = _canonical_company_name_and_country(current_user.company_name)
    resolved = _resolve_company_file(canon or current_user.company_name)
    companies = [{"key": (canon or resolved.stem), "label": (canon or resolved.stem)}] if resolved else []
    return jsonify({"companies": companies})


@app.route("/api/excel_schema/sheets", methods=["GET"])
@login_required
def api_excel_schema_sheets():
    company = request.args.get("company", "").strip()
    company_file = _resolve_company_file(company)
    if not company_file or not company_file.exists():
        return jsonify({"error": "Company Excel file not found"}), 404
    if not _user_can_access_company_file(company_file):
        return jsonify({"error": "Access denied"}), 403

    try:
        sheets = _get_visible_sheet_names(company_file)
    except Exception as e:
        return jsonify({"error": f"Failed to read workbook: {e}"}), 400

    return jsonify({"company": company_file.stem, "sheets": sheets})


@app.route("/api/excel_schema/headers", methods=["GET"])
@login_required
def api_excel_schema_headers():
    company = request.args.get("company", "").strip()
    sheet = request.args.get("sheet", "").strip()
    if not sheet:
        return jsonify({"error": "sheet is required"}), 400
    if _is_hidden_schema_sheet(sheet):
        return jsonify({"error": "This sheet is not available for web data entry"}), 403

    company_file = _resolve_company_file(company)
    if not company_file or not company_file.exists():
        return jsonify({"error": "Company Excel file not found"}), 404
    if not _user_can_access_company_file(company_file):
        return jsonify({"error": "Access denied"}), 403

    try:
        try:
            header_row, headers, rules = _get_sheet_headers_and_rules(company_file, sheet)
        except KeyError:
            return jsonify({"error": "Sheet not found"}), 404
    except Exception as e:
        return jsonify({"error": f"Failed to read sheet headers: {e}"}), 400
    return jsonify({"company": company_file.stem, "sheet": sheet, "header_row": header_row, "headers": headers, "rules": rules})


@app.route("/api/excel_schema/download", methods=["GET"])
@login_required
def api_excel_schema_download():
    company = request.args.get("company", "").strip()
    company_file = _resolve_company_file(company)
    if not company_file or not company_file.exists():
        return jsonify({"error": "Company Excel file not found"}), 404
    if not _user_can_access_company_file(company_file):
        return jsonify({"error": "Access denied"}), 403

    return send_file(str(company_file), as_attachment=True, download_name=company_file.name)


@app.route("/api/excel_schema/save", methods=["POST"])
@login_required
def api_excel_schema_save():
    payload = request.get_json(silent=True) or {}
    company = (payload.get("company") or "").strip()
    sheet = (payload.get("sheet") or "").strip()
    rows = payload.get("rows") or []

    if not company or not sheet:
        return jsonify({"error": "company and sheet are required"}), 400
    if not isinstance(rows, list):
        return jsonify({"error": "rows must be a list"}), 400
    if _is_hidden_schema_sheet(sheet):
        return jsonify({"error": "This sheet is not available for web data entry"}), 403

    company_file = _resolve_company_file(company)
    if not company_file or not company_file.exists():
        return jsonify({"error": "Company Excel file not found"}), 404
    if not _user_can_access_company_file(company_file):
        return jsonify({"error": "Access denied"}), 403

    try:
        _ensure_company_backup_and_initialize(company_file)
    except Exception as e:
        return jsonify({"error": f"Backup/initialize failed: {e}"}), 500

    try:
        wb = load_workbook(company_file, keep_links=False)
        if sheet not in wb.sheetnames:
            return jsonify({"error": "Sheet not found"}), 404

        ws = wb[sheet]
        header_row, headers = _detect_header_row_and_headers(ws)

        # Clear existing web-entered rows for this sheet (keep instructions + header row)
        if ws.max_row and ws.max_row > header_row:
            ws.delete_rows(header_row + 1, ws.max_row - header_row)

        cleaned_rows: list[list[str]] = []
        for r in rows:
            if not isinstance(r, list):
                continue
            rr = [("" if v is None else str(v)) for v in r]
            if not any(v.strip() for v in rr):
                continue
            # Fit to header count
            if len(rr) < len(headers):
                rr = rr + [""] * (len(headers) - len(rr))
            elif len(rr) > len(headers):
                rr = rr[: len(headers)]
            cleaned_rows.append(rr)

        cleaned_rows, validation_errors = _validate_and_normalize_rows(headers, cleaned_rows)
        if validation_errors:
            return jsonify({"error": validation_errors[0], "validation_errors": validation_errors[:20]}), 400

        start_row = header_row + 1
        for i, rr in enumerate(cleaned_rows):
            for j, v in enumerate(rr, start=1):
                vv = (v or "").strip()
                ws.cell(row=start_row + i, column=j).value = vv if vv != "" else None

        wb.save(company_file)
        _invalidate_schema_cache(company_file)

    except Exception as e:
        return jsonify({"error": f"Save failed: {e}"}), 500

    return jsonify({"ok": True, "company": company_file.stem, "sheet": sheet, "saved_rows": len(cleaned_rows), "file": str(company_file)})


@app.route("/api/mapping/run", methods=["POST"])
@login_required
def api_mapping_run():
    _ensure_db_tables()
    payload = request.get_json(silent=True) or {}
    company = (payload.get("company") or "").strip()
    sheet = (payload.get("sheet") or "").strip()
    headers = payload.get("headers") or []
    rows = payload.get("rows") or []

    if not company or not sheet:
        return jsonify({"error": "company and sheet are required"}), 400
    if not isinstance(headers, list) or not all(isinstance(h, str) for h in headers):
        return jsonify({"error": "headers must be a list of strings"}), 400
    if not isinstance(rows, list):
        return jsonify({"error": "rows must be a list"}), 400
    if _is_hidden_schema_sheet(sheet):
        return jsonify({"error": "This sheet is not available for web data entry"}), 403

    company_file = _resolve_company_file(company)
    if not company_file or not company_file.exists():
        return jsonify({"error": "Company Excel file not found"}), 404
    if not _user_can_access_company_file(company_file):
        return jsonify({"error": "Access denied"}), 403

    # Build DataFrame
    safe_headers = [h.strip() if h and str(h).strip() else f"Column {i+1}" for i, h in enumerate(headers)]
    cleaned_rows = []
    for r in rows:
        if not isinstance(r, list):
            continue
        rr = [("" if v is None else str(v)) for v in r]
        if not any(v.strip() for v in rr):
            continue
        if len(rr) < len(safe_headers):
            rr = rr + [""] * (len(safe_headers) - len(rr))
        elif len(rr) > len(safe_headers):
            rr = rr[: len(safe_headers)]
        cleaned_rows.append(rr)

    cleaned_rows, validation_errors = _validate_and_normalize_rows(safe_headers, cleaned_rows)
    if validation_errors:
        return jsonify({"error": validation_errors[0], "validation_errors": validation_errors[:20]}), 400

    df = pd.DataFrame(cleaned_rows, columns=safe_headers)

    run_id = uuid.uuid4().hex[:12]
    mr = MappingRun(
        id=run_id,
        user_id=current_user.id,
        company_name=company,
        sheet_name=sheet,
        status="running",
        created_at=datetime.utcnow(),
    )
    try:
        db.session.add(mr)
        db.session.commit()
    except Exception:
        pass

    try:
        mapped_df, out_path, input_path = run_mapping(company, sheet, df)
    except Exception as e:
        try:
            mr.status = "failed"
            mr.error_message = str(e)
            db.session.commit()
        except Exception:
            pass
        return jsonify({"error": f"Mapping failed: {e}"}), 500

    try:
        mr.status = "succeeded"
        mr.output_path = str(out_path)
        mr.input_path = str(input_path)
        db.session.commit()
    except Exception:
        pass

    # Persist summary totals for dashboards
    try:
        total_tco2e, rows_count, used_col = _sum_tco2e_from_xlsx(out_path, sheet)
        if used_col is None:
            total_tco2e, rows_count, used_col = _sum_tco2e(mapped_df)
        scope = _infer_scope_from_sheet(sheet)
        summ = MappingRunSummary.query.filter_by(run_id=run_id).first()
        if not summ:
            summ = MappingRunSummary(
                run_id=run_id,
                company_name=company,
                sheet_name=sheet,
                created_at=datetime.utcnow(),
            )
            db.session.add(summ)
        summ.company_name = company
        summ.sheet_name = sheet
        summ.scope = scope
        summ.tco2e_total = float(total_tco2e or 0.0)
        summ.rows_count = int(rows_count or 0)
        db.session.commit()
    except Exception:
        pass

    preview = mapped_df.head(40).fillna("").to_dict(orient="records")
    return jsonify(
        {
            "ok": True,
            "run_id": run_id,
            "company": company,
            "sheet": sheet,
            "mapped_columns": list(mapped_df.columns),
            "preview": preview,
            "preview_rows": len(preview),
        }
    )


@app.route("/api/mapping/download/<run_id>", methods=["GET"])
@login_required
def api_mapping_download(run_id: str):
    _ensure_db_tables()
    mr = MappingRun.query.get(run_id)
    if not mr:
        # Fallback to legacy in-memory if present
        meta = _MAPPING_RUNS.get(run_id)
        if not meta:
            flash("Mapping output not found (expired).")
            return redirect(url_for("dashboard"))
        if not bool(getattr(current_user, "is_admin", False)) and meta.get("user_id") != current_user.id:
            flash("Access denied")
            return redirect(url_for("dashboard"))
        p = meta.get("path")
        if not p or not os.path.exists(str(p)):
            flash("Mapping output file not found.")
            return redirect(url_for("dashboard"))
        fn = f"mapped_{meta.get('company')}_{meta.get('sheet')}.xlsx"
        fn = secure_filename(fn) or "mapped_results.xlsx"
        return send_file(str(p), as_attachment=True, download_name=fn)

    if not bool(getattr(current_user, "is_admin", False)) and mr.user_id != current_user.id:
        flash("Access denied")
        return redirect(url_for("dashboard"))
    if not mr.output_path or not os.path.exists(mr.output_path):
        flash("Mapping output file not found.")
        return redirect(url_for("dashboard"))

    fn = secure_filename(f"mapped_{mr.company_name}_{mr.sheet_name}.xlsx") or "mapped_results.xlsx"
    return send_file(str(mr.output_path), as_attachment=True, download_name=fn)


@app.route("/mappings", methods=["GET"])
@login_required
def mapping_runs():
    _ensure_db_tables()
    runs = (
        MappingRun.query.filter_by(user_id=current_user.id)
        .order_by(MappingRun.created_at.desc())
        .limit(200)
        .all()
    )
    return render_template("mapping_runs.html", user=current_user, runs=runs)


@app.route("/admin/mapping_runs", methods=["GET"])
@login_required
def admin_mapping_runs():
    _ensure_db_tables()
    if not bool(getattr(current_user, "is_admin", False)):
        flash("Access denied")
        return redirect(url_for("dashboard"))

    rows = (
        db.session.query(MappingRun, User.email, User.company_name)
        .join(User, MappingRun.user_id == User.id)
        .order_by(MappingRun.created_at.desc())
        .limit(500)
        .all()
    )
    return render_template("mapping_runs_admin.html", user=current_user, rows=rows)


@app.route('/admin', methods=['GET', 'POST'])
@login_required
def admin():
    if not current_user.is_admin:
        flash('Access denied')
        return redirect(url_for('dashboard'))

    _backfill_mapping_summaries()
    users = User.query.all()
    mapping_runs = MappingRun.query.order_by(MappingRun.created_at.desc()).all()

    # Find all months with mapping runs
    all_months = set()
    for run in mapping_runs:
        if run.created_at:
            all_months.add(run.created_at.strftime('%B %Y'))
    if not all_months:
        all_months.add(datetime.now().strftime('%B %Y'))
    available_months = sorted(list(all_months), key=lambda x: datetime.strptime(x, '%B %Y'), reverse=True)
    selected_month = request.args.get('month') or available_months[0]

    # Calculate progress per company based on mapped categories vs available schema sheets
    companies = sorted(_company_canonical_file_map().keys(), key=lambda x: x.lower())
    submission_stats = []
    chart_data = []
    for company in companies:
        expected = _count_company_schema_sheets(company)
        submitted_sheets = set()
        for run in mapping_runs:
            if (run.company_name == company and run.created_at and run.created_at.strftime('%B %Y') == selected_month and run.status == 'succeeded'):
                submitted_sheets.add(run.sheet_name)
        submitted = len(submitted_sheets)
        rate = int(round((submitted / expected) * 100)) if expected > 0 else 0
        submission_stats.append({
            'company': company,
            'month': selected_month,
            'submitted': submitted,
            'expected': expected,
            'rate': rate
        })
        chart_data.append({
            'company': company,
            'submitted': submitted,
            'expected': expected,
            'rate': rate
        })

    # Last 7 days mapping count
    seven_days_ago = datetime.now() - timedelta(days=7)
    recent_admin_submissions_count = MappingRun.query.filter(
        MappingRun.created_at >= seven_days_ago
    ).count()

    return render_template(
        'admin.html',
        users=users,
        submissions=[],
        mapping_runs=mapping_runs,
        recent_admin_submissions_count=recent_admin_submissions_count,
        available_months=available_months,
        selected_month=selected_month,
        submission_stats=submission_stats,
        chart_data=chart_data
    )

@app.route('/logout')
@login_required
def logout():
    logout_user()
    return redirect(url_for('index'))

@app.route('/report', methods=['GET'])
@login_required
def report():
    if current_user.is_admin:
        return redirect(url_for('admin_report', **request.args.to_dict(flat=True)))
    _backfill_mapping_summaries()
    template_filter = request.args.get('template', '').strip().lower()
    month_filter = request.args.get('date', '').strip()
    company_keys = _company_candidate_keys(current_user.company_name)
    summaries = (
        MappingRunSummary.query.filter(MappingRunSummary.company_name.in_(company_keys))
        .order_by(MappingRunSummary.created_at.desc())
        .all()
    )
    filtered_summaries = []
    seen = set()
    run_cache: dict[str, MappingRun | None] = {}
    for sub in summaries:
        key = ((sub.company_name or "").strip().lower(), (sub.sheet_name or "").strip().lower())
        if key in seen:
            continue
        seen.add(key)
        match = True
        if template_filter and template_filter not in (sub.sheet_name or "").lower():
            match = False
        profile = _period_profile_for_summary(sub, run_cache=run_cache)
        points = list(profile.get("points") or [])
        if month_filter:
            points = [p for p in points if str(p.get("key") or "").startswith(month_filter)]
            if not points:
                match = False
        if match:
            filtered_summaries.append((sub, points if month_filter else points))
    emission_results = []
    chart_data = []
    for idx, (sub, points) in enumerate(filtered_summaries):
        filtered_points = list(points or [])
        if not filtered_points:
            profile = _period_profile_for_summary(sub, run_cache=run_cache)
            filtered_points = list(profile.get("points") or [])
        total_t = _safe_float(sum(float(p.get("value") or 0.0) for p in filtered_points) if filtered_points else (sub.tco2e_total or 0.0))
        label = str(sub.sheet_name or "Category")
        emission_results.append({
            'template_name': label,
            'period_label': _format_period_label(filtered_points, getattr(sub, "created_at", None)),
            'total_emission': round(total_t, 6),
            'by_category': {label: total_t}
        })
        chart_data.append({
            'bar_id': f'barChart{idx+1}',
            'line_id': f'lineChart{idx+1}',
            'pie_id': f'pieChart{idx+1}',
            'labels': [str(p.get("label") or label) for p in filtered_points] if filtered_points else [label],
            'values': [_safe_float(p.get("value") or 0.0) for p in filtered_points] if filtered_points else [total_t]
        })
    return render_template('report.html', emission_results=emission_results, user=current_user, chart_data=chart_data)

@app.route('/admin_report', methods=['GET'])
@login_required
def admin_report():
    if not current_user.is_admin:
        flash('Access denied')
        return redirect(url_for('dashboard'))
    _backfill_mapping_summaries()
    company_filter = request.args.get('company', '').strip().lower()
    template_filter = request.args.get('template', '').strip().lower()
    month_filter = request.args.get('date', '').strip()
    summaries = MappingRunSummary.query.order_by(MappingRunSummary.created_at.desc()).all()
    filtered_submissions = []
    seen = set()
    run_cache: dict[str, MappingRun | None] = {}
    for sub in summaries:
        key = ((sub.company_name or "").strip().lower(), (sub.sheet_name or "").strip().lower())
        if key in seen:
            continue
        seen.add(key)
        match = True
        if company_filter and company_filter not in (sub.company_name or "").lower():
            match = False
        if template_filter and template_filter not in (sub.sheet_name or "").lower():
            match = False
        profile = _period_profile_for_summary(sub, run_cache=run_cache)
        points = list(profile.get("points") or [])
        if month_filter:
            points = [p for p in points if str(p.get("key") or "").startswith(month_filter)]
            if not points:
                match = False
        if match:
            filtered_submissions.append((sub, points if month_filter else points))
    emission_results = []
    chart_data = []
    for idx, (sub, points) in enumerate(filtered_submissions):
        filtered_points = list(points or [])
        if not filtered_points:
            profile = _period_profile_for_summary(sub, run_cache=run_cache)
            filtered_points = list(profile.get("points") or [])
        total_t = _safe_float(sum(float(p.get("value") or 0.0) for p in filtered_points) if filtered_points else (sub.tco2e_total or 0.0))
        label = str(sub.sheet_name or "Category")
        emission_results.append({
            'company_name': sub.company_name,
            'template_name': label,
            'period_label': _format_period_label(filtered_points, getattr(sub, "created_at", None)),
            'total_emission': round(total_t, 6),
            'by_category': {label: total_t}
        })
        chart_data.append({
            'bar_id': f'barChart{idx+1}',
            'line_id': f'lineChart{idx+1}',
            'pie_id': f'pieChart{idx+1}',
            'labels': [str(p.get("label") or label) for p in filtered_points] if filtered_points else [label],
            'values': [_safe_float(p.get("value") or 0.0) for p in filtered_points] if filtered_points else [total_t]
        })
    return render_template('admin_report.html', emission_results=emission_results, user=current_user, chart_data=chart_data)

@app.route('/admin/emission_factors', methods=['GET', 'POST'])
@login_required
def manage_emission_factors():
    if not current_user.is_admin:
        flash('Access denied')
        return redirect(url_for('dashboard'))

    if request.method == 'POST':
        flash('Emission factors are now sourced from the mapping workbook and are read-only in the UI.')
        return redirect(url_for('manage_emission_factors'))

    data = _load_stage2_emission_factors()
    all_rows: list[dict[str, object]] = list(data.get("rows") or [])
    sheets: list[str] = list(data.get("sheets") or [])
    scopes: list[str] = list(data.get("scopes") or [])
    sources: list[str] = list(data.get("sources") or [])

    # Filters
    search = request.args.get('search', '').strip()
    sheet_filter = request.args.get('sheet', '').strip()
    scope_filter = request.args.get('scope', '').strip()
    source_filter = request.args.get('ef_source', '').strip()

    def _matches(row: dict[str, object]) -> bool:
        if sheet_filter and str(row.get("sheet") or "") != sheet_filter:
            return False
        if scope_filter and str(row.get("scope") or "") != scope_filter:
            return False
        if source_filter and str(row.get("ef_source") or "") != source_filter:
            return False

        if search:
            hay = " ".join(
                str(row.get(k) or "")
                for k in (
                    "sheet",
                    "ef_name",
                    "ef_description",
                    "scope",
                    "ef_category",
                    "ef_id",
                    "ef_value",
                    "ef_unit",
                    "ef_source",
                    "Emission Factor Category",
                )
            ).lower()
            if search.lower() not in hay:
                return False
        return True

    filtered = [r for r in all_rows if _matches(r)]
    filtered.sort(key=lambda r: (str(r.get("sheet") or ""), str(r.get("ef_name") or ""), str(r.get("ef_id") or "")))

    # Pagination (lightweight object compatible with template)
    page = int(request.args.get('page', 1) or 1)
    per_page = 60
    total = len(filtered)
    pages = max(1, (total + per_page - 1) // per_page)
    page = max(1, min(page, pages))
    start = (page - 1) * per_page
    end = start + per_page
    items = filtered[start:end]

    factors = SimpleNamespace(
        items=items,
        page=page,
        pages=pages,
        has_prev=(page > 1),
        has_next=(page < pages),
        prev_num=(page - 1),
        next_num=(page + 1),
        total=total,
    )

    if request.method == 'POST':
        # handled above (read-only)
        pass

    return render_template(
        'emission_factors.html',
        factors=factors,
        search=search,
        sheet=sheet_filter,
        scope=scope_filter,
        ef_source=source_filter,
        available_sheets=sheets,
        available_scopes=scopes,
        available_sources=sources,
        mapping_path=str(STAGE2_EF_XLSX),
    )

@app.route('/admin/emission_factors/delete/<int:factor_id>', methods=['POST'])
@login_required
def delete_emission_factor(factor_id):
    if not current_user.is_admin:
        flash('Access denied')
        return redirect(url_for('dashboard'))
    flash('Emission factors are sourced from the mapping workbook and cannot be deleted here.')
    return redirect(url_for('manage_emission_factors'))

@app.route('/admin/emission_factors/update/<int:factor_id>', methods=['POST'])
@login_required
def update_emission_factor(factor_id):
    if not current_user.is_admin:
        flash('Access denied')
        return redirect(url_for('dashboard'))
    flash('Emission factors are sourced from the mapping workbook and cannot be updated here.')
    return redirect(url_for('manage_emission_factors'))


def _ef_expected_headers() -> list[str]:
    return [
        "ef_name",
        "ef_description",
        "scope",
        "ef_category",
        "ef_id",
        "ef_value",
        "ef_unit",
        "ef_source",
        "Emission Factor Category",
    ]


def _clear_ef_cache() -> None:
    global _EF_CACHE
    _EF_CACHE = {"mtime_ns": None, "rows": None, "sheets": None, "scopes": None, "sources": None}


@app.route("/admin/emission_factors/mapping/reload", methods=["GET"])
@login_required
def reload_emission_factors_mapping():
    if not current_user.is_admin:
        flash("Access denied")
        return redirect(url_for("dashboard"))
    _clear_ef_cache()
    flash("Emission factors reloaded from mapping workbook.")
    # Preserve current filters if present
    return redirect(
        url_for(
            "manage_emission_factors",
            search=request.args.get("search", ""),
            sheet=request.args.get("sheet", ""),
            scope=request.args.get("scope", ""),
            ef_source=request.args.get("ef_source", ""),
            page=request.args.get("page", 1),
        )
    )


def _cleanup_mapping_runs(max_age_minutes: int = 180) -> None:
    now = datetime.now()
    to_del = []
    for run_id, meta in list(_MAPPING_RUNS.items()):
        created = meta.get("created_at")
        if isinstance(created, datetime) and (now - created).total_seconds() > max_age_minutes * 60:
            to_del.append(run_id)
    for rid in to_del:
        try:
            p = _MAPPING_RUNS.get(rid, {}).get("path")
            if p and os.path.exists(str(p)):
                os.remove(str(p))
        except Exception:
            pass
        _MAPPING_RUNS.pop(rid, None)


def _to_numeric_spend(x: object) -> float | None:
    try:
        if x is None or (isinstance(x, float) and pd.isna(x)):  # type: ignore[arg-type]
            return None
        if isinstance(x, (int, float)) and not (isinstance(x, float) and pd.isna(x)):
            return float(x)

        s = str(x).strip().replace(" ", "")
        if s == "":
            return None
        if s.startswith("(") and s.endswith(")"):
            s = "-" + s[1:-1]
        # handle comma/period formats
        if s.count(",") == 1 and s.count(".") > 0 and s.rfind(",") > s.rfind("."):
            s = s.replace(".", "").replace(",", ".")
        elif s.count(".") == 1 and s.count(",") > 0 and s.rfind(".") > s.rfind(","):
            s = s.replace(",", "")
        elif s.count(",") == 1 and s.count(".") == 0:
            s = s.replace(",", ".")
        else:
            s = s.replace(",", "")
        return float(s)
    except Exception:
        return None


_FX_2025_TO_EUR: dict[str, float] = {
    "USD": 0.884969,
    "JPY": 0.005916,
    "BGN": 0.511300,
    "CZK": 0.040506,
    "DKK": 0.133987,
    "GBP": 1.167144,
    "HUF": 0.002514,
    "PLN": 0.235868,
    "RON": 0.198319,
    "SEK": 0.090364,
    "CHF": 1.067201,
    "ISK": 0.006913,
    "NOK": 0.085344,
    "TRY": 0.022313,
    "AUD": 0.570854,
    "BRL": 0.158550,
    "CAD": 0.633422,
    "CNY": 0.123175,
    "HKD": 0.113502,
    "IDR": 0.000054,
    "ILS": 0.256889,
    "INR": 0.010150,
    "KRW": 0.000623,
    "MXN": 0.046146,
    "MYR": 0.206874,
    "NZD": 0.514888,
    "PHP": 0.015390,
    "SGD": 0.677705,
    "THB": 0.026943,
    "ZAR": 0.049557,
    "EUR": 1.0,
    "EURO": 1.0,
    "EUROS": 1.0,
    "€": 1.0,
}


def _normalize_currency(raw: object) -> str | None:
    if raw is None:
        return None
    try:
        if isinstance(raw, float) and pd.isna(raw):
            return None
    except Exception:
        pass
    s = str(raw).upper()
    if "€" in s:
        return "EUR"
    if "£" in s:
        return "GBP"
    m = re.search(r"\b(USD|JPY|BGN|CZK|DKK|GBP|HUF|PLN|RON|SEK|CHF|ISK|NOK|TRY|AUD|BRL|CAD|CNY|HKD|IDR|ILS|INR|KRW|MXN|MYR|NZD|PHP|SGD|THB|ZAR|EUR|EURO|EUROS|€)\b", s)
    return m.group(1) if m else None


def preprocess_for_mapping(company_name: str, sheet_name: str, df: pd.DataFrame) -> pd.DataFrame:
    """
    Web-friendly 'mini Stage1': keep inputs stable for Stage2 mapping without multi-company merge.
    Does NOT change mapping rules; only prepares a single-sheet dataframe.
    """
    df = df.copy()
    df.columns = [str(c).strip() for c in df.columns]

    # Drop fully empty rows
    df = df.dropna(how="all")

    # Minimal whitespace normalization
    for c in df.columns:
        try:
            if pd.api.types.is_object_dtype(df[c]) or pd.api.types.is_string_dtype(df[c]):
                df[c] = df[c].astype("string").str.strip()
        except Exception:
            pass

    canon_company, canon_country = _canonical_company_name_and_country(company_name)
    company_display = canon_company or company_name

    # Compatibility columns expected by some downstream steps
    if "Source_File" not in df.columns:
        df["Source_File"] = f"{company_display}.xlsx"
    if "subsidiary_name" not in df.columns:
        df["subsidiary_name"] = company_display

    # Country is required by several mapping rules (previously derived from "Company Information")
    # Ensure both "Country" and "country" exist for robust header matching.
    if "Country" not in df.columns:
        df["Country"] = ""
    if "country" not in df.columns:
        df["country"] = ""

    if canon_country:
        # Fill only where blank
        try:
            mask = df["Country"].isna() | (df["Country"].astype(str).str.strip() == "")
            df.loc[mask, "Country"] = canon_country
        except Exception:
            df["Country"] = canon_country
        try:
            mask2 = df["country"].isna() | (df["country"].astype(str).str.strip() == "")
            df.loc[mask2, "country"] = canon_country
        except Exception:
            df["country"] = canon_country

    # Add Spend_Euro if possible (mirrors Currency_converter_17Dec behavior at sheet-level)
    if "Spend_Euro" not in df.columns:
        sheet_key = str(sheet_name).strip().lower()

        def find_col_exact(name: str) -> str | None:
            for col in df.columns:
                if str(col).strip().lower() == name.strip().lower():
                    return str(col)
            return None

        # Spend/currency columns with special cases
        if sheet_key == "scope 3 cat 15 pensions":
            spend_col = next((c for c in df.columns if "employer payment to pension provider" in str(c).lower()), None)
            currency_col = find_col_exact("currency")
        elif sheet_key == "scope 3 cat 6 business travel":
            spend_col = find_col_exact("spend")
            currency_col = find_col_exact("spend currency")
        else:
            spend_col = next((c for c in df.columns if "spend" in str(c).lower() and "euro" not in str(c).lower()), None)
            currency_col = next((c for c in df.columns if "currency" in str(c).lower()), None)

        spend_euro_vals = []
        if spend_col and currency_col:
            for _, row in df.iterrows():
                spend = _to_numeric_spend(row.get(spend_col))
                curr_code = _normalize_currency(row.get(currency_col))
                if spend is None or curr_code is None:
                    spend_euro_vals.append(pd.NA)
                    continue
                rate = _FX_2025_TO_EUR.get(curr_code)
                spend_euro_vals.append(spend * rate if rate else pd.NA)
        else:
            spend_euro_vals = [pd.NA] * len(df)
        df["Spend_Euro"] = spend_euro_vals

    return df


def _import_stage2_main_mapping():
    # Lazy import to keep Flask startup fast
    if str(STAGE2_MAPPING_DIR) not in sys.path:
        sys.path.insert(0, str(STAGE2_MAPPING_DIR))
    import importlib
    return importlib.import_module("main_mapping")


def run_mapping(company_name: str, sheet_name: str, df: pd.DataFrame) -> tuple[pd.DataFrame, Path, Path]:
    """
    Run existing Stage2 mapping logic for a single company + single sheet dataframe.
    Mapping logic is NOT modified; we call main_mapping.process_all_sheets() on a temporary workbook.
    Returns (mapped_df, output_workbook_path).
    """
    _cleanup_mapping_runs()

    df_pre = preprocess_for_mapping(company_name, sheet_name, df)

    run_dir = INSTANCE_DIR / "mapping_runs" / f"{datetime.now().strftime('%Y%m%d_%H%M%S')}_{uuid.uuid4().hex[:10]}"
    run_dir.mkdir(parents=True, exist_ok=True)
    input_xlsx = run_dir / f"{secure_filename(company_name)}__{secure_filename(sheet_name)}.xlsx"

    with pd.ExcelWriter(input_xlsx, engine="openpyxl") as writer:
        df_pre.to_excel(writer, sheet_name=str(sheet_name)[:31], index=False)

    mm = _import_stage2_main_mapping()

    # Run Stage2 mapping under a lock (Stage2 writes to a shared output/ directory)
    start_ts = time.time() if "time" in globals() else None
    import time as _time
    start_ts = _time.time()
    with _STAGE2_MAP_LOCK:
        orig = getattr(mm, "INPUT_WORKBOOK_NAME", None)
        try:
            setattr(mm, "INPUT_WORKBOOK_NAME", str(input_xlsx))
            mm.process_all_sheets()
        finally:
            try:
                setattr(mm, "INPUT_WORKBOOK_NAME", orig)
            except Exception:
                pass

    # Pick newest mapped_results output created after we started
    candidates = sorted(
        STAGE2_MAPPING_OUTPUT_DIR.glob("mapped_results*.xlsx"),
        key=lambda p: p.stat().st_mtime,
        reverse=True,
    )
    out_path = None
    for p in candidates:
        try:
            if p.stat().st_mtime >= start_ts - 2:
                out_path = p
                break
        except Exception:
            continue
    if out_path is None and candidates:
        out_path = candidates[0]
    if out_path is None:
        raise RuntimeError("Stage2 mapping did not produce an output workbook.")

    # Copy output to a stable location for this run (Stage2 output dir is shared)
    stable_out = run_dir / "mapped_results.xlsx"
    try:
        shutil.copy2(out_path, stable_out)
        out_path = stable_out
    except Exception:
        pass

    sheets = pd.read_excel(out_path, sheet_name=None, engine="openpyxl")
    # Find the mapped sheet (safe-name truncation may apply)
    mapped_df = None
    for k, v in sheets.items():
        if str(k).strip().lower() == str(sheet_name).strip().lower():
            mapped_df = v
            break
    if mapped_df is None:
        # fallback: return first sheet if only one
        if len(sheets) == 1:
            mapped_df = next(iter(sheets.values()))
        else:
            # best-effort: match by prefix
            want = str(sheet_name).strip().lower()[:20]
            for k, v in sheets.items():
                if str(k).strip().lower().startswith(want):
                    mapped_df = v
                    break
    if mapped_df is None:
        raise RuntimeError("Could not locate mapped sheet in Stage2 output workbook.")

    return mapped_df, out_path, input_xlsx


def _find_ef_row_in_sheet(ws, ef_id: str) -> int | None:
    ef_id = (ef_id or "").strip()
    if not ef_id:
        return None
    headers = [("" if v is None else str(v).strip()) for v in (next(ws.iter_rows(min_row=1, max_row=1, values_only=True), None) or [])]
    while headers and headers[-1] == "":
        headers.pop()
    if "ef_id" not in headers:
        return None
    ef_col = headers.index("ef_id") + 1
    for idx, row_values in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
        v = row_values[ef_col - 1] if row_values and len(row_values) >= ef_col else None
        if v is None:
            continue
        if str(v).strip() == ef_id:
            return idx
    return None


@app.route("/admin/emission_factors/mapping/update", methods=["POST"])
@login_required
def update_emission_factor_mapping():
    if not current_user.is_admin:
        flash("Access denied")
        return redirect(url_for("dashboard"))

    sheet = (request.form.get("sheet") or "").strip()
    original_ef_id = (request.form.get("original_ef_id") or "").strip()
    ef_id = (request.form.get("ef_id") or "").strip()

    if not sheet or not original_ef_id:
        flash("Missing sheet or ef_id")
        return redirect(url_for("manage_emission_factors"))

    payload = {k: (request.form.get(k) or "").strip() for k in _ef_expected_headers()}
    if ef_id:
        payload["ef_id"] = ef_id

    try:
        wb = load_workbook(STAGE2_EF_XLSX, keep_links=False)
        if sheet not in wb.sheetnames:
            flash("Sheet not found in mapping workbook")
            return redirect(url_for("manage_emission_factors"))
        ws = wb[sheet]

        headers = [("" if v is None else str(v).strip()) for v in (next(ws.iter_rows(min_row=1, max_row=1, values_only=True), None) or [])]
        while headers and headers[-1] == "":
            headers.pop()

        # Validate required headers exist
        missing = [h for h in _ef_expected_headers() if h not in headers]
        if missing:
            flash(f"Mapping sheet headers are missing: {', '.join(missing)}")
            return redirect(url_for("manage_emission_factors"))

        row_idx = _find_ef_row_in_sheet(ws, original_ef_id)
        if row_idx is None:
            flash("Could not locate the emission factor row by ef_id")
            return redirect(url_for("manage_emission_factors"))

        for h in _ef_expected_headers():
            col = headers.index(h) + 1
            v = payload.get(h)
            if h == "ef_value":
                # keep numeric if possible
                try:
                    ws.cell(row=row_idx, column=col).value = float(v) if v != "" else None
                except Exception:
                    ws.cell(row=row_idx, column=col).value = v if v != "" else None
            else:
                ws.cell(row=row_idx, column=col).value = v if v != "" else None

        # Backup then save (protect against accidental edits)
        backup_dir = STAGE2_EF_XLSX.parent / "_ef_backups"
        backup_dir.mkdir(parents=True, exist_ok=True)
        ts = datetime.now().strftime("%Y%m%d_%H%M%S")
        backup_path = backup_dir / f"{STAGE2_EF_XLSX.stem}_before_update_{ts}{STAGE2_EF_XLSX.suffix}"
        try:
            shutil.copy2(STAGE2_EF_XLSX, backup_path)
        except Exception:
            pass

        wb.save(STAGE2_EF_XLSX)
        _clear_ef_cache()
        flash("Emission factor updated in mapping workbook.")
    except PermissionError:
        flash("Mapping workbook is open/locked. Close Excel and try again.")
    except Exception as e:
        flash(f"Update failed: {e}")

    return redirect(url_for("manage_emission_factors", search=request.args.get("search",""), sheet=request.args.get("sheet",""), scope=request.args.get("scope",""), ef_source=request.args.get("ef_source","")))


@app.route("/admin/emission_factors/mapping/delete", methods=["POST"])
@login_required
def delete_emission_factor_mapping():
    if not current_user.is_admin:
        flash("Access denied")
        return redirect(url_for("dashboard"))

    sheet = (request.form.get("sheet") or "").strip()
    ef_id = (request.form.get("ef_id") or "").strip()
    if not sheet or not ef_id:
        flash("Missing sheet or ef_id")
        return redirect(url_for("manage_emission_factors"))

    try:
        wb = load_workbook(STAGE2_EF_XLSX, keep_links=False)
        if sheet not in wb.sheetnames:
            flash("Sheet not found in mapping workbook")
            return redirect(url_for("manage_emission_factors"))
        ws = wb[sheet]
        row_idx = _find_ef_row_in_sheet(ws, ef_id)
        if row_idx is None:
            flash("Could not locate the emission factor row by ef_id")
            return redirect(url_for("manage_emission_factors"))

        backup_dir = STAGE2_EF_XLSX.parent / "_ef_backups"
        backup_dir.mkdir(parents=True, exist_ok=True)
        ts = datetime.now().strftime("%Y%m%d_%H%M%S")
        backup_path = backup_dir / f"{STAGE2_EF_XLSX.stem}_before_delete_{ts}{STAGE2_EF_XLSX.suffix}"
        try:
            shutil.copy2(STAGE2_EF_XLSX, backup_path)
        except Exception:
            pass

        ws.delete_rows(row_idx, 1)
        wb.save(STAGE2_EF_XLSX)
        _clear_ef_cache()
        flash("Emission factor deleted from mapping workbook.")
    except PermissionError:
        flash("Mapping workbook is open/locked. Close Excel and try again.")
    except Exception as e:
        flash(f"Delete failed: {e}")

    return redirect(url_for("manage_emission_factors"))


@app.route("/admin/emission_factors/mapping/import", methods=["POST"])
@login_required
def import_emission_factors_mapping():
    if not current_user.is_admin:
        flash("Access denied")
        return redirect(url_for("dashboard"))

    f = request.files.get("excel_file")
    if not f or not f.filename:
        flash("Please choose an Excel file to import.")
        return redirect(url_for("manage_emission_factors"))

    fn = secure_filename(f.filename)
    if not fn.lower().endswith((".xlsx", ".xlsm")):
        flash("Please upload a .xlsx or .xlsm file.")
        return redirect(url_for("manage_emission_factors"))

    tmp_dir = FRONTEND_UPLOAD_DIR
    tmp_dir.mkdir(parents=True, exist_ok=True)
    tmp_path = tmp_dir / f"ef_import_{datetime.now().strftime('%Y%m%d_%H%M%S')}_{fn}"
    f.save(tmp_path)

    # Validate structure: sheets with expected headers
    try:
        wb = load_workbook(tmp_path, read_only=True, data_only=True, keep_links=False)
        expected = _ef_expected_headers()
        bad = []
        for name in wb.sheetnames:
            ws = wb[name]
            first = next(ws.iter_rows(min_row=1, max_row=1, values_only=True), None)
            headers = [("" if v is None else str(v).strip()) for v in (first or [])]
            while headers and headers[-1] == "":
                headers.pop()
            missing = [h for h in expected if h not in headers]
            if missing:
                bad.append(f"{name}: missing {', '.join(missing)}")
        if bad:
            flash("Import failed. Workbook structure is not compatible: " + " | ".join(bad[:4]) + (" ..." if len(bad) > 4 else ""))
            try:
                tmp_path.unlink()
            except Exception:
                pass
            return redirect(url_for("manage_emission_factors"))
    except Exception as e:
        flash(f"Import failed: {e}")
        try:
            tmp_path.unlink()
        except Exception:
            pass
        return redirect(url_for("manage_emission_factors"))
    
    except Exception as e:
        flash(f"Import failed {e}")
        try:
            tmp_path.unlink()
        except Exception:
            pass
        return redirect(url_for("manage_emission_factors"))
    
    try:
        backup_dir = STAGE2_EF_XLSX.parent / "_ef_backups"
        backup_dir.mkdir(parents=True, exists_ok=True)
        ts = datetime.now().strftime("%Y%m%d_%H%M%S")
        if STAGE2_EF_XLSX.exists():
            shutil.copy(STAGE2_EF_XLSX, backup_dir / f"{STAGE2_EF_XLSX.stem}_before_import_{ts}{STAGE2_EF_XLSX.suffix}" )
        os.replace(str(tmp_path), str(STAGE2_EF_XLSX))
        _clear_ef_cache()
        flash("Emission factors imported successfully (sheet preserved).")
    except PermissionError:
        flash("Mapping workbook is open. Please close Excel file and try it again")
        try:
            tmp_path.unlink()
        except Exception:
            pass
        except Exception as e:
            flash(f"Import failed: {e}")
            try:
                tmp_path.unlink()
            except Exception:
                pass

    # Backup current and replace
    try:
        backup_dir = STAGE2_EF_XLSX.parent / "_ef_backups"
        backup_dir.mkdir(parents=True, exist_ok=True)
        ts = datetime.now().strftime("%Y%m%d_%H%M%S")
        if STAGE2_EF_XLSX.exists():
            shutil.copy2(STAGE2_EF_XLSX, backup_dir / f"{STAGE2_EF_XLSX.stem}_before_import_{ts}{STAGE2_EF_XLSX.suffix}")
        os.replace(str(tmp_path), str(STAGE2_EF_XLSX))
        _clear_ef_cache()
        flash("Emission factors imported successfully (sheets preserved).")
    except PermissionError:
        flash("Mapping workbook is open/locked. Close Excel and try again.")
        try:
            tmp_path.unlink()
        except Exception:
            pass
    except Exception as e:
        flash(f"Import failed: {e}")
        try:
            tmp_path.unlink()
        except Exception:
            pass

    return redirect(url_for("manage_emission_factors"))

@app.route('/admin/emission_factors/export', methods=['GET'])
def export_emission_factors():
    if not current_user.is_authenticated or not getattr(current_user, "is_admin", False):
        return redirect(url_for("dashboard"))

    data = _load_stage2_emission_factors()
    all_rows: list[dict[str, object]] = list(data.get("rows") or [])

    search = request.args.get('search', '').strip()
    sheet_filter = request.args.get('sheet', '').strip()
    scope_filter = request.args.get('scope', '').strip()
    source_filter = request.args.get('ef_source', '').strip()

    def _matches(row: dict[str, object]) -> bool:
        if sheet_filter and str(row.get("sheet") or "") != sheet_filter:
            return False
        if scope_filter and str(row.get("scope") or "") != scope_filter:
            return False
        if source_filter and str(row.get("ef_source") or "") != source_filter:
            return False
        if search:
            hay = " ".join(
                str(row.get(k) or "")
                for k in (
                    "sheet",
                    "ef_name",
                    "ef_description",
                    "scope",
                    "ef_category",
                    "ef_id",
                    "ef_value",
                    "ef_unit",
                    "ef_source",
                    "Emission Factor Category",
                )
            ).lower()
            if search.lower() not in hay:
                return False
        return True

    rows = [r for r in all_rows if _matches(r)]
    rows.sort(key=lambda r: (str(r.get("sheet") or ""), str(r.get("ef_name") or ""), str(r.get("ef_id") or "")))

    
    def _safe_xlsx_sheet_name(name: str) -> str:
        invalid = set(':/\\?*[]')
        cleaned = "".join(ch for ch in (name or "Sheet") if ch not in invalid)
        if not cleaned:
            cleaned = "Sheet"
        return cleaned[:31]

    def _valid_sheet_name(name: str) -> bool:
        if not name:
            return False
        if len(name) > 31:
            return False
        if any(ch in name for ch in ':/\\?*[]'):
            return False
        return True
    # Group by source sheet (category)
    groups: dict[str, list[dict[str, object]]] = {}
    for r in rows:
        groups.setdefault(str(r.get("sheet") or "Sheet"), []).append(r)

    expected_headers = _ef_expected_headers()

    wb = Workbook()
    try:
        wb.remove(wb.active)
    except Exception:
        pass

    used_names: dict[str, int] = {}
    for raw_name in sorted(groups.keys(), key=lambda x: x.lower()):
        base = raw_name if _valid_sheet_name(raw_name) else _safe_xlsx_sheet_name(raw_name)
        n = used_names.get(base, 0) + 1
        used_names[base] = n
        sheet_name = base if n == 1 else _safe_xlsx_sheet_name(f"{base}_{n}")

        ws = wb.create_sheet(title=sheet_name)
        # Header row
        for j, h in enumerate(expected_headers, start=1):
            ws.cell(row=1, column=j).value = h

        # Rows
        for i, r in enumerate(groups[raw_name], start=2):
            for j, h in enumerate(expected_headers, start=1):
                v = r.get(h)
                # Keep ef_value numeric when possible
                if h == "ef_value":
                    try:
                        if v is None or str(v).strip() == "":
                            ws.cell(row=i, column=j).value = None
                        else:
                            ws.cell(row=i, column=j).value = float(v)
                    except Exception:
                        ws.cell(row=i, column=j).value = "" if v is None else str(v)
                else:
                    ws.cell(row=i, column=j).value = None if v is None or str(v).strip() == "" else str(v)

        # Light column sizing
        for col in range(1, len(expected_headers) + 1):
            ws.column_dimensions[chr(64 + col)].width = 22 if col in (1, 4, 5, 6, 7, 8) else 34

        ws.freeze_panes = "A2"

    bio = BytesIO()
    wb.save(bio)
    bio.seek(0)

    filename = f"CTS_Emission_factors_short_list_export_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
    return send_file(
        bio,
        as_attachment=True,
        download_name=filename,
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )

@app.route('/home')
@login_required
def home():
    _ensure_db_tables()
    _backfill_mapping_summaries()
    year = datetime.utcnow().year

    if bool(getattr(current_user, "is_admin", False)):
        # Admin overview across all companies (latest per company+sheet)
        all_rows = MappingRunSummary.query.order_by(MappingRunSummary.created_at.desc()).all()
        latest_by_company_sheet: dict[tuple[str, str], MappingRunSummary] = {}
        for r in all_rows:
            key = ((r.company_name or "").strip(), (r.sheet_name or "").strip().lower())
            if key in latest_by_company_sheet:
                continue
            if not key[0] or not key[1]:
                continue
            latest_by_company_sheet[key] = r

        company_totals: dict[str, dict[str, float]] = defaultdict(lambda: {"scope1": 0.0, "scope2": 0.0, "scope3": 0.0, "total": 0.0})
        for (_c, _s), r in latest_by_company_sheet.items():
            c = (r.company_name or "").strip()
            v = float(getattr(r, "tco2e_total", 0.0) or 0.0)
            sc = int(getattr(r, "scope", 0) or 0)
            if sc == 1:
                company_totals[c]["scope1"] += v
            elif sc == 2:
                company_totals[c]["scope2"] += v
            elif sc == 3:
                company_totals[c]["scope3"] += v
            company_totals[c]["total"] += v

        rows = []
        g = {"scope1": 0.0, "scope2": 0.0, "scope3": 0.0, "total": 0.0}
        for c in sorted(company_totals.keys(), key=lambda x: x.lower()):
            t = company_totals[c]
            rows.append({"company": c, **{k: round(t[k], 3) for k in ("scope1", "scope2", "scope3", "total")}})
            for k in g:
                g[k] += float(t[k] or 0.0)

        return render_template(
            "home.html",
            year=year,
            is_admin=True,
            totals={k: round(g[k], 3) for k in g},
            company_rows=rows,
            breakdown=[],
        )      

    # Regular user: show totals for their company (latest per sheet)
    keys = _company_candidate_keys(getattr(current_user, "company_name", "") or "")
    latest = _latest_sheet_totals_for_company(keys)
    # Prefer year inferred from the most recent mapped output (if available)
    try:
        if latest:
            mr0 = MappingRun.query.get(latest[0].run_id)
            if mr0 and getattr(mr0, "output_path", None):
                y = _detect_data_year_from_xlsx(mr0.output_path, latest[0].sheet_name)
                if y:
                    year = y
    except Exception:
        pass
    scope1 = sum(float(r.tco2e_total or 0.0) for r in latest if int(r.scope or 0) == 1)
    scope2 = sum(float(r.tco2e_total or 0.0) for r in latest if int(r.scope or 0) == 2)
    scope3 = sum(float(r.tco2e_total or 0.0) for r in latest if int(r.scope or 0) == 3)
    total_emission = scope1 + scope2 + scope3
    breakdown = [
        {
            "sheet": r.sheet_name,
            "scope": r.scope,
            "tco2e": round(float(r.tco2e_total or 0.0), 3),
            "updated_at": r.created_at.strftime("%Y-%m-%d %H:%M") if r.created_at else "",
        }
        for r in latest
    ]

    return render_template(
        "home.html",
        year=year,
        is_admin=False,
        totals={
            "total": round(total_emission, 3),
            "scope1": round(scope1, 3),
            "scope2": round(scope2, 3),
            "scope3": round(scope3, 3),
        },
        breakdown=breakdown,
        company_rows=[],
    )
    
@app.route('/Emission-factors')
@login_required
def emission_facor():
    return redirect(url_for('manage_emission_factors'))


@app.route('/carbon-accounting')
@login_required
def carbon_accounting():
    _ensure_db_tables()
    _backfill_mapping_summaries()
    year = datetime.utcnow().year

    keys = _company_candidate_keys(getattr(current_user, "company_name", "") or "")
    latest = _latest_sheet_totals_for_company(keys)
    try:
        if latest:
            mr0 = MappingRun.query.get(latest[0].run_id)
            if mr0 and getattr(mr0, "output_path", None):
                y = _detect_data_year_from_xlsx(mr0.output_path, latest[0].sheet_name)
                if y:
                    year = y
    except Exception:
        pass

    scope1 = sum(float(r.tco2e_total or 0.0) for r in latest if int(r.scope or 0) == 1)
    scope2 = sum(float(r.tco2e_total or 0.0) for r in latest if int(r.scope or 0) == 2)
    scope3 = sum(float(r.tco2e_total or 0.0) for r in latest if int(r.scope or 0) == 3)
    total_emission = scope1 + scope2 + scope3

    by_sheet = [
        {
            "sheet": r.sheet_name,
            "scope": r.scope,
            "tco2e": round(float(r.tco2e_total or 0.0), 3),
            "updated_at": r.created_at.strftime("%Y-%m-%d %H:%M") if r.created_at else "",
        }
        for r in latest
    ]

    return render_template(
        "carbon_accounting.html",
        year=year,
        totals={
            "total": round(total_emission, 3),
            "scope1": round(scope1, 3),
            "scope2": round(scope2, 3),
            "scope3": round(scope3, 3),
        },
        by_sheet=by_sheet,
    )

if __name__ == '__main__':
    with app.app_context():
        db.create_all()
    app.run(debug=True, host='0.0.0.0', port=5000) 
