from __future__ import annotations

from pathlib import Path
from typing import List, Optional, Dict
import glob
import os
import sys
import pandas as pd
import re
from openpyxl.styles import Font, PatternFill
from openpyxl.utils import get_column_letter

PROJECT_ROOT = Path(__file__).resolve().parents[2]
if str(PROJECT_ROOT) not in sys.path:
    sys.path.insert(0, str(PROJECT_ROOT))

from config import STAGE1_KLARAKARBON_OUTPUT_DIR, STAGE2_OUTPUT_DIR, STAGE2_TRAVEL_DIR





BASE_DIR = Path(__file__).resolve().parent
OUTPUT_DIR = STAGE2_OUTPUT_DIR

# External source workbooks now resolve from DATA_DIR-backed locations.
TRAVEL_PATH = STAGE2_TRAVEL_DIR / "analysis_summary.xlsx"
KLARAKARBON_PATH = STAGE1_KLARAKARBON_OUTPUT_DIR / "klarakarbon_categories_mapped_FINAL.xlsx"


def find_latest_final_workbook(base_dir: Path) -> Optional[Path]:
    """
    En güncel final workbook'u bul:
      1) mapped_results_merged_*.xlsx (TERCIH EDILEN)
      2) mapped_results_merged_dc_*.xlsx
      3) mapped_results.xlsx
    'with_sources' yardımcı kopyaları tercih etme.
    """
    out = STAGE2_OUTPUT_DIR
    patterns = [
        str(out / "mapped_results_merged_*.xlsx"),
        str(out / "mapped_results_merged.xlsx"),
        str(out / "mapped_results_merged_dc_*.xlsx"),
        str(out / "mapped_results_merged_dc.xlsx"),
        str(out / "mapped_results.xlsx"),
    ]
    candidates: List[str] = []
    for pat in patterns:
        candidates.extend(glob.glob(pat))
    if not candidates:
        return None
    # Exclude helper copies like *_with_sources_*
    filtered = [c for c in candidates if "with_sources" not in os.path.basename(c).lower()]
    if not filtered:
        filtered = candidates
    # Pick the newest by mtime
    filtered.sort(key=os.path.getmtime, reverse=True)
    return Path(filtered[0])


def concat_all_sheets(xls_path: Path, limit_first_n: Optional[int] = None) -> pd.DataFrame:
    """Read workbook and concatenate its sheets into one DataFrame, preserving sheet name.

    If limit_first_n is provided, only the first N sheets are concatenated.
    """
    try:
        xls = pd.ExcelFile(xls_path)
    except Exception:
        return pd.DataFrame()

    sheet_names = xls.sheet_names
    if limit_first_n is not None:
        sheet_names = sheet_names[:limit_first_n]

    parts: List[pd.DataFrame] = []
    for s in sheet_names:
        try:
            df = pd.read_excel(xls, sheet_name=s)
        except Exception:
            continue
        if df is None or df.empty:
            continue
        temp = df.copy()
        temp["Sheet"] = s
        parts.append(temp)

    if not parts:
        return pd.DataFrame()
    return pd.concat(parts, axis=0, join="outer", ignore_index=True)


def write_with_openpyxl_append(target_path: Path, klar_df: pd.DataFrame, travel_df: pd.DataFrame) -> Path:
    def _style_sheet(writer: pd.ExcelWriter, sheet_name: str, df: pd.DataFrame) -> None:
        try:
            ws = writer.sheets.get(sheet_name)
            if ws is None:
                # Try via workbook
                ws = writer.book[sheet_name]
            # Header style
            header_fill = PatternFill(start_color="D8EAD3", end_color="D8EAD3", fill_type="solid")
            bold_font = Font(bold=True)
            for j, col in enumerate(list(df.columns), start=1):
                cell = ws.cell(row=1, column=j)
                cell.font = bold_font
                cell.fill = header_fill
            # Freeze top row
            ws.freeze_panes = "A2"
            # Column widths
            for j, col in enumerate(list(df.columns), start=1):
                try:
                    series = df[col].astype(str)
                    max_len = max([len(str(col))] + series.str.len().tolist())
                    width = min(max(8, max_len + 1), 40)
                except Exception:
                    width = 16
                ws.column_dimensions[get_column_letter(j)].width = width

            # Date number format for known/date-like columns
            date_like_cols = {
                "reporting period (month, year)",
            }
            for j, col in enumerate(list(df.columns), start=1):
                col_l = str(col).strip().lower()
                is_datetime = pd.api.types.is_datetime64_any_dtype(df[col]) if col in df.columns else False
                if is_datetime or col_l in date_like_cols:
                    for i in range(2, len(df) + 2):
                        try:
                            ws.cell(row=i, column=j).number_format = "yyyy-mm-dd"
                        except Exception:
                            pass
            # Two-decimal number format for co2e columns
            for j, col in enumerate(list(df.columns), start=1):
                col_l = str(col).strip().lower()
                if col_l in {"co2e", "co2e (t)"}:
                    for i in range(2, len(df) + 2):
                        try:
                            ws.cell(row=i, column=j).number_format = "0.00"
                        except Exception:
                            pass
        except Exception:
            # Best effort; ignore styling failures
            pass

    # Try in-place append
    try:
        with pd.ExcelWriter(target_path, engine="openpyxl", mode="a", if_sheet_exists="replace") as writer:
            # Append/replace our two sheets
            klar_df.to_excel(writer, sheet_name="Klarakarbon", index=False)
            _style_sheet(writer, "Klarakarbon", klar_df)
            travel_df.to_excel(writer, sheet_name="Travel", index=False)
            _style_sheet(writer, "Travel", travel_df)

            # Style ALL sheets in the workbook, not only the two above
            try:
                xls_all = pd.ExcelFile(target_path)
                for s in xls_all.sheet_names:
                    try:
                        df_s = pd.read_excel(xls_all, sheet_name=s)
                        if df_s is not None and not df_s.empty:
                            _style_sheet(writer, s[:31], df_s)
                    except Exception:
                        continue
            except Exception:
                pass

        return target_path
    except PermissionError:
        # Fallback: write to a timestamped copy INCLUDING all original sheets
        ts_path = target_path.with_name(f"{target_path.stem}_with_sources_{pd.Timestamp.now().strftime('%Y%m%d_%H%M%S')}{target_path.suffix}")
        try:
            xls = pd.ExcelFile(target_path)
            with pd.ExcelWriter(ts_path, engine="openpyxl") as writer:
                # Copy and style all existing sheets
                for s in xls.sheet_names:
                    try:
                        df_old = pd.read_excel(xls, sheet_name=s)
                        df_old.to_excel(writer, sheet_name=s[:31], index=False)
                        if df_old is not None and not df_old.empty:
                            _style_sheet(writer, s[:31], df_old)
                    except Exception:
                        continue
                # Append/replace new sheets with styling
                klar_df.to_excel(writer, sheet_name="Klarakarbon", index=False)
                _style_sheet(writer, "Klarakarbon", klar_df)
                travel_df.to_excel(writer, sheet_name="Travel", index=False)
                _style_sheet(writer, "Travel", travel_df)
        except Exception:
            # If we fail to read original, write at least the two new sheets (styled)
            with pd.ExcelWriter(ts_path, engine="openpyxl") as writer:
                klar_df.to_excel(writer, sheet_name="Klarakarbon", index=False)
                _style_sheet(writer, "Klarakarbon", klar_df)
                travel_df.to_excel(writer, sheet_name="Travel", index=False)
                _style_sheet(writer, "Travel", travel_df)
        return ts_path


def main() -> None:
    base_wb = find_latest_final_workbook(BASE_DIR)
    if base_wb is None:
        print("No final workbook found under output/.")
        return

    klar_df = concat_all_sheets(KLARAKARBON_PATH, limit_first_n=None)
    # Create 'co2e' in tonnes from 'co2e (kg)' if present (do NOT rename original)
    if not klar_df.empty:
        try:
            lowmap = {str(c).strip().lower(): c for c in klar_df.columns}
            co2e_kg_col = None
            for key in ["co2e (kg)", "co2e(kg)", "co2e kg"]:
                if key in lowmap:
                    co2e_kg_col = lowmap[key]
                    break
            if co2e_kg_col is not None and co2e_kg_col in klar_df.columns:
                klar_df["co2e"] = pd.to_numeric(klar_df[co2e_kg_col], errors="coerce").fillna(0.0) / 1000.0
        except Exception:
            pass

    # Build Company column from source_company using provided mapping
    if not klar_df.empty:
        try:
            lowmap = {str(c).strip().lower(): c for c in klar_df.columns}
            sc_col = lowmap.get("source_company")
            if sc_col is not None and sc_col in klar_df.columns:
                norm = klar_df[sc_col].astype(str).str.strip().str.lower()
                mapping = {
                    "gapit nordics": "Gapit",
                    "nordicepod": "NordicEPOD",
                    "fortica": "Fortica",
                    "gt nordics": "GT Nordics",
                    "nep switchboards": "NEP Switchboards",
                }
                mapped = norm.map(mapping)
                # Fallback to original cleaned casing when not mapped
                fallback = klar_df[sc_col].astype(str).str.strip()
                klar_df["Company"] = mapped.fillna(fallback).astype("object")
        except Exception:
            pass

    # Copy 'ghg category' -> 'GHGP Category' if present (case-insensitive)
    if not klar_df.empty:
        try:
            lowmap = {str(c).strip().lower(): c for c in klar_df.columns}
            ghg_col = lowmap.get("ghg category") or lowmap.get("ghg_category") or lowmap.get("ghgcategory")
            if ghg_col is not None and ghg_col in klar_df.columns:
                klar_df["GHGP Category"] = klar_df[ghg_col].astype("object")
        except Exception:
            pass

    # --- Klarakarbon: drop duplicate rows (ignore the helper 'Sheet' column) ---
    if not klar_df.empty:
        try:
            subset_cols = [c for c in klar_df.columns if c != "Sheet"]
            before_n = len(klar_df)
            klar_df = klar_df.drop_duplicates(subset=subset_cols, keep="first").reset_index(drop=True)
            after_n = len(klar_df)
            removed = before_n - after_n
            if removed > 0:
                print(f"Klarakarbon: removed {removed} duplicate rows")
        except Exception:
            pass

    # Per user: use only the SECOND sheet from Travel
    try:
        txls = pd.ExcelFile(TRAVEL_PATH)
        if len(txls.sheet_names) >= 2:
            tname = txls.sheet_names[1]
        else:
            tname = txls.sheet_names[0]
        travel_df = pd.read_excel(txls, sheet_name=tname)
        # Preserve sheet name column for context
        travel_df["Sheet"] = tname
    except Exception:
        # Fallback to empty if Travel workbook cannot be read
        travel_df = pd.DataFrame()

    # ---------------- Travel transformation: map Cost Center -> Source_file ----------------
    def _detect_cost_center_column(df: pd.DataFrame) -> Optional[str]:
        candidates = [
            "Cost Center", "Cost center", "Cost Centre", "Cost centre",
            "Cost control center", "Cost Control Center", "CCC", "By cost center",
        ]
        lowmap = {str(c).strip().lower(): c for c in df.columns}
        for c in candidates:
            if c in df.columns:
                return c
        for c in candidates:
            k = str(c).strip().lower()
            if k in lowmap:
                return lowmap[k]
        # relaxed: substring
        for c in candidates:
            k = str(c).strip().lower().replace(" ", "")
            for col in df.columns:
                kk = str(col).strip().lower().replace(" ", "")
                if k in kk or kk in k:
                    return col
        return None

    def _norm_key(text: Optional[str]) -> str:
        if text is None:
            return ""
        s = str(text).strip().upper()
        s = s.replace("-", " ").replace("_", " ")
        s = re.sub(r"\s+", " ", s)
        # keep letters, numbers and spaces
        s = re.sub(r"[^A-Z0-9 ]", "", s)
        s = re.sub(r"\s+", " ", s).strip()
        return s


# CORRECT
    # Mapping  (left: Travel Cost Center raw, right: canonical company)
    USER_MAP: Dict[str, str] = {
        _norm_key("BIMMS"): "BIMMS",
        _norm_key("CAERUS"): "Caerus Nordics",
        _norm_key("CTS"): "CTS Nordics",
        _norm_key("CTS DENMARK"): "CTS Denmark",
        _norm_key("CTS EUROPE"): "CTS EU",
        _norm_key("CTS FINLAND"): "CTS Finland",
        _norm_key("CTS SWEDEN"): "CTS Sweden",
        _norm_key("CTS SW"): "CTS Sweden",
        _norm_key("CTS SWEEDEN"): "CTS Sweden",
        _norm_key("SECURITY SOLUTIONS"): "CTS Security Solutions",
        _norm_key("CTS SECURITY"): "CTS Security Solutions",
        _norm_key("CTS PORTUGAL"): "Navitas Portugal",
        _norm_key("CTS Norway"): "CTS Nordics",
        _norm_key("CTS-VDC SERVICES"): "CTS-VDC",
        _norm_key("DC PEOPLE"): "CTS Nordics",
        _norm_key("DC-WORLD-CONSULTANCY"): "CTS Nordics",
        _norm_key("DCS"): "CTS Nordics",
        _norm_key("GAPIT"): "Gapit",
        _norm_key("MECWIDE"): "Mecwide Nordics",
        _norm_key("NAVITAS"): "Navitas Norway",
        _norm_key("NORDICEPOD"): "NordicEPOD",
        _norm_key("QEC"): "QEC",
        _norm_key("SD NORDICS"): "SD Nordics",
        _norm_key("VELOX ELECTRO EUROPE"): "Porvelox",
    }

    if not travel_df.empty:
        cc_col = _detect_cost_center_column(travel_df)
        if cc_col is not None:
            normalized = travel_df[cc_col].astype(str).map(_norm_key)
            mapped = normalized.map(USER_MAP)
            # Default others to CTS Nordics
            travel_df["Source_file"] = mapped.fillna("CTS Nordics").astype("object")
        else:
            # If Cost Center not found, still create Source_file as CTS Nordics default
            travel_df["Source_file"] = "CTS Nordics"

        # ---- Build Reporting period (month, year) as yyyy-mm-dd with day fixed to 09 ----
        def _detect_col(df: pd.DataFrame, names: List[str]) -> Optional[str]:
            lowmap = {str(c).strip().lower(): c for c in df.columns}
            for n in names:
                if n in df.columns:
                    return n
            for n in names:
                k = str(n).strip().lower()
                if k in lowmap:
                    return lowmap[k]
            return None

        year_col = _detect_col(travel_df, ["Year", "year"])  # may contain year or mixed text
        month_col = _detect_col(travel_df, ["Month", "month"])  # may contain month name or mixed

        month_map = {
            "jan": 1, "january": 1,
            "feb": 2, "february": 2,
            "mar": 3, "march": 3,
            "apr": 4, "april": 4,
            "may": 5, "may": 5,
            "jun": 6, "june": 6,
            "jul": 7, "july": 7,
            "aug": 8, "august": 8,
            "sep": 9, "sept": 9, "september": 9,
            "oct": 10, "october": 10,
            "nov": 11, "november": 11,
            "dec": 12, "december": 12,
        }

        years: List[Optional[int]] = []
        months: List[Optional[int]] = []
        for idx in range(len(travel_df)):
            y_raw = travel_df.iloc[idx][year_col] if year_col in travel_df.columns else None
            m_raw = travel_df.iloc[idx][month_col] if month_col in travel_df.columns else None
            blob = f"{y_raw} {m_raw}".strip()
            low = str(blob).lower()
            # year
            y_match = re.search(r"(19|20)\d{2}", low)
            y_val = int(y_match.group(0)) if y_match else None
            # month
            m_val: Optional[int] = None
            for key, num in month_map.items():
                if key in low:
                    m_val = num
                    break
            # numeric month fallback
            if m_val is None:
                try:
                    m_val = int(str(m_raw)) if str(m_raw).isdigit() else None
                except Exception:
                    m_val = None
            years.append(y_val)
            months.append(m_val)

        dt_series = []
        for y, m in zip(years, months):
            try:
                if y is not None and m is not None:
                    dt = pd.Timestamp(year=y, month=m, day=9).date()
                else:
                    dt = None
            except Exception:
                dt = None
            dt_series.append(dt)
        # Use Python date objects to avoid time components
        travel_df["Reporting period (month, year)"] = pd.Series(dt_series, dtype="object")



# Create an option turn off this function for auditor.

        # --- GHGP Category for Travel ---
        try:
            # Business rule (updated):
            # - If Cost Center (normalized) maps via USER_MAP → Cat 6 Business Travel.
            # - Others → Cat 1 Purchased Goods and Services.
            # This ensures all recognized company cost centers are treated as Business Travel.
            travel_df["GHGP Category"] = "Scope 3 Category 1 Purchased Goods and Services"
            if cc_col is not None:
                norm_cc = travel_df[cc_col].astype(str).map(_norm_key)
                recognized_keys = set(USER_MAP.keys())
                mask_bt = norm_cc.isin(recognized_keys)
                travel_df.loc[mask_bt, "GHGP Category"] = "Scope 3 Category 6 Business Travel"
        except Exception:
            # On error, default to Purchased Goods & Services to avoid over-attribution to Business Travel
            travel_df["GHGP Category"] = "Scope 3 Category 1 Purchased Goods and Services"
            
               

        # --- Normalize date to remove time part if any ---
        try:
            col_dt = "Reporting period (month, year)"
            if col_dt in travel_df.columns:
                travel_df[col_dt] = pd.to_datetime(travel_df[col_dt], errors="coerce").dt.date
        except Exception:
            pass

        # --- Travel: ensure 'co2e (kg)' and derived 'co2e' (t) ---
        try:
            lowmap_t = {str(c).strip().lower(): c for c in travel_df.columns}
            # If there is a 'co2e' column, treat it as kg and rename to 'co2e (kg)'
            if "co2e" in lowmap_t and "co2e (kg)" not in lowmap_t:
                travel_df = travel_df.rename(columns={lowmap_t["co2e"]: "co2e (kg)"})
                lowmap_t = {str(c).strip().lower(): c for c in travel_df.columns}
            # If there is 'Total CO2', rename to 'co2e (kg)'
            if "total co2" in lowmap_t:
                travel_df = travel_df.rename(columns={lowmap_t["total co2"]: "co2e (kg)"})
                lowmap_t = {str(c).strip().lower(): c for c in travel_df.columns}
            # Create 'co2e' tonnes column from 'co2e (kg)' if present
            if "co2e (kg)" in lowmap_t:
                col_kg = lowmap_t["co2e (kg)"]
                travel_df["co2e"] = pd.to_numeric(travel_df[col_kg], errors="coerce").fillna(0.0) / 1000.0
        except Exception:
            pass

    # Ensure object dtype for strings to avoid dtype warnings
    for df in (klar_df, travel_df):
        for col in df.columns:
            try:
                if pd.api.types.is_object_dtype(df[col]) or pd.api.types.is_string_dtype(df[col]):
                    df[col] = df[col].astype("object")
            except Exception:
                pass

    written = write_with_openpyxl_append(base_wb, klar_df, travel_df)
    if written == base_wb:
        print(f"Appended sheets into existing workbook: {written.name}")
    else:
        print(f"Base workbook locked; wrote two sheets into new file: {written.name}")


if __name__ == "__main__":
    main()


