from __future__ import annotations

import argparse
import math
import re
import shutil
from datetime import datetime
from pathlib import Path
from typing import Dict, Optional, Tuple
import sys

import pandas as pd
from openpyxl import load_workbook

PROJECT_ROOT = Path(__file__).resolve().parents[2]
if str(PROJECT_ROOT) not in sys.path:
    sys.path.insert(0, str(PROJECT_ROOT))

from config import STAGE2_EF_XLSX


def _norm(s: object) -> str:
    try:
        return str(s).strip().lower()
    except Exception:
        return ""


def _norm_efid(val: object) -> Optional[str]:
    try:
        s = str(val).strip()
    except Exception:
        return None
    if not s:
        return None
    s2 = re.sub(r"[^A-Za-z0-9]", "", s).upper()
    return s2 if s2 else None


def _to_float(v: object) -> Optional[float]:
    if v is None:
        return None
    if isinstance(v, (int, float)):
        try:
            fv = float(v)
            return fv if math.isfinite(fv) else None
        except Exception:
            return None
    s = str(v).strip()
    if s == "":
        return None
    s = s.replace("\u00A0", "").replace(" ", "")
    if "," in s and "." in s:
        s = s.replace(",", "")
    elif "," in s and "." not in s:
        s = s.replace(",", ".")
    try:
        fv = float(s)
        return fv if math.isfinite(fv) else None
    except Exception:
        return None


def _find_col_indices(headers: list[object]) -> Tuple[int, int, int, int, int, int, int, int]:
    """Return 1-based indices for (Company, Sheet_booklets, match_method, Spend_Euro, ef_id, ef_unit, ef_value, co2e (t))."""
    idx_company = idx_booklets = idx_mm = idx_spend = idx_id = idx_unit = idx_val = idx_co2e_t = -1
    for i, h in enumerate(headers, start=1):
        key = _norm(h)
        if key == "company":
            idx_company = i
        elif key == "sheet_booklets":
            idx_booklets = i
        elif key in {"match_method", "match method"}:
            idx_mm = i
        elif key in {"spend_euro", "spend euro", "spend (euro)", "spend eur"}:
            if idx_spend == -1:
                idx_spend = i
        elif key in {"ef_id", "ef id"}:
            idx_id = i
        elif key in {"ef_unit", "ef unit"}:
            idx_unit = i
        elif key in {"ef_value", "ef value"}:
            idx_val = i
        elif key == "co2e (t)":
            idx_co2e_t = i
    if -1 in {idx_company, idx_booklets, idx_mm, idx_spend, idx_id, idx_val, idx_co2e_t}:
        raise RuntimeError(
            "Required columns not found. "
            f"company={idx_company}, sheet_booklets={idx_booklets}, match_method={idx_mm}, "
            f"spend={idx_spend}, ef_id={idx_id}, ef_unit={idx_unit}, ef_value={idx_val}, co2e(t)={idx_co2e_t}"
        )
    return idx_company, idx_booklets, idx_mm, idx_spend, idx_id, idx_unit, idx_val, idx_co2e_t


def _compute_tco2e(spend_eur: Optional[float], ef_value: Optional[float], ef_unit: Optional[str]) -> Optional[float]:
    """Compute tCO2e for EUR-based factors. If unit is kg/EUR -> /1000, else assume tonnes/EUR."""
    if spend_eur is None or ef_value is None:
        return None
    try:
        s = float(spend_eur)
        ef = float(ef_value)
    except Exception:
        return None
    u = (ef_unit or "").strip().lower()
    # If EF is kgCO2e/EUR then convert to tonnes
    if ("kg" in u) and ("eur" in u or "€" in u) and ("/" in u or "per" in u):
        return (s * ef) / 1000.0
    return s * ef


def _build_ef_lookup(ef_workbook: Path) -> Dict[str, Dict[str, object]]:
    """Build EF ID -> {ef_value, ef_unit, ef_source} from allowed EF sheets only."""
    allowed = {"scope 3 purchased service spend", "scope 3 purchased goods spend"}
    sheets = pd.read_excel(ef_workbook, sheet_name=None, engine="openpyxl")
    out: Dict[str, Dict[str, object]] = {}

    def _pick_col(df: pd.DataFrame, cands: list[str]) -> Optional[str]:
        lowmap = {re.sub(r"[^a-z0-9]", "", str(c).lower()): c for c in df.columns}
        for cand in cands:
            k = re.sub(r"[^a-z0-9]", "", cand.lower())
            if k in lowmap:
                return lowmap[k]
        return None

    for name, df in sheets.items():
        if _norm(name) not in allowed:
            continue
        if df is None or df.empty:
            continue
        id_col = _pick_col(df, ["ef_id", "EF ID", "EFID", "ID", "Factor ID", "Emission Factor ID"])
        val_col = _pick_col(df, ["ef_value", "EF Value", "Value", "Emission", "Emission Factor Value"])
        unit_col = _pick_col(df, ["EF Unit", "Unit", "Units"])
        src_col = _pick_col(df, ["Source", "Reference", "Publication", "Provider"])
        if id_col is None or val_col is None:
            continue
        for _, rr in df.iterrows():
            nid = _norm_efid(rr.get(id_col))
            if nid is None or nid in out:
                continue
            fv = _to_float(rr.get(val_col))
            if fv is None:
                continue
            out[nid] = {
                "ef_value": fv,
                "ef_unit": rr.get(unit_col) if unit_col else None,
                "ef_source": rr.get(src_col) if src_col else None,
            }
    return out


def apply_fix(
    input_path: Path,
    output_path: Path,
    ef_workbook: Path,
    sheet_name: str = "S3 Cat 1 Purchased G&S",
    company: str = "BIMMS",
    sheet_booklets_value: str = "Scope 3 Cat 1 Services Spend",
) -> int:
    ef_lookup = _build_ef_lookup(ef_workbook)

    wb = load_workbook(input_path, data_only=False)
    if sheet_name not in wb.sheetnames:
        raise RuntimeError(f"Sheet not found: {sheet_name}")
    ws = wb[sheet_name]
    headers = [c.value for c in ws[1]]
    idx_company, idx_booklets, idx_mm, idx_spend, idx_id, idx_unit, idx_val, idx_co2e_t = _find_col_indices(headers)

    updated = 0
    for r in range(2, ws.max_row + 1):
        if _norm(ws.cell(row=r, column=idx_company).value) != _norm(company):
            continue
        if _norm(ws.cell(row=r, column=idx_booklets).value) != _norm(sheet_booklets_value):
            continue
        if _norm(ws.cell(row=r, column=idx_mm).value) != "boq_exact":
            continue

        # 1) Backfill ef_value/ef_unit if missing
        ef_value_cell = ws.cell(row=r, column=idx_val)
        ef_unit_cell = ws.cell(row=r, column=idx_unit) if idx_unit != -1 else None

        ef_val_now = _to_float(ef_value_cell.value)
        if ef_val_now is None:
            ef_id = ws.cell(row=r, column=idx_id).value
            nid = _norm_efid(ef_id)
            if nid is not None:
                meta = ef_lookup.get(nid)
            else:
                meta = None
            if meta:
                ef_value_cell.value = meta.get("ef_value")
                if ef_unit_cell is not None:
                    cur_unit = ef_unit_cell.value
                    if cur_unit is None or str(cur_unit).strip() == "":
                        ef_unit_cell.value = meta.get("ef_unit")
                updated += 1
                ef_val_now = _to_float(ef_value_cell.value)

        # 2) Compute co2e (t) if missing/zero and inputs exist
        current_co2e = _to_float(ws.cell(row=r, column=idx_co2e_t).value)
        if current_co2e is None or current_co2e == 0.0:
            spend = _to_float(ws.cell(row=r, column=idx_spend).value)
            ef_unit_txt = str(ef_unit_cell.value) if ef_unit_cell is not None and ef_unit_cell.value is not None else None
            out = _compute_tco2e(spend, ef_val_now, ef_unit_txt)
            if out is not None and out != 0.0:
                ws.cell(row=r, column=idx_co2e_t).value = out
                updated += 1

    output_path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(output_path)
    return updated


def main() -> None:
    ap = argparse.ArgumentParser(
        description="Fix BIMMS rows in S3 Cat 1 Purchased G&S where Sheet_booklets=Scope 3 Cat 1 Services Spend and match_method=boq_exact by backfilling ef_value/ef_unit from EF workbook via ef_id."
    )
    ap.add_argument("--input", required=True, help="Path to workbook (.xlsx)")
    ap.add_argument("--output", help="Path to write updated workbook (.xlsx). If omitted, edits in place.")
    ap.add_argument(
        "--ef-workbook",
        default=str(STAGE2_EF_XLSX),
        help="Path to EF workbook (default: merkezi config'teki EF workbook yolu)",
    )
    ap.add_argument("--sheet", default="S3 Cat 1 Purchased G&S", help="Sheet name to modify")
    ap.add_argument("--company", default="BIMMS", help="Company value to target (default: BIMMS)")
    ap.add_argument(
        "--sheet-booklets",
        default="Scope 3 Cat 1 Services Spend",
        help="Sheet_booklets value to target (default: Scope 3 Cat 1 Services Spend)",
    )
    ap.add_argument("--no-backup", action="store_true", help="Do not create a backup before editing in-place.")
    args = ap.parse_args()

    inp = Path(args.input)
    if not inp.exists():
        raise SystemExit(f"Input not found: {inp}")

    efp = Path(args.ef_workbook)
    if not efp.is_absolute():
        efp = Path(__file__).resolve().parent / efp
    if not efp.exists():
        raise SystemExit(f"EF workbook not found: {efp}")

    out = Path(args.output) if args.output else inp

    if (out.resolve() == inp.resolve()) and (not args.no_backup):
        backup_path = inp.with_name(
            f"{inp.stem}_BACKUP_before_bimms_ef_backfill_{datetime.now().strftime('%Y%m%d_%H%M%S')}{inp.suffix}"
        )
        shutil.copy2(inp, backup_path)
        print(f"Backup created: {backup_path}")

    try:
        n = apply_fix(
            inp,
            out,
            ef_workbook=efp,
            sheet_name=args.sheet,
            company=args.company,
            sheet_booklets_value=args.sheet_booklets,
        )
        print(f"Updated rows: {n}")
        print(f"Wrote: {out}")
    except PermissionError as e:
        fallback = inp.with_name(f"{inp.stem}_BIMMS_EF_BACKFILL{inp.suffix}")
        n = apply_fix(
            inp,
            fallback,
            ef_workbook=efp,
            sheet_name=args.sheet,
            company=args.company,
            sheet_booklets_value=args.sheet_booklets,
        )
        print(f"[WARN] Could not overwrite (file locked): {e}")
        print(f"Updated rows: {n}")
        print(f"Wrote fallback: {fallback}")


if __name__ == "__main__":
    main()

