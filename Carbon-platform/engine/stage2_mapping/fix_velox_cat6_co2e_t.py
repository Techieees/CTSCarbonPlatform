from __future__ import annotations

import argparse
import math
import shutil
from datetime import datetime
from pathlib import Path
from typing import Optional, Tuple

from openpyxl import load_workbook


def _norm(s: object) -> str:
    try:
        return str(s).strip().lower()
    except Exception:
        return ""


def _to_float(v: object) -> Optional[float]:
    if v is None:
        return None
    if isinstance(v, (int, float)):
        try:
            fv = float(v)
            return fv if math.isfinite(fv) else None
        except Exception:
            return None
    s = str(v).strip()
    if s == "":
        return None
    # handle "1 234,56" / "1234,56" / "1,234.56"
    s = s.replace(" ", "")
    if "," in s and "." in s:
        # assume commas are thousands separators
        s = s.replace(",", "")
    elif "," in s and "." not in s:
        # assume comma is decimal separator
        s = s.replace(",", ".")
    try:
        fv = float(s)
        return fv if math.isfinite(fv) else None
    except Exception:
        return None


def _find_col_indices(headers: list[object]) -> Tuple[int, int, int]:
    """Return 1-based column indices for (Company, CO2e (kg), co2e (t)).

    Note: Workbook may contain both 'CO2e (kg)' and 'co2e (kg)'. We prefer the first
    matching header, which in this file corresponds to the travel kg column.
    """
    idx_company = idx_kg = idx_t = -1
    for i, h in enumerate(headers, start=1):
        key = _norm(h)
        if key == "company":
            idx_company = i
        elif key == "co2e (kg)":
            # Prefer the exact "CO2e (kg)" header; if duplicates exist, first match wins.
            if idx_kg == -1:
                idx_kg = i
        elif key == "co2e (t)":
            idx_t = i
    if idx_company == -1 or idx_kg == -1 or idx_t == -1:
        raise RuntimeError(
            f"Required columns not found. Found indices company={idx_company}, kg={idx_kg}, t={idx_t}"
        )
    return idx_company, idx_kg, idx_t


def apply_fix(input_path: Path, output_path: Path, sheet_name: str = "S3 Cat 6 Business Travel") -> int:
    # IMPORTANT: data_only=False to preserve formulas in other sheets when saving.
    wb = load_workbook(input_path, data_only=False)
    if sheet_name not in wb.sheetnames:
        raise RuntimeError(f"Sheet not found: {sheet_name}")
    ws = wb[sheet_name]

    # Read header row
    headers = [c.value for c in ws[1]]
    idx_company, idx_kg, idx_t = _find_col_indices(headers)

    updated = 0
    # Iterate rows (skip header)
    for r in range(2, ws.max_row + 1):
        company_val = ws.cell(row=r, column=idx_company).value
        if _norm(company_val) != "velox":
            continue
        kg_val = _to_float(ws.cell(row=r, column=idx_kg).value)
        if kg_val is None:
            continue
        ws.cell(row=r, column=idx_t).value = kg_val / 1000.0
        updated += 1

    # Save to new file (leave original intact)
    output_path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(output_path)
    return updated


def main() -> None:
    ap = argparse.ArgumentParser(description="Fix Velox Cat6 co2e(t) from CO2e(kg)/1000 in S3 Cat 6 Business Travel sheet.")
    ap.add_argument("--input", required=True, help="Path to mapped_results_window workbook (.xlsx)")
    ap.add_argument("--output", help="Path to write updated workbook (.xlsx). If omitted, edits in place.")
    ap.add_argument("--sheet", default="S3 Cat 6 Business Travel", help="Sheet name to modify")
    ap.add_argument(
        "--no-backup",
        action="store_true",
        help="Do not create a backup copy before editing (NOT recommended).",
    )
    args = ap.parse_args()

    inp = Path(args.input)
    if not inp.exists():
        raise SystemExit(f"Input not found: {inp}")

    # In-place if output not provided
    out = Path(args.output) if args.output else inp

    # Backup is optional; default ON so user can revert easily.
    if not args.no_backup:
        backup_path = inp.with_name(
            f"{inp.stem}_BACKUP_before_velox_fix_{datetime.now().strftime('%Y%m%d_%H%M%S')}{inp.suffix}"
        )
        shutil.copy2(inp, backup_path)
        print(f"Backup created: {backup_path}")

    try:
        n = apply_fix(inp, out, sheet_name=args.sheet)
        print(f"Updated rows: {n}")
        print(f"Wrote: {out}")
        return
    except PermissionError as e:
        # Most common reason: the workbook is open in Excel (file lock on Windows).
        # Fallback to writing a new file next to it.
        fallback = inp.with_name(f"{inp.stem}_VELOX_FIX{inp.suffix}")
        n = apply_fix(inp, fallback, sheet_name=args.sheet)
        print(f"[WARN] Could not overwrite (file locked): {e}")
        print(f"Updated rows: {n}")
        print(f"Wrote fallback: {fallback}")

if __name__ == "__main__":
    main()

# bu son yaptigin fix_Velox_cat6_co2e_t.py kodunu da Run_everything.py kismina da entegre eder misin. Herseyi baslat dedigimde bu kisimda baslamali yani. Cok tesekkur ederim.