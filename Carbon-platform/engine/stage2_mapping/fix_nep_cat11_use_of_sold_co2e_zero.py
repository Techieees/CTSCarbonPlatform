from __future__ import annotations

import argparse
import math
import shutil
from datetime import datetime
from pathlib import Path
from typing import Optional, Tuple

from openpyxl import load_workbook


TARGET_STATUS = "NEP SWB emissions rolled into NordicEPOD; set to 0"


def _norm(s: object) -> str:
    try:
        return str(s).strip().lower()
    except Exception:
        return ""


def _norm_status(s: object) -> str:
    # Normalize spaces/case for robust comparisons
    v = _norm(s)
    while "  " in v:
        v = v.replace("  ", " ")
    return v


def _to_float(v: object) -> Optional[float]:
    if v is None:
        return None
    if isinstance(v, (int, float)):
        try:
            fv = float(v)
            return fv if math.isfinite(fv) else None
        except Exception:
            return None
    s = str(v).strip()
    if s == "":
        return None
    s = s.replace(" ", "")
    if "," in s and "." in s:
        s = s.replace(",", "")
    elif "," in s and "." not in s:
        s = s.replace(",", ".")
    try:
        fv = float(s)
        return fv if math.isfinite(fv) else None
    except Exception:
        return None


def _find_sheet_name(wb, preferred: str) -> str:
    if preferred in wb.sheetnames:
        return preferred
    pref_norm = _norm(preferred)
    for name in wb.sheetnames:
        if _norm(name) == pref_norm:
            return name
    # fallback: contains match (handles minor naming differences)
    for name in wb.sheetnames:
        n = _norm(name)
        if "cat 11" in n and "use of sold" in n:
            return name
    raise RuntimeError(f"Sheet not found: {preferred}")


def _find_col_indices(headers: list[object]) -> Tuple[int, int, int]:
    """Return 1-based column indices for (Company, Status, co2e (t)).

    We require 'Company' and 'Status'. For emissions, we prefer 'co2e (t)' but fall back to 'co2e'.
    """
    idx_company = idx_status = idx_co2e = -1
    for i, h in enumerate(headers, start=1):
        key = _norm(h)
        if key == "company":
            idx_company = i
        elif key == "status":
            idx_status = i
        elif key == "co2e (t)":
            idx_co2e = i
    if idx_co2e == -1:
        for i, h in enumerate(headers, start=1):
            if _norm(h) == "co2e":
                idx_co2e = i
                break
    if idx_company == -1 or idx_status == -1 or idx_co2e == -1:
        raise RuntimeError(
            "Required columns not found. "
            f"Found indices company={idx_company}, status={idx_status}, co2e={idx_co2e}"
        )
    return idx_company, idx_status, idx_co2e


def apply_fix(
    input_path: Path,
    output_path: Path,
    sheet_name: str = "S3 Cat 11 Use of Sold",
    company_exact: str = "NEP Switchboards",
    status_exact: str = TARGET_STATUS,
) -> int:
    # IMPORTANT: data_only=False to preserve formulas in other sheets when saving.
    wb = load_workbook(input_path, data_only=False)
    actual_sheet = _find_sheet_name(wb, sheet_name)
    ws = wb[actual_sheet]

    headers = [c.value for c in ws[1]]
    idx_company, idx_status, idx_co2e = _find_col_indices(headers)

    company_norm = _norm(company_exact)
    status_norm = _norm_status(status_exact)

    updated = 0
    for r in range(2, ws.max_row + 1):
        comp_val = ws.cell(row=r, column=idx_company).value
        if _norm(comp_val) != company_norm:
            continue
        st_val = ws.cell(row=r, column=idx_status).value
        if _norm_status(st_val) != status_norm:
            continue

        # Set co2e to 0.0 (even if it's text/blank/formula)
        existing = _to_float(ws.cell(row=r, column=idx_co2e).value)
        if existing is None or existing != 0.0:
            ws.cell(row=r, column=idx_co2e).value = 0.0
            updated += 1

    output_path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(output_path)
    return updated


def main() -> None:
    ap = argparse.ArgumentParser(
        description="Window post-fix: if Company=='NEP Switchboards' AND Status indicates NEP rolled into NordicEPOD, set co2e (t) to 0 in S3 Cat 11 Use of Sold."
    )
    ap.add_argument("--input", required=True, help="Path to mapped_results_window workbook (.xlsx)")
    ap.add_argument("--output", help="Path to write updated workbook (.xlsx). If omitted, edits in place.")
    ap.add_argument("--sheet", default="S3 Cat 11 Use of Sold", help="Sheet name to modify")
    ap.add_argument("--company", default="NEP Switchboards", help="Company value to match (exact after trim)")
    ap.add_argument("--status", default=TARGET_STATUS, help="Status value to match (exact after normalization)")
    ap.add_argument(
        "--no-backup",
        action="store_true",
        help="Do not create a backup copy before editing (NOT recommended).",
    )
    args = ap.parse_args()

    inp = Path(args.input)
    if not inp.exists():
        raise SystemExit(f"Input not found: {inp}")

    out = Path(args.output) if args.output else inp

    if not args.no_backup:
        backup_path = inp.with_name(
            f"{inp.stem}_BACKUP_before_nep_cat11_fix_{datetime.now().strftime('%Y%m%d_%H%M%S')}{inp.suffix}"
        )
        shutil.copy2(inp, backup_path)
        print(f"Backup created: {backup_path}")

    try:
        n = apply_fix(
            inp,
            out,
            sheet_name=args.sheet,
            company_exact=args.company,
            status_exact=args.status,
        )
        print(f"Updated rows: {n}")
        print(f"Wrote: {out}")
    except PermissionError as e:
        fallback = inp.with_name(f"{inp.stem}_NEP_CAT11_FIX{inp.suffix}")
        n = apply_fix(
            inp,
            fallback,
            sheet_name=args.sheet,
            company_exact=args.company,
            status_exact=args.status,
        )
        print(f"[WARN] Could not overwrite (file locked): {e}")
        print(f"Updated rows: {n}")
        print(f"Wrote fallback: {fallback}")


if __name__ == "__main__":
    main()

