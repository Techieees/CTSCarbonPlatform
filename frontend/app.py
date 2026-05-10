import warnings
import importlib.util
warnings.filterwarnings("ignore", category=UserWarning, module="openpyxl")
from flask import Flask, render_template, request, redirect, url_for, flash, jsonify, send_file, send_from_directory, Response, session, abort, g
from flask_sqlalchemy import SQLAlchemy
from flask_login import LoginManager, UserMixin, login_user, login_required, logout_user, current_user
from sqlalchemy import and_, case, desc, exists, func, or_, tuple_, update
from sqlalchemy.exc import IntegrityError
from werkzeug.exceptions import RequestEntityTooLarge
from werkzeug.security import generate_password_hash, check_password_hash
from werkzeug.utils import secure_filename
from openpyxl import load_workbook, Workbook
import pandas as pd
import os
import sys
import threading
import subprocess
import uuid
import shutil
from pathlib import Path
from datetime import date, datetime, timedelta
from types import SimpleNamespace
import hashlib
import html
import json
import mimetypes
from collections import defaultdict, Counter
import csv
import difflib
from io import BytesIO, StringIO
import re
import time
import math
import calendar
import ipaddress
import secrets
import smtplib
import ssl
from functools import lru_cache
from urllib import error as urllib_error
from urllib import request as urllib_request
from urllib.parse import quote
from email.message import EmailMessage
from dotenv import load_dotenv
from markupsafe import Markup, escape
from PIL import Image, ImageDraw, ImageFont

try:
    import fitz  # type: ignore
except Exception:
    fitz = None

PROJECT_ROOT = Path(__file__).resolve().parents[1]
if str(PROJECT_ROOT) not in sys.path:
    sys.path.insert(0, str(PROJECT_ROOT))
    
from config import (
    CCC_API_BASE_URL,
    CCC_GET_ENDPOINTS_PATH,
    CCC_API_PAGE_SIZE,
    CCC_PASSWORD,
    CCC_SHEET_MAPPING_PATH,
    CCC_USERNAME,
    DATA_DIR,
    FRONTEND_DB_PATH,
    FRONTEND_INSTANCE_DIR,
    FRONTEND_UPLOAD_DIR,
    MAIL_DEFAULT_SENDER,
    MAIL_PASSWORD,
    MAIL_PORT,
    MAIL_SERVER,
    MAIL_USERNAME,
    PIPELINE_RUNS_DIR,
    PIPELINE_TEMPLATE_DIR,
    PROFILE_PHOTOS_STORAGE_DIR,
    OPENWEATHER_API_KEY,
    PROJECT_ROOT,
    PUBLIC_APP_BASE_URL,
    SECRET_KEY,
    STAGE1_INPUT_BACKUP_DIR,
    STAGE1_INPUT_DIR,
    STAGE1_KLARAKARBON_OUTPUT_DIR,
    STAGE1_KLARAKARBON_UPLOAD_DIR,
    STAGE2_EF_XLSX,
    STAGE2_KLARAKARBON_DIR,
    STAGE2_MAPPING_DIR,
    STAGE2_OUTPUT_DIR,
    STAGE2_TRAVEL_DIR,
)
from company_slug import company_slug
from preprocess_jobs import (
    klarakarbon_entry_headers,
    klarakarbon_company_supported,
    run_klarakarbon_preprocess,
    run_travel_preprocess,
    validate_klarakarbon_uploads,
    validate_travel_upload,
)
from frontend.evidence_processing import MAX_UPLOAD_BYTES
from frontend.services import messaging_service, notification_service, search_service
from frontend.services.presence_utils import is_online_from_last_seen
from frontend.utils.template_registry import (
    TEMPLATE_MODE_2026,
    TEMPLATE_MODE_LEGACY,
    VALID_TEMPLATE_MODES,
    TemplateRegistry,
    normalize_template_mode,
)

APP_DIR = Path(__file__).resolve().parent
INSTANCE_DIR = FRONTEND_INSTANCE_DIR

# Use centralized configuration so file-system paths do not change between
# local development and server deployments.
app = Flask(__name__, instance_path=str(INSTANCE_DIR), instance_relative_config=True)
app.config['SECRET_KEY'] = SECRET_KEY
INSTANCE_DIR.mkdir(parents=True, exist_ok=True)
app.config['SQLALCHEMY_DATABASE_URI'] = 'sqlite:///' + str(FRONTEND_DB_PATH)
app.config['SQLALCHEMY_TRACK_MODIFICATIONS'] = False
app.config['UPLOAD_FOLDER'] = str(FRONTEND_UPLOAD_DIR)
app.config["MAX_CONTENT_LENGTH"] = MAX_UPLOAD_BYTES

app.logger.info("[PROFILE_PHOTO_STORAGE] Using storage path: %s", PROFILE_PHOTOS_STORAGE_DIR)

TEMPLATES_2026_PATH = APP_DIR / "data" / "templates2026.json"
PRODUCTS_TEMPLATE_PATH = APP_DIR / "data" / "templates_products.json"
ISO_COUNTRIES_PATH = APP_DIR / "data" / "iso_countries.json"
EMPLOYEE_COMMUTING_DATA_DIR = STAGE2_MAPPING_DIR / "Employee Commuting National Averages"
EMPLOYEE_COMMUTING_NATIONAL_AVERAGES_XLSX = (
    EMPLOYEE_COMMUTING_DATA_DIR / "Employee_headcount_national_averages.xlsx"
)
EMPLOYEE_COMMUTING_HEADCOUNT_CSV = (
    EMPLOYEE_COMMUTING_DATA_DIR / "employee_commuting_headcount.csv"
)
EMPLOYEE_COMMUTING_HEADCOUNT_FIELDS: tuple[dict[str, str], ...] = (
    {"key": "company_name", "label": "Company_Name", "input_type": "text"},
    {"key": "headcount", "label": "Headcount", "input_type": "number"},
)
EMPLOYEE_COMMUTING_NATIONAL_AVERAGE_FIELDS: tuple[dict[str, str], ...] = (
    {"key": "company_name", "label": "Company_Name", "input_type": "text"},
    {"key": "country", "label": "Country", "input_type": "text"},
    {"key": "average_one_day", "label": "Average one day", "input_type": "number"},
    {"key": "car_pct", "label": "Car %", "input_type": "number"},
    {"key": "bus_pct", "label": "Bus %", "input_type": "number"},
    {"key": "walking_and_cycling_pct", "label": "Walking and Cycling %", "input_type": "number"},
    {"key": "mixed_pct", "label": "Mixed %", "input_type": "number"},
)
ALLOWED_EMAIL_DOMAIN = "cts-nordics.com"
PROFILE_PHOTO_ALLOWED_EXT = frozenset({".png", ".jpg", ".jpeg", ".webp"})
_PROFILE_PHOTOS_FS_MIGRATED = False


def _migrate_profile_photos_to_storage_once() -> None:
    """Move legacy uploads from frontend/static/profile_photos into storage/profile_photos (once per process)."""
    global _PROFILE_PHOTOS_FS_MIGRATED
    if _PROFILE_PHOTOS_FS_MIGRATED:
        return
    try:
        legacy_dir = APP_DIR / "static" / "profile_photos"
        if legacy_dir.is_dir():
            PROFILE_PHOTOS_STORAGE_DIR.mkdir(parents=True, exist_ok=True)
            for p in list(legacy_dir.iterdir()):
                if not p.is_file():
                    continue
                dest = PROFILE_PHOTOS_STORAGE_DIR / p.name
                if dest.exists():
                    continue
                try:
                    shutil.move(str(p), str(dest))
                except OSError:
                    try:
                        shutil.copy2(str(p), str(dest))
                        p.unlink(missing_ok=True)
                    except Exception:
                        pass
    except Exception:
        pass
    finally:
        _PROFILE_PHOTOS_FS_MIGRATED = True


TRAVEL_ALLOWED_EXT = frozenset({".xlsb", ".xlsx"})
FEED_IMAGE_ALLOWED_EXT = frozenset({".png", ".jpg", ".jpeg", ".webp", ".gif"})
FEED_VIDEO_ALLOWED_EXT = frozenset({".mp4", ".webm", ".mov", ".m4v"})
FEED_FILE_ALLOWED_EXT = frozenset({".pdf", ".doc", ".docx", ".xls", ".xlsx", ".ppt", ".pptx", ".csv", ".txt"})
MODULE_DOCUMENT_ALLOWED_EXT = frozenset({".pdf", ".docx"})
AWARDS_HEADER_IMAGE_ALLOWED_EXT = frozenset({".png", ".jpg", ".jpeg"})
FEED_POST_TYPES: tuple[str, ...] = ("update", "report", "newsletter", "event", "award", "alert")
FEED_POST_TYPES_SET = frozenset(FEED_POST_TYPES)
FEED_FILTER_OPTIONS: tuple[str, ...] = ("all",) + FEED_POST_TYPES
FEED_COMPOSER_TYPES: tuple[str, ...] = ("update", "report", "alert")
REPORT_PREVIEWABLE_EXT = frozenset({".pdf", ".doc", ".docx"})
REPORT_DOCUMENT_EXT = frozenset({".doc", ".docx"})
REPORT_PREVIEW_PAGE_COUNT = 3
FEED_REFERENCE_TYPES = frozenset({"report", "newsletter", "event", "award", "challenge", "challenge_response"})
FEED_REACTION_OPTIONS: tuple[dict[str, str], ...] = (
    {"type": "like", "label": "Like", "icon": "👍"},
    {"type": "celebrate", "label": "Celebrate", "icon": "👏"},
    {"type": "support", "label": "Support", "icon": "❤️"},
    {"type": "insightful", "label": "Insightful", "icon": "💡"},
    {"type": "funny", "label": "Funny", "icon": "😂"},
)
FEED_REACTION_META: dict[str, dict[str, str]] = {
    item["type"]: item for item in FEED_REACTION_OPTIONS
}
FEED_REACTION_TYPES: tuple[str, ...] = tuple(item["type"] for item in FEED_REACTION_OPTIONS)
FEED_REACTION_TYPES_SET = frozenset(FEED_REACTION_TYPES)
AVERAGES_WASTE_TYPES: tuple[str, ...] = ("General", "Plastic", "Metal", "Paper", "Organic", "Hazardous")
AVERAGES_WASTE_UNITS: tuple[str, ...] = ("kg", "tonnes", "lbs")
SCENARIO_COMPANY_OPTIONS: tuple[str, ...] = ("GT Nordics", "Nordic EPOD", "DC Piping")
SCENARIO_CATEGORY_CONFIG: dict[str, tuple[str, ...]] = {
    "GT Nordics": ("9", "11", "12"),
    "Nordic EPOD": ("9", "11", "12"),
    "DC Piping": ("9", "12"),
}
TEMPLATE_MODE_OPTIONS: tuple[str, ...] = (TEMPLATE_MODE_LEGACY, TEMPLATE_MODE_2026)
BUSINESS_TYPE_OPTIONS: tuple[str, ...] = (
    "Service provider",
    "Manufacturer",
    "Construction",
    "Execution",
)
HEATING_SOURCE_OPTIONS: tuple[str, ...] = (
    "District heating",
    "Electricity",
    "Gas",
    "Fuel",
)
TRAVEL_PROVIDER_OPTIONS: tuple[tuple[str, str], ...] = (
    ("", "Select option"),
    ("yes", "Yes"),
    ("no", "No"),
)
AWARDS_QUESTION_TYPES: tuple[str, ...] = ("text", "textarea", "single_choice", "file")
AWARDS_QUESTION_TYPES_SET = frozenset(AWARDS_QUESTION_TYPES)
AUDIT_2025_COLUMNS: tuple[str, ...] = (
    "Month",
    "Company",
    "Scope 3 Category 1 Purchased Goods & Services",
    "Scope 3 Cat 11 Use of Sold of Products",
    "Scope 3 Category 12 End of Life",
    "S3 Cat 3 FERA",
    "Scope 3 Category 4 Upstream Transportation",
    "Scope 3 Category 5 Waste",
    "Scope 3 Category 6 Business Travel",
    "S3 Cat 7 Employee Commute",
    "Scope 3 Category 9 Downstream Transportation",
    "Scope 1",
    "Scope 2",
    "Row Total (t)",
    "Company Share in Total (%)",
    "Company Share in Month (%)",
)
AUDIT_2025_CATEGORY_COLUMNS: tuple[str, ...] = (
    "Scope 3 Category 1 Purchased Goods & Services",
    "Scope 3 Cat 11 Use of Sold of Products",
    "Scope 3 Category 12 End of Life",
    "S3 Cat 3 FERA",
    "Scope 3 Category 4 Upstream Transportation",
    "Scope 3 Category 5 Waste",
    "Scope 3 Category 6 Business Travel",
    "S3 Cat 7 Employee Commute",
    "Scope 3 Category 9 Downstream Transportation",
    "Scope 1",
    "Scope 2",
)
AUDIT_2025_NUMERIC_COLUMNS: tuple[str, ...] = AUDIT_2025_CATEGORY_COLUMNS + (
    "Row Total (t)",
    "Company Share in Total (%)",
    "Company Share in Month (%)",
)
AUDIT_2025_WORKBOOK_PATH = PROJECT_ROOT / "engine" / "stage2_mapping" / "audit_2025" / "Audit 2025 Total Tables.xlsx"
_AUDIT_2025_CACHE: dict[str, object] = {
    "cache_key": None,
    "payload": None,
}
OPERATING_SITE_TYPE_OPTIONS: tuple[str, ...] = ("office", "factory", "warehouse", "other")
STAGE2_2026_SHEET_ALIASES: dict[str, str] = {
    "Scope 3 Category 1 Purchased Goods & Services": "Scope 3 Cat 1 Goods Spend",
    "Scope 3 Category 6 Business Travel": "Scope 3 Cat 6 Business Travel",
    "Scope 3 Category 9 Downstream Transportation": "Scope 3 Cat 4+9 Transport Spend",
    "Scope 3 Category 12 End of Life": "Scope 3 Cat 12 End of Life",
}


def _email_domain_allowed(email: str) -> bool:
    e = (email or "").strip().lower()
    if e.count("@") != 1:
        return False
    local, domain = e.split("@", 1)
    return bool(local) and domain == ALLOWED_EMAIL_DOMAIN


def _hash_password_reset_token(raw_token: str) -> str:
    return hashlib.sha256((raw_token or "").encode("utf-8")).hexdigest()


def _send_plain_email(to_addr: str, subject: str, body: str) -> bool:
    if not MAIL_SERVER:
        app.logger.warning("Password reset: MAIL_SERVER not configured; email not sent to %s", to_addr)
        return False
    try:
        msg = EmailMessage()
        msg["Subject"] = subject
        msg["From"] = MAIL_DEFAULT_SENDER
        msg["To"] = to_addr
        msg.set_content(body)
        ctx = ssl.create_default_context()
        if MAIL_PORT == 465:
            with smtplib.SMTP_SSL(MAIL_SERVER, MAIL_PORT, context=ctx, timeout=30) as smtp:
                if MAIL_USERNAME:
                    smtp.login(MAIL_USERNAME, MAIL_PASSWORD)
                smtp.send_message(msg)
        else:
            with smtplib.SMTP(MAIL_SERVER, MAIL_PORT, timeout=30) as smtp:
                smtp.ehlo()
                smtp.starttls(context=ctx)
                smtp.ehlo()
                if MAIL_USERNAME:
                    smtp.login(MAIL_USERNAME, MAIL_PASSWORD)
                smtp.send_message(msg)
        return True
    except Exception:
        app.logger.exception("Password reset email failed for %s", to_addr)
        return False


_FORGOT_PW_WINDOW_SEC = 600
_FORGOT_PW_MAX_PER_IP = 5
_FORGOT_PW_MAX_PER_EMAIL = 3
_forgot_pw_lock = threading.Lock()
_forgot_pw_ip_ts: dict[str, list[float]] = {}
_forgot_pw_email_ts: dict[str, list[float]] = {}


def _client_ip_for_rate_limit() -> str:
    xff = (request.headers.get("X-Forwarded-For") or "").split(",")[0].strip()
    if xff:
        return xff[:200]
    return str(request.remote_addr or "unknown")[:200]


def _prune_ts_list(ts_list: list[float], now: float) -> None:
    cutoff = now - _FORGOT_PW_WINDOW_SEC
    ts_list[:] = [t for t in ts_list if t > cutoff]


def _forgot_pw_rate_allow_ip(client_ip: str) -> bool:
    now = time.time()
    with _forgot_pw_lock:
        lst = _forgot_pw_ip_ts.setdefault(client_ip, [])
        _prune_ts_list(lst, now)
        if len(lst) >= _FORGOT_PW_MAX_PER_IP:
            return False
        lst.append(now)
        return True


def _forgot_pw_rate_allow_email(email_norm: str) -> bool:
    now = time.time()
    with _forgot_pw_lock:
        lst = _forgot_pw_email_ts.setdefault(email_norm, [])
        _prune_ts_list(lst, now)
        if len(lst) >= _FORGOT_PW_MAX_PER_EMAIL:
            return False
        lst.append(now)
        return True


def _audit_password_reset(event: str, **kwargs: object) -> None:
    extra = " ".join(f"{k}={kwargs[k]}" for k in sorted(kwargs))
    app.logger.info("[AUDIT] %s %s", event, extra)


def _password_reset_meets_policy(pw: str) -> bool:
    if len(pw) < 8:
        return False
    if not re.search(r"[A-Za-z]", pw):
        return False
    if not re.search(r"\d", pw):
        return False
    return True


def _delete_expired_password_reset_tokens() -> None:
    expired = PasswordResetToken.query.filter(PasswordResetToken.expires_at < datetime.utcnow()).all()
    for row in expired:
        db.session.delete(row)


def _load_iso_countries() -> list[tuple[str, str]]:
    try:
        with ISO_COUNTRIES_PATH.open("r", encoding="utf-8") as f:
            raw = json.load(f)
    except Exception:
        return []
    out: list[tuple[str, str]] = []
    if not isinstance(raw, list):
        return out
    for row in raw:
        if not isinstance(row, dict):
            continue
        code = str(row.get("code") or "").strip().upper()
        name = str(row.get("name") or "").strip()
        if code and name:
            out.append((code, name))
    out.sort(key=lambda x: x[1].casefold())
    return out


ISO_COUNTRIES = _load_iso_countries()
ISO_COUNTRY_NAME_BY_CODE = {code: name for code, name in ISO_COUNTRIES}
ISO_COUNTRY_CODE_BY_NAME = {name.casefold(): code for code, name in ISO_COUNTRIES}


def _normalize_template_key(value: str) -> str:
    return " ".join(str(value or "").strip().lower().split())


TEMPLATE_REGISTRY = TemplateRegistry(
    templates2026_path=TEMPLATES_2026_PATH,
)
print("Templates loaded:", len(TEMPLATE_REGISTRY.templates_2026))

KLARAKARBON_SHEET_NAME = "Klarakarbon"
TRAVEL_SHEET_NAME = "Travel"
KLARAKARBON_UPLOAD_COMPANIES = ("Fortica", "Gapit", "GT Nordics", "NEP Switchboards", "NordicEPOD")


def _template_mode_from_request() -> str | None:
    raw = (
        request.args.get("template_mode")
        or request.form.get("template_mode")
        or ((request.get_json(silent=True) or {}).get("template_mode") if request.is_json else None)
    )
    if raw is None:
        return None
    mode = normalize_template_mode(raw)
    if str(raw).strip() not in VALID_TEMPLATE_MODES:
        return None
    return mode


def _current_template_mode() -> str:
    requested = _template_mode_from_request()
    if requested:
        return requested
    session_mode = normalize_template_mode(session.get("template_mode"))
    if str(session.get("template_mode") or "").strip() in VALID_TEMPLATE_MODES:
        return session_mode
    user_mode = normalize_template_mode(getattr(current_user, "template_mode", None))
    if str(getattr(current_user, "template_mode", "") or "").strip() in VALID_TEMPLATE_MODES:
        return user_mode
    return TEMPLATE_MODE_LEGACY


def _persist_template_mode(mode: object) -> str:
    resolved = normalize_template_mode(mode)
    session["template_mode"] = resolved
    return resolved


def _operating_locations_from_json(raw: object) -> list[dict[str, str]]:
    try:
        rows = json.loads(str(raw or "[]"))
    except Exception:
        rows = []
    if not isinstance(rows, list):
        return []
    out: list[dict[str, str]] = []
    valid_country_codes = {code for code, _name in ISO_COUNTRIES}
    valid_site_types = set(OPERATING_SITE_TYPE_OPTIONS)
    for item in rows:
        if not isinstance(item, dict):
            continue
        country = str(item.get("country") or "").strip().upper()
        site_type = str(item.get("site_type") or "").strip().lower()
        if not country and not site_type:
            continue
        if country and country not in valid_country_codes:
            continue
        if site_type and site_type not in valid_site_types:
            continue
        out.append({"country": country, "site_type": site_type})
    return out


def _operating_locations_for_user(u: object) -> list[dict[str, str]]:
    return _operating_locations_from_json(getattr(u, "operating_locations_json", None))


def _profile_template_context(
    *,
    companies: list[str] | None = None,
    resolved_company: str = "",
) -> dict[str, object]:
    return {
        "companies": companies or [],
        "resolved_company": resolved_company,
        "iso_countries": ISO_COUNTRIES,
        "template_mode_options": TEMPLATE_MODE_OPTIONS,
        "business_type_options": BUSINESS_TYPE_OPTIONS,
        "heating_source_options": HEATING_SOURCE_OPTIONS,
        "travel_provider_options": TRAVEL_PROVIDER_OPTIONS,
        "operating_site_type_options": OPERATING_SITE_TYPE_OPTIONS,
        "current_template_mode": _current_template_mode(),
        "operating_locations_initial": _operating_locations_for_user(current_user),
    }


def _current_profile_payload() -> dict[str, object]:
    return {
        "company_name": (getattr(current_user, "company_name", None) or "").strip(),
        "business_type": (getattr(current_user, "business_type", None) or "").strip(),
        "product_type": (getattr(current_user, "product_type", None) or "").strip(),
        "heating_source": (getattr(current_user, "heating_source", None) or "").strip(),
        "travel_provider": (getattr(current_user, "travel_provider", None) or "").strip(),
        "template_mode": _current_template_mode(),
    }


def _template_bundle_for_company(company_name: str) -> dict[str, object]:
    return TEMPLATE_REGISTRY.get_bundle(
        template_mode=_current_template_mode(),
        company_name=company_name,
        profile=_current_profile_payload(),
    )


def _stage2_sheet_name_for_run(sheet_name: str, template_mode: str) -> str:
    if normalize_template_mode(template_mode) != TEMPLATE_MODE_2026:
        return str(sheet_name or "").strip()
    return STAGE2_2026_SHEET_ALIASES.get(str(sheet_name or "").strip(), str(sheet_name or "").strip())


def _restore_env_var(name: str, original_value: str | None) -> None:
    if original_value is None:
        os.environ.pop(name, None)
    else:
        os.environ[name] = original_value

# ---- Stage2 mapping (web single-company runner) ----
STAGE2_MAPPING_OUTPUT_DIR = STAGE2_OUTPUT_DIR
_STAGE2_MAP_LOCK = threading.Lock()
_MAPPING_RUNS: dict[str, dict[str, object]] = {}  # legacy in-memory (kept for backward compatibility)
jobs: dict[str, dict[str, object]] = {}
job_store = jobs
_JOBS_LOCK = threading.Lock()
_JOB_RETENTION_MINUTES = 240
_SUBPROCESS_TIMEOUT_SECONDS = 300

_EVIDENCE_UPLOAD_GUARD = threading.Lock()
_EVIDENCE_UPLOAD_LOCKS: dict[tuple[str, str], threading.Lock] = {}


def _evidence_company_digest_lock(company: str, digest: str) -> threading.Lock:
    """Serialize evidence uploads per (tenant, content hash) to avoid duplicate rows / races."""
    key = (str(company or "").strip().lower(), str(digest or "").strip().lower())
    with _EVIDENCE_UPLOAD_GUARD:
        lk = _EVIDENCE_UPLOAD_LOCKS.get(key)
        if lk is None:
            lk = threading.Lock()
            _EVIDENCE_UPLOAD_LOCKS[key] = lk
        return lk


def _evidence_log(event: str, **fields: object) -> None:
    """Server-side evidence audit trail (no paths returned to clients)."""
    try:
        parts = " ".join(f"{k}={fields[k]!r}" for k in sorted(fields.keys()))
        print(f"[EVIDENCE] {event} {parts}")
    except Exception:
        pass


class JobCancelled(Exception):
    pass


def _job_timestamp() -> str:
    return datetime.utcnow().isoformat() + "Z"


def _cleanup_old_jobs() -> None:
    cutoff = datetime.utcnow() - timedelta(minutes=_JOB_RETENTION_MINUTES)
    with _JOBS_LOCK:
        stale: list[str] = []
        for job_id, job in jobs.items():
            completed_at = str(job.get("completed_at") or "")
            if not completed_at:
                continue
            try:
                completed_dt = datetime.fromisoformat(completed_at.replace("Z", ""))
            except Exception:
                continue
            if completed_dt < cutoff:
                stale.append(job_id)
        for job_id in stale:
            jobs.pop(job_id, None)


def _job_snapshot(job_id: str) -> dict[str, object] | None:
    with _JOBS_LOCK:
        job = jobs.get(job_id)
        return dict(job) if job is not None else None


def _user_can_access_job(job: dict[str, object] | None, u: object | None) -> bool:
    if not job:
        return False
    if bool(getattr(u, "is_admin", False)) or _is_owner_user(u):
        return True
    return int(job.get("user_id") or 0) == int(getattr(u, "id", 0) or 0)


def _is_job_cancel_requested(job_id: str) -> bool:
    with _JOBS_LOCK:
        return bool(jobs.get(job_id, {}).get("cancel_requested"))


def _raise_if_job_cancelled(job_id: str) -> None:
    if _is_job_cancel_requested(job_id):
        raise JobCancelled()


def _update_job(job_id: str, **updates: object) -> None:
    with _JOBS_LOCK:
        job = jobs.get(job_id)
        if not job:
            return
        updates.setdefault("updated_at", _job_timestamp())
        job.update(updates)


def _update_job_progress(job_id: str, progress: int, message: str | None = None) -> None:
    safe_progress = max(0, min(100, int(progress or 0)))
    payload: dict[str, object] = {"progress": safe_progress}
    if message is not None:
        payload["message"] = message
    _update_job(job_id, **payload)
    print(f"[JOB] Progress {job_id}: {safe_progress}%")


def _serialize_job(job: dict[str, object]) -> dict[str, object]:
    response = {
        "job_id": job.get("job_id"),
        "type": job.get("type"),
        "company": job.get("company"),
        "status": job.get("status"),
        "progress": int(job.get("progress") or 0),
        "message": job.get("message") or "",
        "error": job.get("error"),
        "created_at": job.get("created_at"),
        "updated_at": job.get("updated_at"),
        "started_at": job.get("started_at"),
        "completed_at": job.get("completed_at"),
        "cancel_requested": bool(job.get("cancel_requested")),
    }
    if job.get("rows") is not None:
        try:
            response["rows"] = int(job.get("rows") or 0)
        except Exception:
            response["rows"] = job.get("rows")
    if job.get("result") is not None:
        response["result"] = job.get("result")
    return response


def _active_job_id_for_job_type(job_type: str) -> str | None:
    """job_id if a background job of this type is pending or running (in-process registry only)."""
    jt = str(job_type or "").strip()
    if not jt:
        return None
    with _JOBS_LOCK:
        for job in jobs.values():
            if str(job.get("type") or "") != jt:
                continue
            if str(job.get("status") or "") in {"pending", "running"}:
                jid = str(job.get("job_id") or "").strip()
                return jid or None
    return None


def run_in_background(job_type: str, company: str, target, *args, **kwargs) -> str:
    _cleanup_old_jobs()
    job_id = uuid.uuid4().hex[:12]
    user_id = kwargs.pop("job_user_id", None)
    user_email = kwargs.pop("job_user_email", "")
    now = _job_timestamp()
    with _JOBS_LOCK:
        jobs[job_id] = {
            "job_id": job_id,
            "type": job_type,
            "company": company,
            "status": "pending",
            "progress": 0,
            "message": "Queued",
            "error": None,
            "created_at": now,
            "updated_at": now,
            "started_at": None,
            "completed_at": None,
            "cancel_requested": False,
            "user_id": user_id,
            "user_email": user_email,
            "result": None,
        }
    print(f"[JOB] Created {job_id}")
    print(f"[JOB CREATED] {job_id}")

    def runner() -> None:
        _update_job(job_id, status="running", started_at=_job_timestamp(), message="Running")
        print(f"[JOB] Running {job_id}")
        try:
            _raise_if_job_cancelled(job_id)
            with app.app_context():
                print(f"[JOB CONTEXT] Running job {job_id} without request context")
                result = target(job_id=job_id, *args, **kwargs)
            if _is_job_cancel_requested(job_id):
                _update_job(
                    job_id,
                    status="cancelled",
                    message="Cancelled",
                    completed_at=_job_timestamp(),
                )
                print(f"[JOB] Cancelled {job_id}")
                return
            _update_job(
                job_id,
                status="completed",
                progress=100,
                message="Completed",
                completed_at=_job_timestamp(),
                result=result,
            )
            print(f"[JOB] Completed {job_id}")
        except JobCancelled:
            try:
                with app.app_context():
                    db.session.rollback()
            except Exception:
                pass
            _update_job(
                job_id,
                status="cancelled",
                message="Cancelled",
                completed_at=_job_timestamp(),
            )
            print(f"[JOB] Cancelled {job_id}")
        except Exception as exc:
            try:
                with app.app_context():
                    db.session.rollback()
            except Exception:
                pass
            _update_job(
                job_id,
                status="failed",
                error=str(exc),
                message=str(exc),
                completed_at=_job_timestamp(),
            )
            print(f"[JOB] Failed {job_id}: {exc}")
        finally:
            try:
                with app.app_context():
                    db.session.remove()
            except Exception:
                pass

    threading.Thread(target=runner, daemon=True).start()
    return job_id

# ---- Company canonical names + countries (web mapping) ----
# User-provided source-of-truth list (deduplicated).
_COMPANY_COUNTRY_CANONICAL: dict[str, str] = {
    "BIMMS": "Portugal",
    "CTS Finland": "Finland",
    "CTS-VDC Services": "Ireland",
    "DC Piping": "Portugal",
    "GT Nordics": "Norway",
    "Mecwide Nordics": "Norway",
    "Porvelox": "Portugal",
    "Caerus Nordics": "Norway",
    "CTS Sweden": "Sweden",
    "Navitas Portugal": "Portugal",
    "NEP Switchboards": "Norway",
    "CTS Denmark": "Denmark",
    "CTS Group": "Switzerland",
    "CTS Nordics": "Norway",
    "Fortica": "Norway",
    "MC Prefab": "Sweden",
    "Navitas Norway": "Norway",
    "QEC": "Norway",
    "SD Nordics": "Norway",
    "CTS Security Solutions": "Sweden",
    "Velox": "Norway",
    "CTS EU": "Portugal",
    "Gapit": "Norway",
}


def _canonical_company_name_and_country(name: str) -> tuple[str, str | None]:
    raw = (name or "").strip()
    if not raw:
        return "", None

    def norm(s: str) -> str:
        return "".join(ch.lower() for ch in (s or "") if ch.isalnum())

    n = norm(raw)
    # Build synonyms (file stems vs display names)
    synonyms: dict[str, str] = {
        norm("CTS-VDC"): "CTS-VDC Services",
        norm("CTS VDC"): "CTS-VDC Services",
        norm("CTS-VDC Services"): "CTS-VDC Services",
        norm("CTS Group HQ"): "CTS Group",
        norm("CTS Group"): "CTS Group",
        norm("NordicEPOD"): "Nordic EPOD",
        norm("Nordic EPOD"): "Nordic EPOD",
        norm("Caerus Nordics"): "Caerus Nordics",
        norm("CTS EU"): "CTS EU",
        norm("GT Nordics"): "GT Nordics",
        norm("Mecwide Nordics"): "Mecwide Nordics",
        norm("Navitas Norway"): "Navitas Norway",
        norm("Navitas Portugal"): "Navitas Portugal",
        norm("CTS Denmark"): "CTS Denmark",
        norm("CTS Finland"): "CTS Finland",
        norm("CTS Sweden"): "CTS Sweden",
        norm("CTS Nordics"): "CTS Nordics",
        norm("CTS Security Solutions"): "CTS Security Solutions",
        norm("DC Piping"): "DC Piping",
        norm("MC Prefab"): "MC Prefab",
        norm("Porvelox"): "Porvelox",
        norm("Fortica"): "Fortica",
        norm("Gapit"): "Gapit",
        norm("QEC"): "QEC",
        norm("SD Nordics"): "SD Nordics",
        norm("Velox"): "Velox",
        norm("BIMMS"): "BIMMS",
        norm("NEP Switchboards"): "NEP Switchboards",
        norm("NEP Switchboards AS"): "NEP Switchboards",
    }

    canonical = synonyms.get(n, raw)
    # Country lookup uses canonical keys; if missing, best-effort match by normalized canonical names
    country = _COMPANY_COUNTRY_CANONICAL.get(canonical)
    if country is None:
        want = norm(canonical)
        for k, v in _COMPANY_COUNTRY_CANONICAL.items():
            if norm(k) == want:
                country = v
                canonical = k
                break

    return canonical, country

db = SQLAlchemy(app)
login_manager = LoginManager()
login_manager.init_app(app)
login_manager.login_view = "login"

_EVIDENCE_MAX_SIZE_USER_MSG = "File exceeds maximum allowed size (25MB)."


@app.errorhandler(RequestEntityTooLarge)
def handle_request_entity_too_large(_exc):
    if request.path.startswith("/api/"):
        return jsonify(ok=False, error=_EVIDENCE_MAX_SIZE_USER_MSG), 413
    return (
        "The uploaded file is too large.",
        413,
        {"Content-Type": "text/plain; charset=utf-8"},
    )


_ENDPOINTS_WITHOUT_PROFILE = frozenset(
    {
        "login",
        "logout",
        "register",
        "profile_setup",
        "profile_page",
        "static",
        "mouse_symbol_image",
        "index",
        "who_we_are",
        "platform",
        "methodology",
        "methodology_reference",
        "methodology_scope1",
        "methodology_scope2",
        "methodology_scope3",
        "methodology_scope3_category1",
        "methodology_scope3_category2",
        "methodology_scope3_category3",
        "methodology_scope3_category4",
        "methodology_scope3_category5",
        "methodology_scope3_category6",
        "methodology_scope3_category7",
        "methodology_scope3_category9",
        "methodology_scope3_category10",
        "methodology_scope3_category11",
        "methodology_scope3_category12",
        "methodology_scope3_category13",
        "methodology_scope3_category14",
        "methodology_scope3_category15",
        "methodology_scope3_category16",
        "trust",
        "forgot_password",
        "reset_password",
        "request_access",
        "impact",
        "impact_approach",
        "impact_operations",
        "impact_lca_epd",
        "impact_collaboration",
        "impact_corporate",
        "impact_materiality",
        "impact_carbon",
        "impact_stakeholders",
        "impact_sdgs",
        "impact_governance",
        "impact_reporting",
        "esg",
        "csrd",
        "lca",
        "lca_tool",
        "csrd_policies",
        "csrd_policy_add",
        "csrd_policy_update",
        "csrd_policy_delete",
        "csrd_policy_file",
    }
)


@app.before_request
def _require_complete_profile_for_app():
    if not current_user.is_authenticated:
        return
    if _user_profile_complete(current_user):
        return
    ep = request.endpoint
    if ep in _ENDPOINTS_WITHOUT_PROFILE:
        return
    if str(request.path or "").startswith("/api/profile-photo/"):
        return
    if request.path.startswith("/api/"):
        return jsonify({"error": "Complete profile setup before using this feature."}), 403
    return redirect(url_for("profile_setup"))


@app.before_request
def _require_current_month_products_input():
    # Temporarily disabled while the Products Log flow is being tested.
    # Keep the page/API available, but do not block navigation across the app.
    return
    if not current_user.is_authenticated:
        return
    if not _user_profile_complete(current_user) or _is_readonly_user(current_user):
        return
    endpoint = str(request.endpoint or "").strip()
    if not endpoint or endpoint == "static" or request.path.startswith("/static/"):
        return
    allowed_endpoints = {
        "products_input_page",
        "api_products_input_save",
        "api_products_input_export",
        "profile_setup",
        "logout",
        "mouse_symbol_image",
    }
    if endpoint in allowed_endpoints:
        return
    if endpoint.startswith("api_"):
        return jsonify({"error": "Submit this month's product data before using this feature."}), 403
    try:
        if _products_current_month_has_entry(current_user):
            return
    except Exception:
        return
    flash("Please submit this month's Products Log before continuing.", "warning")
    return redirect(url_for("products_input_page"))


@app.before_request
def _ensure_activity_session_for_authenticated_user():
    if current_user.is_authenticated:
        _activity_session_id(create_if_missing=True)


@app.before_request
def _enforce_readonly_auditor_access():
    if not current_user.is_authenticated or not _is_readonly_user(current_user):
        return
    endpoint = str(request.endpoint or "").strip()
    method = str(request.method or "GET").upper()
    blocked_pages = {
        "dashboard",
        "data_sources_averages",
        "data_sources_scenarios",
    }
    blocked_write_endpoints = {
        "create_feed_post",
        "api_feed_post_reaction",
        "create_challenge",
        "submit_challenge_response",
        "api_follow_user",
        "api_unfollow_user",
        "api_profile_cover",
        "api_feed_post_comment",
        "api_feed_comment_like",
        "profile_page",
        "api_excel_schema_save",
        "api_averages_save",
        "api_scenarios_save",
        "api_evidence_upload",
        "api_evidence_link",
        "api_evidence_unlink",
        "api_mapping_run",
        "api_pipeline_append_run",
        "analytics_forecasting",
        "analytics_decarbonization",
        "analytics_mapped_window_output",
        "analytics_emissions_totals",
        "analytics_share_analysis",
        "governance_audit_ready_output",
        "governance_double_counting_check",
        "data_sources_ccc_api",
    }
    if method == "GET" and endpoint in blocked_pages:
        flash("Auditor accounts have read-only access to feed, reports, and mapped outputs.", "warning")
        return redirect(url_for("feed"))
    if method in {"POST", "PUT", "PATCH", "DELETE"} and endpoint in blocked_write_endpoints:
        if request.path.startswith("/api/"):
            return jsonify({"error": "Auditor accounts have read-only access."}), 403
        flash("Auditor accounts have read-only access.", "warning")
        return redirect(url_for("feed"))


@app.after_request
def _log_authenticated_activity(response: Response):
    try:
        if not current_user.is_authenticated:
            return response
        if request.method.upper() == "OPTIONS":
            return response
        if _is_static_or_ignored_activity_path(request.path):
            return response
        if request.environ.get("skip_activity_log"):
            return response
        _write_activity_log_for_user(current_user, action=_classify_activity_action())
        _touch_user_last_seen_throttled(current_user)
    except Exception:
        pass
    return response


@app.context_processor
def _nav_profile_context():
    def nav_profile_photo_url() -> str | None:
        try:
            if not current_user.is_authenticated:
                return None
            return _profile_photo_url_for_user(current_user)
        except Exception:
            return None

    return dict(nav_profile_photo_url=nav_profile_photo_url)


# Database models
class User(UserMixin, db.Model):
    id = db.Column(db.Integer, primary_key=True)
    email = db.Column(db.String(120), unique=True, nullable=False)
    password_hash = db.Column(db.String(200), nullable=False)
    company_name = db.Column(db.String(200), nullable=False)
    is_admin = db.Column(db.Boolean, default=False)
    # owner | super_admin | admin | manager | user — kept in sync with is_admin via sync_user_admin_flag()
    role = db.Column(db.String(32), nullable=True, default="user")
    created_at = db.Column(db.DateTime, default=datetime.utcnow)
    first_name = db.Column(db.String(100), nullable=True)
    last_name = db.Column(db.String(100), nullable=True)
    job_title = db.Column(db.String(200), nullable=True)
    phone = db.Column(db.String(40), nullable=True)
    company_country = db.Column(db.String(100), nullable=True)
    profile_photo_path = db.Column(db.String(500), nullable=True)
    cover_image = db.Column(db.String(500), nullable=True)
    is_profile_complete = db.Column(db.Boolean, default=False)
    template_mode = db.Column(db.String(20), nullable=True, default=TEMPLATE_MODE_LEGACY)
    business_type = db.Column(db.String(100), nullable=True)
    product_type = db.Column(db.String(100), nullable=True)
    quantity = db.Column(db.String(100), nullable=True)
    quantity_unit = db.Column(db.String(100), nullable=True)
    number_of_products_in_use = db.Column(db.String(100), nullable=True)
    end_use_location = db.Column(db.String(200), nullable=True)
    heating_source = db.Column(db.String(100), nullable=True)
    travel_provider = db.Column(db.String(20), nullable=True)
    operating_locations_json = db.Column(db.Text, nullable=True)
    last_seen_at = db.Column(db.DateTime, nullable=True, index=True)


USER_ROLES: tuple[str, ...] = ("owner", "super_admin", "admin", "manager", "auditor", "user")
USER_ROLES_SET = frozenset(USER_ROLES)
# Roles that may use existing admin routes (is_admin=True)
ROLES_WITH_ADMIN_ACCESS = frozenset({"owner", "super_admin", "admin", "manager"})


def normalize_user_role(raw: str | None) -> str:
    r = (raw or "user").strip().lower()
    return r if r in USER_ROLES_SET else "user"


def _split_full_name(full_name: str) -> tuple[str, str]:
    s = (full_name or "").strip()
    if not s:
        return ("", "")
    parts = s.split(None, 1)
    if len(parts) == 1:
        return (parts[0], "")
    return (parts[0], parts[1])


ACCESS_REQUEST_NOTIFY_EMAIL = "florian.d@cts-nordics.com"
# Owner accounts are not created from the access-request flow (use User roles admin only).
ACCESS_REQUEST_APPROVE_ROLES: tuple[str, ...] = tuple(r for r in USER_ROLES if r != "owner")
ACCESS_REQUEST_APPROVE_ROLES_SET = frozenset(ACCESS_REQUEST_APPROVE_ROLES)


def sync_user_admin_flag(user: "User") -> None:
    """Set role + is_admin from canonical role (call after changing role)."""
    r = normalize_user_role(getattr(user, "role", None))
    user.role = r
    user.is_admin = r in ROLES_WITH_ADMIN_ACCESS


def _user_public_dict_for_admin(u: "User") -> dict:
    """Safe user fields for admin panel JSON (no password)."""
    photo = _profile_photo_url_for_user(u)
    r = normalize_user_role(getattr(u, "role", None))
    raw_role = (getattr(u, "role", None) or "").strip() or r
    fn = (getattr(u, "first_name", None) or "").strip()
    ln = (getattr(u, "last_name", None) or "").strip()
    full = " ".join(x for x in (fn, ln) if x).strip()
    return {
        "id": u.id,
        "email": u.email,
        "first_name": fn or None,
        "last_name": ln or None,
        "full_name": full or None,
        "phone": (getattr(u, "phone", None) or "").strip() or None,
        "company_name": (u.company_name or "").strip(),
        "company_country": (getattr(u, "company_country", None) or "").strip() or None,
        "role": r,
        "role_display": raw_role,
        "profile_photo_url": photo,
        "created_at": u.created_at.strftime("%Y-%m-%d") if getattr(u, "created_at", None) else None,
    }


def _user_display_name(u: User | None) -> str:
    if u is None:
        return ""
    first = (getattr(u, "first_name", None) or "").strip()
    last = (getattr(u, "last_name", None) or "").strip()
    full = " ".join(part for part in (first, last) if part).strip()
    return full or (getattr(u, "email", None) or "").strip() or f"User {getattr(u, 'id', '')}"


def _role_badge_label(raw_role: object) -> str:
    role = normalize_user_role(str(raw_role or ""))
    return {
        "owner": "",
        "super_admin": "",
        "admin": "",
        "manager": "",
        "auditor": "",
        "user": "",
    }.get(role, "")


def _user_role_label(u: User | None) -> str:
    if u is None:
        return ""
    raw_role = str(getattr(u, "role", None) or "").strip()
    if raw_role:
        return _role_badge_label(raw_role)
    return ""


def _user_professional_title(u: User | None) -> str:
    if u is None:
        return ""
    title = (getattr(u, "job_title", None) or "").strip()
    if title:
        return title
    return ""


def _profile_photo_url_for_user(u: User | None) -> str | None:
    if u is None:
        return None
    rel = getattr(u, "profile_photo_path", None)
    if not rel:
        return None
    try:
        uid = int(getattr(u, "id", 0) or 0)
        if uid <= 0:
            return None
        return url_for("api_profile_photo", user_id=uid)
    except Exception:
        return None


def _mapping_card_payload_for_pair(company_name: str, sheet_name: str) -> dict[str, object] | None:
    """Live mapping summary for one Data Entry pair (same shape as api_admin_upload_notifications rows)."""
    company_key = str(company_name or "").strip()
    sheet_key = str(sheet_name or "").strip()
    if not company_key or not sheet_key:
        return None
    try:
        batches = _list_admin_data_entry_batches()
    except Exception:
        return None
    for b in batches:
        if str(b.get("company_name") or "").strip() != company_key:
            continue
        if str(b.get("sheet_name") or "").strip() != sheet_key:
            continue
        uploaded_at = b.get("uploaded_at")
        ts = uploaded_at.strftime("%Y-%m-%d %H:%M") if isinstance(uploaded_at, datetime) else ""
        mapped_at = b.get("mapped_at")
        mapped_ts = mapped_at.strftime("%Y-%m-%d %H:%M") if isinstance(mapped_at, datetime) else ""
        return {
            "company_name": company_key,
            "uploaded_by_user": str(b.get("uploaded_by_user") or "Unknown"),
            "uploaded_by_user_id": int(b.get("uploaded_by_user_id") or 0),
            "uploaded_by_job_title": str(b.get("uploaded_by_job_title") or ""),
            "uploaded_by_has_profile_photo": bool(b.get("uploaded_by_has_profile_photo")),
            "upload_timestamp": ts,
            "category": sheet_key,
            "row_count": int(b.get("row_count") or 0),
            "mapping_status": str(b.get("mapping_status_label") or ""),
            "mapping_state": str(b.get("mapping_state") or ""),
            "mapped_by_admin": str(b.get("mapped_by") or ""),
            "mapping_timestamp": mapped_ts,
            "mapped": bool(b.get("mapped")),
        }
    return None


def _notification_payload(row: "Notification") -> dict[str, object]:
    created_at = row.created_at.strftime("%Y-%m-%d %H:%M") if getattr(row, "created_at", None) else ""
    payload: dict[str, object] = {
        "id": int(row.id),
        "title": row.title,
        "message": row.message,
        "type": row.type,
        "link": row.link,
        "is_read": bool(row.is_read),
        "created_at": created_at,
    }
    raw_meta = getattr(row, "meta_json", None)
    if raw_meta:
        try:
            parsed = json.loads(str(raw_meta))
            if isinstance(parsed, dict):
                payload["mapping_card"] = parsed
        except Exception:
            pass
    return payload


def _message_payload(row: "Message", *, viewer_id: int) -> dict[str, object]:
    sender = User.query.get(int(row.sender_id))
    receiver = User.query.get(int(row.receiver_id))
    return {
        "id": int(row.id),
        "thread_id": row.thread_id,
        "sender_id": int(row.sender_id),
        "receiver_id": int(row.receiver_id),
        "sender_name": _user_display_name(sender),
        "receiver_name": _user_display_name(receiver),
        "message": row.message,
        "created_at": row.created_at.strftime("%Y-%m-%d %H:%M") if getattr(row, "created_at", None) else "",
        "is_read": bool(row.is_read),
        "is_mine": int(row.sender_id) == int(viewer_id),
    }


def _contact_payload(u: "User") -> dict[str, object]:
    return {
        "id": int(u.id),
        "name": _user_display_name(u),
        "email": u.email,
        "job_title": (getattr(u, "job_title", None) or "").strip(),
        "company_name": (getattr(u, "company_name", None) or "").strip(),
        "profile_photo_url": _profile_photo_url_for_user(u),
        "is_online": bool(is_online_from_last_seen(getattr(u, "last_seen_at", None))),
    }


_LAST_SEEN_TOUCH_MIN_INTERVAL = timedelta(seconds=60)


def _touch_user_last_seen_throttled(user: "User") -> None:
    """Persist last_seen_at at most once per minute per user (separate DB transaction)."""
    if user is None or not getattr(user, "id", None):
        return
    try:
        uid = int(user.id)
        now = datetime.utcnow()
        cutoff = now - _LAST_SEEN_TOUCH_MIN_INTERVAL
        tbl = User.__table__
        stmt = (
            update(tbl)
            .where(tbl.c.id == uid)
            .where(or_(tbl.c.last_seen_at.is_(None), tbl.c.last_seen_at < cutoff))
            .values(last_seen_at=now)
        )
        with db.engine.begin() as conn:
            conn.execute(stmt)
    except Exception:
        pass


_MESSAGE_TYPING_STATE_LOCK = threading.Lock()
_MESSAGE_TYPING_STATE: dict[str, dict[str, object]] = {}


def _message_typing_state_key(sender_id: int, receiver_id: int) -> str:
    return f"{int(sender_id)}:{int(receiver_id)}"


def _set_message_typing_state(sender_id: int, receiver_id: int, *, is_typing: bool) -> None:
    key = _message_typing_state_key(sender_id, receiver_id)
    now = time.time()
    with _MESSAGE_TYPING_STATE_LOCK:
        expired_keys = [
            item_key
            for item_key, item in _MESSAGE_TYPING_STATE.items()
            if float(item.get("expires_at") or 0.0) <= now
        ]
        for item_key in expired_keys:
            _MESSAGE_TYPING_STATE.pop(item_key, None)
        if not is_typing:
            _MESSAGE_TYPING_STATE.pop(key, None)
            return
        _MESSAGE_TYPING_STATE[key] = {
            "sender_id": int(sender_id),
            "receiver_id": int(receiver_id),
            "expires_at": now + 1.5,
        }


def _message_typing_status(sender_id: int, receiver_id: int) -> bool:
    key = _message_typing_state_key(sender_id, receiver_id)
    now = time.time()
    with _MESSAGE_TYPING_STATE_LOCK:
        item = _MESSAGE_TYPING_STATE.get(key)
        if not item:
            return False
        if float(item.get("expires_at") or 0.0) <= now:
            _MESSAGE_TYPING_STATE.pop(key, None)
            return False
        return True


class Company(db.Model):
    __tablename__ = "companies"
    id = db.Column(db.Integer, primary_key=True)
    company_name = db.Column(db.String(200), unique=True, nullable=False)
    company_logo_path = db.Column(db.String(500), nullable=True)
    created_by_user_id = db.Column(db.Integer, db.ForeignKey("user.id"), nullable=True)


class AveragesData(db.Model):
    __tablename__ = "averages_data"
    id = db.Column(db.Integer, primary_key=True)
    company_name = db.Column(db.String(200), unique=True, nullable=False, index=True)
    saved_by_user_id = db.Column(db.Integer, db.ForeignKey("user.id"), nullable=True)
    electricity_kwh = db.Column(db.Float, nullable=True)
    electricity_country = db.Column(db.String(120), nullable=True)
    electricity_emission_factor = db.Column(db.Float, nullable=True)
    district_heating_kwh = db.Column(db.Float, nullable=True)
    district_heating_supplier = db.Column(db.String(200), nullable=True)
    waste_type = db.Column(db.String(120), nullable=True)
    waste_weight = db.Column(db.Float, nullable=True)
    waste_unit = db.Column(db.String(40), nullable=True)
    water_total_m3 = db.Column(db.Float, nullable=True)
    building_size_m2 = db.Column(db.Float, nullable=True)
    water_per_m2 = db.Column(db.Float, nullable=True)
    created_at = db.Column(db.DateTime, default=datetime.utcnow, nullable=False)
    updated_at = db.Column(db.DateTime, default=datetime.utcnow, onupdate=datetime.utcnow, nullable=False)


class ScenariosData(db.Model):
    __tablename__ = "scenarios_data"
    id = db.Column(db.Integer, primary_key=True)
    company_name = db.Column(db.String(200), unique=True, nullable=False, index=True)
    saved_by_user_id = db.Column(db.Integer, db.ForeignKey("user.id"), nullable=True)
    categories_json = db.Column(db.Text, nullable=True)
    created_at = db.Column(db.DateTime, default=datetime.utcnow, nullable=False)
    updated_at = db.Column(db.DateTime, default=datetime.utcnow, onupdate=datetime.utcnow, nullable=False)


class Report(db.Model):
    __tablename__ = "report"
    id = db.Column(db.Integer, primary_key=True)
    title = db.Column(db.String(255), nullable=False)
    file_path = db.Column(db.String(500), nullable=False)
    preview_paths = db.Column(db.Text, nullable=True)
    uploaded_by = db.Column(db.Integer, db.ForeignKey("user.id"), nullable=False, index=True)
    company_id = db.Column(db.Integer, db.ForeignKey("companies.id"), nullable=False, index=True)
    category_id = db.Column(db.Integer, db.ForeignKey("reports_categories.id"), nullable=True, index=True)
    created_at = db.Column(db.DateTime, default=datetime.utcnow, nullable=False, index=True)
    uploader = db.relationship("User", lazy="joined")
    company = db.relationship("Company", lazy="joined")
    category = db.relationship("ReportCategory", lazy="joined")


class ReportCategory(db.Model):
    __tablename__ = "reports_categories"
    id = db.Column(db.Integer, primary_key=True)
    name = db.Column(db.String(200), nullable=False, unique=True, index=True)
    created_by = db.Column(db.Integer, db.ForeignKey("user.id"), nullable=False, index=True)
    created_at = db.Column(db.DateTime, default=datetime.utcnow, nullable=False, index=True)
    creator = db.relationship("User", lazy="joined")


class Newsletter(db.Model):
    __tablename__ = "newsletter"
    id = db.Column(db.Integer, primary_key=True)
    title = db.Column(db.String(255), nullable=False)
    file_path = db.Column(db.String(500), nullable=False)
    uploaded_by = db.Column(db.Integer, db.ForeignKey("user.id"), nullable=False, index=True)
    created_at = db.Column(db.DateTime, default=datetime.utcnow, nullable=False, index=True)
    uploader = db.relationship("User", lazy="joined")


class Event(db.Model):
    __tablename__ = "event"
    id = db.Column(db.Integer, primary_key=True)
    title = db.Column(db.String(255), nullable=False)
    description = db.Column(db.Text, nullable=False, default="")
    event_date = db.Column(db.DateTime, nullable=False, index=True)
    created_by = db.Column(db.Integer, db.ForeignKey("user.id"), nullable=False, index=True)
    created_at = db.Column(db.DateTime, default=datetime.utcnow, nullable=False, index=True)
    creator = db.relationship("User", lazy="joined")


class AwardsForm(db.Model):
    __tablename__ = "awards_forms"
    id = db.Column(db.Integer, primary_key=True)
    title = db.Column(db.String(255), nullable=False)
    description = db.Column(db.Text, nullable=False, default="")
    header_image = db.Column(db.String(500), nullable=True)
    created_by = db.Column(db.Integer, db.ForeignKey("user.id"), nullable=False, index=True)
    created_at = db.Column(db.DateTime, default=datetime.utcnow, nullable=False, index=True)
    creator = db.relationship("User", lazy="joined")


class AwardsQuestion(db.Model):
    __tablename__ = "awards_questions"
    id = db.Column(db.Integer, primary_key=True)
    form_id = db.Column(db.Integer, db.ForeignKey("awards_forms.id"), nullable=False, index=True)
    question_text = db.Column(db.Text, nullable=False)
    question_type = db.Column(db.String(40), nullable=False)
    required = db.Column(db.Boolean, nullable=False, default=False)
    options = db.Column(db.Text, nullable=True)
    form = db.relationship("AwardsForm", lazy="joined")


class AwardsSubmission(db.Model):
    __tablename__ = "awards_submissions"
    id = db.Column(db.Integer, primary_key=True)
    form_id = db.Column(db.Integer, db.ForeignKey("awards_forms.id"), nullable=False, index=True)
    submitted_by = db.Column(db.Integer, db.ForeignKey("user.id"), nullable=False, index=True)
    created_at = db.Column(db.DateTime, default=datetime.utcnow, nullable=False, index=True)
    form = db.relationship("AwardsForm", lazy="joined")
    submitter = db.relationship("User", lazy="joined")


class AwardsAnswer(db.Model):
    __tablename__ = "awards_answers"
    id = db.Column(db.Integer, primary_key=True)
    submission_id = db.Column(db.Integer, db.ForeignKey("awards_submissions.id"), nullable=False, index=True)
    question_id = db.Column(db.Integer, db.ForeignKey("awards_questions.id"), nullable=False, index=True)
    answer_text = db.Column(db.Text, nullable=True)
    submission = db.relationship("AwardsSubmission", lazy="joined")
    question = db.relationship("AwardsQuestion", lazy="joined")


class Challenge(db.Model):
    __tablename__ = "challenge"
    id = db.Column(db.Integer, primary_key=True)
    title = db.Column(db.String(255), nullable=False)
    description = db.Column(db.Text, nullable=False, default="")
    created_by = db.Column(db.Integer, db.ForeignKey("user.id"), nullable=False, index=True)
    deadline = db.Column(db.DateTime, nullable=True)
    created_at = db.Column(db.DateTime, default=datetime.utcnow, nullable=False, index=True)
    creator = db.relationship("User", lazy="joined")


class ChallengeResponse(db.Model):
    __tablename__ = "challenge_response"
    id = db.Column(db.Integer, primary_key=True)
    challenge_id = db.Column(db.Integer, db.ForeignKey("challenge.id"), nullable=False, index=True)
    user_id = db.Column(db.Integer, db.ForeignKey("user.id"), nullable=False, index=True)
    answer = db.Column(db.Text, nullable=False, default="")
    created_at = db.Column(db.DateTime, default=datetime.utcnow, nullable=False, index=True)
    challenge = db.relationship("Challenge", lazy="joined")
    user = db.relationship("User", lazy="joined")


class UserFollow(db.Model):
    __tablename__ = "user_follows"
    __table_args__ = (db.UniqueConstraint("follower_id", "following_id", name="uq_user_follows_pair"),)

    id = db.Column(db.Integer, primary_key=True)
    follower_id = db.Column(db.Integer, db.ForeignKey("user.id"), nullable=False, index=True)
    following_id = db.Column(db.Integer, db.ForeignKey("user.id"), nullable=False, index=True)
    created_at = db.Column(db.DateTime, default=datetime.utcnow, nullable=False, index=True)


class FeedPost(db.Model):
    __tablename__ = "feed_post"
    id = db.Column(db.Integer, primary_key=True)
    author_user_id = db.Column(db.Integer, db.ForeignKey("user.id"), nullable=False, index=True)
    content = db.Column(db.Text, nullable=False, default="")
    post_type = db.Column(db.String(20), nullable=False, default="update", index=True)
    media_type = db.Column(db.String(20), nullable=True)
    media_path = db.Column(db.String(500), nullable=True)
    reference_id = db.Column(db.Integer, nullable=True, index=True)
    reference_type = db.Column(db.String(40), nullable=True, index=True)
    created_at = db.Column(db.DateTime, default=datetime.utcnow, nullable=False, index=True)
    author = db.relationship("User", lazy="joined")


class PostReaction(db.Model):
    __tablename__ = "post_reactions"
    __table_args__ = (db.UniqueConstraint("post_id", "user_id", name="uq_post_reactions_post_user"),)

    id = db.Column(db.Integer, primary_key=True)
    post_id = db.Column(db.Integer, db.ForeignKey("feed_post.id"), nullable=False, index=True)
    user_id = db.Column(db.Integer, db.ForeignKey("user.id"), nullable=False, index=True)
    reaction_type = db.Column(db.String(20), nullable=False, default="like", index=True)
    created_at = db.Column(db.DateTime, default=datetime.utcnow, nullable=False)
    updated_at = db.Column(db.DateTime, default=datetime.utcnow, onupdate=datetime.utcnow, nullable=False)


class Comment(db.Model):
    __tablename__ = "comments"

    id = db.Column(db.Integer, primary_key=True)
    post_id = db.Column(db.Integer, db.ForeignKey("feed_post.id"), nullable=False, index=True)
    user_id = db.Column(db.Integer, db.ForeignKey("user.id"), nullable=False, index=True)
    content = db.Column(db.Text, nullable=False, default="")
    mentioned_user_ids_json = db.Column(db.Text, nullable=True)
    created_at = db.Column(db.DateTime, default=datetime.utcnow, nullable=False, index=True)
    post = db.relationship("FeedPost", lazy="joined")
    user = db.relationship("User", lazy="joined")


class CommentLike(db.Model):
    __tablename__ = "comment_likes"
    __table_args__ = (db.UniqueConstraint("comment_id", "user_id", name="uq_comment_likes_comment_user"),)

    id = db.Column(db.Integer, primary_key=True)
    comment_id = db.Column(db.Integer, db.ForeignKey("comments.id"), nullable=False, index=True)
    user_id = db.Column(db.Integer, db.ForeignKey("user.id"), nullable=False, index=True)
    created_at = db.Column(db.DateTime, default=datetime.utcnow, nullable=False)


class PasswordResetToken(db.Model):
    __tablename__ = "password_reset_tokens"
    id = db.Column(db.Integer, primary_key=True)
    user_id = db.Column(db.Integer, db.ForeignKey("user.id"), nullable=False)
    token_hash = db.Column(db.String(64), nullable=False, index=True)
    expires_at = db.Column(db.DateTime, nullable=False)
    used = db.Column(db.Boolean, nullable=False, default=False)


class EmissionFactor(db.Model):
    id = db.Column(db.Integer, primary_key=True)
    category = db.Column(db.String(200), nullable=False)
    subcategory = db.Column(db.String(200), nullable=False)
    factor = db.Column(db.Float, nullable=False)
    unit = db.Column(db.String(100))
    year = db.Column(db.Integer)
    description = db.Column(db.Text)
    extra_data = db.Column(db.Text)  # Full row payload as JSON


class PipelineRun(db.Model):
    id = db.Column(db.Integer, primary_key=True)
    user_id = db.Column(db.Integer, db.ForeignKey('user.id'), nullable=False)
    created_at = db.Column(db.DateTime, default=datetime.utcnow)
    status = db.Column(db.String(30), default="queued")  # queued, running, succeeded, failed
    input_dir = db.Column(db.String(800), nullable=False)
    run_dir = db.Column(db.String(800), nullable=False)
    stage1_output = db.Column(db.String(800))
    stage2_output_dir = db.Column(db.String(800))
    log_path = db.Column(db.String(800))
    exit_code = db.Column(db.Integer)
    error_message = db.Column(db.Text)


class MappingRun(db.Model):
    """
    Web-triggered single-company Stage2 mapping runs (persisted).
    """
    id = db.Column(db.String(32), primary_key=True)  # run_id (hex)
    user_id = db.Column(db.Integer, db.ForeignKey('user.id'), nullable=False)
    company_name = db.Column(db.String(200), nullable=False)
    sheet_name = db.Column(db.String(200), nullable=False)
    created_at = db.Column(db.DateTime, default=datetime.utcnow)
    status = db.Column(db.String(30), default="queued")  # queued, running, succeeded, failed
    input_path = db.Column(db.String(900))
    output_path = db.Column(db.String(900))
    error_message = db.Column(db.Text)
    # When set, this run mapped only the given DataEntry batch (entry_group).
    source_entry_group = db.Column(db.String(40), nullable=True)


class MappingRunSummary(db.Model):
    """
    Lightweight persisted totals derived from a MappingRun output workbook.
    Used to power Home / Carbon Accounting dashboards.
    """
    id = db.Column(db.Integer, primary_key=True)
    run_id = db.Column(db.String(32), unique=True, nullable=False)
    company_name = db.Column(db.String(200), nullable=False)
    sheet_name = db.Column(db.String(200), nullable=False)
    scope = db.Column(db.Integer)  # 1/2/3 or NULL
    tco2e_total = db.Column(db.Float, default=0.0)
    rows_count = db.Column(db.Integer, default=0)
    mapped_categories_count = db.Column(db.Integer, default=0)
    total_categories = db.Column(db.Integer, default=0)
    coverage_pct = db.Column(db.Float, default=0.0)
    created_at = db.Column(db.DateTime, default=datetime.utcnow)


class MappingUnmappedRow(db.Model):
    """
    Rows that completed a mapping run but still have status == "No match".
    Kept separate so mapped rows can continue into dashboards while owners review gaps.
    """
    __tablename__ = "mapping_unmapped_row"
    __table_args__ = (
        db.Index("ix_unmapped_company_sheet_status", "company_name", "sheet_name", "review_status"),
        db.Index("ix_unmapped_run_row", "run_id", "row_number"),
    )

    id = db.Column(db.Integer, primary_key=True)
    run_id = db.Column(db.String(32), db.ForeignKey("mapping_run.id"), nullable=False, index=True)
    user_id = db.Column(db.Integer, db.ForeignKey("user.id"), nullable=True, index=True)
    company_name = db.Column(db.String(200), nullable=False, index=True)
    sheet_name = db.Column(db.String(200), nullable=False, index=True)
    source_entry_group = db.Column(db.String(40), nullable=True, index=True)
    row_number = db.Column(db.Integer, nullable=False, default=0)
    row_label = db.Column(db.String(500), nullable=True)
    status_value = db.Column(db.String(120), nullable=False, default="No match")
    row_payload = db.Column(db.Text, nullable=False, default="{}")
    review_status = db.Column(db.String(32), nullable=False, default="open", index=True)
    assigned_ef_id = db.Column(db.String(120), nullable=True, index=True)
    owner_notes = db.Column(db.Text, nullable=True)
    created_at = db.Column(db.DateTime, default=datetime.utcnow, nullable=False, index=True)
    resolved_at = db.Column(db.DateTime, nullable=True)
    resolved_by_user_id = db.Column(db.Integer, db.ForeignKey("user.id"), nullable=True)


class DataEntry(db.Model):
    id = db.Column(db.Integer, primary_key=True)
    company_name = db.Column(db.String(200), nullable=False)
    sheet_name = db.Column(db.String(200), nullable=False)
    entry_group = db.Column(db.String(32), nullable=False, default="")
    uploaded_by_user_id = db.Column(db.Integer, db.ForeignKey('user.id'), nullable=True)
    row_index = db.Column(db.Integer, nullable=False, default=1)
    column_name = db.Column(db.String(200), nullable=False)
    value = db.Column(db.Text)
    created_at = db.Column(db.DateTime, default=datetime.utcnow)


class EvidenceFile(db.Model):
    __tablename__ = "evidence_files"
    __table_args__ = (
        db.UniqueConstraint("company_name", "sha256_hash", name="uq_evidence_company_sha256"),
        db.Index("ix_evidence_files_company_uploaded", "company_name", "uploaded_at"),
        db.Index("ix_evidence_files_sha256", "sha256_hash"),
    )

    id = db.Column(db.Integer, primary_key=True)
    company_name = db.Column(db.String(200), nullable=False, index=True)
    original_filename = db.Column(db.String(500), nullable=False)
    stored_filename = db.Column(db.String(220), nullable=False)
    file_extension = db.Column(db.String(16), nullable=False)
    mime_type = db.Column(db.String(120), nullable=False)
    sha256_hash = db.Column(db.String(64), nullable=False)
    file_size_original = db.Column(db.BigInteger, nullable=False, default=0)
    file_size_optimized = db.Column(db.BigInteger, nullable=True)
    storage_path = db.Column(db.String(600), nullable=False, default="")
    thumbnail_storage_path = db.Column(db.String(600), nullable=True)
    uploaded_by = db.Column(db.Integer, db.ForeignKey("user.id"), nullable=True)
    uploaded_at = db.Column(db.DateTime, default=datetime.utcnow, nullable=False)
    upload_source = db.Column(db.String(64), nullable=False, default="data_entry")
    processing_status = db.Column(db.String(32), nullable=False, default="pending")
    processing_error = db.Column(db.Text, nullable=True)
    is_deleted = db.Column(db.Boolean, nullable=False, default=False)
    deleted_at = db.Column(db.DateTime, nullable=True)
    is_orphaned = db.Column(db.Boolean, nullable=False, default=False)
    orphaned_at = db.Column(db.DateTime, nullable=True)
    relation_count = db.Column(db.Integer, nullable=True)


class DataEntryEvidence(db.Model):
    __tablename__ = "data_entry_evidence"
    __table_args__ = (
        db.UniqueConstraint(
            "company_name",
            "sheet_name",
            "entry_group",
            "evidence_file_id",
            name="uq_data_entry_evidence_row_file",
        ),
        db.Index("ix_data_entry_evidence_row", "company_name", "sheet_name", "entry_group"),
    )

    id = db.Column(db.Integer, primary_key=True)
    company_name = db.Column(db.String(200), nullable=False)
    sheet_name = db.Column(db.String(200), nullable=False)
    entry_group = db.Column(db.String(128), nullable=False)
    evidence_file_id = db.Column(db.Integer, db.ForeignKey("evidence_files.id"), nullable=False)
    linked_by = db.Column(db.Integer, db.ForeignKey("user.id"), nullable=True)
    linked_at = db.Column(db.DateTime, default=datetime.utcnow, nullable=False)


class ProductMonthlyEntry(db.Model):
    __tablename__ = "product_monthly_entry"
    __table_args__ = (
        db.Index("ix_product_monthly_company_period", "company_name", "reporting_period_key"),
    )

    id = db.Column(db.Integer, primary_key=True)
    company_name = db.Column(db.String(200), nullable=False, index=True)
    reporting_period_key = db.Column(db.String(7), nullable=False, index=True)  # YYYY-MM
    reporting_period_label = db.Column(db.String(40), nullable=False)
    row_index = db.Column(db.Integer, nullable=False, default=1)
    product_type = db.Column(db.String(200), nullable=False)
    quantity = db.Column(db.Float, nullable=False)
    quantity_unit = db.Column(db.String(80), nullable=False)
    end_use_location = db.Column(db.String(200), nullable=False)
    product_weight = db.Column(db.Float, nullable=False)
    product_unit = db.Column(db.String(80), nullable=False)
    created_by_user_id = db.Column(db.Integer, db.ForeignKey("user.id"), nullable=False, index=True)
    created_at = db.Column(db.DateTime, default=datetime.utcnow, nullable=False)
    updated_at = db.Column(db.DateTime, default=datetime.utcnow, onupdate=datetime.utcnow, nullable=False)
    creator = db.relationship("User", lazy="joined")


class CsrdPolicy(db.Model):
    __tablename__ = "csrd_policies"
    id = db.Column(db.Integer, primary_key=True)
    title = db.Column(db.String(500), nullable=False)
    short_description = db.Column(db.Text, nullable=False)
    # Relative to FRONTEND_UPLOAD_DIR, e.g. csrd_policies/<uuid>_<name>.pdf
    file_relpath = db.Column(db.String(600), nullable=False)
    created_at = db.Column(db.DateTime, default=datetime.utcnow)
    updated_at = db.Column(db.DateTime, default=datetime.utcnow, onupdate=datetime.utcnow)


class EmployeeCommutingHeadcount(db.Model):
    __tablename__ = "employee_commuting_headcount"
    id = db.Column(db.Integer, primary_key=True)
    company_name = db.Column(db.String(200), nullable=False, unique=True)
    headcount = db.Column(db.Integer, nullable=False)
    created_at = db.Column(db.DateTime, default=datetime.utcnow)
    updated_at = db.Column(db.DateTime, default=datetime.utcnow, onupdate=datetime.utcnow)


class EmployeeCommutingNationalAverage(db.Model):
    __tablename__ = "employee_commuting_national_average"
    id = db.Column(db.Integer, primary_key=True)
    company_name = db.Column(db.String(200), nullable=False, unique=True)
    country = db.Column(db.String(100), nullable=False)
    average_one_day = db.Column(db.Float, nullable=False)
    car_pct = db.Column(db.Float, nullable=False)
    bus_pct = db.Column(db.Float, nullable=False)
    walking_and_cycling_pct = db.Column(db.Float, nullable=False)
    mixed_pct = db.Column(db.Float, nullable=False)
    created_at = db.Column(db.DateTime, default=datetime.utcnow)
    updated_at = db.Column(db.DateTime, default=datetime.utcnow, onupdate=datetime.utcnow)


class Notification(db.Model):
    __tablename__ = "notifications"
    id = db.Column(db.Integer, primary_key=True)
    user_id = db.Column(db.Integer, db.ForeignKey("user.id"), nullable=False, index=True)
    title = db.Column(db.String(255), nullable=False)
    message = db.Column(db.Text, nullable=False)
    type = db.Column(db.String(50), nullable=False, default="info")
    link = db.Column(db.String(500), nullable=True)
    is_read = db.Column(db.Boolean, nullable=False, default=False, index=True)
    created_at = db.Column(db.DateTime, default=datetime.utcnow, index=True)
    meta_json = db.Column(db.Text, nullable=True)


class Message(db.Model):
    __tablename__ = "messages"
    id = db.Column(db.Integer, primary_key=True)
    thread_id = db.Column(db.String(64), nullable=False, index=True)
    sender_id = db.Column(db.Integer, db.ForeignKey("user.id"), nullable=False, index=True)
    receiver_id = db.Column(db.Integer, db.ForeignKey("user.id"), nullable=False, index=True)
    message = db.Column(db.Text, nullable=False)
    created_at = db.Column(db.DateTime, default=datetime.utcnow, index=True)
    is_read = db.Column(db.Boolean, nullable=False, default=False, index=True)


class AccessRequest(db.Model):
    __tablename__ = "access_requests"
    id = db.Column(db.Integer, primary_key=True)
    full_name = db.Column(db.String(200), nullable=False)
    email = db.Column(db.String(120), nullable=False, index=True)
    company = db.Column(db.String(200), nullable=False)
    reason = db.Column(db.Text, nullable=False)
    status = db.Column(db.String(32), nullable=False, default="pending")
    created_at = db.Column(db.DateTime, default=datetime.utcnow)


class GovernanceRegister(db.Model):
    """Admin governance register: access integrations and review metadata only (no secrets)."""

    __tablename__ = "governance_register"
    id = db.Column(db.Integer, primary_key=True)
    request_date = db.Column(db.Date, nullable=False, index=True)
    requested_by = db.Column(db.String(200), nullable=True)
    team_department = db.Column(db.String(200), nullable=True, index=True)
    api_software_name = db.Column(db.String(300), nullable=False)
    vendor_platform = db.Column(db.String(300), nullable=True)
    purpose_business_reason = db.Column(db.Text, nullable=True)
    access_type = db.Column(db.String(64), nullable=False, index=True)
    environment = db.Column(db.String(32), nullable=False, index=True)
    status = db.Column(db.String(64), nullable=False, index=True)
    approved_by = db.Column(db.String(200), nullable=True)
    expiry_review_date = db.Column(db.Date, nullable=True, index=True)
    notes_risks = db.Column(db.Text, nullable=True)
    linked_documentation = db.Column(db.String(2048), nullable=True)
    owner = db.Column(db.String(200), nullable=False, index=True)
    created_by_user_id = db.Column(db.Integer, db.ForeignKey("user.id"), nullable=True, index=True)
    last_updated_by_user_id = db.Column(db.Integer, db.ForeignKey("user.id"), nullable=True, index=True)
    attachments_stub_json = db.Column(db.Text, nullable=True)
    created_at = db.Column(db.DateTime, default=datetime.utcnow)
    updated_at = db.Column(db.DateTime, default=datetime.utcnow, onupdate=datetime.utcnow)


class GovernanceRegisterAuditLog(db.Model):
    """Append-only governance actions for administrators; register row may be deleted but register_id is kept for history."""

    __tablename__ = "governance_register_audit"
    id = db.Column(db.Integer, primary_key=True)
    register_id = db.Column(db.Integer, nullable=True, index=True)
    user_id = db.Column(db.Integer, db.ForeignKey("user.id"), nullable=True, index=True)
    action = db.Column(db.String(32), nullable=False, index=True)
    record_label = db.Column(db.String(400), nullable=True)
    created_at = db.Column(db.DateTime, default=datetime.utcnow, index=True)


GOVERNANCE_ACCESS_TYPES: tuple[str, ...] = (
    "API",
    "Software",
    "Admin Access",
    "Service Account",
    "Database",
    "SharePoint",
    "Power BI",
)
GOVERNANCE_ENVIRONMENTS: tuple[str, ...] = ("Production", "Test", "Sandbox")
GOVERNANCE_STATUSES: tuple[str, ...] = (
    "Requested",
    "Under Review",
    "Approved",
    "Rejected",
    "Revoked",
    "In Progress",
    "Expired",
)
GOVERNANCE_ACCESS_TYPES_SET = frozenset(GOVERNANCE_ACCESS_TYPES)
GOVERNANCE_ENVIRONMENTS_SET = frozenset(GOVERNANCE_ENVIRONMENTS)
GOVERNANCE_STATUSES_SET = frozenset(GOVERNANCE_STATUSES)
GOVERNANCE_REGISTER_STATUSES_CLOSED_FOR_OVERDUE: frozenset[str] = frozenset(
    {"Rejected", "Revoked", "Expired"}
)
_GOVERNANCE_CREDENTIAL_HINT = re.compile(
    r"(?i)(api[_-]?key|secret|password|token|bearer\s|authorization:\s*basic|"
    r"BEGIN (?:RSA |OPENSSH |)PRIVATE|ssh-rsa|eyJ[A-Za-z0-9_-]+\.eyJ|"
    r"xox[baprs]-|ghp_[A-Za-z0-9]{20,}|glpat-[A-Za-z0-9_-]{16,}|"
    r"sk_live_[A-Za-z0-9]{20,}|sk_test_[A-Za-z0-9]{20,}|"
    r"AKIA[0-9A-Z]{16})"
)


def _governance_field_secret_rejection(raw: str | None, *, max_len: int = 8000) -> str | None:
    """Return flash error message if text looks like a credential or is too long; else None."""
    s = (raw or "").strip()
    if not s:
        return None
    if len(s) > max_len:
        return "A text field is too long."
    if _GOVERNANCE_CREDENTIAL_HINT.search(s):
        return "Credential-like content is not allowed in this register. Remove keys, tokens, or passwords."
    if len(s) >= 48:
        alnum_ratio = sum(1 for c in s if c.isalnum()) / max(len(s), 1)
        if alnum_ratio > 0.85 and re.search(r"[A-Za-z0-9+/]{36,}", s):
            return "Credential-like content is not allowed. Use documentation links only—do not paste secrets."
    return None


def _governance_validate_text_fields(
    purpose: str | None, notes: str | None, docs: str | None
) -> str | None:
    for msg in (
        _governance_field_secret_rejection(purpose),
        _governance_field_secret_rejection(notes),
        _governance_field_secret_rejection(docs, max_len=2048),
    ):
        if msg:
            return msg
    return None


def _parse_html_date(value: object) -> date | None:
    if value is None:
        return None
    s = str(value or "").strip()
    if not s:
        return None
    try:
        return datetime.strptime(s[:10], "%Y-%m-%d").date()
    except ValueError:
        return None


def _governance_sort_column(key: str):
    mapping = {
        "request_date": GovernanceRegister.request_date,
        "requested_by": GovernanceRegister.requested_by,
        "team_department": GovernanceRegister.team_department,
        "api_software_name": GovernanceRegister.api_software_name,
        "vendor_platform": GovernanceRegister.vendor_platform,
        "access_type": GovernanceRegister.access_type,
        "environment": GovernanceRegister.environment,
        "status": GovernanceRegister.status,
        "approved_by": GovernanceRegister.approved_by,
        "expiry_review_date": GovernanceRegister.expiry_review_date,
        "owner": GovernanceRegister.owner,
        "created_at": GovernanceRegister.created_at,
        "updated_at": GovernanceRegister.updated_at,
        "last_updated_by_user_id": GovernanceRegister.last_updated_by_user_id,
    }
    return mapping.get((key or "").strip(), GovernanceRegister.request_date)


def _apply_governance_register_filters(q, args) -> object:
    status = (args.get("status") or "").strip()
    environment = (args.get("environment") or "").strip()
    access_type = (args.get("access_type") or "").strip()
    team = (args.get("team") or "").strip()
    search = (args.get("q") or "").strip()

    if status and status in GOVERNANCE_STATUSES_SET:
        q = q.filter(GovernanceRegister.status == status)
    if environment and environment in GOVERNANCE_ENVIRONMENTS_SET:
        q = q.filter(GovernanceRegister.environment == environment)
    if access_type and access_type in GOVERNANCE_ACCESS_TYPES_SET:
        q = q.filter(GovernanceRegister.access_type == access_type)
    if team:
        like = f"%{team}%"
        q = q.filter(GovernanceRegister.team_department.ilike(like))
    if search:
        like = f"%{search}%"
        q = q.filter(
            or_(
                GovernanceRegister.api_software_name.ilike(like),
                GovernanceRegister.requested_by.ilike(like),
                GovernanceRegister.vendor_platform.ilike(like),
                GovernanceRegister.team_department.ilike(like),
                GovernanceRegister.owner.ilike(like),
                GovernanceRegister.purpose_business_reason.ilike(like),
                GovernanceRegister.notes_risks.ilike(like),
                GovernanceRegister.linked_documentation.ilike(like),
                GovernanceRegister.approved_by.ilike(like),
            )
        )
    return q


def _governance_register_stats(base_q) -> dict[str, int]:
    return {
        "total": base_q.count(),
        "approved": base_q.filter(GovernanceRegister.status == "Approved").count(),
        "under_review": base_q.filter(GovernanceRegister.status == "Under Review").count(),
        "expired": base_q.filter(GovernanceRegister.status == "Expired").count(),
        "revoked": base_q.filter(GovernanceRegister.status == "Revoked").count(),
    }


def _governance_register_stats_with_pct(base_q) -> dict[str, object]:
    s = _governance_register_stats(base_q)
    t = int(s.get("total") or 0)
    if t <= 0:
        return {
            **s,
            "pct_approved": None,
            "pct_under_review": None,
            "pct_expired": None,
            "pct_revoked": None,
        }
    return {
        **s,
        "pct_approved": round(100.0 * int(s["approved"]) / t, 1),
        "pct_under_review": round(100.0 * int(s["under_review"]) / t, 1),
        "pct_expired": round(100.0 * int(s["expired"]) / t, 1),
        "pct_revoked": round(100.0 * int(s["revoked"]) / t, 1),
    }


def _governance_user_label_map(user_ids: set[int]) -> dict[int, str]:
    if not user_ids:
        return {}
    clean = {int(x) for x in user_ids if x}
    if not clean:
        return {}
    rows = User.query.filter(User.id.in_(clean)).all()
    return {
        int(u.id): f"{_display_name_for_user(u)} ({(u.email or '').strip() or 'no-email'})".strip()
        for u in rows
    }


def _log_governance_audit(
    register_id: int | None,
    action: str,
    *,
    record_label: str | None = None,
    user_id: int | None = None,
) -> None:
    try:
        uid = user_id
        if uid is None:
            uid = int(getattr(current_user, "id", 0) or 0) or None
        lab = (record_label or "").strip()[:400] or None
        db.session.add(
            GovernanceRegisterAuditLog(
                register_id=register_id,
                user_id=uid,
                action=str(action or "unknown")[:32],
                record_label=lab,
            )
        )
    except Exception:
        pass


def _governance_recent_audit_payload(limit: int = 25) -> list[dict[str, object]]:
    rows = (
        GovernanceRegisterAuditLog.query.order_by(GovernanceRegisterAuditLog.created_at.desc())
        .limit(max(1, min(limit, 100)))
        .all()
    )
    uids = {int(r.user_id) for r in rows if r.user_id}
    um = _governance_user_label_map(uids)
    out: list[dict[str, object]] = []
    for r in rows:
        uid = int(r.user_id) if r.user_id else None
        out.append(
            {
                "at": r.created_at.strftime("%Y-%m-%d %H:%M") if r.created_at else "",
                "action": r.action or "",
                "record_label": (r.record_label or "").strip(),
                "user_label": um.get(uid, "Unknown") if uid else "-",
            }
        )
    return out


def _governance_attachments_stub_label(stub: str | None) -> tuple[str, str]:
    """Return (cell text, title tooltip) for attachment placeholder column."""
    s = (stub or "").strip() or "[]"
    try:
        data = json.loads(s)
        if isinstance(data, list) and len(data) > 0:
            return (f"{len(data)} linked", "Reserved attachment metadata (future upload support)")
    except Exception:
        pass
    return ("-", "File attachments reserved for a future release")


def _governance_row_to_editor_dict(r: GovernanceRegister) -> dict[str, object]:
    return {
        "id": r.id,
        "request_date": r.request_date.isoformat() if r.request_date else "",
        "requested_by": r.requested_by or "",
        "team_department": r.team_department or "",
        "api_software_name": r.api_software_name or "",
        "vendor_platform": r.vendor_platform or "",
        "purpose_business_reason": r.purpose_business_reason or "",
        "access_type": r.access_type or "",
        "environment": r.environment or "",
        "status": r.status or "",
        "approved_by": r.approved_by or "",
        "expiry_review_date": r.expiry_review_date.isoformat() if r.expiry_review_date else "",
        "notes_risks": r.notes_risks or "",
        "linked_documentation": r.linked_documentation or "",
        "owner": r.owner or "",
    }


def _governance_filter_map_from_values(values) -> dict[str, str]:
    """Build stable filter query parts for url_for (omit empties)."""
    out: dict[str, str] = {}
    for key in ("q", "status", "environment", "access_type", "team"):
        raw = (values.get(key) if values is not None else None) or ""
        s = str(raw).strip()
        if s:
            out[key] = s
    return out


def _governance_read_form_fields() -> tuple[dict[str, object] | None, str | None]:
    request_date = _parse_html_date(request.form.get("request_date"))
    if not request_date:
        return None, "Request date is required."
    api_software_name = (request.form.get("api_software_name") or "").strip()
    if not api_software_name:
        return None, "API / software name is required."
    owner = (request.form.get("owner") or "").strip()
    if not owner:
        return None, "Owner is required."
    access_type = (request.form.get("access_type") or "").strip()
    if access_type not in GOVERNANCE_ACCESS_TYPES_SET:
        return None, "Invalid access type."
    environment = (request.form.get("environment") or "").strip()
    if environment not in GOVERNANCE_ENVIRONMENTS_SET:
        return None, "Invalid environment."
    status = (request.form.get("status") or "").strip()
    if status not in GOVERNANCE_STATUSES_SET:
        return None, "Invalid status."
    purpose_raw = request.form.get("purpose_business_reason")
    purpose = (str(purpose_raw).strip() if purpose_raw is not None else "") or None
    notes_raw = request.form.get("notes_risks")
    notes = (str(notes_raw).strip() if notes_raw is not None else "") or None
    docs_raw = request.form.get("linked_documentation")
    docs = (str(docs_raw).strip() if docs_raw is not None else "") or None
    msg = _governance_validate_text_fields(purpose, notes, docs)
    if msg:
        return None, msg
    expiry = _parse_html_date(request.form.get("expiry_review_date"))
    return {
        "request_date": request_date,
        "requested_by": (request.form.get("requested_by") or "").strip() or None,
        "team_department": (request.form.get("team_department") or "").strip() or None,
        "api_software_name": api_software_name,
        "vendor_platform": (request.form.get("vendor_platform") or "").strip() or None,
        "purpose_business_reason": purpose,
        "access_type": access_type,
        "environment": environment,
        "status": status,
        "approved_by": (request.form.get("approved_by") or "").strip() or None,
        "expiry_review_date": expiry,
        "notes_risks": notes,
        "linked_documentation": docs,
        "owner": owner,
    }, None


class UserActivityLog(db.Model):
    __tablename__ = "user_activity_logs"
    id = db.Column(db.Integer, primary_key=True)
    user_id = db.Column(db.Integer, db.ForeignKey("user.id"), nullable=True, index=True)
    company_name = db.Column(db.String(200), nullable=True, index=True)
    email = db.Column(db.String(120), nullable=True, index=True)
    action = db.Column(db.String(80), nullable=False, index=True)
    page = db.Column(db.String(500), nullable=True)
    endpoint = db.Column(db.String(120), nullable=True, index=True)
    method = db.Column(db.String(16), nullable=True)
    ip_address = db.Column(db.String(80), nullable=True, index=True)
    country = db.Column(db.String(120), nullable=True, index=True)
    city = db.Column(db.String(120), nullable=True, index=True)
    device = db.Column(db.String(80), nullable=True, index=True)
    browser = db.Column(db.String(80), nullable=True, index=True)
    os = db.Column(db.String(80), nullable=True, index=True)
    referrer = db.Column(db.String(500), nullable=True)
    session_id = db.Column(db.String(64), nullable=True, index=True)
    created_at = db.Column(db.DateTime, default=datetime.utcnow, index=True)


def _csrd_policy_upload_dir() -> Path:
    d = FRONTEND_UPLOAD_DIR / "csrd_policies"
    d.mkdir(parents=True, exist_ok=True)
    return d


def _csrd_filename_is_pdf(filename: str) -> bool:
    return Path(secure_filename(filename or "")).suffix.lower() == ".pdf"


def _employee_commuting_data_dir() -> Path:
    EMPLOYEE_COMMUTING_DATA_DIR.mkdir(parents=True, exist_ok=True)
    return EMPLOYEE_COMMUTING_DATA_DIR


def _clean_employee_commuting_company(value: object) -> str:
    return " ".join(str(value or "").replace("\u00A0", " ").strip().split())


def _serialize_employee_commuting_headcount_rows() -> list[dict[str, object]]:
    rows = EmployeeCommutingHeadcount.query.order_by(EmployeeCommutingHeadcount.company_name.asc()).all()
    return [
        {
            "company_name": _clean_employee_commuting_company(row.company_name),
            "headcount": int(row.headcount or 0),
        }
        for row in rows
    ]


def _serialize_employee_commuting_national_average_rows() -> list[dict[str, object]]:
    rows = (
        EmployeeCommutingNationalAverage.query
        .order_by(EmployeeCommutingNationalAverage.company_name.asc())
        .all()
    )
    return [
        {
            "company_name": _clean_employee_commuting_company(row.company_name),
            "country": str(row.country or "").strip(),
            "average_one_day": float(row.average_one_day or 0),
            "car_pct": float(row.car_pct or 0),
            "bus_pct": float(row.bus_pct or 0),
            "walking_and_cycling_pct": float(row.walking_and_cycling_pct or 0),
            "mixed_pct": float(row.mixed_pct or 0),
        }
        for row in rows
    ]


def _publish_employee_commuting_files() -> None:
    _employee_commuting_data_dir()

    headcount_rows = _serialize_employee_commuting_headcount_rows()
    national_average_rows = _serialize_employee_commuting_national_average_rows()

    headcount_df = pd.DataFrame(
        [
            {
                "Company_Name": row["company_name"],
                "Headcount": int(row["headcount"] or 0),
            }
            for row in headcount_rows
        ]
    )
    if headcount_df.empty:
        headcount_df = pd.DataFrame(columns=["Company_Name", "Headcount"])
    headcount_df.to_csv(EMPLOYEE_COMMUTING_HEADCOUNT_CSV, index=False)

    national_average_df = pd.DataFrame(
        [
            {
                "Company_Name": row["company_name"],
                "Country": row["country"],
                "Average one day": float(row["average_one_day"] or 0),
                "Car %": float(row["car_pct"] or 0),
                "Bus %": float(row["bus_pct"] or 0),
                "Walking and Cycling %": float(row["walking_and_cycling_pct"] or 0),
                "Mixed %": float(row["mixed_pct"] or 0),
            }
            for row in national_average_rows
        ]
    )
    if national_average_df.empty:
        national_average_df = pd.DataFrame(
            columns=[
                "Company_Name",
                "Country",
                "Average one day",
                "Car %",
                "Bus %",
                "Walking and Cycling %",
                "Mixed %",
            ]
        )

    merged_df = national_average_df.merge(headcount_df, on="Company_Name", how="outer")
    for col in [
        "Country",
        "Average one day",
        "Car %",
        "Bus %",
        "Walking and Cycling %",
        "Mixed %",
        "Headcount",
    ]:
        if col not in merged_df.columns:
            merged_df[col] = ""
    merged_df = merged_df[
        [
            "Company_Name",
            "Country",
            "Average one day",
            "Car %",
            "Bus %",
            "Walking and Cycling %",
            "Mixed %",
            "Headcount",
        ]
    ].sort_values("Company_Name", na_position="last")

    with pd.ExcelWriter(EMPLOYEE_COMMUTING_NATIONAL_AVERAGES_XLSX, engine="openpyxl") as writer:
        merged_df.to_excel(writer, sheet_name="Sheet1", index=False)


def _seed_employee_commuting_defaults() -> None:
    _employee_commuting_data_dir()
    seeded = False

    workbook_exists = EMPLOYEE_COMMUTING_NATIONAL_AVERAGES_XLSX.exists()
    if not workbook_exists:
        _publish_employee_commuting_files()
        return

    try:
        source_df = pd.read_excel(EMPLOYEE_COMMUTING_NATIONAL_AVERAGES_XLSX, engine="openpyxl")
    except Exception:
        return
    if source_df is None or source_df.empty:
        return

    if EmployeeCommutingNationalAverage.query.count() == 0:
        for _, row in source_df.iterrows():
            company_name = _clean_employee_commuting_company(row.get("Company_Name"))
            if not company_name:
                continue
            db.session.add(
                EmployeeCommutingNationalAverage(
                    company_name=company_name,
                    country=str(row.get("Country") or "").strip(),
                    average_one_day=float(pd.to_numeric(row.get("Average one day"), errors="coerce") or 0),
                    car_pct=float(pd.to_numeric(row.get("Car %"), errors="coerce") or 0),
                    bus_pct=float(pd.to_numeric(row.get("Bus %"), errors="coerce") or 0),
                    walking_and_cycling_pct=float(pd.to_numeric(row.get("Walking and Cycling %"), errors="coerce") or 0),
                    mixed_pct=float(pd.to_numeric(row.get("Mixed %"), errors="coerce") or 0),
                )
            )
        seeded = True

    if EmployeeCommutingHeadcount.query.count() == 0:
        for _, row in source_df.iterrows():
            company_name = _clean_employee_commuting_company(row.get("Company_Name"))
            headcount_value = pd.to_numeric(row.get("Headcount"), errors="coerce")
            if not company_name or pd.isna(headcount_value):
                continue
            db.session.add(
                EmployeeCommutingHeadcount(
                    company_name=company_name,
                    headcount=int(round(float(headcount_value))),
                )
            )
        seeded = True

    if seeded:
        db.session.commit()
        _publish_employee_commuting_files()


def _parse_employee_commuting_non_negative_float(value: object, label: str, row_no: int) -> float:
    numeric = pd.to_numeric(value, errors="coerce")
    if pd.isna(numeric):
        raise ValueError(f"Row {row_no}: `{label}` must be numeric.")
    out = float(numeric)
    if out < 0:
        raise ValueError(f"Row {row_no}: `{label}` cannot be negative.")
    return out


def _normalize_employee_commuting_headcount_payload(rows: object) -> list[dict[str, object]]:
    if not isinstance(rows, list):
        raise ValueError("Rows payload must be a list.")

    normalized: list[dict[str, object]] = []
    seen: set[str] = set()
    for idx, raw in enumerate(rows, start=1):
        if not isinstance(raw, dict):
            continue
        company_name = _clean_employee_commuting_company(raw.get("company_name"))
        headcount_raw = str(raw.get("headcount") or "").strip()
        if not company_name and not headcount_raw:
            continue
        if not company_name:
            raise ValueError(f"Row {idx}: `Company_Name` is required.")
        headcount = int(round(_parse_employee_commuting_non_negative_float(headcount_raw, "Headcount", idx)))
        dedup_key = company_name.lower()
        if dedup_key in seen:
            raise ValueError(f"Row {idx}: `{company_name}` appears more than once.")
        seen.add(dedup_key)
        normalized.append({"company_name": company_name, "headcount": headcount})
    return normalized


def _normalize_employee_commuting_national_average_payload(rows: object) -> list[dict[str, object]]:
    if not isinstance(rows, list):
        raise ValueError("Rows payload must be a list.")

    normalized: list[dict[str, object]] = []
    seen: set[str] = set()
    for idx, raw in enumerate(rows, start=1):
        if not isinstance(raw, dict):
            continue
        company_name = _clean_employee_commuting_company(raw.get("company_name"))
        country = str(raw.get("country") or "").strip()
        raw_values = [
            str(raw.get("average_one_day") or "").strip(),
            str(raw.get("car_pct") or "").strip(),
            str(raw.get("bus_pct") or "").strip(),
            str(raw.get("walking_and_cycling_pct") or "").strip(),
            str(raw.get("mixed_pct") or "").strip(),
        ]
        if not company_name and not country and not any(raw_values):
            continue
        if not company_name:
            raise ValueError(f"Row {idx}: `Company_Name` is required.")
        if not country:
            raise ValueError(f"Row {idx}: `Country` is required.")

        average_one_day = _parse_employee_commuting_non_negative_float(raw.get("average_one_day"), "Average one day", idx)
        car_pct = _parse_employee_commuting_non_negative_float(raw.get("car_pct"), "Car %", idx)
        bus_pct = _parse_employee_commuting_non_negative_float(raw.get("bus_pct"), "Bus %", idx)
        walking_and_cycling_pct = _parse_employee_commuting_non_negative_float(
            raw.get("walking_and_cycling_pct"),
            "Walking and Cycling %",
            idx,
        )
        mixed_pct = _parse_employee_commuting_non_negative_float(raw.get("mixed_pct"), "Mixed %", idx)
        total_pct = car_pct + bus_pct + walking_and_cycling_pct + mixed_pct
        if abs(total_pct - 100.0) > 1.0:
            raise ValueError(
                f"Row {idx}: transport shares must add up to 100. Current total: {total_pct:.2f}"
            )

        dedup_key = company_name.lower()
        if dedup_key in seen:
            raise ValueError(f"Row {idx}: `{company_name}` appears more than once.")
        seen.add(dedup_key)
        normalized.append(
            {
                "company_name": company_name,
                "country": country,
                "average_one_day": average_one_day,
                "car_pct": car_pct,
                "bus_pct": bus_pct,
                "walking_and_cycling_pct": walking_and_cycling_pct,
                "mixed_pct": mixed_pct,
            }
        )
    return normalized


def _replace_employee_commuting_headcount_rows(rows: list[dict[str, object]]) -> None:
    EmployeeCommutingHeadcount.query.delete()
    for row in rows:
        db.session.add(
            EmployeeCommutingHeadcount(
                company_name=str(row["company_name"]),
                headcount=int(row["headcount"]),
            )
        )
    db.session.commit()
    _publish_employee_commuting_files()


def _replace_employee_commuting_national_average_rows(rows: list[dict[str, object]]) -> None:
    EmployeeCommutingNationalAverage.query.delete()
    for row in rows:
        db.session.add(
            EmployeeCommutingNationalAverage(
                company_name=str(row["company_name"]),
                country=str(row["country"]),
                average_one_day=float(row["average_one_day"]),
                car_pct=float(row["car_pct"]),
                bus_pct=float(row["bus_pct"]),
                walking_and_cycling_pct=float(row["walking_and_cycling_pct"]),
                mixed_pct=float(row["mixed_pct"]),
            )
        )
    db.session.commit()
    _publish_employee_commuting_files()


def _ensure_db_tables() -> None:
    """
    Best-effort table creation for environments that don't run app.py directly.
    Safe to call within request context.
    """
    try:
        db.create_all()
        _ensure_mapping_run_summary_columns()
        _ensure_data_entry_columns()
        _ensure_mapping_run_source_entry_group_column()
        _ensure_user_profile_columns()
        _ensure_user_last_seen_column()
        _ensure_feed_post_reference_columns()
        _ensure_report_category_columns()
        _ensure_awards_form_columns()
        _ensure_mapping_unmapped_row_columns()
        _ensure_evidence_tables_columns()
        _ensure_governance_register_columns()
        _ensure_notification_meta_json_column()
        _migrate_profile_photos_to_storage_once()
    except Exception:
        pass


def _is_static_or_ignored_activity_path(path: str | None) -> bool:
    raw = str(path or "")
    return (
        not raw
        or raw.startswith("/static/")
        or raw.startswith("/assets/")
        or raw.startswith("/favicon")
        or raw.startswith("/api/profile-photo")
    )


def _client_ip_address() -> str:
    forwarded = str(request.headers.get("X-Forwarded-For", "") or "").strip()
    if forwarded:
        return forwarded.split(",")[0].strip()
    return str(request.remote_addr or "").strip() or "Unknown"


def _normalize_referrer(value: str | None) -> str | None:
    ref = str(value or "").strip()
    return ref[:500] if ref else None


def _normalize_page_path() -> str:
    path = str(request.path or "").strip() or "/"
    query_string = str(request.query_string.decode("utf-8", errors="ignore") if request.query_string else "").strip()
    if request.endpoint == "search_page" and query_string:
        return f"{path}?{query_string}"[:500]
    return path[:500]


def _activity_session_id(create_if_missing: bool = True) -> str | None:
    sid = str(session.get("activity_session_id", "") or "").strip()
    if sid or not create_if_missing:
        return sid or None
    sid = uuid.uuid4().hex
    session["activity_session_id"] = sid
    return sid


def _parse_user_agent(user_agent: str | None) -> tuple[str, str, str]:
    ua = str(user_agent or "").lower()
    if not ua:
        return ("Unknown", "Unknown", "Unknown")

    if "bot" in ua or "spider" in ua or "crawl" in ua:
        device = "Bot"
    elif "tablet" in ua or "ipad" in ua:
        device = "Tablet"
    elif "mobile" in ua or "iphone" in ua or "android" in ua:
        device = "Mobile"
    else:
        device = "Desktop"

    if "edg/" in ua:
        browser = "Edge"
    elif "opr/" in ua or "opera" in ua:
        browser = "Opera"
    elif "chrome/" in ua and "edg/" not in ua:
        browser = "Chrome"
    elif "safari/" in ua and "chrome/" not in ua:
        browser = "Safari"
    elif "firefox/" in ua:
        browser = "Firefox"
    elif "msie" in ua or "trident/" in ua:
        browser = "Internet Explorer"
    else:
        browser = "Other"

    if "windows" in ua:
        os_name = "Windows"
    elif "mac os" in ua or "macintosh" in ua:
        os_name = "macOS"
    elif "android" in ua:
        os_name = "Android"
    elif "iphone" in ua or "ipad" in ua or "ios" in ua:
        os_name = "iOS"
    elif "linux" in ua:
        os_name = "Linux"
    else:
        os_name = "Other"
    return (device, browser, os_name)


@lru_cache(maxsize=512)
def _geo_lookup_for_ip(ip_address_raw: str) -> tuple[str, str]:
    ip = str(ip_address_raw or "").strip()
    if not ip:
        return ("Unknown", "Unknown")
    try:
        addr = ipaddress.ip_address(ip)
        if addr.is_private or addr.is_loopback or addr.is_reserved or addr.is_multicast:
            return ("Local", "Local")
    except ValueError:
        return ("Unknown", "Unknown")

    url = f"https://ipapi.co/{ip}/json/"
    try:
        with urllib_request.urlopen(url, timeout=0.6) as response:
            payload = json.loads(response.read().decode("utf-8"))
        country = str(payload.get("country_name") or payload.get("country") or "Unknown").strip() or "Unknown"
        city = str(payload.get("city") or "Unknown").strip() or "Unknown"
        return (country, city)
    except (urllib_error.URLError, TimeoutError, json.JSONDecodeError, OSError, ValueError):
        return ("Unknown", "Unknown")


def _classify_activity_action() -> str:
    endpoint = str(request.endpoint or "").strip()
    path = str(request.path or "").strip().lower()
    method = str(request.method or "GET").upper()

    if endpoint == "search_page":
        return "search_usage"
    if endpoint == "api_mapping_run":
        return "mapping_run"
    if endpoint == "data_sources_ccc_api" and method == "POST":
        return "ccc_api_sync"
    if endpoint == "analytics_forecasting" and method == "POST":
        return "forecasting_run"
    if endpoint == "analytics_decarbonization" and method == "POST":
        return "decarbonization_run"
    if endpoint == "governance_audit_ready_output" and method == "POST":
        return "audit_dataset_generation"
    if endpoint == "analytics_mapped_window_output" and method == "POST":
        return "mapped_window_generation"
    if endpoint == "analytics_emissions_totals" and method == "POST":
        return "totals_generation"
    if endpoint == "analytics_share_analysis" and method == "POST":
        return "share_analysis_generation"
    if endpoint == "governance_double_counting_check" and method == "POST":
        return "double_counting_generation"
    if "upload" in path or endpoint in {"admin", "dashboard"} and method == "POST":
        return "data_upload"
    if path.startswith("/api/"):
        return "api_request"
    return "page_visit"


def _activity_log_payload_for_user(user: "User", *, action: str) -> dict[str, object]:
    ip_address = _client_ip_address()
    country, city = _geo_lookup_for_ip(ip_address)
    device, browser, os_name = _parse_user_agent(request.headers.get("User-Agent"))
    return {
        "user_id": int(getattr(user, "id", 0) or 0) or None,
        "company_name": (getattr(user, "company_name", None) or "").strip() or None,
        "email": (getattr(user, "email", None) or "").strip() or None,
        "action": action,
        "page": _normalize_page_path(),
        "endpoint": str(request.endpoint or "").strip() or None,
        "method": str(request.method or "").upper() or None,
        "ip_address": ip_address,
        "country": country or "Unknown",
        "city": city or "Unknown",
        "device": device or "Unknown",
        "browser": browser or "Unknown",
        "os": os_name or "Unknown",
        "referrer": _normalize_referrer(request.referrer),
        "session_id": _activity_session_id(create_if_missing=True),
        "created_at": datetime.utcnow(),
    }


def _write_activity_log_for_user(user: "User", *, action: str) -> None:
    if user is None or not getattr(user, "id", None):
        return
    try:
        payload = _activity_log_payload_for_user(user, action=action)
        with db.engine.begin() as conn:
            conn.execute(UserActivityLog.__table__.insert().values(**payload))
    except Exception:
        pass


def _extract_dataset_name(path_or_ref: str | None) -> str | None:
    raw = str(path_or_ref or "").strip()
    if not raw:
        return None
    cleaned = raw.split("?", 1)[0].rstrip("/")
    name = Path(cleaned).name
    if any(name.lower().endswith(ext) for ext in (".xlsx", ".xls", ".csv", ".json", ".pdf", ".xlsb")):
        return name
    return None


def _owner_analytics_context() -> dict[str, object]:
    now = datetime.utcnow()
    logs = UserActivityLog.query.order_by(UserActivityLog.created_at.asc()).all()
    total_users = int(User.query.count())
    if not logs:
        empty_metrics = {
            "total_users": total_users,
            "active_users_24h": 0,
            "active_users_7d": 0,
            "active_users_30d": 0,
            "unique_sessions": 0,
            "total_logins": 0,
            "total_page_views": 0,
            "total_api_requests": 0,
            "total_uploads": 0,
            "total_searches": 0,
            "ccc_api_usage_count": 0,
            "mapping_runs_count": 0,
            "unmapped_open_count": 0,
            "forecast_runs_count": 0,
            "audit_dataset_runs": 0,
            "avg_session_duration_minutes": 0.0,
            "pages_per_session": 0.0,
            "bounce_rate": 0.0,
            "feature_adoption_rate": 0.0,
            "top_company_engagement_score": 0.0,
            "most_used_dataset": "None yet",
            "peak_activity_hour": "None yet",
            "most_visited_pages": [],
            "most_active_companies": [],
            "top_countries": [],
            "top_cities": [],
            "top_referrers": [],
            "top_browsers": [],
            "top_devices": [],
            "top_operating_systems": [],
            "most_active_hours": [],
            "feature_usage_frequency": [],
            "action_distribution": [],
            "dataset_usage_frequency": [],
            "company_engagement": [],
            "recent_visits": [],
            "user_last_seen": [],
        }
        return {"metrics": empty_metrics, "chart_data": empty_metrics, "activity_rows": 0}

    rows: list[dict[str, object]] = []
    for row in logs:
        rows.append(
            {
                "user_id": row.user_id,
                "company_name": row.company_name or "Unknown",
                "email": row.email or "",
                "action": row.action or "",
                "page": row.page or "",
                "endpoint": row.endpoint or "",
                "method": row.method or "",
                "ip_address": row.ip_address or "Unknown",
                "country": row.country or "Unknown",
                "city": row.city or "Unknown",
                "device": row.device or "Unknown",
                "browser": row.browser or "Unknown",
                "os": row.os or "Unknown",
                "referrer": row.referrer or "",
                "session_id": row.session_id or "",
                "created_at": row.created_at or now,
                "dataset_name": _extract_dataset_name(row.page) or _extract_dataset_name(row.referrer) or "",
            }
        )
    frame = pd.DataFrame(rows)
    frame["created_at"] = pd.to_datetime(frame["created_at"])
    frame["date"] = frame["created_at"].dt.strftime("%Y-%m-%d")
    frame["hour"] = frame["created_at"].dt.hour

    last_24h = frame[frame["created_at"] >= (now - timedelta(hours=24))]
    last_7d = frame[frame["created_at"] >= (now - timedelta(days=7))]
    last_30d = frame[frame["created_at"] >= (now - timedelta(days=30))]
    active_users_24h = int(last_24h["user_id"].dropna().nunique())
    active_users_7d = int(last_7d["user_id"].dropna().nunique())
    active_users_30d = int(last_30d["user_id"].dropna().nunique())

    def _top_pairs(series: pd.Series, *, limit: int = 8, exclude_unknown: bool = False) -> list[dict[str, object]]:
        cleaned = series.fillna("").astype(str).str.strip()
        if exclude_unknown:
            cleaned = cleaned[~cleaned.isin(["", "Unknown"])]
        else:
            cleaned = cleaned[cleaned != ""]
        return [{"name": str(idx), "value": int(val)} for idx, val in cleaned.value_counts().head(limit).items()]

    total_logins = int((frame["action"] == "login").sum())
    total_page_views = int(frame["action"].isin(["page_visit", "search_usage"]).sum())
    total_api_requests = int((frame["action"] == "api_request").sum())
    total_uploads = int((frame["action"] == "data_upload").sum())
    total_searches = int((frame["action"] == "search_usage").sum())
    ccc_api_usage_count = int((frame["action"] == "ccc_api_sync").sum())
    mapping_runs_count = int((frame["action"] == "mapping_run").sum())
    forecast_runs_count = int((frame["action"] == "forecasting_run").sum())
    audit_dataset_runs = int((frame["action"] == "audit_dataset_generation").sum())
    try:
        unmapped_open_count = int(MappingUnmappedRow.query.filter_by(review_status="open").count())
    except Exception:
        unmapped_open_count = 0

    session_frame = frame[frame["session_id"].astype(str).str.strip() != ""].copy()
    session_durations: list[float] = []
    session_page_counts: list[int] = []
    if not session_frame.empty:
        grouped_sessions = session_frame.groupby("session_id")["created_at"].agg(["min", "max"])
        session_durations = [
            max(0.0, float((row["max"] - row["min"]).total_seconds() / 60.0))
            for _, row in grouped_sessions.iterrows()
        ]
        session_page_counts = [
            int(value)
            for value in session_frame[session_frame["action"].isin(["page_visit", "search_usage"])]
            .groupby("session_id")
            .size()
            .tolist()
        ]
    unique_sessions = int(session_frame["session_id"].nunique()) if not session_frame.empty else 0
    avg_session_duration_minutes = round(sum(session_durations) / len(session_durations), 1) if session_durations else 0.0
    pages_per_session = round(total_page_views / unique_sessions, 2) if unique_sessions else 0.0
    bounce_rate = round((sum(1 for value in session_page_counts if value <= 1) / len(session_page_counts)) * 100.0, 1) if session_page_counts else 0.0

    feature_user_sets = {
        "mapping": set(frame.loc[frame["action"] == "mapping_run", "user_id"].dropna().astype(int).tolist()),
        "ccc_api": set(frame.loc[frame["action"] == "ccc_api_sync", "user_id"].dropna().astype(int).tolist()),
        "forecasting": set(frame.loc[frame["action"] == "forecasting_run", "user_id"].dropna().astype(int).tolist()),
        "audit_export": set(frame.loc[frame["action"] == "audit_dataset_generation", "user_id"].dropna().astype(int).tolist()),
        "search": set(frame.loc[frame["action"] == "search_usage", "user_id"].dropna().astype(int).tolist()),
    }
    feature_adoption = [
        {
            "name": key.replace("_", " ").title(),
            "value": round((len(user_ids) / total_users) * 100.0, 1) if total_users else 0.0,
        }
        for key, user_ids in feature_user_sets.items()
    ]
    feature_adoption_rate = round(sum(item["value"] for item in feature_adoption) / len(feature_adoption), 1) if feature_adoption else 0.0

    upload_actions = {"data_upload"}
    company_engagement: list[dict[str, object]] = []
    for company_name, group in frame.groupby("company_name"):
        company_sessions = [sid for sid in group["session_id"].astype(str).tolist() if sid]
        session_minutes = 0.0
        if company_sessions:
            company_session_frame = session_frame[session_frame["session_id"].isin(company_sessions)]
            if not company_session_frame.empty:
                local_group = company_session_frame.groupby("session_id")["created_at"].agg(["min", "max"])
                local_minutes = [
                    max(0.0, float((row["max"] - row["min"]).total_seconds() / 60.0))
                    for _, row in local_group.iterrows()
                ]
                session_minutes = sum(local_minutes) / len(local_minutes) if local_minutes else 0.0
        score = (
            int((group["action"] == "login").sum()) * 3
            + int(group["action"].isin(upload_actions).sum()) * 4
            + int((group["action"] == "mapping_run").sum()) * 5
            + int((group["action"] == "ccc_api_sync").sum()) * 4
            + float(session_minutes)
            + int(group["user_id"].dropna().nunique()) * 6
        )
        company_engagement.append(
            {
                "name": company_name,
                "value": round(score, 1),
            }
        )
    company_engagement.sort(key=lambda item: float(item["value"]), reverse=True)
    top_company_engagement_score = float(company_engagement[0]["value"]) if company_engagement else 0.0

    dataset_usage_frequency = _top_pairs(frame["dataset_name"], limit=8, exclude_unknown=True)
    most_used_dataset = dataset_usage_frequency[0]["name"] if dataset_usage_frequency else "None yet"

    feature_usage_counter = Counter(
        {
            "Login": total_logins,
            "Search": int((frame["action"] == "search_usage").sum()),
            "Mapping": mapping_runs_count,
            "Open No Match Rows": unmapped_open_count,
            "CCC API": ccc_api_usage_count,
            "Forecasting": forecast_runs_count,
            "Audit Export": audit_dataset_runs,
            "Page Views": total_page_views,
        }
    )
    feature_usage_frequency = [{"name": name, "value": int(value)} for name, value in feature_usage_counter.items()]

    most_visited_pages = _top_pairs(frame["page"], limit=8)
    most_active_companies = _top_pairs(frame["company_name"], limit=8, exclude_unknown=True)
    top_countries = _top_pairs(frame["country"], limit=8, exclude_unknown=True)
    top_cities = _top_pairs(frame["city"], limit=8, exclude_unknown=True)
    top_referrers = _top_pairs(frame["referrer"], limit=8, exclude_unknown=True)
    top_browsers = _top_pairs(frame["browser"], limit=8, exclude_unknown=True)
    top_devices = _top_pairs(frame["device"], limit=8, exclude_unknown=True)
    top_operating_systems = _top_pairs(frame["os"], limit=8, exclude_unknown=True)
    most_active_hours = [{"name": f"{int(idx):02d}:00", "value": int(val)} for idx, val in frame["hour"].value_counts().sort_index().items()]
    peak_activity_hour = max(most_active_hours, key=lambda item: int(item["value"]))["name"] if most_active_hours else "None yet"
    action_distribution = _top_pairs(frame["action"], limit=12)
    recent_visits = [
        {
            "email": str(row.get("email") or "Unknown"),
            "company_name": str(row.get("company_name") or "Unknown"),
            "action": str(row.get("action") or ""),
            "page": str(row.get("page") or ""),
            "ip_address": str(row.get("ip_address") or "Unknown"),
            "location": ", ".join([v for v in [str(row.get("city") or ""), str(row.get("country") or "")] if v and v != "Unknown"]) or "Unknown",
            "device": str(row.get("device") or "Unknown"),
            "browser": str(row.get("browser") or "Unknown"),
            "created_at": row.get("created_at").strftime("%Y-%m-%d %H:%M") if hasattr(row.get("created_at"), "strftime") else "",
        }
        for row in frame.sort_values("created_at", ascending=False).head(30).to_dict(orient="records")
    ]
    user_last_seen = []
    for email, group in frame[frame["email"].astype(str).str.strip() != ""].groupby("email"):
        latest = group.sort_values("created_at", ascending=False).iloc[0]
        first_seen = group["created_at"].min()
        user_last_seen.append(
            {
                "email": str(email),
                "company_name": str(latest.get("company_name") or "Unknown"),
                "last_seen": latest["created_at"].strftime("%Y-%m-%d %H:%M"),
                "first_seen": first_seen.strftime("%Y-%m-%d %H:%M") if hasattr(first_seen, "strftime") else "",
                "visits": int(len(group.index)),
                "last_page": str(latest.get("page") or ""),
                "location": ", ".join([v for v in [str(latest.get("city") or ""), str(latest.get("country") or "")] if v and v != "Unknown"]) or "Unknown",
                "device": str(latest.get("device") or "Unknown"),
            }
        )
    user_last_seen.sort(key=lambda item: str(item["last_seen"]), reverse=True)

    daily_active_users = (
        frame.groupby("date")["user_id"].nunique().sort_index()
    )
    session_duration_distribution = [
        {"name": "0-5 min", "value": sum(1 for value in session_durations if value <= 5)},
        {"name": "5-15 min", "value": sum(1 for value in session_durations if 5 < value <= 15)},
        {"name": "15-30 min", "value": sum(1 for value in session_durations if 15 < value <= 30)},
        {"name": "30-60 min", "value": sum(1 for value in session_durations if 30 < value <= 60)},
        {"name": "60+ min", "value": sum(1 for value in session_durations if value > 60)},
    ]

    metrics = {
        "total_users": total_users,
        "active_users_24h": active_users_24h,
        "active_users_7d": active_users_7d,
        "active_users_30d": active_users_30d,
        "unique_sessions": unique_sessions,
        "total_logins": total_logins,
        "total_page_views": total_page_views,
        "total_api_requests": total_api_requests,
        "total_uploads": total_uploads,
        "total_searches": total_searches,
        "ccc_api_usage_count": ccc_api_usage_count,
        "mapping_runs_count": mapping_runs_count,
        "unmapped_open_count": unmapped_open_count,
        "forecast_runs_count": forecast_runs_count,
        "audit_dataset_runs": audit_dataset_runs,
        "avg_session_duration_minutes": avg_session_duration_minutes,
        "pages_per_session": pages_per_session,
        "bounce_rate": bounce_rate,
        "feature_adoption_rate": feature_adoption_rate,
        "top_company_engagement_score": top_company_engagement_score,
        "most_used_dataset": most_used_dataset,
        "peak_activity_hour": peak_activity_hour,
        "most_visited_pages": most_visited_pages,
        "most_active_companies": most_active_companies,
        "top_countries": top_countries,
        "top_cities": top_cities,
        "top_referrers": top_referrers,
        "top_browsers": top_browsers,
        "top_devices": top_devices,
        "top_operating_systems": top_operating_systems,
        "most_active_hours": most_active_hours,
        "feature_usage_frequency": feature_usage_frequency,
        "action_distribution": action_distribution,
        "dataset_usage_frequency": dataset_usage_frequency,
        "company_engagement": company_engagement[:8],
        "recent_visits": recent_visits,
        "user_last_seen": user_last_seen[:30],
    }
    chart_data = {
        "daily_active_users": [{"name": str(idx), "value": int(val)} for idx, val in daily_active_users.items()],
        "activity_by_hour": most_active_hours,
        "top_pages": most_visited_pages,
        "country_distribution": top_countries,
        "city_distribution": top_cities,
        "browser_distribution": top_browsers,
        "device_distribution": top_devices,
        "os_distribution": top_operating_systems,
        "company_distribution": most_active_companies,
        "feature_usage": feature_usage_frequency,
        "action_distribution": action_distribution,
        "referrer_distribution": top_referrers,
        "dataset_usage": dataset_usage_frequency,
        "session_duration_distribution": session_duration_distribution,
    }
    return {
        "metrics": metrics,
        "chart_data": chart_data,
        "activity_rows": int(len(frame.index)),
    }


def _create_user_notification(
    user_id: int,
    *,
    title: str,
    message: str,
    notification_type: str = "info",
    link: str | None = None,
    mapping_card: dict[str, object] | None = None,
    feed_event: str | None = None,
    feed_company: str | None = None,
    feed_api_name: str | None = None,
    feed_timestamp: datetime | None = None,
) -> None:
    try:
        notification_service.create_notification(
            db.session,
            Notification,
            user_id=int(user_id),
            title=title,
            message=message,
            notification_type=notification_type,
            link=link,
            meta=mapping_card,
        )
        db.session.commit()
        if str(notification_type or "").strip().lower() == "success":
            _create_system_feed_post_for_event(
                event_key=feed_event,
                company_name=feed_company,
                api_name=feed_api_name,
                event_timestamp=feed_timestamp,
            )
    except Exception:
        db.session.rollback()


def _user_profile_complete(u: object) -> bool:
    if u is None or not getattr(u, "is_authenticated", False):
        return True
    v = getattr(u, "is_profile_complete", None)
    if v is None:
        return True
    return bool(v)


def _company_logo_slug_filename(company_key: str) -> str:
    raw = (company_key or "").strip()
    slug = re.sub(r"[^0-9a-zA-Z]+", "_", raw).strip("_") or "company"
    return f"{slug}.png"


def _static_subdir(*parts: str) -> Path:
    p = APP_DIR / "static"
    for part in parts:
        p = p / part
    p.mkdir(parents=True, exist_ok=True)
    return p


def _save_upload_image(
    storage,
    dest_dir: Path,
    base_name: str,
    allowed_ext: frozenset[str],
) -> str | None:
    if not storage or not getattr(storage, "filename", None):
        return None
    fn = secure_filename(storage.filename or "")
    ext = Path(fn).suffix.lower()
    if ext not in allowed_ext:
        return None
    dest_dir.mkdir(parents=True, exist_ok=True)
    dest = dest_dir / f"{base_name}{ext}"
    storage.save(str(dest))
    rel = dest.relative_to(APP_DIR / "static").as_posix()
    return rel


def _canonical_company_logo_slug_rel(template_company_key: str) -> str | None:
    """Official logo file under static/company_logos/{slug}.png (company-controlled)."""
    key = (template_company_key or "").strip()
    if not key:
        return None
    slug_rel = f"company_logos/{_company_logo_slug_filename(key)}"
    try:
        if (APP_DIR / "static" / slug_rel).is_file():
            return slug_rel
    except Exception:
        pass
    return None


def _company_logo_static_rel(template_company_key: str) -> str | None:
    """Prefer canonical slug file on disk; fall back to Company row path (legacy)."""
    key = (template_company_key or "").strip()
    if not key:
        return None
    slug_rel = _canonical_company_logo_slug_rel(key)
    if slug_rel:
        return slug_rel
    row = Company.query.filter_by(company_name=key).first()
    if not row or not row.company_logo_path:
        return None
    p = APP_DIR / "static" / row.company_logo_path
    try:
        if p.is_file():
            return row.company_logo_path
    except Exception:
        pass
    return None


def _save_profile_photo_file(storage, user_id: int) -> str | None:
    if not storage or not getattr(storage, "filename", None):
        return None
    fn = secure_filename(storage.filename or "")
    ext = Path(fn).suffix.lower()
    if ext not in PROFILE_PHOTO_ALLOWED_EXT:
        return None
    PROFILE_PHOTOS_STORAGE_DIR.mkdir(parents=True, exist_ok=True)
    dest = PROFILE_PHOTOS_STORAGE_DIR / f"user_{user_id}_{uuid.uuid4().hex[:10]}{ext}"
    storage.save(str(dest))
    return f"profile_photos/{dest.name}"


def _save_profile_cover_file(storage, user_id: int) -> str | None:
    if not storage or not getattr(storage, "filename", None):
        return None
    fn = secure_filename(storage.filename or "")
    ext = Path(fn).suffix.lower()
    if ext not in PROFILE_PHOTO_ALLOWED_EXT:
        return None
    dest_dir = _static_subdir("uploads", "covers")
    dest = dest_dir / f"user_{user_id}_{uuid.uuid4().hex[:10]}{ext}"
    storage.save(str(dest))
    return dest.relative_to(APP_DIR / "static").as_posix()


_AVATAR_COLOR_PAIRS: tuple[tuple[str, str], ...] = (
    ("#dbeafe", "#1d4ed8"),
    ("#dcfce7", "#166534"),
    ("#fae8ff", "#9333ea"),
    ("#fee2e2", "#b91c1c"),
    ("#fef3c7", "#b45309"),
    ("#e0f2fe", "#0369a1"),
)


def _initials(value: str) -> str:
    parts = [part for part in str(value or "").strip().split() if part]
    if not parts:
        return "CP"
    return "".join(part[0].upper() for part in parts[:2]) or "CP"


def _svg_data_url(markup: str) -> str:
    return "data:image/svg+xml;utf8," + quote(markup)


def _default_avatar_url(seed_text: str) -> str:
    seed = str(seed_text or "").strip() or "Carbon Platform"
    bg, fg = _AVATAR_COLOR_PAIRS[sum(ord(ch) for ch in seed) % len(_AVATAR_COLOR_PAIRS)]
    initials = _initials(seed)
    svg = (
        "<svg xmlns='http://www.w3.org/2000/svg' width='160' height='160' viewBox='0 0 160 160'>"
        f"<rect width='160' height='160' rx='80' fill='{bg}'/>"
        f"<text x='50%' y='54%' text-anchor='middle' font-family='Arial, sans-serif' font-size='56' font-weight='700' fill='{fg}'>{initials}</text>"
        "</svg>"
    )
    return _svg_data_url(svg)


def _profile_photo_safe_basename(user_id: int, rel: str | None) -> str | None:
    if not rel:
        return None
    name = Path(str(rel).replace("\\", "/")).name
    if not name or name in (".", ".."):
        return None
    ext = Path(name).suffix.lower()
    if ext not in PROFILE_PHOTO_ALLOWED_EXT:
        return None
    prefix = f"user_{int(user_id)}_"
    if not name.startswith(prefix):
        return None
    if ".." in name or "/" in name or "\\" in str(rel):
        return None
    return name


def _profile_photo_disk_path_for_user(u: User) -> Path | None:
    rel = getattr(u, "profile_photo_path", None)
    uid = int(getattr(u, "id", 0) or 0)
    if uid <= 0:
        return None
    name = _profile_photo_safe_basename(uid, rel)
    if not name:
        return None
    try:
        primary = (PROFILE_PHOTOS_STORAGE_DIR / name).resolve()
        root = PROFILE_PHOTOS_STORAGE_DIR.resolve()
        primary.relative_to(root)
    except (ValueError, OSError):
        return None
    if primary.is_file():
        return primary
    try:
        legacy = (APP_DIR / "static" / "profile_photos" / name).resolve()
        leg_root = (APP_DIR / "static" / "profile_photos").resolve()
        legacy.relative_to(leg_root)
    except (ValueError, OSError):
        return None
    if legacy.is_file():
        return legacy
    return None


def _default_avatar_svg_xml(seed_text: str) -> str:
    seed = str(seed_text or "").strip() or "Carbon Platform"
    bg, fg = _AVATAR_COLOR_PAIRS[sum(ord(ch) for ch in seed) % len(_AVATAR_COLOR_PAIRS)]
    ini = html.escape(_initials(seed), quote=True)
    return (
        "<?xml version='1.0' encoding='UTF-8'?>"
        f"<svg xmlns='http://www.w3.org/2000/svg' width='160' height='160' viewBox='0 0 160 160'>"
        f"<rect width='160' height='160' rx='80' fill='{bg}'/>"
        f"<text x='50%' y='54%' text-anchor='middle' font-family='Arial, sans-serif' "
        f"font-size='56' font-weight='700' fill='{fg}'>{ini}</text></svg>"
    )


@app.route("/api/profile-photo/<int:user_id>")
@login_required
def api_profile_photo(user_id: int):
    """Serve uploaded profile photos only to authenticated users (not public static)."""
    target = db.session.get(User, user_id)
    if target is None:
        abort(403)
    disk_path = _profile_photo_disk_path_for_user(target)
    headers = {"Cache-Control": "private, max-age=86400"}
    if disk_path and disk_path.is_file():
        ext = disk_path.suffix.lower()
        mt = {
            ".png": "image/png",
            ".jpg": "image/jpeg",
            ".jpeg": "image/jpeg",
            ".webp": "image/webp",
        }.get(ext) or mimetypes.guess_type(disk_path.name)[0] or "application/octet-stream"
        return send_file(
            str(disk_path),
            mimetype=mt,
            max_age=86400,
            conditional=True,
            etag=True,
        )
    body = _default_avatar_svg_xml(_user_display_name(target))
    return Response(body, mimetype="image/svg+xml; charset=utf-8", headers=headers)


def _default_company_logo_url(company_name: str) -> str:
    seed = str(company_name or "").strip() or "Company"
    initials = _initials(seed[:2] if len(seed.split()) == 1 else seed)
    svg = (
        "<svg xmlns='http://www.w3.org/2000/svg' width='160' height='160' viewBox='0 0 160 160'>"
        "<rect width='160' height='160' rx='28' fill='#eef2f7'/>"
        "<rect x='24' y='32' width='112' height='96' rx='18' fill='#dbe3ee'/>"
        f"<text x='50%' y='57%' text-anchor='middle' font-family='Arial, sans-serif' font-size='42' font-weight='700' fill='#4b5563'>{initials}</text>"
        "</svg>"
    )
    return _svg_data_url(svg)


def _user_avatar_url(u: User | None) -> str:
    photo = _profile_photo_url_for_user(u)
    if photo:
        return photo
    if bool(getattr(u, "is_admin", False)):
        return url_for("static", filename="images/logo.svg.png")
    return _default_avatar_url(_user_display_name(u))


def _company_logo_url(company_name: str | None) -> str:
    key = (company_name or "").strip()
    rel = _company_logo_static_rel(key) if key else None
    if rel:
        return url_for("static", filename=rel)
    return _default_company_logo_url(key)


def _normalize_feed_post_type(raw: object) -> str:
    value = str(raw or "").strip().lower()
    return value if value in FEED_POST_TYPES_SET else "update"


def _normalize_feed_reference_type(raw: object) -> str:
    value = str(raw or "").strip().lower()
    return value if value in FEED_REFERENCE_TYPES else ""


def _normalize_feed_reaction_type(raw: object) -> str:
    value = str(raw or "").strip().lower()
    return value if value in FEED_REACTION_TYPES_SET else ""


def _feed_reaction_button_state(current_reaction: object | None) -> dict[str, object]:
    normalized = _normalize_feed_reaction_type(current_reaction)
    meta = FEED_REACTION_META.get(normalized) or FEED_REACTION_META["like"]
    return {
        "type": normalized,
        "label": str(meta.get("label") or "Like"),
        "icon": str(meta.get("icon") or "👍"),
        "is_active": bool(normalized),
    }


def _build_feed_reaction_summary(counts_by_type: dict[str, int] | Counter[str]) -> list[dict[str, object]]:
    items: list[dict[str, object]] = []
    for index, reaction_type in enumerate(FEED_REACTION_TYPES):
        count = int(counts_by_type.get(reaction_type) or 0)
        if count <= 0:
            continue
        meta = FEED_REACTION_META.get(reaction_type) or {}
        items.append(
            {
                "type": reaction_type,
                "label": str(meta.get("label") or reaction_type.title()),
                "icon": str(meta.get("icon") or ""),
                "count": count,
                "_order": index,
            }
        )
    items.sort(key=lambda item: (-int(item.get("count") or 0), int(item.get("_order") or 0)))
    return [{key: value for key, value in item.items() if key != "_order"} for item in items[:3]]


def _feed_reaction_maps(post_ids: list[int], user_id: int | None = None) -> tuple[dict[int, list[dict[str, object]]], dict[int, str]]:
    if not post_ids:
        return {}, {}

    summary_counts: dict[int, Counter[str]] = defaultdict(Counter)
    aggregate_rows = (
        db.session.query(
            PostReaction.post_id,
            PostReaction.reaction_type,
            db.func.count(PostReaction.id),
        )
        .filter(PostReaction.post_id.in_(post_ids))
        .group_by(PostReaction.post_id, PostReaction.reaction_type)
        .all()
    )
    for post_id, reaction_type, count in aggregate_rows:
        normalized = _normalize_feed_reaction_type(reaction_type)
        if not normalized:
            continue
        summary_counts[int(post_id)][normalized] = int(count or 0)

    summary_map = {
        int(post_id): _build_feed_reaction_summary(counter)
        for post_id, counter in summary_counts.items()
    }

    current_map: dict[int, str] = {}
    if user_id:
        current_rows = (
            PostReaction.query.filter(
                PostReaction.post_id.in_(post_ids),
                PostReaction.user_id == int(user_id),
            )
            .all()
        )
        for row in current_rows:
            normalized = _normalize_feed_reaction_type(getattr(row, "reaction_type", None))
            if normalized:
                current_map[int(row.post_id)] = normalized

    return summary_map, current_map


def _platform_owner_user() -> User | None:
    return (
        User.query.filter(db.func.lower(User.role) == "owner")
        .order_by(User.created_at.asc(), User.id.asc())
        .first()
    )


def _create_feed_post_record(
    *,
    author_id: int,
    content: str,
    post_type: str = "alert",
    reference_id: int | None = None,
    reference_type: str | None = None,
    created_at: datetime | None = None,
) -> FeedPost | None:
    clean_content = str(content or "").strip()
    if not author_id or not clean_content:
        return None
    normalized_post_type = _normalize_feed_post_type(post_type)
    normalized_reference_type = _normalize_feed_reference_type(reference_type)
    post_created_at = created_at if isinstance(created_at, datetime) else datetime.utcnow()
    duplicate_window_start = post_created_at - timedelta(seconds=90)
    duplicate_window_end = post_created_at + timedelta(seconds=90)
    existing = (
        FeedPost.query.filter(
            FeedPost.author_user_id == int(author_id),
            FeedPost.post_type == normalized_post_type,
            FeedPost.content == clean_content,
            FeedPost.reference_id == (int(reference_id) if reference_id else None),
            FeedPost.reference_type == (normalized_reference_type or None),
            FeedPost.created_at >= duplicate_window_start,
            FeedPost.created_at <= duplicate_window_end,
        )
        .order_by(FeedPost.created_at.desc(), FeedPost.id.desc())
        .first()
    )
    if existing is not None:
        return existing

    row = FeedPost(
        author_user_id=int(author_id),
        content=clean_content,
        post_type=normalized_post_type,
        reference_id=int(reference_id) if reference_id else None,
        reference_type=normalized_reference_type or None,
        created_at=post_created_at,
    )
    try:
        db.session.add(row)
        db.session.commit()
        return row
    except Exception:
        db.session.rollback()
        return None


def _create_content_feed_post(
    *,
    author_id: int,
    reference_type: str,
    reference_id: int,
    title: str,
) -> FeedPost | None:
    normalized_reference_type = _normalize_feed_reference_type(reference_type)
    normalized_title = str(title or "").strip() or "Untitled"
    message_map = {
        "report": f"New report published: {normalized_title}",
        "newsletter": f"New newsletter published: {normalized_title}",
        "event": f"New event created: {normalized_title}",
        "award": "New Sustainability Award open for submissions",
    }
    message = message_map.get(normalized_reference_type, "")
    if not message:
        return None
    return _create_feed_post_record(
        author_id=int(author_id),
        content=message,
        post_type=normalized_reference_type,
        reference_id=int(reference_id),
        reference_type=normalized_reference_type,
    )


def _build_system_feed_post_message(
    *,
    event_key: str | None,
    company_name: str | None = None,
    api_name: str | None = None,
) -> str:
    event_name = str(event_key or "").strip().lower()
    company_label = _clean_company_name(company_name or "") or "Platform"
    api_label = str(api_name or "").strip() or "API"
    if event_name == "data_upload":
        return f"{company_label} uploaded new data successfully."
    if event_name == "api_connection":
        return f"{company_label} connected to {api_label}."
    if event_name == "mapping_completed":
        return f"Mapping completed for {company_label}."
    if event_name == "pipeline_completed":
        return f"Data pipeline executed successfully for {company_label}."
    return ""


def _create_system_feed_post_for_event(
    *,
    event_key: str | None,
    company_name: str | None = None,
    api_name: str | None = None,
    event_timestamp: datetime | None = None,
) -> FeedPost | None:
    message = _build_system_feed_post_message(
        event_key=event_key,
        company_name=company_name,
        api_name=api_name,
    )
    if not message:
        return None
    owner = _platform_owner_user()
    if owner is None:
        return None
    return _create_feed_post_record(
        author_id=int(owner.id),
        content=message,
        post_type="alert",
        created_at=event_timestamp,
    )


def _normalize_feed_filter(raw: object) -> str:
    value = str(raw or "").strip().lower()
    return value if value in FEED_FILTER_OPTIONS else "all"


def _clean_company_name(value: object) -> str:
    raw = str(value or "").strip()
    return _resolve_template_company_name(raw) or raw


def _safe_float(value: object) -> float | None:
    if value in (None, ""):
        return None
    try:
        return float(str(value).strip())
    except Exception:
        return None


def _positive_ratio(numerator: float | None, denominator: float | None) -> float | None:
    if numerator is None or denominator is None or denominator <= 0:
        return None
    return numerator / denominator


def _is_owner_user(u: object | None) -> bool:
    return normalize_user_role(getattr(u, "role", None)) == "owner"


def _is_auditor_user(u: object | None) -> bool:
    return normalize_user_role(getattr(u, "role", None)) == "auditor"


def _is_readonly_user(u: object | None) -> bool:
    return _is_auditor_user(u)


def _allowed_companies_for_averages() -> list[str]:
    if _is_owner_user(current_user):
        return list(COMPANIES)
    company_name = _clean_company_name(getattr(current_user, "company_name", "") or "")
    return [company_name] if company_name else []


def _resolve_averages_company(raw_company: object | None = None) -> tuple[str | None, str | None]:
    allowed_companies = _allowed_companies_for_averages()
    if _is_owner_user(current_user):
        company_name = _clean_company_name(raw_company or "")
        if not company_name:
            company_name = _clean_company_name(getattr(current_user, "company_name", "") or "")
        if not company_name and allowed_companies:
            company_name = allowed_companies[0]
        if company_name not in allowed_companies:
            return None, "Select a valid company."
        return company_name, None

    company_name = _clean_company_name(getattr(current_user, "company_name", "") or "")
    if not company_name:
        return None, "Company is required."
    return company_name, None


def _format_relative_datetime(value: datetime | None) -> str:
    if not value:
        return "No data saved yet"
    now = datetime.utcnow()
    delta = now - value
    seconds = max(0, int(delta.total_seconds()))
    if seconds < 60:
        return "Last updated: just now"
    if seconds < 3600:
        minutes = seconds // 60
        return f"Last updated: {minutes} minute{'s' if minutes != 1 else ''} ago"
    if seconds < 86400:
        hours = seconds // 3600
        return f"Last updated: {hours} hour{'s' if hours != 1 else ''} ago"
    days = seconds // 86400
    if days < 7:
        return f"Last updated: {days} day{'s' if days != 1 else ''} ago"
    return "Last updated: " + value.strftime("%d %b %Y %H:%M")


def _averages_payload(row: AveragesData | None) -> dict[str, object]:
    if row is None:
        return {
            "electricity_kwh": "",
            "electricity_country": "",
            "electricity_emission_factor": "",
            "district_heating_kwh": "",
            "district_heating_supplier": "",
            "waste_type": "",
            "waste_weight": "",
            "waste_unit": "",
            "water_total_m3": "",
            "building_size_m2": "",
            "water_per_m2": "",
            "has_data": False,
            "updated_at_label": "No data saved yet",
            "updated_at_iso": "",
            "summary": {
                "electricity_kwh": "--",
                "district_heating_kwh": "--",
                "waste_total": "--",
                "water_per_m2": "--",
            },
        }
    waste_total = "--"
    if row.waste_weight is not None:
        waste_total = f"{row.waste_weight:g} {row.waste_unit or ''}".strip()
    water_per_m2 = row.water_per_m2 if row.water_per_m2 is not None else ""
    has_data = any(
        value not in (None, "", 0)
        for value in (
            row.electricity_kwh,
            row.district_heating_kwh,
            row.waste_weight,
            row.water_total_m3,
            row.building_size_m2,
        )
    )
    return {
        "electricity_kwh": row.electricity_kwh if row.electricity_kwh is not None else "",
        "electricity_country": row.electricity_country or "",
        "electricity_emission_factor": row.electricity_emission_factor if row.electricity_emission_factor is not None else "",
        "district_heating_kwh": row.district_heating_kwh if row.district_heating_kwh is not None else "",
        "district_heating_supplier": row.district_heating_supplier or "",
        "waste_type": row.waste_type or "",
        "waste_weight": row.waste_weight if row.waste_weight is not None else "",
        "waste_unit": row.waste_unit or "",
        "water_total_m3": row.water_total_m3 if row.water_total_m3 is not None else "",
        "building_size_m2": row.building_size_m2 if row.building_size_m2 is not None else "",
        "water_per_m2": water_per_m2,
        "has_data": has_data,
        "updated_at_label": _format_relative_datetime(getattr(row, "updated_at", None)),
        "updated_at_iso": getattr(row, "updated_at", None).isoformat() if getattr(row, "updated_at", None) else "",
        "summary": {
            "electricity_kwh": f"{row.electricity_kwh:g}" if row.electricity_kwh is not None else "--",
            "district_heating_kwh": f"{row.district_heating_kwh:g}" if row.district_heating_kwh is not None else "--",
            "waste_total": waste_total,
            "water_per_m2": f"{row.water_per_m2:.4f}" if row.water_per_m2 is not None else "--",
        },
    }


def _default_scenario_inputs(categories: tuple[str, ...]) -> dict[str, dict[str, object]]:
    data: dict[str, dict[str, object]] = {}
    if "9" in categories:
        data["9"] = {
            "annual_production": "",
            "transport_distance": "",
            "transport_type": "",
        }
    if "11" in categories:
        data["11"] = {
            "number_of_products_in_use": "",
            "usage_factor": "",
        }
    if "12" in categories:
        data["12"] = {
            "product_weight": "",
            "waste_type": "",
            "disposal_method": "",
        }
    return data


def _scenario_rows_payload(rows: list[ScenariosData]) -> dict[str, dict[str, object]]:
    out: dict[str, dict[str, object]] = {}
    for row in rows:
        company_name = _clean_company_name(getattr(row, "company_name", None))
        categories = SCENARIO_CATEGORY_CONFIG.get(company_name, tuple())
        payload = _default_scenario_inputs(categories)
        raw = getattr(row, "categories_json", None)
        try:
            saved = json.loads(str(raw or "{}"))
        except Exception:
            saved = {}
        if isinstance(saved, dict):
            for key, value in saved.items():
                if key in payload and isinstance(value, dict):
                    payload[key].update({k: value.get(k, "") for k in payload[key].keys()})
        summary_values: list[str] = []
        if "9" in payload:
            summary_values.append("Production: " + (str(payload["9"].get("annual_production") or "--")))
        if "11" in payload:
            summary_values.append("Usage factor: " + (str(payload["11"].get("usage_factor") or "--")))
        if "12" in payload:
            summary_values.append("Weight: " + (str(payload["12"].get("product_weight") or "--")))
        has_data = any(
            str(field_value or "").strip()
            for category_data in payload.values()
            for field_value in category_data.values()
        )
        out[company_name] = {
            "inputs": payload,
            "active_categories": list(categories),
            "has_data": has_data,
            "updated_at_label": _format_relative_datetime(getattr(row, "updated_at", None)),
            "updated_at_iso": getattr(row, "updated_at", None).isoformat() if getattr(row, "updated_at", None) else "",
            "summary_values": summary_values,
        }
    for company_name in SCENARIO_COMPANY_OPTIONS:
        if company_name not in out:
            out[company_name] = {
                "inputs": _default_scenario_inputs(SCENARIO_CATEGORY_CONFIG.get(company_name, tuple())),
                "active_categories": list(SCENARIO_CATEGORY_CONFIG.get(company_name, tuple())),
                "has_data": False,
                "updated_at_label": "No data saved yet",
                "updated_at_iso": "",
                "summary_values": [],
            }
    return out


def _make_template_row(company_name: str, sheet_name: str, values: dict[str, object]) -> pd.DataFrame:
    headers = _get_template_sheet_headers(company_name, sheet_name)
    if not headers:
        raise RuntimeError(f"Template headers not found for {sheet_name}.")
    row = {header: "" for header in headers}
    for key, value in values.items():
        if key in row:
            row[key] = value
    return pd.DataFrame([[row.get(header, "") for header in headers]], columns=headers)


def _averages_mapping_frames(company_name: str, payload: dict[str, object]) -> list[tuple[str, pd.DataFrame]]:
    period_label = datetime.utcnow().strftime("%B %Y")
    _canon_company, canon_country = _canonical_company_name_and_country(company_name)
    company_country = str(payload.get("electricity_country") or "").strip() or (canon_country or "")
    shared_source = "Averages input"
    frames: list[tuple[str, pd.DataFrame]] = []

    electricity_kwh = _safe_float(payload.get("electricity_kwh"))
    if electricity_kwh is not None and electricity_kwh > 0:
        frames.append(
            (
                "Scope 2 Electricity",
                _make_template_row(
                    company_name,
                    "Scope 2 Electricity",
                    {
                        "Reporting period (month, year)": period_label,
                        "Consumption": electricity_kwh,
                        "Unit": "kWh",
                        "Country": company_country,
                        "Site Tag": company_name,
                        "Data Source": shared_source,
                    },
                ),
            )
        )

    district_heating_kwh = _safe_float(payload.get("district_heating_kwh"))
    if district_heating_kwh is not None and district_heating_kwh > 0:
        supplier = str(payload.get("district_heating_supplier") or "").strip()
        frames.append(
            (
                "Scope 2 District Heating",
                _make_template_row(
                    company_name,
                    "Scope 2 District Heating",
                    {
                        "Reporting period (month, year)": period_label,
                        "Country": canon_country or company_country,
                        "Consumption": district_heating_kwh,
                        "Unit": "kWh",
                        "Site Tag": company_name,
                        "Data Source": f"{shared_source}{' - ' + supplier if supplier else ''}",
                    },
                ),
            )
        )

    waste_weight = _safe_float(payload.get("waste_weight"))
    waste_type = str(payload.get("waste_type") or "").strip()
    waste_unit = str(payload.get("waste_unit") or "").strip()
    if waste_weight is not None and waste_weight > 0 and waste_type:
        frames.append(
            (
                "Scope 3 Category 5 Waste",
                _make_template_row(
                    company_name,
                    "Scope 3 Category 5 Waste",
                    {
                        "Reporting period (month, year)": period_label,
                        "Site": company_name,
                        "Waste Stream": waste_type,
                        "Weight": waste_weight,
                        "Weight Unit": waste_unit,
                        "Treatment Method": "",
                        "Country": canon_country or company_country,
                        "Data Source": shared_source,
                    },
                ),
            )
        )

    return frames


def _run_mapping_for_virtual_sheet(user_id: int, company_name: str, sheet_name: str, df: pd.DataFrame) -> dict[str, object]:
    run_id = uuid.uuid4().hex[:12]
    mr = MappingRun(
        id=run_id,
        user_id=int(user_id),
        company_name=company_name,
        sheet_name=sheet_name,
        status="running",
        created_at=datetime.utcnow(),
        source_entry_group="averages",
    )
    try:
        db.session.add(mr)
        db.session.commit()
    except Exception:
        db.session.rollback()

    try:
        mapped_df, out_path, input_path = run_mapping(company_name, sheet_name, df)
    except Exception as exc:
        try:
            mr.status = "failed"
            mr.error_message = str(exc)
            db.session.commit()
        except Exception:
            db.session.rollback()
        return {"ok": False, "sheet": sheet_name, "error": str(exc)}

    total_tco2e, rows_count, _used_col = _sum_tco2e(mapped_df)

    try:
        mr.status = "succeeded"
        mr.output_path = str(out_path)
        mr.input_path = str(input_path)
        db.session.commit()
    except Exception:
        db.session.rollback()

    try:
        _upsert_mapping_run_summary(
            run_id=run_id,
            company_name=company_name,
            sheet_name=sheet_name,
            mapped_df=mapped_df,
            output_path=out_path,
        )
        _sync_unmapped_rows_for_mapping_run(
            run_id=run_id,
            user_id=int(user_id),
            company_name=company_name,
            sheet_name=sheet_name,
            source_entry_group="averages",
            mapped_df=mapped_df,
        )
        db.session.commit()
    except Exception:
        db.session.rollback()

    return {
        "ok": True,
        "sheet": sheet_name,
        "run_id": run_id,
        "rows_count": int(rows_count or 0),
        "tco2e_total": round(float(total_tco2e or 0.0), 6),
    }


def _format_feed_timestamp(value: datetime | None) -> str:
    if not value:
        return ""
    now = datetime.utcnow()
    delta = now - value
    seconds = max(0, int(delta.total_seconds()))
    if seconds < 60:
        return "Just now"
    if seconds < 3600:
        return f"{seconds // 60}m ago"
    if seconds < 86400:
        return f"{seconds // 3600}h ago"
    if seconds < 604800:
        return f"{seconds // 86400}d ago"
    return value.strftime("%d %b %Y %H:%M")


def _feed_media_url(rel_path: str | None) -> str | None:
    rel = str(rel_path or "").strip()
    if not rel:
        return None
    return url_for("static", filename=rel)


def _save_feed_media_file(storage, *, user_id: int) -> tuple[str | None, str | None, str | None]:
    if not storage or not getattr(storage, "filename", None):
        return (None, None, None)
    filename = secure_filename(storage.filename or "")
    ext = Path(filename).suffix.lower()
    if not ext:
        return (None, None, "Unsupported file type.")

    media_type = None
    if ext in FEED_IMAGE_ALLOWED_EXT:
        media_type = "image"
    elif ext in FEED_VIDEO_ALLOWED_EXT:
        media_type = "video"
    elif ext in FEED_FILE_ALLOWED_EXT:
        media_type = "file"
    if not media_type:
        return (None, None, "Unsupported file type.")

    dest_dir = _static_subdir("feed_uploads")
    stored_name = f"user_{user_id}_{datetime.utcnow().strftime('%Y%m%d%H%M%S')}_{uuid.uuid4().hex[:10]}{ext}"
    dest = dest_dir / stored_name
    storage.save(str(dest))
    return (dest.relative_to(APP_DIR / "static").as_posix(), media_type, None)


def _json_list(value: object) -> list[str]:
    if isinstance(value, list):
        return [str(item).strip() for item in value if str(item).strip()]
    try:
        data = json.loads(str(value or "[]"))
    except Exception:
        return []
    if not isinstance(data, list):
        return []
    return [str(item).strip() for item in data if str(item).strip()]


def _json_int_list(value: object) -> list[int]:
    values: list[int] = []
    for item in _json_list(value):
        try:
            parsed = int(item)
        except Exception:
            continue
        if parsed > 0 and parsed not in values:
            values.append(parsed)
    return values


def _company_row_for_name(company_name: str, *, created_by_user_id: int | None = None) -> Company | None:
    resolved_name = _clean_company_name(company_name)
    if not resolved_name:
        return None
    row = Company.query.filter_by(company_name=resolved_name).first()
    if row is not None:
        if created_by_user_id and not row.created_by_user_id:
            row.created_by_user_id = int(created_by_user_id)
        return row
    row = Company(company_name=resolved_name, created_by_user_id=created_by_user_id)
    db.session.add(row)
    db.session.flush()
    return row


def _save_report_file(storage, *, user_id: int) -> tuple[str | None, str | None]:
    if not storage or not getattr(storage, "filename", None):
        return (None, "No report file provided.")
    filename = secure_filename(storage.filename or "")
    ext = Path(filename).suffix.lower()
    if ext not in FEED_FILE_ALLOWED_EXT:
        return (None, "Unsupported report file type.")
    dest_dir = _static_subdir("uploads", "reports")
    stored_name = f"report_{user_id}_{datetime.utcnow().strftime('%Y%m%d%H%M%S')}_{uuid.uuid4().hex[:10]}{ext}"
    dest = dest_dir / stored_name
    storage.save(str(dest))
    return (dest.relative_to(APP_DIR / "static").as_posix(), None)


def _save_module_document(storage, *, user_id: int, module_name: str) -> tuple[str | None, str | None]:
    if not storage or not getattr(storage, "filename", None):
        return (None, "No file provided.")
    filename = secure_filename(storage.filename or "")
    ext = Path(filename).suffix.lower()
    if ext not in MODULE_DOCUMENT_ALLOWED_EXT:
        return (None, "Unsupported file type. Please upload a PDF or DOCX file.")
    dest_dir = _static_subdir("uploads", module_name)
    stored_name = f"{module_name}_{user_id}_{datetime.utcnow().strftime('%Y%m%d%H%M%S')}_{uuid.uuid4().hex[:10]}{ext}"
    dest = dest_dir / stored_name
    storage.save(str(dest))
    return (dest.relative_to(APP_DIR / "static").as_posix(), None)


def _make_report_preview_placeholder(title: str, *, page_number: int, base_name: str) -> str:
    preview_dir = _static_subdir("report_previews")
    dest = preview_dir / f"{base_name}_page_{page_number}.png"
    image = Image.new("RGB", (900, 1200), color=(250, 251, 255))
    draw = ImageDraw.Draw(image)
    font = ImageFont.load_default()
    draw.rounded_rectangle((36, 36, 864, 1164), radius=28, outline=(198, 209, 224), width=3, fill=(255, 255, 255))
    draw.rectangle((74, 92, 826, 210), fill=(234, 242, 255))
    draw.text((96, 120), f"Report preview {page_number}", fill=(38, 73, 137), font=font)
    draw.text((96, 260), (title or "Untitled report")[:84], fill=(32, 37, 46), font=font)
    draw.text((96, 320), "Preview generated for document upload.", fill=(92, 103, 119), font=font)
    for idx in range(6):
        top = 420 + idx * 92
        draw.rounded_rectangle((96, top, 804, top + 42), radius=12, fill=(241, 245, 250))
    image.save(dest, format="PNG")
    return dest.relative_to(APP_DIR / "static").as_posix()


def _generate_pdf_report_previews(source_path: Path, *, base_name: str) -> list[str]:
    if fitz is None:
        return []
    previews: list[str] = []
    preview_dir = _static_subdir("report_previews")
    doc = fitz.open(str(source_path))
    try:
        for page_index in range(min(REPORT_PREVIEW_PAGE_COUNT, len(doc))):
            page = doc.load_page(page_index)
            pix = page.get_pixmap(matrix=fitz.Matrix(1.4, 1.4), alpha=False)
            dest = preview_dir / f"{base_name}_page_{page_index + 1}.png"
            pix.save(str(dest))
            previews.append(dest.relative_to(APP_DIR / "static").as_posix())
    finally:
        doc.close()
    return previews


def _generate_report_preview_paths(*, report_title: str, report_rel_path: str, report_id: int) -> list[str]:
    static_path = APP_DIR / "static" / str(report_rel_path or "")
    ext = static_path.suffix.lower()
    base_name = f"report_{report_id}_{uuid.uuid4().hex[:8]}"
    previews: list[str] = []
    if ext == ".pdf" and static_path.is_file():
        try:
            previews = _generate_pdf_report_previews(static_path, base_name=base_name)
        except Exception:
            previews = []
    if previews:
        return previews[:REPORT_PREVIEW_PAGE_COUNT]
    return [
        _make_report_preview_placeholder(report_title, page_number=index + 1, base_name=base_name)
        for index in range(REPORT_PREVIEW_PAGE_COUNT)
    ]


def _report_payload(row: Report | None) -> dict[str, object] | None:
    if row is None:
        return None
    preview_paths = _json_list(getattr(row, "preview_paths", None))
    company_name = getattr(getattr(row, "company", None), "company_name", None) or ""
    file_path = str(getattr(row, "file_path", "") or "")
    ext = Path(file_path).suffix.lower().lstrip(".")
    file_type_label = "PDF" if ext == "pdf" else ("Word" if ext in {"doc", "docx"} else (ext.upper() if ext else "File"))
    file_type_icon = "PDF" if ext == "pdf" else ("DOC" if ext in {"doc", "docx"} else "FILE")
    uploader = getattr(row, "uploader", None)
    return {
        "id": int(row.id),
        "title": str(getattr(row, "title", "") or "").strip() or "Untitled report",
        "detail_url": url_for("report_detail", report_id=int(row.id)),
        "file_url": url_for("open_report", report_id=int(row.id)),
        "preview_urls": [_feed_media_url(path) for path in preview_paths if _feed_media_url(path)],
        "created_at_label": _format_feed_timestamp(getattr(row, "created_at", None)),
        "created_at_display": getattr(row, "created_at", None).strftime("%d %b %Y") if getattr(row, "created_at", None) else "",
        "file_type_label": file_type_label,
        "file_type_icon": file_type_icon,
        "company_name": company_name,
        "category_name": str(getattr(getattr(row, "category", None), "name", "") or "").strip(),
        "author_name": _user_display_name(uploader),
        "author_profile_url": url_for("public_profile", user_id=int(uploader.id)) if getattr(uploader, "id", None) else "",
    }


def _newsletter_payload(row: Newsletter | None) -> dict[str, object] | None:
    if row is None:
        return None
    file_path = str(getattr(row, "file_path", "") or "")
    ext = Path(file_path).suffix.lower().lstrip(".")
    uploader = getattr(row, "uploader", None)
    return {
        "id": int(row.id),
        "title": str(getattr(row, "title", "") or "").strip() or "Untitled newsletter",
        "detail_url": url_for("newsletter_detail", newsletter_id=int(row.id)),
        "file_url": url_for("open_newsletter", newsletter_id=int(row.id)),
        "created_at_label": _format_feed_timestamp(getattr(row, "created_at", None)),
        "created_at_display": getattr(row, "created_at", None).strftime("%d %b %Y") if getattr(row, "created_at", None) else "",
        "file_type_label": "PDF" if ext == "pdf" else "Word",
        "file_type_icon": "PDF" if ext == "pdf" else "DOC",
        "author_name": _user_display_name(uploader),
        "author_profile_url": url_for("public_profile", user_id=int(uploader.id)) if getattr(uploader, "id", None) else "",
    }


def _event_payload(row: Event | None) -> dict[str, object] | None:
    if row is None:
        return None
    creator = getattr(row, "creator", None)
    event_date = getattr(row, "event_date", None)
    return {
        "id": int(row.id),
        "title": str(getattr(row, "title", "") or "").strip() or "Untitled event",
        "description": str(getattr(row, "description", "") or "").strip(),
        "detail_url": url_for("event_detail", event_id=int(row.id)),
        "event_date_label": event_date.strftime("%d %b %Y %H:%M") if isinstance(event_date, datetime) else "",
        "event_date_iso": event_date.isoformat() if isinstance(event_date, datetime) else "",
        "created_at_label": _format_feed_timestamp(getattr(row, "created_at", None)),
        "created_at_display": getattr(row, "created_at", None).strftime("%d %b %Y") if getattr(row, "created_at", None) else "",
        "author_name": _user_display_name(creator),
        "author_profile_url": url_for("public_profile", user_id=int(creator.id)) if getattr(creator, "id", None) else "",
    }


def _award_form_payload(row: AwardsForm | None) -> dict[str, object] | None:
    if row is None:
        return None
    creator = getattr(row, "creator", None)
    submission_count = AwardsSubmission.query.filter_by(form_id=int(row.id)).count()
    question_count = AwardsQuestion.query.filter_by(form_id=int(row.id)).count()
    return {
        "id": int(row.id),
        "title": str(getattr(row, "title", "") or "").strip() or "Untitled award",
        "description": str(getattr(row, "description", "") or "").strip(),
        "header_image_url": url_for("static", filename=str(getattr(row, "header_image", "") or "")) if getattr(row, "header_image", None) else "",
        "detail_url": url_for("awards_form_page", form_id=int(row.id)),
        "admin_url": url_for("awards_admin", form_id=int(row.id)),
        "submissions_url": url_for("awards_submissions_page", form_id=int(row.id)),
        "export_url": url_for("awards_export_csv", form_id=int(row.id)),
        "created_at_label": _format_feed_timestamp(getattr(row, "created_at", None)),
        "created_at_display": getattr(row, "created_at", None).strftime("%d %b %Y") if getattr(row, "created_at", None) else "",
        "author_name": _user_display_name(creator),
        "author_profile_url": url_for("public_profile", user_id=int(creator.id)) if getattr(creator, "id", None) else "",
        "submission_count": int(submission_count or 0),
        "question_count": int(question_count or 0),
    }


def _report_category_payload(row: ReportCategory | None) -> dict[str, object] | None:
    if row is None:
        return None
    return {
        "id": int(row.id),
        "name": str(getattr(row, "name", "") or "").strip() or "Category",
        "created_at_display": getattr(row, "created_at", None).strftime("%d %b %Y") if getattr(row, "created_at", None) else "",
    }


def _normalize_awards_question_type(raw: object) -> str:
    value = str(raw or "").strip().lower()
    return value if value in AWARDS_QUESTION_TYPES_SET else "text"


def _parse_awards_question_options(raw: object) -> list[str]:
    if raw is None:
        return []
    if isinstance(raw, list):
        values = raw
    else:
        try:
            decoded = json.loads(str(raw))
            values = decoded if isinstance(decoded, list) else []
        except Exception:
            values = [line for line in str(raw or "").splitlines()]
    out: list[str] = []
    for item in values:
        value = str(item or "").strip()
        if value and value not in out:
            out.append(value)
    return out


def _awards_question_payload(row: AwardsQuestion | None) -> dict[str, object] | None:
    if row is None:
        return None
    options = _parse_awards_question_options(getattr(row, "options", None))
    question_type = _normalize_awards_question_type(getattr(row, "question_type", None))
    return {
        "id": int(row.id),
        "question_text": str(getattr(row, "question_text", "") or "").strip(),
        "question_type": question_type,
        "required": bool(getattr(row, "required", False)),
        "options": options,
        "options_text": "\n".join(options),
        "type_label": {
            "text": "Single line text",
            "textarea": "Multi-line text",
            "single_choice": "Single choice",
            "file": "File upload",
        }.get(question_type, "Question"),
    }


def _parse_awards_builder_questions(raw: object) -> tuple[list[dict[str, object]], str | None]:
    try:
        decoded = json.loads(str(raw or "[]"))
    except Exception:
        return [], "Questions payload is invalid."
    if not isinstance(decoded, list):
        return [], "Questions payload is invalid."
    questions: list[dict[str, object]] = []
    for item in decoded:
        if not isinstance(item, dict):
            continue
        question_text = str(item.get("question_text") or "").strip()
        question_type = _normalize_awards_question_type(item.get("question_type"))
        required = bool(item.get("required"))
        options = _parse_awards_question_options(item.get("options"))
        if not question_text:
            continue
        if question_type == "single_choice" and not options:
            return [], "Single choice questions require at least one option."
        questions.append(
            {
                "question_text": question_text,
                "question_type": question_type,
                "required": required,
                "options": options,
            }
        )
    if not questions:
        return [], "Add at least one question."
    return questions, None


def _save_awards_answer_file(storage, *, user_id: int, form_id: int, question_id: int) -> tuple[str | None, str | None]:
    if not storage or not getattr(storage, "filename", None):
        return (None, "No file provided.")
    filename = secure_filename(storage.filename or "")
    ext = Path(filename).suffix.lower()
    allowed = MODULE_DOCUMENT_ALLOWED_EXT | PROFILE_PHOTO_ALLOWED_EXT
    if ext not in allowed:
        return (None, "Unsupported file type.")
    dest_dir = _static_subdir("uploads", "awards")
    stored_name = f"award_{form_id}_{question_id}_{user_id}_{datetime.utcnow().strftime('%Y%m%d%H%M%S')}_{uuid.uuid4().hex[:8]}{ext}"
    dest = dest_dir / stored_name
    storage.save(str(dest))
    return (dest.relative_to(APP_DIR / "static").as_posix(), None)


def _audit_2025_safe_cell(value: object) -> object:
    if value is None:
        return None
    try:
        if pd.isna(value):
            return None
    except Exception:
        pass
    if isinstance(value, pd.Timestamp):
        return value.isoformat()
    if hasattr(value, "item"):
        try:
            return value.item()
        except Exception:
            return value
    return value


def _audit_2025_year_from_value(raw_value: object, parsed_value: object) -> int | None:
    if isinstance(parsed_value, pd.Timestamp):
        return int(parsed_value.year)
    match = re.search(r"(20\d{2})", str(raw_value or ""))
    if match:
        try:
            return int(match.group(1))
        except Exception:
            return None
    return None


def load_audit_2025_data() -> dict[str, object]:
    path = AUDIT_2025_WORKBOOK_PATH
    if not path.exists():
        return {"ok": False, "error": "Audit 2025 workbook is not available on this server."}
    try:
        stat = path.stat()
        cache_key = (int(stat.st_mtime_ns), int(stat.st_size))
        if _AUDIT_2025_CACHE.get("cache_key") == cache_key and isinstance(_AUDIT_2025_CACHE.get("payload"), dict):
            return dict(_AUDIT_2025_CACHE["payload"])

        frame = pd.read_excel(path)
        frame.columns = [str(column or "").strip() for column in frame.columns]
        missing_columns = [column for column in AUDIT_2025_COLUMNS if column not in frame.columns]
        if missing_columns:
            return {"ok": False, "error": "Audit 2025 workbook is missing required columns."}

        frame = frame.loc[:, list(AUDIT_2025_COLUMNS)].copy()
        month_parsed = pd.to_datetime(frame["Month"], errors="coerce")
        month_display: list[str] = []
        month_sort: list[str] = []
        year_values: list[int | None] = []
        for raw_value, parsed_value in zip(frame["Month"].tolist(), month_parsed.tolist()):
            if isinstance(parsed_value, pd.Timestamp):
                month_display.append(parsed_value.strftime("%b %Y"))
                month_sort.append(parsed_value.strftime("%Y-%m"))
            else:
                raw_label = str(raw_value or "").strip()
                month_display.append(raw_label)
                month_sort.append(raw_label)
            year_values.append(_audit_2025_year_from_value(raw_value, parsed_value))

        frame["_audit_month_label"] = month_display
        frame["_audit_month_sort"] = month_sort
        frame["_audit_year"] = year_values

        numeric_frame = frame.copy()
        for column in AUDIT_2025_NUMERIC_COLUMNS:
            numeric_frame[column] = pd.to_numeric(numeric_frame[column], errors="coerce")

        records: list[dict[str, object]] = []
        for index in range(len(frame.index)):
            row: dict[str, object] = {}
            for column in AUDIT_2025_COLUMNS:
                if column in AUDIT_2025_NUMERIC_COLUMNS:
                    row[column] = _audit_2025_safe_cell(numeric_frame.iloc[index][column])
                else:
                    row[column] = _audit_2025_safe_cell(frame.iloc[index][column])
            row["_audit_month_label"] = _audit_2025_safe_cell(frame.iloc[index]["_audit_month_label"])
            row["_audit_month_sort"] = _audit_2025_safe_cell(frame.iloc[index]["_audit_month_sort"])
            row["_audit_year"] = _audit_2025_safe_cell(frame.iloc[index]["_audit_year"])
            records.append(row)

        month_order_pairs = sorted(
            {
                (str(record.get("_audit_month_sort") or ""), str(record.get("_audit_month_label") or ""))
                for record in records
                if str(record.get("_audit_month_label") or "").strip()
            },
            key=lambda item: item[0],
        )
        payload = {
            "ok": True,
            "records": records,
            "columns": list(AUDIT_2025_COLUMNS),
            "category_columns": list(AUDIT_2025_CATEGORY_COLUMNS),
            "years": sorted({int(value) for value in year_values if isinstance(value, int)}),
            "months": [label for _sort_key, label in month_order_pairs],
            "companies": sorted(
                {
                    str(value).strip()
                    for value in frame["Company"].tolist()
                    if str(value or "").strip()
                },
                key=lambda item: item.casefold(),
            ),
        }
        _AUDIT_2025_CACHE["cache_key"] = cache_key
        _AUDIT_2025_CACHE["payload"] = payload
        return dict(payload)
    except Exception:
        app.logger.exception("Audit 2025 workbook could not be loaded.")
        return {"ok": False, "error": "Audit 2025 workbook could not be loaded."}


def _save_awards_header_image(storage, *, user_id: int, form_id: int | None = None) -> tuple[str | None, str | None]:
    if not storage or not getattr(storage, "filename", None):
        return (None, None)
    filename = secure_filename(storage.filename or "")
    ext = Path(filename).suffix.lower()
    if ext not in AWARDS_HEADER_IMAGE_ALLOWED_EXT:
        return (None, "Unsupported image type. Please upload a JPG or PNG image.")
    dest_dir = _static_subdir("uploads", "awards")
    form_part = f"form_{int(form_id)}_" if form_id else "form_new_"
    stored_name = f"{form_part}{user_id}_{datetime.utcnow().strftime('%Y%m%d%H%M%S')}_{uuid.uuid4().hex[:8]}{ext}"
    dest = dest_dir / stored_name
    storage.save(str(dest))
    return (dest.relative_to(APP_DIR / "static").as_posix(), None)


def _awards_answer_display_payload(question: AwardsQuestion | None, answer_text: object) -> dict[str, object]:
    payload = _awards_question_payload(question) or {}
    value = str(answer_text or "").strip()
    is_file = str(payload.get("question_type") or "") == "file"
    file_url = url_for("static", filename=value) if is_file and value else ""
    return {
        "question_text": str(payload.get("question_text") or "Question"),
        "answer_text": value,
        "is_file": is_file,
        "file_url": file_url,
    }


def _awards_submission_payload(row: AwardsSubmission | None) -> dict[str, object] | None:
    if row is None:
        return None
    answer_rows = (
        AwardsAnswer.query.filter_by(submission_id=int(row.id))
        .join(AwardsQuestion, AwardsQuestion.id == AwardsAnswer.question_id)
        .order_by(AwardsQuestion.id.asc(), AwardsAnswer.id.asc())
        .all()
    )
    submitter = getattr(row, "submitter", None)
    return {
        "id": int(row.id),
        "user": _user_display_name(submitter),
        "created_at_label": _format_feed_timestamp(getattr(row, "created_at", None)),
        "created_at_display": getattr(row, "created_at", None).strftime("%d %b %Y %H:%M") if getattr(row, "created_at", None) else "",
        "answers": [
            _awards_answer_display_payload(getattr(answer, "question", None), getattr(answer, "answer_text", None))
            for answer in answer_rows
        ],
    }


def _awards_single_choice_analytics(form_id: int) -> list[dict[str, object]]:
    question_rows = (
        AwardsQuestion.query.filter_by(form_id=int(form_id), question_type="single_choice")
        .order_by(AwardsQuestion.id.asc())
        .all()
    )
    analytics: list[dict[str, object]] = []
    for question in question_rows:
        options = _parse_awards_question_options(getattr(question, "options", None))
        counts = {option: 0 for option in options}
        answer_rows = AwardsAnswer.query.filter_by(question_id=int(question.id)).all()
        for answer in answer_rows:
            value = str(getattr(answer, "answer_text", "") or "").strip()
            if value in counts:
                counts[value] += 1
        analytics.append(
            {
                "question_text": str(getattr(question, "question_text", "") or "").strip() or "Question",
                "options": [{"label": option, "count": int(counts.get(option, 0))} for option in options],
            }
        )
    return analytics


def _challenge_payload(row: Challenge | None) -> dict[str, object] | None:
    if row is None:
        return None
    response_count = ChallengeResponse.query.filter_by(challenge_id=int(row.id)).count()
    deadline = getattr(row, "deadline", None)
    deadline_label = deadline.strftime("%d %b %Y %H:%M") if isinstance(deadline, datetime) else ""
    return {
        "id": int(row.id),
        "title": str(getattr(row, "title", "") or "").strip() or "Challenge",
        "description": str(getattr(row, "description", "") or "").strip(),
        "deadline_label": deadline_label,
        "deadline_iso": deadline.isoformat() if isinstance(deadline, datetime) else "",
        "response_count": int(response_count or 0),
        "is_open": bool(deadline is None or deadline >= datetime.utcnow()),
    }


def _challenge_response_payload(row: ChallengeResponse | None) -> dict[str, object] | None:
    if row is None:
        return None
    challenge = getattr(row, "challenge", None)
    return {
        "id": int(row.id),
        "answer": str(getattr(row, "answer", "") or "").strip(),
        "challenge_id": int(getattr(row, "challenge_id", 0) or 0),
        "challenge_title": str(getattr(challenge, "title", "") or "").strip(),
    }


def _normalize_comment_mention_ids(raw_ids: object, content: str) -> list[int]:
    content_value = str(content or "")
    parsed_ids: list[int] = []
    if isinstance(raw_ids, str):
        try:
            raw_ids = json.loads(raw_ids)
        except Exception:
            raw_ids = [part for part in raw_ids.split(",") if str(part).strip()]
    if isinstance(raw_ids, (list, tuple, set)):
        for item in raw_ids:
            try:
                value = int(item)
            except Exception:
                continue
            if value > 0 and value not in parsed_ids:
                parsed_ids.append(value)
    if not parsed_ids:
        return []
    users = User.query.filter(User.id.in_(parsed_ids)).all()
    valid_ids: list[int] = []
    for user in users:
        mention_token = f"@{_user_display_name(user)}".strip()
        if mention_token and mention_token in content_value and int(user.id) not in valid_ids:
            valid_ids.append(int(user.id))
    return valid_ids


def _render_text_with_mentions(content: str, *, mention_users: list[User] | None = None) -> Markup:
    raw_content = str(content or "")
    users = [user for user in (mention_users or []) if user is not None]
    if not users:
        return Markup("<br>".join(escape(part) for part in raw_content.splitlines())) if raw_content else Markup("")
    tokens = []
    token_map: dict[str, User] = {}
    for user in users:
        token = f"@{_user_display_name(user)}".strip()
        if not token:
            continue
        tokens.append(token)
        token_map[token] = user
    if not tokens:
        return Markup("<br>".join(escape(part) for part in raw_content.splitlines())) if raw_content else Markup("")
    tokens.sort(key=len, reverse=True)
    pattern = re.compile("|".join(re.escape(token) for token in tokens))
    cursor = 0
    pieces: list[str] = []
    for match in pattern.finditer(raw_content):
        if match.start() > cursor:
            pieces.append(str(escape(raw_content[cursor:match.start()])))
        token = match.group(0)
        pieces.append(
            '<span class="feed-mention" data-mentioned-user-id="{user_id}">{label}</span>'.format(
                user_id=int(getattr(token_map.get(token), "id", 0) or 0),
                label=escape(token),
            )
        )
        cursor = match.end()
    if cursor < len(raw_content):
        pieces.append(str(escape(raw_content[cursor:])))
    return Markup("".join(pieces).replace("\n", "<br>"))


def _comment_payload(
    row: Comment,
    *,
    like_count: int = 0,
    liked_by_viewer: bool = False,
) -> dict[str, object]:
    author = getattr(row, "user", None)
    mention_ids = _json_int_list(getattr(row, "mentioned_user_ids_json", None))
    mention_users = User.query.filter(User.id.in_(mention_ids)).all() if mention_ids else []
    return {
        "id": int(row.id),
        "post_id": int(getattr(row, "post_id", 0) or 0),
        "user_id": int(getattr(row, "user_id", 0) or 0),
        "author_name": _user_display_name(author) or "User",
        "author_avatar_url": _user_avatar_url(author),
        "author_profile_url": url_for("public_profile", user_id=int(author.id)) if getattr(author, "id", None) else url_for("feed"),
        "author_role_label": _user_role_label(author),
        "content": str(getattr(row, "content", "") or "").strip(),
        "content_html": _render_text_with_mentions(str(getattr(row, "content", "") or ""), mention_users=mention_users),
        "created_at_label": _format_feed_timestamp(getattr(row, "created_at", None)),
        "created_at_iso": getattr(row, "created_at", None).isoformat() if getattr(row, "created_at", None) else "",
        "like_count": int(like_count or 0),
        "liked_by_viewer": bool(liked_by_viewer),
        "like_endpoint": url_for("api_feed_comment_like", comment_id=int(row.id)),
        "mentioned_user_ids": mention_ids,
    }


def _comment_payload_maps(post_ids: list[int], viewer_user_id: int) -> tuple[dict[int, list[dict[str, object]]], dict[int, int]]:
    if not post_ids:
        return {}, {}
    rows = (
        Comment.query.filter(Comment.post_id.in_(post_ids))
        .order_by(Comment.created_at.asc(), Comment.id.asc())
        .all()
    )
    if not rows:
        return {post_id: [] for post_id in post_ids}, {post_id: 0 for post_id in post_ids}
    comment_ids = [int(row.id) for row in rows]
    like_counts: dict[int, int] = {
        int(comment_id): int(total or 0)
        for comment_id, total in (
            db.session.query(CommentLike.comment_id, db.func.count(CommentLike.id))
            .filter(CommentLike.comment_id.in_(comment_ids))
            .group_by(CommentLike.comment_id)
            .all()
        )
    }
    liked_ids = {
        int(row.comment_id)
        for row in CommentLike.query.filter(
            CommentLike.comment_id.in_(comment_ids),
            CommentLike.user_id == int(viewer_user_id or 0),
        ).all()
    } if viewer_user_id else set()
    payload_map: dict[int, list[dict[str, object]]] = {post_id: [] for post_id in post_ids}
    count_map: dict[int, int] = {post_id: 0 for post_id in post_ids}
    for row in rows:
        post_id = int(getattr(row, "post_id", 0) or 0)
        payload = _comment_payload(
            row,
            like_count=like_counts.get(int(row.id), 0),
            liked_by_viewer=int(row.id) in liked_ids,
        )
        payload_map.setdefault(post_id, []).append(payload)
        count_map[post_id] = count_map.get(post_id, 0) + 1
    return payload_map, count_map


def _feed_post_payload(
    row: FeedPost,
    *,
    reaction_summary: list[dict[str, object]] | None = None,
    current_reaction: str | None = None,
    comments: list[dict[str, object]] | None = None,
) -> dict[str, object]:
    author = getattr(row, "author", None)
    normalized_reference_type = _normalize_feed_reference_type(getattr(row, "reference_type", None))
    reference_id = int(getattr(row, "reference_id", 0) or 0)
    report_payload = _report_payload(Report.query.get(reference_id)) if normalized_reference_type == "report" and reference_id else None
    newsletter_payload = _newsletter_payload(Newsletter.query.get(reference_id)) if normalized_reference_type == "newsletter" and reference_id else None
    event_payload = _event_payload(Event.query.get(reference_id)) if normalized_reference_type == "event" and reference_id else None
    award_payload = _award_form_payload(AwardsForm.query.get(reference_id)) if normalized_reference_type == "award" and reference_id else None
    challenge_payload = _challenge_payload(Challenge.query.get(reference_id)) if normalized_reference_type == "challenge" and reference_id else None
    response_payload = _challenge_response_payload(ChallengeResponse.query.get(reference_id)) if normalized_reference_type == "challenge_response" and reference_id else None
    company_name = (getattr(author, "company_name", None) or "").strip() or "CTS Carbon Platform"
    author_title = _user_professional_title(author)
    author_name = _user_display_name(author) or "CTS User"
    summary = list(reaction_summary or [])
    reaction_state = _feed_reaction_button_state(current_reaction)
    comment_rows = list(comments or [])
    return {
        "id": int(row.id),
        "content": str(getattr(row, "content", "") or "").strip(),
        "post_type": _normalize_feed_post_type(getattr(row, "post_type", None)),
        "post_type_label": _normalize_feed_post_type(getattr(row, "post_type", None)).title(),
        "media_type": (getattr(row, "media_type", None) or "").strip() or None,
        "media_url": _feed_media_url(getattr(row, "media_path", None)),
        "media_name": Path(str(getattr(row, "media_path", "") or "")).name if getattr(row, "media_path", None) else None,
        "created_at_label": _format_feed_timestamp(getattr(row, "created_at", None)),
        "created_at_iso": getattr(row, "created_at", None).isoformat() if getattr(row, "created_at", None) else "",
        "author_name": author_name,
        "author_role_label": _user_role_label(author),
        "author_title": author_title,
        "author_company": company_name,
        "author_avatar_url": _user_avatar_url(author),
        "author_profile_url": url_for("public_profile", user_id=int(author.id)) if getattr(author, "id", None) else url_for("feed"),
        "company_logo_url": _company_logo_url(company_name),
        "share_url": url_for("feed") + f"#post-{int(row.id)}",
        "reference_type": normalized_reference_type or None,
        "reference_id": reference_id or None,
        "report": report_payload,
        "newsletter": newsletter_payload,
        "event": event_payload,
        "award": award_payload,
        "challenge": challenge_payload,
        "challenge_response": response_payload,
        "can_respond_to_challenge": bool(challenge_payload and not _is_readonly_user(current_user) and challenge_payload.get("is_open")),
        "comments": comment_rows,
        "comment_count": len(comment_rows),
        "reaction_summary": summary,
        "reaction_total": sum(int(item.get("count") or 0) for item in summary),
        "current_reaction": str(reaction_state.get("type") or ""),
        "current_reaction_label": str(reaction_state.get("label") or "Like"),
        "current_reaction_icon": str(reaction_state.get("icon") or "👍"),
    }


def _ensure_user_profile_columns() -> None:
    try:
        from sqlalchemy import inspect, text

        inspector = inspect(db.engine)
        if not inspector.has_table("user"):
            return
        existing = {col["name"] for col in inspector.get_columns("user")}
        alters: list[str] = []
        backfill_profile_complete = False
        if "first_name" not in existing:
            alters.append("ALTER TABLE user ADD COLUMN first_name VARCHAR(100)")
        if "last_name" not in existing:
            alters.append("ALTER TABLE user ADD COLUMN last_name VARCHAR(100)")
        if "job_title" not in existing:
            alters.append("ALTER TABLE user ADD COLUMN job_title VARCHAR(200)")
        if "phone" not in existing:
            alters.append("ALTER TABLE user ADD COLUMN phone VARCHAR(40)")
        if "company_country" not in existing:
            alters.append("ALTER TABLE user ADD COLUMN company_country VARCHAR(100)")
        if "profile_photo_path" not in existing:
            alters.append("ALTER TABLE user ADD COLUMN profile_photo_path VARCHAR(500)")
        if "cover_image" not in existing:
            alters.append("ALTER TABLE user ADD COLUMN cover_image VARCHAR(500)")
        if "is_profile_complete" not in existing:
            alters.append("ALTER TABLE user ADD COLUMN is_profile_complete BOOLEAN")
            backfill_profile_complete = True
        if "template_mode" not in existing:
            alters.append("ALTER TABLE user ADD COLUMN template_mode VARCHAR(20)")
        if "business_type" not in existing:
            alters.append("ALTER TABLE user ADD COLUMN business_type VARCHAR(100)")
        if "product_type" not in existing:
            alters.append("ALTER TABLE user ADD COLUMN product_type VARCHAR(100)")
        if "quantity" not in existing:
            alters.append("ALTER TABLE user ADD COLUMN quantity VARCHAR(100)")
        if "quantity_unit" not in existing:
            alters.append("ALTER TABLE user ADD COLUMN quantity_unit VARCHAR(100)")
        if "number_of_products_in_use" not in existing:
            alters.append("ALTER TABLE user ADD COLUMN number_of_products_in_use VARCHAR(100)")
        if "end_use_location" not in existing:
            alters.append("ALTER TABLE user ADD COLUMN end_use_location VARCHAR(200)")
        if "heating_source" not in existing:
            alters.append("ALTER TABLE user ADD COLUMN heating_source VARCHAR(100)")
        if "travel_provider" not in existing:
            alters.append("ALTER TABLE user ADD COLUMN travel_provider VARCHAR(20)")
        if "operating_locations_json" not in existing:
            alters.append("ALTER TABLE user ADD COLUMN operating_locations_json TEXT")
        role_added = False
        if "role" not in existing:
            alters.append("ALTER TABLE user ADD COLUMN role VARCHAR(32)")
            role_added = True
        if not alters:
            return
        with db.engine.begin() as conn:
            for stmt in alters:
                conn.execute(text(stmt))
            if backfill_profile_complete:
                conn.execute(text("UPDATE user SET is_profile_complete = 1 WHERE is_profile_complete IS NULL"))
            conn.execute(text(f"UPDATE user SET template_mode = '{TEMPLATE_MODE_LEGACY}' WHERE template_mode IS NULL"))
            if role_added:
                conn.execute(text("UPDATE user SET role = 'admin' WHERE is_admin = 1"))
                conn.execute(text("UPDATE user SET role = 'user' WHERE role IS NULL"))
    except Exception:
        pass


def _ensure_user_last_seen_column() -> None:
    try:
        from sqlalchemy import inspect, text

        inspector = inspect(db.engine)
        if not inspector.has_table("user"):
            return
        existing = {col["name"] for col in inspector.get_columns("user")}
        if "last_seen_at" in existing:
            return
        with db.engine.begin() as conn:
            conn.execute(text("ALTER TABLE user ADD COLUMN last_seen_at DATETIME"))
    except Exception:
        pass


def _ensure_feed_post_reference_columns() -> None:
    try:
        from sqlalchemy import inspect, text

        inspector = inspect(db.engine)
        if not inspector.has_table("feed_post"):
            return
        existing = {col["name"] for col in inspector.get_columns("feed_post")}
        alters: list[str] = []
        if "reference_id" not in existing:
            alters.append("ALTER TABLE feed_post ADD COLUMN reference_id INTEGER")
        if "reference_type" not in existing:
            alters.append("ALTER TABLE feed_post ADD COLUMN reference_type VARCHAR(40)")
        if not alters:
            return
        with db.engine.begin() as conn:
            for stmt in alters:
                conn.execute(text(stmt))
    except Exception:
        pass


def _ensure_notification_meta_json_column() -> None:
    try:
        from sqlalchemy import inspect, text

        inspector = inspect(db.engine)
        if not inspector.has_table("notifications"):
            return
        existing = {col["name"] for col in inspector.get_columns("notifications")}
        if "meta_json" in existing:
            return
        with db.engine.begin() as conn:
            conn.execute(text("ALTER TABLE notifications ADD COLUMN meta_json TEXT"))
    except Exception:
        pass


def _ensure_report_category_columns() -> None:
    try:
        from sqlalchemy import inspect, text

        inspector = inspect(db.engine)
        if not inspector.has_table("report"):
            return
        existing = {col["name"] for col in inspector.get_columns("report")}
        if "category_id" in existing:
            return
        with db.engine.begin() as conn:
            conn.execute(text("ALTER TABLE report ADD COLUMN category_id INTEGER"))
    except Exception:
        pass


def _ensure_awards_form_columns() -> None:
    try:
        from sqlalchemy import inspect, text

        inspector = inspect(db.engine)
        if not inspector.has_table("awards_forms"):
            return
        existing = {col["name"] for col in inspector.get_columns("awards_forms")}
        if "header_image" in existing:
            return
        with db.engine.begin() as conn:
            conn.execute(text("ALTER TABLE awards_forms ADD COLUMN header_image VARCHAR(500)"))
    except Exception:
        pass


def _ensure_mapping_unmapped_row_columns() -> None:
    try:
        from sqlalchemy import inspect, text

        inspector = inspect(db.engine)
        if not inspector.has_table("mapping_unmapped_row"):
            return
        existing = {col["name"] for col in inspector.get_columns("mapping_unmapped_row")}
        alters: list[str] = []
        if "source_entry_group" not in existing:
            alters.append("ALTER TABLE mapping_unmapped_row ADD COLUMN source_entry_group VARCHAR(40)")
        if "row_label" not in existing:
            alters.append("ALTER TABLE mapping_unmapped_row ADD COLUMN row_label VARCHAR(500)")
        if "status_value" not in existing:
            alters.append("ALTER TABLE mapping_unmapped_row ADD COLUMN status_value VARCHAR(120) DEFAULT 'No match'")
        if "review_status" not in existing:
            alters.append("ALTER TABLE mapping_unmapped_row ADD COLUMN review_status VARCHAR(32) DEFAULT 'open'")
        if "assigned_ef_id" not in existing:
            alters.append("ALTER TABLE mapping_unmapped_row ADD COLUMN assigned_ef_id VARCHAR(120)")
        if "owner_notes" not in existing:
            alters.append("ALTER TABLE mapping_unmapped_row ADD COLUMN owner_notes TEXT")
        if "resolved_at" not in existing:
            alters.append("ALTER TABLE mapping_unmapped_row ADD COLUMN resolved_at DATETIME")
        if "resolved_by_user_id" not in existing:
            alters.append("ALTER TABLE mapping_unmapped_row ADD COLUMN resolved_by_user_id INTEGER")
        if not alters:
            return
        with db.engine.begin() as conn:
            for stmt in alters:
                conn.execute(text(stmt))
    except Exception:
        pass


def _ensure_mapping_run_source_entry_group_column() -> None:
    try:
        from sqlalchemy import inspect, text

        inspector = inspect(db.engine)
        if not inspector.has_table("mapping_run"):
            return
        existing_columns = {col["name"] for col in inspector.get_columns("mapping_run")}
        if "source_entry_group" in existing_columns:
            return
        with db.engine.begin() as conn:
            conn.execute(text("ALTER TABLE mapping_run ADD COLUMN source_entry_group VARCHAR(40)"))
    except Exception:
        pass


def _resolve_template_company_name(company_name: str) -> str | None:
    raw = (company_name or "").strip()
    if not raw:
        return None
    canon, _country = _canonical_company_name_and_country(raw)
    return (canon or raw).strip() or None


def _resolve_template_sheet_name(company_name: str, sheet_name: str) -> str | None:
    if str(sheet_name or "").strip().lower() == KLARAKARBON_SHEET_NAME.lower():
        return KLARAKARBON_SHEET_NAME
    resolved = TEMPLATE_REGISTRY.resolve_sheet_name(company_name, sheet_name, template_mode=_current_template_mode())
    return resolved or None


def _resolve_template_sheet_name_with_mode(company_name: str, sheet_name: str, *, template_mode: str) -> str | None:
    """Like _resolve_template_sheet_name but uses an explicit template_mode (no Flask request/session)."""
    if str(sheet_name or "").strip().lower() == KLARAKARBON_SHEET_NAME.lower():
        return KLARAKARBON_SHEET_NAME
    mode = normalize_template_mode(template_mode)
    resolved = TEMPLATE_REGISTRY.resolve_sheet_name(company_name, sheet_name, template_mode=mode)
    return resolved or None


def _get_template_sheet_headers_with_mode(company_name: str, sheet_name: str, *, template_mode: str) -> list[str]:
    """Resolve sheet headers without Flask request/session (background jobs). Empty profile: Cat 1 is unaffected."""
    if str(sheet_name or "").strip().lower() == KLARAKARBON_SHEET_NAME.lower():
        resolved_company = _resolve_template_company_name(company_name)
        if not resolved_company:
            return []
        return list(klarakarbon_entry_headers(resolved_company))
    resolved_sheet = _resolve_template_sheet_name_with_mode(company_name, sheet_name, template_mode=template_mode)
    if not resolved_sheet:
        return []
    bundle = TEMPLATE_REGISTRY.get_bundle(
        template_mode=normalize_template_mode(template_mode),
        company_name=company_name,
        profile={},
    )
    for item in bundle.get("visible_templates") or []:
        if _normalize_template_key(str(item.get("sheet_name") or "")) == _normalize_template_key(resolved_sheet):
            return list(item.get("headers") or [])
    return []


def _list_template_companies_for_user() -> list[dict[str, str]]:
    if bool(getattr(current_user, "is_admin", False)):
        return [{"key": name, "label": name} for name in sorted(_COMPANY_COUNTRY_CANONICAL.keys())]

    resolved = _resolve_template_company_name(getattr(current_user, "company_name", "") or "")
    if not resolved:
        return []
    return [{"key": resolved, "label": resolved}]


def _get_template_company_sheets(company_name: str) -> list[str]:
    resolved_company = _resolve_template_company_name(company_name) or (company_name or "").strip()
    bundle = _template_bundle_for_company(resolved_company)
    return [str(item.get("sheet_name") or "") for item in (bundle.get("visible_templates") or []) if str(item.get("sheet_name") or "").strip()]


def _get_template_sheet_headers(company_name: str, sheet_name: str) -> list[str]:
    if str(sheet_name or "").strip().lower() == KLARAKARBON_SHEET_NAME.lower():
        resolved_company = _resolve_template_company_name(company_name)
        if not resolved_company:
            return []
        return list(klarakarbon_entry_headers(resolved_company))
    resolved_sheet = _resolve_template_sheet_name(company_name, sheet_name)
    if not resolved_sheet:
        return []
    for item in _template_bundle_for_company(company_name).get("visible_templates") or []:
        if _normalize_template_key(str(item.get("sheet_name") or "")) == _normalize_template_key(resolved_sheet):
            return list(item.get("headers") or [])
    return []


def _build_template_workbook(company_name: str) -> BytesIO:
    resolved_company = _resolve_template_company_name(company_name)
    wb = Workbook()
    try:
        wb.remove(wb.active)
    except Exception:
        pass

    for sheet_name in _get_template_company_sheets(resolved_company or company_name):
        ws = wb.create_sheet(title=str(sheet_name)[:31])
        for idx, header in enumerate(_get_template_sheet_headers(resolved_company or company_name, sheet_name), start=1):
            ws.cell(row=1, column=idx).value = header

    bio = BytesIO()
    wb.save(bio)
    bio.seek(0)
    return bio


def _user_can_access_company(company_name: str) -> bool:
    if bool(getattr(current_user, "is_admin", False)):
        return True
    resolved_requested = _resolve_template_company_name(company_name)
    resolved_owned = _resolve_template_company_name(getattr(current_user, "company_name", "") or "")
    if not resolved_requested or not resolved_owned:
        return False
    return _normalize_template_key(resolved_requested) == _normalize_template_key(resolved_owned)


def _get_data_entry_template_schema(company_name: str, sheet_name: str) -> tuple[list[str], dict[str, dict[str, str]]]:
    headers = _get_template_sheet_headers(company_name, sheet_name)
    if not headers:
        return [], {}
    rules = {h: _infer_column_rule(h) for h in headers}
    return headers, rules


def _clean_data_entry_rows(headers: list[str], rows: list[object]) -> tuple[list[list[str]], list[str]]:
    cleaned_rows: list[list[str]] = []
    for r in rows:
        if not isinstance(r, list):
            continue
        rr = [("" if v is None else str(v)) for v in r]
        if not any(v.strip() for v in rr):
            continue
        if len(rr) < len(headers):
            rr = rr + [""] * (len(headers) - len(rr))
        elif len(rr) > len(headers):
            rr = rr[: len(headers)]
        cleaned_rows.append(rr)
    return _validate_and_normalize_rows(headers, cleaned_rows)


def _parse_iso_datetime(value: object) -> datetime | None:
    text = str(value or "").strip()
    if not text:
        return None
    try:
        return datetime.fromisoformat(text.replace("Z", "+00:00"))
    except Exception:
        return None


def _is_data_entry_editable(created_at: datetime | None) -> bool:
    if bool(getattr(current_user, "is_admin", False)):
        return True
    if not isinstance(created_at, datetime):
        return True
    return created_at >= (datetime.utcnow() - timedelta(days=7))


def _normalize_data_entry_rows(headers: list[str], rows: list[object]) -> tuple[list[dict[str, object]], list[str]]:
    normalized_rows: list[dict[str, object]] = []
    raw_values: list[list[str]] = []
    for row_idx, r in enumerate(rows, start=1):
        if isinstance(r, dict):
            cells = r.get("cells") or r.get("values") or []
            entry_group = str(r.get("entry_group") or "").strip()
            created_at = _parse_iso_datetime(r.get("created_at"))
            is_persisted = bool(r.get("is_persisted"))
            source_row_index = int(r.get("row_index") or row_idx)
        elif isinstance(r, list):
            cells = r
            entry_group = ""
            created_at = None
            is_persisted = False
            source_row_index = row_idx
        else:
            continue

        if not isinstance(cells, list):
            continue
        rr = [("" if v is None else str(v)) for v in cells]
        if not any(v.strip() for v in rr):
            continue
        if len(rr) < len(headers):
            rr = rr + [""] * (len(headers) - len(rr))
        elif len(rr) > len(headers):
            rr = rr[: len(headers)]
        normalized_rows.append(
            {
                "entry_group": entry_group,
                "created_at": created_at,
                "is_persisted": is_persisted,
                "row_index": source_row_index,
                "cells": rr,
            }
        )
        raw_values.append(rr)

    cleaned_values, validation_errors = _validate_and_normalize_rows(headers, raw_values)
    for meta, cleaned in zip(normalized_rows, cleaned_values):
        meta["cells"] = cleaned
    return normalized_rows, validation_errors


def _existing_ccc_import_dedup_keys(company_name: str, sheet_name: str, dedup_column: str) -> set[str]:
    rows = db.session.query(DataEntry.value).filter(
        DataEntry.company_name == company_name,
        DataEntry.sheet_name == sheet_name,
        DataEntry.column_name == dedup_column,
    )
    return {str(v or "").strip().lower() for (v,) in rows.all() if str(v or "").strip()}


def _data_entry_content_fingerprint(headers: list[str], cells: list[str]) -> str:
    """Stable fingerprint for duplicate detection (full row vector, trimmed)."""
    padded = list(cells[: len(headers)])
    while len(padded) < len(headers):
        padded.append("")
    normalized = [str(c or "").strip() for c in padded]
    return json.dumps(normalized, ensure_ascii=True, separators=(",", ":"))


def _validate_data_entry_row_requirements(headers: list[str], rows: list[dict[str, object]]) -> list[str]:
    """Required-field validation after normalization (types already checked)."""
    errors: list[str] = []
    rules = [_infer_column_rule(h) for h in headers]
    for ridx, meta in enumerate(rows, start=1):
        cells = list(meta.get("cells") or [])
        while len(cells) < len(headers):
            cells.append("")
        if not any(str(c or "").strip() for c in cells):
            continue
        for cidx, h in enumerate(headers):
            rule = rules[cidx] if cidx < len(rules) else {"type": "text"}
            val = str(cells[cidx]).strip() if cidx < len(cells) else ""
            if rule.get("required") == "1" and not val:
                errors.append(f"Row {ridx}, column '{h}': required field is empty")
    return errors


def _norm_evidence_header_key(x: str) -> str:
    return "".join(ch.lower() for ch in (x or "") if ch.isalnum())


def _header_matches_evidence_column(header: str) -> bool:
    priority_exact = (
        "Source_File",
        "Source file",
        "source_file",
        "Source File",
        "Data Source",
        "Data source",
        "Datasource",
        "Source",
    )
    if (header or "").strip() in priority_exact:
        return True
    n = _norm_evidence_header_key(header)
    return n in ("sourcefile", "sourcefilename", "datasources")


def _sync_evidence_orphan_metadata(evidence_file_id: int, *, commit: bool = True) -> None:
    """Update cached link count and orphan flags (never deletes files)."""
    ef = db.session.get(EvidenceFile, evidence_file_id)
    if ef is None or ef.is_deleted:
        return
    n = int(
        db.session.query(func.count(DataEntryEvidence.id)).filter_by(evidence_file_id=evidence_file_id).scalar()
        or 0
    )
    prev_orphan = bool(getattr(ef, "is_orphaned", False))
    ef.relation_count = n
    if n == 0:
        ef.is_orphaned = True
        if ef.orphaned_at is None:
            ef.orphaned_at = datetime.utcnow()
    else:
        ef.is_orphaned = False
        ef.orphaned_at = None
    db.session.add(ef)
    if commit:
        db.session.commit()
    if prev_orphan != bool(ef.is_orphaned):
        _evidence_log(
            "orphan_state_changed",
            evidence_file_id=evidence_file_id,
            is_orphaned=bool(ef.is_orphaned),
            relation_count=n,
        )


def _evidence_ui_status(ef: EvidenceFile) -> str:
    st = str(getattr(ef, "processing_status", "") or "").lower()
    if st in ("pending", "processing"):
        return "processing"
    if st == "failed":
        return "failed"
    return "ready"


def _delete_evidence_links_for_grid_row(company_name: str, sheet_name: str, grid_row_key: str | None) -> None:
    key = str(grid_row_key or "").strip()
    if not key:
        return
    rows = (
        DataEntryEvidence.query.filter_by(
            company_name=company_name,
            sheet_name=sheet_name,
            entry_group=key,
        ).all()
    )
    if not rows:
        return
    fids = {int(r.evidence_file_id) for r in rows}
    for r in rows:
        db.session.delete(r)
    db.session.flush()
    for fid in fids:
        _sync_evidence_orphan_metadata(fid, commit=False)
    db.session.commit()


def _evidence_local_read_path(rel: str | None):
    from frontend.storage import get_evidence_storage

    return get_evidence_storage().try_get_local_path(rel)


def _safe_evidence_disk_path(rel: str | None) -> Path | None:
    """Legacy name: resolve evidence relative key via storage provider."""
    return _evidence_local_read_path(rel)


def _data_entry_grid_row_exists(company_name: str, sheet_name: str, headers: list[str], grid_row_key: str) -> bool:
    want = str(grid_row_key or "").strip()
    if not want:
        return False
    for row in _load_data_entry_grid_rows(company_name, sheet_name, headers):
        if str(row.get("entry_group") or "").strip() == want:
            return True
    return False


def _evidence_row_attachment_counts(company_name: str, sheet_name: str) -> dict[str, int]:
    rows = (
        db.session.query(DataEntryEvidence.entry_group, func.count(DataEntryEvidence.id))
        .filter(DataEntryEvidence.company_name == company_name, DataEntryEvidence.sheet_name == sheet_name)
        .group_by(DataEntryEvidence.entry_group)
        .all()
    )
    return {str(eg or ""): int(n or 0) for eg, n in rows}


_SHARED_INVOICE_MIN_ROWS = 3


def _evidence_row_shared_invoice_flags(company_name: str, sheet_name: str) -> dict[str, bool]:
    """True when any linked file on that grid row is reused across enough rows (relation_count)."""
    dee = DataEntryEvidence
    ef = EvidenceFile
    agg = (
        db.session.query(dee.entry_group, func.max(func.coalesce(ef.relation_count, 0)).label("mx"))
        .join(ef, ef.id == dee.evidence_file_id)
        .filter(
            dee.company_name == company_name,
            dee.sheet_name == sheet_name,
            ef.is_deleted.is_(False),
        )
        .group_by(dee.entry_group)
        .all()
    )
    return {str(eg or ""): int(mx or 0) >= _SHARED_INVOICE_MIN_ROWS for eg, mx in agg}


def _serialize_evidence_public(
    ef: EvidenceFile,
    *,
    link_id: int | None = None,
    link_meta: dict[str, object] | None = None,
) -> dict[str, object]:
    pid = int(ef.id)
    base: dict[str, object] = {
        "id": pid,
        "original_filename": ef.original_filename,
        "mime_type": ef.mime_type,
        "file_extension": ef.file_extension,
        "processing_status": ef.processing_status,
        "ui_status": _evidence_ui_status(ef),
        "uploaded_at": ef.uploaded_at.isoformat() if ef.uploaded_at else "",
        "file_size_original": int(ef.file_size_original or 0),
        "file_size_optimized": int(ef.file_size_optimized or 0) if ef.file_size_optimized is not None else None,
        "is_orphaned": bool(getattr(ef, "is_orphaned", False)),
        "relation_count": int(getattr(ef, "relation_count", 0) or 0),
        "thumbnail_url": url_for("api_evidence_thumbnail", evidence_id=pid) if ef.thumbnail_storage_path else None,
        "preview_url": url_for("api_evidence_preview", evidence_id=pid),
        "download_url": url_for("api_evidence_download", evidence_id=pid),
    }
    if link_id is not None:
        base["link_id"] = link_id
    if link_meta:
        base.update(link_meta)
    return base


def _user_can_access_evidence_file(ef: EvidenceFile | None) -> bool:
    if ef is None or ef.is_deleted:
        return False
    return bool(_user_can_access_company(ef.company_name))


def _run_evidence_processing_job(
    *,
    job_id: str,
    evidence_file_id: int,
    staging_rel: str,
    upload_company: str,
) -> dict[str, object]:
    from frontend.evidence_processing import process_evidence_file_with_storage
    from frontend.storage import get_evidence_storage

    storage = get_evidence_storage()
    _update_job_progress(job_id, 10, "Optimizing evidence file")
    ef = db.session.get(EvidenceFile, evidence_file_id)
    if ef is None:
        raise RuntimeError("Evidence record missing")

    staging_path = storage.try_get_local_path(staging_rel)
    if staging_path is None or not staging_path.is_file():
        ef.processing_status = "failed"
        ef.processing_error = "Staging file missing"
        db.session.commit()
        _evidence_log("optimization_failed", evidence_file_id=evidence_file_id, reason="staging_missing")
        raise RuntimeError("Staging file missing")

    ef.processing_status = "processing"
    ef.processing_error = None
    db.session.commit()

    try:
        base_ts = ef.uploaded_at or datetime.utcnow()
        result = process_evidence_file_with_storage(
            storage=storage,
            staging_path=staging_path,
            sha256_hex=ef.sha256_hash,
            normalized_ext=str(ef.file_extension or "").lower(),
            base_ts=base_ts,
        )
        _raise_if_job_cancelled(job_id)
        ef.stored_filename = str(result["stored_filename"])
        ef.storage_path = str(result["storage_path"]).replace("\\", "/")
        ef.thumbnail_storage_path = (
            str(result["thumbnail_storage_path"]).replace("\\", "/") if result.get("thumbnail_storage_path") else None
        )
        ef.file_extension = str(result["file_extension"])
        ef.mime_type = str(result["mime_type"])
        ef.file_size_original = int(result["original_size"])
        ef.file_size_optimized = int(result["optimized_size"])
        ef.processing_status = "ready"
        ef.processing_error = None
        db.session.commit()
        storage.delete_file(staging_rel)
        try:
            staging_path.unlink(missing_ok=True)
        except Exception:
            pass
        _update_job_progress(job_id, 100, "Evidence ready")
        _evidence_log(
            "optimization_completed",
            evidence_file_id=evidence_file_id,
            optimized_bytes=int(result["optimized_size"]),
        )
        return {
            "ok": True,
            "evidence_file_id": evidence_file_id,
            "company": upload_company,
            "processing_status": "ready",
        }
    except JobCancelled:
        raise
    except Exception as exc:
        _evidence_log(
            "optimization_failed",
            evidence_file_id=evidence_file_id,
            error=str(exc),
        )
        try:
            ef2 = db.session.get(EvidenceFile, evidence_file_id)
            if ef2 is not None:
                ef2.processing_status = "failed"
                ef2.processing_error = str(exc)
                db.session.commit()
        except Exception:
            db.session.rollback()
        storage.delete_file(staging_rel)
        try:
            staging_path.unlink(missing_ok=True)
        except Exception:
            pass
        raise


def _delete_data_entry_group(
    company_name: str,
    sheet_name: str,
    entry_group: str,
    created_at: datetime | None,
    row_index: int,
) -> None:
    _delete_evidence_links_for_grid_row(company_name, sheet_name, entry_group)

    if entry_group and not entry_group.startswith("legacy:"):
        DataEntry.query.filter_by(
            company_name=company_name,
            sheet_name=sheet_name,
            entry_group=entry_group,
        ).delete(synchronize_session=False)
        return

    query = DataEntry.query.filter_by(
        company_name=company_name,
        sheet_name=sheet_name,
        row_index=row_index,
    )
    if created_at is not None:
        query = query.filter(DataEntry.created_at == created_at)
    query.delete(synchronize_session=False)


_DATA_ENTRY_UPLOAD_USER_UNSPECIFIED = object()


def _upsert_data_entries(
    company_name: str,
    sheet_name: str,
    headers: list[str],
    rows: list[dict[str, object]],
    *,
    uploaded_by_user_id: object = _DATA_ENTRY_UPLOAD_USER_UNSPECIFIED,
) -> dict[str, object]:
    """
    Save rows with duplicate detection (company + sheet + full row content).
    Returns counts and entry_group ids that were newly written.
    """
    grid_snapshot = _load_data_entry_grid_rows_no_request(company_name, sheet_name, headers)
    entry_group_to_fp: dict[str, str] = {}
    all_fps: set[str] = set()
    for r in grid_snapshot:
        eg = str(r.get("entry_group") or "").strip()
        if not eg:
            continue
        fp = _data_entry_content_fingerprint(headers, list(r.get("cells") or []))
        entry_group_to_fp[eg] = fp
        all_fps.add(fp)

    saved_rows_count = 0
    duplicate_rows_count = 0
    saved_entry_groups: list[str] = []
    batch_fps: set[str] = set()

    next_row_index = (
        db.session.query(db.func.max(DataEntry.row_index))
        .filter_by(company_name=company_name, sheet_name=sheet_name)
        .scalar()
        or 0
    )

    for row in rows:
        cells = list(row.get("cells") or [])
        if not any(str(v or "").strip() for v in cells):
            continue

        fp = _data_entry_content_fingerprint(headers, cells)
        entry_group = str(row.get("entry_group") or "").strip()
        created_at = row.get("created_at")
        if not isinstance(created_at, datetime):
            created_at = None
        source_row_index = int(row.get("row_index") or 0)
        is_persisted = bool(row.get("is_persisted"))

        check = set(all_fps) | set(batch_fps)
        old_fp = entry_group_to_fp.get(entry_group) if entry_group else None
        if is_persisted and old_fp is not None:
            check.discard(old_fp)
            if fp == old_fp:
                duplicate_rows_count += 1
                continue

        if fp in check:
            duplicate_rows_count += 1
            continue

        if is_persisted:
            if not _is_data_entry_editable(created_at):
                continue
            if old_fp is not None:
                all_fps.discard(old_fp)
            entry_group_to_fp.pop(entry_group, None)
            _delete_data_entry_group(company_name, sheet_name, entry_group, created_at, source_row_index)
            effective_created_at = created_at or datetime.utcnow()
            effective_row_index = source_row_index or (next_row_index + 1)
            effective_entry_group = uuid.uuid4().hex[:12]
        else:
            next_row_index += 1
            effective_created_at = datetime.utcnow()
            effective_row_index = next_row_index
            effective_entry_group = uuid.uuid4().hex[:12]

        if uploaded_by_user_id is _DATA_ENTRY_UPLOAD_USER_UNSPECIFIED:
            uploader_id = getattr(current_user, "id", None)
        else:
            uploader_id = uploaded_by_user_id

        for column_name, value in zip(headers, cells):
            vv = (value or "").strip()
            if vv == "":
                continue
            db.session.add(
                DataEntry(
                    company_name=company_name,
                    sheet_name=sheet_name,
                    entry_group=effective_entry_group,
                    uploaded_by_user_id=uploader_id,
                    row_index=effective_row_index,
                    column_name=column_name,
                    value=vv,
                    created_at=effective_created_at,
                )
            )

        all_fps.add(fp)
        batch_fps.add(fp)
        entry_group_to_fp[effective_entry_group] = fp
        saved_rows_count += 1
        saved_entry_groups.append(effective_entry_group)

    return {
        "saved_rows_count": saved_rows_count,
        "duplicate_rows_count": duplicate_rows_count,
        "saved_entry_groups": saved_entry_groups,
    }


def _load_data_entry_grid_rows(company_name: str, sheet_name: str, headers: list[str]) -> list[dict[str, object]]:
    entries = (
        DataEntry.query.filter_by(company_name=company_name, sheet_name=sheet_name)
        .order_by(DataEntry.created_at.asc(), DataEntry.row_index.asc(), DataEntry.id.asc())
        .all()
    )
    grouped: dict[str, dict[str, object]] = {}
    order: list[str] = []
    for entry in entries:
        created_at = getattr(entry, "created_at", None)
        row_index = int(getattr(entry, "row_index", 0) or 0)
        entry_group = str(getattr(entry, "entry_group", "") or "").strip()
        group_key = entry_group or f"legacy:{created_at.isoformat() if created_at else ''}:{row_index}"
        if group_key not in grouped:
            grouped[group_key] = {
                "entry_group": group_key,
                "row_index": row_index,
                "created_at": created_at,
                "values": {},
            }
            order.append(group_key)
        values = grouped[group_key]["values"]
        if isinstance(values, dict):
            values[str(entry.column_name)] = "" if entry.value is None else str(entry.value)

    rows: list[dict[str, object]] = []
    for key in order:
        row = grouped[key]
        values = row.get("values") if isinstance(row.get("values"), dict) else {}
        created_at = row.get("created_at") if isinstance(row.get("created_at"), datetime) else None
        rows.append(
            {
                "entry_group": row.get("entry_group") or "",
                "row_index": int(row.get("row_index") or 0),
                "created_at": created_at.isoformat() if created_at else "",
                "is_editable": _is_data_entry_editable(created_at),
                "is_persisted": True,
                "cells": [values.get(header, "") for header in headers],
            }
        )
    return rows


def _load_data_entry_grid_rows_no_request(company_name: str, sheet_name: str, headers: list[str]) -> list[dict[str, object]]:
    entries = (
        DataEntry.query.filter_by(company_name=company_name, sheet_name=sheet_name)
        .order_by(DataEntry.created_at.asc(), DataEntry.row_index.asc(), DataEntry.id.asc())
        .all()
    )
    grouped: dict[str, dict[str, object]] = {}
    order: list[str] = []
    for entry in entries:
        created_at = getattr(entry, "created_at", None)
        row_index = int(getattr(entry, "row_index", 0) or 0)
        entry_group = str(getattr(entry, "entry_group", "") or "").strip()
        group_key = entry_group or f"legacy:{created_at.isoformat() if created_at else ''}:{row_index}"
        if group_key not in grouped:
            grouped[group_key] = {
                "entry_group": group_key,
                "row_index": row_index,
                "created_at": created_at,
                "values": {},
            }
            order.append(group_key)
        values = grouped[group_key]["values"]
        if isinstance(values, dict):
            values[str(entry.column_name)] = "" if entry.value is None else str(entry.value)

    rows: list[dict[str, object]] = []
    for key in order:
        row = grouped[key]
        values = row.get("values") if isinstance(row.get("values"), dict) else {}
        created_at = row.get("created_at") if isinstance(row.get("created_at"), datetime) else None
        rows.append(
            {
                "entry_group": row.get("entry_group") or "",
                "row_index": int(row.get("row_index") or 0),
                "created_at": created_at.isoformat() if created_at else "",
                "is_editable": True,
                "is_persisted": True,
                "cells": [values.get(header, "") for header in headers],
            }
        )
    return rows


def _load_data_entries_dataframe(company_name: str, sheet_name: str, headers: list[str]) -> "pd.DataFrame":
    rows = _load_data_entry_grid_rows(company_name, sheet_name, headers)
    values = [list(row.get("cells") or []) for row in rows]
    return pd.DataFrame(values, columns=headers)


def _load_data_entries_dataframe_no_request(
    company_name: str, sheet_name: str, headers: list[str], entry_groups: set[str] | None = None
) -> "pd.DataFrame":
    rows = _load_data_entry_grid_rows_no_request(company_name, sheet_name, headers)
    if entry_groups:
        rows = [r for r in rows if str(r.get("entry_group") or "").strip() in entry_groups]
    values = [list(row.get("cells") or []) for row in rows]
    if not values:
        return pd.DataFrame(columns=headers)
    return pd.DataFrame(values, columns=headers)


def _load_data_entries_dataframe_for_entry_groups(
    company_name: str, sheet_name: str, headers: list[str], entry_groups: set[str]
) -> "pd.DataFrame":
    """Build a dataframe containing only logical rows whose entry_group is in entry_groups."""
    if not entry_groups:
        return pd.DataFrame(columns=headers)
    rows = _load_data_entry_grid_rows(company_name, sheet_name, headers)
    filtered = [r for r in rows if str(r.get("entry_group") or "").strip() in entry_groups]
    values = [list(row.get("cells") or []) for row in filtered]
    if not values:
        return pd.DataFrame(columns=headers)
    return pd.DataFrame(values, columns=headers)


def _import_translation_module():
    import importlib
    return importlib.import_module("engine.stage1_preprocess.Datas.translate_me_the_chosen_one_30Sep")


def _user_can_run_translation(u: object | None) -> bool:
    return normalize_user_role(getattr(u, "role", None)) in {"owner", "super_admin", "admin"}


def _user_can_run_ccc_data_entry_import(u: object | None) -> bool:
    """CCC → Data Entry bulk import (internal procurement tool)."""
    return normalize_user_role(getattr(u, "role", None)) in {"owner", "super_admin", "admin"}


def _string_cell_value(value: object) -> str:
    try:
        if pd.isna(value):
            return ""
    except Exception:
        pass
    return "" if value is None else str(value)


_DATA_ENTRY_MAPPING_METADATA_COLUMNS = frozenset(
    {
        "ef_id",
        "status",
        "mapping_status",
        "mapped_by",
        "mapping_run_id",
    }
)


def _perf_log(page: str, **parts: float | int) -> None:
    """Lightweight timing for production diagnosis (structured single line)."""
    try:
        extra = " ".join(f"{k}={v}" for k, v in sorted(parts.items()) if v is not None)
        print(f"[PERF] page={page} {extra}".strip())
    except Exception:
        pass


def _norm_mapping_status(value: object) -> str:
    return str(value or "").strip().lower()


def _row_is_fully_ef_mapped(ef_id: object, status_value: object) -> bool:
    """
    A grid row counts as EF-mapped only when a non-empty ef_id is paired with a status
    that is not a definitive non-match. Otherwise stale ef_id cells could incorrectly
    close unmapped snapshots (see [UNMAPPED_SYNC] logs).
    """
    ef = str(ef_id or "").strip()
    if not ef:
        return False
    st = _norm_mapping_status(status_value)
    if not st:
        return True
    if st == "no match" or st.startswith("no match") or st == "unmapped":
        return False
    if "no match" in st and "partial" not in st:
        return False
    if "partial" in st:
        return False
    if "fail" in st or st == "error":
        return False
    return True


def _data_entry_non_metadata_cell_columns(cells: dict[str, str]) -> bool:
    """True if the grid point has at least one populated non-metadata column."""
    for col, raw in cells.items():
        ck = str(col or "").strip().lower()
        if ck in _DATA_ENTRY_MAPPING_METADATA_COLUMNS:
            continue
        if ck == "ccc_import_dedup":
            continue
        if str(raw or "").strip():
            return True
    return False


def _rollup_ef_state_from_grid(grid: dict[tuple[str, int], dict[str, str]]) -> dict[str, int]:
    """Shared rollup used by single-sheet loads and bulk batch aggregation."""
    out = {
        "data_rows": 0,
        "fully_mapped": 0,
        "partial": 0,
        "unmapped": 0,
        "failed": 0,
        "pending": 0,
    }
    for _key, cells in grid.items():
        if not _data_entry_non_metadata_cell_columns(cells):
            continue
        out["data_rows"] += 1
        meta = {k.lower(): v for k, v in cells.items()}
        ef_id = meta.get("ef_id") or ""
        st_raw = meta.get("status") or meta.get("mapping_status") or ""
        st = _norm_mapping_status(st_raw)

        if _row_is_fully_ef_mapped(ef_id, st_raw):
            out["fully_mapped"] += 1
            continue
        if "partial" in st:
            out["partial"] += 1
            continue
        if "fail" in st or st == "error":
            out["failed"] += 1
            continue
        if st == "no match" or st == "unmapped" or ("no match" in st and "partial" not in st):
            out["unmapped"] += 1
            continue
        if not str(ef_id or "").strip() and not st:
            out["pending"] += 1
            continue
        out["unmapped"] += 1
    return out


def _rollup_ef_state_for_data_entry_sheet(company_name: str, sheet_name: str) -> dict[str, int]:
    """
    Aggregate live Data Entry mapping metadata per logical grid row.
    Used so admin batches / notifications do not treat 'MappingRun succeeded' as EF-mapped.
    """
    rows = (
        DataEntry.query.filter(
            DataEntry.company_name == company_name,
            DataEntry.sheet_name == sheet_name,
        )
        .order_by(DataEntry.row_index.asc(), DataEntry.id.asc())
        .all()
    )
    grid: dict[tuple[str, int], dict[str, str]] = {}
    for row in rows:
        key = (str(row.entry_group or "").strip(), int(row.row_index or 0))
        grid.setdefault(key, {})[str(row.column_name or "").strip()] = _string_cell_value(row.value).strip()
    return _rollup_ef_state_from_grid(grid)


def _data_entry_grid_from_entries(entries: list[DataEntry]) -> dict[tuple[str, int], dict[str, str]]:
    grid: dict[tuple[str, int], dict[str, str]] = {}
    for row in sorted(entries, key=lambda e: (int(e.row_index or 0), int(getattr(e, "id", 0) or 0))):
        key = (str(row.entry_group or "").strip(), int(row.row_index or 0))
        grid.setdefault(key, {})[str(row.column_name or "").strip()] = _string_cell_value(row.value).strip()
    return grid


def _mapping_state_for_map_batch(
    *,
    action_type: str,
    rollup: dict[str, int],
    uploaded_at: datetime | None,
    mapped_at: datetime | None,
    pipeline_ready: bool,
) -> tuple[str, str]:
    """
    Returns (mapping_state, mapping_status_label) for admin notifications.
    mapping_state is a stable machine token for the popup (see dashboard base.html).
    """
    if action_type != "map":
        if pipeline_ready:
            return "pipeline_ready", "Pipeline output refreshed"
        return "pending", "Awaiting pipeline / append run"

    try:
        up_ok = bool(uploaded_at and mapped_at and mapped_at >= uploaded_at)
    except Exception:
        up_ok = bool(mapped_at and uploaded_at)

    if not up_ok:
        return "pending", "Not mapped yet (no mapping run after this upload)"

    dr = int(rollup.get("data_rows") or 0)
    full = int(rollup.get("fully_mapped") or 0)
    partial = int(rollup.get("partial") or 0)
    unmapped = int(rollup.get("unmapped") or 0)
    failed = int(rollup.get("failed") or 0)
    pend = int(rollup.get("pending") or 0)

    if dr <= 0:
        return "pending", "No data rows detected for mapping summary"

    if full >= dr:
        return "fully_mapped", f"Fully mapped (EF) — {full}/{dr} rows"

    if full > 0 or partial > 0:
        return "partially_mapped", f"Partially mapped — {full} full, {partial} partial, {unmapped + pend} without EF ({dr} rows)"

    if failed > 0 and unmapped == 0 and pend == 0:
        return "failed", f"Mapping failed for rows — {failed}/{dr} rows"

    if unmapped > 0 or pend > 0:
        return "unmapped", f"Unmapped / no EF match — {unmapped + pend} rows need review ({dr} rows)"

    return "unmapped", f"Unmapped — review {dr} row(s)"


def _upsert_data_entry_cell(
    *,
    company_name: str,
    sheet_name: str,
    entry_group: str,
    row_index: int,
    column_name: str,
    value: object,
    uploaded_by_user_id: int | None = None,
) -> None:
    query = DataEntry.query.filter_by(
        company_name=company_name,
        sheet_name=sheet_name,
        row_index=int(row_index),
        column_name=str(column_name),
    )
    if entry_group and not entry_group.startswith("legacy:"):
        query = query.filter_by(entry_group=entry_group)
    entry = query.order_by(DataEntry.id.asc()).first()
    if entry is None:
        entry = DataEntry(
            company_name=company_name,
            sheet_name=sheet_name,
            entry_group="" if entry_group.startswith("legacy:") else entry_group,
            uploaded_by_user_id=uploaded_by_user_id,
            row_index=int(row_index),
            column_name=str(column_name),
            created_at=datetime.utcnow(),
        )
        db.session.add(entry)
    entry.value = _string_cell_value(value)


def _clear_data_entry_mapping_metadata(
    company_name: str,
    sheet_name: str,
    entry_groups: set[str] | None = None,
) -> int:
    query = DataEntry.query.filter(
        DataEntry.company_name == company_name,
        DataEntry.sheet_name == sheet_name,
        DataEntry.column_name.in_(tuple(_DATA_ENTRY_MAPPING_METADATA_COLUMNS)),
    )
    normalized_groups = {str(g or "").strip() for g in (entry_groups or set()) if str(g or "").strip()}
    if normalized_groups:
        query = query.filter(DataEntry.entry_group.in_(tuple(normalized_groups)))
    deleted = int(query.delete(synchronize_session=False) or 0)
    if deleted:
        print(f"[UNMAPPED] Cleared {deleted} mapping metadata cell(s) for {company_name} / {sheet_name}")
    return deleted


def _clear_data_entry_mapping_metadata_for_grid_keys(
    company_name: str,
    sheet_name: str,
    keys: set[tuple[str, int]],
) -> int:
    """
    Clears mapping metadata only for specific (entry_group, row_index) keys.
    Used after translation so unrelated rows keep snapshots until re-evaluated.
    """
    if not keys:
        return 0
    meta_cols = tuple(_DATA_ENTRY_MAPPING_METADATA_COLUMNS)
    deleted_total = 0
    keys_list = list(keys)
    chunk_size = 48
    for i in range(0, len(keys_list), chunk_size):
        chunk = keys_list[i : i + chunk_size]
        ors = []
        for eg, rix in chunk:
            ors.append(
                and_(
                    DataEntry.entry_group == str(eg or "").strip(),
                    DataEntry.row_index == int(rix),
                )
            )
        q = DataEntry.query.filter(
            DataEntry.company_name == company_name,
            DataEntry.sheet_name == sheet_name,
            DataEntry.column_name.in_(meta_cols),
            or_(*ors),
        )
        deleted_total += int(q.delete(synchronize_session=False) or 0)
    if deleted_total:
        print(
            f"[TRANSLATION_INVALIDATION] Cleared {deleted_total} mapping metadata cell(s) "
            f"for {len(keys)} grid key(s) on {company_name} / {sheet_name}"
        )
    return deleted_total


def _data_entry_row_metadata_lookup(company_name: str, sheet_name: str) -> dict[tuple[str, int], dict[str, str]]:
    rows = (
        DataEntry.query.filter(
            DataEntry.company_name == company_name,
            DataEntry.sheet_name == sheet_name,
            DataEntry.column_name.in_(tuple(_DATA_ENTRY_MAPPING_METADATA_COLUMNS)),
        )
        .order_by(DataEntry.created_at.asc(), DataEntry.row_index.asc(), DataEntry.id.asc())
        .all()
    )
    out: dict[tuple[str, int], dict[str, str]] = {}
    for row in rows:
        key = (str(row.entry_group or "").strip(), int(row.row_index or 0))
        out.setdefault(key, {})[str(row.column_name or "").strip().lower()] = _string_cell_value(row.value).strip()
    return out


def _persist_mapping_metadata_to_data_entry(
    *,
    company_name: str,
    sheet_name: str,
    source_entry_group: str | None,
    mapped_df: pd.DataFrame,
    run_id: str,
    user_id: int | None,
) -> None:
    if mapped_df is None or mapped_df.empty:
        return

    grid_rows = _load_data_entry_grid_rows_no_request(company_name, sheet_name, list(mapped_df.columns))
    entry_group_filter = str(source_entry_group or "").strip()
    if entry_group_filter:
        grid_rows = [r for r in grid_rows if str(r.get("entry_group") or "").strip() == entry_group_filter]
    if not grid_rows:
        return

    ef_col = _df_find_column_casefold(mapped_df, ("ef_id", "EF ID", "Emission Factor ID", "assigned_ef_id"))
    status_col = _df_find_column_casefold(mapped_df, ("status", "Status", "mapping_status", "Mapping Status"))
    mapped_by_col = _df_find_column_casefold(mapped_df, ("mapped_by", "Mapped By"))

    ef_mapped_rows = 0
    no_match_rows = 0
    for pos, row_meta in enumerate(grid_rows):
        if pos >= len(mapped_df.index):
            break
        mapped_row = mapped_df.iloc[pos]
        row_index = int(row_meta.get("row_index") or (pos + 1))
        entry_group = str(row_meta.get("entry_group") or "").strip()
        ef_value = _string_cell_value(mapped_row.get(ef_col) if ef_col else "").strip()
        status_value = _string_cell_value(mapped_row.get(status_col) if status_col else "").strip()
        mapped_by_value = _string_cell_value(mapped_row.get(mapped_by_col) if mapped_by_col else "").strip()

        if _row_is_fully_ef_mapped(ef_value, status_value):
            ef_mapped_rows += 1
        if _norm_mapping_status(status_value) == "no match":
            no_match_rows += 1

        _upsert_data_entry_cell(
            company_name=company_name,
            sheet_name=sheet_name,
            entry_group=entry_group,
            row_index=row_index,
            column_name="ef_id",
            value=ef_value,
            uploaded_by_user_id=user_id,
        )
        _upsert_data_entry_cell(
            company_name=company_name,
            sheet_name=sheet_name,
            entry_group=entry_group,
            row_index=row_index,
            column_name="status",
            value=status_value,
            uploaded_by_user_id=user_id,
        )
        _upsert_data_entry_cell(
            company_name=company_name,
            sheet_name=sheet_name,
            entry_group=entry_group,
            row_index=row_index,
            column_name="mapping_run_id",
            value=run_id,
            uploaded_by_user_id=user_id,
        )
        if mapped_by_value:
            _upsert_data_entry_cell(
                company_name=company_name,
                sheet_name=sheet_name,
                entry_group=entry_group,
                row_index=row_index,
                column_name="mapped_by",
                value=mapped_by_value,
                uploaded_by_user_id=user_id,
            )

    print(
        f"[MAPPING_STATE] run_id={run_id} company={company_name!r} sheet={sheet_name!r} "
        f"grid_targets={min(len(grid_rows), len(mapped_df.index))} ef_mapped_rows={ef_mapped_rows} "
        f"no_match_rows={no_match_rows}"
    )
    print(f"[UNMAPPED] Persisted live mapping metadata for {company_name} / {sheet_name} run {run_id}")


def _delete_stale_open_unmapped_rows_from_live_data_entry(
    company_name: str | None = None,
    sheet_name: str | None = None,
) -> int:
    query = MappingUnmappedRow.query.filter(MappingUnmappedRow.review_status == "open")
    if company_name:
        query = query.filter(MappingUnmappedRow.company_name == company_name)
    if sheet_name:
        query = query.filter(MappingUnmappedRow.sheet_name == sheet_name)

    # Single consistent path: only remove open unmapped when live metadata shows a real EF mapping.
    # Previously: any non-empty ef_id deleted snapshots even when status was still "No Match", which
    # broke trust in the unmapped inbox ([UNMAPPED_SYNC]).
    rows = query.all()
    if not rows:
        return 0

    print(
        f"[UNMAPPED_REFRESH] stale_cleanup scope company={company_name or '*'} sheet={sheet_name or '*'} "
        f"candidates={len(rows)}"
    )

    metadata_cache: dict[tuple[str, str], dict[tuple[str, int], dict[str, str]]] = {}
    to_delete: list[MappingUnmappedRow] = []
    for row in rows:
        cache_key = (str(row.company_name or ""), str(row.sheet_name or ""))
        if cache_key not in metadata_cache:
            metadata_cache[cache_key] = _data_entry_row_metadata_lookup(*cache_key)
        lookup = metadata_cache[cache_key]
        entry_group = str(row.source_entry_group or "").strip()
        row_index = max(1, int(row.row_number or 2) - 1)
        meta = lookup.get((entry_group, row_index)) or lookup.get(("", row_index)) or {}
        ef_id = str(meta.get("ef_id") or "").strip()
        status_raw = str(meta.get("status") or meta.get("mapping_status") or "").strip()
        if _row_is_fully_ef_mapped(ef_id, status_raw):
            to_delete.append(row)
            print(
                f"[UNMAPPED_SYNC] close_open id={row.id} company={row.company_name!r} sheet={row.sheet_name!r} "
                f"row_num={row.row_number} reason=live_ef_mapped ef_id={ef_id!r} status={status_raw!r}"
            )

    for row in to_delete:
        db.session.delete(row)
    if to_delete:
        print(f"[UNMAPPED] Deleted {len(to_delete)} stale open unmapped row(s) (live EF mapped)")
    return len(to_delete)


def _supersede_open_unmapped_rows(
    company_name: str,
    sheet_name: str,
    source_entry_group: str | None = None,
    *,
    reason: str = "data_changed",
    row_numbers: set[int] | None = None,
) -> int:
    query = MappingUnmappedRow.query.filter(
        MappingUnmappedRow.company_name == company_name,
        MappingUnmappedRow.sheet_name == sheet_name,
        MappingUnmappedRow.review_status == "open",
    )
    entry_group = str(source_entry_group or "").strip()
    if entry_group:
        query = query.filter(MappingUnmappedRow.source_entry_group == entry_group)
    if row_numbers is not None:
        nums = sorted({int(x) for x in row_numbers if x is not None})
        if not nums:
            return 0
        query = query.filter(MappingUnmappedRow.row_number.in_(nums))
    rows = query.all()
    for row in rows:
        row.review_status = "superseded"
        row.resolved_at = datetime.utcnow()
        row.owner_notes = (str(row.owner_notes or "").strip() + f" | Superseded: {reason}").strip(" |")
        print(
            f"[UNMAPPED_SYNC] supersede_open id={row.id} row_number={row.row_number} "
            f"entry_group={row.source_entry_group!r} reason={reason!r}"
        )
    if rows:
        print(
            f"[UNMAPPED] Superseded {len(rows)} open row(s) for {company_name} / {sheet_name}"
            f"{(' / ' + entry_group) if entry_group else ''}: {reason}"
        )
    return len(rows)


def _persist_translated_data_entry_columns(
    *,
    company_name: str,
    sheet_name: str,
    headers: list[str],
    original_df: pd.DataFrame,
    translated_df: pd.DataFrame,
    columns: tuple[str, ...],
    uploaded_by_user_id: int | None = None,
) -> tuple[int, int, set[tuple[str, int]]]:
    grid_rows = _load_data_entry_grid_rows_no_request(company_name, sheet_name, headers)
    changed_rows: set[int] = set()
    changed_cells = 0
    affected_keys: set[tuple[str, int]] = set()

    for row_pos, row in enumerate(grid_rows):
        if row_pos >= len(translated_df.index):
            break
        entry_group = str(row.get("entry_group") or "").strip()
        row_index = int(row.get("row_index") or (row_pos + 1))
        is_legacy_group = entry_group.startswith("legacy:")

        for col in columns:
            if col not in headers or col not in translated_df.columns:
                continue
            old_value = _string_cell_value(original_df.iloc[row_pos].get(col) if row_pos < len(original_df.index) else "")
            new_value = _string_cell_value(translated_df.iloc[row_pos].get(col))
            if new_value == old_value:
                continue

            query = DataEntry.query.filter_by(
                company_name=company_name,
                sheet_name=sheet_name,
                row_index=row_index,
                column_name=col,
            )
            if entry_group and not is_legacy_group:
                query = query.filter_by(entry_group=entry_group)
            entry = query.order_by(DataEntry.id.asc()).first()
            if entry is None:
                entry = DataEntry(
                    company_name=company_name,
                    sheet_name=sheet_name,
                    entry_group="" if is_legacy_group else entry_group,
                    uploaded_by_user_id=uploaded_by_user_id,
                    row_index=row_index,
                    column_name=col,
                    created_at=datetime.utcnow(),
                )
                db.session.add(entry)
            entry.value = new_value
            changed_rows.add(row_pos)
            changed_cells += 1
            storage_group = "" if is_legacy_group else entry_group
            affected_keys.add((storage_group, row_index))

    return len(changed_rows), changed_cells, affected_keys


def _batch_action_type_for_sheet(sheet_name: str) -> str:
    sheet_key = str(sheet_name or "").strip().lower()
    if sheet_key in {KLARAKARBON_SHEET_NAME.lower(), TRAVEL_SHEET_NAME.lower()}:
        return "append_run"
    return "map"


def _batch_action_label_for_sheet(sheet_name: str) -> str:
    return "Append & Run pipeline" if _batch_action_type_for_sheet(sheet_name) == "append_run" else "Map"


def _display_name_for_user(user: "User | None") -> str:
    if user is None:
        return "Unknown"
    first = str(getattr(user, "first_name", "") or "").strip()
    last = str(getattr(user, "last_name", "") or "").strip()
    full = " ".join(part for part in [first, last] if part).strip()
    if full:
        return full
    email = str(getattr(user, "email", "") or "").strip()
    if email and "@" in email:
        return email.split("@", 1)[0]
    return email or "Unknown"


def _list_admin_data_entry_batches() -> list[dict[str, object]]:
    """
    Uploaded data grouped to one row per (company, sheet).
    """
    from sqlalchemy import func

    _ensure_db_tables()

    co_company = func.trim(func.coalesce(DataEntry.company_name, ""))
    co_sheet = func.trim(func.coalesce(DataEntry.sheet_name, ""))

    rows = (
        db.session.query(
            co_company.label("company_name"),
            co_sheet.label("sheet_name"),
            func.max(DataEntry.created_at).label("uploaded_at"),
            func.count(func.distinct(DataEntry.row_index)).label("row_count"),
        )
        .group_by(co_company, co_sheet)
        .order_by(func.max(DataEntry.created_at).desc())
        .limit(500)
        .all()
    )

    latest_merged_workbook = _find_latest_merged_mapping_workbook()
    latest_merged_mtime = None
    try:
        latest_merged_mtime = latest_merged_workbook.stat().st_mtime if latest_merged_workbook else None
    except Exception:
        latest_merged_mtime = None

    seen_keys: set[tuple[str, str]] = set()
    work_items: list[tuple[object, str, str, str]] = []
    for r in rows:
        company = str(r.company_name or "").strip()
        sheet = str(r.sheet_name or "").strip()
        dedup_key = (company, sheet)
        if dedup_key in seen_keys:
            continue
        seen_keys.add(dedup_key)

        if _is_hidden_schema_sheet(sheet):
            continue

        action_type = _batch_action_type_for_sheet(sheet)
        work_items.append((r, company, sheet, action_type))

    batch_keys = list({(c, s) for _, c, s, _ in work_items})
    latest_mr: dict[tuple[str, str], MappingRun] = {}
    for mr in MappingRun.query.filter_by(status="succeeded").order_by(MappingRun.created_at.desc()).all():
        k = ((getattr(mr, "company_name", None) or "").strip(), (getattr(mr, "sheet_name", None) or "").strip())
        if k[0] and k[1] and k not in latest_mr:
            latest_mr[k] = mr

    entries_by_pair: dict[tuple[str, str], list[DataEntry]] = defaultdict(list)
    if batch_keys:
        bulk_entries = DataEntry.query.filter(
            tuple_(DataEntry.company_name, DataEntry.sheet_name).in_(batch_keys)
        ).all()
        for ent in bulk_entries:
            entries_by_pair[
                ((getattr(ent, "company_name", None) or "").strip(), (getattr(ent, "sheet_name", None) or "").strip())
            ].append(ent)

    rollups: dict[tuple[str, str], dict[str, int]] = {}
    for key in batch_keys:
        rollups[key] = _rollup_ef_state_from_grid(_data_entry_grid_from_entries(entries_by_pair.get(key, [])))

    mapper_user_ids = {int(mr.user_id) for mr in latest_mr.values() if getattr(mr, "user_id", None) is not None}
    users_by_id: dict[int, User] = {}
    if mapper_user_ids:
        for u in User.query.filter(User.id.in_(tuple(mapper_user_ids))).all():
            users_by_id[int(u.id)] = u

    def _latest_entry_for_pair(pair: tuple[str, str]) -> DataEntry | None:
        ents = entries_by_pair.get(pair) or []
        if not ents:
            return None
        return max(ents, key=lambda e: (getattr(e, "created_at", None) or datetime.min, int(getattr(e, "id", 0) or 0)))

    out: list[dict[str, object]] = []
    for r, company, sheet, action_type in work_items:
        pair = (company, sheet)
        mapped_run = latest_mr.get(pair)
        mapper_email = ""
        mapped_at_dt = getattr(mapped_run, "created_at", None) if mapped_run else None
        if mapped_run is not None:
            u = users_by_id.get(int(getattr(mapped_run, "user_id", 0) or 0))
            if u is not None:
                mapper_email = str(getattr(u, "email", "") or "")

        pipeline_ready = False
        if action_type != "map":
            if latest_merged_mtime is not None and getattr(r, "uploaded_at", None) is not None:
                try:
                    pipeline_ready = latest_merged_mtime >= r.uploaded_at.timestamp()
                except Exception:
                    pipeline_ready = False
            uploaded_at = r.uploaded_at if isinstance(r.uploaded_at, datetime) else None
            mapping_state, mapping_status_label = _mapping_state_for_map_batch(
                action_type="append_run",
                rollup=rollups.get(pair, {}),
                uploaded_at=uploaded_at,
                mapped_at=mapped_at_dt,
                pipeline_ready=pipeline_ready,
            )
            is_mapped = mapping_state == "pipeline_ready"
        else:
            uploaded_at = r.uploaded_at if isinstance(r.uploaded_at, datetime) else None
            mapping_state, mapping_status_label = _mapping_state_for_map_batch(
                action_type="map",
                rollup=rollups.get(pair, {}),
                uploaded_at=uploaded_at,
                mapped_at=mapped_at_dt if isinstance(mapped_at_dt, datetime) else None,
                pipeline_ready=False,
            )
            is_mapped = mapping_state == "fully_mapped"

        latest_entry = _latest_entry_for_pair(pair)
        uploaded_by_user_id = int(getattr(latest_entry, "uploaded_by_user_id", 0) or 0) if latest_entry else 0
        uploaded_by_name = ""
        uploaded_user = None
        if uploaded_by_user_id:
            uploaded_user = db.session.get(User, uploaded_by_user_id)
            uploaded_by_name = _display_name_for_user(uploaded_user)
        uid = hashlib.sha1(f"{company}\x00{sheet}".encode("utf-8")).hexdigest()[:16]
        out.append(
            {
                "batch_uid": uid,
                "company_name": company,
                "sheet_name": sheet,
                "entry_group": "",
                "uploaded_at": r.uploaded_at,
                "row_count": int(r.row_count or 0),
                "mappable": True,
                "action_type": action_type,
                "action_label": _batch_action_label_for_sheet(sheet),
                "mapped": is_mapped,
                "mapping_state": mapping_state,
                "mapping_status_label": mapping_status_label,
                "mapping_counts": dict(rollups.get(pair, {})),
                "mapped_at": mapped_at_dt,
                "mapped_by": mapper_email,
                "uploaded_by_user": uploaded_by_name,
                "uploaded_by_user_id": uploaded_by_user_id,
                "uploaded_by_job_title": _user_professional_title(uploaded_user),
                "uploaded_by_has_profile_photo": bool(
                    (getattr(uploaded_user, "profile_photo_path", None) or "").strip()
                )
                if uploaded_user
                else False,
                "download_run_id": str(getattr(mapped_run, "id", "") or "") if mapped_run else "",
            }
        )

    travel_path = STAGE2_TRAVEL_DIR / "analysis_summary.xlsx"
    if travel_path.exists():
        companies_for_travel = sorted({str(b.get("company_name") or "").strip() for b in out if str(b.get("company_name") or "").strip()})
        uploaded_at = None
        try:
            uploaded_at = datetime.utcfromtimestamp(travel_path.stat().st_mtime)
        except Exception:
            uploaded_at = None
        existing_travel = {(str(b.get("company_name") or "").strip(), str(b.get("sheet_name") or "").strip()) for b in out}
        for company in companies_for_travel:
            key = (company, TRAVEL_SHEET_NAME)
            if key in existing_travel:
                continue
            uid = hashlib.sha1(f"{company}\x00{TRAVEL_SHEET_NAME}".encode("utf-8")).hexdigest()[:16]
            is_mapped = False
            if latest_merged_mtime is not None and uploaded_at is not None:
                try:
                    is_mapped = latest_merged_mtime >= uploaded_at.timestamp()
                except Exception:
                    is_mapped = False
            mapping_state, mapping_status_label = _mapping_state_for_map_batch(
                action_type="append_run",
                rollup={},
                uploaded_at=uploaded_at,
                mapped_at=None,
                pipeline_ready=is_mapped,
            )
            out.append(
                {
                    "batch_uid": uid,
                    "company_name": company,
                    "sheet_name": TRAVEL_SHEET_NAME,
                    "entry_group": "",
                    "uploaded_at": uploaded_at,
                    "row_count": 0,
                    "mappable": True,
                    "action_type": "append_run",
                    "action_label": _batch_action_label_for_sheet(TRAVEL_SHEET_NAME),
                    "mapped": is_mapped,
                    "mapping_state": mapping_state,
                    "mapping_status_label": mapping_status_label,
                    "mapping_counts": {},
                    "mapped_at": None,
                    "mapped_by": "",
                    "uploaded_by_user": "System",
                    "uploaded_by_user_id": 0,
                    "uploaded_by_job_title": "",
                    "uploaded_by_has_profile_photo": False,
                    "download_run_id": "",
                }
            )
    return out


def _batches_for_admin_mapping_json(batches: list[dict[str, object]]) -> list[dict[str, object]]:
    """JSON-serializable batch list for the admin mapping page scripts."""
    out: list[dict[str, object]] = []
    for b in batches:
        ua = b.get("uploaded_at")
        ma = b.get("mapped_at")
        out.append(
            {
                "batch_uid": str(b.get("batch_uid") or ""),
                "company_name": str(b.get("company_name") or ""),
                "sheet_name": str(b.get("sheet_name") or ""),
                "entry_group": str(b.get("entry_group") or ""),
                "row_count": int(b.get("row_count") or 0),
                "mappable": bool(b.get("mappable")),
                "action_type": str(b.get("action_type") or ""),
                "action_label": str(b.get("action_label") or ""),
                "mapped": bool(b.get("mapped")),
                "mapping_state": str(b.get("mapping_state") or ""),
                "mapping_status_label": str(b.get("mapping_status_label") or ""),
                "mapping_counts": b.get("mapping_counts") if isinstance(b.get("mapping_counts"), dict) else {},
                "uploaded_at": ua.isoformat() + "Z" if isinstance(ua, datetime) else "",
                "mapped_at": ma.isoformat() + "Z" if isinstance(ma, datetime) else "",
                "mapped_by": str(b.get("mapped_by") or ""),
                "uploaded_by_user": str(b.get("uploaded_by_user") or ""),
                "download_run_id": str(b.get("download_run_id") or ""),
            }
        )
    return out


def _ensure_mapping_run_summary_columns() -> None:
    try:
        from sqlalchemy import inspect, text

        inspector = inspect(db.engine)
        if not inspector.has_table("mapping_run_summary"):
            return

        existing_columns = {col["name"] for col in inspector.get_columns("mapping_run_summary")}
        alter_statements = []
        if "mapped_categories_count" not in existing_columns:
            alter_statements.append(
                "ALTER TABLE mapping_run_summary ADD COLUMN mapped_categories_count INTEGER DEFAULT 0"
            )
        if "total_categories" not in existing_columns:
            alter_statements.append(
                "ALTER TABLE mapping_run_summary ADD COLUMN total_categories INTEGER DEFAULT 0"
            )
        if "coverage_pct" not in existing_columns:
            alter_statements.append(
                "ALTER TABLE mapping_run_summary ADD COLUMN coverage_pct FLOAT DEFAULT 0"
            )

        if not alter_statements:
            return

        with db.engine.begin() as conn:
            for stmt in alter_statements:
                conn.execute(text(stmt))
    except Exception:
        pass


def _ensure_data_entry_columns() -> None:
    try:
        from sqlalchemy import inspect, text

        inspector = inspect(db.engine)
        if not inspector.has_table("data_entry"):
            return

        existing_columns = {col["name"] for col in inspector.get_columns("data_entry")}
        alters: list[str] = []
        if "entry_group" not in existing_columns:
            alters.append("ALTER TABLE data_entry ADD COLUMN entry_group VARCHAR(32) DEFAULT ''")
        if "uploaded_by_user_id" not in existing_columns:
            alters.append("ALTER TABLE data_entry ADD COLUMN uploaded_by_user_id INTEGER")
        if not alters:
            return
        with db.engine.begin() as conn:
            for stmt in alters:
                conn.execute(text(stmt))
    except Exception:
        pass


def _backfill_evidence_orphan_metadata() -> None:
    """Recompute relation_count / orphan flags for all evidence rows."""
    try:
        ids = [int(r[0]) for r in db.session.query(EvidenceFile.id).filter(EvidenceFile.is_deleted.is_(False)).all()]
        for eid in ids:
            _sync_evidence_orphan_metadata(eid, commit=False)
        db.session.commit()
    except Exception:
        db.session.rollback()


def _ensure_evidence_tables_columns() -> None:
    """SQLite-friendly additive DDL for evidence tables."""
    try:
        from sqlalchemy import inspect, text

        inspector = inspect(db.engine)
        if not inspector.has_table("evidence_files"):
            return

        existing = {col["name"] for col in inspector.get_columns("evidence_files")}
        alters: list[str] = []
        needs_orphan_backfill = False

        if "thumbnail_storage_path" not in existing:
            alters.append("ALTER TABLE evidence_files ADD COLUMN thumbnail_storage_path VARCHAR(600)")
        if "processing_status" not in existing:
            alters.append("ALTER TABLE evidence_files ADD COLUMN processing_status VARCHAR(32) DEFAULT 'pending'")
        if "processing_error" not in existing:
            alters.append("ALTER TABLE evidence_files ADD COLUMN processing_error TEXT")
        if "is_orphaned" not in existing:
            alters.append("ALTER TABLE evidence_files ADD COLUMN is_orphaned BOOLEAN DEFAULT 0")
            needs_orphan_backfill = True
        if "orphaned_at" not in existing:
            alters.append("ALTER TABLE evidence_files ADD COLUMN orphaned_at DATETIME")
            needs_orphan_backfill = True
        if "relation_count" not in existing:
            alters.append("ALTER TABLE evidence_files ADD COLUMN relation_count INTEGER")
            needs_orphan_backfill = True

        if not alters:
            return
        with db.engine.begin() as conn:
            for stmt in alters:
                conn.execute(text(stmt))

        if needs_orphan_backfill:
            _backfill_evidence_orphan_metadata()
    except Exception:
        pass


def _ensure_governance_register_columns() -> None:
    """SQLite-friendly additive DDL for governance_register."""
    try:
        from sqlalchemy import inspect, text

        inspector = inspect(db.engine)
        if not inspector.has_table("governance_register"):
            return
        existing = {col["name"] for col in inspector.get_columns("governance_register")}
        alters: list[str] = []
        if "created_by_user_id" not in existing:
            alters.append("ALTER TABLE governance_register ADD COLUMN created_by_user_id INTEGER")
        if "last_updated_by_user_id" not in existing:
            alters.append("ALTER TABLE governance_register ADD COLUMN last_updated_by_user_id INTEGER")
        if "attachments_stub_json" not in existing:
            alters.append("ALTER TABLE governance_register ADD COLUMN attachments_stub_json TEXT")
        if not alters:
            return
        with db.engine.begin() as conn:
            for stmt in alters:
                conn.execute(text(stmt))
    except Exception:
        pass


def _infer_scope_from_sheet(sheet_name: str) -> int | None:
    s = (sheet_name or "").strip().lower()
    if "scope 1" in s or s.startswith("scope1"):
        return 1
    if "scope 2" in s or s.startswith("scope2"):
        return 2
    if "scope 3" in s or s.startswith("scope3"):
        return 3
    return None


def _infer_scope3_ghg_category(sheet_name: str) -> int | None:
    """
    Map template / mapped sheet name to GHG Protocol Scope 3 category number (1–15).
    Used to filter analytics per methodology category page.
    """
    if _infer_scope_from_sheet(sheet_name) != 3:
        return None
    s = (sheet_name or "").strip().lower()
    if "cat 4+9" in s or re.search(r"\b4\s*\+\s*9\b", s):
        return 9 if "downstream" in s else 4
    for cat in (15, 14, 13, 12, 11, 10, 9, 8, 7, 6, 5, 4, 3, 2, 1):
        if re.search(rf"cat(?:egory)?\s*{cat}\b", s) or re.search(rf"\bs3c{cat}\b", s):
            return cat
    if "scope 3 services spend" in s and "cat 1" not in s:
        return 1
    return None


def _find_data_source_column(df: "pd.DataFrame") -> str | None:
    if df is None or getattr(df, "columns", None) is None:
        return None

    def norm(x: str) -> str:
        return "".join(ch.lower() for ch in (x or "") if ch.isalnum())

    priority = (
        "Source_File",
        "Source file",
        "source_file",
        "Source File",
        "Data Source",
        "Data source",
        "Datasource",
        "Source",
    )
    for p in priority:
        if p in df.columns:
            return str(p)
    for c in df.columns:
        n = norm(str(c))
        if n in ("sourcefile", "sourcefilename", "datasources"):
            return str(c)
    return None


def _find_emission_factor_label_column(df: "pd.DataFrame") -> str | None:
    """Column used to group emission factor distribution (name, id, or description)."""
    if df is None or getattr(df, "columns", None) is None:
        return None

    def norm(x: str) -> str:
        return "".join(ch.lower() for ch in (x or "") if ch.isalnum())

    priority = (
        "ef_name",
        "EF Name",
        "Matched EF",
        "Emission Factor Name",
        "ef_id",
        "EF ID",
        "Emission Factor ID",
        "ef_description",
        "EF Description",
    )
    for p in priority:
        if p in df.columns:
            return str(p)
    for c in df.columns:
        n = norm(str(c))
        if n in ("efname", "emissionfactorname", "matchedef"):
            return str(c)
        if "emissionfactor" in n and "value" not in n and "unit" not in n and "category" not in n:
            return str(c)
    return None


def _df_row_emission_and_labels(
    df: "pd.DataFrame",
    tco2e_col: str | None,
    ef_col: str | None,
    src_col: str | None,
) -> tuple[dict[str, float], dict[str, float], int]:
    """Returns (ef_buckets, source_buckets, valid_row_count) weighted by tCO₂e."""
    ef_acc: dict[str, float] = defaultdict(float)
    src_acc: dict[str, float] = defaultdict(float)
    n_valid = 0
    if df is None or not tco2e_col:
        return ef_acc, src_acc, 0
    try:
        cols = [tco2e_col]
        if ef_col and ef_col in df.columns:
            cols.append(ef_col)
        if src_col and src_col in df.columns:
            cols.append(src_col)
        sub = df.loc[:, [c for c in cols if c in df.columns]]
    except Exception:
        return ef_acc, src_acc, 0

    for _, row in sub.iterrows():
        try:
            v = _parse_float_loose(row.get(tco2e_col))
        except Exception:
            v = None
        if v is None or abs(float(v)) <= 1e-18:
            continue
        n_valid += 1
        val = float(v)
        ef_key = "Unspecified"
        if ef_col and ef_col in df.columns:
            raw = row.get(ef_col)
            if raw is not None and str(raw).strip() != "" and not (isinstance(raw, float) and pd.isna(raw)):
                ef_key = str(raw).strip()[:80]
        ef_acc[ef_key] += val

        src_key = "Unspecified"
        if src_col and src_col in df.columns:
            raw_s = row.get(src_col)
            if raw_s is not None and str(raw_s).strip() != "" and not (isinstance(raw_s, float) and pd.isna(raw_s)):
                src_key = str(raw_s).strip()[:120]
        src_acc[src_key] += val

    return ef_acc, src_acc, n_valid


def _chart_top_buckets(raw: dict[str, float], top_n: int = 10) -> tuple[list[str], list[float]]:
    items = sorted(raw.items(), key=lambda x: abs(x[1]), reverse=True)
    if not items:
        return [], []
    if len(items) <= top_n:
        return [k for k, _ in items], [float(v) for _, v in items]
    head = items[: top_n - 1]
    rest = sum(float(v) for _, v in items[top_n - 1 :])
    labels = [k for k, _ in head] + ["Other"]
    values = [float(v) for _, v in head] + [rest]
    return labels, values


def _scope3_category_charts_payload(category_num: int) -> dict[str, object]:
    """
    JSON for Scope 3 methodology category dashboards: mapped outputs only.
    """
    empty: dict[str, object] = {
        "has_data": False,
        "total_tco2e": 0.0,
        "record_count": 0,
        "monthly_labels": [],
        "monthly_values": [],
        "ef_labels": [],
        "ef_values": [],
        "source_labels": [],
        "source_values": [],
    }
    try:
        if not getattr(current_user, "is_authenticated", False):
            return empty
    except Exception:
        return empty

    keys = _company_candidate_keys(getattr(current_user, "company_name", "") or "")
    if not keys:
        return empty

    latest = _latest_sheet_totals_for_company(keys)
    subs = [
        s
        for s in latest
        if int(getattr(s, "scope", 0) or 0) == 3 and _infer_scope3_ghg_category(str(getattr(s, "sheet_name", "") or "")) == category_num
    ]
    if not subs:
        return empty

    total_tco2e = 0.0
    record_count = 0
    monthly_map: dict[str, float] = defaultdict(float)
    monthly_label: dict[str, str] = {}
    ef_all: dict[str, float] = defaultdict(float)
    src_all: dict[str, float] = defaultdict(float)

    for sub in subs:
        total_tco2e += float(getattr(sub, "tco2e_total", 0.0) or 0.0)

        for r in _build_reporting_rows_from_summary(sub):
            sk = str(r.get("sortKey") or "").strip()
            if not sk:
                continue
            monthly_map[sk] += float(r.get("emissions") or 0.0)
            dl = str(r.get("dateLabel") or sk).strip()
            if sk not in monthly_label or len(dl) >= len(monthly_label.get(sk, "")):
                monthly_label[sk] = dl

        rid = str(getattr(sub, "run_id", "") or "")
        mr = MappingRun.query.get(rid) if rid else None
        op = getattr(mr, "output_path", None) if mr else None
        if not op or not os.path.exists(str(op)):
            continue
        df = _read_sheet_df_from_workbook(op, getattr(sub, "sheet_name", None))
        if df is None or getattr(df, "empty", True):
            continue
        tcol = _find_tco2e_column(df)
        ef_c = _find_emission_factor_label_column(df)
        src_c = _find_data_source_column(df)
        ef_part, src_part, n_ok = _df_row_emission_and_labels(df, tcol, ef_c, src_c)
        record_count += int(n_ok)
        for k, v in ef_part.items():
            ef_all[k] += float(v)
        for k, v in src_part.items():
            src_all[k] += float(v)

    if record_count <= 0 and subs:
        record_count = sum(int(getattr(s, "rows_count", 0) or 0) for s in subs)

    has_any = (
        total_tco2e > 1e-12
        or bool(monthly_map)
        or bool(ef_all)
        or bool(src_all)
        or int(record_count or 0) > 0
    )
    if not has_any:
        return empty

    month_keys = sorted(monthly_map.keys(), key=lambda x: str(x))
    monthly_labels = [monthly_label.get(k, k) for k in month_keys]
    monthly_values = [round(float(monthly_map[k]), 6) for k in month_keys]

    ef_labels, ef_values = _chart_top_buckets(dict(ef_all), 10)
    src_labels, src_values = _chart_top_buckets(dict(src_all), 10)

    return {
        "has_data": True,
        "total_tco2e": round(float(total_tco2e), 6),
        "record_count": int(record_count),
        "monthly_labels": monthly_labels,
        "monthly_values": monthly_values,
        "ef_labels": ef_labels,
        "ef_values": [round(float(v), 6) for v in ef_values],
        "source_labels": src_labels,
        "source_values": [round(float(v), 6) for v in src_values],
    }


def _count_company_mapped_categories(company_name: str) -> int:
    company_keys = _company_candidate_keys(company_name)
    if not company_keys:
        return 0

    runs = (
        MappingRun.query.filter(
            MappingRun.status == "succeeded",
            MappingRun.company_name.in_(company_keys),
        )
        .order_by(MappingRun.created_at.desc())
        .all()
    )
    seen: set[str] = set()
    for run in runs:
        sheet_key = (getattr(run, "sheet_name", "") or "").strip().lower()
        if sheet_key:
            seen.add(sheet_key)
    return len(seen)


def _calculate_mapping_coverage(company_name: str) -> tuple[int, int, float]:
    mapped_categories_count = int(_count_company_mapped_categories(company_name) or 0)
    total_categories = int(_count_company_schema_sheets(company_name) or 0)
    coverage_pct = round((mapped_categories_count / total_categories) * 100, 2) if total_categories > 0 else 0.0
    return mapped_categories_count, total_categories, coverage_pct


def _upsert_mapping_run_summary(
    run_id: str,
    company_name: str,
    sheet_name: str,
    mapped_df: "pd.DataFrame",
    output_path: str | Path | None,
) -> None:
    total_tco2e = 0.0
    rows_count = 0
    used_col = None
    if output_path:
        total_tco2e, rows_count, used_col = _sum_tco2e_from_xlsx(output_path, sheet_name)
    if used_col is None:
        total_tco2e, rows_count, used_col = _sum_tco2e(mapped_df)

    mapped_categories_count, total_categories, coverage_pct = _calculate_mapping_coverage(company_name)
    print("SUMMARY:", mapped_categories_count, total_categories, coverage_pct)

    summ = MappingRunSummary.query.filter_by(run_id=run_id).first()
    if not summ:
        company_key = (company_name or "").strip()
        sheet_key = (sheet_name or "").strip().lower()
        existing_for_sheet = (
            MappingRunSummary.query.filter_by(company_name=company_key)
            .order_by(MappingRunSummary.created_at.desc())
            .all()
        )
        for row in existing_for_sheet:
            row_sheet_key = (getattr(row, "sheet_name", "") or "").strip().lower()
            if row_sheet_key == sheet_key:
                summ = row
                break

    if not summ:
        summ = MappingRunSummary(run_id=run_id)
        db.session.add(summ)

    summ.run_id = run_id
    summ.company_name = company_name
    summ.sheet_name = sheet_name
    summ.scope = _infer_scope_from_sheet(sheet_name)
    summ.tco2e_total = float(total_tco2e or 0.0)
    summ.rows_count = int(rows_count or 0)
    summ.mapped_categories_count = int(mapped_categories_count or 0)
    summ.total_categories = int(total_categories or 0)
    summ.coverage_pct = float(coverage_pct or 0.0)
    summ.created_at = datetime.now()


def _df_find_column_casefold(df: "pd.DataFrame", candidates: tuple[str, ...]) -> str | None:
    if df is None or getattr(df, "columns", None) is None:
        return None
    lowered = {str(c).strip().lower(): str(c) for c in df.columns}
    for candidate in candidates:
        hit = lowered.get(str(candidate).strip().lower())
        if hit:
            return hit
    return None


def _json_safe_value(value: object) -> object:
    try:
        if pd.isna(value):
            return ""
    except Exception:
        pass
    if isinstance(value, (datetime, pd.Timestamp)):
        return value.isoformat()
    if isinstance(value, date):
        return value.isoformat()
    try:
        if hasattr(value, "item"):
            return value.item()
    except Exception:
        pass
    return value


def _unmapped_row_label(row: "pd.Series") -> str:
    candidates = (
        "Product Type",
        "Product type",
        "Service Provided",
        "Service provided",
        "Purchased Goods and Services",
        "Spend Category",
        "Description",
        "Supplier",
        "Vendor",
        "BillOfQuantity",
        "Bill of Quantity",
        "Activity",
    )
    parts: list[str] = []
    for col in candidates:
        if col not in row.index:
            continue
        value = str(row.get(col) or "").strip()
        if value and value.lower() not in {"nan", "none"} and value not in parts:
            parts.append(value)
        if len(parts) >= 3:
            break
    return " / ".join(parts)[:500] if parts else "Unmapped row"


def _mapping_unmapped_context_value(row: "pd.Series", candidates: tuple[str, ...]) -> str:
    for col in candidates:
        if col in row.index:
            value = str(row.get(col) or "").strip()
            if value and value.lower() not in {"nan", "none"}:
                return value
    return ""


def _sync_unmapped_rows_for_mapping_run(
    *,
    run_id: str,
    user_id: int | None,
    company_name: str,
    sheet_name: str,
    source_entry_group: str | None,
    mapped_df: "pd.DataFrame",
) -> int:
    if mapped_df is None or mapped_df.empty:
        return 0
    entry_group = str(source_entry_group or "").strip() or None
    previous_query = MappingUnmappedRow.query.filter(
        MappingUnmappedRow.company_name == company_name,
        MappingUnmappedRow.sheet_name == sheet_name,
        MappingUnmappedRow.review_status == "open",
    )
    if entry_group:
        previous_query = previous_query.filter(MappingUnmappedRow.source_entry_group == entry_group)
    superseded_count = 0
    for previous in previous_query.all():
        previous.review_status = "superseded"
        previous.resolved_at = datetime.utcnow()
        superseded_count += 1

    print(
        f"[UNMAPPED_SYNC] Superseded {superseded_count} open row(s) for "
        f"{company_name} / {sheet_name}{' / ' + entry_group if entry_group else ''}"
    )

    status_col = _df_find_column_casefold(mapped_df, ("status", "Status", "mapping_status", "Mapping Status"))
    if not status_col:
        MappingUnmappedRow.query.filter_by(run_id=run_id).delete()
        print(f"[UNMAPPED_SYNC] Mapping output has no status column for run {run_id}; no open rows inserted")
        return 0

    status_series = mapped_df[status_col].fillna("").astype(str).str.strip().str.lower()
    no_match_mask = status_series == "no match"
    if not bool(getattr(no_match_mask, "any", lambda: False)()):
        MappingUnmappedRow.query.filter_by(run_id=run_id).delete()
        print(f"[UNMAPPED_SYNC] No open No match rows remain for run {run_id}")
        return 0

    MappingUnmappedRow.query.filter_by(run_id=run_id).delete()
    inserted = 0
    for idx in mapped_df[no_match_mask].index.tolist():
        row = mapped_df.loc[idx]
        payload = {str(col): _json_safe_value(row.get(col)) for col in mapped_df.columns}
        db.session.add(
            MappingUnmappedRow(
                run_id=run_id,
                user_id=user_id,
                company_name=company_name,
                sheet_name=sheet_name,
                source_entry_group=entry_group,
                row_number=int(idx) + 2,
                row_label=_unmapped_row_label(row),
                status_value=str(row.get(status_col) or "No match").strip() or "No match",
                row_payload=json.dumps(payload, ensure_ascii=True, default=str),
                review_status="open",
            )
        )
        inserted += 1
    print(f"[UNMAPPED_SYNC] Inserted {inserted} open No match row(s) for run {run_id}")
    return inserted


def _unmapped_row_preview(row: MappingUnmappedRow) -> dict[str, object]:
    try:
        payload = json.loads(row.row_payload or "{}")
    except Exception:
        payload = {}
    visible_keys = [
        "Product Type",
        "Service Provided",
        "Supplier",
        "Vendor",
        "Description",
        "Spend Category",
        "BillOfQuantity",
        "Bill of Quantity",
        "Spend_Euro",
        "Spend",
        "Quantity",
        "Unit",
    ]
    details = [
        {"name": key, "value": str(payload.get(key) or "")}
        for key in visible_keys
        if str(payload.get(key) or "").strip()
    ]
    if not details:
        details = [
            {"name": str(key), "value": str(value)}
            for key, value in list(payload.items())[:8]
            if str(value or "").strip()
        ]
    return {
        "id": row.id,
        "run_id": row.run_id,
        "company_name": row.company_name,
        "sheet_name": row.sheet_name,
        "source_entry_group": row.source_entry_group or "",
        "row_number": row.row_number,
        "row_label": row.row_label or "Unmapped row",
        "status_value": row.status_value or "No match",
        "review_status": row.review_status or "open",
        "assigned_ef_id": row.assigned_ef_id or "",
        "owner_notes": row.owner_notes or "",
        "created_at": row.created_at,
        "details": details[:10],
        "ef_search": row.row_label or _mapping_unmapped_context_value(pd.Series(payload), ("Product Type", "Service Provided", "Description")),
    }


def _unmapped_payload(row: MappingUnmappedRow) -> dict[str, object]:
    try:
        payload = json.loads(row.row_payload or "{}")
    except Exception:
        payload = {}
    return payload if isinstance(payload, dict) else {}


def _unmapped_description(row: MappingUnmappedRow) -> str:
    payload = _unmapped_payload(row)
    for key in ("Description", "description"):
        value = str(payload.get(key) or "").strip()
        if value and value.lower() not in {"nan", "none"}:
            return value
    return ""


def _unmapped_description_key(row: MappingUnmappedRow) -> str:
    value = _unmapped_description(row)
    return re.sub(r"\s+", " ", value).strip().lower()


def _dedupe_unmapped_rows_by_description(rows: list[MappingUnmappedRow]) -> tuple[list[MappingUnmappedRow], dict[int, int]]:
    kept: list[MappingUnmappedRow] = []
    seen: dict[str, int] = {}
    duplicate_counts: dict[int, int] = {}
    for row in rows:
        desc_key = _unmapped_description_key(row)
        if not desc_key:
            kept.append(row)
            duplicate_counts[int(row.id)] = 1
            continue
        group_key = desc_key
        if group_key in seen:
            duplicate_counts[seen[group_key]] = duplicate_counts.get(seen[group_key], 1) + 1
            continue
        seen[group_key] = int(row.id)
        kept.append(row)
        duplicate_counts[int(row.id)] = 1
    return kept, duplicate_counts


def _is_cat1_purchased_goods_services_sheet(sheet_name: str) -> bool:
    normalized = str(sheet_name or "").strip().lower()
    return normalized in {
        "scope 3 category 1 purchased goods & services",
        "scope 3 cat 1 goods spend",
        "scope 3 cat 1 services spend",
        "scope 3 cat 1 common purchases",
        "scope 3 cat 1 goods activity",
        "scope 3 cat 1 services activity",
        "scope 3 cat 1 supplier summary",
        "scope 3 cat 1 goods service",
        "scope 3 cat 1 goods services",
    }


def _unmapped_category_sheet_names(sheet_name: str) -> set[str]:
    raw = str(sheet_name or "").strip()
    canonical = STAGE2_2026_SHEET_ALIASES.get(raw, raw)
    names = {raw, canonical}
    if _is_cat1_purchased_goods_services_sheet(raw):
        names.update({"Scope 3 Purchased Goods Spend", "Scope 3 Purchased Service Spend"})
    return {name for name in names if name}


def _is_klarakarbon_ef_row(row: dict[str, object]) -> bool:
    category = str(row.get("ef_category") or "").strip().lower()
    return category.startswith("klarakarbon")


def _ef_option_key(option: dict[str, object]) -> str:
    return f"{option.get('sheet') or ''}||{option.get('ef_id') or ''}"


def _ef_option_label(option: dict[str, object]) -> str:
    name = str(option.get("ef_name") or option.get("Emission Factor Category") or option.get("ef_description") or "").strip()
    ef_id = str(option.get("ef_id") or "").strip()
    if name and ef_id:
        return f"{name} ({ef_id})"
    return ef_id or name


def _load_unmapped_ef_options_for_sheet(sheet_name: str) -> list[dict[str, object]]:
    data = _load_stage2_emission_factors()
    all_rows: list[dict[str, object]] = list(data.get("rows") or [])
    category_sheets = _unmapped_category_sheet_names(sheet_name)
    category_norms = {name.strip().lower() for name in category_sheets}
    is_cat1 = _is_cat1_purchased_goods_services_sheet(sheet_name)

    options: list[dict[str, object]] = []
    for row in all_rows:
        if _is_klarakarbon_ef_row(row):
            continue
        ef_id = str(row.get("ef_id") or "").strip()
        if not ef_id:
            continue
        row_sheet = str(row.get("sheet") or "").strip()
        row_sheet_norm = row_sheet.lower()
        ef_category_norm = str(row.get("ef_category") or "").strip().lower()
        if is_cat1:
            if row_sheet_norm not in {"scope 3 purchased goods spend", "scope 3 purchased service spend"}:
                continue
        elif category_norms:
            matches_sheet = row_sheet_norm in category_norms
            matches_category = any(name and name in ef_category_norm for name in category_norms)
            if not matches_sheet and not matches_category:
                continue
        item = dict(row)
        item["key"] = _ef_option_key(item)
        item["label"] = _ef_option_label(item)
        if item["label"]:
            options.append(item)

    seen: set[tuple[str, str]] = set()
    unique: list[dict[str, object]] = []
    for option in sorted(options, key=lambda r: (str(r.get("ef_id") or ""), str(r.get("ef_name") or "").lower())):
        key = (str(option.get("sheet") or ""), str(option.get("ef_id") or ""))
        if key in seen:
            continue
        seen.add(key)
        unique.append(option)
    return unique[:500]


def _find_unmapped_ef_option(ef_key_or_id: str, preferred_sheet: str | None = None) -> dict[str, object] | None:
    value = (ef_key_or_id or "").strip()
    if not value:
        return None
    wanted_sheet = (preferred_sheet or "").strip()
    wanted_ef_id = value
    if "||" in value:
        wanted_sheet, wanted_ef_id = value.split("||", 1)
        wanted_sheet = wanted_sheet.strip()
        wanted_ef_id = wanted_ef_id.strip()
    data = _load_stage2_emission_factors()
    for row in list(data.get("rows") or []):
        if _is_klarakarbon_ef_row(row):
            continue
        if str(row.get("ef_id") or "").strip() != wanted_ef_id:
            continue
        if wanted_sheet and str(row.get("sheet") or "").strip() != wanted_sheet:
            continue
        option = dict(row)
        option["key"] = _ef_option_key(option)
        option["label"] = _ef_option_label(option)
        return option
    return None


def _apply_unmapped_sheet_filters(query, company_filter: str, sheet_filter: str, search: str):
    """Company / sheet / text filters shared by the admin unmapped table and summary tiles."""
    company_filter = (company_filter or "").strip()
    sheet_filter = (sheet_filter or "").strip()
    search = (search or "").strip()
    if company_filter:
        query = query.filter(MappingUnmappedRow.company_name.ilike(f"%{company_filter}%"))
    if sheet_filter:
        query = query.filter(MappingUnmappedRow.sheet_name.ilike(f"%{sheet_filter}%"))
    if search:
        like = f"%{search}%"
        query = query.filter(
            or_(
                MappingUnmappedRow.row_label.ilike(like),
                MappingUnmappedRow.row_payload.ilike(like),
                MappingUnmappedRow.assigned_ef_id.ilike(like),
                MappingUnmappedRow.owner_notes.ilike(like),
            )
        )
    return query


def _unmapped_query_from_request():
    status_filter = (request.args.get("status") or "open").strip().lower()
    company_filter = (request.args.get("company") or "").strip()
    sheet_filter = (request.args.get("sheet") or "").strip()
    search = (request.args.get("search") or "").strip()

    query = MappingUnmappedRow.query
    query = _apply_unmapped_sheet_filters(query, company_filter, sheet_filter, search)
    if status_filter and status_filter != "all":
        query = query.filter(MappingUnmappedRow.review_status == status_filter)
    return query, status_filter, company_filter, sheet_filter, search


def _refresh_open_unmapped_against_live_rows() -> int:
    """
    Reconcile open MappingUnmappedRow records with live Data Entry EF status.
    Only touches (company, sheet) pairs that still have open rows — avoids scanning the
    full unmapped table against all Data Entry on every admin page load ([UNMAPPED_REFRESH]).
    """
    pairs = (
        db.session.query(MappingUnmappedRow.company_name, MappingUnmappedRow.sheet_name)
        .filter(MappingUnmappedRow.review_status == "open")
        .distinct()
        .all()
    )
    total = 0
    for c, s in pairs:
        co = str(c or "").strip()
        sh = str(s or "").strip()
        if not co or not sh:
            continue
        total += _delete_stale_open_unmapped_rows_from_live_data_entry(co, sh)
    if pairs:
        print(f"[UNMAPPED_REFRESH] scoped_stale_cleanup pairs_checked={len(pairs)} rows_removed={total}")
    return total


def _unmapped_review_status_counts(
    *,
    company_filter: str = "",
    sheet_filter: str = "",
    search: str = "",
) -> dict[str, int]:
    """
    Grouped counts for summary tiles — must apply the same list filters as the visible table.
    Uses one GROUP BY round-trip when no dedupe-distinct correction is requested.
    """
    base = MappingUnmappedRow.query
    base = _apply_unmapped_sheet_filters(base, company_filter, sheet_filter, search)
    rows = (
        base.with_entities(MappingUnmappedRow.review_status, func.count(MappingUnmappedRow.id))
        .group_by(MappingUnmappedRow.review_status)
        .all()
    )
    by_status: dict[str, int] = defaultdict(int)
    total = 0
    for st, n in rows:
        key = str(st or "").strip().lower() or "open"
        c = int(n or 0)
        total += c
        by_status[key] += c
    return {
        "open": int(by_status.get("open", 0)),
        "resolved": int(by_status.get("resolved", 0)),
        "ignored": int(by_status.get("ignored", 0)),
        "superseded": int(by_status.get("superseded", 0)),
        "all": total,
    }


def _unmapped_open_distinct_description_count(
    *,
    company_filter: str = "",
    sheet_filter: str = "",
    search: str = "",
) -> int:
    """Open rows only: unique descriptions (aligned with dedupe=description table view semantics)."""
    q = MappingUnmappedRow.query
    q = _apply_unmapped_sheet_filters(q, company_filter, sheet_filter, search)
    q = q.filter(MappingUnmappedRow.review_status == "open")
    desc_keys: set[str] = set()
    for row in q.all():
        desc_keys.add(_unmapped_description_key(row))
    return len(desc_keys)


def _admin_unmapped_page_counts(
    *,
    company_filter: str,
    sheet_filter: str,
    search: str,
    dedupe_mode: str,
) -> dict[str, int]:
    """Summary tiles — same scoped filters as the table; dedupe mode adjusts Open to distinct descriptions."""
    counts = _unmapped_review_status_counts(
        company_filter=company_filter,
        sheet_filter=sheet_filter,
        search=search,
    )
    if (dedupe_mode or "").strip().lower() == "description":
        distinct_open = _unmapped_open_distinct_description_count(
            company_filter=company_filter,
            sheet_filter=sheet_filter,
            search=search,
        )
        counts["open"] = int(distinct_open)
        counts["all"] = (
            int(counts.get("resolved", 0))
            + int(counts.get("ignored", 0))
            + int(counts.get("superseded", 0))
            + int(distinct_open)
        )
    print(
        f"[UNMAPPED_COUNTERS]\n"
        f"open={counts['open']}\n"
        f"resolved={counts['resolved']}\n"
        f"superseded={counts['superseded']}\n"
        f"ignored={counts['ignored']}"
    )
    return counts


def _backup_stage2_ef_workbook(action: str) -> None:
    backup_dir = STAGE2_EF_XLSX.parent / "_ef_backups"
    backup_dir.mkdir(parents=True, exist_ok=True)
    ts = datetime.now().strftime("%Y%m%d_%H%M%S")
    backup_path = backup_dir / f"{STAGE2_EF_XLSX.stem}_before_{action}_{ts}{STAGE2_EF_XLSX.suffix}"
    try:
        shutil.copy2(STAGE2_EF_XLSX, backup_path)
    except Exception:
        pass


def _coerce_excel_number(value: object) -> object:
    try:
        text = str(value or "").strip()
        if text == "":
            return None
        return float(text)
    except Exception:
        return value


def _append_unmapped_mapping_to_workbook(row: MappingUnmappedRow, ef_option: dict[str, object]) -> tuple[str, bool]:
    description = _unmapped_description(row)
    if not description:
        raise ValueError("Description column is empty for this unmapped row.")

    is_cat1 = _is_cat1_purchased_goods_services_sheet(row.sheet_name or "")
    target_sheet = "All together" if is_cat1 else str(ef_option.get("sheet") or "").strip()
    if not target_sheet:
        raise ValueError("Could not determine the target emission factor sheet.")

    wb = load_workbook(STAGE2_EF_XLSX, keep_links=False)
    if target_sheet not in wb.sheetnames:
        raise ValueError(f"Target sheet not found in mapping workbook: {target_sheet}")
    ws = wb[target_sheet]
    headers = [("" if v is None else str(v).strip()) for v in (next(ws.iter_rows(min_row=1, max_row=1, values_only=True), None) or [])]
    while headers and headers[-1] == "":
        headers.pop()
    header_idx = {h: idx + 1 for idx, h in enumerate(headers) if h}

    if is_cat1:
        required = ["Product type", "ef_description", "ef_id", "ef_name", "ef_value", "ef_unit", "ef_source", "scope"]
        missing = [h for h in required if h not in header_idx]
        if missing:
            raise ValueError(f"All together sheet headers are missing: {', '.join(missing)}")
        product_col = header_idx["Product type"]
        ef_id_col = header_idx["ef_id"]
        assigned_ef_id = str(ef_option.get("ef_id") or "").strip()
        desc_key = re.sub(r"\s+", " ", description).strip().lower()
        for values in ws.iter_rows(min_row=2, values_only=True):
            existing_desc = values[product_col - 1] if product_col - 1 < len(values) else None
            existing_ef_id = values[ef_id_col - 1] if ef_id_col - 1 < len(values) else None
            if re.sub(r"\s+", " ", str(existing_desc or "")).strip().lower() == desc_key and str(existing_ef_id or "").strip() == assigned_ef_id:
                return target_sheet, False
        new_row = ws.max_row + 1
        values = {
            "Product type": description,
            "ef_description": description,
            "ef_id": assigned_ef_id,
            "ef_name": str(ef_option.get("ef_name") or ef_option.get("Emission Factor Category") or "").strip(),
            "ef_value": _coerce_excel_number(ef_option.get("ef_value")),
            "ef_unit": str(ef_option.get("ef_unit") or "").strip(),
            "ef_source": str(ef_option.get("ef_source") or "").strip(),
            "scope": str(ef_option.get("scope") or "3").strip() or "3",
        }
    else:
        required = _ef_expected_headers()
        missing = [h for h in required if h not in header_idx]
        if missing:
            raise ValueError(f"{target_sheet} sheet headers are missing: {', '.join(missing)}")
        ef_id_col = header_idx["ef_id"]
        desc_col = header_idx["ef_description"]
        assigned_ef_id = str(ef_option.get("ef_id") or "").strip()
        desc_key = re.sub(r"\s+", " ", description).strip().lower()
        for values in ws.iter_rows(min_row=2, values_only=True):
            existing_desc = values[desc_col - 1] if desc_col - 1 < len(values) else None
            existing_ef_id = values[ef_id_col - 1] if ef_id_col - 1 < len(values) else None
            if re.sub(r"\s+", " ", str(existing_desc or "")).strip().lower() == desc_key and str(existing_ef_id or "").strip() == assigned_ef_id:
                return target_sheet, False
        new_row = ws.max_row + 1
        values = {
            "ef_name": str(ef_option.get("ef_name") or ef_option.get("Emission Factor Category") or "").strip(),
            "ef_description": description,
            "scope": str(ef_option.get("scope") or "").strip(),
            "ef_category": str(ef_option.get("ef_category") or row.sheet_name or "").strip(),
            "ef_id": assigned_ef_id,
            "ef_value": _coerce_excel_number(ef_option.get("ef_value")),
            "ef_unit": str(ef_option.get("ef_unit") or "").strip(),
            "ef_source": str(ef_option.get("ef_source") or "").strip(),
            "Emission Factor Category": str(ef_option.get("Emission Factor Category") or ef_option.get("ef_name") or "").strip(),
        }

    _backup_stage2_ef_workbook("unmapped_mapping")
    for header, value in values.items():
        ws.cell(row=new_row, column=header_idx[header]).value = value
    wb.save(STAGE2_EF_XLSX)
    _clear_ef_cache()
    return target_sheet, True


def _open_duplicate_rows_for_unmapped(row: MappingUnmappedRow) -> list[MappingUnmappedRow]:
    desc_key = _unmapped_description_key(row)
    if not desc_key:
        return [row]
    candidates = MappingUnmappedRow.query.filter(
        MappingUnmappedRow.sheet_name == row.sheet_name,
        MappingUnmappedRow.review_status == "open",
    ).all()
    rows = [candidate for candidate in candidates if _unmapped_description_key(candidate) == desc_key]
    if row not in rows:
        rows.append(row)
    return rows


def _apply_unmapped_mapping(
    row: MappingUnmappedRow,
    ef_option: dict[str, object],
    *,
    owner_notes: str | None = None,
    resolve_duplicates: bool = True,
) -> tuple[str, int, bool]:
    target_sheet, inserted = _append_unmapped_mapping_to_workbook(row, ef_option)
    assigned_ef_id = str(ef_option.get("ef_id") or "").strip() or None
    now = datetime.utcnow()
    user_id = int(getattr(current_user, "id", 0) or 0) or None
    rows_to_update = _open_duplicate_rows_for_unmapped(row) if resolve_duplicates else [row]
    for item in rows_to_update:
        item.assigned_ef_id = assigned_ef_id
        if owner_notes is not None:
            item.owner_notes = owner_notes.strip() or item.owner_notes
        item.review_status = "resolved"
        item.resolved_at = now
        item.resolved_by_user_id = user_id
    return target_sheet, len(rows_to_update), inserted


def _import_stage2_compute_emissions_tco2e():
    mod = _import_stage2_main_mapping()
    return getattr(mod, "_compute_emissions_tco2e")


def _unmapped_data_entry_header_lookup(headers: list[str]) -> dict[str, str]:
    """Map canonical field keys -> exact header string from the active template."""
    canon = {str(h).strip().lower(): str(h) for h in headers if str(h).strip()}
    wanted: dict[str, tuple[str, ...]] = {
        "ef_id": ("ef_id", "ef id"),
        "ef_name": ("ef_name", "ef name"),
        "ef_unit": ("ef_unit", "ef unit"),
        "ef_value": ("ef_value", "ef value"),
        "ef_source": ("ef_source", "ef source", "source"),
        "emissions_tco2e": ("emissions_tco2e", "emissions (tco2e)", "emissions tco2e", "emissions"),
        "match_method": ("match_method", "match method"),
        "status": ("status", "mapping status"),
    }
    out: dict[str, str] = {}
    for key, candidates in wanted.items():
        hit: str | None = None
        for cand in candidates:
            if cand.lower() in canon:
                hit = canon[cand.lower()]
                break
        if hit:
            out[key] = hit
    return out


def _spend_euro_from_unmapped_payload(payload: dict[str, object]) -> float | None:
    for col in ("Spend_Euro", "Spend Euro", "spend_euro", "Spend"):
        if col in payload:
            val = _to_numeric_spend(payload.get(col))
            if val is not None:
                return float(val)
    return None


def _locate_data_entry_grid_row_for_unmapped(row: MappingUnmappedRow, headers: list[str]) -> dict[str, object] | None:
    company = _resolve_template_company_name(row.company_name) or str(row.company_name or "").strip()
    sheet = _resolve_template_sheet_name(company, row.sheet_name) or str(row.sheet_name or "").strip()
    grid_rows = _load_data_entry_grid_rows(company, sheet, headers)
    entry_group = str(row.source_entry_group or "").strip()
    if entry_group:
        grid_rows = [r for r in grid_rows if str(r.get("entry_group") or "").strip() == entry_group]

    pos = int(row.row_number or 0) - 2
    if pos >= 0 and pos < len(grid_rows):
        return grid_rows[pos]

    desc = _unmapped_description(row)
    if not desc:
        return None
    desc_key = re.sub(r"\s+", " ", desc).strip().lower()
    lookup = {str(h).strip().lower() for h in headers}
    desc_header = None
    for cand in ("Description", "description"):
        if cand.lower() in lookup:
            desc_header = cand
            break
    if not desc_header:
        return None
    try:
        col_idx = headers.index(desc_header)
    except Exception:
        return None
    for candidate in grid_rows:
        cells = list(candidate.get("cells") or [])
        if col_idx >= len(cells):
            continue
        cell_val = str(cells[col_idx] or "").strip()
        if re.sub(r"\s+", " ", cell_val).strip().lower() == desc_key:
            return candidate
    return None


def _set_data_entry_column_values(
    *,
    company_name: str,
    sheet_name: str,
    entry_group: str,
    row_index: int,
    created_at: datetime | None,
    updates: dict[str, str],
) -> None:
    if not updates:
        return
    query = DataEntry.query.filter_by(
        company_name=company_name,
        sheet_name=sheet_name,
        entry_group=entry_group,
        row_index=int(row_index or 0),
    )
    if created_at is not None:
        query = query.filter(DataEntry.created_at == created_at)
    entries = query.all()
    if not entries:
        ts = created_at or datetime.utcnow()
        uid = int(getattr(current_user, "id", 0) or 0) or None
        for column_name, text in updates.items():
            if str(text or "").strip() == "":
                continue
            db.session.add(
                DataEntry(
                    company_name=company_name,
                    sheet_name=sheet_name,
                    entry_group=entry_group,
                    uploaded_by_user_id=uid,
                    row_index=int(row_index or 0),
                    column_name=str(column_name),
                    value=str(text).strip(),
                    created_at=ts,
                )
            )
        return

    for entry in entries:
        column_name = str(getattr(entry, "column_name", "") or "")
        if column_name not in updates:
            continue
        new_val = updates[column_name]
        if str(new_val or "").strip() == "":
            entry.value = None
        else:
            entry.value = str(new_val).strip()


def _apply_unmapped_mapping_to_data_entry_row(
    row: MappingUnmappedRow,
    ef_option: dict[str, object],
    *,
    match_method: str,
    status_text: str = "Mapped",
) -> None:
    company = _resolve_template_company_name(row.company_name) or str(row.company_name or "").strip()
    sheet = _resolve_template_sheet_name(company, row.sheet_name) or str(row.sheet_name or "").strip()
    headers, _rules = _get_data_entry_template_schema(company, sheet)
    if not headers:
        raise ValueError("Data entry headers not found for this company and sheet.")

    target = _locate_data_entry_grid_row_for_unmapped(row, headers)
    if not target:
        raise ValueError("Could not locate the matching data entry row for this unmapped record.")

    header_by_key = _unmapped_data_entry_header_lookup(headers)
    payload = _unmapped_payload(row)
    spend_eur = _spend_euro_from_unmapped_payload(payload)
    compute = _import_stage2_compute_emissions_tco2e()
    emissions = compute(
        spend_eur,
        _coerce_excel_number(ef_option.get("ef_value")),
        str(ef_option.get("ef_unit") or "").strip() or None,
    )

    values: dict[str, str] = {}
    mapping = {
        "ef_id": str(ef_option.get("ef_id") or "").strip(),
        "ef_name": str(ef_option.get("ef_name") or ef_option.get("Emission Factor Category") or "").strip(),
        "ef_unit": str(ef_option.get("ef_unit") or "").strip(),
        "ef_value": "" if ef_option.get("ef_value") is None else str(_coerce_excel_number(ef_option.get("ef_value"))),
        "ef_source": str(ef_option.get("ef_source") or "").strip(),
        "match_method": str(match_method or "").strip(),
        "status": str(status_text or "").strip(),
        "emissions_tco2e": "" if emissions is None else str(round(float(emissions), 10)).rstrip("0").rstrip("."),
    }
    for key, text in mapping.items():
        header_name = header_by_key.get(key)
        if not header_name:
            continue
        values[header_name] = text

    entry_group = str(target.get("entry_group") or "").strip()
    if entry_group.startswith("legacy:"):
        raise ValueError("Legacy data entry rows cannot be auto-updated from this tool.")
    row_index = int(target.get("row_index") or 0)
    created_at = _parse_iso_datetime(target.get("created_at"))

    _set_data_entry_column_values(
        company_name=company,
        sheet_name=sheet,
        entry_group=entry_group,
        row_index=row_index,
        created_at=created_at,
        updates=values,
    )


def _apply_unmapped_map_full(
    row: MappingUnmappedRow,
    ef_option: dict[str, object],
    *,
    owner_notes: str | None = None,
    match_method: str = "manual:unmapped_map",
) -> tuple[str, int, bool]:
    target_sheet, resolved_count, inserted = _apply_unmapped_mapping(
        row, ef_option, owner_notes=owner_notes, resolve_duplicates=True
    )
    # Data entry row applies to the representative unmapped row only (same grid row for duplicates)
    _apply_unmapped_mapping_to_data_entry_row(row, ef_option, match_method=match_method, status_text="Mapped")
    return target_sheet, resolved_count, inserted


def _all_together_cell(row_values: tuple[object, ...], header_idx: dict[str, int], header: str) -> object:
    idx = header_idx.get(header)
    if idx is None or idx >= len(row_values):
        return ""
    return row_values[idx]


def _load_all_together_fuzzy_options() -> list[dict[str, object]]:
    if not STAGE2_EF_XLSX.exists():
        return []
    wb = load_workbook(STAGE2_EF_XLSX, read_only=True, data_only=True, keep_links=False)
    if "All together" not in wb.sheetnames:
        return []
    ws = wb["All together"]
    first = next(ws.iter_rows(min_row=1, max_row=1, values_only=True), None)
    headers = [("" if v is None else str(v).strip()) for v in (first or [])]
    header_idx = {h: idx for idx, h in enumerate(headers) if h}
    required = {"Product type", "ef_id"}
    if not required.issubset(header_idx):
        return []

    options: list[dict[str, object]] = []
    for row_values in ws.iter_rows(min_row=2, values_only=True):
        product_type = str(_all_together_cell(row_values, header_idx, "Product type") or "").strip()
        ef_id = str(_all_together_cell(row_values, header_idx, "ef_id") or "").strip()
        if not product_type or not ef_id:
            continue
        option = {
            "sheet": "All together",
            "Product type": product_type,
            "ef_description": str(_all_together_cell(row_values, header_idx, "ef_description") or "").strip(),
            "ef_id": ef_id,
            "ef_name": str(_all_together_cell(row_values, header_idx, "ef_name") or "").strip(),
            "ef_value": _all_together_cell(row_values, header_idx, "ef_value"),
            "ef_unit": str(_all_together_cell(row_values, header_idx, "ef_unit") or "").strip(),
            "ef_source": str(_all_together_cell(row_values, header_idx, "ef_source") or "").strip(),
            "scope": str(_all_together_cell(row_values, header_idx, "scope") or "3").strip() or "3",
        }
        option["key"] = _ef_option_key(option)
        option["label"] = _ef_option_label(option)
        options.append(option)
    return options


def _parse_unmapped_fuzzy_threshold(raw: object, default: float = 0.92) -> float:
    try:
        value = float(str(raw or "").strip().replace(",", "."))
    except Exception:
        return default
    if value <= 0:
        return default
    return max(0.0, min(1.0, value))


def _build_unmapped_fuzzy_suggestions(rows: list[MappingUnmappedRow], threshold: float = 0.92) -> list[dict[str, object]]:
    suggestions: list[dict[str, object]] = []
    all_together_options = _load_all_together_fuzzy_options()
    for row in rows:
        description = _unmapped_description(row)
        if not description:
            continue
        desc_norm = _norm_name(description)
        if not desc_norm:
            continue
        best_option: dict[str, object] | None = None
        best_score = 0.0
        for option in all_together_options:
            candidate = str(option.get("Product type") or "").strip()
            cand_norm = _norm_name(candidate)
            if not cand_norm:
                continue
            score = difflib.SequenceMatcher(None, desc_norm, cand_norm).ratio()
            if score > best_score:
                best_score = score
                best_option = option
        if best_option is not None and best_score >= threshold:
            preview = _unmapped_row_preview(row)
            preview["description"] = description
            suggestions.append(
                {
                    "row": preview,
                    "ef": best_option,
                    "ef_key": _ef_option_key(best_option),
                    "score": best_score,
                    "score_pct": f"{best_score:.2f}",
                    "match_method": f"Cat1: All together fuzzy:{best_score:.2f}",
                }
            )
    return suggestions


def _find_tco2e_column(df: "pd.DataFrame") -> str | None:
    if df is None or getattr(df, "columns", None) is None:
        return None

    def norm(x: str) -> str:
        return "".join(ch.lower() for ch in (x or "") if ch.isalnum())

    candidates: list[str] = []
    for c in list(df.columns):
        n = norm(str(c))
        if n in {"emissionstco2e", "emissionstco2etonnes", "tco2e", "totalemissionstco2e"}:
            candidates.append(str(c))
            continue
        if n in ("co2e", "co2et", "co2etonnes", "co2eton", "co2etonne"):
            candidates.append(str(c))
            continue
        if n.startswith("emissionstco2e") or n.startswith("co2et") or n.startswith("co2e"):
            candidates.append(str(c))
            continue
        if ("co2e" in n) and ("calculated" in n):
            candidates.append(str(c))

    if not candidates:
        return None

    best_col = None
    best_score = (-1, -1.0)
    for c in candidates:
        try:
            ser = df[c]
        except Exception:
            continue

        numeric_count = 0
        total_abs = 0.0
        try:
            for v in ser.tolist():
                f = _parse_float_loose(v)
                if f is None:
                    continue
                numeric_count += 1
                total_abs += abs(float(f))
        except Exception:
            continue

        if (numeric_count, total_abs) > best_score:
            best_col = str(c)
            best_score = (numeric_count, total_abs)

    return best_col


def _sum_tco2e(df: "pd.DataFrame") -> tuple[float, int, str | None]:
    """
    Returns (total_tco2e, rows_count, used_column).
    """
    col = _find_tco2e_column(df)
    if not col:
        return 0.0, int(len(df.index)) if df is not None else 0, None
    try:
        ser_raw = df[col]
        # Robust numeric parsing for Excel exports (comma decimals, NBSP, etc.)
        try:
            if getattr(ser_raw, "dtype", None) == "object":
                ser_raw = (
                    ser_raw.astype(str)
                    .str.replace("\u00a0", "", regex=False)
                    .str.replace(" ", "", regex=False)
                    .str.replace(",", ".", regex=False)
                )
        except Exception:
            pass
        ser = pd.to_numeric(ser_raw, errors="coerce")
        total = float(ser.fillna(0).sum())
    except Exception:
        total = 0.0
    return total, int(len(df.index)) if df is not None else 0, col


def _parse_float_loose(v) -> float | None:
    if v is None:
        return None
    try:
        if pd.isna(v):
            return None
    except Exception:
        pass
    if isinstance(v, (int, float)):
        try:
            out = float(v)
            if math.isnan(out) or math.isinf(out):
                return None
            return out
        except Exception:
            return None
    try:
        s = str(v).strip()
    except Exception:
        return None
    if not s:
        return None
    s = s.replace("\u00a0", "").replace(" ", "")
    # Extract first numeric token (handles "12.3t", "1,23", "≈ 5.1", etc.)
    import re as _re
    m = _re.search(r"[+-]?\d+(?:[.,]\d+)?(?:e[+-]?\d+)?", s, flags=_re.IGNORECASE)
    if not m:
        return None
    token = m.group(0).replace(",", ".")
    try:
        out = float(token)
        if math.isnan(out) or math.isinf(out):
            return None
        return out
    except Exception:
        return None


def _sum_tco2e_from_xlsx(xlsx_path: str | Path, sheet_name: str | None) -> tuple[float, int, str | None]:
    """
    Sum tCO2e values from an Excel workbook using openpyxl with data_only=True.
    This is more reliable than pandas when cells contain formulas (Excel-visible numbers).
    Returns (total_tco2e, rows_count, used_header_name).
    """
    try:
        p = Path(xlsx_path)
    except Exception:
        p = None
    if not p or not str(p) or not os.path.exists(str(p)):
        return 0.0, 0, None

    try:
        wb = load_workbook(str(p), read_only=True, data_only=True, keep_links=False)
    except Exception:
        return 0.0, 0, None

    target_ws = None
    want = (sheet_name or "").strip().lower()
    if want:
        for n in wb.sheetnames:
            if str(n).strip().lower() == want:
                target_ws = wb[n]
                break
    if target_ws is None and wb.sheetnames:
        target_ws = wb[wb.sheetnames[0]]
    if target_ws is None:
        return 0.0, 0, None

    def norm(x: str) -> str:
        return "".join(ch.lower() for ch in (x or "") if ch.isalnum())

    header_row, headers = _detect_header_row_and_headers(target_ws, max_scan_rows=50)
    candidates: list[tuple[int, str]] = []
    for i, h in enumerate(headers, start=1):
        nh = norm(str(h))
        if nh in (
            "emissionstco2e",
            "emissionstco2etonnes",
            "tco2e",
            "totalemissionstco2e",
            "co2e",
            "co2et",
            "co2etonnes",
        ):
            candidates.append((i, str(h)))
            continue
        if nh.startswith("emissionstco2e") or nh.startswith("co2et") or nh.startswith("co2e"):
            candidates.append((i, str(h)))
            continue
        # Some providers include long headers containing "Calculated CO2e"
        if ("co2e" in nh) and ("calculated" in nh):
            candidates.append((i, str(h)))

    if not candidates:
        return 0.0, 0, None

    rows_count = 0
    max_r = int(getattr(target_ws, "max_row", 1) or 1)
    # First pass: count non-empty rows (for display) + compute per-candidate sums.
    best = {"used": None, "total": 0.0, "numeric_count": 0}
    for r in range(header_row + 1, max_r + 1):
        # Skip fully empty rows to keep counts consistent
        row_has_any = False
        try:
            for c in range(1, min(int(getattr(target_ws, "max_column", 1) or 1), 40) + 1):
                v2 = target_ws.cell(row=r, column=c).value
                if v2 is not None and str(v2).strip() != "":
                    row_has_any = True
                    break
        except Exception:
            row_has_any = True
        if not row_has_any:
            continue

        rows_count += 1
        for col_idx, used in candidates:
            val = target_ws.cell(row=r, column=col_idx).value
            f = _parse_float_loose(val)
            if f is None:
                continue
            # Accumulate into a temporary bucket per candidate by encoding in dict key
            k = f"{col_idx}:{used}"
            # stash totals in best dict dynamically
            # (avoid allocating large per-column arrays)
            tot_key = f"tot::{k}"
            cnt_key = f"cnt::{k}"
            best[tot_key] = float(best.get(tot_key, 0.0)) + float(f)
            best[cnt_key] = int(best.get(cnt_key, 0)) + 1

    # Choose the candidate column with the highest numeric_count; ties broken by total magnitude.
    chosen = None
    chosen_total = 0.0
    chosen_cnt = 0
    for col_idx, used in candidates:
        k = f"{col_idx}:{used}"
        cnt = int(best.get(f"cnt::{k}", 0))
        tot = float(best.get(f"tot::{k}", 0.0))
        if cnt > chosen_cnt or (cnt == chosen_cnt and abs(tot) > abs(chosen_total)):
            chosen = used
            chosen_total = tot
            chosen_cnt = cnt

    if chosen is None:
        return 0.0, int(rows_count), None

    return float(chosen_total), int(rows_count), str(chosen)


def _detect_data_year_from_xlsx(xlsx_path: str | Path, sheet_name: str | None, scan_rows: int = 500) -> int | None:
    """
    Best-effort: infer reporting year from a period column like
    'Reporting period (month, year)' with values like \"Jan-2026\" (legacy \"Jan'-2025\" tolerated).
    Returns the most common year found (or max if tie), else None.
    """
    try:
        p = Path(xlsx_path)
    except Exception:
        return None
    if not p or not os.path.exists(str(p)):
        return None
    try:
        wb = load_workbook(str(p), read_only=True, data_only=True, keep_links=False)
    except Exception:
        return None

    want = (sheet_name or "").strip().lower()
    ws = None
    if want:
        for n in wb.sheetnames:
            if str(n).strip().lower() == want:
                ws = wb[n]
                break
    if ws is None and wb.sheetnames:
        ws = wb[wb.sheetnames[0]]
    if ws is None:
        return None

    def norm(x: str) -> str:
        return "".join(ch.lower() for ch in (x or "") if ch.isalnum())

    header_row, headers = _detect_header_row_and_headers(ws, max_scan_rows=50)
    period_col = None
    for i, h in enumerate(headers, start=1):
        nh = norm(str(h))
        if nh in ("reportingperiodmonthyear", "reportingperiod", "reportingperiodmonth", "monthyear", "periodmonthyear"):
            period_col = i
            break
        if ("period" in nh and "year" in nh) or nh.startswith("reportingperiod"):
            period_col = i
            break
    if period_col is None:
        return None

    import re as _re
    counts: dict[int, int] = {}
    max_r = int(getattr(ws, "max_row", 1) or 1)
    end = min(max_r, header_row + int(scan_rows))
    for r in range(header_row + 1, end + 1):
        v = ws.cell(row=r, column=period_col).value
        if v is None:
            continue
        s = str(v)
        m = _re.search(r"(19|20)\d{2}", s)
        if not m:
            continue
        y = int(m.group(0))
        counts[y] = counts.get(y, 0) + 1
    if not counts:
        return None
    best = sorted(counts.items(), key=lambda kv: (kv[1], kv[0]), reverse=True)[0][0]
    return int(best)


def _parse_period_value(value) -> datetime | None:
    if value is None:
        return None
    try:
        if isinstance(value, pd.Timestamp):
            if pd.isna(value):
                return None
            return datetime(int(value.year), int(value.month), 1)
    except Exception:
        pass
    if isinstance(value, datetime):
        return datetime(value.year, value.month, 1)

    s = str(value).strip()
    if not s:
        return None

    import re as _re
    month_map = {
        "jan": 1, "feb": 2, "mar": 3, "apr": 4, "may": 5, "jun": 6,
        "jul": 7, "aug": 8, "sep": 9, "oct": 10, "nov": 11, "dec": 12,
    }

    # Many uploaded workbooks store "reporting period (month, year)"
    # as the first day of each month, e.g. 01/03/2025 for Mar 2025.
    m = _re.match(r"^\s*(\d{1,2})[/-](\d{1,2})[/-]((?:19|20)\d{2})\s*$", s)
    if m:
        first = int(m.group(1))
        second = int(m.group(2))
        year = int(m.group(3))
        if first == 1 and 1 <= second <= 12:
            return datetime(year, second, 1)
        if second == 1 and 1 <= first <= 12:
            return datetime(year, first, 1)

    m = _re.search(r"(?i)\b(jan|feb|mar|apr|may|jun|jul|aug|sep|oct|nov|dec)[a-z]*[\'\s\-_/,]*(19|20)\d{2}\b", s)
    if m:
        mon = month_map[m.group(1).lower()[:3]]
        year_match = _re.search(r"(19|20)\d{2}", s)
        if year_match:
            return datetime(int(year_match.group(0)), mon, 1)

    m = _re.search(r"\b((?:19|20)\d{2})[-/](0?[1-9]|1[0-2])\b", s)
    if m:
        return datetime(int(m.group(1)), int(m.group(2)), 1)

    m = _re.match(r"^\s*((?:19|20)\d{2})\s*$", s)
    if m:
        return datetime(int(m.group(1)), 1, 1)

    try:
        dt = pd.to_datetime([s], errors="coerce", dayfirst=False)[0]
        if pd.isna(dt):
            return None
        return datetime(int(dt.year), int(dt.month), 1)
    except Exception:
        return None


def _column_name_period_priority(name: str) -> int:
    """Lower is higher priority (reporting > period > date > …)."""
    n = "".join(ch.lower() for ch in (name or "") if ch.isalnum())
    low = (name or "").strip().lower()
    if "reporting period" in low:
        return 1
    if "purchase date" in low:
        return 2
    if "reporting" in n:
        return 1
    if "period" in n:
        return 2
    if "date" in n:
        return 3
    if "month" in n:
        return 4
    if n in ("activitydate", "consumptiondate", "invoicedate"):
        return 5
    if "activity" in n and "date" in n:
        return 5
    if "invoice" in n and "date" in n:
        return 5
    if "consumption" in n and "date" in n:
        return 5
    return 99


def _find_period_column(df: "pd.DataFrame", sample_rows: int = 120) -> str | None:
    if df is None or getattr(df, "columns", None) is None or len(df.columns) == 0:
        return None

    cols = list(df.columns)

    # 0) Data-entry templates: prefer explicit headers used in CTS forms
    for c in cols:
        low = str(c).strip().lower()
        if "reporting period" in low:
            return str(c)
    for c in cols:
        low = str(c).strip().lower()
        if "purchase date" in low:
            return str(c)

    first_col = str(cols[0])

    # 1) First column: if most values parse as reporting dates, use it
    try:
        sample = [v for v in df[first_col].tolist()[:sample_rows] if str(v).strip() != ""]
        if sample:
            parsed = sum(1 for v in sample if _parse_period_value(v) is not None)
            need = max(1, int(len(sample) * 0.45))
            if parsed >= max(need, min(2, len(sample))):
                return first_col
    except Exception:
        pass

    # 2) Named columns: priority reporting → period → date → month → activity/invoice dates
    ranked: list[tuple[tuple[int, float, int], str]] = []
    for c in cols:
        sc = str(c)
        pri = _column_name_period_priority(sc)
        if pri >= 99:
            continue
        try:
            sample = [v for v in df[sc].tolist()[:sample_rows] if str(v).strip() != ""]
        except Exception:
            continue
        if not sample:
            continue
        parsed = sum(1 for v in sample if _parse_period_value(v) is not None)
        ratio = parsed / max(len(sample), 1)
        min_parsed = 1 if len(sample) <= 8 else 2
        min_ratio = 0.2 if len(sample) <= 8 else 0.35
        if pri <= 2:
            min_parsed = 1
            min_ratio = 0.12
        if parsed >= min_parsed and ratio >= min_ratio:
            ranked.append(((pri, -ratio, -parsed), sc))
    if ranked:
        ranked.sort(key=lambda x: x[0])
        return ranked[0][1]

    # 3) Heuristic: any column with strong date-like values
    best_col = None
    best_score = (-1, -1.0)
    for c in cols:
        try:
            sample = [v for v in df[str(c)].tolist()[:sample_rows] if str(v).strip() != ""]
        except Exception:
            continue
        if not sample:
            continue
        parsed = sum(1 for v in sample if _parse_period_value(v) is not None)
        ratio = parsed / max(len(sample), 1)
        min_ok = 1 if len(sample) <= 6 else 3
        rmin = 0.34 if len(sample) <= 6 else 0.55
        if parsed >= min_ok and ratio >= rmin and (parsed, ratio) > best_score:
            best_col = str(c)
            best_score = (parsed, ratio)
    return best_col


def _read_sheet_df_from_workbook(xlsx_path: str | Path, sheet_name: str | None) -> "pd.DataFrame | None":
    try:
        sheets = pd.read_excel(xlsx_path, sheet_name=None, engine="openpyxl")
    except Exception:
        return None

    want = (sheet_name or "").strip().lower()
    if want:
        for k, v in sheets.items():
            if str(k).strip().lower() == want:
                return v

    if len(sheets) == 1:
        return next(iter(sheets.values()))

    if want:
        want_prefix = want[:20]
        for k, v in sheets.items():
            if str(k).strip().lower().startswith(want_prefix):
                return v

    return next(iter(sheets.values())) if sheets else None


def _build_period_profile_from_df(df: "pd.DataFrame") -> dict[str, object]:
    total_tco2e, rows_count, used_col = _sum_tco2e(df)
    if not used_col:
        return {
            "points": [],
            "total": 0.0,
            "rows_count": int(rows_count or 0),
            "min_date": None,
            "max_date": None,
        }

    period_col = _find_period_column(df)
    if not period_col:
        return {
            "points": [],
            "total": float(total_tco2e or 0.0),
            "rows_count": int(rows_count or 0),
            "min_date": None,
            "max_date": None,
        }

    buckets: dict[str, float] = defaultdict(float)
    for _, row in df.iterrows():
        try:
            raw_date = row.get(period_col)
            raw_value = row.get(used_col)
        except Exception:
            continue
        dt = _parse_period_value(raw_date)
        val = _parse_float_loose(raw_value)
        if dt is None or val is None:
            continue
        key = dt.strftime("%Y-%m")
        buckets[key] += float(val)

    if not buckets:
        return {
            "points": [],
            "total": float(total_tco2e or 0.0),
            "rows_count": int(rows_count or 0),
            "min_date": None,
            "max_date": None,
        }

    points = []
    for key in sorted(buckets.keys()):
        year, month = key.split("-")
        dt = datetime(int(year), int(month), 1)
        points.append(
            {
                "key": key,
                "label": dt.strftime("%b %Y"),
                "date": dt,
                "value": round(float(buckets[key]), 6),
            }
        )

    return {
        "points": points,
        "total": round(sum(float(p["value"]) for p in points), 6),
        "rows_count": int(rows_count or 0),
        "min_date": points[0]["date"] if points else None,
        "max_date": points[-1]["date"] if points else None,
    }


def _format_period_label(points: list[dict[str, object]], fallback_dt: datetime | None) -> str:
    if points:
        start = points[0].get("date")
        end = points[-1].get("date")
        if isinstance(start, datetime) and isinstance(end, datetime):
            if start.year == end.year and start.month == end.month:
                return start.strftime("%Y-%m")
            return f"{start.strftime('%Y-%m')} to {end.strftime('%Y-%m')}"
    if isinstance(fallback_dt, datetime):
        return fallback_dt.strftime("%Y-%m-%d")
    return ""


def _build_reporting_rows_from_summary(sub: "MappingRunSummary") -> list[dict[str, object]]:
    """
    Chart rows keyed by reporting period from the mapped workbook (not upload / mapping time).
    Emits sortKey (YYYY-MM) and dateLabel for MonthlyTrend / stacked charts.
    """
    out: list[dict[str, object]] = []
    rid = str(getattr(sub, "run_id", "") or "")
    if not rid:
        return out
    try:
        mr = MappingRun.query.get(rid)
    except Exception:
        mr = None
    if not mr:
        return out
    op = getattr(mr, "output_path", None)
    if not op or not os.path.exists(str(op)):
        return out
    df = _read_sheet_df_from_workbook(op, getattr(sub, "sheet_name", None))
    if df is None or getattr(df, "empty", True):
        return out
    tco2e_col = _find_tco2e_column(df)
    period_col = _find_period_column(df)
    if not tco2e_col or not period_col:
        return out
    sheet_name = str(getattr(sub, "sheet_name", "") or "Category")
    scope = _effective_scope(getattr(sub, "scope", None), sheet_name)
    for _, row in df.iterrows():
        try:
            raw_date = row.get(period_col)
            raw_val = row.get(tco2e_col)
        except Exception:
            continue
        dt = _parse_period_value(raw_date)
        val = _parse_float_loose(raw_val)
        if dt is None or val is None:
            continue
        sraw = str(raw_date).strip() if raw_date is not None else ""
        m_yonly = re.match(r"^\s*((?:19|20)\d{2})(?:\.0)?\s*$", sraw)
        if m_yonly:
            sk = m_yonly.group(1)
            label = m_yonly.group(1)
        else:
            sk = dt.strftime("%Y-%m")
            label = dt.strftime("%b %Y")
        out.append(
            {
                "scope": scope,
                "sheet": sheet_name,
                "category": sheet_name,
                "company": str(getattr(sub, "company_name", "") or ""),
                "emissions": float(val),
                "sortKey": sk,
                "dateLabel": label,
            }
        )
    if not out and tco2e_col and period_col:
        total, _, _ = _sum_tco2e(df)
        if total > 0:
            sk = None
            label = None
            for _, row in df.iterrows():
                try:
                    raw_date = row.get(period_col)
                except Exception:
                    continue
                dt = _parse_period_value(raw_date)
                if dt:
                    sk = dt.strftime("%Y-%m")
                    label = dt.strftime("%b %Y")
                    break
            if not sk:
                ca = getattr(sub, "created_at", None)
                if isinstance(ca, datetime):
                    sk = ca.strftime("%Y-%m")
                    label = ca.strftime("%b %Y")
                else:
                    sk = datetime.utcnow().strftime("%Y-%m")
                    label = sk
            out.append(
                {
                    "scope": scope,
                    "sheet": sheet_name,
                    "category": sheet_name,
                    "company": str(getattr(sub, "company_name", "") or ""),
                    "emissions": float(total),
                    "sortKey": sk,
                    "dateLabel": label,
                }
            )
    return out


def _period_profile_for_summary(sub: "MappingRunSummary", run_cache: dict[str, "MappingRun | None"] | None = None) -> dict[str, object]:
    rid = str(getattr(sub, "run_id", "") or "")
    cache_key = "|".join(
        [
            "v3",
            rid,
            str(getattr(sub, "sheet_name", "") or ""),
            str(getattr(sub, "created_at", "") or ""),
            str(getattr(sub, "tco2e_total", "") or ""),
            str(getattr(sub, "rows_count", "") or ""),
        ]
    )
    with _PERIOD_PROFILE_CACHE_LOCK:
        cached = _PERIOD_PROFILE_CACHE.get(cache_key)
        if cached is not None:
            return cached

    profile: dict[str, object] | None = None
    try:
        mr = None
        if run_cache is not None:
            if rid in run_cache:
                mr = run_cache[rid]
            else:
                mr = MappingRun.query.get(rid)
                run_cache[rid] = mr
        else:
            mr = MappingRun.query.get(rid)
        op = getattr(mr, "output_path", None) if mr else None
        if op and os.path.exists(str(op)):
            df = _read_sheet_df_from_workbook(op, getattr(sub, "sheet_name", None))
            if df is not None and not getattr(df, "empty", True):
                profile = _build_period_profile_from_df(df)
    except Exception:
        profile = None

    if profile is None:
        profile = _fallback_period_profile_for_summary(sub)
    with _PERIOD_PROFILE_CACHE_LOCK:
        _PERIOD_PROFILE_CACHE[cache_key] = profile
    return profile


def _company_candidate_keys(raw_company_name: str) -> list[str]:
    keys: list[str] = []
    raw = (raw_company_name or "").strip()
    if raw:
        keys.append(raw)
    canon, _country = _canonical_company_name_and_country(raw)
    if canon and canon not in keys:
        keys.append(canon)
    try:
        p = _resolve_company_file(canon or raw)
        if p and p.stem and p.stem not in keys:
            keys.append(p.stem)
    except Exception:
        pass
    return keys


def _count_company_schema_sheets(company_name: str) -> int:
    sheets = [s for s in _get_template_company_sheets(company_name) if str(s).strip() != KLARAKARBON_SHEET_NAME]
    return int(len(sheets))


HIDDEN_SCHEMA_SHEETS = {"readme", "company information"}


def _is_hidden_schema_sheet(sheet_name: str) -> bool:
    return (sheet_name or "").strip().lower() in HIDDEN_SCHEMA_SHEETS


def _infer_column_rule(header: str) -> dict[str, str]:
    raw = (header or "").strip()
    low = raw.lower()
    compact = "".join(ch for ch in low if ch.isalnum())

    date_markers = (
        "reportingperiod",
        "release date",
        "attachment date",
        "start date",
        "end date",
        "travel date",
        "date duration",
        "purchase date",
    )
    number_keywords = (
        "spend",
        "amount",
        "consumption",
        "activity volume",
        "activity amount",
        "fuel consumption",
        "distance",
        "km",
        "kwh",
        "mwh",
        "litre",
        "liter",
        "volume",
        "quantity",
        "qty",
        "weight",
        "ton",
        "tonne",
        "hours",
        "days",
        "months",
        "employee count",
        "headcount",
        "ef_value",
        "emissions",
        "tco2e",
        "co2e",
    )
    text_exclusions = ("id", "unit", "currency", "supplier", "source", "description", "name", "country")

    if any(marker in low for marker in date_markers) or compact == "date":
        return {"type": "date", "format": "YYYY-MM-DD", "placeholder": "YYYY-MM-DD", "required": "1"}

    if any(ex in low for ex in text_exclusions):
        return {"type": "text"}

    if any(key in low for key in number_keywords):
        return {"type": "number", "format": "decimal", "placeholder": "0"}

    return {"type": "text"}


def _normalize_date_like(value: str) -> str | None:
    s = (value or "").strip()
    if not s:
        return ""
    try:
        dt = pd.to_datetime([s], errors="coerce", dayfirst=False)[0]
        if pd.isna(dt):
            return None
        return dt.strftime("%Y-%m-%d")
    except Exception:
        return None


def _normalize_number_like(value: str) -> str | None:
    s = (value or "").strip()
    if not s:
        return ""
    f = _parse_float_loose(s)
    if f is None:
        return None
    return str(f)


def _validate_and_normalize_rows(headers: list[str], rows: list[list[str]]) -> tuple[list[list[str]], list[str]]:
    rules = [_infer_column_rule(h) for h in headers]
    out: list[list[str]] = []
    errors: list[str] = []
    for ridx, row in enumerate(rows, start=1):
        normalized: list[str] = []
        for cidx, value in enumerate(row):
            rule = rules[cidx] if cidx < len(rules) else {"type": "text"}
            text = "" if value is None else str(value).strip()
            if rule.get("type") == "date":
                norm = _normalize_date_like(text)
                if norm is None:
                    errors.append(f"Row {ridx}, column '{headers[cidx]}': expected date format YYYY-MM-DD")
                    normalized.append(text)
                else:
                    normalized.append(norm)
            elif rule.get("type") == "number":
                norm = _normalize_number_like(text)
                if norm is None:
                    errors.append(f"Row {ridx}, column '{headers[cidx]}': expected a numeric value")
                    normalized.append(text)
                else:
                    normalized.append(norm)
            else:
                normalized.append(text)
        out.append(normalized)
    return out, errors


def _latest_sheet_totals_for_company(company_keys: list[str]) -> list[MappingRunSummary]:
    """
    For a company (possibly multiple aliases), return the latest summary per sheet.
    """
    if not company_keys:
        return []
    q = (
        MappingRunSummary.query.filter(MappingRunSummary.company_name.in_(company_keys))
        .order_by(MappingRunSummary.created_at.desc())
        .all()
    )
    seen: set[str] = set()
    out: list[MappingRunSummary] = []
    for r in q:
        k = (r.sheet_name or "").strip().lower()
        if not k or k in seen:
            continue
        seen.add(k)
        out.append(r)
    return out


def _backfill_mapping_summaries(max_runs: int = 200) -> None:
    """
    Best-effort backfill for existing MappingRun rows created before summaries existed.
    Safe to call frequently; only fills missing summaries.
    """
    try:
        _ensure_db_tables()
        if not _should_attempt_backfill_mapping_summaries():
            return
        runs = (
            MappingRun.query.filter_by(status="succeeded")
            .order_by(MappingRun.created_at.desc())
            .limit(int(max_runs))
            .all()
        )
        changed = False
        for mr in runs:
            rid = getattr(mr, "id", None)
            if not rid:
                continue
            summ_existing = MappingRunSummary.query.filter_by(run_id=rid).first()
            if summ_existing:
                continue
            p = getattr(mr, "output_path", None)
            if not p or not os.path.exists(str(p)):
                continue
            total_tco2e, rows_count, _used_col = _sum_tco2e_from_xlsx(p, getattr(mr, "sheet_name", None))
            if _used_col is None:
                # Fallback: try pandas if workbook read fails
                try:
                    sheets = pd.read_excel(p, sheet_name=None, engine="openpyxl")
                except Exception:
                    continue
                mapped_df = None
                want = (getattr(mr, "sheet_name", "") or "").strip().lower()
                if want:
                    for k, v in sheets.items():
                        if str(k).strip().lower() == want:
                            mapped_df = v
                            break
                if mapped_df is None:
                    mapped_df = next(iter(sheets.values())) if sheets else None
                if mapped_df is None:
                    continue
                total_tco2e, rows_count, _used_col = _sum_tco2e(mapped_df)
            scope = _infer_scope_from_sheet(getattr(mr, "sheet_name", "") or "")
            if summ_existing:
                summ_existing.company_name = getattr(mr, "company_name", "") or summ_existing.company_name
                summ_existing.sheet_name = getattr(mr, "sheet_name", "") or summ_existing.sheet_name
                summ_existing.scope = scope
                summ_existing.tco2e_total = float(total_tco2e or 0.0)
                summ_existing.rows_count = int(rows_count or 0)
                if not summ_existing.created_at:
                    summ_existing.created_at = getattr(mr, "created_at", None) or datetime.utcnow()
                changed = True
            else:
                summ = MappingRunSummary(
                    run_id=rid,
                    company_name=getattr(mr, "company_name", "") or "",
                    sheet_name=getattr(mr, "sheet_name", "") or "",
                    scope=scope,
                    tco2e_total=float(total_tco2e or 0.0),
                    rows_count=int(rows_count or 0),
                    created_at=getattr(mr, "created_at", None) or datetime.utcnow(),
                )
                db.session.add(summ)
                changed = True
        if changed:
            db.session.commit()
    except Exception:
        try:
            db.session.rollback()
        except Exception:
            pass

@login_manager.user_loader
def load_user(user_id):
    return db.session.get(User, int(user_id))


def _norm_name(s: str) -> str:
    return "".join(ch.lower() for ch in (s or "") if ch.isalnum())


def _list_stage1_input_files() -> list[Path]:
    out: list[Path] = []
    try:
        for p in sorted(STAGE1_INPUT_DIR.glob("*.xls*"), key=lambda x: x.name.lower()):
            if p.name.startswith("~$"):
                continue
            if p.suffix.lower() not in (".xlsx", ".xlsm"):
                continue
            out.append(p)
    except Exception:
        out = []
    return out


def _company_file_map() -> dict[str, Path]:
    """
    Map a stable company key -> input Excel path.
    Key is the filename stem (e.g. "BIMMS" from "BIMMS.xlsx").
    """
    m: dict[str, Path] = {}
    for p in _list_stage1_input_files():
        m[p.stem] = p
    return m


def _company_canonical_file_map() -> dict[str, Path]:
    """
    Map canonical company name -> input Excel path.
    If multiple files map to the same canonical name, keep the first.
    """
    out: dict[str, Path] = {}
    for stem, p in _company_file_map().items():
        canon, _country = _canonical_company_name_and_country(stem)
        key = canon or stem
        if key not in out:
            out[key] = p
    return out


def _resolve_company_file(company_key: str) -> Path | None:
    company_key = (company_key or "").strip()
    if not company_key:
        return None

    # First: canonical map (supports web display names)
    canon, _country = _canonical_company_name_and_country(company_key)
    cm = _company_canonical_file_map()
    if canon and canon in cm:
        return cm[canon]

    m = _company_file_map()
    if company_key in m:
        return m[company_key]

    want = _norm_name(company_key)
    if not want:
        return None

    # Best-effort normalized match (handles spaces/dashes differences)
    for k, p in m.items():
        if _norm_name(k) == want:
            return p

    return None


def _detect_header_row_and_headers(ws, max_scan_rows: int = 20) -> tuple[int, list[str]]:
    """
    Return (header_row_index_1based, headers[]).
    Scans first N rows and picks the earliest row with the highest number of non-empty cells (>=2).
    This helps when sheets have a few instruction rows before the actual header row.
    """
    best_row_idx = 1
    best_values = None
    best_score = -1

    max_r = min(int(getattr(ws, "max_row", 1) or 1), max_scan_rows)
    max_c = int(getattr(ws, "max_column", 1) or 1)
    max_c = min(max_c, 200)  # guard

    for r in range(1, max_r + 1):
        values = []
        score = 0
        last_non_empty = 0
        for c in range(1, max_c + 1):
            v = ws.cell(row=r, column=c).value
            s = "" if v is None else str(v).strip()
            values.append(s)
            if s:
                score += 1
                last_non_empty = c
        if score >= 2 and score > best_score:
            best_score = score
            best_row_idx = r
            best_values = values[:last_non_empty] if last_non_empty else values

    if best_values is None:
        # Fallback: use first row up to last non-empty cell
        r = 1
        last_non_empty = 0
        values = []
        for c in range(1, max_c + 1):
            v = ws.cell(row=r, column=c).value
            s = "" if v is None else str(v).strip()
            values.append(s)
            if s:
                last_non_empty = c
        best_values = values[:last_non_empty] if last_non_empty else values

    headers: list[str] = []
    for i, h in enumerate(best_values, start=1):
        hh = (h or "").strip()
        headers.append(hh if hh else f"Column {i}")

    return best_row_idx, headers


def _read_header_row_raw(ws, header_row: int, max_columns: int = 200) -> list[str]:
    """
    Read header row values as strings, up to the last non-empty cell.
    Does not invent placeholder names.
    """
    max_c = int(getattr(ws, "max_column", 1) or 1)
    max_c = min(max_c, max_columns)
    values: list[str] = []
    last_non_empty = 0
    for c in range(1, max_c + 1):
        v = ws.cell(row=header_row, column=c).value
        s = "" if v is None else str(v).strip()
        values.append(s)
        if s:
            last_non_empty = c
    return values[:last_non_empty] if last_non_empty else values


def _write_schema_only_workbook(source_file: Path, dest_file: Path) -> None:
    """
    Create a schema-only workbook:
    - same sheet names
    - blank sheet contents except the detected header row (written at the same row index)
    This avoids slow row-deletion on very large historical files.
    """
    src_wb = load_workbook(source_file, read_only=True, data_only=True, keep_links=False)
    out_wb = Workbook()
    try:
        out_wb.remove(out_wb.active)
    except Exception:
        pass

    for name in src_wb.sheetnames:
        src_ws = src_wb[name]
        header_row, _headers = _detect_header_row_and_headers(src_ws)
        raw_headers = _read_header_row_raw(src_ws, header_row)
        ws = out_wb.create_sheet(title=name)
        for j, v in enumerate(raw_headers, start=1):
            vv = (v or "").strip()
            ws.cell(row=header_row, column=j).value = vv if vv else None

    out_wb.save(dest_file)


def _ensure_company_backup_and_initialize(company_file: Path) -> None:
    """
    One-time operation per company file:
    - Create a backup of the original (historical 2025 filled) workbook
    - Clear all data rows below detected headers (keeps instruction rows + header row)
    This ensures Stage1 pipeline reads only newly submitted web rows.
    """
    backup_path = STAGE1_INPUT_BACKUP_DIR / company_file.name
    if not backup_path.exists():
        try:
            shutil.copy2(company_file, backup_path)
        except Exception:
            # If backup fails, do not proceed with destructive changes.
            raise

        tmp_path = company_file.with_suffix(".webtmp.xlsx")
        try:
            if tmp_path.exists():
                tmp_path.unlink()
        except Exception:
            pass

        _write_schema_only_workbook(backup_path, tmp_path)
        os.replace(str(tmp_path), str(company_file))


_EF_CACHE: dict[str, object] = {"mtime_ns": None, "rows": None, "sheets": None, "scopes": None, "sources": None}
_SCHEMA_CACHE_LOCK = threading.Lock()
_SCHEMA_CACHE: dict[tuple[str, int | None], dict[str, object]] = {}
_BACKFILL_STATE_LOCK = threading.Lock()
_BACKFILL_STATE: dict[str, float] = {"last_attempt_at": 0.0}
_PERIOD_PROFILE_CACHE_LOCK = threading.Lock()
_PERIOD_PROFILE_CACHE: dict[str, dict[str, object]] = {}


def _schema_cache_key(company_file: Path) -> tuple[str, int | None]:
    try:
        st = company_file.stat()
        mtime_ns = getattr(st, "st_mtime_ns", None) or int(st.st_mtime * 1_000_000_000)
    except Exception:
        mtime_ns = None
    return (str(company_file.resolve()), mtime_ns)


def _invalidate_schema_cache(company_file: Path | None = None) -> None:
    global _SCHEMA_CACHE
    with _SCHEMA_CACHE_LOCK:
        if company_file is None:
            _SCHEMA_CACHE = {}
            return
        target = str(company_file.resolve())
        _SCHEMA_CACHE = {k: v for k, v in _SCHEMA_CACHE.items() if k[0] != target}


def _get_schema_cache_entry(company_file: Path) -> dict[str, object]:
    key = _schema_cache_key(company_file)
    with _SCHEMA_CACHE_LOCK:
        cached = _SCHEMA_CACHE.get(key)
        if cached is None:
            cached = {"sheets": None, "headers": {}}
            _SCHEMA_CACHE[key] = cached
        return cached


def _get_visible_sheet_names(company_file: Path) -> list[str]:
    cache_entry = _get_schema_cache_entry(company_file)
    cached_sheets = cache_entry.get("sheets")
    if isinstance(cached_sheets, list):
        return cached_sheets

    wb = load_workbook(company_file, read_only=True, data_only=True, keep_links=False)
    sheets = [s for s in list(wb.sheetnames) if not _is_hidden_schema_sheet(s)]
    cache_entry["sheets"] = sheets
    return sheets


def _get_sheet_headers_and_rules(company_file: Path, sheet: str) -> tuple[int, list[str], dict[str, str]]:
    cache_entry = _get_schema_cache_entry(company_file)
    headers_cache = cache_entry.setdefault("headers", {})
    if isinstance(headers_cache, dict) and sheet in headers_cache:
        cached = headers_cache[sheet]
        if isinstance(cached, tuple) and len(cached) == 3:
            return cached  # type: ignore[return-value]

    wb = load_workbook(company_file, read_only=True, data_only=True, keep_links=False)
    if sheet not in wb.sheetnames:
        raise KeyError(sheet)
    ws = wb[sheet]
    header_row, headers = _detect_header_row_and_headers(ws)
    rules = {h: _infer_column_rule(h) for h in headers}
    result = (header_row, headers, rules)
    if isinstance(headers_cache, dict):
        headers_cache[sheet] = result
    return result


def _should_attempt_backfill_mapping_summaries(ttl_seconds: int = 60) -> bool:
    now = time.time()
    with _BACKFILL_STATE_LOCK:
        last_attempt_at = float(_BACKFILL_STATE.get("last_attempt_at", 0.0) or 0.0)
        if (now - last_attempt_at) < float(ttl_seconds):
            return False
        _BACKFILL_STATE["last_attempt_at"] = now

    try:
        succeeded_count = MappingRun.query.filter_by(status="succeeded").count()
        summary_count = MappingRunSummary.query.count()
        return int(summary_count) < int(succeeded_count)
    except Exception:
        return False


def _fallback_period_profile_for_summary(sub: "MappingRunSummary") -> dict[str, object]:
    fallback_dt = getattr(sub, "created_at", None)
    fallback_point = None
    if isinstance(fallback_dt, datetime):
        fallback_point = {
            "key": fallback_dt.strftime("%Y-%m"),
            "label": fallback_dt.strftime("%b %Y"),
            "date": datetime(fallback_dt.year, fallback_dt.month, 1),
            "value": round(float(getattr(sub, "tco2e_total", 0.0) or 0.0), 6),
        }

    return {
        "points": [fallback_point] if fallback_point else [],
        "total": round(float(getattr(sub, "tco2e_total", 0.0) or 0.0), 6),
        "rows_count": int(getattr(sub, "rows_count", 0) or 0),
        "min_date": fallback_point["date"] if fallback_point else None,
        "max_date": fallback_point["date"] if fallback_point else None,
    }


def _load_stage2_emission_factors() -> dict[str, object]:
    """
    Load mapping emission factors from the Stage2 Excel file (each sheet = category).
    Returns a dict with:
      - rows: list[dict] where keys match the mapping headers
      - sheets: list[str]
      - scopes: list[str]
      - sources: list[str]
    """
    global _EF_CACHE
    try:
        st = STAGE2_EF_XLSX.stat()
        mtime_ns = getattr(st, "st_mtime_ns", None) or int(st.st_mtime * 1_000_000_000)
    except Exception:
        mtime_ns = None

    if _EF_CACHE.get("rows") is not None and _EF_CACHE.get("mtime_ns") == mtime_ns:
        return _EF_CACHE  # type: ignore[return-value]

    rows: list[dict[str, object]] = []
    sheets: list[str] = []
    scopes_set: set[str] = set()
    sources_set: set[str] = set()

    if not STAGE2_EF_XLSX.exists():
        _EF_CACHE = {"mtime_ns": mtime_ns, "rows": [], "sheets": [], "scopes": [], "sources": []}
        return _EF_CACHE  # type: ignore[return-value]

    wb = load_workbook(STAGE2_EF_XLSX, read_only=True, data_only=True, keep_links=False)
    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        sheets.append(sheet_name)

        # Read headers from first row (fast in read_only mode)
        first = next(ws.iter_rows(min_row=1, max_row=1, values_only=True), None)
        headers = [("" if v is None else str(v).strip()) for v in (first or [])]
        while headers and headers[-1] == "":
            headers.pop()

        if not headers:
            continue

        # Expect the mapping headers; tolerate extra columns
        wanted = {
            "ef_name",
            "ef_description",
            "scope",
            "ef_category",
            "ef_id",
            "ef_value",
            "ef_unit",
            "ef_source",
            "Emission Factor Category",
        }

        col_idx = {h: i for i, h in enumerate(headers) if h}

        # Stream rows efficiently
        for row_idx, row_values in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
            values = list(row_values[: len(headers)]) if row_values else []
            if not values:
                continue

            any_val = False
            for v in values:
                if v is None:
                    continue
                if str(v).strip() != "":
                    any_val = True
                    break
            if not any_val:
                continue

            row: dict[str, object] = {"sheet": sheet_name, "_row": row_idx}
            for h, idx in col_idx.items():
                if h in wanted:
                    row[h] = values[idx] if idx < len(values) else None

            # Normalize some fields to strings for filtering/search
            for k in ("ef_name", "ef_description", "scope", "ef_category", "ef_id", "ef_unit", "ef_source", "Emission Factor Category"):
                if k in row and row[k] is not None:
                    row[k] = str(row[k]).strip()

            if row.get("scope"):
                scopes_set.add(str(row["scope"]))
            if row.get("ef_source"):
                sources_set.add(str(row["ef_source"]))

            rows.append(row)

    scopes = sorted([s for s in scopes_set if s], key=lambda x: x.lower())
    sources = sorted([s for s in sources_set if s], key=lambda x: x.lower())
    sheets_sorted = sorted([s for s in sheets if s], key=lambda x: x.lower())

    _EF_CACHE = {"mtime_ns": mtime_ns, "rows": rows, "sheets": sheets_sorted, "scopes": scopes, "sources": sources}
    return _EF_CACHE  # type: ignore[return-value]


def list_pipeline_templates_for_user(user_company: str, is_admin: bool):
    files = []
    try:
        for p in sorted(PIPELINE_TEMPLATE_DIR.glob("*.xls*"), key=lambda x: x.name.lower()):
            if p.name.startswith("~$"):
                continue
            files.append(p)
    except Exception:
        files = []

    if is_admin:
        return files

    want = _norm_name(user_company)
    if not want:
        return []
    # Match by filename containing normalized company name (best-effort)
    out = []
    for p in files:
        if want in _norm_name(p.stem):
            out.append(p)
    return out


def _safe_user_can_access_run(run: "PipelineRun") -> bool:
    return bool(current_user.is_admin) or (run.user_id == current_user.id)


def _run_pipeline_background(run_id: int) -> None:
    with app.app_context():
        run = PipelineRun.query.get(run_id)
        if not run:
            return

        run.status = "running"
        db.session.commit()

        run_dir = Path(run.run_dir)
        input_dir = Path(run.input_dir)
        log_path = run_dir / "pipeline.log"
        run.log_path = str(log_path)
        db.session.commit()

        stage2_out_dir = PROJECT_ROOT / "engine" / "stage2_mapping" / "output"
        before = set()
        try:
            if stage2_out_dir.exists():
                before = {p.resolve() for p in stage2_out_dir.glob("*.xlsx")}
        except Exception:
            before = set()

        cmd = [
            sys.executable,
            str(PROJECT_ROOT / "run_pipeline.py"),
            "all",
            "--stage1-input-folder",
            str(input_dir),
            "--stage1-work-dir",
            str(run_dir / "stage1"),
        ]

        rc = 1
        try:
            run_dir.mkdir(parents=True, exist_ok=True)
            with open(log_path, "w", encoding="utf-8") as f:
                f.write("CMD: " + " ".join(cmd) + "\n")
                f.write("START: " + datetime.now().isoformat() + "\n\n")
                f.flush()
                proc = subprocess.run(
                    cmd,
                    cwd=str(PROJECT_ROOT),
                    stdout=f,
                    stderr=subprocess.STDOUT,
                    text=True,
                    timeout=_SUBPROCESS_TIMEOUT_SECONDS,
                    env={
                        **os.environ,
                        "PYTHONUTF8": os.environ.get("PYTHONUTF8", "1"),
                        "PYTHONIOENCODING": os.environ.get("PYTHONIOENCODING", "utf-8"),
                        "PYTHONUNBUFFERED": os.environ.get("PYTHONUNBUFFERED", "1"),
                    },
                )
                rc = int(proc.returncode)
        except Exception as exc:
            run.status = "failed"
            run.exit_code = -1
            run.error_message = str(exc)
            db.session.commit()
            return

        run.exit_code = rc

        # Stage1 output: last chained file (if present)
        try:
            stage1_final = run_dir / "stage1" / "stage1_05_translated.xlsx"
            if stage1_final.exists():
                run.stage1_output = str(stage1_final)
        except Exception:
            pass

        # Collect new Stage2 outputs and copy into run dir
        try:
            after = set()
            if stage2_out_dir.exists():
                after = {p.resolve() for p in stage2_out_dir.glob("*.xlsx")}
            new_files = sorted([p for p in after if p not in before], key=lambda p: p.stat().st_mtime)
            if new_files:
                dst = run_dir / "stage2_output"
                dst.mkdir(parents=True, exist_ok=True)
                for p in new_files:
                    shutil.copy2(p, dst / p.name)
                run.stage2_output_dir = str(dst)
        except Exception:
            pass

        run.status = "succeeded" if rc == 0 else "failed"
        db.session.commit()

# Registration dropdown: canonical company names only
COMPANIES = sorted(_COMPANY_COUNTRY_CANONICAL.keys())

# Utility: Calculate emissions from excel data (loads factors from the database)
def calculate_emissions_from_excel(file_path, template_name):
    from sqlalchemy import and_
    total_emission = 0
    by_category = {}
    try:
        xls = pd.ExcelFile(file_path)
        # Check whether an 'Activity Based' sheet exists first
        if 'Activity Based' in xls.sheet_names:
            sheet = 'Activity Based'
        else:
            sheet = 0  # ilk sheet
        header_row = 0
        if template_name == 'Scope 1 Fuel Usage (Mobile Combustion)':
            header_row = 4
        elif template_name == 'Scope 2 Electricity':
            header_row = 0
        elif template_name == 'Scope 2 District Heating':
            header_row = 0
        elif template_name == 'Scope 3 Category 7 Employee Commuting':
            header_row = 0
        elif template_name.startswith('Scope 3'):
            header_row = 0
        elif template_name == 'Water Tracker':
            header_row = 0
        else:
            header_row = 0
        df = xls.parse(sheet, header=header_row)
        if not isinstance(df, pd.DataFrame):
            return {
                'total_emission': 0,
                'by_category': {}
            }
        df.columns = [str(c).strip().lower() for c in df.columns]
        # Scope 1 Fuel Usage
        if template_name == 'Scope 1 Fuel Usage (Mobile Combustion)':
            fuel_col = next((c for c in df.columns if 'fuel' in c), None)
            litre_col = next((c for c in df.columns if 'litre' in c), None)
            if fuel_col and litre_col:
                for _, row in df.iterrows():
                    fuel = str(row[fuel_col]).strip()
                    try:
                        litres = float(row[litre_col])
                    except (ValueError, TypeError):
                        litres = 0
                    # Load factor from database
                    factor_obj = EmissionFactor.query.filter_by(category=template_name, subcategory=fuel).first()
                    factor = factor_obj.factor if factor_obj else 0
                    emission = litres * factor
                    if fuel not in by_category:
                        by_category[fuel] = 0
                    by_category[fuel] += emission
                    total_emission += emission
        # Scope 2 Electricity
        elif template_name == 'Scope 2 Electricity':
            cons_col = next((c for c in df.columns if 'consumption' in c or 'kwh' in c), None)
            factor_col = next((c for c in df.columns if 'emission factor' in c), None)
            if cons_col:
                for _, row in df.iterrows():
                    try:
                        cons = float(row[cons_col])
                    except (ValueError, TypeError):
                        cons = 0
                    if factor_col and row.get(factor_col):
                        factor = float(row[factor_col])
                    else:
                        factor_obj = EmissionFactor.query.filter_by(category=template_name, subcategory='Electricity').first()
                        factor = factor_obj.factor if factor_obj else 0
                    emission = cons * factor
                    by_category['Electricity'] = by_category.get('Electricity', 0) + emission
                    total_emission += emission
        # Scope 2 District Heating
        elif template_name == 'Scope 2 District Heating':
            cons_col = next((c for c in df.columns if 'consumption' in c), None)
            factor_col = next((c for c in df.columns if 'emission factor' in c), None)
            if cons_col:
                for _, row in df.iterrows():
                    try:
                        cons = float(row[cons_col])
                    except (ValueError, TypeError):
                        cons = 0
                    if factor_col and row.get(factor_col):
                        factor = float(row[factor_col])
                    else:
                        factor_obj = EmissionFactor.query.filter_by(category=template_name, subcategory='District Heating').first()
                        factor = factor_obj.factor if factor_obj else 0
                    emission = cons * factor
                    by_category['District Heating'] = by_category.get('District Heating', 0) + emission
                    total_emission += emission
        # Scope 3 Category 7 Employee Commuting
        elif template_name == 'Scope 3 Category 7 Employee Commuting':
            mode_col = next((c for c in df.columns if 'mode of transport' in c), None)
            km_col = next((c for c in df.columns if 'km travelled per month' in c), None)
            if mode_col and km_col:
                for _, row in df.iterrows():
                    mode = str(row[mode_col]).strip().lower()
                    try:
                        km = float(row[km_col])
                    except (ValueError, TypeError):
                        km = 0
                    factor_obj = EmissionFactor.query.filter_by(category=template_name, subcategory=mode).first()
                    factor = factor_obj.factor if factor_obj else 0
                    emission = km * factor
                    if mode not in by_category:
                        by_category[mode] = 0
                    by_category[mode] += emission
                    total_emission += emission
        # Scope 3 (Spend-based veya Activity-based)
        elif template_name.startswith('Scope 3'):
            spend_col = next((c for c in df.columns if 'spend' in c), None)
            factor_col = next((c for c in df.columns if 'emission factor' in c), None)
            cons_col = next((c for c in df.columns if 'consumption' in c or 'amount' in c), None)
            if spend_col and factor_col:
                for _, row in df.iterrows():
                    try:
                        spend = float(row[spend_col])
                        factor = float(row[factor_col])
                        emission = spend * factor
                        by_category['Spend-based'] = by_category.get('Spend-based', 0) + emission
                        total_emission += emission
                    except (ValueError, TypeError):
                        continue
            elif cons_col and factor_col:
                for _, row in df.iterrows():
                    try:
                        cons = float(row[cons_col])
                        factor = float(row[factor_col])
                        emission = cons * factor
                        by_category['Activity-based'] = by_category.get('Activity-based', 0) + emission
                        total_emission += emission
                    except (ValueError, TypeError):
                        continue
        # Water Tracker
        elif template_name == 'Water Tracker':
            cons_col = next((c for c in df.columns if 'consumption' in c or 'amount' in c), None)
            factor_col = next((c for c in df.columns if 'emission factor' in c), None)
            if cons_col:
                for _, row in df.iterrows():
                    try:
                        cons = float(row[cons_col])
                    except (ValueError, TypeError):
                        cons = 0
                    if factor_col and row.get(factor_col):
                        factor = float(row[factor_col])
                    else:
                        factor_obj = EmissionFactor.query.filter_by(category=template_name, subcategory='Water').first()
                        factor = factor_obj.factor if factor_obj else 0
                    emission = cons * factor
                    by_category['Water'] = by_category.get('Water', 0) + emission
                    total_emission += emission
        else:
            # Dummy value for other templates
            return {
                'total_emission': 1234.56,
                'by_category': {
                    'Dummy': 1234.56
                }
            }
    except Exception as e:
        print(f'Excel emission calculation error: {e}')
        return {'total_emission': 0, 'by_category': {}}
    return {'total_emission': round(total_emission, 2), 'by_category': {k: round(v, 2) for k, v in by_category.items()}}

def _safe_float(v) -> float:
    try:
        out = float(v or 0)
        if math.isnan(out) or math.isinf(out):
            return 0.0
        return out
    except Exception:
        return 0.0

def _infer_scopes(template_name: str, total_emission: float) -> tuple[float, float, float]:
    """Infer scope1/2/3 totals for a single submission based on template name."""
    name = (template_name or "").lower()
    s1 = s2 = s3 = 0.0
    if "scope 1" in name:
        s1 = total_emission
    elif "scope 2" in name:
        s2 = total_emission
    elif "scope 3" in name:
        s3 = total_emission
    return s1, s2, s3

def _parse_date_yyyy_mm_dd(s: str | None) -> datetime | None:
    s = (s or "").strip()
    if not s:
        return None
    try:
        return datetime.strptime(s, "%Y-%m-%d")
    except Exception:
        return None

def get_emission_factor_name(extra_data_json):
    if not extra_data_json or extra_data_json == 'null':
        return ''
    try:
        data = json.loads(extra_data_json) if isinstance(extra_data_json, str) else extra_data_json
        for key in data:
            if key.strip().lower() == 'item prettyid':
                return data[key]
        return ''
    except Exception:
        return ''

app.jinja_env.filters['get_emission_factor_name'] = get_emission_factor_name

@app.route('/')
def index():
    return render_template('index.html')

@app.route('/who-we-are')
def who_we_are():
    return render_template('who_we_are.html')

@app.route('/platform')
def platform():
    return render_template('platform.html')


@app.route('/locations')
def locations():
    t0 = time.perf_counter()
    rv = render_template(
        "locations.html",
        mapbox_token=(os.getenv("MAPBOX_TOKEN") or "").strip(),
        OPENWEATHER_API_KEY=OPENWEATHER_API_KEY,
    )
    _perf_log("locations", page_render_ms=(time.perf_counter() - t0) * 1000.0)
    return rv

def _render_methodology_scope3_category(category_num, category_title, category_slug):
    chart_payload = _scope3_category_charts_payload(int(category_num))
    return render_template(
        "methodology_scope3_category.html",
        category_slug=category_slug,
        category_num=category_num,
        category_title=category_title,
        chart_payload=chart_payload,
    )


@app.route("/methodology")
def methodology():
    return render_template("methodology.html")


@app.route("/methodology/reference")
def methodology_reference():
    return render_template("methodology_reference.html")


@app.route("/methodology/scope1")
def methodology_scope1():
    return render_template("methodology_scope1.html")


@app.route("/methodology/scope2")
def methodology_scope2():
    return render_template("methodology_scope2.html")


@app.route("/methodology/scope3")
def methodology_scope3():
    return render_template("methodology_scope3.html")


@app.route("/methodology/scope3/category1")
def methodology_scope3_category1():
    return _render_methodology_scope3_category(
        1, "Purchased Goods and Services", "category1"
    )


@app.route("/methodology/scope3/category2")
def methodology_scope3_category2():
    return _render_methodology_scope3_category(2, "Capital Goods", "category2")


@app.route("/methodology/scope3/category3")
def methodology_scope3_category3():
    return _render_methodology_scope3_category(
        3, "Fuel and Energy Related Activities", "category3"
    )


@app.route("/methodology/scope3/category4")
def methodology_scope3_category4():
    return _render_methodology_scope3_category(
        4, "Upstream Transportation", "category4"
    )


@app.route("/methodology/scope3/category5")
def methodology_scope3_category5():
    return _render_methodology_scope3_category(5, "Waste", "category5")


@app.route("/methodology/scope3/category6")
def methodology_scope3_category6():
    return _render_methodology_scope3_category(
        6, "Business Travel", "category6"
    )


@app.route("/methodology/scope3/category7")
def methodology_scope3_category7():
    return _render_methodology_scope3_category(
        7, "Employee Commuting", "category7"
    )


@app.route("/methodology/scope3/category9")
def methodology_scope3_category9():
    return _render_methodology_scope3_category(
        9, "Downstream Transportation", "category9"
    )


@app.route("/methodology/scope3/category10")
def methodology_scope3_category10():
    return render_template("methodology_scope3_category10.html")


@app.route("/methodology/scope3/category11")
def methodology_scope3_category11():
    return render_template("methodology_scope3_category11.html")


@app.route("/methodology/scope3/category12")
def methodology_scope3_category12():
    return render_template("methodology_scope3_category12.html")


@app.route("/methodology/scope3/category13")
def methodology_scope3_category13():
    return render_template("methodology_scope3_category13.html")


@app.route("/methodology/scope3/category14")
def methodology_scope3_category14():
    return render_template("methodology_scope3_category14.html")


@app.route("/methodology/scope3/category15")
def methodology_scope3_category15():
    return render_template("methodology_scope3_category15.html")


@app.route("/methodology/scope3/category16")
def methodology_scope3_category16():
    return render_template("methodology_scope3_category16.html")


@app.route('/trust')
def trust():
    return render_template('trust.html')


@app.route("/privacy-policy")
def legal_privacy():
    return render_template("legal_privacy.html")


@app.route("/terms-of-service")
def legal_terms():
    return render_template("legal_terms.html")


@app.route("/cookie-policy")
def legal_cookies():
    return render_template("legal_cookies.html")


@app.route('/impact')
def impact():
    return render_template('impact.html')


@app.route('/impact/approach')
def impact_approach():
    return render_template('impact_approach.html')


@app.route('/impact/operations')
def impact_operations():
    return render_template('impact_operations.html')


@app.route('/impact/lca-epd')
def impact_lca_epd():
    return render_template('impact_lca_epd.html')


@app.route('/impact/collaboration')
def impact_collaboration():
    return render_template('impact_collaboration.html')


@app.route('/impact/corporate')
def impact_corporate():
    return render_template('impact_corporate.html')


@app.route('/impact/materiality')
def impact_materiality():
    return render_template('impact_materiality.html')


@app.route('/impact/carbon')
def impact_carbon():
    return render_template('impact_carbon.html')


@app.route('/impact/stakeholders')
def impact_stakeholders():
    return render_template('impact_stakeholders.html')


@app.route('/impact/sdgs')
def impact_sdgs():
    return render_template('impact_sdgs.html')


@app.route('/impact/governance')
def impact_governance():
    return render_template('impact_governance.html')


@app.route('/impact/reporting')
def impact_reporting():
    return render_template('impact_reporting.html')


@app.route('/esg', endpoint='esg')
def esg():
    return render_template('esg.html')


@app.route('/esg/csrd', endpoint='csrd')
def csrd():
    return render_template('csrd_overview.html')


@app.route('/lca')
def lca():
    return render_template('lca.html')


@app.route('/lca/tool')
def lca_tool():
    return render_template('lca_tool.html')


@app.route('/lca-tool')
def lca_tool_legacy_redirect():
    return redirect(url_for('lca_tool'), 301)


@app.route('/csrd')
def csrd_legacy_redirect():
    return redirect(url_for('csrd_policies'), 301)


@app.route('/esg/csrd/policies', endpoint='csrd_policies')
@login_required
def csrd_policies():
    _ensure_db_tables()
    edit_id = request.args.get('edit', type=int)
    edit_policy = None
    if edit_id and current_user.is_admin:
        edit_policy = CsrdPolicy.query.get(edit_id)
    policies = CsrdPolicy.query.order_by(CsrdPolicy.created_at.desc()).all()
    return render_template(
        'csrd_policies.html',
        policies=policies,
        edit_policy=edit_policy,
    )


@app.route('/esg/csrd/policies/add', methods=['POST'])
@app.route('/csrd/add', methods=['POST'])
@login_required
def csrd_policy_add():
    if not current_user.is_admin:
        flash('Access denied')
        return redirect(url_for('csrd_policies'))
    _ensure_db_tables()
    title = (request.form.get('title') or '').strip()
    short_description = (request.form.get('short_description') or '').strip()
    upload = request.files.get('file')
    if not title or not short_description:
        flash('Title and description are required.')
        return redirect(url_for('csrd_policies'))
    if not upload or not getattr(upload, 'filename', None):
        flash('Please upload a PDF file.')
        return redirect(url_for('csrd_policies'))
    if not _csrd_filename_is_pdf(upload.filename):
        flash('Only PDF files are allowed.')
        return redirect(url_for('csrd_policies'))
    upload_dir = _csrd_policy_upload_dir()
    base = secure_filename(upload.filename) or 'policy.pdf'
    if not base.lower().endswith('.pdf'):
        base = f'{base}.pdf'
    unique_name = f'{uuid.uuid4().hex}_{base}'
    dest = upload_dir / unique_name
    upload.save(str(dest))
    rel = f'csrd_policies/{unique_name}'
    row = CsrdPolicy(title=title, short_description=short_description, file_relpath=rel)
    db.session.add(row)
    db.session.commit()
    flash('Policy saved.')
    return redirect(url_for('csrd_policies'))


@app.route('/esg/csrd/policies/<int:policy_id>/update', methods=['POST'])
@app.route('/csrd/<int:policy_id>/update', methods=['POST'])
@login_required
def csrd_policy_update(policy_id: int):
    if not current_user.is_admin:
        flash('Access denied')
        return redirect(url_for('csrd_policies'))
    _ensure_db_tables()
    row = CsrdPolicy.query.get_or_404(policy_id)
    title = (request.form.get('title') or '').strip()
    short_description = (request.form.get('short_description') or '').strip()
    if not title or not short_description:
        flash('Title and description are required.')
        return redirect(url_for('csrd_policies', edit=policy_id))
    row.title = title
    row.short_description = short_description
    upload = request.files.get('file')
    if upload and getattr(upload, 'filename', None):
        if not _csrd_filename_is_pdf(upload.filename):
            flash('Only PDF files are allowed.')
            return redirect(url_for('csrd_policies', edit=policy_id))
        old_path = FRONTEND_UPLOAD_DIR / row.file_relpath
        try:
            if old_path.is_file():
                old_path.unlink()
        except Exception:
            pass
        upload_dir = _csrd_policy_upload_dir()
        base = secure_filename(upload.filename) or 'policy.pdf'
        if not base.lower().endswith('.pdf'):
            base = f'{base}.pdf'
        unique_name = f'{uuid.uuid4().hex}_{base}'
        dest = upload_dir / unique_name
        upload.save(str(dest))
        row.file_relpath = f'csrd_policies/{unique_name}'
    row.updated_at = datetime.utcnow()
    db.session.commit()
    flash('Policy updated.')
    return redirect(url_for('csrd_policies'))


@app.route('/esg/csrd/policies/<int:policy_id>/delete', methods=['POST'])
@app.route('/csrd/<int:policy_id>/delete', methods=['POST'])
@login_required
def csrd_policy_delete(policy_id: int):
    if not current_user.is_admin:
        flash('Access denied')
        return redirect(url_for('csrd_policies'))
    _ensure_db_tables()
    row = CsrdPolicy.query.get_or_404(policy_id)
    disk_path = FRONTEND_UPLOAD_DIR / row.file_relpath
    db.session.delete(row)
    db.session.commit()
    try:
        if disk_path.is_file():
            disk_path.unlink()
    except Exception:
        pass
    flash('Policy removed.')
    return redirect(url_for('csrd_policies'))


@app.route('/esg/csrd/policies/<int:policy_id>/file')
@app.route('/csrd/<int:policy_id>/file')
@login_required
def csrd_policy_file(policy_id: int):
    _ensure_db_tables()
    row = CsrdPolicy.query.get_or_404(policy_id)
    path = FRONTEND_UPLOAD_DIR / row.file_relpath
    if not path.is_file():
        return ('File not found', 404)
    dl_name = Path(row.file_relpath).name
    return send_file(str(path), mimetype='application/pdf', as_attachment=False, download_name=dl_name)


@app.route('/assets/mouse-symbol')
def mouse_symbol_image():
    symbol_path = APP_DIR / "images" / "Symbol for mouse.png"
    if not symbol_path.exists():
        return ("Mouse symbol image not found.", 404)
    return send_file(str(symbol_path), mimetype="image/png")

@app.route('/login', methods=['GET', 'POST'])
def login():
    if request.method == 'POST':
        email = (request.form.get('email') or '').strip().lower()
        password = request.form.get('password') or ''
        if not _email_domain_allowed(email):
            flash('Only CTS company email addresses are allowed')
            return render_template('login.html')

        user = User.query.filter(db.func.lower(User.email) == email).first()
        if user and check_password_hash(user.password_hash, password):
            login_user(user)
            session["activity_session_id"] = uuid.uuid4().hex
            _persist_template_mode(getattr(user, "template_mode", None))
            request.environ["skip_activity_log"] = True
            _write_activity_log_for_user(user, action="login")
            if not _user_profile_complete(user):
                return redirect(url_for('profile_setup'))
            return redirect(url_for('feed'))
        flash('Invalid email or password')

    return render_template('login.html')

@app.route("/register", methods=["GET", "POST"])
def register():
    """Public self-registration is disabled; use the access request flow."""
    return redirect(url_for("request_access"))


@app.route("/request-access", methods=["GET", "POST"])
def request_access():
    if current_user.is_authenticated:
        return redirect(url_for("home"))
    _ensure_db_tables()
    if request.method == "POST":
        full_name = (request.form.get("full_name") or "").strip()
        email = (request.form.get("email") or "").strip().lower()
        company = (request.form.get("company") or "").strip()
        reason = (request.form.get("reason") or "").strip()

        if not full_name or not email or not company or not reason:
            flash("Please fill in all fields.")
            return render_template("request_access.html")

        if not _email_domain_allowed(email):
            flash("Only CTS company email addresses are allowed (@cts-nordics.com).")
            return render_template("request_access.html")

        if User.query.filter(db.func.lower(User.email) == email).first():
            flash("This email is already registered. Please sign in or use Forgot password.")
            return render_template("request_access.html")

        existing_pending = (
            AccessRequest.query.filter(db.func.lower(AccessRequest.email) == email)
            .filter(AccessRequest.status == "pending")
            .first()
        )
        if existing_pending:
            flash("You already have a pending access request. We will contact you when it is reviewed.")
            return render_template("request_access.html")

        row = AccessRequest(
            full_name=full_name,
            email=email,
            company=company,
            reason=reason,
            status="pending",
        )
        db.session.add(row)
        db.session.commit()

        admin_body = (
            "New access request received.\n\n"
            f"Name:\n{full_name}\n\n"
            f"Company:\n{company}\n\n"
            f"Email:\n{email}\n\n"
            f"Reason:\n{reason}\n\n"
            "Review request in admin panel.\n"
        )
        _send_plain_email(
            ACCESS_REQUEST_NOTIFY_EMAIL,
            "New Access Request - CTS Platform",
            admin_body,
        )

        flash("Thank you. Your request has been submitted. We will notify you when your access is approved.")
        return redirect(url_for("request_access"))

    return render_template("request_access.html")


@app.route("/admin/access-requests", methods=["GET", "POST"])
@login_required
def admin_access_requests():
    if not getattr(current_user, "is_admin", False):
        flash("Access denied")
        return redirect(url_for("dashboard"))

    _ensure_db_tables()

    if request.method == "POST":
        action = (request.form.get("action") or "").strip()
        rid = request.form.get("request_id", type=int)
        if action != "approve" or not rid:
            flash("Invalid request.")
            return redirect(url_for("admin_access_requests"))

        row = db.session.get(AccessRequest, rid)
        if not row:
            flash("Request not found.")
            return redirect(url_for("admin_access_requests"))

        if row.status != "pending":
            flash("This request is no longer pending.")
            return redirect(url_for("admin_access_requests"))

        email = (row.email or "").strip().lower()
        if User.query.filter(db.func.lower(User.email) == email).first():
            flash("A user with this email already exists. Remove or reject the request manually if needed.")
            return redirect(url_for("admin_access_requests"))

        password = request.form.get("password") or ""
        if len(password) < 8:
            flash("Password must be at least 8 characters.")
            return redirect(url_for("admin_access_requests"))

        new_role = normalize_user_role(request.form.get("role"))
        if new_role not in ACCESS_REQUEST_APPROVE_ROLES_SET:
            flash("Invalid role selected.")
            return redirect(url_for("admin_access_requests"))

        first_name, last_name = _split_full_name(row.full_name)
        company_name = (row.company or "").strip() or "Unknown"

        user = User(
            email=email,
            password_hash=generate_password_hash(password),
            company_name=company_name,
            first_name=first_name or None,
            last_name=last_name or None,
            is_profile_complete=False,
            role=new_role,
        )
        sync_user_admin_flag(user)
        db.session.add(user)

        row.status = "approved"
        db.session.commit()

        flash(
            f"Account created for {email} (role: {new_role}). No email was sent — share credentials with them directly."
        )
        return redirect(url_for("admin_access_requests"))

    rows = AccessRequest.query.order_by(AccessRequest.created_at.desc()).all()
    return render_template(
        "admin_access_requests.html",
        requests=rows,
        approve_roles=ACCESS_REQUEST_APPROVE_ROLES,
    )


def _admin_governance_denied_response():
    flash("Access denied")
    return redirect(url_for("dashboard"))


@app.route("/admin/governance-register", methods=["GET"])
@login_required
def admin_governance_register():
    if not getattr(current_user, "is_admin", False):
        return _admin_governance_denied_response()
    _ensure_db_tables()
    page = request.args.get("page", 1, type=int) or 1
    if page < 1:
        page = 1
    sort_key = (request.args.get("sort") or "request_date").strip()
    direction = (request.args.get("dir") or "desc").strip().lower()
    if direction not in ("asc", "desc"):
        direction = "desc"
    col = _governance_sort_column(sort_key)
    args_map = _governance_filter_map_from_values(request.args)
    base = GovernanceRegister.query
    base = _apply_governance_register_filters(base, args_map)
    stats = _governance_register_stats_with_pct(base)
    ordered = base.order_by(col.asc() if direction == "asc" else col.desc(), GovernanceRegister.id.asc() if direction == "asc" else GovernanceRegister.id.desc())
    pagination = ordered.paginate(page=page, per_page=25, error_out=False)
    rows_payload = [_governance_row_to_editor_dict(r) for r in pagination.items]
    page_uids: set[int] = set()
    for r in pagination.items:
        if r.last_updated_by_user_id:
            page_uids.add(int(r.last_updated_by_user_id))
    gr_last_user_labels = _governance_user_label_map(page_uids)
    recent_audit = _governance_recent_audit_payload(25)
    return render_template(
        "admin/governance_register.html",
        pagination=pagination,
        stats=stats,
        rows_payload=rows_payload,
        filter_args=args_map,
        sort_key=sort_key,
        sort_dir=direction,
        governance_access_types=GOVERNANCE_ACCESS_TYPES,
        governance_environments=GOVERNANCE_ENVIRONMENTS,
        governance_statuses=GOVERNANCE_STATUSES,
        default_request_date=date.today().isoformat(),
        gr_url_placeholder_id=837265194,
        gr_last_user_labels=gr_last_user_labels,
        gr_today=date.today(),
        gr_closed_statuses=tuple(GOVERNANCE_REGISTER_STATUSES_CLOSED_FOR_OVERDUE),
        recent_audit=recent_audit,
        governance_attachments_stub_label=_governance_attachments_stub_label,
    )


@app.route("/admin/governance-register/export", methods=["GET"])
@login_required
def admin_governance_register_export():
    if not getattr(current_user, "is_admin", False):
        return _admin_governance_denied_response()
    _ensure_db_tables()
    args_map = _governance_filter_map_from_values(request.args)
    q = _apply_governance_register_filters(GovernanceRegister.query, args_map)
    q = q.order_by(GovernanceRegister.request_date.desc(), GovernanceRegister.id.desc())
    rows = q.all()
    wb = Workbook()
    ws = wb.active
    ws.title = "Governance Register"
    headers = [
        "Request Date",
        "Requested By",
        "Team / Department",
        "API / Software Name",
        "Vendor / Platform",
        "Purpose / Business Reason",
        "Access Type",
        "Environment",
        "Status",
        "Approved By",
        "Expiry / Review Date",
        "Notes / Risks",
        "Linked Documentation",
        "Owner",
        "Created By",
        "Last Updated By",
        "Attachments (Stub)",
        "Created At",
        "Updated At",
    ]
    ws.append(headers)
    uids: set[int] = set()
    for r in rows:
        if r.created_by_user_id:
            uids.add(int(r.created_by_user_id))
        if r.last_updated_by_user_id:
            uids.add(int(r.last_updated_by_user_id))
    um = _governance_user_label_map(uids)
    for r in rows:
        cb = um.get(int(r.created_by_user_id)) if r.created_by_user_id else ""
        lu = um.get(int(r.last_updated_by_user_id)) if r.last_updated_by_user_id else ""
        att_txt, _ = _governance_attachments_stub_label(r.attachments_stub_json)
        ws.append(
            [
                r.request_date.isoformat() if r.request_date else "",
                r.requested_by or "",
                r.team_department or "",
                r.api_software_name or "",
                r.vendor_platform or "",
                r.purpose_business_reason or "",
                r.access_type or "",
                r.environment or "",
                r.status or "",
                r.approved_by or "",
                r.expiry_review_date.isoformat() if r.expiry_review_date else "",
                r.notes_risks or "",
                r.linked_documentation or "",
                r.owner or "",
                cb,
                lu,
                att_txt,
                r.created_at.strftime("%Y-%m-%d %H:%M") if r.created_at else "",
                r.updated_at.strftime("%Y-%m-%d %H:%M") if r.updated_at else "",
            ]
        )
    bio = BytesIO()
    wb.save(bio)
    bio.seek(0)
    fn = f"governance_register_{datetime.utcnow().strftime('%Y%m%d_%H%M')}.xlsx"
    return send_file(
        bio,
        as_attachment=True,
        download_name=fn,
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )


@app.route("/admin/governance-register/create", methods=["POST"])
@login_required
def admin_governance_register_create():
    if not getattr(current_user, "is_admin", False):
        return _admin_governance_denied_response()
    _ensure_db_tables()
    fields, err = _governance_read_form_fields()
    if err:
        flash(err)
        return redirect(url_for("admin_governance_register"))
    row = GovernanceRegister(**fields)
    uid = int(current_user.id)
    row.created_by_user_id = uid
    row.last_updated_by_user_id = uid
    row.attachments_stub_json = row.attachments_stub_json or "[]"
    db.session.add(row)
    db.session.flush()
    _log_governance_audit(row.id, "create", record_label=row.api_software_name, user_id=uid)
    db.session.commit()
    flash("Governance record created.")
    return redirect(url_for("admin_governance_register"))


@app.route("/admin/governance-register/update/<int:row_id>", methods=["POST"])
@login_required
def admin_governance_register_update(row_id: int):
    if not getattr(current_user, "is_admin", False):
        return _admin_governance_denied_response()
    _ensure_db_tables()
    row = db.session.get(GovernanceRegister, row_id)
    if not row:
        flash("Record not found.")
        return redirect(url_for("admin_governance_register"))
    fields, err = _governance_read_form_fields()
    if err:
        flash(err)
        return redirect(url_for("admin_governance_register"))
    for k, v in fields.items():
        setattr(row, k, v)
    uid = int(current_user.id)
    row.last_updated_by_user_id = uid
    row.updated_at = datetime.utcnow()
    _log_governance_audit(row.id, "update", record_label=row.api_software_name, user_id=uid)
    db.session.commit()
    flash("Governance record updated.")
    return redirect(url_for("admin_governance_register"))


@app.route("/admin/governance-register/delete/<int:row_id>", methods=["POST"])
@login_required
def admin_governance_register_delete(row_id: int):
    if not getattr(current_user, "is_admin", False):
        return _admin_governance_denied_response()
    _ensure_db_tables()
    row = db.session.get(GovernanceRegister, row_id)
    if row:
        label = (row.api_software_name or "")[:400]
        rid = int(row.id)
        uid = int(current_user.id)
        _log_governance_audit(rid, "delete", record_label=label, user_id=uid)
        db.session.delete(row)
        db.session.commit()
        flash("Governance record deleted.")
    else:
        flash("Record not found.")
    return redirect(url_for("admin_governance_register"))


@app.route("/forgot-password", methods=["GET", "POST"])
def forgot_password():
    _ensure_db_tables()
    generic_ok = "If the email exists, a reset link will be sent."
    if request.method == "POST":
        email = (request.form.get("email") or "").strip().lower()
        client_ip = _client_ip_for_rate_limit()
        if not _forgot_pw_rate_allow_ip(client_ip):
            _audit_password_reset("password_reset_request_blocked", reason="rate_limit_ip", ip=client_ip)
            flash("Too many password reset attempts. Please try again later.")
            return redirect(url_for("forgot_password"))
        if not _email_domain_allowed(email):
            flash("Only CTS company email addresses are allowed")
            return render_template("forgot_password.html")
        if not _forgot_pw_rate_allow_email(email):
            _audit_password_reset("password_reset_request_blocked", reason="rate_limit_email", ip=client_ip, email=email)
            flash("Too many password reset attempts. Please try again later.")
            return redirect(url_for("forgot_password"))
        user = User.query.filter(db.func.lower(User.email) == email).first()
        if user:
            _delete_expired_password_reset_tokens()
            raw = secrets.token_urlsafe(32)
            token_hash = _hash_password_reset_token(raw)
            for old in PasswordResetToken.query.filter_by(user_id=user.id).all():
                db.session.delete(old)
            db.session.add(
                PasswordResetToken(
                    user_id=user.id,
                    token_hash=token_hash,
                    expires_at=datetime.utcnow() + timedelta(hours=1),
                    used=False,
                )
            )
            db.session.commit()
            reset_url = f"{PUBLIC_APP_BASE_URL}/reset-password/{raw}"
            body = (
                "You requested a password reset.\n\n"
                "Click the link below to create a new password:\n\n"
                f"{reset_url}\n\n"
                "This link expires in 1 hour.\n\n"
                "If you did not request this, ignore this email.\n"
            )
            _send_plain_email(str(user.email), "Password reset request", body)
            _audit_password_reset("password_reset_requested", ip=client_ip, email=email, user_id=user.id)
        else:
            _audit_password_reset("password_reset_requested", ip=client_ip, email=email, user_found="false")
        flash(generic_ok)
        return redirect(url_for("forgot_password"))
    return render_template("forgot_password.html")


@app.route("/reset-password/<token>", methods=["GET", "POST"])
def reset_password(token):
    _ensure_db_tables()
    raw = (token or "").strip()
    if len(raw) < 20:
        flash("This reset link is invalid or has expired.")
        return redirect(url_for("login"))
    th = _hash_password_reset_token(raw)
    row = (
        PasswordResetToken.query.filter_by(token_hash=th, used=False)
        .filter(PasswordResetToken.expires_at > datetime.utcnow())
        .first()
    )
    if request.method == "GET":
        if not row:
            flash("This reset link is invalid or has expired.")
            return redirect(url_for("login"))
        return render_template("reset_password.html", token=raw)

    if not row:
        flash("This reset link is invalid or has expired.")
        return redirect(url_for("login"))

    p1 = request.form.get("password") or ""
    p2 = request.form.get("password_confirm") or ""
    if not _password_reset_meets_policy(p1):
        flash("Password must be at least 8 characters and include at least one letter and one number.")
        return render_template("reset_password.html", token=raw)
    if p1 != p2:
        flash("Passwords do not match.")
        return render_template("reset_password.html", token=raw)

    user = User.query.filter_by(id=row.user_id).first()
    if not user:
        flash("This reset link is invalid or has expired.")
        return redirect(url_for("login"))

    client_ip = _client_ip_for_rate_limit()
    user.password_hash = generate_password_hash(p1)
    row.used = True
    db.session.commit()
    _audit_password_reset("password_reset_completed", ip=client_ip, user_id=user.id)
    flash("Password updated successfully.")
    return redirect(url_for("login"))


def _normalize_profile_select(raw: object, allowed: tuple[str, ...]) -> str | None:
    value = str(raw or "").strip()
    if not value:
        return None
    return value if value in set(allowed) else None


def _normalize_yes_no(raw: object) -> str | None:
    value = str(raw or "").strip().lower()
    if value in {"yes", "no"}:
        return value
    return None


def _parse_operating_locations_form(raw: object) -> tuple[list[dict[str, str]], str | None]:
    try:
        rows = json.loads(str(raw or "[]"))
    except Exception:
        return [], "Operating locations payload is invalid."
    if not isinstance(rows, list):
        return [], "Operating locations payload is invalid."
    valid_country_codes = {code for code, _name in ISO_COUNTRIES}
    valid_site_types = set(OPERATING_SITE_TYPE_OPTIONS)
    out: list[dict[str, str]] = []
    for idx, item in enumerate(rows, start=1):
        if not isinstance(item, dict):
            return [], f"Operating location {idx} is invalid."
        country = str(item.get("country") or "").strip().upper()
        site_type = str(item.get("site_type") or "").strip().lower()
        if not country and not site_type:
            continue
        if country and country not in valid_country_codes:
            return [], f"Operating location {idx} has an invalid country."
        if site_type and site_type not in valid_site_types:
            return [], f"Operating location {idx} has an invalid site type."
        out.append({"country": country, "site_type": site_type})
    return out, None


def _apply_profile_form_fields(user: User, form) -> str | None:
    mode_raw = str(form.get("template_mode") or "").strip()
    mode = normalize_template_mode(mode_raw) if mode_raw in VALID_TEMPLATE_MODES else TEMPLATE_MODE_LEGACY
    user.template_mode = mode

    product_profile_keys = {
        "business_type",
        "product_type",
        "quantity",
        "quantity_unit",
        "number_of_products_in_use",
        "end_use_location",
        "heating_source",
        "travel_provider",
        "operating_locations_json",
    }
    if any(key in form for key in product_profile_keys):
        business_type = _normalize_profile_select(form.get("business_type"), BUSINESS_TYPE_OPTIONS)
        product_type = (form.get("product_type") or "").strip() or None
        quantity = (form.get("quantity") or "").strip() or None
        quantity_unit = (form.get("quantity_unit") or "").strip() or None
        number_of_products_in_use = (form.get("number_of_products_in_use") or "").strip() or None
        end_use_location = (form.get("end_use_location") or "").strip() or None
        heating_source = _normalize_profile_select(form.get("heating_source"), HEATING_SOURCE_OPTIONS)
        travel_provider = _normalize_yes_no(form.get("travel_provider"))
        operating_locations, operating_error = _parse_operating_locations_form(form.get("operating_locations_json"))
        if operating_error:
            return operating_error
        if business_type != "Manufacturer":
            product_type = None
        if travel_provider is None and str(form.get("travel_provider") or "").strip():
            return "Travel provider must be Yes or No."

        user.business_type = business_type
        user.product_type = product_type
        user.quantity = quantity
        user.quantity_unit = quantity_unit
        user.number_of_products_in_use = number_of_products_in_use
        user.end_use_location = end_use_location
        user.heating_source = heating_source
        user.travel_provider = travel_provider
        user.operating_locations_json = json.dumps(operating_locations, ensure_ascii=True)
    _persist_template_mode(mode)
    return None


PRODUCTS_INPUT_COLUMNS: tuple[str, ...] = (
    "Reporting period (month, year)",
    "Product Type",
    "Quantity",
    "Quantity Unit",
    "End Use Location",
    "Product Weight",
    "Product Unit",
)


def _products_current_period(today: date | None = None) -> tuple[str, str]:
    value = today or date.today()
    return value.strftime("%Y-%m"), value.strftime("%B %Y")


def _load_products_template_config() -> dict[str, object]:
    try:
        with PRODUCTS_TEMPLATE_PATH.open("r", encoding="utf-8") as f:
            data = json.load(f)
    except Exception:
        data = {}
    if not isinstance(data, dict):
        data = {}
    data.setdefault("columns", list(PRODUCTS_INPUT_COLUMNS))
    data.setdefault("dropdowns", {})
    data.setdefault("validation", {})
    return data


def _products_company_name(user: object) -> str:
    return str(getattr(user, "company_name", None) or "").strip()


def _products_current_month_has_entry(user: object) -> bool:
    _ensure_db_tables()
    company_name = _products_company_name(user)
    if not company_name:
        return False
    period_key, _period_label = _products_current_period()
    return (
        ProductMonthlyEntry.query.filter_by(
            company_name=company_name,
            reporting_period_key=period_key,
        )
        .limit(1)
        .first()
        is not None
    )


def _product_entry_payload(row: ProductMonthlyEntry) -> dict[str, object]:
    return {
        "id": int(row.id),
        "Reporting period (month, year)": row.reporting_period_label,
        "Product Type": row.product_type,
        "Quantity": row.quantity,
        "Quantity Unit": row.quantity_unit,
        "End Use Location": row.end_use_location,
        "Product Weight": row.product_weight,
        "Product Unit": row.product_unit,
    }


def _products_profile_payload(user: object) -> dict[str, object]:
    return {
        "business_type": str(getattr(user, "business_type", None) or "").strip(),
        "product_type": str(getattr(user, "product_type", None) or "").strip(),
        "quantity": str(getattr(user, "quantity", None) or "").strip(),
        "quantity_unit": str(getattr(user, "quantity_unit", None) or "").strip(),
        "number_of_products_in_use": str(getattr(user, "number_of_products_in_use", None) or "").strip(),
        "end_use_location": str(getattr(user, "end_use_location", None) or "").strip(),
        "heating_source": str(getattr(user, "heating_source", None) or "").strip(),
        "travel_provider": str(getattr(user, "travel_provider", None) or "").strip(),
        "operating_locations": _operating_locations_for_user(user),
    }


def _float_from_payload(value: object) -> float | None:
    raw = str(value if value is not None else "").strip().replace(",", ".")
    if not raw:
        return None
    try:
        parsed = float(raw)
    except Exception:
        return None
    return parsed if parsed > 0 else None


def _country_code_from_location(value: object) -> str:
    raw = str(value or "").strip()
    if not raw:
        return ""
    upper = raw.upper()
    if upper in ISO_COUNTRY_NAME_BY_CODE:
        return upper
    match = re.search(r"\(([A-Z]{2})\)", raw)
    if match and match.group(1) in ISO_COUNTRY_NAME_BY_CODE:
        return match.group(1)
    return ISO_COUNTRY_CODE_BY_NAME.get(raw.casefold(), "")


def _products_distance_for_location(end_use_location: object, company_country: object) -> tuple[float, float]:
    target_code = _country_code_from_location(end_use_location)
    origin_code = _country_code_from_location(company_country)
    if target_code and origin_code and target_code == origin_code:
        return 300.0, 0.0
    european_codes = {
        "AT", "BE", "BG", "HR", "CY", "CZ", "DK", "EE", "FI", "FR", "DE", "GR", "HU",
        "IE", "IT", "LV", "LT", "LU", "MT", "NL", "PL", "PT", "RO", "SK", "SI", "ES",
        "SE", "NO", "IS", "CH", "GB",
    }
    if target_code in european_codes:
        return 1500.0, 300.0
    return 250.0, 8000.0


def _normalize_products_profile_payload(profile: dict[str, object]) -> tuple[dict[str, object], list[str]]:
    errors: list[str] = []
    business_type = _normalize_profile_select(profile.get("business_type"), BUSINESS_TYPE_OPTIONS)
    if not business_type:
        errors.append("Business Type is required.")
    heating_source = _normalize_profile_select(profile.get("heating_source"), HEATING_SOURCE_OPTIONS)
    travel_provider = _normalize_yes_no(profile.get("travel_provider"))
    if str(profile.get("travel_provider") or "").strip() and travel_provider is None:
        errors.append("Travel Provider must be Yes or No.")

    locations_raw = profile.get("operating_locations")
    if isinstance(locations_raw, list):
        locations_json = json.dumps(locations_raw, ensure_ascii=True)
    else:
        locations_json = str(profile.get("operating_locations_json") or "[]")
    operating_locations, operating_error = _parse_operating_locations_form(locations_json)
    if operating_error:
        errors.append(operating_error)
    if not operating_locations:
        errors.append("At least one Operating Location is required.")

    number_of_products = str(profile.get("number_of_products_in_use") or "").strip()
    if business_type == "Manufacturer" and not number_of_products:
        errors.append("Number of products in use is required for manufacturers.")

    return (
        {
            "business_type": business_type or "",
            "number_of_products_in_use": number_of_products,
            "heating_source": heating_source or "",
            "travel_provider": travel_provider or "",
            "operating_locations": operating_locations,
        },
        errors,
    )


def _normalize_products_rows(rows: object, period_key: str, period_label: str) -> tuple[list[dict[str, object]], list[str]]:
    config = _load_products_template_config()
    dropdowns = config.get("dropdowns") if isinstance(config.get("dropdowns"), dict) else {}
    quantity_units = set(dropdowns.get("quantity_units") or ())
    product_units = set(dropdowns.get("product_units") or ())
    normalized: list[dict[str, object]] = []
    errors: list[str] = []
    if not isinstance(rows, list):
        return [], ["Rows must be a list."]
    for idx, row in enumerate(rows, start=1):
        if not isinstance(row, dict):
            errors.append(f"Row {idx} is invalid.")
            continue
        product_type = str(row.get("Product Type") or row.get("product_type") or "").strip()
        quantity = _float_from_payload(row.get("Quantity") or row.get("quantity"))
        quantity_unit = str(row.get("Quantity Unit") or row.get("quantity_unit") or "").strip()
        end_use_location = str(row.get("End Use Location") or row.get("end_use_location") or "").strip()
        product_weight = _float_from_payload(row.get("Product Weight") or row.get("product_weight"))
        product_unit = str(row.get("Product Unit") or row.get("product_unit") or "").strip()

        if not product_type:
            errors.append(f"Row {idx}: Product Type is required.")
        if quantity is None:
            errors.append(f"Row {idx}: Quantity must be greater than 0.")
        if not quantity_unit:
            errors.append(f"Row {idx}: Quantity Unit is required.")
        elif quantity_units and quantity_unit not in quantity_units:
            errors.append(f"Row {idx}: Quantity Unit is invalid.")
        if not end_use_location:
            errors.append(f"Row {idx}: End Use Location is required.")
        if product_weight is None:
            errors.append(f"Row {idx}: Product Weight must be greater than 0.")
        if not product_unit:
            errors.append(f"Row {idx}: Product Unit is required.")
        elif product_units and product_unit not in product_units:
            errors.append(f"Row {idx}: Product Unit is invalid.")

        normalized.append(
            {
                "reporting_period_key": period_key,
                "reporting_period_label": period_label,
                "product_type": product_type,
                "quantity": float(quantity or 0),
                "quantity_unit": quantity_unit,
                "end_use_location": end_use_location,
                "product_weight": float(product_weight or 0),
                "product_unit": product_unit,
            }
        )
    if not normalized:
        errors.append("Add at least one product row for the current month.")
    return normalized, errors


def _generated_products_category_datasets(user: object, rows: list[dict[str, object]]) -> dict[str, list[dict[str, object]]]:
    category9: list[dict[str, object]] = []
    category11: list[dict[str, object]] = []
    category12: list[dict[str, object]] = []
    for row in rows:
        total_weight = float(row["product_weight"]) * float(row["quantity"])
        road_km, sea_km = _products_distance_for_location(row["end_use_location"], getattr(user, "company_country", None))
        base = {
            "Reporting period (month, year)": row["reporting_period_label"],
            "Product Type": row["product_type"],
            "Quantity": row["quantity"],
            "Quantity Unit": row["quantity_unit"],
            "End Use Location": row["end_use_location"],
        }
        category9.append(
            {
                **base,
                "Product Weight": row["product_weight"],
                "Product Unit": row["product_unit"],
                "Total Weight": total_weight,
                "Road km": road_km,
                "Sea km": sea_km,
                "Road emission factor": 0.000059,
                "Sea emission factor": 0.000016,
                "Calculated emissions": (total_weight * (road_km * 0.000059)) + (total_weight * (sea_km * 0.000016)),
            }
        )
        category11.append(dict(base))
        category12.append(dict(base))
    return {
        "Scope 3 Category 9 Downstream Transport": category9,
        "Scope 3 Category 11 Use of Sold Products": category11,
        "Scope 3 Category 12 End of Life": category12,
    }


def _publish_products_structured_input(user: object, period_key: str, period_label: str, rows: list[dict[str, object]]) -> Path | None:
    try:
        out_dir = DATA_DIR / "products_input"
        out_dir.mkdir(parents=True, exist_ok=True)
        company_name = _products_company_name(user)
        out_path = out_dir / f"{company_slug(company_name)}_{period_key}.json"
        payload = {
            "company_name": company_name,
            "reporting_period_key": period_key,
            "reporting_period_label": period_label,
            "source": "Business Data Input - Products Log",
            "columns": list(PRODUCTS_INPUT_COLUMNS),
            "rows": rows,
            "generated_categories": _generated_products_category_datasets(user, rows),
        }
        out_path.write_text(json.dumps(payload, ensure_ascii=True, indent=2), encoding="utf-8")
        return out_path
    except Exception:
        return None


@app.route("/profile/setup", methods=["GET", "POST"])
@login_required
def profile_setup():
    _ensure_db_tables()
    if _user_profile_complete(current_user):
        return redirect(url_for("dashboard"))

    companies = list(COMPANIES)
    resolved_company = _resolve_template_company_name(current_user.company_name or "") or (current_user.company_name or "").strip()
    template_ctx = _profile_template_context(
        companies=companies,
        resolved_company=resolved_company,
    )

    if request.method == "POST":
        first_name = (request.form.get("first_name") or "").strip()
        last_name = (request.form.get("last_name") or "").strip()
        job_title = (request.form.get("job_title") or "").strip()
        phone = (request.form.get("phone") or "").strip()
        co = (request.form.get("company_name") or "").strip()
        country_code = (request.form.get("company_country") or "").strip().upper()

        if not first_name or not last_name or not job_title:
            flash("First name, last name, and job title are required.")
            return render_template("profile_setup.html", **template_ctx)

        if co not in COMPANIES:
            flash("Invalid company selection.")
            return render_template("profile_setup.html", **template_ctx)

        valid_codes = {c for c, _n in ISO_COUNTRIES}
        if country_code not in valid_codes:
            flash("Please select a valid country.")
            return render_template("profile_setup.html", **template_ctx)

        current_user.first_name = first_name
        current_user.last_name = last_name
        current_user.job_title = job_title
        current_user.phone = phone or None
        current_user.company_name = co
        current_user.company_country = country_code
        profile_error = _apply_profile_form_fields(current_user, request.form)
        if profile_error:
            flash(profile_error)
            return render_template("profile_setup.html", **template_ctx)

        pfile = request.files.get("profile_photo")
        rel_photo = _save_profile_photo_file(pfile, int(current_user.id))
        if rel_photo:
            current_user.profile_photo_path = rel_photo

        current_user.is_profile_complete = True
        db.session.commit()
        flash("Profile saved.")
        return redirect(url_for("dashboard"))

    return render_template("profile_setup.html", **template_ctx)


@app.route("/settings/profile", methods=["GET", "POST"], endpoint="settings_profile_page")
@app.route("/profile", methods=["GET", "POST"])
@login_required
def profile_page():
    _ensure_db_tables()
    if not _user_profile_complete(current_user):
        return redirect(url_for("profile_setup"))
    if _is_readonly_user(current_user):
        return redirect(url_for("public_profile", user_id=int(current_user.id)))

    if request.method == "POST":
        current_user.first_name = (request.form.get("first_name") or "").strip() or None
        current_user.last_name = (request.form.get("last_name") or "").strip() or None
        current_user.job_title = (request.form.get("job_title") or "").strip() or None
        current_user.phone = (request.form.get("phone") or "").strip() or None
        profile_error = _apply_profile_form_fields(current_user, request.form)
        if profile_error:
            flash(profile_error)
            return redirect(url_for("settings_profile_page"))

        if current_user.is_admin:
            country_code = (request.form.get("company_country") or "").strip().upper()
            valid_codes = {c for c, _n in ISO_COUNTRIES}
            if country_code in valid_codes:
                current_user.company_country = country_code

        pfile = request.files.get("profile_photo")
        rel_photo = _save_profile_photo_file(pfile, int(current_user.id))
        if rel_photo:
            current_user.profile_photo_path = rel_photo

        db.session.commit()
        flash("Profile updated.")
        return redirect(url_for("settings_profile_page"))

    cc = (current_user.company_country or "").strip().upper()
    country_readonly_label = cc
    for code, name in ISO_COUNTRIES:
        if code == cc:
            country_readonly_label = f"{name} ({code})"
            break

    return render_template(
        "profile.html",
        company_display=(current_user.company_name or ""),
        country_readonly_label=country_readonly_label,
        **_profile_template_context(),
    )


@app.route("/data-input/products", methods=["GET"], endpoint="products_input_page")
@login_required
def products_input_page():
    _ensure_db_tables()
    period_key, period_label = _products_current_period()
    company_name = _products_company_name(current_user)
    rows = (
        ProductMonthlyEntry.query.filter_by(
            company_name=company_name,
            reporting_period_key=period_key,
        )
        .order_by(ProductMonthlyEntry.row_index.asc(), ProductMonthlyEntry.id.asc())
        .all()
    )
    return render_template(
        "products_input.html",
        config=_load_products_template_config(),
        period_key=period_key,
        period_label=period_label,
        company_name=company_name,
        product_rows=[_product_entry_payload(row) for row in rows],
        product_profile=_products_profile_payload(current_user),
        business_type_options=BUSINESS_TYPE_OPTIONS,
        heating_source_options=HEATING_SOURCE_OPTIONS,
        travel_provider_options=TRAVEL_PROVIDER_OPTIONS,
        operating_site_type_options=OPERATING_SITE_TYPE_OPTIONS,
        iso_countries=ISO_COUNTRIES,
    )


@app.route("/api/products-input/save", methods=["POST"], endpoint="api_products_input_save")
@login_required
def api_products_input_save():
    _ensure_db_tables()
    if _is_readonly_user(current_user):
        return jsonify({"error": "Auditor accounts cannot submit product data."}), 403
    payload = request.get_json(silent=True) or {}
    period_key, period_label = _products_current_period()
    profile_payload, profile_errors = _normalize_products_profile_payload(payload.get("profile") if isinstance(payload.get("profile"), dict) else {})
    rows, row_errors = _normalize_products_rows(payload.get("rows"), period_key, period_label)
    errors = profile_errors + row_errors
    if errors:
        return jsonify({"error": errors[0], "validation_errors": errors[:30]}), 400

    company_name = _products_company_name(current_user)
    if not company_name:
        return jsonify({"error": "Company name is required before saving product data."}), 400

    try:
        ProductMonthlyEntry.query.filter_by(
            company_name=company_name,
            reporting_period_key=period_key,
        ).delete()
        for idx, row in enumerate(rows, start=1):
            db.session.add(
                ProductMonthlyEntry(
                    company_name=company_name,
                    reporting_period_key=period_key,
                    reporting_period_label=period_label,
                    row_index=idx,
                    product_type=str(row["product_type"]),
                    quantity=float(row["quantity"]),
                    quantity_unit=str(row["quantity_unit"]),
                    end_use_location=str(row["end_use_location"]),
                    product_weight=float(row["product_weight"]),
                    product_unit=str(row["product_unit"]),
                    created_by_user_id=int(current_user.id),
                )
            )

        current_user.business_type = str(profile_payload["business_type"] or "") or None
        current_user.number_of_products_in_use = str(profile_payload["number_of_products_in_use"] or "") or None
        current_user.heating_source = str(profile_payload["heating_source"] or "") or None
        current_user.travel_provider = str(profile_payload["travel_provider"] or "") or None
        current_user.operating_locations_json = json.dumps(profile_payload["operating_locations"], ensure_ascii=True)
        if rows:
            first = rows[0]
            current_user.product_type = str(first["product_type"] or "") or None
            current_user.quantity = str(first["quantity"] or "") or None
            current_user.quantity_unit = str(first["quantity_unit"] or "") or None
            current_user.end_use_location = str(first["end_use_location"] or "") or None

        db.session.commit()
    except Exception as exc:
        db.session.rollback()
        return jsonify({"error": f"Save failed: {exc}"}), 500

    out_path = _publish_products_structured_input(current_user, period_key, period_label, rows)
    return jsonify(
        {
            "ok": True,
            "message": f"{len(rows)} product row(s) saved for {period_label}.",
            "reporting_period_key": period_key,
            "reporting_period_label": period_label,
            "structured_input_path": str(out_path) if out_path else "",
            "generated_categories": _generated_products_category_datasets(current_user, rows),
        }
    )


@app.route("/api/products-input/export", methods=["GET"], endpoint="api_products_input_export")
@login_required
def api_products_input_export():
    _ensure_db_tables()
    period_key = str(request.args.get("period") or _products_current_period()[0]).strip()
    if not re.fullmatch(r"\d{4}-\d{2}", period_key):
        period_key = _products_current_period()[0]
    company_name = _products_company_name(current_user)
    rows = (
        ProductMonthlyEntry.query.filter_by(
            company_name=company_name,
            reporting_period_key=period_key,
        )
        .order_by(ProductMonthlyEntry.row_index.asc(), ProductMonthlyEntry.id.asc())
        .all()
    )
    normalized_rows = [
        {
            "reporting_period_key": row.reporting_period_key,
            "reporting_period_label": row.reporting_period_label,
            "product_type": row.product_type,
            "quantity": float(row.quantity),
            "quantity_unit": row.quantity_unit,
            "end_use_location": row.end_use_location,
            "product_weight": float(row.product_weight),
            "product_unit": row.product_unit,
        }
        for row in rows
    ]
    return jsonify(
        {
            "company_name": company_name,
            "reporting_period_key": period_key,
            "rows": [_product_entry_payload(row) for row in rows],
            "generated_categories": _generated_products_category_datasets(current_user, normalized_rows),
        }
    )





def _country_label_from_code(raw_code: object) -> str:
    code = str(raw_code or "").strip().upper()
    if not code:
        return ""
    for item_code, item_name in ISO_COUNTRIES:
        if str(item_code).strip().upper() == code:
            return f"{item_name} ({code})"
    return code


def _country_flag_emoji(raw_code: object) -> str:
    code = str(raw_code or "").strip().upper()
    if not re.fullmatch(r"[A-Z]{2}", code):
        return ""
    return "".join(chr(0x1F1E6 + ord(char) - ord("A")) for char in code)


def _profile_location_label(u: User | None) -> str:
    if u is None:
        return ""
    end_use_location = str(getattr(u, "end_use_location", None) or "").strip()
    if end_use_location:
        return end_use_location
    return _country_label_from_code(getattr(u, "company_country", None))


def _profile_about_text(u: User | None) -> str:
    if u is None:
        return ""
    name = _user_display_name(u)
    title = str(getattr(u, "job_title", None) or "").strip()
    company = str(getattr(u, "company_name", None) or "").strip()
    location = _profile_location_label(u)
    fragments: list[str] = []
    if title and company:
        fragments.append(f"{name} works as {title} at {company}.")
    elif title:
        fragments.append(f"{name} works as {title}.")
    elif company:
        fragments.append(f"{name} is part of {company}.")
    business_type = str(getattr(u, "business_type", None) or "").strip()
    if business_type:
        fragments.append(f"Business type: {business_type}.")
    if location:
        fragments.append(f"Location: {location}.")
    return " ".join(fragment for fragment in fragments if fragment).strip() or f"{name} is part of the platform team."


def _suggested_profile_users(user_id: int, *, company_name: str) -> list[dict[str, object]]:
    rows = (
        User.query.filter(User.id != int(user_id))
        .order_by(
            db.case((db.func.lower(User.company_name) == str(company_name or "").strip().lower(), 0), else_=1),
            User.created_at.desc(),
        )
        .limit(4)
        .all()
    )
    suggestions: list[dict[str, object]] = []
    for row in rows:
        suggestions.append(
            {
                "id": int(row.id),
                "name": _user_display_name(row),
                "title": _user_professional_title(row),
                "company": str(getattr(row, "company_name", None) or "").strip(),
                "avatar_url": _user_avatar_url(row),
                "profile_url": url_for("public_profile", user_id=int(row.id)),
            }
        )
    return suggestions


def _feed_payloads_for_rows(rows: list[FeedPost]) -> list[dict[str, object]]:
    post_ids = [int(row.id) for row in rows]
    reaction_summary_map, current_reaction_map = _feed_reaction_maps(post_ids, int(getattr(current_user, "id", 0) or 0))
    comment_payload_map, _comment_count_map = _comment_payload_maps(post_ids, int(getattr(current_user, "id", 0) or 0))
    return [
        _feed_post_payload(
            row,
            reaction_summary=reaction_summary_map.get(int(row.id), []),
            current_reaction=current_reaction_map.get(int(row.id), ""),
            comments=comment_payload_map.get(int(row.id), []),
        )
        for row in rows
    ]


@app.route("/profile/<int:user_id>", methods=["GET"])
@login_required
def public_profile(user_id: int):
    _ensure_db_tables()
    user_row = User.query.get_or_404(int(user_id))
    posts = FeedPost.query.filter_by(author_user_id=int(user_row.id)).order_by(FeedPost.created_at.desc(), FeedPost.id.desc()).all()
    reports = Report.query.filter_by(uploaded_by=int(user_row.id)).order_by(Report.created_at.desc(), Report.id.desc()).all()
    company_name = str(getattr(user_row, "company_name", None) or "").strip()
    is_own_profile = int(getattr(current_user, "id", 0) or 0) == int(user_row.id)
    follower_count = (
        db.session.query(db.func.count(UserFollow.id))
        .filter(UserFollow.following_id == int(user_row.id))
        .scalar()
        or 0
    )
    following_count = (
        db.session.query(db.func.count(UserFollow.id))
        .filter(UserFollow.follower_id == int(user_row.id))
        .scalar()
        or 0
    )
    is_following = False
    if not is_own_profile and getattr(current_user, "is_authenticated", False):
        is_following = (
            db.session.query(UserFollow.id)
            .filter(
                UserFollow.follower_id == int(current_user.id),
                UserFollow.following_id == int(user_row.id),
            )
            .first()
            is not None
        )
    activity_candidates = [
        getattr(posts[0], "created_at", None) if posts else None,
        getattr(reports[0], "created_at", None) if reports else None,
    ]
    last_activity_at = max((value for value in activity_candidates if value is not None), default=None)
    active_recently = bool(last_activity_at and last_activity_at >= (datetime.utcnow() - timedelta(days=14)))
    return render_template(
        "user_profile.html",
        profile_user=user_row,
        profile_name=_user_display_name(user_row),
        profile_title=_user_professional_title(user_row),
        profile_role_label=_user_role_label(user_row),
        profile_avatar_url=_user_avatar_url(user_row),
        profile_company=company_name or "CTS Carbon Platform",
        profile_country_flag=_country_flag_emoji(getattr(user_row, "company_country", None)),
        profile_cover_url=url_for("static", filename=user_row.cover_image) if getattr(user_row, "cover_image", None) else "",
        profile_location=_profile_location_label(user_row),
        profile_about=_profile_about_text(user_row),
        profile_company_logo_url=_company_logo_url(company_name),
        profile_stats={
            "posts": len(posts),
            "reports": len(reports),
            "followers": int(follower_count),
            "following": int(following_count),
            "joined": getattr(user_row, "created_at", None).strftime("%b %Y") if getattr(user_row, "created_at", None) else "",
            "template_mode": normalize_template_mode(getattr(user_row, "template_mode", None)),
            "active_recently": active_recently,
            "recent_activity_label": "Active recently" if active_recently else "Latest activity",
        },
        suggested_users=_suggested_profile_users(int(user_row.id), company_name=company_name),
        profile_posts=_feed_payloads_for_rows(posts),
        profile_reports=[payload for payload in (_report_payload(row) for row in reports) if payload],
        feed_profile={
            "name": _user_display_name(current_user),
            "title": _user_professional_title(current_user),
            "company_name": (getattr(current_user, "company_name", None) or "").strip() or "CTS Carbon Platform",
            "avatar_url": _user_avatar_url(current_user),
        },
        feed_post_types=FEED_COMPOSER_TYPES,
        feed_post_next_url=url_for("public_profile", user_id=int(user_row.id)),
        feed_reaction_options=FEED_REACTION_OPTIONS,
        can_create_posts=not _is_readonly_user(current_user),
        comment_actor_avatar_url=_user_avatar_url(current_user),
        is_own_profile=is_own_profile,
        is_following=is_following,
        can_profile_interact=not _is_readonly_user(current_user),
    )


@app.route("/reports/<int:report_id>/open", methods=["GET"])
@login_required
def open_report(report_id: int):
    _ensure_db_tables()
    row = Report.query.get_or_404(int(report_id))
    disk_path = APP_DIR / "static" / str(getattr(row, "file_path", "") or "")
    if not disk_path.is_file():
        abort(404)
    return send_file(str(disk_path), as_attachment=False, download_name=Path(disk_path).name)


@app.route("/reports", methods=["GET"])
@login_required
def reports_page():
    _ensure_db_tables()
    selected_category_raw = str(request.args.get("category") or "").strip()
    selected_category_id = int(selected_category_raw) if selected_category_raw.isdigit() else None
    category_rows = ReportCategory.query.order_by(ReportCategory.name.asc(), ReportCategory.id.asc()).all()
    query = Report.query.order_by(Report.created_at.desc(), Report.id.desc())
    if selected_category_id:
        query = query.filter(Report.category_id == int(selected_category_id))
    rows = query.all()
    category_items = [
        {
            "id": None,
            "name": "All Reports",
            "count": Report.query.count(),
            "href": url_for("reports_page"),
            "is_active": selected_category_id is None,
        }
    ]
    for category in category_rows:
        category_items.append(
            {
                "id": int(category.id),
                "name": str(category.name or "").strip() or "Category",
                "count": Report.query.filter_by(category_id=int(category.id)).count(),
                "href": url_for("reports_page", category=int(category.id)),
                "is_active": selected_category_id == int(category.id),
            }
        )
    return render_template(
        "reports.html",
        page_title="Reports",
        page_subtitle="Shared reports published by the sustainability team.",
        items=[payload for payload in (_report_payload(row) for row in rows) if payload],
        can_manage=bool(current_user.is_admin and not _is_readonly_user(current_user)),
        upload_action=url_for("upload_report"),
        create_category_action=url_for("create_report_category"),
        categories=category_items,
        report_categories=[payload for payload in (_report_category_payload(row) for row in category_rows) if payload],
        selected_category_id=selected_category_id,
    )


@app.route("/reports/<int:report_id>", methods=["GET"])
@login_required
def report_detail(report_id: int):
    _ensure_db_tables()
    row = Report.query.get_or_404(int(report_id))
    payload = _report_payload(row)
    return render_template(
        "module_detail.html",
        module_key="report",
        page_title=payload["title"] if payload else "Report",
        item=payload,
    )


@app.route("/newsletters", methods=["GET"])
@login_required
def newsletters_page():
    _ensure_db_tables()
    rows = Newsletter.query.order_by(Newsletter.created_at.desc(), Newsletter.id.desc()).all()
    return render_template(
        "module_list.html",
        module_key="newsletter",
        page_title="Newsletters",
        page_subtitle="Internal newsletters shared across the platform.",
        items=[payload for payload in (_newsletter_payload(row) for row in rows) if payload],
        can_manage=bool(current_user.is_admin and not _is_readonly_user(current_user)),
        upload_action=url_for("upload_newsletter"),
    )


@app.route("/newsletters/<int:newsletter_id>", methods=["GET"])
@login_required
def newsletter_detail(newsletter_id: int):
    _ensure_db_tables()
    row = Newsletter.query.get_or_404(int(newsletter_id))
    payload = _newsletter_payload(row)
    return render_template(
        "module_detail.html",
        module_key="newsletter",
        page_title=payload["title"] if payload else "Newsletter",
        item=payload,
    )


@app.route("/newsletters/<int:newsletter_id>/open", methods=["GET"])
@login_required
def open_newsletter(newsletter_id: int):
    _ensure_db_tables()
    row = Newsletter.query.get_or_404(int(newsletter_id))
    disk_path = APP_DIR / "static" / str(getattr(row, "file_path", "") or "")
    if not disk_path.is_file():
        abort(404)
    return send_file(str(disk_path), as_attachment=False, download_name=Path(disk_path).name)


@app.route("/events", methods=["GET"])
@login_required
def events_page():
    _ensure_db_tables()
    rows = Event.query.order_by(Event.event_date.desc(), Event.created_at.desc(), Event.id.desc()).all()
    return render_template(
        "module_list.html",
        module_key="event",
        page_title="Events",
        page_subtitle="Upcoming and recent sustainability events.",
        items=[payload for payload in (_event_payload(row) for row in rows) if payload],
        can_manage=bool(current_user.is_admin and not _is_readonly_user(current_user)),
        upload_action=url_for("create_event"),
    )


@app.route("/events/<int:event_id>", methods=["GET"])
@login_required
def event_detail(event_id: int):
    _ensure_db_tables()
    row = Event.query.get_or_404(int(event_id))
    payload = _event_payload(row)
    return render_template(
        "module_detail.html",
        module_key="event",
        page_title=payload["title"] if payload else "Event",
        item=payload,
    )


@app.route("/reports/categories", methods=["POST"])
@login_required
def create_report_category():
    _ensure_db_tables()
    if not bool(current_user.is_admin) or _is_readonly_user(current_user):
        abort(403)
    name = " ".join(str(request.form.get("name") or "").strip().split())
    if not name:
        flash("Category name is required.", "warning")
        return redirect(url_for("reports_page"))
    existing = ReportCategory.query.filter(db.func.lower(ReportCategory.name) == name.lower()).first()
    if existing is not None:
        flash("Category already exists.", "warning")
        return redirect(url_for("reports_page", category=int(existing.id)))
    row = ReportCategory(name=name, created_by=int(current_user.id))
    db.session.add(row)
    db.session.commit()
    flash("Report category created.", "success")
    return redirect(url_for("reports_page", category=int(row.id)))


@app.route("/upload/report", methods=["POST"])
@login_required
def upload_report():
    _ensure_db_tables()
    if not bool(current_user.is_admin) or _is_readonly_user(current_user):
        abort(403)
    upload = request.files.get("file")
    upload_name = secure_filename(getattr(upload, "filename", "") or "")
    title = (request.form.get("title") or "").strip() or Path(upload_name).stem.replace("_", " ").strip() or "Untitled report"
    file_path, error = _save_module_document(upload, user_id=int(current_user.id), module_name="reports")
    if error:
        flash(error, "warning")
        return redirect(url_for("reports_page"))
    company_label = _clean_company_name(getattr(current_user, "company_name", "") or "") or "CTS Carbon Platform"
    company_row = _company_row_for_name(company_label, created_by_user_id=int(current_user.id))
    if company_row is None:
        flash("A company is required before uploading a report.", "warning")
        return redirect(url_for("reports_page"))
    category_raw = str(request.form.get("category_id") or "").strip()
    category_id = int(category_raw) if category_raw.isdigit() else None
    category_row = ReportCategory.query.get(category_id) if category_id else None
    row = Report(
        title=title,
        file_path=str(file_path),
        preview_paths="[]",
        uploaded_by=int(current_user.id),
        company_id=int(company_row.id),
        category_id=int(category_row.id) if category_row is not None else None,
    )
    db.session.add(row)
    db.session.flush()
    row.preview_paths = json.dumps(
        _generate_report_preview_paths(
            report_title=title,
            report_rel_path=str(file_path),
            report_id=int(row.id),
        )
    )
    db.session.commit()
    _create_content_feed_post(
        author_id=int(current_user.id),
        reference_type="report",
        reference_id=int(row.id),
        title=row.title,
    )
    flash("Report uploaded successfully.", "success")
    return redirect(url_for("reports_page"))


@app.route("/upload/newsletter", methods=["POST"])
@login_required
def upload_newsletter():
    _ensure_db_tables()
    if not bool(current_user.is_admin) or _is_readonly_user(current_user):
        abort(403)
    upload = request.files.get("file")
    upload_name = secure_filename(getattr(upload, "filename", "") or "")
    title = (request.form.get("title") or "").strip() or Path(upload_name).stem.replace("_", " ").strip() or "Untitled newsletter"
    file_path, error = _save_module_document(upload, user_id=int(current_user.id), module_name="newsletters")
    if error:
        flash(error, "warning")
        return redirect(url_for("newsletters_page"))
    row = Newsletter(
        title=title,
        file_path=str(file_path),
        uploaded_by=int(current_user.id),
    )
    db.session.add(row)
    db.session.commit()
    _create_content_feed_post(
        author_id=int(current_user.id),
        reference_type="newsletter",
        reference_id=int(row.id),
        title=row.title,
    )
    flash("Newsletter uploaded successfully.", "success")
    return redirect(url_for("newsletters_page"))


@app.route("/create/event", methods=["POST"])
@login_required
def create_event():
    _ensure_db_tables()
    if not bool(current_user.is_admin) or _is_readonly_user(current_user):
        abort(403)
    title = (request.form.get("title") or "").strip()
    description = (request.form.get("description") or "").strip()
    event_date_raw = (request.form.get("event_date") or "").strip()
    if not title or not description or not event_date_raw:
        flash("Title, description, and event date are required.", "warning")
        return redirect(url_for("events_page"))
    try:
        event_date = datetime.fromisoformat(event_date_raw)
    except Exception:
        flash("Use a valid event date.", "warning")
        return redirect(url_for("events_page"))
    row = Event(
        title=title,
        description=description,
        event_date=event_date,
        created_by=int(current_user.id),
    )
    db.session.add(row)
    db.session.commit()
    _create_content_feed_post(
        author_id=int(current_user.id),
        reference_type="event",
        reference_id=int(row.id),
        title=row.title,
    )
    flash("Event created successfully.", "success")
    return redirect(url_for("events_page"))


@app.route("/awards", methods=["GET"])
@login_required
def awards_index():
    _ensure_db_tables()
    rows = AwardsForm.query.order_by(AwardsForm.created_at.desc(), AwardsForm.id.desc()).all()
    return render_template(
        "awards_index.html",
        forms=[payload for payload in (_award_form_payload(row) for row in rows) if payload],
        can_manage=bool(current_user.is_admin and not _is_readonly_user(current_user)),
    )


@app.route("/awards/admin", methods=["GET", "POST"])
@login_required
def awards_admin():
    _ensure_db_tables()
    if not bool(current_user.is_admin) or _is_readonly_user(current_user):
        abort(403)
    form_id_raw = str(request.values.get("form_id") or "").strip()
    edit_form = AwardsForm.query.get(int(form_id_raw)) if form_id_raw.isdigit() else None
    if request.method == "POST":
        title = (request.form.get("title") or "").strip()
        description = (request.form.get("description") or "").strip()
        questions, question_error = _parse_awards_builder_questions(request.form.get("questions_payload"))
        if not title:
            flash("Form title is required.", "warning")
            return redirect(url_for("awards_admin", form_id=int(edit_form.id)) if edit_form else url_for("awards_admin"))
        if question_error:
            flash(question_error, "warning")
            return redirect(url_for("awards_admin", form_id=int(edit_form.id)) if edit_form else url_for("awards_admin"))
        is_new_form = edit_form is None
        if edit_form is None:
            edit_form = AwardsForm(
                title=title,
                description=description,
                created_by=int(current_user.id),
            )
            db.session.add(edit_form)
            db.session.flush()
        else:
            edit_form.title = title
            edit_form.description = description
        header_image_path, header_image_error = _save_awards_header_image(
            request.files.get("header_image"),
            user_id=int(current_user.id),
            form_id=int(edit_form.id) if getattr(edit_form, "id", None) else None,
        )
        if header_image_error:
            flash(header_image_error, "warning")
            return redirect(url_for("awards_admin", form_id=int(edit_form.id)) if edit_form else url_for("awards_admin"))
        if header_image_path:
            edit_form.header_image = header_image_path
        if not is_new_form:
            AwardsQuestion.query.filter_by(form_id=int(edit_form.id)).delete()
        for question in questions:
            db.session.add(
                AwardsQuestion(
                    form_id=int(edit_form.id),
                    question_text=str(question["question_text"]),
                    question_type=str(question["question_type"]),
                    required=bool(question["required"]),
                    options=json.dumps(question["options"]) if question["options"] else None,
                )
            )
        db.session.commit()
        if is_new_form:
            _create_content_feed_post(
                author_id=int(current_user.id),
                reference_type="award",
                reference_id=int(edit_form.id),
                title=edit_form.title,
            )
            flash("Awards form created.", "success")
        else:
            flash("Awards form updated.", "success")
        return redirect(url_for("awards_admin", form_id=int(edit_form.id)))
    form_rows = AwardsForm.query.order_by(AwardsForm.created_at.desc(), AwardsForm.id.desc()).all()
    edit_questions = (
        AwardsQuestion.query.filter_by(form_id=int(edit_form.id)).order_by(AwardsQuestion.id.asc()).all()
        if edit_form is not None
        else []
    )
    return render_template(
        "awards_admin.html",
        forms=[payload for payload in (_award_form_payload(row) for row in form_rows) if payload],
        edit_form=_award_form_payload(edit_form) if edit_form is not None else None,
        edit_questions=[payload for payload in (_awards_question_payload(row) for row in edit_questions) if payload],
        awards_question_types=AWARDS_QUESTION_TYPES,
    )


@app.route("/awards/<int:form_id>/submissions", methods=["GET"])
@login_required
def awards_submissions_page(form_id: int):
    _ensure_db_tables()
    if not bool(current_user.is_admin) or _is_readonly_user(current_user):
        abort(403)
    form_row = AwardsForm.query.get_or_404(int(form_id))
    submission_rows = (
        AwardsSubmission.query.filter_by(form_id=int(form_row.id))
        .order_by(AwardsSubmission.created_at.desc(), AwardsSubmission.id.desc())
        .all()
    )
    return render_template(
        "awards_submissions.html",
        form_item=_award_form_payload(form_row),
        submissions=[payload for payload in (_awards_submission_payload(row) for row in submission_rows) if payload],
        analytics=_awards_single_choice_analytics(int(form_row.id)),
        total_submissions=len(submission_rows),
    )


@app.route("/awards/<int:form_id>/export", methods=["GET"])
@login_required
def awards_export_csv(form_id: int):
    _ensure_db_tables()
    if not bool(current_user.is_admin) or _is_readonly_user(current_user):
        abort(403)
    form_row = AwardsForm.query.get_or_404(int(form_id))
    question_rows = AwardsQuestion.query.filter_by(form_id=int(form_row.id)).order_by(AwardsQuestion.id.asc()).all()
    submission_rows = (
        AwardsSubmission.query.filter_by(form_id=int(form_row.id))
        .order_by(AwardsSubmission.created_at.asc(), AwardsSubmission.id.asc())
        .all()
    )
    output = StringIO()
    writer = csv.writer(output)
    headers = ["Submission ID", "User", "Date"] + [str(getattr(question, "question_text", "") or f"Question {index + 1}") for index, question in enumerate(question_rows)]
    writer.writerow(headers)
    for submission in submission_rows:
        answer_map: dict[int, str] = {}
        for answer in AwardsAnswer.query.filter_by(submission_id=int(submission.id)).all():
            answer_map[int(answer.question_id)] = str(getattr(answer, "answer_text", "") or "")
        writer.writerow(
            [
                int(submission.id),
                _user_display_name(getattr(submission, "submitter", None)),
                getattr(submission, "created_at", None).strftime("%Y-%m-%d %H:%M:%S") if getattr(submission, "created_at", None) else "",
                *[answer_map.get(int(question.id), "") for question in question_rows],
            ]
        )
    csv_text = output.getvalue()
    output.close()
    filename_slug = re.sub(r"[^0-9a-zA-Z]+", "_", str(getattr(form_row, "title", "") or "awards")).strip("_") or "awards"
    return Response(
        csv_text,
        mimetype="text/csv",
        headers={"Content-Disposition": f'attachment; filename="{filename_slug}_submissions.csv"'},
    )


@app.route("/awards/<int:form_id>", methods=["GET", "POST"])
@login_required
def awards_form_page(form_id: int):
    _ensure_db_tables()
    form_row = AwardsForm.query.get_or_404(int(form_id))
    question_rows = AwardsQuestion.query.filter_by(form_id=int(form_row.id)).order_by(AwardsQuestion.id.asc()).all()
    if request.method == "POST":
        submission = AwardsSubmission(form_id=int(form_row.id), submitted_by=int(current_user.id))
        db.session.add(submission)
        db.session.flush()
        for question in question_rows:
            payload = _awards_question_payload(question) or {}
            question_type = str(payload.get("question_type") or "text")
            answer_text = ""
            if question_type == "file":
                file_answer, file_error = _save_awards_answer_file(
                    request.files.get(f"question_{int(question.id)}"),
                    user_id=int(current_user.id),
                    form_id=int(form_row.id),
                    question_id=int(question.id),
                )
                if file_error and bool(payload.get("required")):
                    db.session.rollback()
                    flash(file_error, "warning")
                    return redirect(url_for("awards_form_page", form_id=int(form_row.id)))
                answer_text = str(file_answer or "")
            else:
                answer_text = (request.form.get(f"question_{int(question.id)}") or "").strip()
            if payload.get("required") and not answer_text:
                db.session.rollback()
                flash("Please complete all required questions.", "warning")
                return redirect(url_for("awards_form_page", form_id=int(form_row.id)))
            if question_type == "single_choice" and answer_text and answer_text not in list(payload.get("options") or []):
                db.session.rollback()
                flash("One of the selected answers is invalid.", "warning")
                return redirect(url_for("awards_form_page", form_id=int(form_row.id)))
            db.session.add(
                AwardsAnswer(
                    submission_id=int(submission.id),
                    question_id=int(question.id),
                    answer_text=answer_text or None,
                )
            )
        db.session.commit()
        flash("Submission received successfully.", "success")
        return redirect(url_for("awards_form_page", form_id=int(form_row.id)))
    return render_template(
        "awards_form.html",
        form_item=_award_form_payload(form_row),
        questions=[payload for payload in (_awards_question_payload(row) for row in question_rows) if payload],
        can_manage=bool(current_user.is_admin and not _is_readonly_user(current_user)),
    )


@app.route("/api/profile/cover", methods=["POST"])
@login_required
def api_profile_cover():
    _ensure_db_tables()
    upload = request.files.get("cover_image")
    rel_cover = _save_profile_cover_file(upload, int(current_user.id))
    if not rel_cover:
        return jsonify({"ok": False, "error": "Please upload a PNG, JPG, JPEG, or WEBP image."}), 400
    current_user.cover_image = rel_cover
    db.session.commit()
    return jsonify(
        {
            "ok": True,
            "cover_url": url_for("static", filename=rel_cover, v=uuid.uuid4().hex[:8]),
        }
    )


@app.route("/api/follow/<int:user_id>", methods=["POST"])
@login_required
def api_follow_user(user_id: int):
    _ensure_db_tables()
    target_user = User.query.get_or_404(int(user_id))
    if int(target_user.id) == int(current_user.id):
        return jsonify({"error": "You cannot follow yourself."}), 400
    existing = UserFollow.query.filter_by(
        follower_id=int(current_user.id),
        following_id=int(target_user.id),
    ).first()
    if existing is None:
        db.session.add(UserFollow(follower_id=int(current_user.id), following_id=int(target_user.id)))
        db.session.commit()
    follower_count = (
        db.session.query(db.func.count(UserFollow.id))
        .filter(UserFollow.following_id == int(target_user.id))
        .scalar()
        or 0
    )
    return jsonify({"ok": True, "following": True, "follower_count": int(follower_count)})


@app.route("/api/unfollow/<int:user_id>", methods=["POST"])
@login_required
def api_unfollow_user(user_id: int):
    _ensure_db_tables()
    target_user = User.query.get_or_404(int(user_id))
    UserFollow.query.filter_by(
        follower_id=int(current_user.id),
        following_id=int(target_user.id),
    ).delete(synchronize_session=False)
    db.session.commit()
    follower_count = (
        db.session.query(db.func.count(UserFollow.id))
        .filter(UserFollow.following_id == int(target_user.id))
        .scalar()
        or 0
    )
    return jsonify({"ok": True, "following": False, "follower_count": int(follower_count)})


def _render_dashboard_admin_analytics():
    t_agg0 = time.perf_counter()
    company_filter = request.args.get('company', '').strip().lower()
    template_filter = request.args.get('template', '').strip().lower()
    include_categories = (request.args.get('include_categories', '1').strip() != '0')
    all_time = (request.args.get('all_time', '').strip() in ('1', 'true', 'on', 'yes'))
    date_from = _parse_date_yyyy_mm_dd(request.args.get('from'))
    date_to = _parse_date_yyyy_mm_dd(request.args.get('to'))

    # Default window: last 365 days (unless all_time is explicitly requested)
    if not all_time and not date_from and not date_to:
        date_to = datetime.now()
        date_from = date_to - timedelta(days=365)

    q = MappingRunSummary.query.order_by(MappingRunSummary.created_at.desc())
    if company_filter:
        q = q.filter(MappingRunSummary.company_name.ilike(f"%{company_filter}%"))
    if template_filter:
        q = q.filter(MappingRunSummary.sheet_name.ilike(f"%{template_filter}%"))

    summaries = q.all()
    # Only count the latest run per company+sheet
    latest_summaries: list[MappingRunSummary] = []
    seen_latest: set[tuple[str, str]] = set()
    for s in summaries:
        key = ((s.company_name or "").strip().lower(), (s.sheet_name or "").strip().lower())
        if not key[0] or not key[1] or key in seen_latest:
            continue
        seen_latest.add(key)
        latest_summaries.append(s)

    company_totals = defaultdict(lambda: {"total": 0.0, "scope1": 0.0, "scope2": 0.0, "scope3": 0.0, "count": 0})
    month_totals = defaultdict(float)  # YYYY-MM -> total
    year_totals = defaultdict(float)   # YYYY -> total
    category_totals = defaultdict(float)
    run_cache: dict[str, MappingRun | None] = {}

    grand = {"total": 0.0, "scope1": 0.0, "scope2": 0.0, "scope3": 0.0, "count": 0}
    for sub in latest_summaries:
        profile = _period_profile_for_summary(sub, run_cache=run_cache)
        points = list(profile.get("points") or [])
        if date_from:
            points = [p for p in points if isinstance(p.get("date"), datetime) and p["date"] >= date_from]
        if date_to:
            points = [p for p in points if isinstance(p.get("date"), datetime) and p["date"] < (date_to + timedelta(days=1))]
        if (date_from or date_to) and not points:
            continue

        total = _safe_float(sum(float(p.get("value") or 0.0) for p in points) if points else profile.get("total", 0.0))
        scope_num = int(getattr(sub, "scope", 0) or 0)
        s1 = total if scope_num == 1 else 0.0
        s2 = total if scope_num == 2 else 0.0
        s3 = total if scope_num == 3 else 0.0

        company = (sub.company_name or "").strip() or "(unknown)"
        company_totals[company]["total"] += total
        company_totals[company]["scope1"] += s1
        company_totals[company]["scope2"] += s2
        company_totals[company]["scope3"] += s3
        company_totals[company]["count"] += 1

        grand["total"] += total
        grand["scope1"] += s1
        grand["scope2"] += s2
        grand["scope3"] += s3
        grand["count"] += 1

        for point in points:
            dt = point.get("date")
            if not isinstance(dt, datetime):
                continue
            month_key = dt.strftime("%Y-%m")
            year_key = dt.strftime("%Y")
            month_totals[month_key] += float(point.get("value") or 0.0)
            year_totals[year_key] += float(point.get("value") or 0.0)

        if include_categories:
            category_totals[str(sub.sheet_name or "(unknown)")] += total

    company_rows = [
        {
            "company": k,
            "total": round(v["total"], 2),
            "scope1": round(v["scope1"], 2),
            "scope2": round(v["scope2"], 2),
            "scope3": round(v["scope3"], 2),
            "count": v["count"],
        }
        for k, v in company_totals.items()
    ]
    company_rows.sort(key=lambda r: r["total"], reverse=True)

    month_rows = [{"month": k, "total": round(v, 2)} for k, v in month_totals.items()]
    month_rows.sort(key=lambda r: r["month"])

    year_rows = [{"year": k, "total": round(v, 2)} for k, v in year_totals.items()]
    year_rows.sort(key=lambda r: r["year"])

    category_rows = [{"category": k, "total": round(v, 2)} for k, v in category_totals.items()]
    category_rows.sort(key=lambda r: r["total"], reverse=True)

    # Chart payloads (keep top N for readability)
    top_companies = company_rows[:12]
    top_categories = category_rows[:12]
    chart_payload = {
        "companies": {
            "labels": [r["company"] for r in top_companies],
            "values": [r["total"] for r in top_companies],
        },
        "months": {
            "labels": [r["month"] for r in month_rows][-24:],
            "values": [r["total"] for r in month_rows][-24:],
        },
        "years": {
            "labels": [r["year"] for r in year_rows],
            "values": [r["total"] for r in year_rows],
        },
        "categories": {
            "labels": [r["category"] for r in top_categories],
            "values": [r["total"] for r in top_categories],
        },
    }

    _perf_log("analytics_aggregation", analytics_aggregation_ms=(time.perf_counter() - t_agg0) * 1000.0)

    return render_template(
        "dashboard_admin.html",
        user=current_user,
        filters={
            "company": request.args.get("company", ""),
            "template": request.args.get("template", ""),
            "from": request.args.get("from", ""),
            "to": request.args.get("to", ""),
            "include_categories": "1" if include_categories else "0",
            "all_time": "1" if all_time else "0",
        },
        grand={
            "total": round(grand["total"], 2),
            "scope1": round(grand["scope1"], 2),
            "scope2": round(grand["scope2"], 2),
            "scope3": round(grand["scope3"], 2),
            "count": grand["count"],
            "companies": len(company_totals),
        },
        company_rows=company_rows,
        month_rows=month_rows,
        year_rows=year_rows,
        category_rows=category_rows,
        chart_payload=chart_payload,
    )


@app.route("/dashboard/analytics")
@login_required
def dashboard_analytics():
    if not current_user.is_admin:
        return redirect(url_for("dashboard"))
    t0 = time.perf_counter()
    _ensure_db_tables()
    _backfill_mapping_summaries()
    rv = _render_dashboard_admin_analytics()
    _perf_log("dashboard_analytics", page_render_ms=(time.perf_counter() - t0) * 1000.0)
    return rv


@app.route('/dashboard')
@login_required
def dashboard():
    t0 = time.perf_counter()
    t_db0 = time.perf_counter()
    _ensure_db_tables()
    if current_user.is_admin:
        _backfill_mapping_summaries()

    thirty_days_ago = datetime.now() - timedelta(days=30)
    agg = (
        db.session.query(
            func.count(MappingRun.id),
            func.sum(case((MappingRun.created_at >= thirty_days_ago, 1), else_=0)),
        )
        .filter(MappingRun.user_id == current_user.id)
        .one()
    )
    total_mapping_runs = int(agg[0] or 0)
    recent_mapping_runs_count = int(agg[1] or 0)
    mapping_runs = (
        MappingRun.query.filter_by(user_id=current_user.id)
        .order_by(MappingRun.created_at.desc())
        .limit(10)
        .all()
    )
    t_db_ms = (time.perf_counter() - t_db0) * 1000.0

    companies = _list_template_companies_for_user()
    default_company = companies[0]["key"] if companies else None

    if not current_user.is_admin:
        rk = _resolve_template_company_name(current_user.company_name or "")
    else:
        rk = _resolve_template_company_name(default_company) if default_company else None
    company_logo_rel = _company_logo_static_rel(rk) if rk else None
    klarakarbon_supported = klarakarbon_company_supported(rk or "")

    rv = render_template(
        "dashboard.html",
        user=current_user,
        companies=companies,
        default_company=default_company,
        mapping_runs=mapping_runs,
        total_mapping_runs=total_mapping_runs,
        recent_mapping_runs_count=recent_mapping_runs_count,
        company_logo_rel=company_logo_rel,
        klarakarbon_company=rk,
        klarakarbon_supported=klarakarbon_supported,
        can_run_translation=_user_can_run_translation(current_user),
    )
    _perf_log("dashboard", page_render_ms=(time.perf_counter() - t0) * 1000.0, db_ms=t_db_ms)
    return rv


@app.route("/preprocess/klarakarbon/upload", methods=["POST"])
@login_required
def upload_klarakarbon_preprocess():
    flash("Direct Klarakarbon Excel upload is disabled. Use the 'Klarakarbon' category in Data Entry.")
    return redirect(url_for("dashboard"))


def _run_klarakarbon_preprocess_job(*, job_id: str, company_name: str, run_dir: str, upload_paths: list[str]) -> dict[str, object]:
    def progress(progress_value: int, message: str) -> None:
        _update_job_progress(job_id, progress_value, message)

    _update_job_progress(job_id, 5, "Starting Klarakarbon preprocessing")
    run_klarakarbon_preprocess(
        company_name,
        Path(run_dir),
        [Path(path) for path in upload_paths],
        progress_callback=progress,
    )
    output_path = STAGE2_KLARAKARBON_DIR / "klarakarbon_summary.xlsx"
    row_count = 0
    if output_path.exists():
        df = pd.read_excel(output_path, sheet_name=0, engine="openpyxl")
        row_count = int(len(df))
    _update_job(job_id, rows=row_count)
    return {
        "ok": True,
        "company": company_name,
        "run_dir": run_dir,
        "output": str(output_path),
        "rows": row_count,
    }


@app.route("/data-sources/klarakarbon", methods=["GET", "POST"])
@login_required
def data_sources_klarakarbon():
    if not bool(getattr(current_user, "is_admin", False)):
        flash("Access denied")
        return redirect(url_for("dashboard"))

    if request.method == "POST":
        company_name = str(request.form.get("company") or "").strip()
        if company_name not in KLARAKARBON_UPLOAD_COMPANIES:
            flash("Unsupported Klarakarbon company.")
            return redirect(url_for("data_sources_klarakarbon"))
        uploads = [f for f in request.files.getlist("klarakarbon_files") if getattr(f, "filename", "")]
        if not uploads:
            flash("Please choose at least one Klarakarbon .xlsx file.")
            return redirect(url_for("data_sources_klarakarbon"))

        slug = company_slug(company_name)
        run_id = f"{datetime.now().strftime('%Y%m%d_%H%M%S')}_{uuid.uuid4().hex[:10]}"
        run_dir = FRONTEND_UPLOAD_DIR / "preprocess" / "klarakarbon" / slug / run_id
        company_input_dir = STAGE1_KLARAKARBON_UPLOAD_DIR / slug
        company_input_dir.mkdir(parents=True, exist_ok=True)
        for existing_file in company_input_dir.iterdir():
            if existing_file.is_file():
                existing_file.unlink()
        upload_paths: list[Path] = []
        for upload in uploads:
            safe_name = secure_filename(upload.filename or "")
            if not safe_name or Path(safe_name).suffix.lower() != ".xlsx":
                flash("Only .xlsx files are allowed for Klarakarbon uploads.")
                return redirect(url_for("data_sources_klarakarbon"))
            upload_path = company_input_dir / safe_name
            upload.save(str(upload_path))
            upload_paths.append(upload_path)

        validation_errors = validate_klarakarbon_uploads(company_name, upload_paths)
        if validation_errors:
            flash(validation_errors[0])
            return redirect(url_for("data_sources_klarakarbon"))

        job_id = run_in_background(
            "preprocess",
            company_name,
            _run_klarakarbon_preprocess_job,
            company_name=company_name,
            run_dir=str(run_dir),
            upload_paths=[str(path) for path in upload_paths],
            job_user_id=int(current_user.id),
            job_user_email=str(getattr(current_user, "email", "") or ""),
        )
        flash(f"Klarakarbon preprocessing started for {company_name}. Job ID: {job_id}")
        return redirect(url_for("data_sources_klarakarbon"))

    company_cards = [
        {
            "name": company,
            "slug": company_slug(company),
            "upload_dir": str(STAGE1_KLARAKARBON_UPLOAD_DIR / company_slug(company)),
        }
        for company in KLARAKARBON_UPLOAD_COMPANIES
    ]
    return render_template(
        "data_source_klarakarbon.html",
        user=current_user,
        companies=company_cards,
        output_exists=(STAGE2_KLARAKARBON_DIR / "klarakarbon_summary.xlsx").exists(),
    )


def _user_can_access_company_file(company_file: Path) -> bool:
    if not company_file:
        return False
    if bool(getattr(current_user, "is_admin", False)):
        return True
    user_file = _resolve_company_file(getattr(current_user, "company_name", "") or "")
    if not user_file:
        return False
    try:
        return user_file.resolve() == company_file.resolve()
    except Exception:
        return str(user_file) == str(company_file)


@app.route("/api/excel_schema/companies", methods=["GET"])
@login_required
def api_excel_schema_companies():
    mode = _current_template_mode()
    if current_user.is_admin:
        companies = list(COMPANIES)
        return jsonify({"companies": companies, "template_mode": mode})

    company_name = _resolve_template_company_name(getattr(current_user, "company_name", "") or "")
    companies = [company_name] if company_name else []
    return jsonify({"companies": companies, "template_mode": mode})


@app.route("/api/excel_schema/sheets", methods=["GET"])
@login_required
def api_excel_schema_sheets():
    company = request.args.get("company", "").strip()
    if not _user_can_access_company(company):
        return jsonify({"error": "Access denied"}), 403
    resolved_company = _resolve_template_company_name(company)
    if not resolved_company:
        return jsonify({"error": "Company not found"}), 404
    bundle = _template_bundle_for_company(resolved_company)
    visible_templates = [str(item.get("sheet_name") or "") for item in (bundle.get("visible_templates") or []) if str(item.get("sheet_name") or "").strip()]
    print("TEMPLATES SOURCE:", "2026 ONLY")
    print("VISIBLE TEMPLATES:", visible_templates)
    return jsonify(
        {
            "company": resolved_company,
            "template_mode": bundle.get("template_mode"),
            "sheets": visible_templates,
            "visible_templates": bundle.get("visible_templates", []),
            "enabled_categories": bundle.get("enabled_categories", []),
            "disabled_categories": bundle.get("disabled_categories", []),
            "metadata_validation": bundle.get("metadata_validation", []),
        }
    )


@app.route("/debug/templates", methods=["GET"])
@login_required
def debug_templates():
    _ensure_db_tables()
    requested_company = (request.args.get("company") or "").strip()
    company_name = requested_company or _resolve_template_company_name(getattr(current_user, "company_name", "") or "") or ""
    if company_name and not _user_can_access_company(company_name):
        return jsonify({"error": "Access denied"}), 403
    bundle = _template_bundle_for_company(company_name)
    return jsonify(
        {
            "company": company_name,
            "template_mode": bundle.get("template_mode"),
            "loaded_templates_count": len(TEMPLATE_REGISTRY.templates_2026),
            "visible_templates": [item.get("sheet_name") for item in (bundle.get("visible_templates") or [])],
            "missing_metadata": bundle.get("metadata_validation", []),
        }
    )


@app.route("/api/excel_schema/headers", methods=["GET"])
@login_required
def api_excel_schema_headers():
    company = request.args.get("company", "").strip()
    sheet = request.args.get("sheet", "").strip()
    if not sheet:
        return jsonify({"error": "sheet is required"}), 400
    if _is_hidden_schema_sheet(sheet):
        return jsonify({"error": "This sheet is not available for web data entry"}), 403

    if not _user_can_access_company(company):
        return jsonify({"error": "Access denied"}), 403
    resolved_company = _resolve_template_company_name(company)
    if not resolved_company:
        return jsonify({"error": "Company not found"}), 404
    resolved_sheet = _resolve_template_sheet_name(resolved_company, sheet)
    if not resolved_sheet:
        return jsonify({"error": "Sheet not found"}), 404
    headers = _get_template_sheet_headers(resolved_company, resolved_sheet)
    bundle = _template_bundle_for_company(resolved_company)
    metadata_status = [
        item for item in (bundle.get("metadata_validation") or [])
        if _normalize_template_key(str(item.get("sheet_name") or "")) == _normalize_template_key(resolved_sheet)
    ]
    return jsonify(
        {
            "company": resolved_company,
            "sheet": resolved_sheet,
            "template_mode": bundle.get("template_mode"),
            "header_row": 1,
            "headers": headers,
            "rules": {h: _infer_column_rule(h) for h in headers},
            "enabled_categories": bundle.get("enabled_categories", []),
            "disabled_categories": bundle.get("disabled_categories", []),
            "metadata_status": metadata_status,
        }
    )


@app.route("/api/data_entry/rows", methods=["GET"])
@login_required
def api_data_entry_rows():
    company = request.args.get("company", "").strip()
    sheet = request.args.get("sheet", "").strip()
    if not company or not sheet:
        return jsonify({"error": "company and sheet are required"}), 400
    if not _user_can_access_company(company):
        return jsonify({"error": "Access denied"}), 403

    resolved_company = _resolve_template_company_name(company)
    if not resolved_company:
        return jsonify({"error": "Company not found"}), 404
    resolved_sheet = _resolve_template_sheet_name(resolved_company, sheet)
    if not resolved_sheet:
        return jsonify({"error": "Sheet not found"}), 404

    headers, _rules = _get_data_entry_template_schema(resolved_company, resolved_sheet)
    bundle = _template_bundle_for_company(resolved_company)
    return jsonify(
        {
            "company": resolved_company,
            "sheet": resolved_sheet,
            "template_mode": bundle.get("template_mode"),
            "headers": headers,
            "rows": _load_data_entry_grid_rows(resolved_company, resolved_sheet, headers),
        }
    )


@app.route("/api/data-entry/site-tags", methods=["GET"])
@login_required
def api_data_entry_site_tags():
    from frontend.services.site_tag_service import get_site_tags_for_company

    company = request.args.get("company", "").strip()
    if not company:
        return jsonify({"error": "company is required"}), 400
    if not _user_can_access_company(company):
        return jsonify({"error": "Access denied"}), 403
    resolved_company = _resolve_template_company_name(company) or company
    tags = get_site_tags_for_company(resolved_company)
    if not tags:
        tags = get_site_tags_for_company(company)
    return jsonify({"ok": True, "site_tags": tags})


@app.route("/api/data-entry/reporting-periods", methods=["GET"])
@login_required
def api_data_entry_reporting_periods():
    from frontend.services.reporting_period_service import get_reporting_period_options_2026, reporting_period_sort_key

    periods = list(get_reporting_period_options_2026())
    periods = sorted(set(periods), key=reporting_period_sort_key)
    return jsonify({"ok": True, "periods": periods})


@app.route("/api/ccc/import-to-data-entry", methods=["POST"])
@login_required
def api_ccc_import_to_data_entry():
    """Queue background job: CCC purchase orders → Data Entry (Scope 3 Cat 1)."""
    if not _user_can_run_ccc_data_entry_import(current_user):
        return jsonify({"error": "Forbidden", "detail": "CCC Data Entry import requires admin, super_admin, or owner role."}), 403
    from frontend.services.ccc_data_entry_import_service import CCC_TARGET_SHEET
    from frontend.services.site_tag_service import resolve_registered_project

    payload = request.get_json(silent=True) or {}
    projects = payload.get("projects")
    if not isinstance(projects, list) or not projects:
        return jsonify({"error": "projects must be a non-empty array"}), 400

    labels = [str(p).strip() for p in projects if str(p).strip()]
    if not labels:
        return jsonify({"error": "No valid project labels"}), 400

    unknown: list[str] = []
    for lab in labels:
        reg = resolve_registered_project(lab)
        if not reg:
            unknown.append(lab)
            continue
        rc = str(reg.get("responsible_company") or "").strip()
        resolved_company = _resolve_template_company_name(rc) or rc
        if not resolved_company:
            unknown.append(lab)
            continue
        if not _user_can_access_company(resolved_company):
            return jsonify({"error": "Access denied for one or more target companies", "project": lab}), 403

        resolved_sheet = _resolve_template_sheet_name(resolved_company, CCC_TARGET_SHEET)
        if not resolved_sheet:
            return jsonify({"error": "Target sheet not available for company", "company": resolved_company}), 400
        hdrs = _get_template_sheet_headers(resolved_company, resolved_sheet)
        if not hdrs:
            return jsonify({"error": "Target sheet has no template headers", "company": resolved_company}), 400

    if unknown:
        return jsonify(
            {
                "error": "Unknown project(s): not found in site_tags_2026 registry",
                "unknown_projects": unknown[:50],
            }
        ), 400

    active_ccc = _active_job_id_for_job_type("ccc_import")
    if active_ccc:
        return jsonify(
            {
                "error": "CCC import already running",
                "detail": "Wait for the current CCC import job to finish before starting another.",
                "job_id": active_ccc,
            }
        ), 409

    tm = _current_template_mode()
    job_id = run_in_background(
        "ccc_import",
        "CCC API",
        _run_ccc_import_job,
        project_labels=labels,
        user_id=int(getattr(current_user, "id", 0) or 0) or None,
        template_mode=tm,
        job_user_id=int(getattr(current_user, "id", 0) or 0) or None,
        job_user_email=str(getattr(current_user, "email", "") or ""),
    )
    return jsonify({"ok": True, "job_id": job_id})


@app.route("/api/evidence/row-summary", methods=["GET"])
@login_required
def api_evidence_row_summary():
    company = request.args.get("company", "").strip()
    sheet = request.args.get("sheet", "").strip()
    if not company or not sheet:
        return jsonify({"error": "company and sheet are required"}), 400
    if not _user_can_access_company(company):
        return jsonify({"error": "Access denied"}), 403
    resolved_company = _resolve_template_company_name(company)
    if not resolved_company:
        return jsonify({"error": "Company not found"}), 404
    resolved_sheet = _resolve_template_sheet_name(resolved_company, sheet)
    if not resolved_sheet:
        return jsonify({"error": "Sheet not found"}), 404
    counts = _evidence_row_attachment_counts(resolved_company, resolved_sheet)
    shared_rows = _evidence_row_shared_invoice_flags(resolved_company, resolved_sheet)
    return jsonify(
        {
            "company": resolved_company,
            "sheet": resolved_sheet,
            "counts": counts,
            "shared_invoice_rows": shared_rows,
        }
    )


@app.route("/api/evidence/for-row", methods=["GET"])
@login_required
def api_evidence_for_row():
    company = request.args.get("company", "").strip()
    sheet = request.args.get("sheet", "").strip()
    entry_group = str(request.args.get("entry_group") or "").strip()
    if not company or not sheet or not entry_group:
        return jsonify({"error": "company, sheet, and entry_group are required"}), 400
    if not _user_can_access_company(company):
        return jsonify({"error": "Access denied"}), 403
    resolved_company = _resolve_template_company_name(company)
    if not resolved_company:
        return jsonify({"error": "Company not found"}), 404
    resolved_sheet = _resolve_template_sheet_name(resolved_company, sheet)
    if not resolved_sheet:
        return jsonify({"error": "Sheet not found"}), 404
    headers, _rules = _get_data_entry_template_schema(resolved_company, resolved_sheet)
    if not _data_entry_grid_row_exists(resolved_company, resolved_sheet, headers, entry_group):
        return jsonify({"error": "Row not found"}), 404

    links = (
        db.session.query(DataEntryEvidence, EvidenceFile)
        .join(EvidenceFile, EvidenceFile.id == DataEntryEvidence.evidence_file_id)
        .filter(
            DataEntryEvidence.company_name == resolved_company,
            DataEntryEvidence.sheet_name == resolved_sheet,
            DataEntryEvidence.entry_group == entry_group,
            EvidenceFile.is_deleted.is_(False),
        )
        .order_by(DataEntryEvidence.linked_at.desc())
        .all()
    )
    uploader_ids = sorted({int(ef.uploaded_by or 0) for _, ef in links if ef.uploaded_by})
    uploaders: dict[int, str] = {}
    if uploader_ids:
        for u in User.query.filter(User.id.in_(uploader_ids)).all():
            uploaders[int(u.id)] = str(u.email or "").strip()

    items: list[dict[str, object]] = []
    for link, ef in links:
        row = _serialize_evidence_public(ef, link_id=int(link.id))
        row["uploaded_by_email"] = uploaders.get(int(ef.uploaded_by or 0), "")
        items.append(row)
    return jsonify({"company": resolved_company, "sheet": resolved_sheet, "entry_group": entry_group, "items": items})


@app.route("/api/evidence/search", methods=["GET"])
@login_required
def api_evidence_search():
    company = request.args.get("company", "").strip()
    q_raw = (request.args.get("q") or "").strip()
    if not company:
        return jsonify({"error": "company is required"}), 400
    if not _user_can_access_company(company):
        return jsonify({"error": "Access denied"}), 403
    resolved_company = _resolve_template_company_name(company)
    if not resolved_company:
        return jsonify({"error": "Company not found"}), 404
    try:
        limit = int(request.args.get("limit") or 25)
    except Exception:
        limit = 25
    limit = max(1, min(limit, 100))
    sort_raw = (request.args.get("sort") or "newest").strip().lower()
    sort_key = sort_raw if sort_raw in ("newest", "filename", "most_linked") else "newest"

    query = EvidenceFile.query.filter(
        EvidenceFile.company_name == resolved_company,
        EvidenceFile.is_deleted.is_(False),
        EvidenceFile.processing_status == "ready",
    )
    q_lower = q_raw.lower()
    if q_lower:
        safe_frag = q_raw.replace("\\", "").replace("%", "").replace("_", "")
        if safe_frag.strip():
            query = query.filter(func.lower(EvidenceFile.original_filename).like(f"%{safe_frag.lower()}%"))

    if sort_key == "filename":
        query = query.order_by(func.lower(EvidenceFile.original_filename).asc(), EvidenceFile.uploaded_at.desc())
    elif sort_key == "most_linked":
        lc = (
            db.session.query(
                DataEntryEvidence.evidence_file_id.label("eid"),
                func.count(DataEntryEvidence.id).label("lcnt"),
            )
            .group_by(DataEntryEvidence.evidence_file_id)
            .subquery()
        )
        query = query.outerjoin(lc, EvidenceFile.id == lc.c.eid).order_by(
            desc(func.coalesce(lc.c.lcnt, 0)),
            EvidenceFile.uploaded_at.desc(),
        )
    else:
        query = query.order_by(EvidenceFile.uploaded_at.desc())

    items = query.limit(limit).all()

    uploader_ids = sorted({int(e.uploaded_by) for e in items if e.uploaded_by})
    uploaders: dict[int, str] = {}
    if uploader_ids:
        for u in User.query.filter(User.id.in_(uploader_ids)).all():
            uploaders[int(u.id)] = str(u.email or "").strip()

    enriched = []
    for ef in items:
        row = _serialize_evidence_public(ef)
        uid = int(ef.uploaded_by or 0)
        row["uploaded_by_email"] = uploaders.get(uid, "")
        enriched.append(row)

    return jsonify({"company": resolved_company, "items": enriched})


@app.route("/api/evidence/upload", methods=["POST"])
@login_required
def api_evidence_upload():
    from frontend.evidence_processing import MAX_UPLOAD_BYTES, sha256_file, validate_upload_file
    from frontend.storage import get_evidence_storage
    from frontend.storage.providers.local import LocalStorageProvider

    company = (request.form.get("company") or "").strip()
    if not company:
        return jsonify({"error": "company is required"}), 400
    if not _user_can_access_company(company):
        return jsonify({"error": "Access denied"}), 403
    resolved_company = _resolve_template_company_name(company)
    if not resolved_company:
        return jsonify({"error": "Company not found"}), 404
    if "file" not in request.files:
        return jsonify({"error": "file is required"}), 400
    uf = request.files["file"]
    raw_name = secure_filename(uf.filename or "") or "upload"
    suffix = Path(raw_name).suffix.lower()[:12]

    storage = get_evidence_storage()
    if not isinstance(storage, LocalStorageProvider):
        return jsonify({"error": "Evidence uploads require filesystem-backed storage in this release"}), 501

    sid = uuid.uuid4().hex
    staging_fn = f"{sid}{suffix}" if suffix else sid
    staging_rel = storage.generate_path("evidence", "_staging", staging_fn)
    staging_abs = storage.absolute_path_for_write(staging_rel)
    uf.save(str(staging_abs))

    def _staging_cleanup() -> None:
        storage.delete_file(staging_rel)
        try:
            staging_abs.unlink(missing_ok=True)
        except Exception:
            pass

    try:
        sz = staging_abs.stat().st_size
        if sz > MAX_UPLOAD_BYTES:
            _staging_cleanup()
            return jsonify(ok=False, error=_EVIDENCE_MAX_SIZE_USER_MSG), 400
        normalized_ext, mime = validate_upload_file(staging_abs, suffix.lstrip("."), sz)
    except ValueError as exc:
        _staging_cleanup()
        return jsonify({"error": str(exc)}), 400

    digest = sha256_file(staging_abs)

    with _evidence_company_digest_lock(resolved_company, digest):
        existing = EvidenceFile.query.filter_by(
            sha256_hash=digest, company_name=resolved_company, is_deleted=False
        ).first()
        if existing is not None:
            if existing.processing_status == "ready":
                _staging_cleanup()
                _evidence_log(
                    "duplicate_reused",
                    evidence_file_id=int(existing.id),
                    company=resolved_company,
                    status="ready",
                )
                return jsonify(
                    {
                        "ok": True,
                        "deduplicated": True,
                        "evidence_file": _serialize_evidence_public(existing),
                        "processing_job_id": None,
                    }
                )
            if existing.processing_status == "pending":
                _staging_cleanup()
                _evidence_log(
                    "duplicate_reused",
                    evidence_file_id=int(existing.id),
                    company=resolved_company,
                    status="pending",
                )
                return jsonify(
                    {
                        "ok": True,
                        "deduplicated": True,
                        "pending": True,
                        "evidence_file": _serialize_evidence_public(existing),
                        "processing_job_id": None,
                    }
                )
            if existing.processing_status == "failed":
                existing.processing_status = "pending"
                existing.processing_error = None
                db.session.commit()
                jid = run_in_background(
                    "evidence_processing",
                    resolved_company,
                    _run_evidence_processing_job,
                    evidence_file_id=int(existing.id),
                    staging_rel=staging_rel,
                    upload_company=resolved_company,
                    job_user_id=current_user.id,
                    job_user_email=str(getattr(current_user, "email", "") or ""),
                )
                _evidence_log(
                    "optimization_retry_enqueued",
                    evidence_file_id=int(existing.id),
                    company=resolved_company,
                    job_id=jid,
                )
                return jsonify(
                    {
                        "ok": True,
                        "retry": True,
                        "evidence_file": _serialize_evidence_public(existing),
                        "processing_job_id": jid,
                    }
                )

        now_u = datetime.utcnow()
        ef = EvidenceFile(
            company_name=resolved_company,
            original_filename=str(raw_name)[:500],
            stored_filename=f"{digest}.pending",
            file_extension=normalized_ext,
            mime_type=mime,
            sha256_hash=digest,
            file_size_original=int(sz),
            uploaded_by=current_user.id,
            upload_source="data_entry",
            processing_status="pending",
            storage_path="",
            is_orphaned=True,
            orphaned_at=now_u,
            relation_count=0,
        )
        db.session.add(ef)
        try:
            db.session.commit()
        except IntegrityError:
            db.session.rollback()
            _staging_cleanup()
            dup = EvidenceFile.query.filter_by(sha256_hash=digest, company_name=resolved_company).first()
            if dup is None:
                return jsonify({"error": "Duplicate handling failed"}), 409
            _evidence_log(
                "duplicate_race_resolved",
                evidence_file_id=int(dup.id),
                company=resolved_company,
            )
            return jsonify(
                {
                    "ok": True,
                    "deduplicated": True,
                    "evidence_file": _serialize_evidence_public(dup),
                    "processing_job_id": None,
                }
            )

        jid = run_in_background(
            "evidence_processing",
            resolved_company,
            _run_evidence_processing_job,
            evidence_file_id=int(ef.id),
            staging_rel=staging_rel,
            upload_company=resolved_company,
            job_user_id=current_user.id,
            job_user_email=str(getattr(current_user, "email", "") or ""),
        )
        _evidence_log("upload_accepted", evidence_file_id=int(ef.id), company=resolved_company, job_id=jid)
    return jsonify({"ok": True, "evidence_file": _serialize_evidence_public(ef), "processing_job_id": jid})


@app.route("/api/evidence/link", methods=["POST"])
@login_required
def api_evidence_link():
    payload = request.get_json(silent=True) or {}
    company = (payload.get("company") or "").strip()
    sheet = (payload.get("sheet") or "").strip()
    groups_payload = payload.get("entry_groups")
    entry_single = str(payload.get("entry_group") or "").strip()
    try:
        evidence_file_id = int(payload.get("evidence_file_id") or 0)
    except Exception:
        evidence_file_id = 0

    groups: list[str] = []
    if isinstance(groups_payload, list):
        seen_g: set[str] = set()
        for raw in groups_payload:
            g = str(raw or "").strip()
            if not g or g in seen_g:
                continue
            seen_g.add(g)
            groups.append(g)
    if not groups and entry_single:
        groups = [entry_single]

    if not company or not sheet or not groups or evidence_file_id <= 0:
        return (
            jsonify({"error": "company, sheet, entry_groups (or entry_group), and evidence_file_id are required"}),
            400,
        )
    if not _user_can_access_company(company):
        return jsonify({"error": "Access denied"}), 403
    resolved_company = _resolve_template_company_name(company)
    if not resolved_company:
        return jsonify({"error": "Company not found"}), 404
    resolved_sheet = _resolve_template_sheet_name(resolved_company, sheet)
    if not resolved_sheet:
        return jsonify({"error": "Sheet not found"}), 404
    headers, _rules = _get_data_entry_template_schema(resolved_company, resolved_sheet)

    for eg in groups:
        if not _data_entry_grid_row_exists(resolved_company, resolved_sheet, headers, eg):
            return jsonify({"error": "Row not found", "entry_group": eg}), 404

    ef = db.session.get(EvidenceFile, evidence_file_id)
    if not _user_can_access_evidence_file(ef) or str(ef.company_name).strip() != resolved_company:
        return jsonify({"error": "Evidence not found"}), 404
    if ef.processing_status != "ready":
        return jsonify({"error": "Evidence is not ready yet"}), 409

    results: list[dict[str, object]] = []
    for eg in groups:
        exists_link = DataEntryEvidence.query.filter_by(
            company_name=resolved_company,
            sheet_name=resolved_sheet,
            entry_group=eg,
            evidence_file_id=evidence_file_id,
        ).first()
        if exists_link:
            results.append(
                {"entry_group": eg, "link_id": int(exists_link.id), "deduplicated": True},
            )
            continue

        link = DataEntryEvidence(
            company_name=resolved_company,
            sheet_name=resolved_sheet,
            entry_group=eg,
            evidence_file_id=evidence_file_id,
            linked_by=current_user.id,
        )
        db.session.add(link)
        db.session.flush()
        _sync_evidence_orphan_metadata(evidence_file_id, commit=False)
        results.append({"entry_group": eg, "link_id": int(link.id), "deduplicated": False})

    db.session.commit()
    _evidence_log(
        "evidence_linked",
        evidence_file_id=evidence_file_id,
        link_count=len(results),
        entry_groups=[str(r.get("entry_group") or "") for r in results],
    )
    first_id = int(results[0]["link_id"]) if results else 0
    return jsonify({"ok": True, "links": results, "link_id": first_id})


def _api_evidence_unlink_impl() -> tuple[dict, int]:
    payload = request.get_json(silent=True) or {}
    try:
        link_id = int(payload.get("link_id") or 0)
    except Exception:
        link_id = 0
    if link_id <= 0:
        return {"error": "link_id is required"}, 400
    link = db.session.get(DataEntryEvidence, link_id)
    if link is None:
        return {"error": "Link not found"}, 404
    if not _user_can_access_company(link.company_name):
        return {"error": "Access denied"}, 403
    evidence_file_id = int(link.evidence_file_id)
    db.session.delete(link)
    db.session.flush()
    _sync_evidence_orphan_metadata(evidence_file_id, commit=False)
    db.session.commit()
    _evidence_log("evidence_unlinked", evidence_file_id=evidence_file_id, link_id=link_id)
    return {"ok": True}, 200


@app.route("/api/evidence/unlink", methods=["DELETE", "POST"])
@login_required
def api_evidence_unlink():
    body, status = _api_evidence_unlink_impl()
    return jsonify(body), status


@app.route("/api/evidence/<int:evidence_id>/audit-summary", methods=["GET"])
@login_required
def api_evidence_audit_summary(evidence_id: int):
    ef = db.session.get(EvidenceFile, evidence_id)
    if not _user_can_access_evidence_file(ef):
        return jsonify({"error": "Not found"}), 404

    uid = int(ef.uploaded_by or 0)
    uploaded_by_email = ""
    if uid:
        u_row = db.session.get(User, uid)
        if u_row is not None:
            uploaded_by_email = str(u_row.email or "").strip()

    breakdown_rows = (
        db.session.query(
            DataEntryEvidence.sheet_name,
            func.count(func.distinct(DataEntryEvidence.entry_group)),
        )
        .filter(
            DataEntryEvidence.evidence_file_id == evidence_id,
            DataEntryEvidence.company_name == ef.company_name,
        )
        .group_by(DataEntryEvidence.sheet_name)
        .order_by(DataEntryEvidence.sheet_name.asc())
        .all()
    )
    sheet_breakdown = [{"sheet_name": str(sn or ""), "row_count": int(rc or 0)} for sn, rc in breakdown_rows]

    return jsonify(
        {
            "evidence_file": _serialize_evidence_public(ef),
            "uploaded_by_email": uploaded_by_email,
            "link_row_count": int(getattr(ef, "relation_count", 0) or 0),
            "sheet_breakdown": sheet_breakdown,
        }
    )


@app.route("/api/evidence/<int:evidence_id>", methods=["GET"])
@login_required
def api_evidence_detail(evidence_id: int):
    ef = db.session.get(EvidenceFile, evidence_id)
    if not _user_can_access_evidence_file(ef):
        return jsonify({"error": "Not found"}), 404
    return jsonify({"evidence_file": _serialize_evidence_public(ef)})


@app.route("/api/evidence/<int:evidence_id>/download", methods=["GET"])
@login_required
def api_evidence_download(evidence_id: int):
    ef = db.session.get(EvidenceFile, evidence_id)
    if not _user_can_access_evidence_file(ef) or ef.processing_status != "ready":
        abort(404)
    disk = _safe_evidence_disk_path(ef.storage_path)
    if not disk or not disk.is_file():
        abort(404)
    mt = str(ef.mime_type or "application/octet-stream")
    return send_file(disk, as_attachment=True, download_name=ef.original_filename or ef.stored_filename, mimetype=mt)


@app.route("/api/evidence/<int:evidence_id>/preview", methods=["GET"])
@login_required
def api_evidence_preview(evidence_id: int):
    ef = db.session.get(EvidenceFile, evidence_id)
    if not _user_can_access_evidence_file(ef) or ef.processing_status != "ready":
        abort(404)
    disk = _safe_evidence_disk_path(ef.storage_path)
    if not disk or not disk.is_file():
        abort(404)
    mt = str(ef.mime_type or "application/octet-stream")
    return send_file(disk, as_attachment=False, mimetype=mt, download_name=ef.original_filename or ef.stored_filename)


@app.route("/api/evidence/<int:evidence_id>/thumbnail", methods=["GET"])
@login_required
def api_evidence_thumbnail(evidence_id: int):
    ef = db.session.get(EvidenceFile, evidence_id)
    if not _user_can_access_evidence_file(ef) or ef.processing_status != "ready":
        abort(404)
    thumb = _safe_evidence_disk_path(ef.thumbnail_storage_path)
    if not thumb or not thumb.is_file():
        abort(404)
    return send_file(thumb, as_attachment=False, mimetype="image/webp")


@app.route("/api/excel_schema/download", methods=["GET"])
@login_required
def api_excel_schema_download():
    company = request.args.get("company", "").strip()
    if not _user_can_access_company(company):
        return jsonify({"error": "Access denied"}), 403
    resolved_company = _resolve_template_company_name(company)
    if not resolved_company:
        return jsonify({"error": "Company not found"}), 404
    bio = _build_template_workbook(resolved_company)
    filename = secure_filename(f"{resolved_company}_template.xlsx") or "template.xlsx"
    return send_file(
        bio,
        as_attachment=True,
        download_name=filename,
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )


@app.route("/api/excel_schema/save", methods=["POST"])
@login_required
def api_excel_schema_save():
    payload = request.get_json(silent=True) or {}
    company = (payload.get("company") or "").strip()
    sheet = (payload.get("sheet") or "").strip()
    rows = payload.get("rows") or []

    if not company or not sheet:
        return jsonify({"error": "company and sheet are required"}), 400
    if not isinstance(rows, list):
        return jsonify({"error": "rows must be a list"}), 400
    if _is_hidden_schema_sheet(sheet):
        return jsonify({"error": "This sheet is not available for web data entry"}), 403
    if not _user_can_access_company(company):
        return jsonify({"error": "Access denied"}), 403
    resolved_company = _resolve_template_company_name(company)
    if not resolved_company:
        return jsonify({"error": "Company not found"}), 404
    resolved_sheet = _resolve_template_sheet_name(resolved_company, sheet)
    if not resolved_sheet:
        return jsonify({"error": "Sheet not found"}), 404
    headers, _rules = _get_data_entry_template_schema(resolved_company, resolved_sheet)
    if not headers:
        return jsonify({"error": "Sheet not found"}), 404
    template_mode = _current_template_mode()
    normalized_rows, validation_errors = _normalize_data_entry_rows(headers, rows)
    if validation_errors:
        return jsonify({"error": validation_errors[0], "validation_errors": validation_errors[:20]}), 400
    requirement_errors = _validate_data_entry_row_requirements(headers, normalized_rows)
    if requirement_errors:
        return jsonify({"error": requirement_errors[0], "validation_errors": requirement_errors[:20]}), 400
    try:
        result = _upsert_data_entries(resolved_company, resolved_sheet, headers, normalized_rows)
        db.session.commit()
    except Exception as e:
        db.session.rollback()
        return jsonify({"error": f"Save failed: {e}"}), 500

    saved_rows_count = int(result.get("saved_rows_count") or 0)
    duplicate_rows_count = int(result.get("duplicate_rows_count") or 0)
    saved_entry_groups = list(result.get("saved_entry_groups") or [])
    if saved_rows_count > 0:
        card_meta = _mapping_card_payload_for_pair(resolved_company, resolved_sheet)
        if not card_meta:
            u = current_user
            card_meta = {
                "company_name": resolved_company,
                "uploaded_by_user": _display_name_for_user(u),
                "uploaded_by_user_id": int(getattr(u, "id", 0) or 0),
                "uploaded_by_job_title": _user_professional_title(u),
                "uploaded_by_has_profile_photo": bool((getattr(u, "profile_photo_path", None) or "").strip()),
                "upload_timestamp": datetime.utcnow().strftime("%Y-%m-%d %H:%M"),
                "category": resolved_sheet,
                "row_count": saved_rows_count,
                "mapping_status": "",
                "mapping_state": "",
                "mapped_by_admin": "",
                "mapping_timestamp": "",
                "mapped": False,
            }
        _create_user_notification(
            current_user.id,
            title="Data upload completed",
            message=f"{resolved_company} uploaded new data successfully.",
            notification_type="success",
            link=url_for("home"),
            mapping_card=card_meta,
            feed_event="data_upload",
            feed_company=resolved_company,
            feed_timestamp=datetime.utcnow(),
        )
    return jsonify(
        {
            "ok": True,
            "company": resolved_company,
            "sheet": resolved_sheet,
            "saved_rows": saved_rows_count,
            "saved_rows_count": saved_rows_count,
            "duplicate_rows_count": duplicate_rows_count,
            "saved_entry_groups": saved_entry_groups,
            "message": f"{saved_rows_count} rows saved, {duplicate_rows_count} duplicates skipped",
        }
    )


@app.route("/api/job-status/<job_id>", methods=["GET"])
@app.route("/job-status/<job_id>", methods=["GET"])
@login_required
def api_job_status(job_id: str):
    jid = str(job_id or "").strip()
    print(f"[JOB STATUS CHECK] {jid}")
    job = _job_snapshot(jid)
    if not _user_can_access_job(job, current_user):
        print(f"[JOB ERROR] Job not found: {jid}")
        return jsonify(
            {
                "job_id": jid,
                "status": "not_found",
                "progress": 0,
                "message": "Job not found or server restarted",
                "error": None,
            }
        )
    return jsonify(_serialize_job(job))


@app.route("/api/jobs", methods=["GET"])
@app.route("/job-history", methods=["GET"])
@login_required
def api_job_history():
    _cleanup_old_jobs()
    with _JOBS_LOCK:
        visible = [dict(job) for job in jobs.values() if _user_can_access_job(job, current_user)]
    visible.sort(key=lambda item: str(item.get("created_at") or ""), reverse=True)
    history = [_serialize_job(job) for job in visible[:100]]
    return jsonify({"jobs": history})


@app.route("/cancel-job/<job_id>", methods=["POST"])
@login_required
def api_cancel_job(job_id: str):
    jid = str(job_id or "").strip()
    job = _job_snapshot(jid)
    if not _user_can_access_job(job, current_user):
        return jsonify({"error": "Job not found"}), 404
    with _JOBS_LOCK:
        live = jobs.get(jid)
        if not live:
            return jsonify({"error": "Job not found"}), 404
        if str(live.get("status") or "") in {"completed", "failed", "cancelled"}:
            return jsonify({"ok": True, "job_id": jid, "status": live.get("status")})
        live["cancel_requested"] = True
        live["message"] = "Cancellation requested"
    print(f"[JOB] Cancelled {jid}")
    return jsonify({"ok": True, "job_id": jid, "status": "cancelling"})


def _run_translation_job(
    *,
    job_id: str,
    resolved_company: str,
    translation_plan: list[dict[str, object]],
    user_id: int | None = None,
) -> dict[str, object]:
    translation_mod = _import_translation_module()
    target_companies = set(getattr(translation_mod, "TARGET_COMPANIES", set()))
    run_translation_fn = getattr(translation_mod, "run_translation")

    if resolved_company not in target_companies:
        _update_job_progress(job_id, 100, "Translation skipped: company is not configured")
        return {
            "ok": True,
            "company": resolved_company,
            "message": "Translation skipped: company is not configured for translation",
            "rows_affected": 0,
            "cells_updated": 0,
            "sheets": [],
        }

    print(f"[TRANSLATION] Started for {resolved_company}")
    _update_job_progress(job_id, 5, f"Preparing translation for {resolved_company}")
    total_rows_affected = 0
    total_cells_updated = 0
    translated_sheets: list[dict[str, object]] = []
    items = list(translation_plan or [])
    total_items = max(1, len(items))

    for idx, item in enumerate(items, start=1):
        _raise_if_job_cancelled(job_id)
        resolved_sheet = str(item.get("sheet") or "").strip()
        headers = list(item.get("headers") or [])
        columns = tuple(str(c) for c in (item.get("columns") or []))
        if not resolved_sheet or not headers:
            continue
        if not headers:
            continue
        df = _load_data_entries_dataframe_no_request(resolved_company, resolved_sheet, headers)
        if df.empty:
            continue

        base_progress = 10 + int(((idx - 1) / total_items) * 75)
        _update_job_progress(job_id, base_progress, f"Translating {resolved_sheet}...")
        translated_df = run_translation_fn(df, resolved_sheet, resolved_company)
        _raise_if_job_cancelled(job_id)
        rows_affected = int(getattr(translated_df, "attrs", {}).get("translation_rows_affected", 0) or 0)
        _update_job_progress(job_id, min(90, base_progress + 20), f"Saving translated {resolved_sheet}...")
        changed_rows, changed_cells, affected_keys = _persist_translated_data_entry_columns(
            company_name=resolved_company,
            sheet_name=resolved_sheet,
            headers=headers,
            original_df=df,
            translated_df=translated_df,
            columns=tuple(columns),
            uploaded_by_user_id=user_id,
        )
        if changed_cells:
            _clear_data_entry_mapping_metadata_for_grid_keys(resolved_company, resolved_sheet, affected_keys)
            by_group: dict[str | None, set[int]] = defaultdict(set)
            for eg, rix in affected_keys:
                gkey = str(eg or "").strip() or None
                by_group[gkey].add(int(rix) + 1)
            for gkey, row_nums in by_group.items():
                _supersede_open_unmapped_rows(
                    resolved_company,
                    resolved_sheet,
                    source_entry_group=gkey,
                    reason="translation_updated_source_values",
                    row_numbers=row_nums,
                )
        total_rows_affected += rows_affected
        total_cells_updated += changed_cells
        translated_sheets.append(
            {
                "sheet": resolved_sheet,
                "rows_affected": rows_affected,
                "rows_updated": changed_rows,
                "cells_updated": changed_cells,
            }
        )

    _raise_if_job_cancelled(job_id)
    db.session.commit()
    _update_job_progress(job_id, 95, "Finalizing translation...")
    print(f"[TRANSLATION] Rows affected: {total_rows_affected}")
    print("[TRANSLATION] Completed")
    return {
        "ok": True,
        "company": resolved_company,
        "message": "Translation completed",
        "rows_affected": total_rows_affected,
        "cells_updated": total_cells_updated,
        "sheets": translated_sheets,
    }


@app.route("/run-translation", methods=["POST"])
@login_required
def api_run_translation():
    if not _user_can_run_translation(current_user):
        return jsonify({"error": "Translation is only available for administrators and owners"}), 403

    _ensure_db_tables()
    payload = request.get_json(silent=True) or {}
    company = str(payload.get("company") or "").strip()
    if not company:
        return jsonify({"error": "company is required"}), 400
    if not _user_can_access_company(company):
        return jsonify({"error": "Access denied"}), 403

    resolved_company = _resolve_template_company_name(company)
    if not resolved_company:
        return jsonify({"error": "Company not found"}), 404

    translation_mod = _import_translation_module()
    columns_by_sheet = dict(getattr(translation_mod, "TRANSLATION_COLUMNS_BY_SHEET", {}))
    translation_plan: list[dict[str, object]] = []
    for sheet_name, columns in columns_by_sheet.items():
        resolved_sheet = _resolve_template_sheet_name(resolved_company, str(sheet_name))
        if not resolved_sheet:
            continue
        headers, _rules = _get_data_entry_template_schema(resolved_company, resolved_sheet)
        if headers:
            translation_plan.append(
                {
                    "sheet": resolved_sheet,
                    "headers": headers,
                    "columns": list(columns),
                }
            )

    job_id = run_in_background(
        "translation",
        resolved_company,
        _run_translation_job,
        resolved_company=resolved_company,
        translation_plan=translation_plan,
        user_id=int(getattr(current_user, "id", 0) or 0) or None,
        job_user_id=int(getattr(current_user, "id", 0) or 0) or None,
        job_user_email=str(getattr(current_user, "email", "") or ""),
    )
    return jsonify({"job_id": job_id, "status": "started"})


def _run_ccc_import_job(
    *,
    job_id: str,
    project_labels: list[str],
    user_id: int | None = None,
    template_mode: str = TEMPLATE_MODE_LEGACY,
) -> dict[str, object]:
    from frontend.services.ccc_data_entry_import_service import (
        CCC_IMPORT_DEDUP_COLUMN,
        CCC_TARGET_SHEET,
        log_duplicate,
        log_import,
        log_inserted,
        log_warning,
        prepare_ccc_data_entry_rows,
        resolve_ccc_project_id,
    )
    from frontend.services.site_tag_service import resolve_registered_project

    stats: dict[str, object] = {
        "projects_processed": 0,
        "rows_fetched": 0,
        "rows_inserted": 0,
        "duplicates_skipped": 0,
        "fingerprint_duplicates_skipped": 0,
        "validation_skipped": 0,
        "status_skipped": 0,
        "errors": [],
    }

    uid = int(user_id or 0) or None
    ccc_po = _load_stage1_api_source_module("ccc_purchase_orders")
    runtime = ccc_po.resolve_runtime_config()

    log_import(f"starting import job projects={len(project_labels)}")
    _update_job_progress(job_id, 2, "Resolving CCC API projects…")

    if not str(runtime.get("base_url") or "").strip():
        raise RuntimeError("CCC API base URL is not configured.")
    if not str(runtime.get("username") or "").strip() or not str(runtime.get("password") or "").strip():
        raise RuntimeError("CCC API credentials are not configured.")

    projects = list(
        ccc_po.load_available_projects(
            base_url=str(runtime["base_url"]),
            username=str(runtime["username"]),
            password=str(runtime["password"]),
            force_refresh=False,
        )
    )
    token = ccc_po.login(str(runtime["base_url"]), str(runtime["username"]), str(runtime["password"]))
    _raise_if_job_cancelled(job_id)
    _update_job(
        job_id,
        result={
            "phase": "running",
            "projects_processed": 0,
            "rows_fetched": 0,
            "rows_inserted": 0,
            "duplicates_skipped": 0,
            "validation_skipped": 0,
            "status_skipped": 0,
            "fingerprint_duplicates_skipped": 0,
            "errors_count": 0,
        },
    )

    total_labels = max(1, len(project_labels))

    for idx, label in enumerate(project_labels):
        _raise_if_job_cancelled(job_id)
        pct = int((idx / total_labels) * 30)
        _update_job_progress(job_id, 5 + pct, "Fetching CCC API rows…")

        reg = resolve_registered_project(label)
        if not reg:
            log_warning(f"unknown project label={label!r}")
            stats["errors"].append(f"unknown_project:{label}")
            continue

        rc = str(reg.get("responsible_company") or "").strip()
        resolved_company = _resolve_template_company_name(rc) or rc
        if not resolved_company:
            log_warning(f"no template company for label={label!r}")
            stats["errors"].append(f"no_company:{label}")
            continue

        resolved_sheet = (
            _resolve_template_sheet_name_with_mode(resolved_company, CCC_TARGET_SHEET, template_mode=template_mode)
            or CCC_TARGET_SHEET
        )
        headers = _get_template_sheet_headers_with_mode(resolved_company, resolved_sheet, template_mode=template_mode)
        if not headers:
            msg = f"no_sheet_headers:{resolved_company}:{resolved_sheet}"
            log_warning(msg)
            stats["errors"].append(msg)
            continue

        pid = resolve_ccc_project_id(projects, label)
        if pid is None:
            projects = list(
                ccc_po.load_available_projects(
                    base_url=str(runtime["base_url"]),
                    username=str(runtime["username"]),
                    password=str(runtime["password"]),
                    force_refresh=True,
                )
            )
            pid = resolve_ccc_project_id(projects, label)
        if pid is None:
            log_warning(f"CCC project id not found for label={label!r}")
            stats["errors"].append(f"project_id:{label}")
            continue

        try:
            items = ccc_po.fetch_purchase_orders(
                str(runtime["base_url"]),
                token,
                project_id=int(pid),
                sorting=ccc_po.PURCHASE_ORDER_SORTING,
                page_size=ccc_po.PURCHASE_ORDER_PAGE_SIZE,
                query_params=None,
            )
        except Exception as exc:
            stats["errors"].append(str(exc))
            log_warning(f"fetch_failed label={label!r}: {exc}")
            continue

        df = ccc_po.normalize_purchase_orders_for_cache(items)
        stats["projects_processed"] = int(stats["projects_processed"]) + 1
        stats["rows_fetched"] = int(stats["rows_fetched"]) + int(len(df.index))
        log_import(f"labels ok={label!r} fetched={len(df.index)} → {resolved_company}")

        _update_job_progress(job_id, 40 + int((idx / total_labels) * 20), "Resolving site mappings…")
        pairs, skipped_status_n = prepare_ccc_data_entry_rows(df, headers, registration=reg)
        stats["status_skipped"] = int(stats["status_skipped"]) + int(skipped_status_n)
        _raise_if_job_cancelled(job_id)
        _update_job_progress(job_id, 60 + int((idx / total_labels) * 15), "Normalizing reporting periods…")

        existing_keys = _existing_ccc_import_dedup_keys(resolved_company, resolved_sheet, CCC_IMPORT_DEDUP_COLUMN)
        _update_job_progress(job_id, 75 + int((idx / total_labels) * 12), "Checking duplicates…")

        for pk, cells in pairs:
            _raise_if_job_cancelled(job_id)
            if pk.lower() in existing_keys:
                stats["duplicates_skipped"] = int(stats["duplicates_skipped"]) + 1
                log_duplicate(f"ccc key={pk[:16]}…")
                continue

            if not any(str(c or "").strip() for c in cells):
                stats["validation_skipped"] = int(stats["validation_skipped"]) + 1
                continue

            payload_row = {
                "cells": cells,
                "is_persisted": False,
                "row_index": 0,
                "entry_group": "",
                "created_at": "",
            }
            normalized_rows, verr = _normalize_data_entry_rows(headers, [payload_row])
            if verr:
                stats["validation_skipped"] = int(stats["validation_skipped"]) + 1
                log_warning(f"normalize skipped: {verr[0]}")
                continue
            if not normalized_rows:
                stats["validation_skipped"] = int(stats["validation_skipped"]) + 1
                continue

            req_err = _validate_data_entry_row_requirements(headers, normalized_rows)
            if req_err:
                stats["validation_skipped"] = int(stats["validation_skipped"]) + 1
                log_warning(f"requirements skipped: {req_err[0]}")
                continue

            try:
                result = _upsert_data_entries(
                    resolved_company, resolved_sheet, headers, normalized_rows, uploaded_by_user_id=uid
                )
                db.session.flush()
            except Exception as exc:
                db.session.rollback()
                stats["errors"].append(str(exc))
                log_warning(f"save failed: {exc}")
                continue

            saved = int(result.get("saved_rows_count") or 0)
            fp_dup = int(result.get("duplicate_rows_count") or 0)
            saved_groups = list(result.get("saved_entry_groups") or [])

            if saved > 0 and saved_groups:
                eg = saved_groups[-1]
                leader = (
                    DataEntry.query.filter_by(
                        company_name=resolved_company,
                        sheet_name=resolved_sheet,
                        entry_group=eg,
                    )
                    .order_by(DataEntry.id.asc())
                    .first()
                )
                row_index_used = int(leader.row_index) if leader else 0
                _upsert_data_entry_cell(
                    company_name=resolved_company,
                    sheet_name=resolved_sheet,
                    entry_group=eg,
                    row_index=row_index_used,
                    column_name=CCC_IMPORT_DEDUP_COLUMN,
                    value=pk,
                    uploaded_by_user_id=uid,
                )
                db.session.commit()
                existing_keys.add(pk.lower())
                stats["rows_inserted"] = int(stats["rows_inserted"]) + 1
                _update_job(job_id, rows=int(stats["rows_inserted"]))
                log_inserted(f"{resolved_company} row_index={row_index_used} group={eg}")
            else:
                db.session.rollback()
                if fp_dup:
                    stats["fingerprint_duplicates_skipped"] = int(stats["fingerprint_duplicates_skipped"]) + 1

        _update_job_progress(job_id, 88 + int(((idx + 1) / total_labels) * 10), "Saving Data Entry rows…")
        errs = stats["errors"] if isinstance(stats["errors"], list) else []
        _update_job(
            job_id,
            result={
                "phase": "running",
                "projects_processed": int(stats["projects_processed"]),
                "rows_fetched": int(stats["rows_fetched"]),
                "rows_inserted": int(stats["rows_inserted"]),
                "duplicates_skipped": int(stats["duplicates_skipped"]),
                "validation_skipped": int(stats["validation_skipped"]),
                "status_skipped": int(stats["status_skipped"]),
                "fingerprint_duplicates_skipped": int(stats["fingerprint_duplicates_skipped"]),
                "errors_count": len(errs),
            },
        )

    _update_job_progress(
        job_id,
        100,
        (
            "Completed CCC import • inserted "
            + str(stats["rows_inserted"])
            + ", skipped duplicates "
            + str(stats["duplicates_skipped"])
        ),
    )
    log_import(
        "summary inserted=%s dedup_skip=%s fp_dup=%s val_skip=%s status_skip=%s projects=%s"
        % (
            stats["rows_inserted"],
            stats["duplicates_skipped"],
            stats["fingerprint_duplicates_skipped"],
            stats["validation_skipped"],
            stats["status_skipped"],
            stats["projects_processed"],
        )
    )
    _update_job(job_id, rows=int(stats["rows_inserted"]))
    errs_final = stats["errors"] if isinstance(stats["errors"], list) else []
    out = dict(stats)
    out["ok"] = True
    out["errors_count"] = len(errs_final)
    out["phase"] = "completed"
    print(
        "[MAPPING_STATE] ccc_import_completed "
        f"inserted={out.get('rows_inserted')} note=no_EF_until_map_job"
    )
    if uid:
        _create_user_notification(
            int(uid),
            title="CCC import completed",
            message=(
                f"Inserted {int(stats.get('rows_inserted') or 0)} row(s) into Data Entry. "
                "This step does not assign emission factors — use Map to set EF / status."
            ),
            notification_type="info",
            link="/",
        )
    return out


def _run_mapping_job(
    *,
    job_id: str,
    user_id: int,
    user_email: str,
    resolved_company: str,
    resolved_sheet: str,
    headers: list[str],
    template_mode: str,
    entry_group_filter: str = "",
) -> dict[str, object]:
    _update_job_progress(job_id, 5, "Loading saved data...")
    if not headers:
        raise RuntimeError("Sheet not found")

    if entry_group_filter:
        df = _load_data_entries_dataframe_no_request(
            resolved_company, resolved_sheet, headers, {entry_group_filter}
        )
        if df.empty:
            raise RuntimeError("No saved rows found for this entry batch")
    else:
        df = _load_data_entries_dataframe_no_request(resolved_company, resolved_sheet, headers)
        if df.empty:
            raise RuntimeError("No saved rows found for this company and sheet")

    _raise_if_job_cancelled(job_id)
    _update_job_progress(job_id, 15, "Creating mapping run...")
    run_id = uuid.uuid4().hex[:12]
    mr = MappingRun(
        id=run_id,
        user_id=int(user_id),
        company_name=resolved_company,
        sheet_name=resolved_sheet,
        status="running",
        created_at=datetime.utcnow(),
        source_entry_group=entry_group_filter or None,
    )
    db.session.add(mr)
    db.session.commit()

    try:
        _raise_if_job_cancelled(job_id)
        _update_job_progress(job_id, 30, "Running mapping engine...")
        mapped_df, out_path, input_path = run_mapping(
            resolved_company,
            resolved_sheet,
            df,
            template_mode=template_mode,
        )
    except JobCancelled:
        mr.status = "cancelled"
        db.session.commit()
        raise
    except Exception as e:
        try:
            mr.status = "failed"
            mr.error_message = str(e)
            db.session.commit()
        except Exception:
            db.session.rollback()
        raise

    if _is_job_cancel_requested(job_id):
        mr.status = "cancelled"
        db.session.commit()
        raise JobCancelled()
    _update_job_progress(job_id, 75, "Saving mapping output...")
    mr.status = "succeeded"
    mr.output_path = str(out_path)
    mr.input_path = str(input_path)
    _persist_mapping_metadata_to_data_entry(
        company_name=resolved_company,
        sheet_name=resolved_sheet,
        source_entry_group=entry_group_filter or None,
        mapped_df=mapped_df,
        run_id=run_id,
        user_id=int(user_id),
    )
    db.session.commit()

    unmapped_count = 0
    _update_job_progress(job_id, 85, "Updating mapping summaries...")
    try:
        _upsert_mapping_run_summary(
            run_id=run_id,
            company_name=resolved_company,
            sheet_name=resolved_sheet,
            mapped_df=mapped_df,
            output_path=out_path,
        )
        unmapped_count = _sync_unmapped_rows_for_mapping_run(
            run_id=run_id,
            user_id=int(user_id),
            company_name=resolved_company,
            sheet_name=resolved_sheet,
            source_entry_group=entry_group_filter or None,
            mapped_df=mapped_df,
        )
        db.session.commit()
    except Exception:
        unmapped_count = 0
        db.session.rollback()

    if _is_job_cancel_requested(job_id):
        mr.status = "cancelled"
        db.session.commit()
        raise JobCancelled()
    _update_job_progress(job_id, 95, "Finalizing mapping...")
    um_n = int(unmapped_count or 0)
    notif_msg = f"{resolved_company} / {resolved_sheet} mapping finished."
    if um_n > 0:
        notif_msg += f" {um_n} row(s) still have no EF match — review Admin ▸ Unmapped rows."
    map_card = _mapping_card_payload_for_pair(resolved_company, resolved_sheet)
    if map_card is None:
        map_card = {
            "company_name": resolved_company,
            "uploaded_by_user": "",
            "uploaded_by_user_id": 0,
            "uploaded_by_job_title": "",
            "uploaded_by_has_profile_photo": False,
            "upload_timestamp": "",
            "category": resolved_sheet,
            "row_count": 0,
            "mapping_status": "",
            "mapping_state": "",
            "mapped_by_admin": "",
            "mapping_timestamp": "",
            "mapped": True,
        }
    else:
        map_card = dict(map_card)
    map_card["mapped_by_admin"] = str(user_email or "").strip()
    map_card["mapping_timestamp"] = datetime.utcnow().strftime("%Y-%m-%d %H:%M")
    _create_user_notification(
        int(user_id),
        title="Mapping run completed",
        message=notif_msg,
        notification_type="success",
        link="/",
        mapping_card=map_card,
        feed_event="mapping_completed",
        feed_company=resolved_company,
        feed_timestamp=datetime.utcnow(),
    )

    preview = mapped_df.head(40).fillna("").to_dict(orient="records")
    resp: dict[str, object] = {
        "ok": True,
        "run_id": run_id,
        "company": resolved_company,
        "sheet": resolved_sheet,
        "mapped_columns": list(mapped_df.columns),
        "preview": preview,
        "preview_rows": len(preview),
        "mapped_at": mr.created_at.isoformat() + "Z" if getattr(mr, "created_at", None) else "",
        "mapped_by": user_email,
        "unmapped_count": int(unmapped_count or 0),
    }
    if entry_group_filter:
        resp["entry_group"] = entry_group_filter
    return resp


@app.route("/run-mapping", methods=["POST"])
@app.route("/api/mapping/run", methods=["POST"])
@login_required
def api_mapping_run():
    if not bool(getattr(current_user, "is_admin", False)):
        return jsonify({"error": "Mapping is only available for administrators"}), 403
    _ensure_db_tables()
    payload = request.get_json(silent=True) or {}
    company = (payload.get("company") or "").strip()
    sheet = (payload.get("sheet") or "").strip()
    rows = payload.get("rows") or []
    entry_group_filter = (payload.get("entry_group") or "").strip()

    if not company or not sheet:
        return jsonify({"error": "company and sheet are required"}), 400
    if not isinstance(rows, list):
        return jsonify({"error": "rows must be a list"}), 400
    if _is_hidden_schema_sheet(sheet):
        return jsonify({"error": "This sheet is not available for web data entry"}), 403
    if not _user_can_access_company(company):
        return jsonify({"error": "Access denied"}), 403

    resolved_company = _resolve_template_company_name(company)
    if not resolved_company:
        return jsonify({"error": "Company not found"}), 404
    resolved_sheet = _resolve_template_sheet_name(resolved_company, sheet)
    if not resolved_sheet:
        return jsonify({"error": "Sheet not found"}), 404
    if resolved_sheet == KLARAKARBON_SHEET_NAME:
        return jsonify({"error": "Klarakarbon entries are not mapped from the Data Entry page."}), 403
    headers, _rules = _get_data_entry_template_schema(resolved_company, resolved_sheet)
    if not headers:
        return jsonify({"error": "Sheet not found"}), 404
    template_mode = _current_template_mode()

    if entry_group_filter:
        existing_batch_map = MappingRun.query.filter_by(
            company_name=resolved_company,
            sheet_name=resolved_sheet,
            status="succeeded",
            source_entry_group=entry_group_filter,
        ).first()
        if existing_batch_map is not None:
            ma = existing_batch_map.created_at.isoformat() if existing_batch_map.created_at else ""
            return jsonify(
                {
                    "error": "This batch was already mapped",
                    "already_mapped": True,
                    "mapped_at": ma,
                }
            ), 409

    if rows:
        normalized_rows, validation_errors = _normalize_data_entry_rows(headers, rows)
        if validation_errors:
            return jsonify({"error": validation_errors[0], "validation_errors": validation_errors[:20]}), 400
        requirement_errors = _validate_data_entry_row_requirements(headers, normalized_rows)
        if requirement_errors:
            return jsonify({"error": requirement_errors[0], "validation_errors": requirement_errors[:20]}), 400
        try:
            _upsert_data_entries(resolved_company, resolved_sheet, headers, normalized_rows)
            db.session.commit()
        except Exception as e:
            db.session.rollback()
            return jsonify({"error": f"Save failed: {e}"}), 500

    if entry_group_filter:
        df_check = _load_data_entries_dataframe_for_entry_groups(
            resolved_company, resolved_sheet, headers, {entry_group_filter}
        )
        if df_check.empty:
            return jsonify({"error": "No saved rows found for this entry batch"}), 400
    else:
        df_check = _load_data_entries_dataframe(resolved_company, resolved_sheet, headers)
        if df_check.empty:
            return jsonify({"error": "No saved rows found for this company and sheet"}), 400

    job_id = run_in_background(
        "mapping",
        resolved_company,
        _run_mapping_job,
        user_id=int(current_user.id),
        user_email=str(getattr(current_user, "email", "") or ""),
        resolved_company=resolved_company,
        resolved_sheet=resolved_sheet,
        headers=headers,
        template_mode=template_mode,
        entry_group_filter=entry_group_filter,
        job_user_id=int(current_user.id),
        job_user_email=str(getattr(current_user, "email", "") or ""),
    )
    return jsonify({"job_id": job_id, "status": "started"})


def _run_append_pipeline_job(
    *,
    job_id: str,
    user_id: int,
    resolved_company: str,
    resolved_sheet: str,
) -> dict[str, object]:
    _update_job_progress(job_id, 10, "Starting append & run pipeline...")
    result = _run_append_and_pipeline(resolved_company, resolved_sheet)
    _update_job_progress(job_id, 90, "Creating completion notification...")
    _create_user_notification(
        int(user_id),
        title="Pipeline run completed",
        message=f"Data pipeline executed successfully for {resolved_company}.",
        notification_type="success",
        link="/",
        feed_event="pipeline_completed",
        feed_company=resolved_company,
        feed_timestamp=datetime.utcnow(),
    )
    _update_job_progress(job_id, 100, "Pipeline completed")
    return result


@app.route("/api/pipeline/append_run", methods=["POST"])
@login_required
def api_pipeline_append_run():
    if not bool(getattr(current_user, "is_admin", False)):
        return jsonify({"error": "Append & Run pipeline is only available for administrators"}), 403

    payload = request.get_json(silent=True) or {}
    company = (payload.get("company") or "").strip()
    sheet = (payload.get("sheet") or "").strip()
    if not company or not sheet:
        return jsonify({"error": "company and sheet are required"}), 400
    if not _user_can_access_company(company):
        return jsonify({"error": "Access denied"}), 403

    resolved_company = _resolve_template_company_name(company)
    if not resolved_company:
        return jsonify({"error": "Company not found"}), 404

    sheet_key = str(sheet).strip()
    if _batch_action_type_for_sheet(sheet_key) != "append_run":
        return jsonify({"error": "This category must use the normal Map action."}), 400
    if sheet_key == KLARAKARBON_SHEET_NAME:
        resolved_sheet = KLARAKARBON_SHEET_NAME
    elif sheet_key == TRAVEL_SHEET_NAME:
        resolved_sheet = TRAVEL_SHEET_NAME
    else:
        return jsonify({"error": "Unsupported pipeline category."}), 400

    job_id = run_in_background(
        "pipeline",
        resolved_company,
        _run_append_pipeline_job,
        user_id=int(current_user.id),
        resolved_company=resolved_company,
        resolved_sheet=resolved_sheet,
        job_user_id=int(current_user.id),
        job_user_email=str(getattr(current_user, "email", "") or ""),
    )
    return jsonify({"job_id": job_id, "status": "started"})


def _find_latest_merged_mapping_workbook() -> Path | None:
    patterns = [
        "mapped_results_merged_*.xlsx",
        "mapped_results_merged.xlsx",
    ]
    candidates: list[Path] = []
    for pattern in patterns:
        candidates.extend(STAGE2_OUTPUT_DIR.glob(pattern))
    candidates = [p for p in candidates if p.is_file() and not p.name.startswith("~$")]
    if not candidates:
        return None
    candidates.sort(key=lambda p: p.stat().st_mtime, reverse=True)
    return candidates[0]


def _find_latest_stage2_output(*patterns: str) -> Path | None:
    candidates: list[Path] = []
    for pattern in patterns:
        candidates.extend(STAGE2_OUTPUT_DIR.rglob(pattern))
    files = [p for p in candidates if p.is_file() and not p.name.startswith("~$")]
    if not files:
        return None
    deduped = list({str(p.resolve()): p for p in files}.values())
    deduped.sort(key=lambda p: p.stat().st_mtime, reverse=True)
    return deduped[0]


def _load_stage2_module(module_basename: str):
    module_path = STAGE2_MAPPING_DIR / f"{module_basename}.py"
    module_name = f"ui_stage2_{module_basename}_{uuid.uuid4().hex}"
    spec = importlib.util.spec_from_file_location(module_name, module_path)
    if spec is None or spec.loader is None:
        raise RuntimeError(f"Could not load module: {module_path.name}")
    module = importlib.util.module_from_spec(spec)
    sys.modules[module_name] = module
    spec.loader.exec_module(module)
    return module


def _load_stage1_api_source_module(module_basename: str):
    module_path = PROJECT_ROOT / "engine" / "stage1_preprocess" / "api_sources" / f"{module_basename}.py"
    module_name = f"ui_stage1_api_{module_basename}_{uuid.uuid4().hex}"
    spec = importlib.util.spec_from_file_location(module_name, module_path)
    if spec is None or spec.loader is None:
        raise RuntimeError(f"Could not load module: {module_path.name}")
    module = importlib.util.module_from_spec(spec)
    sys.modules[module_name] = module
    spec.loader.exec_module(module)
    return module


def _format_output_preview_value(value: object) -> str:
    try:
        if pd.isna(value):
            return ""
    except Exception:
        pass
    if isinstance(value, (pd.Timestamp, datetime)):
        return value.strftime("%Y-%m-%d")
    if isinstance(value, float):
        if math.isfinite(value):
            if value.is_integer():
                return f"{value:,.0f}"
            return f"{value:,.4f}".rstrip("0").rstrip(".")
        return ""
    return str(value)


def _dataframe_preview_payload(sheet_name: str, df: pd.DataFrame, limit: int = 20) -> dict[str, object]:
    preview = df.head(limit).copy()
    columns = [str(c) for c in preview.columns]
    rows: list[list[str]] = []
    for ridx in range(len(preview.index)):
        rows.append([
            _format_output_preview_value(preview.iloc[ridx, cidx])
            for cidx in range(len(columns))
        ])
    return {
        "name": sheet_name,
        "columns": columns,
        "rows": rows,
        "row_count": int(len(df.index)),
        "column_count": int(len(df.columns)),
        "truncated": bool(len(df.index) > limit),
    }


ANALYTICS_OUTPUT_VIEWS: dict[str, dict[str, object]] = {
    "forecasting": {
        "title": "Forecasting",
        "group": "Analytics",
        "summary": "Run the existing forecasting script on demand and preview the newly generated workbook.",
        "preferred_sheets": ("Forecast_Total", "Forecast_Company", "Forecast_Sheet", "Backtest", "Meta"),
        "empty_message": "No forecasting workbook has been generated from the UI in this environment yet.",
    },
    "decarbonization": {
        "title": "Decarbonization",
        "group": "Analytics",
        "summary": "Run the existing decarbonization scripts with UI-supplied parameters and preview the fresh scenario output.",
        "preferred_sheets": ("Yearly", "Monthly", "Delta_vs_BAU", "Meta"),
        "empty_message": "No decarbonization workbook has been generated from the UI in this environment yet.",
    },
    "ccc_api_source": {
        "title": "CCC API",
        "group": "Data Sources",
        "summary": "Test the CCC connection and ingest any configured GET endpoint into stage1-compatible workbook outputs without changing pipeline calculations.",
        "preferred_sheets": ("CCC Purchase Orders Raw", "Scope 3 Cat 1 Common Purchases", "Scope 3 Cat 1 Services Spend", "CCC Waste Suppliers Review"),
        "empty_message": "No CCC API workbook has been generated from the UI in this environment yet.",
    },
    "mapped_window_output": {
        "title": "Mapped Window Output",
        "group": "Data Outputs",
        "summary": "Generate a fresh auditor-style mapped window workbook for a selected reporting period using the existing stage2 filter step.",
        "preferred_sheets": ("Company Totals Window", "Company by GHGP Totals Window", "GHGP sheet Totals Window", "Company Stacked Months Window"),
        "empty_message": "No mapped window workbook has been generated from the UI in this environment yet.",
    },
    "emissions_totals": {
        "title": "Totals Tables",
        "group": "Data Outputs",
        "summary": "Generate fresh totals tables on demand using the existing window workbook and totals-table script.",
        "preferred_sheets": ("Company Totals Window", "Company by GHGP Totals Window", "GHGP sheet Totals Window"),
        "empty_message": "No totals workbook has been generated from the UI in this environment yet.",
    },
    "share_analysis": {
        "title": "Share Analysis",
        "group": "Data Outputs",
        "summary": "Generate fresh share-analysis tables using the existing totals-table stage and preview the resulting workbook.",
        "preferred_sheets": ("Company Totals Window", "Company by GHGP Totals Window", "Company Stacked Data Window", "Company Stacked Months Window"),
        "empty_message": "No share-analysis workbook has been generated from the UI in this environment yet.",
    },
    "double_counting_check": {
        "title": "Double Counting Check",
        "group": "Governance",
        "summary": "Generate a fresh double-counting report using the existing stage2 review script.",
        "preferred_sheets": ("DC Log", "Anomalies"),
        "empty_message": "No double-counting workbook has been generated from the UI in this environment yet.",
    },
    "audit_ready_output": {
        "title": "Audit Ready Dataset",
        "group": "Governance",
        "summary": "Generate a fresh mapped window dataset for external audit review using the existing filtering step only.",
        "preferred_sheets": ("Company Totals Window", "Company by GHGP Totals Window", "GHGP sheet Totals Window", "Company Stacked Months Window"),
        "empty_message": "No audit-ready dataset has been generated from the UI in this environment yet.",
    },
}

ANALYTICS_RUN_LOG_CONFIG: dict[str, dict[str, object]] = {
    "forecasting": {
        "path": APP_DIR / "run_logs" / "forecasting_runs.json",
        "types": {"forecasting"},
    },
    "decarbonization": {
        "path": APP_DIR / "run_logs" / "decarbonization_runs.json",
        "types": {"decarbonization"},
    },
    "ccc_api_source": {
        "path": APP_DIR / "run_logs" / "ccc_api_sync.json",
        "types": {"ccc_api_sync", "ccc_api_test"},
    },
    "mapped_window_output": {
        "path": APP_DIR / "run_logs" / "data_output_runs.json",
        "types": {"mapped_window"},
    },
    "emissions_totals": {
        "path": APP_DIR / "run_logs" / "data_output_runs.json",
        "types": {"totals_tables"},
    },
    "share_analysis": {
        "path": APP_DIR / "run_logs" / "data_output_runs.json",
        "types": {"share_analysis"},
    },
    "double_counting_check": {
        "path": APP_DIR / "run_logs" / "data_output_runs.json",
        "types": {"double_counting"},
    },
    "audit_ready_output": {
        "path": APP_DIR / "run_logs" / "audit_runs.json",
        "types": {"audit_output"},
    },
}


def _analytics_history_download_url(filename: str) -> str | None:
    safe_name = Path(str(filename or "")).name
    if not safe_name:
        return None
    return url_for("analytics_output_download_file", filename=safe_name)


def _analytics_run_log_config(view_key: str) -> dict[str, object] | None:
    cfg = ANALYTICS_RUN_LOG_CONFIG.get(view_key)
    if cfg is not None:
        return cfg
    return None


def _read_run_history(view_key: str, limit: int = 20) -> list[dict[str, object]]:
    cfg = _analytics_run_log_config(view_key)
    path = Path(cfg.get("path")) if cfg and cfg.get("path") else None
    allowed_types = {str(v) for v in (cfg.get("types") or set())} if cfg else set()
    if path is None or not path.exists():
        return []
    try:
        data = json.loads(path.read_text(encoding="utf-8"))
    except Exception:
        return []
    if not isinstance(data, list):
        return []
    rows: list[dict[str, object]] = []
    for item in data:
        if not isinstance(item, dict):
            continue
        row = dict(item)
        row_type = str(row.get("type") or "")
        if allowed_types and row_type not in allowed_types:
            continue
        output_file = str(row.get("output_file") or "")
        row["download_url"] = _analytics_history_download_url(output_file) if output_file else None
        rows.append(row)
    rows.sort(key=lambda item: str(item.get("timestamp") or ""), reverse=True)
    return rows[:limit]


def _append_run_history(view_key: str, payload: dict[str, object]) -> None:
    cfg = _analytics_run_log_config(view_key)
    path = Path(cfg.get("path")) if cfg and cfg.get("path") else None
    if path is None:
        return
    path.parent.mkdir(parents=True, exist_ok=True)
    try:
        existing = json.loads(path.read_text(encoding="utf-8")) if path.exists() else []
    except Exception:
        existing = []
    rows = existing if isinstance(existing, list) else []
    clean_rows = []
    for row in rows:
        if not isinstance(row, dict):
            continue
        r = dict(row)
        r.pop("download_url", None)
        clean_rows.append(r)
    clean_rows.insert(0, payload)
    path.write_text(json.dumps(clean_rows[:100], indent=2), encoding="utf-8")


def _latest_run_history_entry(view_key: str, *, statuses: set[str] | None = None, types: set[str] | None = None) -> dict[str, object] | None:
    rows = _read_run_history(view_key, limit=100)
    for row in rows:
        row_status = str(row.get("status") or "")
        row_type = str(row.get("type") or "")
        if statuses and row_status not in statuses:
            continue
        if types and row_type not in types:
            continue
        return row
    return None


def _load_output_sheet_previews(path: Path, preferred_sheets: tuple[str, ...], *, max_rows: int = 20, max_sheets: int = 4) -> tuple[list[str], list[dict[str, object]]]:
    xls = pd.ExcelFile(path, engine="openpyxl")
    available_sheets = [str(name) for name in xls.sheet_names]
    selected: list[str] = []
    for name in preferred_sheets:
        if name in available_sheets and name not in selected:
            selected.append(name)
    if not selected:
        selected = available_sheets[:max_sheets]
    previews: list[dict[str, object]] = []
    for sheet_name in selected[:max_sheets]:
        df = pd.read_excel(xls, sheet_name=sheet_name, engine="openpyxl")
        previews.append(_dataframe_preview_payload(sheet_name, df, limit=max_rows))
    return available_sheets, previews


def _find_output_file_by_name(filename: str) -> Path | None:
    safe_name = Path(str(filename or "")).name
    if not safe_name:
        return None
    candidates: list[Path] = []
    for root in (STAGE2_OUTPUT_DIR, STAGE1_INPUT_DIR, DATA_DIR):
        candidates.extend([p for p in root.rglob(safe_name) if p.is_file()])
    if not candidates:
        return None
    candidates.sort(key=lambda p: p.stat().st_mtime, reverse=True)
    return candidates[0]


def _latest_logged_output_path(view_key: str) -> Path | None:
    history = _read_run_history(view_key, limit=100)
    for row in history:
        output_file = str(row.get("output_file") or "")
        if not output_file:
            continue
        path = _find_output_file_by_name(output_file)
        if path is not None:
            return path
    return None


def _analytics_output_context_for_path(view_key: str, path: Path | None) -> dict[str, object]:
    cfg = ANALYTICS_OUTPUT_VIEWS[view_key]
    preferred_sheets = tuple(str(s) for s in (cfg.get("preferred_sheets") or ()))
    context: dict[str, object] = {
        "page_key": view_key,
        "page_title": str(cfg.get("title") or "Output"),
        "page_group": str(cfg.get("group") or "Analytics"),
        "page_summary": str(cfg.get("summary") or ""),
        "empty_message": str(cfg.get("empty_message") or "No output found."),
        "output_exists": False,
        "output_file_name": "",
        "output_file_mtime": "",
        "output_sheet_count": 0,
        "output_sheet_names": [],
        "sheet_previews": [],
        "download_url": None,
        "run_notice": "",
        "run_error": "",
        "form_state": {},
        "companion_outputs": [],
        "run_history": _read_run_history(view_key),
    }
    if path is None:
        return context
    try:
        available_sheets, previews = _load_output_sheet_previews(path, preferred_sheets)
    except Exception as exc:
        context["empty_message"] = f"Output file was found but could not be read: {exc}"
        return context
    stat = path.stat()
    context.update(
        {
            "output_exists": True,
            "output_file_name": path.name,
            "output_file_mtime": datetime.fromtimestamp(stat.st_mtime).strftime("%Y-%m-%d %H:%M"),
            "output_sheet_count": len(available_sheets),
            "output_sheet_names": available_sheets,
            "sheet_previews": previews,
            "download_url": url_for("analytics_output_download", kind=view_key),
        }
    )
    return context


def _analytics_output_context(view_key: str) -> dict[str, object]:
    return _analytics_output_context_for_path(view_key, _latest_logged_output_path(view_key))


def _build_companion_output_payload(title: str, path: Path, preferred_sheets: tuple[str, ...]) -> dict[str, object]:
    available_sheets, previews = _load_output_sheet_previews(path, preferred_sheets)
    return {
        "title": title,
        "file_name": path.name,
        "updated_at": datetime.fromtimestamp(path.stat().st_mtime).strftime("%Y-%m-%d %H:%M"),
        "sheet_count": len(available_sheets),
        "sheet_names": available_sheets,
        "sheet_previews": previews,
    }


def _parse_int_form(name: str, default: int, *, minimum: int | None = None, maximum: int | None = None) -> int:
    raw = str(request.form.get(name, "") or "").strip()
    try:
        value = int(raw)
    except Exception:
        value = default
    if minimum is not None:
        value = max(minimum, value)
    if maximum is not None:
        value = min(maximum, value)
    return value


def _parse_int_arg(name: str, default: int | None, *, minimum: int | None = None, maximum: int | None = None) -> int | None:
    raw = str(request.args.get(name, "") or "").strip()
    if not raw:
        return default
    try:
        value = int(raw)
    except (TypeError, ValueError):
        return default
    if minimum is not None:
        value = max(minimum, value)
    if maximum is not None:
        value = min(maximum, value)
    return value


def _parse_float_form(name: str, default: float, *, minimum: float | None = None, maximum: float | None = None) -> float:
    raw = str(request.form.get(name, "") or "").strip()
    try:
        value = float(raw)
    except Exception:
        value = default
    if minimum is not None:
        value = max(minimum, value)
    if maximum is not None:
        value = min(maximum, value)
    return value


def _scenario_growth_profile(scenario_type: str) -> tuple[float, float]:
    st = str(scenario_type or "").strip().lower()
    if st == "expansion":
        return 0.10, 2.5
    if st == "growth":
        return 0.05, 2.0
    if st == "decarbonization":
        return 0.0, 1.5
    return 0.0, 2.0


def _decarbonization_scope_mapping(scope_key: str) -> tuple[list[str], bool]:
    scope = str(scope_key or "").strip().lower()
    if scope == "scope1":
        return (["Scope 1"], False)
    if scope == "scope2":
        return (["Scope 2"], False)
    if scope == "scope3":
        return (["S3 Cat 1 Purchased G&S"], False)
    return ([], True)


def _write_ui_lever_csv(*, target_year: int, reduction_pct: float, scope_key: str, scenario_type: str) -> Path:
    applies_to_cols, scale_all = _decarbonization_scope_mapping(scope_key)
    ramp_years = max(1, target_year - 2025)
    out_path = STAGE2_OUTPUT_DIR / f"ui_decarb_levers_{uuid.uuid4().hex}.csv"
    row = {
        "lever_key": "ui_custom",
        "name": f"UI custom {scenario_type}",
        "reduction_pct": max(0.0, min(1.0, reduction_pct)),
        "start_year_offset": 0,
        "ramp_years": ramp_years,
        "scale_all": bool(scale_all),
        "applies_to_cols": ", ".join(applies_to_cols),
        "notes": f"Generated from UI for scope={scope_key}, target_year={target_year}",
    }
    with out_path.open("w", newline="", encoding="utf-8-sig") as f:
        writer = csv.DictWriter(
            f,
            fieldnames=[
                "lever_key",
                "name",
                "reduction_pct",
                "start_year_offset",
                "ramp_years",
                "scale_all",
                "applies_to_cols",
                "notes",
            ],
        )
        writer.writeheader()
        writer.writerow(row)
    return out_path


def _run_forecasting_from_ui(*, target_year: int) -> Path:
    forecasting = _load_stage2_module("Forecasting")
    cfg = forecasting.ForecastConfig(target_year=int(target_year))
    out_path = forecasting.run_forecasting(None, cfg)
    if out_path is None:
        raise RuntimeError("Forecasting did not produce an output workbook.")
    return Path(out_path)


def _run_decarbonization_from_ui(*, target_year: int, reduction_pct: float, scope_key: str, scenario_type: str) -> tuple[Path, Path]:
    years = max(1, int(target_year) - 2025)
    annual_growth, pod_multiplier_end = _scenario_growth_profile(scenario_type)
    levers_csv = _write_ui_lever_csv(
        target_year=int(target_year),
        reduction_pct=float(reduction_pct),
        scope_key=scope_key,
        scenario_type=scenario_type,
    )

    decarb = _load_stage2_module("Decarbonization")
    scenario_output = decarb.run_scenarios(
        years=years,
        annual_growth=annual_growth,
        pod_multiplier_end=pod_multiplier_end,
        baseline_source="additive_tabs",
        levers_csv=str(levers_csv),
        s3_cat1_multiplier=1.0,
        dirty_region_multiplier=1.0,
    )

    decarb_scenarios = _load_stage2_module("Decarbonization_Scenarios")
    rollout_end = f"{int(target_year)}-01-01"
    companion_output = decarb_scenarios.run(
        None,
        input_window=None,
        user_scenarios=True,
        annual_growth_default=annual_growth,
        annual_growth_10pct=max(annual_growth, 0.10),
        annual_growth_expansion=max(annual_growth, 0.15),
        rollout_end=rollout_end,
    )
    return Path(scenario_output), Path(companion_output)


def _window_period_end_date(year: int, month: int) -> str:
    month = max(1, min(12, int(month)))
    _, day_count = calendar.monthrange(int(year), month)
    return f"{int(year)}-{int(month):02d}-{day_count:02d}"


def _run_mapped_window_from_ui(*, year: int, start_month: int = 1, end_month: int = 12) -> Path:
    year = int(year)
    start_month = max(1, min(12, int(start_month)))
    end_month = max(start_month, min(12, int(end_month)))
    window = _load_stage2_module("filter_run_output_by_period")
    out_path = window.filter_workbook(
        f"{year}-{start_month:02d}-01",
        _window_period_end_date(year, end_month),
        None,
        None,
    )
    if out_path is None:
        raise RuntimeError("Mapped window generation did not produce an output workbook.")
    return Path(out_path)


def _run_totals_tables_from_ui(*, year: int, start_month: int = 1, end_month: int = 12) -> Path:
    window_path = _run_mapped_window_from_ui(year=year, start_month=start_month, end_month=end_month)
    totals = _load_stage2_module("Window_total_tables")
    out_path = totals.main(str(window_path))
    return Path(out_path) if out_path else window_path


def _capture_existing_output_names(*patterns: str) -> set[str]:
    names: set[str] = set()
    for pattern in patterns:
        for path in STAGE2_OUTPUT_DIR.rglob(pattern):
            if path.is_file():
                names.add(path.name)
    return names


def _find_new_output_after_run(before_names: set[str], *patterns: str) -> Path | None:
    candidates: list[Path] = []
    for pattern in patterns:
        for path in STAGE2_OUTPUT_DIR.rglob(pattern):
            if path.is_file():
                candidates.append(path)
    if not candidates:
        return None
    candidates.sort(key=lambda p: p.stat().st_mtime, reverse=True)
    for path in candidates:
        if path.name not in before_names:
            return path
    return candidates[0]


def _run_double_counting_from_ui() -> Path:
    before_names = _capture_existing_output_names("mapped_results_merged_dc_*.xlsx", "mapped_results_merged_dc.xlsx")
    double_counting = _load_stage2_module("double_countin_booklets")
    double_counting.main()
    out_path = _find_new_output_after_run(before_names, "mapped_results_merged_dc_*.xlsx", "mapped_results_merged_dc.xlsx")
    if out_path is None:
        raise RuntimeError("Double counting report did not produce an output workbook.")
    return out_path


def _ccc_runtime_defaults() -> dict[str, object]:
    return {
        "base_url": str(CCC_API_BASE_URL or "").strip(),
        "username": str(CCC_USERNAME or "").strip(),
        "page_size": int(CCC_API_PAGE_SIZE or 100),
        "has_password": bool(str(CCC_PASSWORD or "").strip()),
    }


def _load_ccc_mapping_rules() -> dict[str, object]:
    try:
        data = json.loads(CCC_SHEET_MAPPING_PATH.read_text(encoding="utf-8"))
    except Exception:
        data = {}
    return data if isinstance(data, dict) else {}


def _load_ccc_get_endpoints() -> dict[str, str]:
    try:
        data = json.loads(CCC_GET_ENDPOINTS_PATH.read_text(encoding="utf-8"))
    except Exception:
        data = {}
    if not isinstance(data, dict):
        return {}
    return {
        str(key): str(value)
        for key, value in data.items()
        if str(key).strip() and str(value).strip()
    }


def _parse_json_object_form_field(name: str) -> dict[str, object]:
    raw = str(request.form.get(name, "") or "").strip()
    if not raw:
        return {}
    try:
        data = json.loads(raw)
    except Exception as exc:
        raise RuntimeError(f"{name.replace('_', ' ').title()} must be valid JSON.") from exc
    if not isinstance(data, dict):
        raise RuntimeError(f"{name.replace('_', ' ').title()} must be a JSON object.")
    return data


def _ccc_connection_status_payload(*, state: str = "unknown", message: str = "", tested_at: str = "") -> dict[str, str]:
    labels = {
        "connected": "Connected",
        "configured": "Configured",
        "missing": "Missing credentials",
        "failed": "Connection failed",
        "unknown": "Not tested",
    }
    tones = {
        "connected": "success",
        "configured": "neutral",
        "missing": "error",
        "failed": "error",
        "unknown": "neutral",
    }
    return {
        "state": state,
        "label": labels.get(state, "Unknown"),
        "tone": tones.get(state, "neutral"),
        "message": message,
        "tested_at": tested_at,
    }


def _coerce_ccc_project_id(value: object, default: int | None = None) -> int | None:
    try:
        project_id = int(value)  # type: ignore[arg-type]
    except Exception:
        return default
    return project_id if project_id > 0 else default


def _load_ccc_available_projects(
    *,
    base_url: str | None = None,
    username: str | None = None,
    password: str | None = None,
    force_refresh: bool = False,
) -> list[dict[str, object]]:
    ccc_api = _load_stage1_api_source_module("ccc_purchase_orders")
    return list(
        ccc_api.load_available_projects(
            base_url=str(base_url or "").strip(),
            username=str(username or "").strip(),
            password=str(password or ""),
            force_refresh=bool(force_refresh),
        )
    )


def _load_ccc_purchase_orders_summary(project_id: int | None) -> dict[str, object]:
    default_preview = {
        "name": "Latest purchase orders",
        "columns": ["Supplier", "Total Price", "Currency", "CreatedOn", "Status"],
        "rows": [],
        "row_count": 0,
        "column_count": 5,
        "truncated": False,
    }
    default_payload: dict[str, object] = {
        "available": False,
        "has_rows": False,
        "output_file": "purchase_orders.csv",
        "records_synced": 0,
        "records_synced_display": "0",
        "last_sync_time": "",
        "total_amount": 0.0,
        "total_amount_display": "0.00",
        "total_spend_lines": [],
        "supplier_count": 0,
        "supplier_count_display": "0",
        "preview": default_preview,
        "message": "No purchase orders found",
    }
    try:
        ccc_api = _load_stage1_api_source_module("ccc_purchase_orders")
        raw = dict(ccc_api.load_purchase_orders_cache_summary(project_id=project_id))
    except Exception as exc:
        default_payload["message"] = f"Could not read cached purchase orders: {exc}"
        return default_payload
    records_synced = int(raw.get("records_synced") or 0)
    total_amount = _safe_float(raw.get("total_amount") or 0.0)
    supplier_count = int(raw.get("supplier_count") or 0)
    tb = raw.get("totals_by_currency")
    spend_lines: list[str] = []
    if isinstance(tb, list):
        rows_cur = [x for x in tb if isinstance(x, dict)]
        rows_cur.sort(key=lambda x: str(x.get("currency") or ""))
        for item in rows_cur:
            cur = str(item.get("currency") or "").strip() or "—"
            amt = _safe_float(item.get("total"))
            spend_lines.append(f"{cur} {amt:,.2f}")
    preview = raw.get("preview") if isinstance(raw.get("preview"), dict) else default_preview
    available = bool(raw.get("available"))
    return {
        "available": available,
        "has_rows": records_synced > 0,
        "output_file": str(raw.get("output_file") or "purchase_orders.csv"),
        "records_synced": records_synced,
        "records_synced_display": f"{records_synced:,}",
        "last_sync_time": str(raw.get("last_sync_time") or ""),
        "total_amount": total_amount,
        "total_amount_display": spend_lines[0] if len(spend_lines) == 1 else (" · ".join(spend_lines) if spend_lines else "—"),
        "total_spend_lines": spend_lines,
        "supplier_count": supplier_count,
        "supplier_count_display": f"{supplier_count:,}",
        "preview": preview,
        "message": "" if records_synced > 0 else "No purchase orders found",
    }


def _resolve_ccc_selected_project_id(projects: list[dict[str, object]], preferred_project_id: int | None) -> int | None:
    if preferred_project_id is not None:
        for project in projects:
            project_id = _coerce_ccc_project_id(project.get("id"))
            if project_id == preferred_project_id:
                return project_id
    if projects:
        return _coerce_ccc_project_id(projects[0].get("id"))
    return preferred_project_id


def _ccc_de_import_panel_projects(projects: list[dict[str, object]] | None) -> list[dict[str, object]]:
    """Enrich CCC API project dropdown rows with site-tag registry company/country for admin import UI."""
    from frontend.services.site_tag_service import resolve_registered_project

    rows: list[dict[str, object]] = []
    for p in projects or []:
        lab = str(p.get("label") or "").strip()
        if not lab:
            continue
        reg = resolve_registered_project(lab)
        site = str(reg.get("platform_site_tag") or "").strip() if reg else ""
        company = str(reg.get("responsible_company") or "").strip() if reg else ""
        country = str(reg.get("project_location") or "").strip() if reg else ""
        subtitle = " • ".join(part for part in (company, country) if part)
        if not subtitle:
            subtitle = "Not in site registry" if not reg else ""
        rows.append(
            {
                "label": lab,
                "project_id": p.get("id"),
                "primary": site or lab,
                "subtitle": subtitle,
                "registered": bool(reg),
                "company": company,
                "country": country,
            }
        )
    return rows


def _ccc_ui_context(
    *,
    selected_project_id: int | None = None,
    base_url: str | None = None,
    username: str | None = None,
    password: str | None = None,
    refresh_projects: bool = False,
) -> dict[str, object]:
    defaults = _ccc_runtime_defaults()
    endpoints = _load_ccc_get_endpoints()
    projects = _load_ccc_available_projects(
        base_url=base_url or defaults["base_url"],
        username=username or defaults["username"],
        password=password or "",
        force_refresh=refresh_projects,
    )
    resolved_project_id = _resolve_ccc_selected_project_id(projects, selected_project_id)
    latest_sync = _latest_run_history_entry("ccc_api_source", statuses={"success"}, types={"ccc_api_sync"})
    latest_sync_endpoint = str((latest_sync or {}).get("endpoint") or "").strip().lower()
    context = (
        _analytics_output_context_for_path("ccc_api_source", None)
        if latest_sync_endpoint == "purchase_order"
        else _analytics_output_context("ccc_api_source")
    )
    latest_test = _latest_run_history_entry("ccc_api_source", types={"ccc_api_test", "ccc_api_sync"})
    purchase_orders_summary = _load_ccc_purchase_orders_summary(resolved_project_id)
    if latest_test:
        test_status = _ccc_connection_status_payload(
            state="connected" if str(latest_test.get("status") or "") == "success" else "failed",
            message=str(latest_test.get("message") or ""),
            tested_at=str(latest_test.get("timestamp") or ""),
        )
    elif defaults["base_url"] and defaults["username"] and defaults["has_password"]:
        test_status = _ccc_connection_status_payload(
            state="configured",
            message="Stored credentials found in environment configuration.",
        )
    else:
        test_status = _ccc_connection_status_payload(
            state="missing",
            message="Set CCC credentials in config/api_credentials.env or environment variables.",
        )
    context.update(
        {
            "form_state": {
                "base_url": defaults["base_url"],
                "username": defaults["username"],
                "page_size": defaults["page_size"],
                "has_password": defaults["has_password"],
                "project_id": resolved_project_id,
                "endpoint_name": next(iter(endpoints.keys()), "purchase_order"),
                "path_params": "",
                "query_params": "",
            },
            "ccc_connection_status": test_status,
            "ccc_last_sync": latest_sync,
            "ccc_available_projects": projects,
            "ccc_selected_project_id": resolved_project_id,
            "ccc_purchase_orders": purchase_orders_summary,
            "ccc_mapping_rules": _load_ccc_mapping_rules(),
            "ccc_get_endpoints": endpoints,
        }
    )
    return context


def _test_ccc_connection_from_ui(*, base_url: str, username: str, password: str) -> dict[str, object]:
    ccc_client = _load_stage1_api_source_module("ccc_client")
    return dict(
        ccc_client.test_connection(
            base_url=str(base_url or "").strip(),
            username=str(username or "").strip(),
            password=str(password or ""),
        )
    )


def _run_ccc_generic_ingest_from_ui(*, endpoint_name: str, base_url: str, username: str, password: str, page_size: int, path_params: dict[str, object], query_params: dict[str, object]) -> tuple[Path, dict[str, object]]:
    ccc_ingest = _load_stage1_api_source_module("ccc_generic_ingest")
    result = ccc_ingest.ingest_endpoint(
        str(endpoint_name or "").strip(),
        base_url=str(base_url or "").strip(),
        username=str(username or "").strip(),
        password=str(password or ""),
        page_size=int(page_size or 100),
        path_params=path_params,
        query_params=query_params,
    )
    output_path = Path(result.get("output_path"))
    return output_path, dict(result)


def _run_ccc_purchase_orders_from_ui(
    *,
    base_url: str,
    username: str,
    password: str,
    page_size: int,
    project_id: int | None,
    query_params: dict[str, object] | None,
) -> tuple[Path, dict[str, object]]:
    ccc_api = _load_stage1_api_source_module("ccc_purchase_orders")
    result = ccc_api.sync_purchase_orders_cache(
        base_url=str(base_url or "").strip(),
        username=str(username or "").strip(),
        password=str(password or ""),
        project_id=project_id,
        query_params=dict(query_params or {}),
    )
    output_path = Path(result.get("output_path"))
    return output_path, dict(result)


def _scope_label(scope_key: str) -> str:
    labels = {
        "scope1": "Scope 1",
        "scope2": "Scope 2",
        "scope3": "Scope 3",
        "total": "Total footprint",
    }
    return labels.get(str(scope_key or "").strip().lower(), str(scope_key or ""))


def _resolve_country_code_and_name(raw_country: str | None) -> tuple[str | None, str | None]:
    raw = str(raw_country or "").strip()
    if not raw:
        return None, None
    code = raw.upper()
    if code in ISO_COUNTRY_NAME_BY_CODE:
        return code, ISO_COUNTRY_NAME_BY_CODE[code]
    by_name = ISO_COUNTRY_CODE_BY_NAME.get(raw.casefold())
    if by_name:
        return by_name, ISO_COUNTRY_NAME_BY_CODE.get(by_name)
    return None, raw


def _company_country_lookup() -> dict[str, tuple[str | None, str | None]]:
    lookup: dict[str, tuple[str | None, str | None]] = {}
    rows = User.query.filter(User.company_name.isnot(None)).all()
    for user in rows:
        company_name = str(getattr(user, "company_name", "") or "").strip()
        if not company_name:
            continue
        keys = {_normalize_template_key(company_name)}
        canon, inferred_country = _canonical_company_name_and_country(company_name)
        if canon:
            keys.add(_normalize_template_key(canon))
        country_code, country_name = _resolve_country_code_and_name(getattr(user, "company_country", None))
        if country_code is None and inferred_country:
            country_code, country_name = _resolve_country_code_and_name(inferred_country)
        if country_code is None and country_name is None:
            continue
        for key in keys:
            if key and key not in lookup:
                lookup[key] = (country_code, country_name)
    return lookup


def _load_emissions_map_points() -> dict[str, object]:
    path = _latest_logged_output_path("emissions_totals")
    points: list[dict[str, object]] = []
    context: dict[str, object] = {
        "output_exists": False,
        "output_file_name": "",
        "output_file_mtime": "",
        "points": points,
        "total_emissions": 0.0,
        "companies_count": 0,
        "countries_count": 0,
        "empty_message": "No totals workbook has been generated from the UI yet. Run Totals Tables first to refresh the map.",
    }
    if path is None or not path.exists():
        return context
    try:
        xls = pd.ExcelFile(path, engine="openpyxl")
        preferred_sheet = None
        for candidate in ("Company Totals Window", "Company Totals"):
            if candidate in xls.sheet_names:
                preferred_sheet = candidate
                break
        if preferred_sheet is None:
            raise RuntimeError("No company totals sheet found in the latest workbook.")
        df = pd.read_excel(xls, sheet_name=preferred_sheet, engine="openpyxl")
    except Exception as exc:
        context["empty_message"] = f"Map output was found but could not be read: {exc}"
        return context

    company_col = next((c for c in df.columns if str(c).strip().lower() == "company"), None)
    emissions_col = next((c for c in df.columns if str(c).strip().lower() == "co2e (t)"), None)
    share_col = next((c for c in df.columns if str(c).strip().lower() == "share (%)"), None)
    if company_col is None or emissions_col is None:
        context["empty_message"] = "Latest workbook is missing the expected Company Totals columns."
        return context

    country_lookup = _company_country_lookup()
    country_seen: set[str] = set()
    total_emissions = 0.0
    for _, row in df.iterrows():
        company_name = str(row.get(company_col) or "").strip()
        if not company_name:
            continue
        emissions = _safe_float(row.get(emissions_col) or 0.0)
        if emissions <= 0:
            continue
        share_pct = _safe_float(row.get(share_col) or 0.0) if share_col else 0.0
        norm = _normalize_template_key(company_name)
        country_code, country_name = country_lookup.get(norm, (None, None))
        if country_code is None and country_name is None:
            _canon, inferred_country = _canonical_company_name_and_country(company_name)
            country_code, country_name = _resolve_country_code_and_name(inferred_country)
        if not country_name:
            continue
        total_emissions += emissions
        if country_code:
            country_seen.add(country_code)
        points.append(
            {
                "company_name": company_name,
                "country_code": country_code or "",
                "country_name": country_name,
                "emissions": emissions,
                "share_pct": share_pct,
            }
        )

    if not points:
        return context
    stat = path.stat()
    context.update(
        {
            "output_exists": True,
            "output_file_name": path.name,
            "output_file_mtime": datetime.fromtimestamp(stat.st_mtime).strftime("%Y-%m-%d %H:%M"),
            "points": points,
            "total_emissions": total_emissions,
            "companies_count": len(points),
            "countries_count": len(country_seen),
            "empty_message": "",
        }
    )
    return context


@app.route("/api/mapping/download_merged", methods=["GET"])
@login_required
def api_mapping_download_merged():
    if not bool(getattr(current_user, "is_admin", False)):
        flash("Access denied")
        return redirect(url_for("dashboard"))
    p = _find_latest_merged_mapping_workbook()
    if p is None or not p.exists():
        flash("Merged mapping output file not found.")
        return redirect(url_for("admin_mapping_panel"))
    return send_file(str(p), as_attachment=True, download_name=p.name)


@app.route("/api/admin/upload_notifications", methods=["GET"])
@login_required
def api_admin_upload_notifications():
    if not bool(getattr(current_user, "is_admin", False)):
        return jsonify({"error": "Access denied"}), 403

    _ensure_db_tables()
    one_week_ago = datetime.utcnow() - timedelta(days=7)
    batches = _list_admin_data_entry_batches()
    notifications: list[dict[str, object]] = []
    for b in batches:
        uploaded_at = b.get("uploaded_at")
        if not isinstance(uploaded_at, datetime):
            continue
        if uploaded_at < one_week_ago:
            continue

        mapped = bool(b.get("mapped"))
        mapped_at = b.get("mapped_at")
        mapped_by = str(b.get("mapped_by") or "").strip()

        notifications.append(
            {
                "company_name": str(b.get("company_name") or ""),
                "uploaded_by_user": str(b.get("uploaded_by_user") or "Unknown"),
                "uploaded_by_user_id": int(b.get("uploaded_by_user_id") or 0),
                "uploaded_by_job_title": str(b.get("uploaded_by_job_title") or ""),
                "uploaded_by_has_profile_photo": bool(b.get("uploaded_by_has_profile_photo")),
                "upload_timestamp": uploaded_at.strftime("%Y-%m-%d %H:%M"),
                "category": str(b.get("sheet_name") or ""),
                "row_count": int(b.get("row_count") or 0),
                "mapping_status": str(b.get("mapping_status_label") or ""),
                "mapping_state": str(b.get("mapping_state") or ""),
                "mapping_counts": b.get("mapping_counts") if isinstance(b.get("mapping_counts"), dict) else {},
                "mapped_by_admin": mapped_by,
                "mapping_timestamp": mapped_at.strftime("%Y-%m-%d %H:%M") if isinstance(mapped_at, datetime) else "",
                "mapped": mapped,
            }
        )

    notifications.sort(key=lambda item: str(item.get("upload_timestamp") or ""), reverse=True)
    return jsonify({"notifications": notifications, "count": len(notifications)})


@app.route("/api/mapping/download/<run_id>", methods=["GET"])
@login_required
def api_mapping_download(run_id: str):
    _ensure_db_tables()
    mr = MappingRun.query.get(run_id)
    if not mr:
        # Fallback to legacy in-memory if present
        meta = _MAPPING_RUNS.get(run_id)
        if not meta:
            flash("Mapping output not found (expired).")
            return redirect(url_for("dashboard"))
        if not bool(getattr(current_user, "is_admin", False)) and meta.get("user_id") != current_user.id:
            flash("Access denied")
            return redirect(url_for("dashboard"))
        p = meta.get("path")
        if not p or not os.path.exists(str(p)):
            flash("Mapping output file not found.")
            return redirect(url_for("dashboard"))
        fn = f"mapped_{meta.get('company')}_{meta.get('sheet')}.xlsx"
        fn = secure_filename(fn) or "mapped_results.xlsx"
        return send_file(str(p), as_attachment=True, download_name=fn)

    if not bool(getattr(current_user, "is_admin", False)) and mr.user_id != current_user.id:
        flash("Access denied")
        return redirect(url_for("dashboard"))
    if not mr.output_path or not os.path.exists(mr.output_path):
        flash("Mapping output file not found.")
        return redirect(url_for("dashboard"))

    fn = secure_filename(f"mapped_{mr.company_name}_{mr.sheet_name}.xlsx") or "mapped_results.xlsx"
    return send_file(str(mr.output_path), as_attachment=True, download_name=fn)


@app.route("/mappings", methods=["GET"])
@login_required
def mapping_runs():
    _ensure_db_tables()
    runs = (
        MappingRun.query.filter_by(user_id=current_user.id)
        .order_by(MappingRun.created_at.desc())
        .limit(200)
        .all()
    )
    return render_template("mapping_runs.html", user=current_user, runs=runs)


@app.route("/admin/mapping_runs", methods=["GET"])
@login_required
def admin_mapping_runs():
    _ensure_db_tables()
    if not bool(getattr(current_user, "is_admin", False)):
        flash("Access denied")
        return redirect(url_for("dashboard"))

    rows = (
        db.session.query(MappingRun, User.email, User.company_name)
        .join(User, MappingRun.user_id == User.id)
        .order_by(MappingRun.created_at.desc())
        .limit(500)
        .all()
    )
    return render_template("mapping_runs_admin.html", user=current_user, rows=rows)


@app.route("/admin/tower-defense", methods=["GET"])
@login_required
def admin_tower_defense():
    if not bool(getattr(current_user, "is_admin", False)):
        flash("Access denied")
        return redirect(url_for("dashboard"))
    return render_template("admin_tower_defense.html", user=current_user)


@app.route("/admin/mapping/unmapped", methods=["GET"])
@login_required
def admin_unmapped_mappings():
    _ensure_db_tables()
    if not bool(getattr(current_user, "is_admin", False)):
        flash("Access denied")
        return redirect(url_for("dashboard"))

    t0 = time.perf_counter()
    t_db0 = time.perf_counter()
    removed_stale = _refresh_open_unmapped_against_live_rows()
    if removed_stale:
        db.session.commit()
    t_db = (time.perf_counter() - t_db0) * 1000.0

    query, status_filter, company_filter, sheet_filter, search = _unmapped_query_from_request()
    dedupe_mode = (request.args.get("dedupe") or "").strip().lower()
    rows = query.order_by(MappingUnmappedRow.created_at.desc(), MappingUnmappedRow.id.desc()).limit(500).all()
    duplicate_counts: dict[int, int] = {}
    if dedupe_mode == "description":
        rows, duplicate_counts = _dedupe_unmapped_rows_by_description(rows)
    previews = []
    ef_options_by_row: dict[int, list[dict[str, object]]] = {}
    ef_cache_by_sheet: dict[str, list[dict[str, object]]] = {}
    t_unmapped0 = time.perf_counter()
    for row in rows:
        preview = _unmapped_row_preview(row)
        preview["description"] = _unmapped_description(row)
        preview["duplicate_count"] = duplicate_counts.get(int(row.id), 1)
        previews.append(preview)
        sn = str(row.sheet_name or "").strip()
        if sn not in ef_cache_by_sheet:
            ef_cache_by_sheet[sn] = _load_unmapped_ef_options_for_sheet(sn)
        ef_options_by_row[int(row.id)] = ef_cache_by_sheet[sn]
    t_unmapped_ms = (time.perf_counter() - t_unmapped0) * 1000.0

    counts = _admin_unmapped_page_counts(
        company_filter=company_filter,
        sheet_filter=sheet_filter,
        search=search,
        dedupe_mode=dedupe_mode,
    )
    companies = [
        item[0]
        for item in db.session.query(MappingUnmappedRow.company_name)
        .distinct()
        .order_by(MappingUnmappedRow.company_name.asc())
        .all()
        if item[0]
    ]
    sheets = [
        item[0]
        for item in db.session.query(MappingUnmappedRow.sheet_name)
        .distinct()
        .order_by(MappingUnmappedRow.sheet_name.asc())
        .all()
        if item[0]
    ]
    _perf_log(
        "admin_unmapped_mappings",
        render_ms=(time.perf_counter() - t0) * 1000.0,
        db_ms=t_db,
        unmapped_refresh_ms=t_unmapped_ms,
    )
    return render_template(
        "admin_unmapped_mappings.html",
        user=current_user,
        rows=previews,
        counts=counts,
        status_filter=status_filter,
        company_filter=company_filter,
        sheet_filter=sheet_filter,
        search=search,
        dedupe_mode=dedupe_mode,
        ef_options_by_row=ef_options_by_row,
        fuzzy_suggestions=[],
        fuzzy_scanned_count=0,
        fuzzy_unmatched_count=0,
        fuzzy_threshold=0.92,
        companies=companies,
        sheets=sheets,
    )


@app.route("/admin/mapping/unmapped/<int:row_id>/update", methods=["POST"])
@login_required
def admin_unmapped_mapping_update(row_id: int):
    _ensure_db_tables()
    if not bool(getattr(current_user, "is_admin", False)):
        flash("Access denied")
        return redirect(url_for("dashboard"))

    row = MappingUnmappedRow.query.get_or_404(row_id)
    review_status = (request.form.get("review_status") or "open").strip().lower()
    if review_status not in {"open", "resolved", "ignored"}:
        review_status = "open"
    owner_notes = (request.form.get("owner_notes") or "").strip() or None
    ef_key = (request.form.get("assigned_ef_key") or request.form.get("assigned_ef_id") or "").strip()
    ef_option = _find_unmapped_ef_option(ef_key)
    action = (request.form.get("action") or "save").strip().lower()

    try:
        if action == "map":
            if ef_option is None:
                flash("Choose a valid emission factor before running Map.")
            else:
                target_sheet, resolved_count, inserted = _apply_unmapped_map_full(
                    row,
                    ef_option,
                    owner_notes=owner_notes,
                    match_method="manual:unmapped_map",
                )
                db.session.commit()
                action_note = "added to" if inserted else "already existed in"
                flash(
                    f"Mapped {resolved_count} unmapped row(s); data entry updated; EF mapping {action_note} {target_sheet}."
                )
        else:
            row.review_status = review_status
            row.assigned_ef_id = str(ef_option.get("ef_id") or "").strip() if ef_option else (request.form.get("assigned_ef_id") or "").strip() or None
            row.owner_notes = owner_notes
            if review_status in {"resolved", "ignored"}:
                row.resolved_at = datetime.utcnow()
                row.resolved_by_user_id = int(getattr(current_user, "id", 0) or 0) or None
            else:
                row.resolved_at = None
                row.resolved_by_user_id = None
            db.session.commit()
            if review_status == "resolved" and not row.assigned_ef_id:
                flash("Marked resolved without an EF assignment.")
            elif review_status == "resolved" and row.assigned_ef_id:
                flash("Saved. Use Map to write the EF mapping, recalculate emissions, and update the data entry row.")
            else:
                flash("Unmapped row updated.")
    except PermissionError:
        db.session.rollback()
        flash("Mapping workbook is open/locked. Close Excel and try again.")
    except Exception as e:
        db.session.rollback()
        flash(f"Unmapped mapping failed: {e}")
    return redirect(
        url_for(
            "admin_unmapped_mappings",
            status=request.args.get("status", "open"),
            company=request.args.get("company", ""),
            sheet=request.args.get("sheet", ""),
            search=request.args.get("search", ""),
            dedupe=request.args.get("dedupe", ""),
        )
    )


@app.route("/admin/mapping/unmapped/fuzzy", methods=["GET"])
@login_required
def admin_unmapped_mappings_fuzzy():
    _ensure_db_tables()
    if not bool(getattr(current_user, "is_admin", False)):
        flash("Access denied")
        return redirect(url_for("dashboard"))

    removed_stale = _refresh_open_unmapped_against_live_rows()
    if removed_stale:
        db.session.commit()

    query, status_filter, company_filter, sheet_filter, search = _unmapped_query_from_request()
    rows = query.order_by(MappingUnmappedRow.created_at.desc(), MappingUnmappedRow.id.desc()).limit(500).all()
    dedupe_mode = (request.args.get("dedupe") or "").strip().lower()
    fuzzy_threshold = _parse_unmapped_fuzzy_threshold(request.args.get("threshold"), default=0.92)
    duplicate_counts: dict[int, int] = {}
    if dedupe_mode == "description":
        rows, duplicate_counts = _dedupe_unmapped_rows_by_description(rows)
    suggestions = _build_unmapped_fuzzy_suggestions(rows, threshold=fuzzy_threshold)
    suggested_ids = {int(suggestion["row"]["id"]) for suggestion in suggestions}
    fuzzy_scanned_count = len(rows)
    fuzzy_unmatched_count = max(0, fuzzy_scanned_count - len(suggested_ids))
    previews = []
    ef_options_by_row: dict[int, list[dict[str, object]]] = {}
    ef_cache_by_sheet: dict[str, list[dict[str, object]]] = {}
    for row in rows:
        preview = _unmapped_row_preview(row)
        preview["description"] = _unmapped_description(row)
        preview["duplicate_count"] = duplicate_counts.get(int(row.id), 1)
        previews.append(preview)
        sn = str(row.sheet_name or "").strip()
        if sn not in ef_cache_by_sheet:
            ef_cache_by_sheet[sn] = _load_unmapped_ef_options_for_sheet(sn)
        ef_options_by_row[int(row.id)] = ef_cache_by_sheet[sn]
    counts = _admin_unmapped_page_counts(
        company_filter=company_filter,
        sheet_filter=sheet_filter,
        search=search,
        dedupe_mode=dedupe_mode,
    )
    companies = [
        item[0]
        for item in db.session.query(MappingUnmappedRow.company_name)
        .distinct()
        .order_by(MappingUnmappedRow.company_name.asc())
        .all()
        if item[0]
    ]
    sheets = [
        item[0]
        for item in db.session.query(MappingUnmappedRow.sheet_name)
        .distinct()
        .order_by(MappingUnmappedRow.sheet_name.asc())
        .all()
        if item[0]
    ]
    return render_template(
        "admin_unmapped_mappings.html",
        user=current_user,
        rows=previews,
        counts=counts,
        status_filter=status_filter,
        company_filter=company_filter,
        sheet_filter=sheet_filter,
        search=search,
        dedupe_mode=dedupe_mode,
        ef_options_by_row=ef_options_by_row,
        fuzzy_suggestions=suggestions,
        fuzzy_scanned_count=fuzzy_scanned_count,
        fuzzy_unmatched_count=fuzzy_unmatched_count,
        fuzzy_threshold=fuzzy_threshold,
        companies=companies,
        sheets=sheets,
    )


@app.route("/admin/mapping/unmapped/fuzzy/approve", methods=["POST"])
@login_required
def admin_unmapped_mappings_fuzzy_approve():
    _ensure_db_tables()
    if not bool(getattr(current_user, "is_admin", False)):
        flash("Access denied")
        return redirect(url_for("dashboard"))

    selections = request.form.getlist("suggestion")
    if not selections:
        flash("Choose at least one fuzzy suggestion to approve.")
        return redirect(url_for("admin_unmapped_mappings_fuzzy", status=request.args.get("status", "open"), company=request.args.get("company", ""), sheet=request.args.get("sheet", ""), search=request.args.get("search", ""), dedupe=request.args.get("dedupe", ""), threshold=request.args.get("threshold", "0.92")))

    approved = 0
    try:
        for selection in selections:
            row_id_text, sep, tail = selection.partition("|")
            if not sep:
                continue
            ef_key, sep2, match_method = tail.partition("|")
            try:
                row_id = int(row_id_text)
            except Exception:
                continue
            row = MappingUnmappedRow.query.get(row_id)
            ef_option = _find_unmapped_ef_option(ef_key)
            if row is None or ef_option is None:
                continue
            mm = str(match_method or "").strip() or "Cat1: All together fuzzy"
            _apply_unmapped_map_full(row, ef_option, match_method=mm)
            approved += 1
        db.session.commit()
        flash(f"Approved {approved} fuzzy mapping suggestion(s).")
    except PermissionError:
        db.session.rollback()
        flash("Mapping workbook is open/locked. Close Excel and try again.")
    except Exception as e:
        db.session.rollback()
        flash(f"Fuzzy approval failed: {e}")
    return redirect(url_for("admin_unmapped_mappings", status=request.args.get("status", "open"), company=request.args.get("company", ""), sheet=request.args.get("sheet", ""), search=request.args.get("search", ""), dedupe=request.args.get("dedupe", "")))


@app.route("/admin/mapping/unmapped/export", methods=["GET"])
@login_required
def admin_unmapped_mappings_export():
    _ensure_db_tables()
    if not bool(getattr(current_user, "is_admin", False)):
        flash("Access denied")
        return redirect(url_for("dashboard"))

    query, _status_filter, _company_filter, _sheet_filter, _search = _unmapped_query_from_request()
    rows = query.order_by(MappingUnmappedRow.created_at.desc(), MappingUnmappedRow.id.desc()).all()
    dedupe_mode = (request.args.get("dedupe") or "").strip().lower()
    duplicate_counts: dict[int, int] = {}
    if dedupe_mode == "description":
        rows, duplicate_counts = _dedupe_unmapped_rows_by_description(rows)
    export_rows: list[dict[str, object]] = []
    for row in rows:
        preview = _unmapped_row_preview(row)
        export_rows.append(
            {
                "id": row.id,
                "review_status": row.review_status,
                "company_name": row.company_name,
                "sheet_name": row.sheet_name,
                "source_entry_group": row.source_entry_group or "",
                "run_id": row.run_id,
                "row_number": row.row_number,
                "row_label": row.row_label or "",
                "status_value": row.status_value or "",
                "assigned_ef_id": row.assigned_ef_id or "",
                "owner_notes": row.owner_notes or "",
                "created_at": row.created_at.isoformat() if row.created_at else "",
                "description": _unmapped_description(row),
                "duplicate_count": duplicate_counts.get(int(row.id), 1),
                "details": "; ".join(f"{d['name']}={d['value']}" for d in preview["details"]),
            }
        )
    out = StringIO()
    fieldnames = [
        "id",
        "review_status",
        "company_name",
        "sheet_name",
        "source_entry_group",
        "run_id",
        "row_number",
        "row_label",
        "status_value",
        "assigned_ef_id",
        "owner_notes",
        "created_at",
        "description",
        "duplicate_count",
        "details",
    ]
    writer = csv.DictWriter(out, fieldnames=fieldnames)
    writer.writeheader()
    writer.writerows(export_rows)
    return Response(
        out.getvalue(),
        mimetype="text/csv",
        headers={
            "Content-Disposition": (
                "attachment; filename="
                + ("unmapped_mapping_rows_without_duplications.csv" if dedupe_mode == "description" else "unmapped_mapping_rows.csv")
            )
        },
    )


@app.route("/admin/mapping", methods=["GET"])
@login_required
def admin_mapping_panel():
    _ensure_db_tables()
    if not bool(getattr(current_user, "is_admin", False)):
        flash("Access denied")
        return redirect(url_for("dashboard"))
    _ensure_mapping_run_source_entry_group_column()
    batches = _list_admin_data_entry_batches()
    batches_json = _batches_for_admin_mapping_json(batches)
    return render_template(
        "admin_mapping.html",
        user=current_user,
        batches=batches,
        batches_json=batches_json,
    )


@app.route("/admin/background-jobs", methods=["GET"])
@login_required
def admin_background_jobs():
    if not bool(getattr(current_user, "is_admin", False)):
        flash("Access denied")
        return redirect(url_for("dashboard"))
    return render_template("admin_background_jobs.html", user=current_user)


@app.route("/analytics/output/download/<kind>", methods=["GET"])
@login_required
def analytics_output_download(kind: str):
    cfg = ANALYTICS_OUTPUT_VIEWS.get(kind)
    if not cfg:
        flash("Output not found.")
        return redirect(url_for("home"))
    path = _latest_logged_output_path(kind)
    if path is None or not path.exists():
        flash("No UI-generated output file found.")
        return redirect(url_for("home"))
    return send_file(str(path), as_attachment=True, download_name=path.name)


@app.route("/analytics/output/file/<path:filename>", methods=["GET"])
@login_required
def analytics_output_download_file(filename: str):
    safe_name = Path(str(filename or "")).name
    if not safe_name:
        flash("Output file not found.")
        return redirect(url_for("home"))
    candidates: list[Path] = []
    for root in (STAGE2_OUTPUT_DIR, STAGE1_INPUT_DIR, DATA_DIR):
        candidates.extend([p for p in root.rglob(safe_name) if p.is_file()])
    if not candidates:
        flash("Output file not found.")
        return redirect(url_for("home"))
    candidates.sort(key=lambda p: p.stat().st_mtime, reverse=True)
    path = candidates[0]
    return send_file(str(path), as_attachment=True, download_name=path.name)


def _employee_commuting_data_source_context(
    *,
    page_title: str,
    page_heading: str,
    page_description: str,
    save_url: str,
    fields: tuple[dict[str, str], ...],
    rows: list[dict[str, object]],
) -> dict[str, object]:
    return {
        "page_title": page_title,
        "page_heading": page_heading,
        "page_description": page_description,
        "save_url": save_url,
        "fields": list(fields),
        "rows": rows,
        "published_workbook_name": EMPLOYEE_COMMUTING_NATIONAL_AVERAGES_XLSX.name,
    }


@app.route("/data-sources/employee-commuting/headcount", methods=["GET", "POST"])
@login_required
def data_sources_employee_commuting_headcount():
    if not bool(getattr(current_user, "is_admin", False)):
        flash("Access denied")
        return redirect(url_for("dashboard"))

    _ensure_db_tables()
    _seed_employee_commuting_defaults()

    if request.method == "POST":
        payload = request.get_json(silent=True) or {}
        try:
            rows = _normalize_employee_commuting_headcount_payload(payload.get("rows"))
            _replace_employee_commuting_headcount_rows(rows)
            return jsonify(
                {
                    "ok": True,
                    "saved_rows": len(rows),
                    "rows": _serialize_employee_commuting_headcount_rows(),
                    "published_workbook": EMPLOYEE_COMMUTING_NATIONAL_AVERAGES_XLSX.name,
                }
            )
        except ValueError as exc:
            db.session.rollback()
            return jsonify({"error": str(exc)}), 400
        except Exception as exc:
            db.session.rollback()
            return jsonify({"error": f"Failed to save headcount: {exc}"}), 500

    context = _employee_commuting_data_source_context(
        page_title="Employee Commuting Headcount",
        page_heading="Employee Commuting Headcount",
        page_description=(
            "Manage monthly employee commuting headcount values here. "
            "The national averages workbook is updated automatically after each save."
        ),
        save_url=url_for("data_sources_employee_commuting_headcount"),
        fields=EMPLOYEE_COMMUTING_HEADCOUNT_FIELDS,
        rows=_serialize_employee_commuting_headcount_rows(),
    )
    return render_template("data_source_table.html", user=current_user, **context)


@app.route("/data-sources/employee-commuting/national-averages", methods=["GET", "POST"])
@login_required
def data_sources_employee_commuting_national_averages():
    if not bool(getattr(current_user, "is_admin", False)):
        flash("Access denied")
        return redirect(url_for("dashboard"))

    _ensure_db_tables()
    _seed_employee_commuting_defaults()

    if request.method == "POST":
        payload = request.get_json(silent=True) or {}
        try:
            rows = _normalize_employee_commuting_national_average_payload(payload.get("rows"))
            _replace_employee_commuting_national_average_rows(rows)
            return jsonify(
                {
                    "ok": True,
                    "saved_rows": len(rows),
                    "rows": _serialize_employee_commuting_national_average_rows(),
                    "published_workbook": EMPLOYEE_COMMUTING_NATIONAL_AVERAGES_XLSX.name,
                }
            )
        except ValueError as exc:
            db.session.rollback()
            return jsonify({"error": str(exc)}), 400
        except Exception as exc:
            db.session.rollback()
            return jsonify({"error": f"Failed to save national averages: {exc}"}), 500

    context = _employee_commuting_data_source_context(
        page_title="Employee Commuting National Averages",
        page_heading="Employee Commuting National Averages",
        page_description=(
            "Manage country, average one day, and mode of transport shares here. "
            "These values are combined with the headcount page and used in the Category 7 calculation."
        ),
        save_url=url_for("data_sources_employee_commuting_national_averages"),
        fields=EMPLOYEE_COMMUTING_NATIONAL_AVERAGE_FIELDS,
        rows=_serialize_employee_commuting_national_average_rows(),
    )
    return render_template("data_source_table.html", user=current_user, **context)


@app.route("/data-sources/ccc-api", methods=["GET", "POST"])
@login_required
def data_sources_ccc_api():
    selected_project_id = _coerce_ccc_project_id(request.args.get("project_id"))
    context = _ccc_ui_context(selected_project_id=selected_project_id)
    form_state = dict(context.get("form_state") or {})
    if request.method == "POST":
        action = str(request.form.get("action", "sync") or "sync").strip().lower()
        form_state = {
            "base_url": str(request.form.get("base_url", "") or "").strip(),
            "username": str(request.form.get("username", "") or "").strip(),
            "page_size": _parse_int_form("page_size", 100, minimum=1, maximum=500),
            "project_id": _coerce_ccc_project_id(request.form.get("project_id")),
            "endpoint_name": str(request.form.get("endpoint_name", "") or "").strip(),
            "path_params": str(request.form.get("path_params", "") or "").strip(),
            "query_params": str(request.form.get("query_params", "") or "").strip(),
            "has_password": bool(str(request.form.get("password", "") or "").strip()) or bool(_ccc_runtime_defaults().get("has_password")),
        }
        password = str(request.form.get("password", "") or "")
        try:
            if action == "test_connection":
                result = _test_ccc_connection_from_ui(
                    base_url=form_state["base_url"],
                    username=form_state["username"],
                    password=password,
                )
                context = _ccc_ui_context(
                    selected_project_id=form_state["project_id"],
                    base_url=form_state["base_url"],
                    username=form_state["username"],
                    password=password,
                    refresh_projects=True,
                )
                context["run_notice"] = "CCC API connection test succeeded. JWT token received."
                context["ccc_connection_status"] = _ccc_connection_status_payload(
                    state="connected",
                    message="JWT token received successfully.",
                    tested_at=datetime.now().strftime("%Y-%m-%d %H:%M"),
                )
                _append_run_history(
                    "ccc_api_source",
                    {
                        "timestamp": datetime.now().strftime("%Y-%m-%d %H:%M"),
                        "type": "ccc_api_test",
                        "status": "success",
                        "scenario": "Connection test",
                        "parameters_summary": f"Base URL: {form_state['base_url']}",
                        "base_url": form_state["base_url"],
                        "page_size": int(form_state["page_size"]),
                        "records_imported": 0,
                        "output_filename": "",
                        "output_file": "",
                        "message": str(result.get("token_received") and "JWT token received successfully." or "Connection test completed."),
                    },
                )
                _create_user_notification(
                    current_user.id,
                    title="API connection established",
                    message=f"{_clean_company_name(getattr(current_user, 'company_name', '') or '') or 'Platform'} connected to CCC API.",
                    notification_type="success",
                    link=url_for("data_sources_ccc_api"),
                    feed_event="api_connection",
                    feed_company=_clean_company_name(getattr(current_user, "company_name", "") or "") or "Platform",
                    feed_api_name="CCC API",
                    feed_timestamp=datetime.utcnow(),
                )
            else:
                endpoint_name = str(form_state["endpoint_name"] or "").strip()
                if endpoint_name == "purchase_order":
                    query_params = _parse_json_object_form_field("query_params")
                    project_id = _coerce_ccc_project_id(form_state.get("project_id"))
                    if project_id is None:
                        project_id = _coerce_ccc_project_id((context.get("ccc_selected_project_id")))
                    out_path, result = _run_ccc_purchase_orders_from_ui(
                        base_url=form_state["base_url"],
                        username=form_state["username"],
                        password=password,
                        page_size=int(form_state["page_size"]),
                        project_id=project_id,
                        query_params=query_params,
                    )
                    context = _ccc_ui_context(
                        selected_project_id=_coerce_ccc_project_id(result.get("project_id"), project_id),
                        base_url=form_state["base_url"],
                        username=form_state["username"],
                        password=password,
                        refresh_projects=True,
                    )
                    imported = int(result.get("records_imported") or 0)
                    total_amount = _safe_float(result.get("total_amount") or 0.0)
                    supplier_count = int(result.get("supplier_count") or 0)
                    selected_project_id = _coerce_ccc_project_id(result.get("project_id"), project_id)
                    if not imported:
                        context["run_notice"] = "No purchase orders found"
                    else:
                        context["run_notice"] = (
                            f"Purchase orders synced successfully. Imported {imported} records for project {selected_project_id}."
                        )
                    context["ccc_connection_status"] = _ccc_connection_status_payload(
                        state="connected",
                        message="Last sync authenticated successfully.",
                        tested_at=datetime.now().strftime("%Y-%m-%d %H:%M"),
                    )
                    _append_run_history(
                        "ccc_api_source",
                        {
                            "timestamp": datetime.now().strftime("%Y-%m-%d %H:%M"),
                            "type": "ccc_api_sync",
                            "status": "success",
                            "scenario": endpoint_name,
                            "parameters_summary": f"Project ID: {selected_project_id}, Sorting: D, Page size: 200",
                            "base_url": form_state["base_url"],
                            "endpoint": endpoint_name,
                            "project_id": selected_project_id,
                            "page_size": 200,
                            "records_imported": imported,
                            "output_filename": out_path.name,
                            "output_file": out_path.name,
                            "message": (
                                "No purchase orders found"
                                if not imported
                                else f"Imported {imported} records, total spend {total_amount:,.2f}, suppliers {supplier_count}."
                            ),
                        },
                    )
                    _create_user_notification(
                        current_user.id,
                        title="CCC purchase orders synced",
                        message=(
                            f"No purchase orders found for project {selected_project_id}."
                            if not imported
                            else f"Purchase orders synced successfully with {imported} imported records."
                        ),
                        notification_type="success",
                        link=url_for("data_sources_ccc_api"),
                    )
                else:
                    path_params = _parse_json_object_form_field("path_params")
                    query_params = _parse_json_object_form_field("query_params")
                    out_path, result = _run_ccc_generic_ingest_from_ui(
                        endpoint_name=endpoint_name,
                        base_url=form_state["base_url"],
                        username=form_state["username"],
                        password=password,
                        page_size=int(form_state["page_size"]),
                        path_params=path_params,
                        query_params=query_params,
                    )
                    context = _ccc_ui_context(
                        selected_project_id=form_state["project_id"],
                        base_url=form_state["base_url"],
                        username=form_state["username"],
                        password=password,
                    )
                    context = {**context, **_analytics_output_context_for_path("ccc_api_source", out_path)}
                    imported = int(result.get("records_imported") or 0)
                    if not imported:
                        context["run_notice"] = f"Endpoint `{endpoint_name}` synced successfully but returned no rows."
                    else:
                        context["run_notice"] = (
                            f"Endpoint `{endpoint_name}` synced successfully. "
                            f"Imported {imported} records and saved {out_path.name}."
                        )
                    context["run_notice"] = (
                        context["run_notice"]
                    )
                    context["ccc_connection_status"] = _ccc_connection_status_payload(
                        state="connected",
                        message="Last sync authenticated successfully.",
                        tested_at=datetime.now().strftime("%Y-%m-%d %H:%M"),
                    )
                    _append_run_history(
                        "ccc_api_source",
                        {
                            "timestamp": datetime.now().strftime("%Y-%m-%d %H:%M"),
                            "type": "ccc_api_sync",
                            "status": "success",
                            "scenario": endpoint_name,
                            "parameters_summary": (
                                f"Base URL: {form_state['base_url']}, "
                                f"Endpoint: {endpoint_name}, "
                                f"Page size: {int(form_state['page_size'])}, "
                                f"Rows: {imported}"
                            ),
                            "base_url": form_state["base_url"],
                            "endpoint": endpoint_name,
                            "page_size": int(form_state["page_size"]),
                            "records_imported": imported,
                            "output_filename": out_path.name,
                            "output_file": out_path.name,
                            "message": f"Imported {imported} records.",
                        },
                    )
                    _create_user_notification(
                        current_user.id,
                        title="CCC API sync completed",
                        message=f"{endpoint_name or 'Selected endpoint'} synced successfully with {imported} imported records.",
                        notification_type="success",
                        link=url_for("data_sources_ccc_api"),
                    )
            context["run_history"] = _read_run_history("ccc_api_source")
            context["ccc_last_sync"] = _latest_run_history_entry("ccc_api_source", statuses={"success"}, types={"ccc_api_sync"})
        except Exception as exc:
            context = _ccc_ui_context(
                selected_project_id=form_state.get("project_id"),
                base_url=form_state.get("base_url"),
                username=form_state.get("username"),
                password=password,
            )
            error_message = f"CCC API {'connection test' if action == 'test_connection' else 'sync'} failed: {exc}"
            context["run_error"] = error_message
            context["ccc_connection_status"] = _ccc_connection_status_payload(
                state="failed",
                message=str(exc),
                tested_at=datetime.now().strftime("%Y-%m-%d %H:%M"),
            )
            _append_run_history(
                "ccc_api_source",
                {
                    "timestamp": datetime.now().strftime("%Y-%m-%d %H:%M"),
                    "type": "ccc_api_test" if action == "test_connection" else "ccc_api_sync",
                    "status": "failed",
                    "scenario": "Connection test" if action == "test_connection" else (form_state["endpoint_name"] or "CCC endpoint"),
                    "parameters_summary": (
                        f"Base URL: {form_state['base_url']}, "
                        f"Endpoint: {form_state['endpoint_name'] or 'n/a'}, "
                        f"Page size: {int(form_state['page_size'])}"
                    ),
                    "base_url": form_state["base_url"],
                    "endpoint": form_state["endpoint_name"] or "",
                    "page_size": int(form_state["page_size"]),
                    "records_imported": 0,
                    "output_filename": "",
                    "output_file": "",
                    "message": str(exc),
                },
            )
            context["run_history"] = _read_run_history("ccc_api_source")
        context["form_state"] = form_state
    context["can_ccc_data_entry_import"] = _user_can_run_ccc_data_entry_import(current_user)
    context["ccc_de_import_projects"] = (
        _ccc_de_import_panel_projects(context.get("ccc_available_projects"))
        if context["can_ccc_data_entry_import"]
        else []
    )
    context["ccc_import_api_url"] = url_for("api_ccc_import_to_data_entry")
    context["dashboard_url"] = url_for("dashboard")
    context["ccc_de_target_sheet"] = "Scope 3 Category 1 Purchased Goods & Services"
    return render_template("analytics_output.html", user=current_user, **context)


@app.route("/data-sources/averages", methods=["GET"])
@login_required
def data_sources_averages():
    _ensure_db_tables()
    company_name, company_error = _resolve_averages_company(request.args.get("company"))
    if company_error:
        flash(company_error)
        return redirect(url_for("dashboard"))
    row = AveragesData.query.filter_by(company_name=company_name).first() if company_name else None
    return render_template(
        "averages.html",
        user=current_user,
        template_mode=_current_template_mode(),
        company_name=company_name,
        averages_company_options=_allowed_companies_for_averages(),
        averages_owner_can_select_company=_is_owner_user(current_user),
        averages_data=_averages_payload(row),
        country_options=ISO_COUNTRIES,
        waste_type_options=AVERAGES_WASTE_TYPES,
        waste_unit_options=AVERAGES_WASTE_UNITS,
    )


@app.route("/data-sources/scenarios", methods=["GET"])
@login_required
def data_sources_scenarios():
    _ensure_db_tables()
    bundle = _template_bundle_for_company(_resolve_template_company_name(getattr(current_user, "company_name", "") or "") or "")
    rows = ScenariosData.query.filter(ScenariosData.company_name.in_(SCENARIO_COMPANY_OPTIONS)).all()
    initial_company = _clean_company_name(getattr(current_user, "company_name", "") or "")
    if initial_company not in SCENARIO_COMPANY_OPTIONS:
        initial_company = SCENARIO_COMPANY_OPTIONS[0]
    return render_template(
        "scenarios.html",
        user=current_user,
        template_mode=bundle.get("template_mode"),
        enabled_categories=bundle.get("enabled_categories", []),
        disabled_categories=bundle.get("disabled_categories", []),
        scenario_company_options=SCENARIO_COMPANY_OPTIONS,
        scenario_category_config=SCENARIO_CATEGORY_CONFIG,
        scenario_saved_data=_scenario_rows_payload(rows),
        scenario_initial_company=initial_company,
    )


@app.route("/api/averages/save", methods=["GET", "POST"])
@login_required
def api_averages_save():
    _ensure_db_tables()
    company_name, company_error = _resolve_averages_company((request.args.get("company") if request.method == "GET" else (request.get_json(silent=True) or {}).get("company")))
    if company_error or not company_name:
        return jsonify({"error": company_error or "Company is required."}), 400
    if request.method == "GET":
        row = AveragesData.query.filter_by(company_name=company_name).first()
        return jsonify({"ok": True, "data": _averages_payload(row)})

    payload = request.get_json(silent=True) or {}

    waste_type = str(payload.get("waste_type") or "").strip()
    waste_unit = str(payload.get("waste_unit") or "").strip()
    electricity_country = str(payload.get("electricity_country") or "").strip()

    if waste_type and waste_type not in AVERAGES_WASTE_TYPES:
        return jsonify({"error": "Invalid waste type."}), 400
    if waste_unit and waste_unit not in AVERAGES_WASTE_UNITS:
        return jsonify({"error": "Invalid waste unit."}), 400

    row = AveragesData.query.filter_by(company_name=company_name).first()
    if row is None:
        row = AveragesData(company_name=company_name)
        db.session.add(row)

    row.saved_by_user_id = int(current_user.id)
    row.electricity_kwh = _safe_float(payload.get("electricity_kwh"))
    row.electricity_country = electricity_country or None
    row.electricity_emission_factor = _safe_float(payload.get("electricity_emission_factor"))
    row.district_heating_kwh = _safe_float(payload.get("district_heating_kwh"))
    row.district_heating_supplier = str(payload.get("district_heating_supplier") or "").strip() or None
    row.waste_type = waste_type or None
    row.waste_weight = _safe_float(payload.get("waste_weight"))
    row.waste_unit = waste_unit or None
    row.water_total_m3 = _safe_float(payload.get("water_total_m3"))
    row.building_size_m2 = _safe_float(payload.get("building_size_m2"))
    row.water_per_m2 = _positive_ratio(row.water_total_m3, row.building_size_m2)
    db.session.commit()

    mapping_frames = _averages_mapping_frames(company_name, payload)
    mapping_results: list[dict[str, object]] = []
    mapping_errors: list[str] = []
    for sheet_name, frame in mapping_frames:
        result = _run_mapping_for_virtual_sheet(int(current_user.id), company_name, sheet_name, frame)
        if result.get("ok"):
            mapping_results.append(result)
        else:
            mapping_errors.append(str(result.get("error") or f"{sheet_name} mapping failed."))

    if mapping_results:
        _create_user_notification(
            current_user.id,
            title="Averages mapping completed",
            message=f"Calculated emissions based on averages for {company_name}.",
            notification_type="success",
            link=url_for("data_sources_averages", company=company_name),
            feed_event="mapping_completed",
            feed_company=company_name,
            feed_timestamp=datetime.utcnow(),
        )

    return jsonify(
        {
            "ok": True,
            "message": "Saved successfully",
            "mapping_message": (
                "Calculated emissions based on averages"
                if mapping_results and not mapping_errors
                else ("Averages saved, but some mapping steps failed." if mapping_errors else "Saved successfully")
            ),
            "mapping_results": mapping_results,
            "mapping_errors": mapping_errors,
            "water_per_m2": row.water_per_m2,
            "data": _averages_payload(row),
        }
    )


@app.route("/api/scenarios/save", methods=["GET", "POST"])
@login_required
def api_scenarios_save():
    _ensure_db_tables()
    if request.method == "GET":
        company_name = _clean_company_name(request.args.get("company"))
        categories = SCENARIO_CATEGORY_CONFIG.get(company_name)
        if not categories:
            return jsonify({"error": "Select a valid company."}), 400
        row = ScenariosData.query.filter_by(company_name=company_name).first()
        payload = _scenario_rows_payload([row] if row else [])
        return jsonify({"ok": True, "data": payload.get(company_name)})

    payload = request.get_json(silent=True) or {}
    company_name = _clean_company_name(payload.get("company"))
    categories = SCENARIO_CATEGORY_CONFIG.get(company_name)
    if not categories:
        return jsonify({"error": "Select a valid company."}), 400

    raw_inputs = payload.get("inputs")
    if not isinstance(raw_inputs, dict):
        raw_inputs = {}

    clean_inputs = _default_scenario_inputs(categories)
    for category in categories:
        incoming = raw_inputs.get(category)
        if not isinstance(incoming, dict):
            continue
        for field_name in clean_inputs[category].keys():
            clean_inputs[category][field_name] = str(incoming.get(field_name) or "").strip()

    row = ScenariosData.query.filter_by(company_name=company_name).first()
    if row is None:
        row = ScenariosData(company_name=company_name)
        db.session.add(row)

    row.saved_by_user_id = int(current_user.id)
    row.categories_json = json.dumps(clean_inputs)
    db.session.commit()

    return jsonify(
        {
            "ok": True,
            "message": "Saved successfully",
            "company": company_name,
            "data": _scenario_rows_payload([row]).get(company_name),
        }
    )


@app.route("/analytics/forecasting", methods=["GET", "POST"])
@login_required
def analytics_forecasting():
    target_year = _parse_int_form("target_year", 2030, minimum=2026, maximum=2050) if request.method == "POST" else 2030
    context = _analytics_output_context("forecasting")
    context["form_state"] = {"target_year": target_year}
    if request.method == "POST":
        try:
            out_path = _run_forecasting_from_ui(target_year=target_year)
            context = _analytics_output_context_for_path("forecasting", out_path)
            context["form_state"] = {"target_year": target_year}
            context["run_notice"] = f"Forecasting completed successfully. Loaded fresh output: {out_path.name}"
            _append_run_history(
                "forecasting",
                {
                    "timestamp": datetime.now().strftime("%Y-%m-%d %H:%M"),
                    "type": "forecasting",
                    "scenario": "Forecasting",
                    "parameters_summary": f"Target year: {target_year}",
                    "target_year": target_year,
                    "output_file": out_path.name,
                },
            )
            _create_user_notification(
                current_user.id,
                title="Forecasting completed",
                message=f"Forecasting output {out_path.name} is ready.",
                notification_type="success",
                link=url_for("analytics_forecasting"),
            )
            context["run_history"] = _read_run_history("forecasting")
        except Exception as exc:
            context["run_error"] = f"Forecasting run failed: {exc}"
    return render_template("analytics_output.html", user=current_user, **context)


@app.route("/data-output/travel", methods=["GET"])
@login_required
def data_output_travel():
    output_path = STAGE2_TRAVEL_DIR / "analysis_summary.xlsx"
    sheet_data: dict[str, dict[str, object]] = {}
    sheet_names: list[str] = []
    error_message = ""
    updated_at = ""

    if output_path.exists():
        try:
            xls = pd.ExcelFile(output_path, engine="openpyxl")
            sheet_names = list(xls.sheet_names)
            for sheet_name in sheet_names:
                df = pd.read_excel(xls, sheet_name=sheet_name)
                df = df.where(pd.notna(df), "")
                sheet_data[sheet_name] = {
                    "columns": df.columns.tolist(),
                    "rows": df.to_dict(orient="records"),
                    "row_count": int(len(df)),
                }
            try:
                updated_at = datetime.utcfromtimestamp(output_path.stat().st_mtime).strftime("%Y-%m-%d %H:%M UTC")
            except Exception:
                updated_at = ""
        except Exception as exc:
            error_message = f"Could not read Travel MGMT output: {exc}"

    return render_template(
        "data_output_travel.html",
        user=current_user,
        output_exists=output_path.exists() and not error_message,
        sheet_names=sheet_names,
        sheet_data=sheet_data,
        row_count=sum(int(sheet.get("row_count") or 0) for sheet in sheet_data.values()),
        updated_at=updated_at,
        error_message=error_message,
        download_url=url_for("data_output_travel_download") if output_path.exists() else "",
    )


@app.route("/data-output/travel/download", methods=["GET"])
@login_required
def data_output_travel_download():
    output_path = STAGE2_TRAVEL_DIR / "analysis_summary.xlsx"
    if not output_path.exists():
        flash("Travel MGMT output file not found.")
        return redirect(url_for("data_output_travel"))
    return send_file(str(output_path), as_attachment=True, download_name="travel_mgmt_analysis_summary.xlsx")


@app.route("/data-output/klarakarbon", methods=["GET"])
@login_required
def data_output_klarakarbon():
    output_path = STAGE2_KLARAKARBON_DIR / "klarakarbon_summary.xlsx"
    sheet_data: dict[str, dict[str, object]] = {}
    sheet_names: list[str] = []
    error_message = ""
    updated_at = ""

    if output_path.exists():
        try:
            xls = pd.ExcelFile(output_path, engine="openpyxl")
            sheet_names = list(xls.sheet_names)
            for sheet_name in sheet_names:
                df = pd.read_excel(xls, sheet_name=sheet_name)
                df = df.where(pd.notna(df), "")
                sheet_data[sheet_name] = {
                    "columns": df.columns.tolist(),
                    "rows": df.to_dict(orient="records"),
                    "row_count": int(len(df)),
                }
            try:
                updated_at = datetime.utcfromtimestamp(output_path.stat().st_mtime).strftime("%Y-%m-%d %H:%M UTC")
            except Exception:
                updated_at = ""
        except Exception as exc:
            error_message = f"Could not read Klarakarbon output: {exc}"

    return render_template(
        "data_output_klarakarbon.html",
        user=current_user,
        output_exists=output_path.exists() and not error_message,
        sheet_names=sheet_names,
        sheet_data=sheet_data,
        row_count=sum(int(sheet.get("row_count") or 0) for sheet in sheet_data.values()),
        updated_at=updated_at,
        error_message=error_message,
        download_url=url_for("data_output_klarakarbon_download") if output_path.exists() else "",
    )


@app.route("/data-output/klarakarbon/download", methods=["GET"])
@login_required
def data_output_klarakarbon_download():
    output_path = STAGE2_KLARAKARBON_DIR / "klarakarbon_summary.xlsx"
    if not output_path.exists():
        flash("Klarakarbon output file not found.")
        return redirect(url_for("data_output_klarakarbon"))
    return send_file(str(output_path), as_attachment=True, download_name="klarakarbon_summary.xlsx")


@app.route("/analytics/mapped-window-output", methods=["GET", "POST"])
@login_required
def analytics_mapped_window_output():
    default_year = max(2025, datetime.now().year)
    form_state = {
        "year": default_year,
        "start_month": 1,
        "end_month": 12,
    }
    context = _analytics_output_context("mapped_window_output")
    if request.method == "POST":
        form_state = {
            "year": _parse_int_form("year", default_year, minimum=2020, maximum=2100),
            "start_month": _parse_int_form("start_month", 1, minimum=1, maximum=12),
            "end_month": _parse_int_form("end_month", 12, minimum=1, maximum=12),
        }
        form_state["end_month"] = max(int(form_state["start_month"]), int(form_state["end_month"]))
        try:
            out_path = _run_mapped_window_from_ui(
                year=int(form_state["year"]),
                start_month=int(form_state["start_month"]),
                end_month=int(form_state["end_month"]),
            )
            context = _analytics_output_context_for_path("mapped_window_output", out_path)
            context["run_notice"] = f"Mapped window output generated successfully: {out_path.name}"
            _append_run_history(
                "mapped_window_output",
                {
                    "timestamp": datetime.now().strftime("%Y-%m-%d %H:%M"),
                    "type": "mapped_window",
                    "scenario": "Mapped window output",
                    "parameters_summary": (
                        f"Year: {form_state['year']}, "
                        f"Period: {int(form_state['start_month']):02d}-{int(form_state['end_month']):02d}"
                    ),
                    "year": int(form_state["year"]),
                    "start_month": int(form_state["start_month"]),
                    "end_month": int(form_state["end_month"]),
                    "output_file": out_path.name,
                },
            )
            context["run_history"] = _read_run_history("mapped_window_output")
        except Exception as exc:
            context["run_error"] = f"Mapped window generation failed: {exc}"
    context["form_state"] = form_state
    return render_template("analytics_output.html", user=current_user, **context)


@app.route("/analytics/decarbonization", methods=["GET", "POST"])
@login_required
def analytics_decarbonization():
    context = _analytics_output_context("decarbonization")
    form_state = {
        "preset": "custom",
        "target_year": 2030,
        "reduction_pct": 20,
        "scope": "scope3",
        "scenario_type": "decarbonization",
    }
    if request.method == "POST":
        form_state = {
            "preset": str(request.form.get("scenario_preset", "custom") or "custom"),
            "target_year": _parse_int_form("target_year", 2030, minimum=2026, maximum=2050),
            "reduction_pct": _parse_float_form("reduction_pct", 20.0, minimum=0.0, maximum=100.0),
            "scope": str(request.form.get("scope", "scope3") or "scope3"),
            "scenario_type": str(request.form.get("scenario_type", "decarbonization") or "decarbonization"),
        }
        try:
            scenario_output, companion_output = _run_decarbonization_from_ui(
                target_year=int(form_state["target_year"]),
                reduction_pct=float(form_state["reduction_pct"]) / 100.0,
                scope_key=str(form_state["scope"]),
                scenario_type=str(form_state["scenario_type"]),
            )
            context = _analytics_output_context_for_path("decarbonization", companion_output)
            context["run_notice"] = (
                "Decarbonization run completed successfully. "
                f"Loaded fresh scenario output: {companion_output.name}"
            )
            context["companion_outputs"] = [
                _build_companion_output_payload(
                    "Decarbonization lever workbook",
                    scenario_output,
                    ("Scenarios_Yearly", "Scenarios_Monthly", "S4_Delta", "Meta"),
                )
            ]
            _append_run_history(
                "decarbonization",
                {
                    "timestamp": datetime.now().strftime("%Y-%m-%d %H:%M"),
                    "type": "decarbonization",
                    "scenario": str(form_state["preset"] or "custom"),
                    "parameters_summary": (
                        f"Target year: {form_state['target_year']}, "
                        f"Reduction: {form_state['reduction_pct']}%, "
                        f"Scope: {_scope_label(str(form_state['scope']))}, "
                        f"Type: {form_state['scenario_type']}"
                    ),
                    "target_year": int(form_state["target_year"]),
                    "reduction_pct": float(form_state["reduction_pct"]),
                    "scope": _scope_label(str(form_state["scope"])),
                    "scenario_type": str(form_state["scenario_type"]),
                    "output_file": companion_output.name,
                    "companion_output_file": scenario_output.name,
                },
            )
            _create_user_notification(
                current_user.id,
                title="Decarbonization completed",
                message=f"Scenario output {companion_output.name} is ready for review.",
                notification_type="success",
                link=url_for("analytics_decarbonization"),
            )
            context["run_history"] = _read_run_history("decarbonization")
        except Exception as exc:
            context["run_error"] = f"Decarbonization run failed: {exc}"
    context["form_state"] = form_state
    return render_template("analytics_output.html", user=current_user, **context)


@app.route("/analytics/emissions-totals", methods=["GET", "POST"])
@login_required
def analytics_emissions_totals():
    default_year = max(2025, datetime.now().year)
    form_state = {
        "year": default_year,
        "start_month": 1,
        "end_month": 12,
    }
    context = _analytics_output_context("emissions_totals")
    if request.method == "POST":
        form_state = {
            "year": _parse_int_form("year", default_year, minimum=2020, maximum=2100),
            "start_month": _parse_int_form("start_month", 1, minimum=1, maximum=12),
            "end_month": _parse_int_form("end_month", 12, minimum=1, maximum=12),
        }
        form_state["end_month"] = max(int(form_state["start_month"]), int(form_state["end_month"]))
        try:
            out_path = _run_totals_tables_from_ui(
                year=int(form_state["year"]),
                start_month=int(form_state["start_month"]),
                end_month=int(form_state["end_month"]),
            )
            context = _analytics_output_context_for_path("emissions_totals", out_path)
            context["run_notice"] = f"Totals tables generated successfully: {out_path.name}"
            _append_run_history(
                "emissions_totals",
                {
                    "timestamp": datetime.now().strftime("%Y-%m-%d %H:%M"),
                    "type": "totals_tables",
                    "scenario": "Totals tables",
                    "parameters_summary": (
                        f"Year: {form_state['year']}, "
                        f"Period: {int(form_state['start_month']):02d}-{int(form_state['end_month']):02d}"
                    ),
                    "year": int(form_state["year"]),
                    "start_month": int(form_state["start_month"]),
                    "end_month": int(form_state["end_month"]),
                    "output_file": out_path.name,
                },
            )
            context["run_history"] = _read_run_history("emissions_totals")
        except Exception as exc:
            context["run_error"] = f"Totals generation failed: {exc}"
    context["form_state"] = form_state
    return render_template("analytics_output.html", user=current_user, **context)


@app.route("/analytics/share-analysis", methods=["GET", "POST"])
@login_required
def analytics_share_analysis():
    default_year = max(2025, datetime.now().year)
    form_state = {
        "year": default_year,
        "start_month": 1,
        "end_month": 12,
    }
    context = _analytics_output_context("share_analysis")
    if request.method == "POST":
        form_state = {
            "year": _parse_int_form("year", default_year, minimum=2020, maximum=2100),
            "start_month": _parse_int_form("start_month", 1, minimum=1, maximum=12),
            "end_month": _parse_int_form("end_month", 12, minimum=1, maximum=12),
        }
        form_state["end_month"] = max(int(form_state["start_month"]), int(form_state["end_month"]))
        try:
            out_path = _run_totals_tables_from_ui(
                year=int(form_state["year"]),
                start_month=int(form_state["start_month"]),
                end_month=int(form_state["end_month"]),
            )
            context = _analytics_output_context_for_path("share_analysis", out_path)
            context["run_notice"] = f"Share analysis generated successfully: {out_path.name}"
            _append_run_history(
                "share_analysis",
                {
                    "timestamp": datetime.now().strftime("%Y-%m-%d %H:%M"),
                    "type": "share_analysis",
                    "scenario": "Share analysis",
                    "parameters_summary": (
                        f"Year: {form_state['year']}, "
                        f"Period: {int(form_state['start_month']):02d}-{int(form_state['end_month']):02d}"
                    ),
                    "year": int(form_state["year"]),
                    "start_month": int(form_state["start_month"]),
                    "end_month": int(form_state["end_month"]),
                    "output_file": out_path.name,
                },
            )
            context["run_history"] = _read_run_history("share_analysis")
        except Exception as exc:
            context["run_error"] = f"Share analysis generation failed: {exc}"
    context["form_state"] = form_state
    return render_template("analytics_output.html", user=current_user, **context)


@app.route("/analytics/emissions-map", methods=["GET"])
@login_required
def analytics_emissions_map():
    return render_template("emissions_map.html", user=current_user, **_load_emissions_map_points())


@app.route("/governance/double-counting-check", methods=["GET", "POST"])
@login_required
def governance_double_counting_check():
    context = _analytics_output_context("double_counting_check")
    if request.method == "POST":
        try:
            out_path = _run_double_counting_from_ui()
            context = _analytics_output_context_for_path("double_counting_check", out_path)
            context["run_notice"] = f"Double counting report generated successfully: {out_path.name}"
            _append_run_history(
                "double_counting_check",
                {
                    "timestamp": datetime.now().strftime("%Y-%m-%d %H:%M"),
                    "type": "double_counting",
                    "scenario": "Double counting report",
                    "parameters_summary": "Generated from the latest merged mapping workbook.",
                    "output_file": out_path.name,
                },
            )
            context["run_history"] = _read_run_history("double_counting_check")
        except Exception as exc:
            context["run_error"] = f"Double counting generation failed: {exc}"
    return render_template("analytics_output.html", user=current_user, **context)


@app.route("/governance/audit-ready-output", methods=["GET", "POST"])
@login_required
def governance_audit_ready_output():
    default_year = max(2025, datetime.now().year)
    form_state = {
        "year": default_year,
        "start_month": 1,
        "end_month": 12,
    }
    context = _analytics_output_context("audit_ready_output")
    if request.method == "POST":
        form_state = {
            "year": _parse_int_form("year", default_year, minimum=2020, maximum=2100),
            "start_month": _parse_int_form("start_month", 1, minimum=1, maximum=12),
            "end_month": _parse_int_form("end_month", 12, minimum=1, maximum=12),
        }
        form_state["end_month"] = max(int(form_state["start_month"]), int(form_state["end_month"]))
        try:
            out_path = _run_mapped_window_from_ui(
                year=int(form_state["year"]),
                start_month=int(form_state["start_month"]),
                end_month=int(form_state["end_month"]),
            )
            context = _analytics_output_context_for_path("audit_ready_output", out_path)
            context["run_notice"] = f"Audit-ready dataset generated successfully: {out_path.name}"
            _append_run_history(
                "audit_ready_output",
                {
                    "timestamp": datetime.now().strftime("%Y-%m-%d %H:%M"),
                    "type": "audit_output",
                    "scenario": "Audit ready dataset",
                    "parameters_summary": (
                        f"Year: {form_state['year']}, "
                        f"Period: {int(form_state['start_month']):02d}-{int(form_state['end_month']):02d}"
                    ),
                    "year": int(form_state["year"]),
                    "start_month": int(form_state["start_month"]),
                    "end_month": int(form_state["end_month"]),
                    "output_file": out_path.name,
                },
            )
            _create_user_notification(
                current_user.id,
                title="Audit dataset ready",
                message=f"Audit-ready dataset {out_path.name} has been generated.",
                notification_type="success",
                link=url_for("governance_audit_ready_output"),
            )
            context["run_history"] = _read_run_history("audit_ready_output")
        except Exception as exc:
            context["run_error"] = f"Audit dataset generation failed: {exc}"
    context["form_state"] = form_state
    return render_template("analytics_output.html", user=current_user, **context)


@app.route('/admin', methods=['GET', 'POST'])
@login_required
def admin():
    if not current_user.is_admin:
        flash('Access denied')
        return redirect(url_for('dashboard'))

    users = User.query.all()
    mapping_runs = MappingRun.query.order_by(MappingRun.created_at.desc()).all()

    # Find all months with mapping runs
    all_months = set()
    for run in mapping_runs:
        if run.created_at:
            all_months.add(run.created_at.strftime('%B %Y'))
    if not all_months:
        all_months.add(datetime.now().strftime('%B %Y'))
    available_months = sorted(list(all_months), key=lambda x: datetime.strptime(x, '%B %Y'), reverse=True)
    selected_month = request.args.get('month') or available_months[0]

    # Calculate progress per company based on mapped categories vs available schema sheets
    companies = list(COMPANIES)
    submission_stats = []
    chart_data = []
    for company in companies:
        expected = _count_company_schema_sheets(company)
        submitted_sheets = set()
        for run in mapping_runs:
            if (run.company_name == company and run.created_at and run.created_at.strftime('%B %Y') == selected_month and run.status == 'succeeded'):
                submitted_sheets.add(run.sheet_name)
        submitted = len(submitted_sheets)
        rate = int(round((submitted / expected) * 100)) if expected > 0 else 0
        submission_stats.append({
            'company': company,
            'month': selected_month,
            'submitted': submitted,
            'expected': expected,
            'rate': rate
        })
        chart_data.append({
            'company': company,
            'submitted': submitted,
            'expected': expected,
            'rate': rate
        })

    # Last 7 days mapping count
    seven_days_ago = datetime.now() - timedelta(days=7)
    recent_admin_submissions_count = MappingRun.query.filter(
        MappingRun.created_at >= seven_days_ago
    ).count()

    users_for_admin = [_user_public_dict_for_admin(u) for u in users]
    is_owner = normalize_user_role(getattr(current_user, "role", None)) == "owner"
    return render_template(
        'admin.html',
        users=users,
        users_for_admin=users_for_admin,
        is_owner=is_owner,
        user_roles=list(USER_ROLES),
        submissions=[],
        mapping_runs=mapping_runs,
        recent_admin_submissions_count=recent_admin_submissions_count,
        available_months=available_months,
        selected_month=selected_month,
        submission_stats=submission_stats,
        chart_data=chart_data,
        logos=[],
    )


@app.route("/admin/user/<int:user_id>/profile", methods=["POST"])
@login_required
def admin_update_user_profile(user_id):
    """Owner-only: update another user's profile fields (existing columns only)."""
    if not current_user.is_admin:
        flash("Access denied")
        return redirect(url_for("dashboard"))
    if normalize_user_role(getattr(current_user, "role", None)) != "owner":
        flash("Only Owner can modify user details.")
        return redirect(url_for("admin"))
    _ensure_db_tables()
    target = db.session.get(User, user_id)
    if not target:
        flash("User not found.")
        return redirect(url_for("admin"))

    target.first_name = (request.form.get("first_name") or "").strip() or None
    target.last_name = (request.form.get("last_name") or "").strip() or None
    target.phone = (request.form.get("phone") or "").strip() or None
    cn = (request.form.get("company_name") or "").strip()
    if not cn:
        flash("Company name is required.")
        return redirect(url_for("admin"))
    target.company_name = cn
    target.company_country = (request.form.get("company_country") or "").strip() or None

    new_role = request.form.get("role")
    if new_role is not None and str(new_role).strip():
        new_role = normalize_user_role(new_role)
        prev = normalize_user_role(getattr(target, "role", None))
        if prev == "owner" and new_role != "owner":
            owners = User.query.filter(User.role == "owner").count()
            if owners <= 1:
                flash("Cannot remove the only owner account.")
                return redirect(url_for("admin"))
        target.role = new_role
        sync_user_admin_flag(target)

    pfile = request.files.get("profile_photo")
    if pfile and getattr(pfile, "filename", None):
        rel = _save_profile_photo_file(pfile, user_id)
        if rel:
            target.profile_photo_path = rel

    db.session.commit()
    flash(f"User {target.email} updated.")
    return redirect(url_for("admin"))


def _admin_user_company_options() -> list[str]:
    return list(COMPANIES)


def _render_admin_users_page(
    *,
    create_errors: dict[str, str] | None = None,
    create_form: dict[str, str] | None = None,
):
    users = User.query.order_by(User.email.asc()).all()
    return render_template(
        "admin_users.html",
        users=users,
        user_roles=USER_ROLES,
        companies=_admin_user_company_options(),
        create_errors=create_errors or {},
        create_form=create_form or {},
    )


@app.route("/admin/create-user", methods=["POST"])
@login_required
def admin_create_user():
    _ensure_db_tables()
    if normalize_user_role(getattr(current_user, "role", None)) != "owner":
        flash("Access denied")
        return redirect(url_for("dashboard"))

    email = (request.form.get("email") or "").strip().lower()
    password = request.form.get("password") or ""
    full_name = (request.form.get("full_name") or "").strip()
    title = (request.form.get("title") or "").strip()
    company = (_resolve_template_company_name(request.form.get("company") or "") or (request.form.get("company") or "").strip())
    role = normalize_user_role(request.form.get("role"))

    form_state = {
        "email": email,
        "full_name": full_name,
        "title": title,
        "company": company,
        "role": role,
    }
    errors: dict[str, str] = {}

    if not email:
        errors["email"] = "Email is required."
    elif "@" not in email:
        errors["email"] = "Enter a valid email address."
    elif User.query.filter(db.func.lower(User.email) == email).first():
        errors["email"] = "A user with this email already exists."

    if not password:
        errors["password"] = "Password is required."

    if company not in _admin_user_company_options():
        errors["company"] = "Select a valid company."

    if role not in USER_ROLES_SET:
        errors["role"] = "Select a valid role."

    if errors:
        return _render_admin_users_page(create_errors=errors, create_form=form_state)

    first_name, last_name = _split_full_name(full_name)
    user = User(
        email=email,
        password_hash=generate_password_hash(password),
        company_name=company,
        first_name=first_name or None,
        last_name=last_name or None,
        job_title=title or None,
        role=role,
    )
    sync_user_admin_flag(user)
    db.session.add(user)
    db.session.commit()

    flash("User created successfully")
    return redirect(url_for("admin_users"))


@app.route("/admin/users", methods=["GET", "POST"])
@login_required
def admin_users():
    """Owner-only: list users and change roles."""
    _ensure_db_tables()
    if normalize_user_role(getattr(current_user, "role", None)) != "owner":
        flash("Access denied")
        return redirect(url_for("dashboard"))

    if request.method == "POST":
        uid = request.form.get("user_id", type=int)
        new_role = normalize_user_role(request.form.get("role"))
        if not uid:
            flash("Invalid request.")
            return redirect(url_for("admin_users"))
        target = db.session.get(User, uid)
        if not target:
            flash("User not found.")
            return redirect(url_for("admin_users"))

        prev = normalize_user_role(getattr(target, "role", None))
        if prev == "owner" and new_role != "owner":
            owners = User.query.filter(User.role == "owner").count()
            if owners <= 1:
                flash("Cannot remove the only owner account.")
                return redirect(url_for("admin_users"))

        target.role = new_role
        sync_user_admin_flag(target)
        db.session.commit()
        flash(f"Role updated for {target.email}.")
        return redirect(url_for("admin_users"))

    return _render_admin_users_page()


@app.route("/owner-analytics", methods=["GET"])
@login_required
def owner_analytics():
    if not bool(getattr(current_user, "is_admin", False)):
        abort(403)
    _ensure_db_tables()
    return render_template("owner_analytics.html", user=current_user, **_owner_analytics_context())


def _run_travel_preprocess_job(*, job_id: str, run_dir: str, raw_path: str) -> dict[str, object]:
    def progress(progress_value: int, message: str) -> None:
        _update_job_progress(job_id, progress_value, message)

    _update_job_progress(job_id, 5, "Starting Travel preprocessing")
    run_travel_preprocess(Path(run_dir), Path(raw_path), progress_callback=progress)
    analysis_summary = STAGE2_TRAVEL_DIR / "analysis_summary.xlsx"
    row_count = 0
    if analysis_summary.exists():
        final_df = pd.read_excel(analysis_summary, sheet_name=0, engine="openpyxl")
        row_count = int(len(final_df))
    _update_job(job_id, rows=row_count)
    return {
        "ok": True,
        "run_dir": run_dir,
        "analysis_summary": str(analysis_summary),
        "rows": row_count,
    }


@app.route("/start-travel-preprocess", methods=["POST"])
@app.route("/admin/preprocess/travel/upload", methods=["POST"])
@login_required
def upload_travel_preprocess():
    wants_json = request.path == "/start-travel-preprocess" or "application/json" in str(request.headers.get("Accept") or "")
    if not current_user.is_admin:
        if wants_json:
            return jsonify({"error": "Access denied"}), 403
        flash("Access denied")
        return redirect(url_for("dashboard"))

    upload = request.files.get("travel_file")
    if not upload or not getattr(upload, "filename", None):
        if wants_json:
            return jsonify({"error": "Please choose a Travel Management file (.xlsb or .xlsx)."}), 400
        flash("Please choose a Travel Management file (.xlsb or .xlsx).")
        return redirect(url_for("admin"))

    ext = Path(secure_filename(upload.filename or "")).suffix.lower()
    if ext not in TRAVEL_ALLOWED_EXT:
        if wants_json:
            return jsonify({"error": "Only .xlsb or .xlsx files are allowed"}), 400
        flash("Only .xlsb or .xlsx files are allowed")
        return redirect(url_for("admin"))

    run_id = f"{datetime.now().strftime('%Y%m%d_%H%M%S')}_{uuid.uuid4().hex[:10]}"
    run_dir = FRONTEND_UPLOAD_DIR / "preprocess" / "travel" / run_id
    raw_dir = run_dir / "raw"
    raw_dir.mkdir(parents=True, exist_ok=True)

    safe_name = secure_filename(upload.filename or "travel_source.xlsx") or "travel_source.xlsx"
    raw_path = raw_dir / safe_name
    upload.save(str(raw_path))

    validation_errors = validate_travel_upload(raw_path)
    validation_path = run_dir / "validation.json"
    if validation_errors:
        validation_path.write_text(
            json.dumps({"status": "failed", "errors": validation_errors}, indent=2),
            encoding="utf-8",
        )
        if wants_json:
            return jsonify({"error": validation_errors[0], "errors": validation_errors}), 400
        flash(validation_errors[0])
        return redirect(url_for("admin"))

    validation_path.write_text(
        json.dumps({"status": "passed", "file": raw_path.name}, indent=2),
        encoding="utf-8",
    )

    job_id = run_in_background(
        "preprocess",
        "Travel",
        _run_travel_preprocess_job,
        run_dir=str(run_dir),
        raw_path=str(raw_path),
        job_user_id=int(current_user.id),
        job_user_email=str(getattr(current_user, "email", "") or ""),
    )
    if wants_json:
        return jsonify({"job_id": job_id, "status": "started"})
    flash(f"Travel preprocessing started. Job ID: {job_id}")
    return redirect(url_for("admin"))


@app.route('/logout')
@login_required
def logout():
    request.environ["skip_activity_log"] = True
    _write_activity_log_for_user(current_user, action="logout")
    session.pop("activity_session_id", None)
    logout_user()
    return redirect(url_for('index'))

@app.route('/report', methods=['GET'])
@login_required
def report():
    if current_user.is_admin:
        return redirect(url_for('admin_report', **request.args.to_dict(flat=True)))
    template_filter = request.args.get('template', '').strip().lower()
    month_filter = request.args.get('date', '').strip()
    company_keys = _company_candidate_keys(current_user.company_name)
    summaries = (
        MappingRunSummary.query.filter(MappingRunSummary.company_name.in_(company_keys))
        .order_by(MappingRunSummary.created_at.desc())
        .all()
    )
    filtered_summaries = []
    seen = set()
    run_cache: dict[str, MappingRun | None] = {}
    for sub in summaries:
        key = ((sub.company_name or "").strip().lower(), (sub.sheet_name or "").strip().lower())
        if key in seen:
            continue
        seen.add(key)
        match = True
        if template_filter and template_filter not in (sub.sheet_name or "").lower():
            match = False
        profile = _period_profile_for_summary(sub, run_cache=run_cache)
        points = list(profile.get("points") or [])
        if month_filter:
            points = [p for p in points if str(p.get("key") or "").startswith(month_filter)]
            if not points:
                match = False
        if match:
            filtered_summaries.append((sub, points if month_filter else points))
    emission_results = []
    chart_data = []
    for idx, (sub, points) in enumerate(filtered_summaries):
        filtered_points = list(points or [])
        if not filtered_points:
            profile = _period_profile_for_summary(sub, run_cache=run_cache)
            filtered_points = list(profile.get("points") or [])
        total_t = _safe_float(sum(float(p.get("value") or 0.0) for p in filtered_points) if filtered_points else (sub.tco2e_total or 0.0))
        label = str(sub.sheet_name or "Category")
        emission_results.append({
            'template_name': label,
            'period_label': _format_period_label(filtered_points, getattr(sub, "created_at", None)),
            'total_emission': round(total_t, 6),
            'by_category': {label: total_t}
        })
        chart_data.append({
            'bar_id': f'barChart{idx+1}',
            'line_id': f'lineChart{idx+1}',
            'pie_id': f'pieChart{idx+1}',
            'labels': [str(p.get("label") or label) for p in filtered_points] if filtered_points else [label],
            'values': [_safe_float(p.get("value") or 0.0) for p in filtered_points] if filtered_points else [total_t]
        })
    return render_template('report.html', emission_results=emission_results, user=current_user, chart_data=chart_data)

@app.route('/admin_report', methods=['GET'])
@login_required
def admin_report():
    if not current_user.is_admin:
        flash('Access denied')
        return redirect(url_for('dashboard'))
    company_filter = request.args.get('company', '').strip().lower()
    template_filter = request.args.get('template', '').strip().lower()
    month_filter = request.args.get('date', '').strip()
    summaries = MappingRunSummary.query.order_by(MappingRunSummary.created_at.desc()).all()
    filtered_submissions = []
    seen = set()
    run_cache: dict[str, MappingRun | None] = {}
    for sub in summaries:
        key = ((sub.company_name or "").strip().lower(), (sub.sheet_name or "").strip().lower())
        if key in seen:
            continue
        seen.add(key)
        match = True
        if company_filter and company_filter not in (sub.company_name or "").lower():
            match = False
        if template_filter and template_filter not in (sub.sheet_name or "").lower():
            match = False
        profile = _period_profile_for_summary(sub, run_cache=run_cache)
        points = list(profile.get("points") or [])
        if month_filter:
            points = [p for p in points if str(p.get("key") or "").startswith(month_filter)]
            if not points:
                match = False
        if match:
            filtered_submissions.append((sub, points if month_filter else points))
    emission_results = []
    chart_data = []
    for idx, (sub, points) in enumerate(filtered_submissions):
        filtered_points = list(points or [])
        if not filtered_points:
            profile = _period_profile_for_summary(sub, run_cache=run_cache)
            filtered_points = list(profile.get("points") or [])
        total_t = _safe_float(sum(float(p.get("value") or 0.0) for p in filtered_points) if filtered_points else (sub.tco2e_total or 0.0))
        label = str(sub.sheet_name or "Category")
        emission_results.append({
            'company_name': sub.company_name,
            'template_name': label,
            'period_label': _format_period_label(filtered_points, getattr(sub, "created_at", None)),
            'total_emission': round(total_t, 6),
            'by_category': {label: total_t}
        })
        chart_data.append({
            'bar_id': f'barChart{idx+1}',
            'line_id': f'lineChart{idx+1}',
            'pie_id': f'pieChart{idx+1}',
            'labels': [str(p.get("label") or label) for p in filtered_points] if filtered_points else [label],
            'values': [_safe_float(p.get("value") or 0.0) for p in filtered_points] if filtered_points else [total_t]
        })
    return render_template('admin_report.html', emission_results=emission_results, user=current_user, chart_data=chart_data)

@app.route('/admin/emission_factors', methods=['GET', 'POST'])
@login_required
def manage_emission_factors():
    if not current_user.is_admin:
        flash('Access denied')
        return redirect(url_for('dashboard'))

    if request.method == 'POST':
        flash('Emission factors are now sourced from the mapping workbook and are read-only in the UI.')
        return redirect(url_for('manage_emission_factors'))

    data = _load_stage2_emission_factors()
    all_rows: list[dict[str, object]] = list(data.get("rows") or [])
    sheets: list[str] = list(data.get("sheets") or [])
    scopes: list[str] = list(data.get("scopes") or [])
    sources: list[str] = list(data.get("sources") or [])

    # Filters
    search = request.args.get('search', '').strip()
    sheet_filter = request.args.get('sheet', '').strip()
    scope_filter = request.args.get('scope', '').strip()
    source_filter = request.args.get('ef_source', '').strip()

    def _matches(row: dict[str, object]) -> bool:
        if sheet_filter and str(row.get("sheet") or "") != sheet_filter:
            return False
        if scope_filter and str(row.get("scope") or "") != scope_filter:
            return False
        if source_filter and str(row.get("ef_source") or "") != source_filter:
            return False

        if search:
            hay = " ".join(
                str(row.get(k) or "")
                for k in (
                    "sheet",
                    "ef_name",
                    "ef_description",
                    "scope",
                    "ef_category",
                    "ef_id",
                    "ef_value",
                    "ef_unit",
                    "ef_source",
                    "Emission Factor Category",
                )
            ).lower()
            if search.lower() not in hay:
                return False
        return True

    filtered = [r for r in all_rows if _matches(r)]
    filtered.sort(key=lambda r: (str(r.get("sheet") or ""), str(r.get("ef_name") or ""), str(r.get("ef_id") or "")))

    # Pagination (lightweight object compatible with template)
    page = int(request.args.get('page', 1) or 1)
    per_page = 60
    total = len(filtered)
    pages = max(1, (total + per_page - 1) // per_page)
    page = max(1, min(page, pages))
    start = (page - 1) * per_page
    end = start + per_page
    items = filtered[start:end]

    factors = SimpleNamespace(
        items=items,
        page=page,
        pages=pages,
        has_prev=(page > 1),
        has_next=(page < pages),
        prev_num=(page - 1),
        next_num=(page + 1),
        total=total,
    )

    if request.method == 'POST':
        # handled above (read-only)
        pass

    return render_template(
        'emission_factors.html',
        factors=factors,
        search=search,
        sheet=sheet_filter,
        scope=scope_filter,
        ef_source=source_filter,
        available_sheets=sheets,
        available_scopes=scopes,
        available_sources=sources,
        mapping_path=str(STAGE2_EF_XLSX),
    )

@app.route('/admin/emission_factors/delete/<int:factor_id>', methods=['POST'])
@login_required
def delete_emission_factor(factor_id):
    if not current_user.is_admin:
        flash('Access denied')
        return redirect(url_for('dashboard'))
    flash('Emission factors are sourced from the mapping workbook and cannot be deleted here.')
    return redirect(url_for('manage_emission_factors'))

@app.route('/admin/emission_factors/update/<int:factor_id>', methods=['POST'])
@login_required
def update_emission_factor(factor_id):
    if not current_user.is_admin:
        flash('Access denied')
        return redirect(url_for('dashboard'))
    flash('Emission factors are sourced from the mapping workbook and cannot be updated here.')
    return redirect(url_for('manage_emission_factors'))


def _ef_expected_headers() -> list[str]:
    return [
        "ef_name",
        "ef_description",
        "scope",
        "ef_category",
        "ef_id",
        "ef_value",
        "ef_unit",
        "ef_source",
        "Emission Factor Category",
    ]


def _clear_ef_cache() -> None:
    global _EF_CACHE
    _EF_CACHE = {"mtime_ns": None, "rows": None, "sheets": None, "scopes": None, "sources": None}


@app.route("/admin/emission_factors/mapping/reload", methods=["GET"])
@login_required
def reload_emission_factors_mapping():
    if not current_user.is_admin:
        flash("Access denied")
        return redirect(url_for("dashboard"))
    _clear_ef_cache()
    flash("Emission factors reloaded from mapping workbook.")
    # Preserve current filters if present
    return redirect(
        url_for(
            "manage_emission_factors",
            search=request.args.get("search", ""),
            sheet=request.args.get("sheet", ""),
            scope=request.args.get("scope", ""),
            ef_source=request.args.get("ef_source", ""),
            page=request.args.get("page", 1),
        )
    )


def _cleanup_mapping_runs(max_age_minutes: int = 180) -> None:
    now = datetime.now()
    to_del = []
    for run_id, meta in list(_MAPPING_RUNS.items()):
        created = meta.get("created_at")
        if isinstance(created, datetime) and (now - created).total_seconds() > max_age_minutes * 60:
            to_del.append(run_id)
    for rid in to_del:
        try:
            p = _MAPPING_RUNS.get(rid, {}).get("path")
            if p and os.path.exists(str(p)):
                os.remove(str(p))
        except Exception:
            pass
        _MAPPING_RUNS.pop(rid, None)


def _to_numeric_spend(x: object) -> float | None:
    try:
        if x is None or (isinstance(x, float) and pd.isna(x)):  # type: ignore[arg-type]
            return None
        if isinstance(x, (int, float)) and not (isinstance(x, float) and pd.isna(x)):
            return float(x)

        s = str(x).strip().replace(" ", "")
        if s == "":
            return None
        if s.startswith("(") and s.endswith(")"):
            s = "-" + s[1:-1]
        # handle comma/period formats
        if s.count(",") == 1 and s.count(".") > 0 and s.rfind(",") > s.rfind("."):
            s = s.replace(".", "").replace(",", ".")
        elif s.count(".") == 1 and s.count(",") > 0 and s.rfind(".") > s.rfind(","):
            s = s.replace(",", "")
        elif s.count(",") == 1 and s.count(".") == 0:
            s = s.replace(",", ".")
        else:
            s = s.replace(",", "")
        return float(s)
    except Exception:
        return None


_FX_2025_TO_EUR: dict[str, float] = {
    "USD": 0.884969,
    "JPY": 0.005916,
    "BGN": 0.511300,
    "CZK": 0.040506,
    "DKK": 0.133987,
    "GBP": 1.167144,
    "HUF": 0.002514,
    "PLN": 0.235868,
    "RON": 0.198319,
    "SEK": 0.090364,
    "CHF": 1.067201,
    "ISK": 0.006913,
    "NOK": 0.085344,
    "TRY": 0.022313,
    "AUD": 0.570854,
    "BRL": 0.158550,
    "CAD": 0.633422,
    "CNY": 0.123175,
    "HKD": 0.113502,
    "IDR": 0.000054,
    "ILS": 0.256889,
    "INR": 0.010150,
    "KRW": 0.000623,
    "MXN": 0.046146,
    "MYR": 0.206874,
    "NZD": 0.514888,
    "PHP": 0.015390,
    "SGD": 0.677705,
    "THB": 0.026943,
    "ZAR": 0.049557,
    "EUR": 1.0,
    "EURO": 1.0,
    "EUROS": 1.0,
    "€": 1.0,
}


def _normalize_currency(raw: object) -> str | None:
    if raw is None:
        return None
    try:
        if isinstance(raw, float) and pd.isna(raw):
            return None
    except Exception:
        pass
    s = str(raw).upper()
    if "€" in s:
        return "EUR"
    if "£" in s:
        return "GBP"
    m = re.search(r"\b(USD|JPY|BGN|CZK|DKK|GBP|HUF|PLN|RON|SEK|CHF|ISK|NOK|TRY|AUD|BRL|CAD|CNY|HKD|IDR|ILS|INR|KRW|MXN|MYR|NZD|PHP|SGD|THB|ZAR|EUR|EURO|EUROS|€)\b", s)
    return m.group(1) if m else None


def preprocess_for_mapping(company_name: str, sheet_name: str, df: pd.DataFrame) -> pd.DataFrame:
    """
    Web-friendly 'mini Stage1': keep inputs stable for Stage2 mapping without multi-company merge.
    Does NOT change mapping rules; only prepares a single-sheet dataframe.
    """
    df = df.copy()
    df.columns = [str(c).strip() for c in df.columns]

    # Drop fully empty rows
    df = df.dropna(how="all")

    # Minimal whitespace normalization
    for c in df.columns:
        try:
            if pd.api.types.is_object_dtype(df[c]) or pd.api.types.is_string_dtype(df[c]):
                df[c] = df[c].astype("string").str.strip()
        except Exception:
            pass

    canon_company, canon_country = _canonical_company_name_and_country(company_name)
    company_display = canon_company or company_name

    # Compatibility columns expected by some downstream steps
    if "Source_File" not in df.columns:
        df["Source_File"] = f"{company_display}.xlsx"
    if "subsidiary_name" not in df.columns:
        df["subsidiary_name"] = company_display

    # Country is required by several mapping rules (previously derived from "Company Information")
    # Ensure both "Country" and "country" exist for robust header matching.
    if "Country" not in df.columns:
        df["Country"] = ""
    if "country" not in df.columns:
        df["country"] = ""

    if canon_country:
        # Fill only where blank
        try:
            mask = df["Country"].isna() | (df["Country"].astype(str).str.strip() == "")
            df.loc[mask, "Country"] = canon_country
        except Exception:
            df["Country"] = canon_country
        try:
            mask2 = df["country"].isna() | (df["country"].astype(str).str.strip() == "")
            df.loc[mask2, "country"] = canon_country
        except Exception:
            df["country"] = canon_country

    # Add Spend_Euro if possible (mirrors Currency_converter_17Dec behavior at sheet-level)
    if "Spend_Euro" not in df.columns:
        sheet_key = str(sheet_name).strip().lower()

        def find_col_exact(name: str) -> str | None:
            for col in df.columns:
                if str(col).strip().lower() == name.strip().lower():
                    return str(col)
            return None

        # Spend/currency columns with special cases
        if sheet_key == "scope 3 cat 15 pensions":
            spend_col = next((c for c in df.columns if "employer payment to pension provider" in str(c).lower()), None)
            currency_col = find_col_exact("currency")
        elif sheet_key == "scope 3 cat 6 business travel":
            spend_col = find_col_exact("spend")
            currency_col = find_col_exact("spend currency")
        else:
            spend_col = next((c for c in df.columns if "spend" in str(c).lower() and "euro" not in str(c).lower()), None)
            currency_col = next((c for c in df.columns if "currency" in str(c).lower()), None)

        spend_euro_vals = []
        if spend_col and currency_col:
            for _, row in df.iterrows():
                spend = _to_numeric_spend(row.get(spend_col))
                curr_code = _normalize_currency(row.get(currency_col))
                if spend is None or curr_code is None:
                    spend_euro_vals.append(pd.NA)
                    continue
                rate = _FX_2025_TO_EUR.get(curr_code)
                spend_euro_vals.append(spend * rate if rate else pd.NA)
        else:
            spend_euro_vals = [pd.NA] * len(df)
        df["Spend_Euro"] = spend_euro_vals

    return df


def _import_stage2_main_mapping():
    # Lazy import to keep Flask startup fast
    if str(STAGE2_MAPPING_DIR) not in sys.path:
        sys.path.insert(0, str(STAGE2_MAPPING_DIR))
    import importlib
    return importlib.import_module("main_mapping")


def _import_stage2_append_sources():
    if str(STAGE2_MAPPING_DIR) not in sys.path:
        sys.path.insert(0, str(STAGE2_MAPPING_DIR))
    import importlib
    return importlib.import_module("append_sources_to_mapped")


def _publish_klarakarbon_data_entry_output(company_name: str) -> Path:
    headers, _rules = _get_data_entry_template_schema(company_name, KLARAKARBON_SHEET_NAME)
    if not headers:
        raise RuntimeError("Klarakarbon headers are not configured for this company.")
    df = _load_data_entries_dataframe(company_name, KLARAKARBON_SHEET_NAME, headers)
    if df.empty:
        raise RuntimeError("No saved Klarakarbon rows found for this company.")
    publish_dir = STAGE1_KLARAKARBON_OUTPUT_DIR / company_slug(company_name)
    publish_dir.mkdir(parents=True, exist_ok=True)
    publish_path = publish_dir / "klarakarbon_categories_mapped_FINAL.xlsx"
    with pd.ExcelWriter(publish_path, engine="openpyxl") as writer:
        df.to_excel(writer, sheet_name=KLARAKARBON_SHEET_NAME[:31], index=False)
    return publish_path


def _run_append_and_pipeline(company_name: str, sheet_name: str) -> dict[str, object]:
    sheet_key = str(sheet_name or "").strip()
    if _batch_action_type_for_sheet(sheet_key) != "append_run":
        raise RuntimeError("This sheet does not support append-and-run orchestration.")

    published_path: Path | None = None
    if sheet_key == KLARAKARBON_SHEET_NAME:
        published_path = STAGE2_KLARAKARBON_DIR / "klarakarbon_summary.xlsx"
        if not published_path.exists():
            raise RuntimeError("Klarakarbon summary workbook was not found. Upload Klarakarbon data first.")
    elif sheet_key == TRAVEL_SHEET_NAME:
        published_path = STAGE2_TRAVEL_DIR / "analysis_summary.xlsx"
        if not published_path.exists():
            raise RuntimeError("Travel analysis_summary.xlsx was not found. Upload Travel data first.")

    append_src = _import_stage2_append_sources()
    with _STAGE2_MAP_LOCK:
        append_src.main(company_name=company_name)
        proc = subprocess.run(
            [sys.executable, str(STAGE2_MAPPING_DIR / "Run_Everything.py")],
            cwd=str(STAGE2_MAPPING_DIR),
            capture_output=True,
            text=True,
            encoding="utf-8",
            errors="replace",
            timeout=_SUBPROCESS_TIMEOUT_SECONDS,
        )
    if proc.returncode != 0:
        detail = (proc.stdout or proc.stderr or "").strip()
        if detail:
            detail = detail[-1200:]
        raise RuntimeError(detail or f"Run_Everything.py failed with exit code {proc.returncode}")

    return {
        "ok": True,
        "company": company_name,
        "sheet": sheet_key,
        "action_type": "append_run",
        "published_path": str(published_path) if published_path else "",
        "processed_at": datetime.utcnow().isoformat() + "Z",
    }


def run_mapping(
    company_name: str,
    sheet_name: str,
    df: pd.DataFrame,
    *,
    template_mode: str | None = None,
) -> tuple[pd.DataFrame, Path, Path]:
    """
    Run existing Stage2 mapping logic for a single company + single sheet dataframe.
    Mapping logic is NOT modified; we call main_mapping.process_all_sheets() on a temporary workbook.
    Returns (mapped_df, output_workbook_path).
    """
    _cleanup_mapping_runs()

    df_pre = preprocess_for_mapping(company_name, sheet_name, df)
    resolved_template_mode = normalize_template_mode(template_mode) if template_mode else _current_template_mode()
    internal_sheet_name = _stage2_sheet_name_for_run(sheet_name, resolved_template_mode)

    run_dir = INSTANCE_DIR / "mapping_runs" / f"{datetime.now().strftime('%Y%m%d_%H%M%S')}_{uuid.uuid4().hex[:10]}"
    run_dir.mkdir(parents=True, exist_ok=True)
    input_xlsx = run_dir / f"{secure_filename(company_name)}__{secure_filename(sheet_name)}.xlsx"

    with pd.ExcelWriter(input_xlsx, engine="openpyxl") as writer:
        df_pre.to_excel(writer, sheet_name=str(internal_sheet_name)[:31], index=False)

    mm = _import_stage2_main_mapping()
    append_src = _import_stage2_append_sources()

    # Run Stage2 mapping under a lock (Stage2 writes to a shared output/ directory)
    start_ts = time.time() if "time" in globals() else None
    import time as _time
    start_ts = _time.time()
    with _STAGE2_MAP_LOCK:
        orig = getattr(mm, "INPUT_WORKBOOK_NAME", None)
        orig_env = {
            "CTS_TEMPLATE_MODE": os.environ.get("CTS_TEMPLATE_MODE"),
            "CTS_STAGE2_SOURCE_SHEET": os.environ.get("CTS_STAGE2_SOURCE_SHEET"),
            "CTS_STAGE2_INTERNAL_SHEET": os.environ.get("CTS_STAGE2_INTERNAL_SHEET"),
        }
        try:
            setattr(mm, "INPUT_WORKBOOK_NAME", str(input_xlsx))
            os.environ["CTS_TEMPLATE_MODE"] = normalize_template_mode(resolved_template_mode)
            os.environ["CTS_STAGE2_SOURCE_SHEET"] = str(sheet_name or "").strip()
            os.environ["CTS_STAGE2_INTERNAL_SHEET"] = str(internal_sheet_name or "").strip()
            mm.process_all_sheets()
            append_src.main(company_name=company_name)
        finally:
            try:
                setattr(mm, "INPUT_WORKBOOK_NAME", orig)
            except Exception:
                pass
            for env_name, env_value in orig_env.items():
                _restore_env_var(env_name, env_value)

    # Pick newest mapped_results output created after we started
    candidates = sorted(
        STAGE2_MAPPING_OUTPUT_DIR.glob("mapped_results*.xlsx"),
        key=lambda p: p.stat().st_mtime,
        reverse=True,
    )
    out_path = None
    for p in candidates:
        try:
            if p.stat().st_mtime >= start_ts - 2:
                out_path = p
                break
        except Exception:
            continue
    if out_path is None and candidates:
        out_path = candidates[0]
    if out_path is None:
        raise RuntimeError("Stage2 mapping did not produce an output workbook.")

    # Copy output to a stable location for this run (Stage2 output dir is shared)
    stable_out = run_dir / "mapped_results.xlsx"
    try:
        shutil.copy2(out_path, stable_out)
        out_path = stable_out
    except Exception:
        pass

    sheets = pd.read_excel(out_path, sheet_name=None, engine="openpyxl")
    # Find the mapped sheet (safe-name truncation may apply)
    mapped_df = None
    for k, v in sheets.items():
        if str(k).strip().lower() == str(sheet_name).strip().lower():
            mapped_df = v
            break
    if mapped_df is None:
        for k, v in sheets.items():
            if str(k).strip().lower() == str(internal_sheet_name).strip().lower():
                mapped_df = v
                break
    if mapped_df is None:
        # fallback: return first sheet if only one
        if len(sheets) == 1:
            mapped_df = next(iter(sheets.values()))
        else:
            # best-effort: match by prefix
            want = str(sheet_name).strip().lower()[:20]
            for k, v in sheets.items():
                if str(k).strip().lower().startswith(want):
                    mapped_df = v
                    break
    if mapped_df is None:
        raise RuntimeError("Could not locate mapped sheet in Stage2 output workbook.")

    return mapped_df, out_path, input_xlsx


def _find_ef_row_in_sheet(ws, ef_id: str) -> int | None:
    ef_id = (ef_id or "").strip()
    if not ef_id:
        return None
    headers = [("" if v is None else str(v).strip()) for v in (next(ws.iter_rows(min_row=1, max_row=1, values_only=True), None) or [])]
    while headers and headers[-1] == "":
        headers.pop()
    if "ef_id" not in headers:
        return None
    ef_col = headers.index("ef_id") + 1
    for idx, row_values in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
        v = row_values[ef_col - 1] if row_values and len(row_values) >= ef_col else None
        if v is None:
            continue
        if str(v).strip() == ef_id:
            return idx
    return None


@app.route("/admin/emission_factors/mapping/update", methods=["POST"])
@login_required
def update_emission_factor_mapping():
    if not current_user.is_admin:
        flash("Access denied")
        return redirect(url_for("dashboard"))

    sheet = (request.form.get("sheet") or "").strip()
    original_ef_id = (request.form.get("original_ef_id") or "").strip()
    ef_id = (request.form.get("ef_id") or "").strip()

    if not sheet or not original_ef_id:
        flash("Missing sheet or ef_id")
        return redirect(url_for("manage_emission_factors"))

    payload = {k: (request.form.get(k) or "").strip() for k in _ef_expected_headers()}
    if ef_id:
        payload["ef_id"] = ef_id

    try:
        wb = load_workbook(STAGE2_EF_XLSX, keep_links=False)
        if sheet not in wb.sheetnames:
            flash("Sheet not found in mapping workbook")
            return redirect(url_for("manage_emission_factors"))
        ws = wb[sheet]

        headers = [("" if v is None else str(v).strip()) for v in (next(ws.iter_rows(min_row=1, max_row=1, values_only=True), None) or [])]
        while headers and headers[-1] == "":
            headers.pop()

        # Validate required headers exist
        missing = [h for h in _ef_expected_headers() if h not in headers]
        if missing:
            flash(f"Mapping sheet headers are missing: {', '.join(missing)}")
            return redirect(url_for("manage_emission_factors"))

        row_idx = _find_ef_row_in_sheet(ws, original_ef_id)
        if row_idx is None:
            flash("Could not locate the emission factor row by ef_id")
            return redirect(url_for("manage_emission_factors"))

        for h in _ef_expected_headers():
            col = headers.index(h) + 1
            v = payload.get(h)
            if h == "ef_value":
                # keep numeric if possible
                try:
                    ws.cell(row=row_idx, column=col).value = float(v) if v != "" else None
                except Exception:
                    ws.cell(row=row_idx, column=col).value = v if v != "" else None
            else:
                ws.cell(row=row_idx, column=col).value = v if v != "" else None

        # Backup then save (protect against accidental edits)
        backup_dir = STAGE2_EF_XLSX.parent / "_ef_backups"
        backup_dir.mkdir(parents=True, exist_ok=True)
        ts = datetime.now().strftime("%Y%m%d_%H%M%S")
        backup_path = backup_dir / f"{STAGE2_EF_XLSX.stem}_before_update_{ts}{STAGE2_EF_XLSX.suffix}"
        try:
            shutil.copy2(STAGE2_EF_XLSX, backup_path)
        except Exception:
            pass

        wb.save(STAGE2_EF_XLSX)
        _clear_ef_cache()
        flash("Emission factor updated in mapping workbook.")
    except PermissionError:
        flash("Mapping workbook is open/locked. Close Excel and try again.")
    except Exception as e:
        flash(f"Update failed: {e}")

    return redirect(url_for("manage_emission_factors", search=request.args.get("search",""), sheet=request.args.get("sheet",""), scope=request.args.get("scope",""), ef_source=request.args.get("ef_source","")))


@app.route("/admin/emission_factors/mapping/create", methods=["POST"])
@login_required
def create_emission_factor_mapping():
    if not current_user.is_admin:
        flash("Access denied")
        return redirect(url_for("dashboard"))

    sheet = (request.form.get("sheet") or "").strip()
    payload = {k: (request.form.get(k) or "").strip() for k in _ef_expected_headers()}
    required = ("ef_name", "scope", "ef_category", "ef_id", "ef_value", "ef_unit", "ef_source")
    missing_values = [field for field in required if not payload.get(field)]
    if not sheet or missing_values:
        flash("Please complete all required emission factor fields.")
        return redirect(url_for("manage_emission_factors"))

    try:
        ef_value = float(payload["ef_value"])
    except Exception:
        flash("ef_value must be numeric.")
        return redirect(url_for("manage_emission_factors"))

    try:
        wb = load_workbook(STAGE2_EF_XLSX, keep_links=False)
        if sheet not in wb.sheetnames:
            flash("Sheet not found in mapping workbook")
            return redirect(url_for("manage_emission_factors"))
        ws = wb[sheet]

        headers = [("" if v is None else str(v).strip()) for v in (next(ws.iter_rows(min_row=1, max_row=1, values_only=True), None) or [])]
        while headers and headers[-1] == "":
            headers.pop()

        missing_headers = [h for h in _ef_expected_headers() if h not in headers]
        if missing_headers:
            flash(f"Mapping sheet headers are missing: {', '.join(missing_headers)}")
            return redirect(url_for("manage_emission_factors"))

        if _find_ef_row_in_sheet(ws, payload["ef_id"]) is not None:
            flash("An emission factor with this ef_id already exists on the selected sheet.")
            return redirect(url_for("manage_emission_factors"))

        backup_dir = STAGE2_EF_XLSX.parent / "_ef_backups"
        backup_dir.mkdir(parents=True, exist_ok=True)
        ts = datetime.now().strftime("%Y%m%d_%H%M%S")
        backup_path = backup_dir / f"{STAGE2_EF_XLSX.stem}_before_create_{ts}{STAGE2_EF_XLSX.suffix}"
        try:
            shutil.copy2(STAGE2_EF_XLSX, backup_path)
        except Exception:
            pass

        row_idx = ws.max_row + 1
        for h in _ef_expected_headers():
            col = headers.index(h) + 1
            value = ef_value if h == "ef_value" else (payload.get(h) or None)
            ws.cell(row=row_idx, column=col).value = value

        wb.save(STAGE2_EF_XLSX)
        _clear_ef_cache()
        flash("Emission factor created in mapping workbook.")
    except PermissionError:
        flash("Mapping workbook is open/locked. Close Excel and try again.")
    except Exception as e:
        flash(f"Create failed: {e}")

    return redirect(url_for("manage_emission_factors", sheet=sheet))


@app.route("/admin/emission_factors/mapping/delete", methods=["POST"])
@login_required
def delete_emission_factor_mapping():
    if not current_user.is_admin:
        flash("Access denied")
        return redirect(url_for("dashboard"))

    sheet = (request.form.get("sheet") or "").strip()
    ef_id = (request.form.get("ef_id") or "").strip()
    if not sheet or not ef_id:
        flash("Missing sheet or ef_id")
        return redirect(url_for("manage_emission_factors"))

    try:
        wb = load_workbook(STAGE2_EF_XLSX, keep_links=False)
        if sheet not in wb.sheetnames:
            flash("Sheet not found in mapping workbook")
            return redirect(url_for("manage_emission_factors"))
        ws = wb[sheet]
        row_idx = _find_ef_row_in_sheet(ws, ef_id)
        if row_idx is None:
            flash("Could not locate the emission factor row by ef_id")
            return redirect(url_for("manage_emission_factors"))

        backup_dir = STAGE2_EF_XLSX.parent / "_ef_backups"
        backup_dir.mkdir(parents=True, exist_ok=True)
        ts = datetime.now().strftime("%Y%m%d_%H%M%S")
        backup_path = backup_dir / f"{STAGE2_EF_XLSX.stem}_before_delete_{ts}{STAGE2_EF_XLSX.suffix}"
        try:
            shutil.copy2(STAGE2_EF_XLSX, backup_path)
        except Exception:
            pass

        ws.delete_rows(row_idx, 1)
        wb.save(STAGE2_EF_XLSX)
        _clear_ef_cache()
        flash("Emission factor deleted from mapping workbook.")
    except PermissionError:
        flash("Mapping workbook is open/locked. Close Excel and try again.")
    except Exception as e:
        flash(f"Delete failed: {e}")

    return redirect(url_for("manage_emission_factors"))


@app.route("/admin/emission_factors/mapping/import", methods=["POST"])
@login_required
def import_emission_factors_mapping():
    if not current_user.is_admin:
        flash("Access denied")
        return redirect(url_for("dashboard"))

    f = request.files.get("excel_file")
    if not f or not f.filename:
        flash("Please choose an Excel file to import.")
        return redirect(url_for("manage_emission_factors"))

    fn = secure_filename(f.filename)
    if not fn.lower().endswith((".xlsx", ".xlsm")):
        flash("Please upload a .xlsx or .xlsm file.")
        return redirect(url_for("manage_emission_factors"))

    tmp_dir = FRONTEND_UPLOAD_DIR
    tmp_dir.mkdir(parents=True, exist_ok=True)
    tmp_path = tmp_dir / f"ef_import_{datetime.now().strftime('%Y%m%d_%H%M%S')}_{fn}"
    f.save(tmp_path)

    # Validate structure: sheets with expected headers
    try:
        wb = load_workbook(tmp_path, read_only=True, data_only=True, keep_links=False)
        expected = _ef_expected_headers()
        bad = []
        for name in wb.sheetnames:
            ws = wb[name]
            first = next(ws.iter_rows(min_row=1, max_row=1, values_only=True), None)
            headers = [("" if v is None else str(v).strip()) for v in (first or [])]
            while headers and headers[-1] == "":
                headers.pop()
            missing = [h for h in expected if h not in headers]
            if missing:
                bad.append(f"{name}: missing {', '.join(missing)}")
        if bad:
            flash("Import failed. Workbook structure is not compatible: " + " | ".join(bad[:4]) + (" ..." if len(bad) > 4 else ""))
            try:
                tmp_path.unlink()
            except Exception:
                pass
            return redirect(url_for("manage_emission_factors"))
    except Exception as e:
        flash(f"Import failed: {e}")
        try:
            tmp_path.unlink()
        except Exception:
            pass
        return redirect(url_for("manage_emission_factors"))
    
    except Exception as e:
        flash(f"Import failed {e}")
        try:
            tmp_path.unlink()
        except Exception:
            pass
        return redirect(url_for("manage_emission_factors"))
    
    try:
        backup_dir = STAGE2_EF_XLSX.parent / "_ef_backups"
        backup_dir.mkdir(parents=True, exists_ok=True)
        ts = datetime.now().strftime("%Y%m%d_%H%M%S")
        if STAGE2_EF_XLSX.exists():
            shutil.copy(STAGE2_EF_XLSX, backup_dir / f"{STAGE2_EF_XLSX.stem}_before_import_{ts}{STAGE2_EF_XLSX.suffix}" )
        os.replace(str(tmp_path), str(STAGE2_EF_XLSX))
        _clear_ef_cache()
        flash("Emission factors imported successfully (sheet preserved).")
    except PermissionError:
        flash("Mapping workbook is open. Please close Excel file and try it again")
        try:
            tmp_path.unlink()
        except Exception:
            pass
        except Exception as e:
            flash(f"Import failed: {e}")
            try:
                tmp_path.unlink()
            except Exception:
                pass

    # Backup current and replace
    try:
        backup_dir = STAGE2_EF_XLSX.parent / "_ef_backups"
        backup_dir.mkdir(parents=True, exist_ok=True)
        ts = datetime.now().strftime("%Y%m%d_%H%M%S")
        if STAGE2_EF_XLSX.exists():
            shutil.copy2(STAGE2_EF_XLSX, backup_dir / f"{STAGE2_EF_XLSX.stem}_before_import_{ts}{STAGE2_EF_XLSX.suffix}")
        os.replace(str(tmp_path), str(STAGE2_EF_XLSX))
        _clear_ef_cache()
        flash("Emission factors imported successfully (sheets preserved).")
    except PermissionError:
        flash("Mapping workbook is open/locked. Close Excel and try again.")
        try:
            tmp_path.unlink()
        except Exception:
            pass
    except Exception as e:
        flash(f"Import failed: {e}")
        try:
            tmp_path.unlink()
        except Exception:
            pass

    return redirect(url_for("manage_emission_factors"))

@app.route('/admin/emission_factors/export', methods=['GET'])
def export_emission_factors():
    if not current_user.is_authenticated or not getattr(current_user, "is_admin", False):
        return redirect(url_for("dashboard"))

    data = _load_stage2_emission_factors()
    all_rows: list[dict[str, object]] = list(data.get("rows") or [])

    search = request.args.get('search', '').strip()
    sheet_filter = request.args.get('sheet', '').strip()
    scope_filter = request.args.get('scope', '').strip()
    source_filter = request.args.get('ef_source', '').strip()

    def _matches(row: dict[str, object]) -> bool:
        if sheet_filter and str(row.get("sheet") or "") != sheet_filter:
            return False
        if scope_filter and str(row.get("scope") or "") != scope_filter:
            return False
        if source_filter and str(row.get("ef_source") or "") != source_filter:
            return False
        if search:
            hay = " ".join(
                str(row.get(k) or "")
                for k in (
                    "sheet",
                    "ef_name",
                    "ef_description",
                    "scope",
                    "ef_category",
                    "ef_id",
                    "ef_value",
                    "ef_unit",
                    "ef_source",
                    "Emission Factor Category",
                )
            ).lower()
            if search.lower() not in hay:
                return False
        return True

    rows = [r for r in all_rows if _matches(r)]
    rows.sort(key=lambda r: (str(r.get("sheet") or ""), str(r.get("ef_name") or ""), str(r.get("ef_id") or "")))

    
    def _safe_xlsx_sheet_name(name: str) -> str:
        invalid = set(':/\\?*[]')
        cleaned = "".join(ch for ch in (name or "Sheet") if ch not in invalid)
        if not cleaned:
            cleaned = "Sheet"
        return cleaned[:31]

    def _valid_sheet_name(name: str) -> bool:
        if not name:
            return False
        if len(name) > 31:
            return False
        if any(ch in name for ch in ':/\\?*[]'):
            return False
        return True
    # Group by source sheet (category)
    groups: dict[str, list[dict[str, object]]] = {}
    for r in rows:
        groups.setdefault(str(r.get("sheet") or "Sheet"), []).append(r)

    expected_headers = _ef_expected_headers()

    wb = Workbook()
    try:
        wb.remove(wb.active)
    except Exception:
        pass

    used_names: dict[str, int] = {}
    for raw_name in sorted(groups.keys(), key=lambda x: x.lower()):
        base = raw_name if _valid_sheet_name(raw_name) else _safe_xlsx_sheet_name(raw_name)
        n = used_names.get(base, 0) + 1
        used_names[base] = n
        sheet_name = base if n == 1 else _safe_xlsx_sheet_name(f"{base}_{n}")

        ws = wb.create_sheet(title=sheet_name)
        # Header row
        for j, h in enumerate(expected_headers, start=1):
            ws.cell(row=1, column=j).value = h

        # Rows
        for i, r in enumerate(groups[raw_name], start=2):
            for j, h in enumerate(expected_headers, start=1):
                v = r.get(h)
                # Keep ef_value numeric when possible
                if h == "ef_value":
                    try:
                        if v is None or str(v).strip() == "":
                            ws.cell(row=i, column=j).value = None
                        else:
                            ws.cell(row=i, column=j).value = float(v)
                    except Exception:
                        ws.cell(row=i, column=j).value = "" if v is None else str(v)
                else:
                    ws.cell(row=i, column=j).value = None if v is None or str(v).strip() == "" else str(v)

        # Light column sizing
        for col in range(1, len(expected_headers) + 1):
            ws.column_dimensions[chr(64 + col)].width = 22 if col in (1, 4, 5, 6, 7, 8) else 34

        ws.freeze_panes = "A2"

    bio = BytesIO()
    wb.save(bio)
    bio.seek(0)

    filename = f"CTS_Emission_factors_short_list_export_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
    return send_file(
        bio,
        as_attachment=True,
        download_name=filename,
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )


def _reporting_row_year(row: dict[str, object]) -> int | None:
    sk = str(row.get("sortKey") or "").strip()
    if len(sk) >= 4 and sk[:4].isdigit():
        y = int(sk[:4])
        if 1990 <= y <= 2100:
            return y
    return None


def _infer_scope_from_sheet_name(sheet_name: str | None) -> int:
    """
    When MappingRunSummary.scope is NULL or wrong, many templates encode scope in the sheet title
    (e.g. 'Scope 1 Fuel Usage', 'Scope 3 Cat 1 Goods Spend'). Used so scope detail pages match Home.
    """
    s = (sheet_name or "").strip().lower()
    if not s:
        return 0
    if re.search(r"\bscope\s*3\b", s):
        return 3
    if re.search(r"\bscope\s*2\b", s):
        return 2
    if re.search(r"\bscope\s*1\b", s):
        return 1
    return 0


def _effective_scope(scope_val: object, sheet_name: str | None) -> int:
    try:
        n = int(scope_val) if scope_val is not None else 0
    except (TypeError, ValueError):
        n = 0
    if n in (1, 2, 3):
        return n
    inferred = _infer_scope_from_sheet_name(sheet_name)
    return inferred if inferred else n


def _admin_aggregates_from_reporting_rows(
    rows: list[dict[str, object]],
) -> tuple[dict[str, float], list[dict[str, object]], list[dict[str, object]]]:
    company_totals: dict[str, dict[str, float]] = defaultdict(
        lambda: {"scope1": 0.0, "scope2": 0.0, "scope3": 0.0, "total": 0.0}
    )
    g = {"scope1": 0.0, "scope2": 0.0, "scope3": 0.0, "total": 0.0}
    sheet_agg: dict[tuple[int, str], float] = defaultdict(float)
    for r in rows:
        v = float(r.get("emissions") or 0.0)
        sheet = str(r.get("sheet") or r.get("category") or "Category")
        sc = _effective_scope(r.get("scope"), sheet)
        c = (str(r.get("company") or "").strip()) or "—"
        if sc == 1:
            company_totals[c]["scope1"] += v
        elif sc == 2:
            company_totals[c]["scope2"] += v
        elif sc == 3:
            company_totals[c]["scope3"] += v
        company_totals[c]["total"] += v
        if sc == 1:
            g["scope1"] += v
        elif sc == 2:
            g["scope2"] += v
        elif sc == 3:
            g["scope3"] += v
        g["total"] += v
        sheet_agg[(sc, sheet)] += v
    out_rows: list[dict[str, object]] = []
    for c in sorted(company_totals.keys(), key=lambda x: x.lower()):
        t = company_totals[c]
        out_rows.append({"company": c, **{k: round(t[k], 3) for k in ("scope1", "scope2", "scope3", "total")}})
    breakdown = [
        {"sheet": sheet, "scope": sc, "tco2e": round(float(v), 3), "updated_at": ""}
        for (sc, sheet), v in sorted(sheet_agg.items(), key=lambda x: (-x[1], str(x[0][1]).lower()))
    ]
    return {k: round(g[k], 3) for k in g}, out_rows, breakdown


def _user_totals_and_breakdown_from_reporting_rows(
    rows: list[dict[str, object]],
) -> tuple[dict[str, float], list[dict[str, object]]]:
    s1 = s2 = s3 = 0.0
    sheet_agg: dict[tuple[int, str], float] = defaultdict(float)
    for r in rows:
        v = float(r.get("emissions") or 0.0)
        sheet = str(r.get("sheet") or r.get("category") or "Category")
        sc = _effective_scope(r.get("scope"), sheet)
        if sc == 1:
            s1 += v
        elif sc == 2:
            s2 += v
        elif sc == 3:
            s3 += v
        sheet_agg[(sc, sheet)] += v
    totals = {
        "total": round(s1 + s2 + s3, 3),
        "scope1": round(s1, 3),
        "scope2": round(s2, 3),
        "scope3": round(s3, 3),
    }
    breakdown = [
        {"sheet": sheet, "scope": sc, "tco2e": round(float(v), 3), "updated_at": ""}
        for (sc, sheet), v in sorted(sheet_agg.items(), key=lambda x: (-x[1], str(x[0][1]).lower()))
    ]
    return totals, breakdown


def _home_overview_context():
    """Shared context for Overview (home) and per-scope detail pages."""
    _ensure_db_tables()
    req_year = request.args.get("year", type=int)
    default_year = datetime.utcnow().year

    if bool(getattr(current_user, "is_admin", False)):
        all_rows = MappingRunSummary.query.order_by(MappingRunSummary.created_at.desc()).all()
        latest_by_company_sheet: dict[tuple[str, str], MappingRunSummary] = {}
        for r in all_rows:
            key = ((r.company_name or "").strip(), (r.sheet_name or "").strip().lower())
            if key in latest_by_company_sheet:
                continue
            if not key[0] or not key[1]:
                continue
            latest_by_company_sheet[key] = r

        latest_list = list(latest_by_company_sheet.values())
        reporting_rows_full: list[dict[str, object]] = []
        for r in latest_list:
            try:
                reporting_rows_full.extend(_build_reporting_rows_from_summary(r))
            except Exception:
                continue

        years_in_data = sorted({_reporting_row_year(x) for x in reporting_rows_full if _reporting_row_year(x)})
        run_years = sorted({r.created_at.year for r in latest_list if getattr(r, "created_at", None)})

        if years_in_data:
            selected_year = req_year if req_year in years_in_data else max(years_in_data)
            reporting_rows_adm = [x for x in reporting_rows_full if _reporting_row_year(x) == selected_year]
            if reporting_rows_adm:
                g, rows, breakdown_adm = _admin_aggregates_from_reporting_rows(reporting_rows_adm)
            else:
                g = {"scope1": 0.0, "scope2": 0.0, "scope3": 0.0, "total": 0.0}
                rows = []
                breakdown_adm = []
            return {
                "year": selected_year,
                "selected_year": selected_year,
                "available_years": years_in_data,
                "is_admin": True,
                "totals": g,
                "company_rows": rows,
                "breakdown": breakdown_adm,
                "reporting_rows": reporting_rows_adm,
            }

        company_totals: dict[str, dict[str, float]] = defaultdict(
            lambda: {"scope1": 0.0, "scope2": 0.0, "scope3": 0.0, "total": 0.0}
        )
        for (_c, _s), r in latest_by_company_sheet.items():
            c = (r.company_name or "").strip()
            v = float(getattr(r, "tco2e_total", 0.0) or 0.0)
            sc = _effective_scope(getattr(r, "scope", None), getattr(r, "sheet_name", None))
            if sc == 1:
                company_totals[c]["scope1"] += v
            elif sc == 2:
                company_totals[c]["scope2"] += v
            elif sc == 3:
                company_totals[c]["scope3"] += v
            company_totals[c]["total"] += v

        rows_out = []
        g = {"scope1": 0.0, "scope2": 0.0, "scope3": 0.0, "total": 0.0}
        for c in sorted(company_totals.keys(), key=lambda x: x.lower()):
            t = company_totals[c]
            rows_out.append({"company": c, **{k: round(t[k], 3) for k in ("scope1", "scope2", "scope3", "total")}})
            for k in g:
                g[k] += float(t[k] or 0.0)

        breakdown_adm = [
            {
                "sheet": r.sheet_name,
                "scope": _effective_scope(getattr(r, "scope", None), getattr(r, "sheet_name", None)),
                "tco2e": round(float(r.tco2e_total or 0.0), 3),
                "updated_at": r.created_at.strftime("%Y-%m-%d %H:%M") if r.created_at else "",
            }
            for r in latest_list
        ]
        available_years = run_years or [default_year]
        selected_year = req_year if req_year in available_years else max(available_years)
        display_year = selected_year
        return {
            "year": display_year,
            "selected_year": selected_year,
            "available_years": available_years,
            "is_admin": True,
            "totals": {k: round(g[k], 3) for k in g},
            "company_rows": rows_out,
            "breakdown": breakdown_adm,
            "reporting_rows": reporting_rows_full,
        }

    keys = _company_candidate_keys(getattr(current_user, "company_name", "") or "")
    latest = _latest_sheet_totals_for_company(keys)
    year = default_year
    if latest and getattr(latest[0], "created_at", None):
        year = latest[0].created_at.year

    scope1 = sum(
        float(r.tco2e_total or 0.0)
        for r in latest
        if _effective_scope(getattr(r, "scope", None), getattr(r, "sheet_name", None)) == 1
    )
    scope2 = sum(
        float(r.tco2e_total or 0.0)
        for r in latest
        if _effective_scope(getattr(r, "scope", None), getattr(r, "sheet_name", None)) == 2
    )
    scope3 = sum(
        float(r.tco2e_total or 0.0)
        for r in latest
        if _effective_scope(getattr(r, "scope", None), getattr(r, "sheet_name", None)) == 3
    )
    total_emission = scope1 + scope2 + scope3
    breakdown = [
        {
            "sheet": r.sheet_name,
            "scope": _effective_scope(getattr(r, "scope", None), getattr(r, "sheet_name", None)),
            "tco2e": round(float(r.tco2e_total or 0.0), 3),
            "updated_at": r.created_at.strftime("%Y-%m-%d %H:%M") if r.created_at else "",
        }
        for r in latest
    ]

    reporting_rows_full: list[dict[str, object]] = []
    for r in latest:
        try:
            reporting_rows_full.extend(_build_reporting_rows_from_summary(r))
        except Exception:
            continue

    years_in_data = sorted({_reporting_row_year(x) for x in reporting_rows_full if _reporting_row_year(x)})
    if years_in_data:
        selected_year = req_year if req_year in years_in_data else max(years_in_data)
        reporting_rows = [x for x in reporting_rows_full if _reporting_row_year(x) == selected_year]
        if reporting_rows:
            totals, breakdown = _user_totals_and_breakdown_from_reporting_rows(reporting_rows)
        else:
            totals = {"total": 0.0, "scope1": 0.0, "scope2": 0.0, "scope3": 0.0}
            breakdown = []
        return {
            "year": selected_year,
            "selected_year": selected_year,
            "available_years": years_in_data,
            "is_admin": False,
            "totals": totals,
            "breakdown": breakdown,
            "company_rows": [],
            "reporting_rows": reporting_rows,
        }

    if reporting_rows_full:
        try:
            yrs = [int(str(x.get("sortKey", ""))[:4]) for x in reporting_rows_full if x.get("sortKey")]
            if yrs:
                year = max(yrs)
        except Exception:
            pass

    return {
        "year": year,
        "selected_year": year,
        "available_years": [year],
        "is_admin": False,
        "totals": {
            "total": round(total_emission, 3),
            "scope1": round(scope1, 3),
            "scope2": round(scope2, 3),
            "scope3": round(scope3, 3),
        },
        "breakdown": breakdown,
        "company_rows": [],
        "reporting_rows": reporting_rows_full,
    }


@app.route('/home')
@login_required
def home():
    ctx = _home_overview_context()
    return render_template("home.html", **ctx)


@app.route("/audit-2025")
@login_required
def audit_2025_page():
    return render_template("audit_2025.html")


@app.route("/api/audit-2025")
@login_required
def audit_2025_api():
    payload = load_audit_2025_data()
    status_code = 200 if payload.get("ok") else 503
    return jsonify(payload), status_code


@app.route("/feed")
@login_required
def feed():
    _ensure_db_tables()
    selected_filter = _normalize_feed_filter(request.args.get("type"))
    is_my_updates_view = str(request.args.get("view") or "").strip().lower() == "mine"
    query = FeedPost.query.order_by(FeedPost.created_at.desc(), FeedPost.id.desc())
    if is_my_updates_view:
        query = query.filter(FeedPost.author_user_id == int(current_user.id))
    if selected_filter != "all":
        query = query.filter(FeedPost.post_type == selected_filter)
    rows = query.all()
    posts = _feed_payloads_for_rows(rows)
    can_create_posts = not _is_readonly_user(current_user)
    return render_template(
        "feed.html",
        posts=posts,
        selected_filter=selected_filter,
        feed_filters=FEED_FILTER_OPTIONS,
        feed_post_types=FEED_COMPOSER_TYPES,
        feed_reaction_options=FEED_REACTION_OPTIONS,
        can_create_posts=can_create_posts,
        comment_actor_avatar_url=_user_avatar_url(current_user),
        can_create_challenges=bool(current_user.is_admin and not _is_readonly_user(current_user)),
        is_my_updates_view=is_my_updates_view,
        feed_profile={
            "name": _user_display_name(current_user),
            "title": _user_professional_title(current_user),
            "company_name": (getattr(current_user, "company_name", None) or "").strip() or "CTS Carbon Platform",
            "avatar_url": _user_avatar_url(current_user),
            "company_logo_url": _company_logo_url(getattr(current_user, "company_name", None)),
            "profile_url": url_for("public_profile", user_id=int(current_user.id)),
            "reports_url": url_for("public_profile", user_id=int(current_user.id)) + "#reports",
            "updates_url": url_for("feed", view="mine"),
            "analytics_url": url_for("analytics_emissions_totals"),
            "quick_reports_url": url_for("reports_page"),
            "quick_newsletters_url": url_for("newsletters_page"),
            "quick_events_url": url_for("events_page"),
        },
    )


@app.route("/feed/posts", methods=["POST"])
@login_required
def create_feed_post():
    _ensure_db_tables()
    def redirect_after_post():
        next_url = str(request.form.get("next") or "").strip()
        if next_url.startswith("/") and not next_url.startswith("//"):
            return redirect(next_url)
        return redirect(url_for("feed", type=_normalize_feed_filter(request.form.get("current_filter"))))

    post_type = _normalize_feed_post_type(request.form.get("post_type"))
    content = (request.form.get("content") or "").strip()
    report_title = (request.form.get("report_title") or "").strip()
    image_file = request.files.get("image_file")
    video_file = request.files.get("video_file")
    report_file = request.files.get("report_file")
    selected_files = [
        item for item in (image_file, video_file, report_file)
        if item and getattr(item, "filename", None)
    ]

    if not content and not selected_files:
        flash("Add some text or attach a file before posting.", "warning")
        return redirect_after_post()

    if len(selected_files) > 1:
        flash("Please upload only one image, video, or file per post.", "warning")
        return redirect_after_post()

    media_path = None
    media_type = None
    reference_id = None
    reference_type = None
    if selected_files:
        selected = selected_files[0]
        selected_name = secure_filename(getattr(selected, "filename", "") or "")
        selected_ext = Path(selected_name).suffix.lower()
        if selected_ext in REPORT_PREVIEWABLE_EXT:
            resolved_company = _clean_company_name(getattr(current_user, "company_name", "") or "")
            company_row = _company_row_for_name(resolved_company, created_by_user_id=int(current_user.id))
            if company_row is None:
                flash("A company is required before publishing a report.", "warning")
                return redirect_after_post()
            report_file_path, report_error = _save_report_file(selected, user_id=int(current_user.id))
            if report_error:
                flash(report_error, "warning")
                return redirect_after_post()
            title = report_title or Path(selected_name).stem.replace("_", " ").strip() or "Untitled report"
            report_row = Report(
                title=title,
                file_path=str(report_file_path),
                preview_paths="[]",
                uploaded_by=int(current_user.id),
                company_id=int(company_row.id),
            )
            db.session.add(report_row)
            db.session.flush()
            report_row.preview_paths = json.dumps(
                _generate_report_preview_paths(
                    report_title=title,
                    report_rel_path=str(report_file_path),
                    report_id=int(report_row.id),
                )
            )
            reference_id = int(report_row.id)
            reference_type = "report"
            post_type = "report"
        else:
            media_path, media_type, media_error = _save_feed_media_file(selected, user_id=int(current_user.id))
            if media_error:
                flash(media_error, "warning")
                return redirect_after_post()
    if post_type in {"report", "newsletter", "event"} and reference_type != post_type:
        post_type = "update"

    row = FeedPost(
        author_user_id=int(current_user.id),
        content=content,
        post_type=post_type,
        media_type=media_type,
        media_path=media_path,
        reference_id=reference_id,
        reference_type=reference_type,
    )
    db.session.add(row)
    db.session.commit()
    flash("Post shared successfully.", "success")
    return redirect_after_post()


@app.route("/feed/challenges", methods=["POST"])
@login_required
def create_challenge():
    _ensure_db_tables()
    if not bool(current_user.is_admin):
        abort(403)
    title = (request.form.get("title") or "").strip()
    description = (request.form.get("description") or "").strip()
    deadline_raw = (request.form.get("deadline") or "").strip()
    if not title or not description:
        flash("Challenge title and description are required.", "warning")
        return redirect(url_for("feed"))
    deadline_value = None
    if deadline_raw:
        try:
            deadline_value = datetime.fromisoformat(deadline_raw)
        except Exception:
            flash("Use a valid challenge deadline.", "warning")
            return redirect(url_for("feed"))
    row = Challenge(
        title=title,
        description=description,
        created_by=int(current_user.id),
        deadline=deadline_value,
    )
    db.session.add(row)
    db.session.flush()
    db.session.add(
        FeedPost(
            author_user_id=int(current_user.id),
            content=description,
            post_type="alert",
            reference_id=int(row.id),
            reference_type="challenge",
        )
    )
    db.session.commit()
    flash("Challenge published to the feed.", "success")
    return redirect(url_for("feed", type="alert"))


@app.route("/feed/challenges/<int:challenge_id>/responses", methods=["POST"])
@login_required
def submit_challenge_response(challenge_id: int):
    _ensure_db_tables()
    challenge_row = Challenge.query.get_or_404(int(challenge_id))
    if getattr(challenge_row, "deadline", None) and challenge_row.deadline < datetime.utcnow():
        flash("This challenge is closed.", "warning")
        return redirect(url_for("feed"))
    answer = (request.form.get("answer") or "").strip()
    if not answer:
        flash("Add a response before submitting.", "warning")
        return redirect(url_for("feed"))
    response_row = ChallengeResponse(
        challenge_id=int(challenge_row.id),
        user_id=int(current_user.id),
        answer=answer,
    )
    db.session.add(response_row)
    db.session.flush()
    challenge_post = (
        FeedPost.query.filter_by(reference_type="challenge", reference_id=int(challenge_row.id))
        .order_by(FeedPost.created_at.desc(), FeedPost.id.desc())
        .first()
    )
    if challenge_post is not None:
        mention_ids = _normalize_comment_mention_ids(request.form.getlist("mentioned_user_ids"), answer)
        db.session.add(
            Comment(
                post_id=int(challenge_post.id),
                user_id=int(current_user.id),
                content=answer,
                mentioned_user_ids_json=json.dumps(mention_ids),
            )
        )
    db.session.commit()
    flash("Challenge response added as a comment.", "success")
    return redirect(url_for("feed", type="alert"))


@app.route("/api/feed/posts/<int:post_id>/comments", methods=["POST"])
@login_required
def api_feed_post_comment(post_id: int):
    _ensure_db_tables()
    post = FeedPost.query.get(post_id)
    if post is None:
        return jsonify({"error": "Post not found."}), 404
    payload = request.get_json(silent=True) or {}
    content = str(payload.get("content", "") or "").strip()
    if not content:
        return jsonify({"error": "Add a comment before posting."}), 400
    mention_ids = _normalize_comment_mention_ids(payload.get("mentioned_user_ids"), content)
    row = Comment(
        post_id=int(post.id),
        user_id=int(current_user.id),
        content=content,
        mentioned_user_ids_json=json.dumps(mention_ids),
    )
    db.session.add(row)
    db.session.commit()
    return jsonify({"ok": True, "comment": _comment_payload(row, like_count=0, liked_by_viewer=False)})


@app.route("/api/feed/comments/<int:comment_id>/like", methods=["POST"])
@login_required
def api_feed_comment_like(comment_id: int):
    _ensure_db_tables()
    comment = Comment.query.get(comment_id)
    if comment is None:
        return jsonify({"error": "Comment not found."}), 404
    row = CommentLike.query.filter_by(comment_id=int(comment.id), user_id=int(current_user.id)).first()
    if row is None:
        db.session.add(CommentLike(comment_id=int(comment.id), user_id=int(current_user.id)))
        liked = True
    else:
        db.session.delete(row)
        liked = False
    db.session.commit()
    like_count = (
        db.session.query(db.func.count(CommentLike.id))
        .filter(CommentLike.comment_id == int(comment.id))
        .scalar()
        or 0
    )
    return jsonify({"ok": True, "liked": liked, "like_count": int(like_count)})


@app.route("/api/feed/posts/<int:post_id>/reaction", methods=["POST"])
@login_required
def api_feed_post_reaction(post_id: int):
    _ensure_db_tables()
    payload = request.get_json(silent=True) or {}
    reaction_type = _normalize_feed_reaction_type(payload.get("reaction_type"))
    if not reaction_type:
        return jsonify({"error": "Select a valid reaction."}), 400

    post = FeedPost.query.get(post_id)
    if post is None:
        return jsonify({"error": "Post not found."}), 404

    row = PostReaction.query.filter_by(post_id=int(post.id), user_id=int(current_user.id)).first()
    now = datetime.utcnow()
    if row is None:
        row = PostReaction(
            post_id=int(post.id),
            user_id=int(current_user.id),
            reaction_type=reaction_type,
            created_at=now,
            updated_at=now,
        )
        db.session.add(row)
    else:
        row.reaction_type = reaction_type
        row.updated_at = now

    try:
        db.session.commit()
    except Exception:
        db.session.rollback()
        return jsonify({"error": "Could not save reaction."}), 500

    summary_map, current_map = _feed_reaction_maps([int(post.id)], int(current_user.id))
    button_state = _feed_reaction_button_state(current_map.get(int(post.id)))
    summary = summary_map.get(int(post.id), [])
    return jsonify(
        {
            "ok": True,
            "post_id": int(post.id),
            "reaction_summary": summary,
            "reaction_total": sum(int(item.get("count") or 0) for item in summary),
            "current_reaction": str(button_state.get("type") or ""),
            "current_reaction_label": str(button_state.get("label") or "Like"),
            "current_reaction_icon": str(button_state.get("icon") or "👍"),
        }
    )


@app.route("/scope1")
@login_required
def scope1_detail():
    ctx = _home_overview_context()
    return render_template(
        "scope_detail.html",
        scope_num=1,
        scope_title="Scope 1 — Direct Emissions",
        scope_subtitle="Fuel Usage (Diesel on site, fuels from owned and leased vehicles)",
        **ctx,
    )


@app.route("/scope2")
@login_required
def scope2_detail():
    ctx = _home_overview_context()
    return render_template(
        "scope_detail.html",
        scope_num=2,
        scope_title="Scope 2 - Indirect Energy Emissions",
        scope_subtitle="Electricity and purchased energy with time and category views. Electricity for facilities. District heating for some offices.",
        **ctx,
    )


@app.route("/scope3")
@login_required
def scope3_detail():
    ctx = _home_overview_context()
    return render_template(
        "scope_detail.html",
        scope_num=3,
        scope_title="Scope 3 - Value Chain Emissions",
        scope_subtitle="Purchased goods, transport, travel, waste, and other upstream/downstream categories.",
        **ctx,
    )

    
@app.route('/Emission-factors')
@login_required
def emission_facor():
    return redirect(url_for('manage_emission_factors'))


@app.route('/carbon-accounting')
@login_required
def carbon_accounting():
    _ensure_db_tables()
    year = datetime.utcnow().year

    keys = _company_candidate_keys(getattr(current_user, "company_name", "") or "")
    latest = _latest_sheet_totals_for_company(keys)
    if latest and getattr(latest[0], "created_at", None):
        year = latest[0].created_at.year

    scope1 = sum(float(r.tco2e_total or 0.0) for r in latest if int(r.scope or 0) == 1)
    scope2 = sum(float(r.tco2e_total or 0.0) for r in latest if int(r.scope or 0) == 2)
    scope3 = sum(float(r.tco2e_total or 0.0) for r in latest if int(r.scope or 0) == 3)
    total_emission = scope1 + scope2 + scope3

    by_sheet = [
        {
            "sheet": r.sheet_name,
            "scope": r.scope,
            "tco2e": round(float(r.tco2e_total or 0.0), 3),
            "updated_at": r.created_at.strftime("%Y-%m-%d %H:%M") if r.created_at else "",
        }
        for r in latest
    ]

    return render_template(
        "carbon_accounting.html",
        year=year,
        totals={
            "total": round(total_emission, 3),
            "scope1": round(scope1, 3),
            "scope2": round(scope2, 3),
            "scope3": round(scope3, 3),
        },
        by_sheet=by_sheet,
    )


@app.route("/search", methods=["GET"])
@login_required
def search_page():
    query = str(request.args.get("q", "") or "").strip()
    groups = search_service.search_all(
        query,
        CompanyModel=Company,
        UserModel=User,
        EmissionFactorModel=EmissionFactor,
        stage1_input_dir=STAGE1_INPUT_DIR,
        stage2_output_dir=STAGE2_OUTPUT_DIR,
        run_logs_dir=APP_DIR / "run_logs",
    )
    total_results = sum(len(rows) for rows in groups.values())
    return render_template(
        "search_results.html",
        user=current_user,
        query=query,
        results=groups,
        total_results=total_results,
    )


@app.route("/api/notifications/recent", methods=["GET"])
@login_required
def api_notifications_recent():
    _ensure_db_tables()
    rows = notification_service.recent_notifications(
        Notification,
        user_id=current_user.id,
        limit=_parse_int_arg("limit", 8, minimum=1, maximum=30),
    )
    unread = notification_service.unread_count(Notification, user_id=current_user.id)
    return jsonify({"notifications": [_notification_payload(row) for row in rows], "unread_count": unread})


@app.route("/api/notifications/unread_count", methods=["GET"])
@login_required
def api_notifications_unread_count():
    _ensure_db_tables()
    return jsonify({"unread_count": notification_service.unread_count(Notification, user_id=current_user.id)})


@app.route("/api/notifications/mark-read", methods=["POST"])
@login_required
def api_notifications_mark_read():
    _ensure_db_tables()
    payload = request.get_json(silent=True) or {}
    notification_id = payload.get("notification_id")
    if notification_id in (None, "", "all"):
        updated = notification_service.mark_all_read(db.session, Notification, user_id=current_user.id)
    else:
        updated = 1 if notification_service.mark_one_read(
            db.session,
            Notification,
            user_id=current_user.id,
            notification_id=int(notification_id),
        ) else 0
    db.session.commit()
    return jsonify(
        {
            "ok": True,
            "updated": int(updated),
            "unread_count": notification_service.unread_count(Notification, user_id=current_user.id),
        }
    )


@app.route("/api/messages/contacts", methods=["GET"])
@login_required
def api_message_contacts():
    _ensure_db_tables()
    contacts = messaging_service.available_contacts(User, current_user, limit=50)
    return jsonify({"contacts": [_contact_payload(row) for row in contacts]})


@app.route("/api/messages/conversations", methods=["GET"])
@login_required
def api_message_conversations():
    _ensure_db_tables()
    conversations = messaging_service.list_conversations(Message, User, current_user, limit=30)
    payload = []
    for item in conversations:
        payload.append(
            {
                "thread_id": item["thread_id"],
                "other_user": _contact_payload(item["other_user"]),
                "last_message": _message_payload(item["last_message"], viewer_id=current_user.id),
                "unread_count": int(item["unread_count"]),
            }
        )
    return jsonify(
        {
            "conversations": payload,
            "unread_count": messaging_service.unread_count(Message, current_user),
        }
    )


@app.route("/api/messages/thread", methods=["GET"])
@login_required
def api_message_thread():
    _ensure_db_tables()
    other_user_id = _parse_int_arg("user_id", None, minimum=1)
    after_id = _parse_int_arg("after_id", 0, minimum=0)
    if not other_user_id:
        return jsonify({"error": "user_id is required"}), 400
    try:
        other_user, rows = messaging_service.fetch_thread_messages(
            Message,
            User,
            current_user,
            other_user_id=other_user_id,
            limit=250,
        )
    except ValueError as exc:
        return jsonify({"error": str(exc)}), 403
    if after_id:
        rows = [row for row in rows if int(getattr(row, "id", 0) or 0) > int(after_id)]
    return jsonify(
        {
            "contact": _contact_payload(other_user),
            "messages": [_message_payload(row, viewer_id=current_user.id) for row in rows],
        }
    )


@app.route("/api/messages/send", methods=["POST"])
@login_required
def api_message_send():
    _ensure_db_tables()
    payload = request.get_json(silent=True) or {}
    receiver_id = payload.get("receiver_id")
    if receiver_id in (None, ""):
        return jsonify({"error": "receiver_id is required"}), 400
    try:
        row = messaging_service.send_message(
            db.session,
            Message,
            User,
            current_user,
            receiver_id=int(receiver_id),
            body=str(payload.get("message", "") or ""),
        )
        _set_message_typing_state(int(current_user.id), int(receiver_id), is_typing=False)
        db.session.commit()
        _create_user_notification(
            row.receiver_id,
            title=f"New message from {_user_display_name(current_user)}",
            message=row.message[:180],
            notification_type="message",
            link=None,
        )
    except ValueError as exc:
        db.session.rollback()
        return jsonify({"error": str(exc)}), 400
    except Exception as exc:
        db.session.rollback()
        return jsonify({"error": f"Message send failed: {exc}"}), 500
    return jsonify(
        {
            "ok": True,
            "message": _message_payload(row, viewer_id=current_user.id),
            "unread_count": messaging_service.unread_count(Message, current_user),
        }
    )


@app.route("/api/messages/mark-read", methods=["POST"])
@login_required
def api_message_mark_read():
    _ensure_db_tables()
    payload = request.get_json(silent=True) or {}
    other_user_id = payload.get("user_id")
    if other_user_id in (None, ""):
        return jsonify({"error": "user_id is required"}), 400
    updated = messaging_service.mark_thread_read(
        db.session,
        Message,
        User,
        current_user,
        other_user_id=int(other_user_id),
    )
    db.session.commit()
    return jsonify(
        {
            "ok": True,
            "updated": int(updated),
            "unread_count": messaging_service.unread_count(Message, current_user),
        }
    )


@app.route("/api/messages/typing", methods=["POST"])
@login_required
def api_message_typing():
    _ensure_db_tables()
    payload = request.get_json(silent=True) or {}
    receiver_id = payload.get("receiver_id")
    if receiver_id in (None, ""):
        return jsonify({"error": "receiver_id is required"}), 400
    if int(receiver_id) == int(current_user.id):
        return jsonify({"error": "Cannot target the current user."}), 400
    receiver = User.query.get(int(receiver_id))
    if receiver is None:
        return jsonify({"error": "User not found."}), 404
    _set_message_typing_state(
        int(current_user.id),
        int(receiver_id),
        is_typing=bool(payload.get("is_typing")),
    )
    return jsonify({"ok": True})


@app.route("/api/messages/typing_status", methods=["GET"])
@login_required
def api_message_typing_status():
    _ensure_db_tables()
    other_user_id = _parse_int_arg("user_id", None, minimum=1)
    if not other_user_id:
        return jsonify({"error": "user_id is required"}), 400
    other_user = User.query.get(int(other_user_id))
    if other_user is None:
        return jsonify({"error": "User not found."}), 404
    is_typing = _message_typing_status(int(other_user_id), int(current_user.id))
    return jsonify(
        {
            "ok": True,
            "is_typing": bool(is_typing),
            "user_name": _user_display_name(other_user),
        }
    )


@app.route("/api/messages/unread_count", methods=["GET"])
@login_required
def api_message_unread_count():
    _ensure_db_tables()
    return jsonify({"unread_count": messaging_service.unread_count(Message, current_user)})


@app.route('/admin_data', methods=['GET'])
@login_required
def admin_data():
    rows = MappingRunSummary.query.order_by(MappingRunSummary.created_at.desc()).all()
    data = [
        {
            "company_name": row.company_name,
            "created_at": row.created_at.strftime("%Y-%m") if row.created_at else "",
            "mapped_categories_count": int(getattr(row, "mapped_categories_count", 0) or 0),
            "total_categories": int(getattr(row, "total_categories", 0) or 0),
            "coverage_pct": float(getattr(row, "coverage_pct", 0) or 0),
        }
        for row in rows
    ]
    return jsonify({"data": data})


with app.app_context():
    _ensure_db_tables()


if __name__ == '__main__':
    with app.app_context():
        db.create_all()
    app.run(debug=True, host='0.0.0.0', port=5000) 
