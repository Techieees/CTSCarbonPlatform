from __future__ import annotations

from dataclasses import dataclass
from pathlib import Path
from typing import Dict, Iterable, List, Optional, Tuple
import sys

import pandas as pd

PROJECT_ROOT = Path(__file__).resolve().parents[2]
if str(PROJECT_ROOT) not in sys.path:
    sys.path.insert(0, str(PROJECT_ROOT))

from config import STAGE2_OUTPUT_DIR


OLD_SHEETS_TO_REMOVE = [
    "Company Totals",
    "Company by GHGP Sheet Totals",
    "GHGP Sheet Totals",
    "Company Stacked Data",
    "Company Stacked Data by Months",
]

"""
NOTE (Excel constraint):
Excel sheet names are limited to 31 characters.
Some user-requested names exceed this limit (e.g. 'Company by GHGP Sheet Totals Window').
To keep the output workbook readable in Excel, we write those few sheets using
shortened, deterministic names below.
"""

SHEET_NAME_MAP: Dict[str, str] = {
    # co2e (t)
    "Company Totals Window": "Company Totals Window",
    "Company by GHGP Sheet Totals Window": "Company by GHGP Totals Window",  # <=31
    "GHGP sheet Totals Window": "GHGP sheet Totals Window",
    "Company Stacked Data Window": "Company Stacked Data Window",
    "Company Stacked Data by Months Window": "Company Stacked Months Window",  # <=31
    # Spend_Euro
    "Company Totals Window Spend": "Company Totals Window Spend",
    "Company by GHGP Sheet Totals Window Spend": "Co by GHGP Totals Win Spend",  # <=31
    "GHGP sheet Totals Window Spend": "GHGP sheet Totals Window Spend",
    "Company Stacked Data Window Spend": "Company Stacked Win Spend",  # <=31
    "Company Stacked Data by Months Window Spend": "Co Stacked Months Win Spend",  # <=31
}


# Canonical GHGP sheets (exact names requested by the user)
ALL_GHGP_SHEETS = [
    "S3 Cat 1 Purchased G&S",
    "S3 Cat 11 Use of Sold",
    "S3 Cat 12 End of Life",
    "S3 Cat 3 FERA",
    "S3 Cat 4 Upstream Transport",
    "S3 Cat 5 Waste",
    "S3 Cat 6 Business Travel",
    "S3 Cat 7 Employee Commute",
    "S3 Cat 9 Downstream Transport",
    "Scope 1",
    "Scope 2",
    "Water",
]

# For table #1 (explicit list, Water excluded and Cat 7 not included per user instruction)
TOTALS_SHEETS_EXCLUDING_WATER = [
    "Scope 1",
    "Scope 2",
    "S3 Cat 1 Purchased G&S",
    "S3 Cat 5 Waste",
    "S3 Cat 9 Downstream Transport",
    "S3 Cat 12 End of Life",
    "S3 Cat 6 Business Travel",
    "S3 Cat 11 Use of Sold",
    "S3 Cat 4 Upstream Transport",
    "S3 Cat 3 FERA",
]


CO2E_COL = "co2e (t)"
SPEND_COL = "Spend_Euro"
COMPANY_COL = "Company"
DATE_COL = "Date"


@dataclass(frozen=True)
class _MetricSpec:
    metric_col: str
    # Column name to use in output tables (for totals tables where the metric column is explicit)
    output_metric_name: str


def _safe_to_numeric(series: pd.Series) -> pd.Series:
    # Robust to mixed formatting (commas/dots/spaces)
    try:
        s = series
        if s is None:
            return pd.Series(dtype="float64")
        if pd.api.types.is_numeric_dtype(s):
            return pd.to_numeric(s, errors="coerce")
        txt = s.astype(str).str.replace("\u00A0", "", regex=False).str.replace(" ", "", regex=False)
        # If both separators exist, treat last occurrence as decimal separator
        def _parse_one(v: str) -> Optional[float]:
            try:
                vv = str(v).strip()
                if vv == "" or vv.lower() == "nan":
                    return None
                if "," in vv and "." in vv:
                    last_c = vv.rfind(",")
                    last_d = vv.rfind(".")
                    if last_c > last_d:
                        vv = vv.replace(".", "").replace(",", ".")
                    else:
                        vv = vv.replace(",", "")
                else:
                    vv = vv.replace(",", ".")
                return float(vv)
            except Exception:
                return None

        parsed = txt.map(_parse_one)
        return pd.to_numeric(parsed, errors="coerce")
    except Exception:
        return pd.to_numeric(series, errors="coerce")


def _month_label_from_date(df: pd.DataFrame) -> pd.Series:
    # Required format: "January 2025"
    dt = pd.to_datetime(df[DATE_COL], errors="coerce")
    return dt.dt.strftime("%B %Y")


def _read_sheet(window_path: Path, sheet_name: str) -> Optional[pd.DataFrame]:
    try:
        return pd.read_excel(window_path, sheet_name=sheet_name)
    except Exception:
        return None


def _group_company_sum(df: pd.DataFrame, metric_col: str) -> pd.DataFrame:
    if df is None or df.empty:
        return pd.DataFrame(columns=[COMPANY_COL, metric_col])
    if COMPANY_COL not in df.columns:
        return pd.DataFrame(columns=[COMPANY_COL, metric_col])
    if metric_col not in df.columns:
        # Missing metric column → treat as zeros (no rows)
        out = df[[COMPANY_COL]].copy()
        out[metric_col] = 0.0
        return out.groupby(COMPANY_COL, dropna=False)[metric_col].sum().reset_index()
    tmp = df[[COMPANY_COL, metric_col]].copy()
    tmp[metric_col] = _safe_to_numeric(tmp[metric_col]).fillna(0.0)
    return tmp.groupby(COMPANY_COL, dropna=False)[metric_col].sum().reset_index()


def _group_month_company_sum(df: pd.DataFrame, metric_col: str) -> pd.DataFrame:
    if df is None or df.empty:
        return pd.DataFrame(columns=["Month", COMPANY_COL, metric_col])
    if COMPANY_COL not in df.columns or DATE_COL not in df.columns:
        return pd.DataFrame(columns=["Month", COMPANY_COL, metric_col])
    if metric_col not in df.columns:
        tmp = df[[DATE_COL, COMPANY_COL]].copy()
        tmp[metric_col] = 0.0
    else:
        tmp = df[[DATE_COL, COMPANY_COL, metric_col]].copy()
        tmp[metric_col] = _safe_to_numeric(tmp[metric_col]).fillna(0.0)
    tmp["Month"] = _month_label_from_date(tmp)
    tmp = tmp.dropna(subset=["Month"])
    return tmp.groupby(["Month", COMPANY_COL], dropna=False)[metric_col].sum().reset_index()


def _company_totals_table(window_path: Path, metric: _MetricSpec) -> pd.DataFrame:
    # Table #1: uses explicit list and excludes Water
    parts: List[pd.DataFrame] = []
    for sh in TOTALS_SHEETS_EXCLUDING_WATER:
        df = _read_sheet(window_path, sh)
        parts.append(_group_company_sum(df, metric.metric_col))
    combined = pd.concat(parts, ignore_index=True) if parts else pd.DataFrame(columns=[COMPANY_COL, metric.metric_col])
    if combined.empty:
        out = pd.DataFrame(columns=[COMPANY_COL, metric.output_metric_name, "Share (%)"])
        return out
    totals = combined.groupby(COMPANY_COL, dropna=False)[metric.metric_col].sum().reset_index()
    totals = totals.rename(columns={metric.metric_col: metric.output_metric_name})
    grand = _safe_to_numeric(totals[metric.output_metric_name]).sum()
    if grand and float(grand) != 0.0:
        totals["Share (%)"] = (_safe_to_numeric(totals[metric.output_metric_name]) / float(grand)) * 100.0
    else:
        totals["Share (%)"] = 0.0
    totals = totals.sort_values(metric.output_metric_name, ascending=False, kind="mergesort")
    return totals[[COMPANY_COL, metric.output_metric_name, "Share (%)"]]


def _company_by_ghgp_sheet_totals_table(window_path: Path, metric: _MetricSpec) -> pd.DataFrame:
    # Table #2: all GHGP sheets (as provided)
    rows: List[pd.DataFrame] = []
    for sh in ALL_GHGP_SHEETS:
        df = _read_sheet(window_path, sh)
        g = _group_company_sum(df, metric.metric_col)
        if g.empty:
            continue
        g = g.rename(columns={metric.metric_col: metric.output_metric_name})
        g.insert(0, "GHGP_Sheet", sh)
        rows.append(g)
    out = pd.concat(rows, ignore_index=True) if rows else pd.DataFrame(columns=["GHGP_Sheet", COMPANY_COL, metric.output_metric_name])
    if out.empty:
        out["Share in GHGP (%)"] = pd.Series(dtype="float64")
        out["Share in Company (%)"] = pd.Series(dtype="float64")
        return out[["GHGP_Sheet", COMPANY_COL, metric.output_metric_name, "Share in GHGP (%)", "Share in Company (%)"]]

    # Shares
    sheet_tot = out.groupby("GHGP_Sheet", dropna=False)[metric.output_metric_name].transform("sum")
    comp_tot = out.groupby(COMPANY_COL, dropna=False)[metric.output_metric_name].transform("sum")

    denom_sheet = sheet_tot.replace(0, pd.NA)
    denom_comp = comp_tot.replace(0, pd.NA)

    out["Share in GHGP (%)"] = (_safe_to_numeric(out[metric.output_metric_name]) / denom_sheet) * 100.0
    out["Share in Company (%)"] = (_safe_to_numeric(out[metric.output_metric_name]) / denom_comp) * 100.0
    # Ensure stable float dtype (avoid pandas FutureWarning on silent downcast)
    out["Share in GHGP (%)"] = pd.to_numeric(out["Share in GHGP (%)"], errors="coerce").fillna(0.0)
    out["Share in Company (%)"] = pd.to_numeric(out["Share in Company (%)"], errors="coerce").fillna(0.0)

    # Order: GHGP sheet order, then descending metric
    order_map = {name: i for i, name in enumerate(ALL_GHGP_SHEETS)}
    out["_ord"] = out["GHGP_Sheet"].map(order_map).fillna(9999).astype(int)
    out = out.sort_values(["_ord", metric.output_metric_name], ascending=[True, False], kind="mergesort").drop(columns=["_ord"])
    return out[["GHGP_Sheet", COMPANY_COL, metric.output_metric_name, "Share in GHGP (%)", "Share in Company (%)"]]


def _ghgp_sheet_totals_table(window_path: Path, metric: _MetricSpec) -> pd.DataFrame:
    # Table #3
    totals: List[Dict[str, object]] = []
    for sh in ALL_GHGP_SHEETS:
        df = _read_sheet(window_path, sh)
        if df is None or df.empty or metric.metric_col not in df.columns:
            total_val = 0.0
        else:
            total_val = float(_safe_to_numeric(df[metric.metric_col]).fillna(0.0).sum())
        totals.append({"GHGP_Sheet": sh, metric.output_metric_name: total_val})
    out = pd.DataFrame(totals)
    grand = float(_safe_to_numeric(out[metric.output_metric_name]).sum())
    if grand != 0.0:
        out["Share (%)"] = (_safe_to_numeric(out[metric.output_metric_name]) / grand) * 100.0
    else:
        out["Share (%)"] = 0.0
    return out[["GHGP_Sheet", metric.output_metric_name, "Share (%)"]]


def _company_stacked_table(window_path: Path, metric: _MetricSpec, include_water: bool) -> pd.DataFrame:
    # Table #4 (include_water=True) or base for #5 (include_water=False)
    ghgp_cols = [
        "S3 Cat 1 Purchased G&S",
        "S3 Cat 11 Use of Sold",
        "S3 Cat 12 End of Life",
        "S3 Cat 3 FERA",
        "S3 Cat 4 Upstream Transport",
        "S3 Cat 5 Waste",
        "S3 Cat 6 Business Travel",
        "S3 Cat 7 Employee Commute",
        "S3 Cat 9 Downstream Transport",
        "Scope 1",
        "Scope 2",
    ]
    if include_water:
        ghgp_cols.append("Water")

    # Build wide by merging company sums per sheet
    base: Optional[pd.DataFrame] = None
    for sh in ghgp_cols:
        df = _read_sheet(window_path, sh)
        g = _group_company_sum(df, metric.metric_col)
        g = g.rename(columns={metric.metric_col: sh})
        if base is None:
            base = g
        else:
            base = base.merge(g, on=COMPANY_COL, how="outer")

    if base is None:
        base = pd.DataFrame(columns=[COMPANY_COL] + ghgp_cols)

    for c in ghgp_cols:
        if c not in base.columns:
            base[c] = 0.0
        base[c] = _safe_to_numeric(base[c]).fillna(0.0)

    base["Row Total (t)"] = base[ghgp_cols].sum(axis=1)
    grand = float(_safe_to_numeric(base["Row Total (t)"]).sum())
    if grand != 0.0:
        base["Company Share in Total (%)"] = (_safe_to_numeric(base["Row Total (t)"]) / grand) * 100.0
    else:
        base["Company Share in Total (%)"] = 0.0

    base = base.sort_values("Row Total (t)", ascending=False, kind="mergesort")
    return base[[COMPANY_COL] + ghgp_cols + ["Row Total (t)", "Company Share in Total (%)"]]


def _company_stacked_by_month_table(window_path: Path, metric: _MetricSpec) -> pd.DataFrame:
    # Table #5 (no Water column)
    ghgp_cols = [
        "S3 Cat 1 Purchased G&S",
        "S3 Cat 11 Use of Sold",
        "S3 Cat 12 End of Life",
        "S3 Cat 3 FERA",
        "S3 Cat 4 Upstream Transport",
        "S3 Cat 5 Waste",
        "S3 Cat 6 Business Travel",
        "S3 Cat 7 Employee Commute",
        "S3 Cat 9 Downstream Transport",
        "Scope 1",
        "Scope 2",
    ]

    base: Optional[pd.DataFrame] = None
    for sh in ghgp_cols:
        df = _read_sheet(window_path, sh)
        g = _group_month_company_sum(df, metric.metric_col)
        g = g.rename(columns={metric.metric_col: sh})
        if base is None:
            base = g
        else:
            base = base.merge(g, on=["Month", COMPANY_COL], how="outer")

    if base is None:
        base = pd.DataFrame(columns=["Month", COMPANY_COL] + ghgp_cols)

    for c in ghgp_cols:
        if c not in base.columns:
            base[c] = 0.0
        base[c] = _safe_to_numeric(base[c]).fillna(0.0)

    base["Row Total (t)"] = base[ghgp_cols].sum(axis=1)

    grand = float(_safe_to_numeric(base["Row Total (t)"]).sum())
    if grand != 0.0:
        base["Company Share in Total (%)"] = (_safe_to_numeric(base["Row Total (t)"]) / grand) * 100.0
    else:
        base["Company Share in Total (%)"] = 0.0

    month_tot = base.groupby("Month", dropna=False)["Row Total (t)"].transform("sum").replace(0, pd.NA)
    base["Company Share in Month (%)"] = (_safe_to_numeric(base["Row Total (t)"]) / month_tot) * 100.0
    base["Company Share in Month (%)"] = base["Company Share in Month (%)"].fillna(0.0)

    # Sort months chronologically when possible
    try:
        sort_key = pd.to_datetime(base["Month"], format="%B %Y", errors="coerce")
        base["_m_sort"] = sort_key
        base = base.sort_values(["_m_sort", "Month", "Row Total (t)"], ascending=[True, True, False], kind="mergesort")
        base = base.drop(columns=["_m_sort"])
    except Exception:
        base = base.sort_values(["Month", "Row Total (t)"], ascending=[True, False], kind="mergesort")

    return base[
        ["Month", COMPANY_COL]
        + ghgp_cols
        + ["Row Total (t)", "Company Share in Total (%)", "Company Share in Month (%)"]
    ]


def _pick_window_workbook(window_path: Optional[Path]) -> Optional[Path]:
    if window_path is not None and window_path.exists():
        return window_path
    out_dir = STAGE2_OUTPUT_DIR
    try:
        candidates = [
            p
            for p in out_dir.rglob("mapped_results_window_*.xlsx")
            if p.is_file() and (not p.name.startswith("~$"))
        ]
        if not candidates:
            return None
        candidates.sort(key=lambda p: p.stat().st_mtime, reverse=True)
        return candidates[0]
    except Exception:
        return None


def build_and_write_tables(window_path: Path) -> Path:
    """
    In-place update:
    - Removes the old (incorrect) totals sheets from the window workbook.
    - Writes 5 new window totals sheets for co2e (t).
    - Writes 5 new window totals sheets for Spend_Euro.
    """
    import datetime as _dt
    import shutil

    window_path = Path(window_path)
    if not window_path.exists():
        raise FileNotFoundError(f"Window workbook not found: {window_path}")

    # If the file is locked (open in Excel), fall back to writing a copy.
    target_path = window_path
    try:
        # Quick write-test on same directory (no real changes)
        _ = target_path.stat()
    except Exception:
        pass

    try:
        from openpyxl import load_workbook
        from openpyxl.styles import PatternFill, Font, Border, Side
        from openpyxl.utils import get_column_letter
        from openpyxl.formatting.rule import FormulaRule
    except Exception as exc:  # pragma: no cover
        raise RuntimeError("openpyxl is required to modify Excel files.") from exc

    def _style_totals_sheet(path: Path, sheet_name: str) -> None:
        """Apply green bold header + zebra rows like other outputs."""
        wb_local = load_workbook(path)
        if sheet_name not in wb_local.sheetnames:
            wb_local.close()
            return
        ws = wb_local[sheet_name]

        max_row = ws.max_row or 1
        max_col = ws.max_column or 1

        # NOTE: openpyxl expects ARGB hex; use FF-prefixed colors
        header_fill = PatternFill("solid", fgColor="FFD8EAD3")  # light green
        header_font = Font(bold=True)
        zebra_fill = PatternFill("solid", fgColor="FFF2F2F2")  # light gray
        thin = Side(style="thin", color="000000")
        border = Border(left=thin, right=thin, top=thin, bottom=thin)

        # Header formatting (row 1)
        for c in range(1, max_col + 1):
            cell = ws.cell(row=1, column=c)
            cell.fill = header_fill
            cell.font = header_font
            cell.border = border

        # Freeze header row
        try:
            ws.freeze_panes = ws["A2"]
        except Exception:
            ws.freeze_panes = "A2"

        # Zebra striping: apply to data range only
        if max_row >= 2 and max_col >= 1:
            last_col = get_column_letter(max_col)
            rng = f"A2:{last_col}{max_row}"
            # Use Excel row() to alternate
            ws.conditional_formatting.add(
                rng,
                FormulaRule(formula=["MOD(ROW(),2)=0"], fill=zebra_fill),
            )

        wb_local.save(path)
        wb_local.close()

    try:
        wb = load_workbook(target_path)
    except PermissionError:
        ts = _dt.datetime.now().strftime("%Y%m%d_%H%M%S")
        copy_path = target_path.with_name(f"{target_path.stem}_window_totals_{ts}{target_path.suffix}")
        shutil.copy2(target_path, copy_path)
        target_path = copy_path
        wb = load_workbook(target_path)

    # Remove old sheets if present (and any previously written Window totals sheets)
    for name in OLD_SHEETS_TO_REMOVE:
        if name in wb.sheetnames:
            ws = wb[name]
            wb.remove(ws)
    for desired, actual in SHEET_NAME_MAP.items():
        if actual in wb.sheetnames:
            ws = wb[actual]
            wb.remove(ws)
    # Saving can fail if the file is open/locked in Excel.
    # In that case, write to a new timestamped copy and continue from there.
    try:
        wb.save(target_path)
    except PermissionError:
        ts = _dt.datetime.now().strftime("%Y%m%d_%H%M%S")
        fallback_path = target_path.with_name(f"{target_path.stem}_window_totals_{ts}{target_path.suffix}")
        wb.save(fallback_path)
        target_path = fallback_path
    wb.close()

    # Build tables (co2e)
    co2 = _MetricSpec(metric_col=CO2E_COL, output_metric_name=CO2E_COL)
    spend = _MetricSpec(metric_col=SPEND_COL, output_metric_name=SPEND_COL)

    tables: List[Tuple[str, pd.DataFrame]] = []

    tables.append((SHEET_NAME_MAP["Company Totals Window"], _company_totals_table(target_path, co2)))
    tables.append((SHEET_NAME_MAP["Company by GHGP Sheet Totals Window"], _company_by_ghgp_sheet_totals_table(target_path, co2)))
    tables.append((SHEET_NAME_MAP["GHGP sheet Totals Window"], _ghgp_sheet_totals_table(target_path, co2)))
    tables.append((SHEET_NAME_MAP["Company Stacked Data Window"], _company_stacked_table(target_path, co2, include_water=True)))
    tables.append((SHEET_NAME_MAP["Company Stacked Data by Months Window"], _company_stacked_by_month_table(target_path, co2)))

    # Spend versions (sheet names not provided by user; use explicit suffix to avoid collisions)
    tables.append((SHEET_NAME_MAP["Company Totals Window Spend"], _company_totals_table(target_path, spend)))
    tables.append((SHEET_NAME_MAP["Company by GHGP Sheet Totals Window Spend"], _company_by_ghgp_sheet_totals_table(target_path, spend)))
    tables.append((SHEET_NAME_MAP["GHGP sheet Totals Window Spend"], _ghgp_sheet_totals_table(target_path, spend)))
    tables.append((SHEET_NAME_MAP["Company Stacked Data Window Spend"], _company_stacked_table(target_path, spend, include_water=True)))
    tables.append((SHEET_NAME_MAP["Company Stacked Data by Months Window Spend"], _company_stacked_by_month_table(target_path, spend)))

    # Write (append at end)
    with pd.ExcelWriter(target_path, engine="openpyxl", mode="a", if_sheet_exists="replace") as writer:
        for sheet_name, df in tables:
            df.to_excel(writer, sheet_name=sheet_name, index=False)

    # Apply styling to the newly written totals sheets
    for sheet_name, _df in tables:
        try:
            _style_totals_sheet(target_path, sheet_name)
        except Exception:
            # best-effort formatting; do not block the pipeline
            pass

    return target_path


def main(window_path: Optional[str] = None) -> Optional[Path]:
    """
    Entry point used by Run_Everything.py.
    - If `window_path` is provided, uses it.
    - Else picks the newest `mapped_results_window_*.xlsx` under output/ (recursive).
    """
    wp = _pick_window_workbook(Path(window_path) if window_path else None)
    if wp is None:
        print("[WARN] Window_total_tables: No window workbook found; skipping.")
        return None
    out = build_and_write_tables(wp)
    print(f"[info] Window_total_tables: Updated window workbook -> {out.name}")
    return out

