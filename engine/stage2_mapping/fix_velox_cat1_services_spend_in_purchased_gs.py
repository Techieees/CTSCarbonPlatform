from __future__ import annotations

import argparse
import math
import shutil
from datetime import datetime
from pathlib import Path
from typing import Optional, Tuple

from openpyxl import load_workbook


def _norm(s: object) -> str:
    try:
        return str(s).strip().lower()
    except Exception:
        return ""


def _to_float(v: object) -> Optional[float]:
    if v is None:
        return None
    if isinstance(v, (int, float)):
        try:
            fv = float(v)
            return fv if math.isfinite(fv) else None
        except Exception:
            return None
    s = str(v).strip()
    if s == "":
        return None
    s = s.replace("\u00A0", "").replace(" ", "")
    # handle "1 234,56" / "1234,56" / "1,234.56"
    if "," in s and "." in s:
        # assume commas are thousands separators
        s = s.replace(",", "")
    elif "," in s and "." not in s:
        # assume comma is decimal separator
        s = s.replace(",", ".")
    try:
        fv = float(s)
        return fv if math.isfinite(fv) else None
    except Exception:
        return None


def _find_col_indices(headers: list[object]) -> Tuple[int, int, int, int, int, int]:
    """Return 1-based indices for (Company, Sheet_booklets, Spend_Euro, ef_value, ef_unit, co2e (t))."""
    idx_company = idx_booklets = idx_spend = idx_ef = idx_unit = idx_co2e_t = -1
    for i, h in enumerate(headers, start=1):
        key = _norm(h)
        if key == "company":
            idx_company = i
        elif key == "sheet_booklets":
            idx_booklets = i
        elif key in {"spend_euro", "spend euro", "spend (euro)", "spend eur"}:
            if idx_spend == -1:
                idx_spend = i
        elif key in {"ef_value", "ef value"}:
            idx_ef = i
        elif key in {"ef_unit", "ef unit"}:
            idx_unit = i
        elif key == "co2e (t)":
            idx_co2e_t = i
    if -1 in {idx_company, idx_booklets, idx_spend, idx_ef, idx_co2e_t}:
        raise RuntimeError(
            "Required columns not found. "
            f"company={idx_company}, sheet_booklets={idx_booklets}, spend={idx_spend}, "
            f"ef_value={idx_ef}, ef_unit={idx_unit}, co2e(t)={idx_co2e_t}"
        )
    return idx_company, idx_booklets, idx_spend, idx_ef, idx_unit, idx_co2e_t


def _compute_tco2e(spend_eur: Optional[float], ef_value: Optional[float], ef_unit: Optional[str]) -> Optional[float]:
    if spend_eur is None or ef_value is None:
        return None
    try:
        s = float(spend_eur)
        ef = float(ef_value)
    except Exception:
        return None
    u = (ef_unit or "").strip().lower()
    # If EF is kgCO2e/EUR then convert to tonnes
    if ("kg" in u) and ("eur" in u or "€" in u) and ("/" in u or "per" in u):
        return (s * ef) / 1000.0
    return s * ef


def apply_fix(
    input_path: Path,
    output_path: Path,
    sheet_name: str = "S3 Cat 1 Purchased G&S",
    company: str = "Velox",
    sheet_booklets_value: str = "Scope 3 Services Spend",
) -> int:
    # data_only=False to preserve formulas elsewhere
    wb = load_workbook(input_path, data_only=False)
    if sheet_name not in wb.sheetnames:
        raise RuntimeError(f"Sheet not found: {sheet_name}")
    ws = wb[sheet_name]

    headers = [c.value for c in ws[1]]
    idx_company, idx_booklets, idx_spend, idx_ef, idx_unit, idx_co2e_t = _find_col_indices(headers)

    updated = 0
    for r in range(2, ws.max_row + 1):
        if _norm(ws.cell(row=r, column=idx_company).value) != _norm(company):
            continue
        if _norm(ws.cell(row=r, column=idx_booklets).value) != _norm(sheet_booklets_value):
            continue

        current = _to_float(ws.cell(row=r, column=idx_co2e_t).value)
        # Only fix missing/zero values to avoid overwriting valid results
        if current is not None and current != 0.0:
            continue

        spend = _to_float(ws.cell(row=r, column=idx_spend).value)
        efv = _to_float(ws.cell(row=r, column=idx_ef).value)
        ef_unit = ws.cell(row=r, column=idx_unit).value if idx_unit != -1 else None

        out = _compute_tco2e(spend, efv, str(ef_unit) if ef_unit is not None else None)
        if out is None:
            continue
        ws.cell(row=r, column=idx_co2e_t).value = out
        updated += 1

    output_path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(output_path)
    return updated


def main() -> None:
    ap = argparse.ArgumentParser(
        description="Fix Velox rows in S3 Cat 1 Purchased G&S where Sheet_booklets=Scope 3 Services Spend by setting co2e (t)=Spend_Euro*ef_value."
    )
    ap.add_argument("--input", required=True, help="Path to mapped_results_window workbook (.xlsx)")
    ap.add_argument("--output", help="Path to write updated workbook (.xlsx). If omitted, edits in place.")
    ap.add_argument("--sheet", default="S3 Cat 1 Purchased G&S", help="Sheet name to modify")
    ap.add_argument("--company", default="Velox", help="Company value to target (default: Velox)")
    ap.add_argument(
        "--sheet-booklets",
        default="Scope 3 Services Spend",
        help="Sheet_booklets value to target (default: Scope 3 Services Spend)",
    )
    ap.add_argument("--no-backup", action="store_true", help="Do not create a backup before editing in-place.")
    args = ap.parse_args()

    inp = Path(args.input)
    if not inp.exists():
        raise SystemExit(f"Input not found: {inp}")

    out = Path(args.output) if args.output else inp

    if (out.resolve() == inp.resolve()) and (not args.no_backup):
        backup_path = inp.with_name(
            f"{inp.stem}_BACKUP_before_velox_cat1_services_fix_{datetime.now().strftime('%Y%m%d_%H%M%S')}{inp.suffix}"
        )
        shutil.copy2(inp, backup_path)
        print(f"Backup created: {backup_path}")

    try:
        n = apply_fix(
            inp,
            out,
            sheet_name=args.sheet,
            company=args.company,
            sheet_booklets_value=args.sheet_booklets,
        )
        print(f"Updated rows: {n}")
        print(f"Wrote: {out}")
    except PermissionError as e:
        fallback = inp.with_name(f"{inp.stem}_VELOX_CAT1_SERVICES_FIX{inp.suffix}")
        n = apply_fix(
            inp,
            fallback,
            sheet_name=args.sheet,
            company=args.company,
            sheet_booklets_value=args.sheet_booklets,
        )
        print(f"[WARN] Could not overwrite (file locked): {e}")
        print(f"Updated rows: {n}")
        print(f"Wrote fallback: {fallback}")


if __name__ == "__main__":
    main()

