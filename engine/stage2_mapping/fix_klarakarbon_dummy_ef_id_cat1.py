from __future__ import annotations

import argparse
import shutil
from datetime import datetime
from pathlib import Path
from typing import Dict, Iterable, Optional, Tuple
import unicodedata
import re
import sys

import pandas as pd
from openpyxl import load_workbook

PROJECT_ROOT = Path(__file__).resolve().parents[2]
if str(PROJECT_ROOT) not in sys.path:
    sys.path.insert(0, str(PROJECT_ROOT))

from config import STAGE2_EF_XLSX


# EF database defaults
DEFAULT_EF_WORKBOOK = STAGE2_EF_XLSX
KLARAKARBON_EF_SHEET = "Klarakarbon Emission Factors"


def _norm(s: object) -> str:
    try:
        return str(s).strip().lower()
    except Exception:
        return ""


def _norm_key(s: object) -> Optional[str]:
    if s is None:
        return None
    try:
        v = _normalize_name(s)
    except Exception:
        return None
    if v == "":
        return None
    return v


_ZW_CHARS = {
    "\u200b",  # zero width space
    "\u200c",  # zero width non-joiner
    "\u200d",  # zero width joiner
    "\ufeff",  # BOM
}


def _normalize_name(s: object) -> str:
    """
    Canonicalize names deterministically to make "same visible text" match reliably.
    This is NOT fuzzy matching: after normalization we still require exact equality.
    """
    if s is None:
        return ""
    v = unicodedata.normalize("NFKC", str(s))
    # Remove zero-width/invisible chars
    for ch in _ZW_CHARS:
        v = v.replace(ch, "")
    # Normalize common dash variants
    v = (
        v.replace("\u2010", "-")
        .replace("\u2011", "-")
        .replace("\u2012", "-")
        .replace("\u2013", "-")
        .replace("\u2014", "-")
        .replace("\u2212", "-")
    )
    # Normalize whitespace & slash spacing
    v = re.sub(r"\s+", " ", v)
    v = re.sub(r"\s*/\s*", " / ", v)
    v = v.strip()
    return v.casefold()


def _build_klarakarbon_lookup(
    ef_workbook: Path,
) -> Tuple[Dict[str, Tuple[str, str]], Dict[str, pd.DataFrame]]:
    """
    Build mapping from EF name (lowercased) -> (ef_id, ef_name) for Klarakarbon factors.

    Source:
    - 'Klarakarbon Emission Factors' sheet in EF workbook.

    Deterministic rule:
    - Exact match on normalised ef_name (case-insensitive, trimmed).
    - If the same ef_name maps to multiple different ef_id values in the Klarakarbon sheet,
      treat it as ambiguous and DO NOT include it (so we never guess).
    """
    if not ef_workbook.exists():
        raise RuntimeError(f"EF workbook not found: {ef_workbook}")

    try:
        df = pd.read_excel(ef_workbook, sheet_name=KLARAKARBON_EF_SHEET, usecols=["ef_name", "ef_id"])
    except Exception as e:
        raise RuntimeError(
            f"Could not read required columns from sheet '{KLARAKARBON_EF_SHEET}' in EF workbook: {ef_workbook}. "
            f"Expected columns: ef_name, ef_id. Original error: {e}"
        )

    if df is None or df.empty:
        return {}, {}

    # Keep extra columns if present for reporting (not required for mapping).
    extra_cols = [c for c in ["scope", "ef_category", "ef_unit", "ef_source", "Emission Factor Category"] if c in df.columns]
    df = df[["ef_name", "ef_id", *extra_cols]].copy()

    def _clean_id(x: object) -> str:
        if x is None:
            return ""
        try:
            s = str(x).strip()
        except Exception:
            return ""
        return "" if s.lower() == "nan" else s

    df["__key"] = df["ef_name"].map(_normalize_name)
    df["__ef_id"] = df["ef_id"].map(_clean_id)
    df["__ef_name_disp"] = df["ef_name"].astype(str).map(lambda s: str(s).strip())

    # Valid rows for mapping require non-empty ef_id and non-empty key
    df_valid = df[(df["__key"] != "") & (df["__ef_id"] != "")].copy()

    out: Dict[str, Tuple[str, str]] = {}
    ambiguous_rows: Dict[str, pd.DataFrame] = {}

    # Group by key and detect ambiguity deterministically
    for k, g in df_valid.groupby("__key", dropna=False):
        ids = sorted(set(g["__ef_id"].tolist()))
        if len(ids) == 1:
            # Unambiguous: pick the first display name for ef_name fill
            first = g.iloc[0]
            out[k] = (ids[0], str(first["__ef_name_disp"]).strip())
        else:
            # Ambiguous: keep the candidate rows for reporting
            show_cols = ["ef_name", "ef_id", *extra_cols]
            ambiguous_rows[str(k)] = g[show_cols].copy()

    return out, ambiguous_rows

    out: Dict[str, Tuple[str, str]] = {}
    seen_ids: Dict[str, set[str]] = {}
    for _, r in df.iterrows():
        k = _norm_key(r.get("ef_name"))
        efid = r.get("ef_id")
        if k is None:
            continue
        if efid is None or str(efid).strip() == "":
            continue
        efid_s = str(efid).strip()
        if k not in seen_ids:
            seen_ids[k] = set()
        seen_ids[k].add(efid_s)
        if k not in out:
            out[k] = (efid_s, str(r.get("ef_name")).strip())

    ambiguous = {k for k, ids in seen_ids.items() if len(ids) > 1}
    for k in ambiguous:
        out.pop(k, None)
    return out


def _find_col_indices(headers: list[object]) -> Tuple[int, int, int, int]:
    """
    Return 1-based indices for:
    - source column: Sheet_booklets OR Data Source sheet
    - ef_id
    - ef_name
    - emission factor name (input factor name for Klarakarbon rows)
    """
    idx_src = idx_efid = idx_efname = idx_emf_name = -1
    for i, h in enumerate(headers, start=1):
        key = _norm(h)
        if key in {"sheet_booklets", "data source sheet"} and idx_src == -1:
            idx_src = i
        elif key in {"ef_id", "ef id"} and idx_efid == -1:
            idx_efid = i
        elif key in {"ef_name", "ef name"} and idx_efname == -1:
            idx_efname = i
        elif key == "emission factor name" and idx_emf_name == -1:
            idx_emf_name = i
    if -1 in {idx_src, idx_efid, idx_efname, idx_emf_name}:
        raise RuntimeError(
            "Required columns not found. "
            f"source={idx_src}, ef_id={idx_efid}, ef_name={idx_efname}, emission factor name={idx_emf_name}"
        )
    return idx_src, idx_efid, idx_efname, idx_emf_name


def _is_empty(v: object) -> bool:
    if v is None:
        return True
    try:
        return str(v).strip() == ""
    except Exception:
        return True


def apply_fix(
    input_path: Path,
    output_path: Path,
    sheet_name: str = "S3 Cat 1 Purchased G&S",
    ef_workbook: Optional[Path] = None,
) -> Tuple[int, int]:
    """
    Applies (NO calculations):
    - STEP 1 (deterministic, no fallback):
      For Klarakarbon rows, fill missing ef_id (+ ef_name if missing) by matching window 'emission factor name'
      against EF DB Klarakarbon sheet 'Klarakarbon Emission Factors' column 'ef_name' (exact match after trim+lower).

    Does NOT modify existing ef_id values (only fills missing).
    Does NOT modify Travel rows.

    Returns: (n_filled_ef_id, n_dummy_set)  # dummy is always 0 in this script now
    """
    ef_path = ef_workbook or DEFAULT_EF_WORKBOOK
    ef_path = Path(ef_path)
    kbk_lookup, kbk_ambiguous = _build_klarakarbon_lookup(ef_path)

    wb = load_workbook(input_path, data_only=False)
    if sheet_name not in wb.sheetnames:
        raise RuntimeError(f"Sheet not found: {sheet_name}")
    ws = wb[sheet_name]

    headers = [c.value for c in ws[1]]
    idx_src, idx_efid, idx_efname, idx_emf_name = _find_col_indices(headers)

    filled_efid = 0
    dummy_set = 0  # kept for backward compatibility with earlier pipeline prints
    not_found_counts: Dict[str, int] = {}
    ambiguous_counts: Dict[str, int] = {}

    for r in range(2, ws.max_row + 1):
        src = ws.cell(row=r, column=idx_src).value
        src_norm = _norm(src)

        # Do not modify Travel rows at all
        if "travel" in src_norm:
            continue

        # STEP 1: Only for Klarakarbon rows; use 'emission factor name' to match EF database
        if "klarakarbon" in src_norm:
            cur_id = ws.cell(row=r, column=idx_efid).value
            if _is_empty(cur_id):
                emf_name = ws.cell(row=r, column=idx_emf_name).value
                k = _norm_key(emf_name)
                hit = kbk_lookup.get(k) if k is not None else None
                if hit:
                    ws.cell(row=r, column=idx_efid).value = hit[0]
                    filled_efid += 1
                    # Fill ef_name too if missing
                    if _is_empty(ws.cell(row=r, column=idx_efname).value):
                        ws.cell(row=r, column=idx_efname).value = hit[1]
                else:
                    raw = ""
                    try:
                        raw = str(emf_name).strip()
                    except Exception:
                        raw = ""
                    if k is not None and k in kbk_ambiguous:
                        ambiguous_counts[raw] = ambiguous_counts.get(raw, 0) + 1
                    else:
                        not_found_counts[raw] = not_found_counts.get(raw, 0) + 1

    output_path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(output_path)

    # Write deterministic reports for any remaining missing mappings
    try:
        if not_found_counts:
            nf = (
                pd.DataFrame(
                    [{"emission factor name": k, "row_count": v} for k, v in not_found_counts.items() if str(k).strip() != ""],
                )
                .sort_values("row_count", ascending=False)
                .reset_index(drop=True)
            )
            nf_path = output_path.with_name(f"{output_path.stem}_KBK_NOT_FOUND.csv")
            nf.to_csv(nf_path, index=False, encoding="utf-8-sig")
            print(f"[info] Wrote not-found report: {nf_path}")

        if ambiguous_counts:
            amb = (
                pd.DataFrame(
                    [{"emission factor name": k, "row_count": v} for k, v in ambiguous_counts.items() if str(k).strip() != ""],
                )
                .sort_values("row_count", ascending=False)
                .reset_index(drop=True)
            )
            amb_path = output_path.with_name(f"{output_path.stem}_KBK_AMBIGUOUS_NAMES.csv")
            amb.to_csv(amb_path, index=False, encoding="utf-8-sig")
            print(f"[info] Wrote ambiguous-names report: {amb_path}")

            # Candidate rows for each ambiguous key
            cand_rows = []
            for raw_name in amb["emission factor name"].tolist():
                key = _normalize_name(raw_name)
                g = kbk_ambiguous.get(key)
                if g is None or g.empty:
                    continue
                gg = g.copy()
                gg.insert(0, "emission factor name (window)", raw_name)
                cand_rows.append(gg)
            if cand_rows:
                cand = pd.concat(cand_rows, ignore_index=True)
                cand_path = output_path.with_name(f"{output_path.stem}_KBK_AMBIGUOUS_CANDIDATES.csv")
                cand.to_csv(cand_path, index=False, encoding="utf-8-sig")
                print(f"[info] Wrote ambiguous-candidates report: {cand_path}")
    except Exception as e:
        print(f"[warn] Could not write KBK mapping reports: {e}")

    return filled_efid, dummy_set


def main() -> None:
    ap = argparse.ArgumentParser(
        description="Post-fix for window workbook: fill missing Klarakarbon ef_id/ef_name from EF DB sheet 'Klarakarbon Emission Factors' in S3 Cat 1 Purchased G&S (no fallback, no calculations)."
    )
    ap.add_argument("--input", required=True, help="Path to mapped_results_window workbook (.xlsx)")
    ap.add_argument("--output", help="Path to write updated workbook (.xlsx). If omitted, edits in place.")
    ap.add_argument("--sheet", default="S3 Cat 1 Purchased G&S", help="Sheet name to modify")
    ap.add_argument(
        "--ef-workbook",
        default=str(DEFAULT_EF_WORKBOOK),
        help="Path to EF database workbook (default: merkezi config'teki EF workbook yolu).",
    )
    ap.add_argument("--no-backup", action="store_true", help="Do not create a backup before editing in-place.")
    args = ap.parse_args()

    inp = Path(args.input)
    if not inp.exists():
        raise SystemExit(f"Input not found: {inp}")
    out = Path(args.output) if args.output else inp

    if (out.resolve() == inp.resolve()) and (not args.no_backup):
        backup_path = inp.with_name(
            f"{inp.stem}_BACKUP_before_klarakarbon_dummy_efid_fix_{datetime.now().strftime('%Y%m%d_%H%M%S')}{inp.suffix}"
        )
        shutil.copy2(inp, backup_path)
        print(f"Backup created: {backup_path}")

    try:
        n_efid, n_dummy = apply_fix(
            inp,
            out,
            sheet_name=args.sheet,
            ef_workbook=Path(args.ef_workbook),
        )
        print(f"Filled ef_id (Klarakarbon): {n_efid}")
        print(f"Wrote: {out}")
    except PermissionError as e:
        fallback = inp.with_name(f"{inp.stem}_KLARAKARBON_DUMMY_EFID_FIX{inp.suffix}")
        n_efid, n_dummy = apply_fix(
            inp,
            fallback,
            sheet_name=args.sheet,
            ef_workbook=Path(args.ef_workbook),
        )
        print(f"[WARN] Could not overwrite (file locked): {e}")
        print(f"Filled ef_id (Klarakarbon): {n_efid}")
        print(f"Wrote fallback: {fallback}")


if __name__ == "__main__":
    main()

