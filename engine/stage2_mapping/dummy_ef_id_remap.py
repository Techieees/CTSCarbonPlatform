from __future__ import annotations

import argparse
import re
import unicodedata
from pathlib import Path
from typing import Dict, Optional, Tuple

import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Border, Font, PatternFill, Side


_ZW_CHARS = {"\u200b", "\u200c", "\u200d", "\ufeff"}


def normalize_name(s: object) -> str:
    """
    Deterministic normalization (NOT fuzzy matching).
    We still require exact equality after this normalization.
    """
    if s is None:
        return ""
    v = unicodedata.normalize("NFKC", str(s))
    for ch in _ZW_CHARS:
        v = v.replace(ch, "")
    v = (
        v.replace("\u2010", "-")
        .replace("\u2011", "-")
        .replace("\u2012", "-")
        .replace("\u2013", "-")
        .replace("\u2014", "-")
        .replace("\u2212", "-")
    )
    v = re.sub(r"\s+", " ", v).strip()
    return v.casefold()


def _pick_col(cols: list[object], want: set[str]) -> Optional[str]:
    for c in cols:
        if str(c).strip().lower() in want:
            return str(c)
    return None


def _fill_for_source_value(src_value: object) -> PatternFill:
    """
    Auditor color-coding based on Sheet_booklets value:
    - Klarakarbon: light blue
    - Travel: grey
    - Transportation %10: yellow
    - Everything else (booklets): light orange
    """
    src_l = str(src_value).strip().lower() if src_value is not None else ""
    if src_l == "klarakarbon":
        return PatternFill("solid", fgColor="CFE2F3")  # light blue
    if src_l == "travel":
        return PatternFill("solid", fgColor="D9D9D9")  # grey
    if src_l == "transportation %10":
        return PatternFill("solid", fgColor="FFD966")  # yellow
    # More vivid orange than the previous very light shade
    return PatternFill("solid", fgColor="F4B084")  # orange


def apply_source_row_coloring(*, wb, source_col_name: str = "Sheet_booklets") -> int:
    """
    Apply the source color-coding to ALL sheets in a workbook that contain the given column.
    Returns the number of rows colored across all matching sheets.
    """
    colored_rows = 0
    want = str(source_col_name).strip().lower()
    thin = Side(style="thin", color="000000")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    for ws in wb.worksheets:
        try:
            headers = [c.value for c in ws[1]]
        except Exception:
            continue

        idx_src = -1
        for i, h in enumerate(headers, start=1):
            key = str(h).strip().lower() if h is not None else ""
            if key == want:
                idx_src = i
                break
        if idx_src == -1:
            continue

        last_col = ws.max_column
        max_row = ws.max_row

        # Header row: ensure borders exist (do not change header fill)
        try:
            for c in range(1, last_col + 1):
                ws.cell(row=1, column=c).border = border
        except Exception:
            pass

        for r in range(2, max_row + 1):
            fill = _fill_for_source_value(ws.cell(row=r, column=idx_src).value)
            for c in range(1, last_col + 1):
                cell = ws.cell(row=r, column=c)
                cell.fill = fill
                cell.border = border
                try:
                    cell.font = cell.font.copy(color="000000")
                except Exception:
                    cell.font = Font(color="000000")
            colored_rows += 1

    return colored_rows


def generate_mapping_from_unique_lists(
    kbk_csv: Path,
    non_kbk_csv: Path,
    out_mapping_csv: Path,
    out_summary_csv: Path,
) -> Tuple[Path, Path]:
    kbk = pd.read_csv(kbk_csv)
    non = pd.read_csv(non_kbk_csv)
    kbk["source_group"] = "klarakarbon"
    non["source_group"] = "non_klarakarbon"
    df = pd.concat([kbk, non], ignore_index=True)

    for c in ["ef_id", "ef_name", "row_count", "source_group"]:
        if c not in df.columns:
            raise RuntimeError(f"Missing required column '{c}' in inputs")

    df["key"] = df["ef_name"].map(normalize_name)
    keys = sorted([k for k in df["key"].unique().tolist() if isinstance(k, str) and k != ""])
    key_to_dummy = {k: f"DUMMY_{i:05d}" for i, k in enumerate(keys, start=1)}
    df["dummy_ef_id"] = df["key"].map(key_to_dummy)

    mapping = df[["source_group", "ef_id", "ef_name", "key", "dummy_ef_id", "row_count"]].copy()
    mapping = mapping.sort_values(["dummy_ef_id", "source_group", "ef_id"])
    out_mapping_csv.parent.mkdir(parents=True, exist_ok=True)
    mapping.to_csv(out_mapping_csv, index=False, encoding="utf-8-sig")

    summary = (
        mapping.groupby(["dummy_ef_id", "key"])
        .agg(
            total_rows=("row_count", "sum"),
            distinct_ef_id=("ef_id", lambda s: s.nunique()),
            distinct_ef_name=("ef_name", lambda s: s.nunique()),
            sources=("source_group", lambda s: " | ".join(sorted(set(s.tolist())))),
        )
        .reset_index()
        .sort_values(["total_rows"], ascending=False)
    )
    summary.to_csv(out_summary_csv, index=False, encoding="utf-8-sig")
    return out_mapping_csv, out_summary_csv


def apply_dummy_mapping(
    *,
    workbook_in: Path,
    workbook_out: Path,
    mapping_csv: Path,
    sheet_name: str = "S3 Cat 1 Purchased G&S",
    source_col_name: str = "Sheet_booklets",
    ef_id_col_name: str = "ef_id",
    ef_name_col_name: str = "ef_name",
    dummy_col_name: str = "dummy_ef_id",
    dummy_name_col_name: str = "dummy_ef_id_name",
) -> None:
    mapping = pd.read_csv(mapping_csv)
    if not {"key", "dummy_ef_id", "ef_name", "row_count"}.issubset(set(mapping.columns)):
        raise RuntimeError(
            f"mapping_csv missing required columns. Got: {sorted(mapping.columns.tolist())}"
        )

    key_to_dummy: Dict[str, str] = dict(
        zip(mapping["key"].astype(str).tolist(), mapping["dummy_ef_id"].astype(str).tolist())
    )

    # Canonical name per key: highest row_count wins (deterministic).
    tmp = mapping.copy()
    tmp["ef_name"] = tmp["ef_name"].astype(str).map(lambda s: str(s).strip())
    tmp["row_count"] = pd.to_numeric(tmp["row_count"], errors="coerce").fillna(0).astype(int)
    tmp = tmp.sort_values(["key", "row_count", "ef_name"], ascending=[True, False, True])
    key_to_name: Dict[str, str] = {}
    for k, g in tmp.groupby("key", dropna=False):
        k = str(k)
        if k == "" or k.lower() == "nan":
            continue
        key_to_name[k] = str(g.iloc[0]["ef_name"]).strip()

    # next dummy id if we need to create new keys on the fly
    nums = []
    for v in key_to_dummy.values():
        try:
            if str(v).startswith("DUMMY_"):
                nums.append(int(str(v).split("_")[1]))
        except Exception:
            pass
    next_id = (max(nums) + 1) if nums else 1

    wb = load_workbook(workbook_in, data_only=False)
    if sheet_name not in wb.sheetnames:
        raise RuntimeError(f"Sheet not found: {sheet_name}")
    ws = wb[sheet_name]

    headers = [c.value for c in ws[1]]
    idx_src = idx_efid = idx_efname = idx_dummy = idx_dummy_name = -1
    for i, h in enumerate(headers, start=1):
        key = str(h).strip().lower() if h is not None else ""
        if key == source_col_name.strip().lower() and idx_src == -1:
            idx_src = i
        elif key in {ef_id_col_name.strip().lower(), "ef id"} and idx_efid == -1:
            idx_efid = i
        elif key in {ef_name_col_name.strip().lower(), "ef name"} and idx_efname == -1:
            idx_efname = i
        elif key == dummy_col_name.strip().lower() and idx_dummy == -1:
            idx_dummy = i
        elif key == dummy_name_col_name.strip().lower() and idx_dummy_name == -1:
            idx_dummy_name = i

    if -1 in {idx_src, idx_efid, idx_efname}:
        raise RuntimeError(f"Required columns not found. src={idx_src}, ef_id={idx_efid}, ef_name={idx_efname}")

    # Add dummy_ef_id if missing (append at end)
    if idx_dummy == -1:
        idx_dummy = ws.max_column + 1
        ws.cell(row=1, column=idx_dummy).value = dummy_col_name

    # Add dummy_ef_id_name if missing (append at end)
    if idx_dummy_name == -1:
        idx_dummy_name = ws.max_column + 1
        ws.cell(row=1, column=idx_dummy_name).value = dummy_name_col_name

    set_count = 0
    created_on_fly = 0
    for r in range(2, ws.max_row + 1):
        src = ws.cell(row=r, column=idx_src).value
        src_l = str(src).strip().lower() if src is not None else ""

        # Only skip real Travel rows (exact match), not booklets that contain the word "Travel"
        if src_l == "travel":
            continue

        ef_name = ws.cell(row=r, column=idx_efname).value
        k = normalize_name(ef_name)
        if k == "":
            efid = ws.cell(row=r, column=idx_efid).value
            efid_s = "" if efid is None else str(efid).strip()
            if efid_s and efid_s.lower() != "nan":
                k = f"__efid__:{efid_s}"
            else:
                k = "__missing__"

        dummy = key_to_dummy.get(k)
        if not dummy:
            dummy = f"DUMMY_{next_id:05d}"
            next_id += 1
            key_to_dummy[k] = dummy
            # best-effort canonical name for new key
            key_to_name[k] = str(ef_name).strip() if ef_name is not None else ""
            created_on_fly += 1

        dummy_name = key_to_name.get(k, "")
        ws.cell(row=r, column=idx_dummy).value = dummy
        ws.cell(row=r, column=idx_dummy_name).value = dummy_name
        set_count += 1

    colored_rows = apply_source_row_coloring(wb=wb, source_col_name=source_col_name)

    workbook_out.parent.mkdir(parents=True, exist_ok=True)
    wb.save(workbook_out)
    print(f"Wrote: {workbook_out}")
    print(f"dummy set (non-travel rows): {set_count}")
    print(f"new dummy keys created on fly: {created_on_fly}")
    print(f"rows colored (all sheets with '{source_col_name}'): {colored_rows}")


def main() -> None:
    ap = argparse.ArgumentParser(description="Generate/apply dummy_ef_id + dummy_ef_id_name (no calculations).")
    ap.add_argument("--workbook-in", required=True, help="Input window workbook (.xlsx)")
    ap.add_argument("--workbook-out", required=True, help="Output workbook (.xlsx)")
    ap.add_argument("--mapping-csv", required=True, help="dummy mapping csv (from generate step)")
    ap.add_argument("--sheet", default="S3 Cat 1 Purchased G&S", help="Sheet name")
    args = ap.parse_args()

    apply_dummy_mapping(
        workbook_in=Path(args.workbook_in),
        workbook_out=Path(args.workbook_out),
        mapping_csv=Path(args.mapping_csv),
        sheet_name=args.sheet,
    )


if __name__ == "__main__":
    main()

